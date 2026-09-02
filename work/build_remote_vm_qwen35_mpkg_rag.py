#!/usr/bin/env python3
"""Build the self-contained remote-VM MP-KG-RAG notebook."""

from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path


CORE_PATH = Path(__file__).with_name("mpkg_rag_core.py")
CORE_SOURCE = CORE_PATH.read_text(encoding="utf-8")
CORE_SOURCE_SHA256 = hashlib.sha256(CORE_SOURCE.encode("utf-8")).hexdigest()
EVAL_CORE_PATH = Path(__file__).with_name("mpkg_eval_core.py")
EVAL_CORE_SOURCE = EVAL_CORE_PATH.read_text(encoding="utf-8")
EVAL_CORE_SOURCE_SHA256 = hashlib.sha256(EVAL_CORE_SOURCE.encode("utf-8")).hexdigest()
LOCK_PATH = Path(__file__).resolve().parents[1] / "requirements-remote-vm.lock"
LOCK_SOURCE = LOCK_PATH.read_text(encoding="utf-8")
LOCKFILE_SHA256 = hashlib.sha256(LOCK_SOURCE.encode("utf-8")).hexdigest()
EMITTED_APPLICATION_PACKAGE_COUNT = sum(
    1 for raw_line in LOCK_SOURCE.splitlines()
    if (raw_line.strip() and not raw_line.lstrip().startswith(("#", "-")) and "==" in raw_line)
)
MANAGED_PACKAGE_COUNT = 23
RESOLVED_PACKAGE_COUNT = EMITTED_APPLICATION_PACKAGE_COUNT + MANAGED_PACKAGE_COUNT
if (EMITTED_APPLICATION_PACKAGE_COUNT, MANAGED_PACKAGE_COUNT, RESOLVED_PACKAGE_COUNT) != (115, 23, 138):
    raise RuntimeError("lock_package_count_mismatch")
LOCK_PACKAGE_COUNTS = {
    "emitted_application_package_count": EMITTED_APPLICATION_PACKAGE_COUNT,
    "managed_package_count": MANAGED_PACKAGE_COUNT,
    "resolved_package_count": RESOLVED_PACKAGE_COUNT,
}


def markdown(source: str) -> dict:
    return {"cell_type": "markdown", "metadata": {}, "source": [line + "\n" for line in source.splitlines()]}


def code(source: str) -> dict:
    return {"cell_type": "code", "execution_count": None, "metadata": {}, "outputs": [], "source": [line + "\n" for line in source.splitlines()]}


def build_notebook(output_path: Path) -> None:
    cells = [
        markdown("""# Remote-VM Qwen3.5 Multi-Perspective MP-KG-RAG

This generated notebook embeds the exact tested `work/mpkg_rag_core.py` source.
The generator model is configured once and reused sequentially.
The same Qwen model is loaded sequentially for extraction/signatures and final
generation; retrieval models are released between those phases. Graph-on and
graph-off use identical frozen record IDs and query signatures."""),
code(f"""# 01 - Install the checked-in exact lock on the remote VM. Run once, then restart the kernel.
from pathlib import Path as _Path
import hashlib as _hashlib
_LOCKFILE_SHA256_EXPECTED = {LOCKFILE_SHA256!r}
_REMOTE_LOCK_PATH = _Path('/tmp/mpkg-rag-requirements-remote-vm.lock')
_REMOTE_LOCK_PATH.write_text({LOCK_SOURCE!r}, encoding='utf-8')
assert _hashlib.sha256(_REMOTE_LOCK_PATH.read_bytes()).hexdigest() == _LOCKFILE_SHA256_EXPECTED, 'lockfile_source_hash_mismatch'
# The hashed lock contains every non-managed transitive application package.
# --no-deps prevents pip from replacing the managed torch/CUDA base runtime.
# --require-hashes makes every emitted application artifact hash-verified.
%pip install --requirement /tmp/mpkg-rag-requirements-remote-vm.lock --require-hashes --no-input --disable-pip-version-check --no-deps
"""),
        code(CORE_SOURCE + f"\nCORE_SOURCE_SHA256 = {CORE_SOURCE_SHA256!r}\n"),
        code(EVAL_CORE_SOURCE + f"\nEVAL_CORE_SOURCE_SHA256 = {EVAL_CORE_SOURCE_SHA256!r}\n"),
        code("""# 02 - Imports, reproducibility, run layout, and memory diagnostics.
import os, re, gc, json, time, math, random, hashlib, unicodedata, tempfile, platform, sys, importlib.metadata, fcntl, atexit
from pathlib import Path
from packaging.version import Version
import numpy as np, pandas as pd, torch
from tqdm.auto import tqdm

LOCKFILE_SHA256 = _LOCKFILE_SHA256_EXPECTED
LOCK_PACKAGE_COUNTS = {"emitted_application_package_count": 115, "managed_package_count": 23, "resolved_package_count": 138}
DEFAULT_GENERATOR_MODEL_REVISION = "3764fa359b9082ea5a1e4a5e3ac3aaf6e9671636"
DEFAULT_EMBEDDING_MODEL_REVISION = "5617a9f61b028005a4858fdac845db406aefb181"
DEFAULT_RERANKER_MODEL_REVISION = "953dc6f6f85a1b2dbfca4c34a2796e7dde08d41e"
def require_model_revision(revision, model_role):
    value = str(revision or "").strip()
    if re.fullmatch(r"[0-9a-fA-F]{40}", value) is None:
        raise RuntimeError(f"immutable_model_revision_unavailable:{model_role}")
    return value.lower()
GENERATOR_MODEL_REVISION = require_model_revision(os.environ.get("MODEL_REVISION", DEFAULT_GENERATOR_MODEL_REVISION), "generator")
EMBEDDING_MODEL_REVISION = require_model_revision(os.environ.get("EMBEDDING_MODEL_REVISION", DEFAULT_EMBEDDING_MODEL_REVISION), "embedding")
RERANKER_MODEL_REVISION = require_model_revision(os.environ.get("RERANKER_MODEL_REVISION", DEFAULT_RERANKER_MODEL_REVISION), "reranker")
QWEN35_MIN_TRANSFORMERS_VERSION = "5.2.0"
def validate_qwen35_transformers_compatibility():
    observed = importlib.metadata.version("transformers")
    if Version(observed) < Version(QWEN35_MIN_TRANSFORMERS_VERSION):
        raise RuntimeError(f"qwen35_transformers_incompatible:observed={observed}:minimum={QWEN35_MIN_TRANSFORMERS_VERSION}")
    return {"observed": observed, "minimum": QWEN35_MIN_TRANSFORMERS_VERSION, "status": "compatible"}
QWEN35_TRANSFORMERS_COMPATIBILITY = validate_qwen35_transformers_compatibility()
SEED = 3407
BOOTSTRAP_SEED = SEED
os.environ.setdefault(\"PYTHONHASHSEED\", str(SEED))
random.seed(SEED); np.random.seed(SEED); torch.manual_seed(SEED)
if torch.cuda.is_available(): torch.cuda.manual_seed_all(SEED)
assert torch.cuda.is_available(), \"An NVIDIA GPU is required.\"
MANAGED_ACCELERATOR_CONTRACT = {
    \"torch\": \"2.11.0+cu128\",
    \"torchvision\": \"0.26.0+cu128\",
    \"triton\": \"3.6.0\",
    \"xformers\": \"0.0.35\",
    \"torchao\": \"0.17.0+cu128\",
    \"nvidia-cuda-runtime-cu12\": \"12.8.90\",
    \"nvidia-nvjitlink-cu12\": \"12.8.93\",
    \"nvidia-nvtx-cu12\": \"12.8.90\",
    \"torch_cuda\": \"12.8\",
}
MANAGED_ACCELERATOR_CONTRACT_HASH = stable_id(\"managed-accelerator-contract.v1\", json.dumps(MANAGED_ACCELERATOR_CONTRACT, sort_keys=True))
def validate_managed_accelerator():
    observed = {}
    for package_name, expected_version in MANAGED_ACCELERATOR_CONTRACT.items():
        if package_name == \"torch_cuda\": continue
        try:
            observed[package_name] = importlib.metadata.version(package_name)
        except importlib.metadata.PackageNotFoundError as error:
            raise RuntimeError(f\"managed_accelerator_distribution_missing:{package_name}\") from error
        if observed[package_name] != expected_version:
            raise RuntimeError(f\"managed_accelerator_distribution_mismatch:{package_name}:{observed[package_name]}:{expected_version}\")
    torch_cuda = str(getattr(torch.version, \"cuda\", \"\") or \"\")
    if torch_cuda != MANAGED_ACCELERATOR_CONTRACT[\"torch_cuda\"]:
        raise RuntimeError(f\"managed_torch_cuda_mismatch:torch_cuda={torch_cuda}:expected={MANAGED_ACCELERATOR_CONTRACT['torch_cuda']}\")
    try:
        import torchvision, triton, xformers, torchao, bitsandbytes
        if not callable(getattr(bitsandbytes, \"matmul_4bit\", None)) or getattr(getattr(bitsandbytes, \"cextension\", None), \"lib\", None) is None:
            raise RuntimeError(\"bitsandbytes_kernel_unavailable\")
        import xformers.ops
        import triton.language as tl
    except Exception as error:
        raise RuntimeError(f\"managed_accelerator_import_failed:{type(error).__name__}:{error}\") from error
    try:
        probe = torch.ones((2, 2), device=\"cuda\", dtype=torch.float32)
        result = probe @ probe
        torch.cuda.synchronize()
        if not torch.allclose(result, torch.full((2, 2), 2.0, device=\"cuda\")):
            raise RuntimeError(\"matmul_result_mismatch\")
        bnb_input = torch.ones((1, 2), device=\"cuda\", dtype=torch.float16)
        bnb_weight = torch.ones((2, 2), device=\"cuda\", dtype=torch.float16)
        bnb_quantized, bnb_state = bitsandbytes.functional.quantize_4bit(bnb_weight, quant_type=\"nf4\")
        bnb_result = bitsandbytes.matmul_4bit(bnb_input, bnb_quantized.t(), bnb_state)
        if tuple(bnb_result.shape) != (1, 2) or not bool(torch.isfinite(bnb_result).all()):
            raise RuntimeError(\"bitsandbytes_smoke_failed\")
        xformers_query = torch.randn((1, 4, 1, 8), device=\"cuda\", dtype=torch.float16)
        xformers_key = torch.randn((1, 4, 1, 8), device=\"cuda\", dtype=torch.float16)
        xformers_value = torch.randn((1, 4, 1, 8), device=\"cuda\", dtype=torch.float16)
        xformers_result = xformers.ops.memory_efficient_attention(xformers_query, xformers_key, xformers_value)
        torch.cuda.synchronize()
        if tuple(xformers_result.shape) != tuple(xformers_query.shape) or not bool(torch.isfinite(xformers_result).all()):
            raise RuntimeError(\"xformers_smoke_failed\")
        @triton.jit
        def _managed_contract_add_kernel(input_ptr, output_ptr, n_elements, BLOCK_SIZE: tl.constexpr):
            pid = tl.program_id(0)
            offsets = pid * BLOCK_SIZE + tl.arange(0, BLOCK_SIZE)
            mask = offsets < n_elements
            values = tl.load(input_ptr + offsets, mask=mask)
            tl.store(output_ptr + offsets, values + 1.0, mask=mask)
        triton_input = torch.ones(128, device=\"cuda\", dtype=torch.float32)
        triton_output = torch.empty_like(triton_input)
        _managed_contract_add_kernel[lambda meta: (triton.cdiv(triton_input.numel(), meta[\"BLOCK_SIZE\"]),)](triton_input, triton_output, triton_input.numel(), BLOCK_SIZE=128)
        torch.cuda.synchronize()
        if not bool(torch.allclose(triton_output, torch.full_like(triton_output, 2.0))):
            raise RuntimeError(\"triton_kernel_smoke_failed\")
        del bnb_input, bnb_weight, bnb_quantized, bnb_state, bnb_result, probe, result, xformers_query, xformers_key, xformers_value, xformers_result, triton_input, triton_output
        torch.cuda.empty_cache()
    except Exception as error:
        raise RuntimeError(f\"managed_accelerator_smoke_failed:{type(error).__name__}:{error}\") from error
    return {\"contract\": MANAGED_ACCELERATOR_CONTRACT, \"contract_hash\": MANAGED_ACCELERATOR_CONTRACT_HASH, \"observed_distributions\": observed, \"torch_cuda\": torch_cuda, \"bitsandbytes_kernel\": \"matmul_4bit\", \"xformers_kernel\": \"memory_efficient_attention\", \"triton_kernel\": \"managed_contract_add_kernel\", \"cuda_smoke\": \"passed\"}
MANAGED_ACCELERATOR = validate_managed_accelerator()
PROJECT_ROOT = Path(os.environ.get(\"PROJECT_ROOT\", \"/workspace/mp_kg_rag\")).expanduser().resolve()
SHARD_INDEX = int(os.environ.get(\"SHARD_INDEX\", \"0\")); SHARD_COUNT = int(os.environ.get(\"SHARD_COUNT\", \"1\"))
MAX_EXPERIMENT_ROWS = int(os.environ.get(\"MAX_EXPERIMENT_ROWS\", \"1550\")); CACHE_BASE_CAPACITY = int(os.environ.get(\"CACHE_MAX_RECORDS\", \"256\")); EFFECTIVE_CACHE_CAPACITY = derive_effective_cache_capacity(CACHE_BASE_CAPACITY, row_limit=MAX_EXPERIMENT_ROWS, shard_count=SHARD_COUNT)
BASE_RUN_NAME = os.environ.get(\"RUN_NAME\", \"qwen35_mp_kg_rag_v2_semantic\")
RUN_NAME = derive_shard_run_name(BASE_RUN_NAME, SHARD_INDEX, SHARD_COUNT)
RUN = PROJECT_ROOT / \"runs\" / RUN_NAME
RUN_LOCK_HANDLE = None
def acquire_run_lock():
    global RUN_LOCK_HANDLE
    RUN.mkdir(parents=True, exist_ok=True)
    RUN_LOCK_HANDLE = (RUN / \".run.lock\").open(\"a+\", encoding=\"utf-8\")
    try:
        fcntl.flock(RUN_LOCK_HANDLE.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
    except BlockingIOError:
        RUN_LOCK_HANDLE.close(); RUN_LOCK_HANDLE = None
        raise RuntimeError(\"run_already_active\")
def release_run_lock():
    global RUN_LOCK_HANDLE
    if RUN_LOCK_HANDLE is None: return
    try: fcntl.flock(RUN_LOCK_HANDLE.fileno(), fcntl.LOCK_UN)
    finally: RUN_LOCK_HANDLE.close(); RUN_LOCK_HANDLE = None
acquire_run_lock(); atexit.register(release_run_lock)
for folder in [RUN / x for x in [\"artifacts\", \"checkpoints\", \"exports\", \"logs\", \"review_queue\"]]: folder.mkdir(parents=True, exist_ok=True)
def now(): return time.strftime(\"%Y-%m-%dT%H:%M:%SZ\", time.gmtime())
def write_json(path, value): Path(path).write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding=\"utf-8\")
def write_json_atomic(path, value):
    path = Path(path); path.parent.mkdir(parents=True, exist_ok=True); temporary = None
    try:
        with tempfile.NamedTemporaryFile(\"w\", encoding=\"utf-8\", dir=path.parent, prefix=f\".{path.name}.\", suffix=\".tmp\", delete=False) as stream:
            temporary = Path(stream.name); json.dump(value, stream, ensure_ascii=False, indent=2); stream.flush(); os.fsync(stream.fileno())
        os.replace(temporary, path)
    except Exception:
        if temporary is not None: temporary.unlink(missing_ok=True)
        raise
def _locked_append_line(path, line):
    path = Path(path); path.parent.mkdir(parents=True, exist_ok=True); lock_path = path.with_name(path.name + \".lock\")
    with lock_path.open(\"a+\", encoding=\"utf-8\") as lock:
        fcntl.flock(lock.fileno(), fcntl.LOCK_EX)
        try:
            with path.open(\"a\", encoding=\"utf-8\") as stream: stream.write(line); stream.flush(); os.fsync(stream.fileno())
        finally: fcntl.flock(lock.fileno(), fcntl.LOCK_UN)
def append_jsonl(path, row):
    _locked_append_line(path, json.dumps(row, ensure_ascii=False, sort_keys=True) + \"\\n\")
def load_jsonl(path):
    path = Path(path); return [json.loads(x) for x in path.read_text(encoding=\"utf-8\").splitlines() if x.strip()] if path.exists() else []
def log_event(stage, event, **data):
    append_jsonl(RUN / \"logs\" / \"events.jsonl\", {\"at\": now(), \"stage\": stage, \"event\": event, \"run_name\": RUN_NAME, \"config_hash\": globals().get(\"CONFIG_HASH\"), \"run_identity_hash\": globals().get(\"RUN_IDENTITY_HASH\"), \"shard\": {\"index\": SHARD_INDEX, \"count\": SHARD_COUNT}, \"manifest_hashes\": {key: globals().get(key) for key in [\"CORPUS_MANIFEST_HASH\", \"CHUNK_MANIFEST_HASH\", \"GRAPH_MANIFEST_HASH\", \"SPLIT_MEMBERSHIP_HASH\"]}, **data})
def _resolved_identity(name):
    value = globals().get(name)
    return value if value is not None else f"pending:{name}"
def log_event(stage, event, **data):
    append_jsonl(RUN / "logs" / "events.jsonl", {"at": now(), "stage": stage, "event": event, "run_name": RUN_NAME, "config_hash": _resolved_identity("CONFIG_HASH"), "run_identity_hash": _resolved_identity("RUN_IDENTITY_HASH"), "identity_status": "ready" if globals().get("RUN_IDENTITY_HASH") else "pending", "shard": {"index": SHARD_INDEX, "count": SHARD_COUNT}, "manifest_hashes": {key: _resolved_identity(key) for key in ["CORPUS_MANIFEST_HASH", "CHUNK_MANIFEST_HASH", "GRAPH_MANIFEST_HASH", "SPLIT_MEMBERSHIP_HASH"]}, **data})
def gpu_snapshot(label):
    if not torch.cuda.is_available(): return {\"label\": label, \"cuda\": False}
    torch.cuda.synchronize(); return {\"label\": label, \"cuda\": True, \"device\": torch.cuda.get_device_name(0), \"bf16_supported\": bool(torch.cuda.is_bf16_supported()), \"allocated_gb\": torch.cuda.memory_allocated() / 2**30, \"reserved_gb\": torch.cuda.memory_reserved() / 2**30, \"peak_allocated_gb\": torch.cuda.max_memory_allocated() / 2**30, \"peak_reserved_gb\": torch.cuda.max_memory_reserved() / 2**30}
MEMORY_SNAPSHOTS = [gpu_snapshot(\"before_models\")]
def package_version(name):
    try: return importlib.metadata.version(name)
    except importlib.metadata.PackageNotFoundError: return None
def _lock_requirements(lock_source):
    requirements = {}
    for raw_line in lock_source.splitlines():
        line = raw_line.strip()
        if not line or line.startswith(\"#\") or line.startswith(\"--\"): continue
        if \"==\" not in line or line.count(\"==\") != 1: raise RuntimeError(\"invalid_exact_lock_entry\")
        name, version = (part.strip() for part in line.split(\"==\"))
        version = version.rstrip(\"\\\\\").strip()
        if not name or not version or any(char in name for char in \"<>!=~[]; \"): raise RuntimeError(\"invalid_exact_lock_entry\")
        requirements[name.lower().replace(\"_\", \"-\")] = version
    if not requirements: raise RuntimeError(\"empty_dependency_lock\")
    return dict(sorted(requirements.items()))
APPLICATION_CONSTRAINTS = _lock_requirements(__LOCK_SOURCE_REPR__)
LOCKED_REQUIREMENTS = APPLICATION_CONSTRAINTS
def installed_package_versions():
    return {str(dist.metadata.get(\"Name\") or \"\").lower().replace(\"_\", \"-\"): str(dist.version) for dist in importlib.metadata.distributions() if dist.metadata.get(\"Name\")}
def verify_locked_environment(expected=None):
    expected = dict(expected or LOCKED_REQUIREMENTS)
    installed = installed_package_versions()
    missing = sorted(name for name in expected if name not in installed)
    mismatched = {name: {\"expected\": version, \"installed\": installed.get(name)} for name, version in expected.items() if installed.get(name) != version}
    if missing or mismatched: raise RuntimeError(json.dumps({\"locked_environment_mismatch\": True, \"missing\": missing, \"mismatched\": mismatched}, sort_keys=True))
    return {\"locked_requirements\": expected, \"installed_package_versions\": installed, \"verified\": True}
LOCK_ENVIRONMENT = verify_locked_environment()
def _cuda_driver_version():
    try: return torch._C._cuda_getDriverVersion()
    except Exception: return None
ENVIRONMENT_FINGERPRINT = {\"python_version\": platform.python_version(), \"python_implementation\": platform.python_implementation(), \"python_executable\": sys.executable, \"python_hash_seed\": os.environ.get(\"PYTHONHASHSEED\"), \"hash_randomization\": bool(sys.flags.hash_randomization), \"locked_requirements\": LOCK_ENVIRONMENT[\"locked_requirements\"], \"installed_package_versions\": LOCK_ENVIRONMENT[\"installed_package_versions\"], \"lock_package_counts\": LOCK_PACKAGE_COUNTS, \"pytorch_version\": getattr(torch, \"__version__\", None), \"torch_cuda_version\": getattr(torch.version, \"cuda\", None), \"cuda_driver_version\": _cuda_driver_version(), \"cudnn_version\": torch.backends.cudnn.version() if torch.cuda.is_available() else None, \"cuda_available\": bool(torch.cuda.is_available()), \"managed_accelerator\": MANAGED_ACCELERATOR, \"qwen35_transformers_compatibility\": QWEN35_TRANSFORMERS_COMPATIBILITY, \"gpu_name\": torch.cuda.get_device_name(0) if torch.cuda.is_available() else None, \"gpu_count\": int(torch.cuda.device_count()) if torch.cuda.is_available() else 0, \"model_revisions\": {\"bertscore_model_id\": BERTSCORE_MODEL_ID, \"bertscore_model_revision\": BERTSCORE_MODEL_REVISION, \"nli_model_id\": NLI_MODEL_ID, \"nli_model_revision\": NLI_MODEL_REVISION, \"generator_model_id\": (os.environ.get(\"MODEL_ID\") or \"unsloth/\" + \"Qwen3.5-4B\"), \"generator_model_revision\": os.environ.get(\"MODEL_REVISION\"), \"embedding_model_id\": \"BAAI/bge-m3\", \"embedding_model_revision\": os.environ.get(\"EMBEDDING_MODEL_REVISION\"), \"reranker_model_id\": \"BAAI/bge-reranker-v2-m3\", \"reranker_model_revision\": os.environ.get(\"RERANKER_MODEL_REVISION\")}}
ENVIRONMENT_FINGERPRINT[\"managed_accelerator_contract_hash\"] = MANAGED_ACCELERATOR_CONTRACT_HASH
ENVIRONMENT_FINGERPRINT[\"detoxify_metric_contract\"] = {\"device\": \"cpu\", \"required\": True, \"language\": \"en\"}
ENVIRONMENT_FINGERPRINT_HASH = stable_id(\"environment-fingerprint.v1\", json.dumps(ENVIRONMENT_FINGERPRINT, ensure_ascii=False, sort_keys=True, default=str))
""".replace("__LOCK_SOURCE_REPR__", repr(LOCK_SOURCE))),
        code("""# 03 - Configuration and frozen run identity.
ENVIRONMENT_FINGERPRINT["model_revisions"].update({"generator_model_revision": GENERATOR_MODEL_REVISION, "embedding_model_revision": EMBEDDING_MODEL_REVISION, "reranker_model_revision": RERANKER_MODEL_REVISION})
ENVIRONMENT_FINGERPRINT_HASH = stable_id("environment-fingerprint.v1", json.dumps(ENVIRONMENT_FINGERPRINT, ensure_ascii=False, sort_keys=True, default=str))
CONFIG = {
    \"dataset_xlsx\": PROJECT_ROOT / \"input\" / \"dataset (1).xlsx\", \"corpus_root\": PROJECT_ROOT / \"input\" / \"lgbt_hate_speech_kg_sources\", \"sheet_name\": \"Final_Dataset\",
    \"evaluation_categories\": [\"Homophobic\", \"Non-Homophobic\"], \"max_experiment_rows\": MAX_EXPERIMENT_ROWS, \"split_name\": os.environ.get(\"SPLIT_NAME\", \"test\"), \"smoke_test\": False, \"smoke_rows\": 8,
    \"minimum_parse_rate\": 0.98, \"require_citation_nli\": True, \"require_detoxify\": True, \"detoxify_device\": \"cpu\", \"evidence_char_budget\": 9000, \"embedding_model\": \"BAAI/bge-m3\", \"reranker_model\": \"BAAI/bge-reranker-v2-m3\", \"generator_model\": os.environ.get(\"MODEL_ID\", \"unsloth/Qwen3.5-4B\"), \"load_in_4bit\": True, \"max_seq_length\": 4096,
    \"perspective_batch_size\": 5, \"extraction_batch_size\": 24, \"perspective_max_new_tokens\": 768, \"plan_max_new_tokens\": 768, \"answer_max_new_tokens\": 768, \"reasoning_max_new_tokens\": 192, \"thinking_enabled\": True, \"minimum_healthy_factual_documents\": 40, \"minimum_factual_document_coverage\": 0.80, \"dense_top_k\": 12, \"bm25_top_k\": 12, \"graph_top_k\": 8, \"rerank_top_k\": 5,
    \"max_chunk_chars\": 1800, \"chunk_overlap_chars\": 220, \"minimum_authority\": 0.60, \"minimum_dense_score\": 0.25, \"minimum_graph_score\": 0.0, \"minimum_rerank_probability\": 0.55, \"max_graph_hops\": 2, \"rrf_constant\": 60.0,
    \"precompute_evidence\": True, \"graph_ablation\": False, \"generation_variants\": [\"qwen_zero_shot\", \"qwen_few_shot\", \"kg_rag\", \"mp_kg_rag\"], \"shard_index\": SHARD_INDEX, \"shard_count\": SHARD_COUNT, \"cache_max_records\": EFFECTIVE_CACHE_CAPACITY,
    \"evaluation_languages\": [\"en\", \"hi\", \"ta\"], \"bertscore_model_id\": BERTSCORE_MODEL_ID, \"bertscore_model_revision\": BERTSCORE_MODEL_REVISION, \"lockfile_sha256\": LOCKFILE_SHA256, \"lock_package_counts\": LOCK_PACKAGE_COUNTS, \"environment_fingerprint_hash\": ENVIRONMENT_FINGERPRINT_HASH,
    \"nli_model_id\": NLI_MODEL_ID, \"nli_model_revision\": NLI_MODEL_REVISION, \"nli_model_label_mapping\": MODEL_LABEL_MAPPING, \"nli_dataset_label_mapping\": DATASET_LABEL_MAPPING, \"nli_calibration_examples\": 600, \"nli_calibration_min_support\": 50, \"nli_calibration_bootstrap\": 2000, \"nli_min_accuracy\": 0.70, \"nli_min_entailment_precision\": 0.70, \"nli_min_entailment_recall\": 0.70, \"nli_min_per_label_support\": 50, \"nli_calibration_provenance\": {language: {**NLI_DATASET_PROVENANCE[language], \"dataset_content_hash\": None, \"code_hash\": CORE_SOURCE_SHA256, \"eval_core_hash\": EVAL_CORE_SOURCE_SHA256, \"model_id\": NLI_MODEL_ID, \"model_revision\": NLI_MODEL_REVISION, \"model_label_mapping\": MODEL_LABEL_MAPPING, \"dataset_label_mapping\": DATASET_LABEL_MAPPING} for language in [\"en\", \"hi\", \"ta\"]}, \"annotation_max_ids\": 200,
}
CONFIG["generator_model_revision"] = GENERATOR_MODEL_REVISION
CONFIG["embedding_model_revision"] = EMBEDDING_MODEL_REVISION
CONFIG["reranker_model_revision"] = RERANKER_MODEL_REVISION
CONFIG["managed_accelerator_contract"] = MANAGED_ACCELERATOR_CONTRACT
CONFIG["managed_accelerator_contract_hash"] = MANAGED_ACCELERATOR_CONTRACT_HASH
CONFIG["qwen35_transformers_compatibility"] = QWEN35_TRANSFORMERS_COMPATIBILITY
CONFIG["nli_model_label_mapping"] = MODEL_LABEL_MAPPING
CONFIG["nli_dataset_label_mapping"] = DATASET_LABEL_MAPPING
CONFIG["nli_min_per_label_support"] = 50
CONFIG["nli_min_accuracy_lower"] = 0.70
CONFIG["nli_min_entailment_precision_lower"] = 0.70
CONFIG["nli_min_entailment_recall_lower"] = 0.70
for _language in ["en", "hi", "ta"]:
    CONFIG["nli_calibration_provenance"][_language]["dataset_content_hash"] = None
    CONFIG["nli_calibration_provenance"][_language]["model_label_mapping"] = MODEL_LABEL_MAPPING
    CONFIG["nli_calibration_provenance"][_language]["dataset_label_mapping"] = DATASET_LABEL_MAPPING
IDENTITY = identity_config(CONFIG)
CONFIG_HASH = IDENTITY[\"config_hash\"]
SPLIT_IDENTITY_HASH = IDENTITY[\"split_identity_hash\"]
RUN_IDENTITY_HASH = IDENTITY[\"run_identity_hash\"]
PROMPT_TEMPLATE_HASH = stable_id(\"mp_kg_rag_prompt_templates_v4_closed_catalog_fail_closed\")
FEW_SHOT_PROMPT_REVISION = \"few-shot.v1.contamination-safe.static\"
FEW_SHOT_EXAMPLES = [
    {\"post\": \"A claim says a group is dangerous because of identity.\", \"target\": \"identity-based hate\", \"response\": \"Identity does not determine a person's character. Respond to people as individuals and avoid generalizing about a protected group.\"},
    {\"post\": \"Someone says respect means staying silent about discrimination.\", \"target\": \"respect and equality\", \"response\": \"Respect includes listening to affected people and rejecting discrimination while keeping the conversation constructive.\"},
]
EXTRACTION_MODEL = CONFIG[\"generator_model\"]; EXTRACTION_PROMPT_REVISION = \"semantic-claims.v1-qwen-constrained-repair1\"
RETRIEVAL_THRESHOLDS = {k: CONFIG[k] for k in [\"minimum_authority\", \"minimum_dense_score\", \"minimum_graph_score\", \"minimum_rerank_probability\", \"max_graph_hops\"]}
GRAPH_CONFIG = {\"max_hops\": CONFIG[\"max_graph_hops\"], \"hop_decay\": 0.75, \"minimum_dense_score\": CONFIG[\"minimum_dense_score\"], \"minimum_graph_score\": CONFIG[\"minimum_graph_score\"], \"minimum_rerank_probability\": CONFIG[\"minimum_rerank_probability\"], \"max_evidence\": CONFIG[\"rerank_top_k\"], \"weights\": {\"query_entity\": 2.0, \"predicate\": 1.0, \"polarity\": 0.5, \"modality\": 0.4, \"stance\": 0.8, \"review_state\": 0.8, \"authority\": 0.8, \"extraction_confidence\": 0.0, \"seed_score\": 1.0, \"hop_decay\": 0.6}, \"review_state_scores\": {\"accepted\": 1.0, \"reviewed\": 0.85, \"unknown\": 0.0}, \"extraction_confidence_status\": \"diagnostic_only_unweighted\"}
GRAPH_CONFIG_HASH = stable_id(json.dumps(GRAPH_CONFIG, sort_keys=True))
SCORING_CALIBRATION_STATUS = "uncalibrated_diagnostic_only"
SELF_CONFIDENCE_STATUS = "uncalibrated_diagnostic_only"
QUALITY_THRESHOLDS = {"minimum_graph_linked_claim_rate": 0.02, "minimum_query_linked_entity_rate": 0.02, "minimum_linked_claims": 1, "minimum_linked_queries": 1}
def validate_checkpoint_identity(saved, expected):
    if saved != expected:
        raise RuntimeError(\"resume_identity_mismatch\")
    return True
def checkpoint_identity(record, variant):
    return {\"record_id\": str(record[\"ID\"]), \"variant\": variant, \"input_text_sha256\": str(record[\"input_text_sha256\"]), \"target_sha256\": hashlib.sha256(str(record.get(\"Target\", \"\")).encode()).hexdigest(), \"category_sha256\": hashlib.sha256(str(record.get(\"Category\", \"\")).encode()).hexdigest(), \"reference_answer_sha256\": hashlib.sha256((normalize_optional_text(record.get(\"Counter Narrative\")) or \"\").encode()).hexdigest(), \"corpus_manifest_hash\": CORPUS_MANIFEST_HASH, \"audit_manifest_hash\": AUDIT_MANIFEST_HASH, \"chunk_manifest_hash\": CHUNK_MANIFEST_HASH, \"graph_manifest_hash\": GRAPH_MANIFEST_HASH, \"extraction_model\": EXTRACTION_MODEL, \"extraction_prompt_revision\": EXTRACTION_PROMPT_REVISION, \"retrieval_thresholds\": RETRIEVAL_THRESHOLDS, \"graph_config\": GRAPH_CONFIG, \"config_hash\": CONFIG_HASH, \"run_identity_hash\": RUN_IDENTITY_HASH, \"lockfile_sha256\": LOCKFILE_SHA256, \"environment_fingerprint_hash\": ENVIRONMENT_FINGERPRINT_HASH, \"prompt_template_hash\": PROMPT_TEMPLATE_HASH, \"variant_prompt_hash\": stable_id(PROMPT_TEMPLATE_HASH, FEW_SHOT_PROMPT_REVISION if variant == \"qwen_few_shot\" else variant), \"split_name\": CONFIG[\"split_name\"], \"split_membership_hash\": SPLIT_MEMBERSHIP_HASH, \"shard_index\": SHARD_INDEX, \"shard_count\": SHARD_COUNT, \"run_name\": RUN_NAME}
"""),
        code("""# 04 - Dataset audit and manifest-authoritative collision-safe corpus registry.
from bs4 import BeautifulSoup
import fitz
def audit_dataset(path):
    required = [\"ID\", \"Text\", \"Category\", \"Target\", \"Counter Narrative\"]; frame = pd.read_excel(path, sheet_name=CONFIG[\"sheet_name\"]); assert set(required).issubset(frame.columns)
    frame = frame[required].copy()
    raw_rows = frame.to_dict(\"records\")
    base = filter_audit(raw_rows, [normalize_optional_text(row.get(\"ID\")) is not None and normalize_optional_text(row.get(\"Text\")) is not None for row in raw_rows], filter_name=\"required_id_text\", reasons=[None if normalize_optional_text(row.get(\"ID\")) is not None and normalize_optional_text(row.get(\"Text\")) is not None else \"missing_id_or_text\" for row in raw_rows])
    frame = pd.DataFrame(base[\"kept_rows\"], columns=required)
    frame[\"ID\"] = frame[\"ID\"].astype(str).str.replace(r\"\\.0$\", \"\", regex=True); assert not frame.ID.duplicated().any()
    category_result = filter_rows_by_category(frame.to_dict(\"records\"), CONFIG[\"evaluation_categories\"]); frame = pd.DataFrame(category_result[\"rows\"])
    reference_result = quarantine_missing_references(frame.to_dict(\"records\"), reference_key=\"Counter Narrative\")
    frame[\"reference_available\"] = frame[\"ID\"].astype(str).isin({row[\"ID\"] for row in reference_result[\"scorable_rows\"]})
    frame[\"CategoryNormalized\"] = frame[\"Category\"].map(normalize_category)
    frame[\"input_text_sha256\"] = frame.Text.astype(str).map(lambda x: hashlib.sha256(x.encode()).hexdigest()); frame[\"script_bucket\"] = frame.Text.astype(str).map(lambda x: \"tamil\" if any(\"\\u0b80\" <= c <= \"\\u0bff\" for c in x) else (\"devanagari\" if any(\"\\u0900\" <= c <= \"\\u097f\" for c in x) else \"latin_or_mixed\")); frame[\"stratify_key\"] = frame.CategoryNormalized.astype(str) + \"|\" + frame.Target.fillna(\"No Target\").astype(str) + \"|\" + frame.script_bucket
    write_json(RUN / \"artifacts\" / \"dataset_filter_manifest.json\", {\"filters\": [base, category_result[\"manifest\"], reference_result[\"manifest\"]], \"reference_quarantine_count\": len(reference_result[\"quarantined_rows\"]), \"reference_quarantine_policy\": \"generation_preserved_scoring_excluded\"})
    log_event(\"dataset\", \"filter_manifest_written\", reference_quarantine_count=len(reference_result[\"quarantined_rows\"]), dataset_rows=len(frame))
    return frame.reset_index(drop=True)
def make_frozen_split(frame):
    path = RUN / \"artifacts\" / \"frozen_split.json\"
    membership = [{\"ID\": str(row[\"ID\"]), \"input_text_sha256\": hashlib.sha256(str(row[\"Text\"]).encode()).hexdigest(), \"target_sha256\": hashlib.sha256(str(row.get(\"Target\", \"\")).encode()).hexdigest(), \"category_sha256\": hashlib.sha256(str(row.get(\"Category\", \"\")).encode()).hexdigest(), \"reference_answer_sha256\": hashlib.sha256((normalize_optional_text(row.get(\"Counter Narrative\")) or \"\").encode()).hexdigest()} for _, row in frame.sort_values(\"ID\").iterrows()]
    membership_hash = stable_id(\"frozen-membership.v2\", json.dumps(membership, ensure_ascii=False, sort_keys=True))
    if path.exists():
        frozen = json.loads(path.read_text())
        if frozen.get(\"config_hash\") != CONFIG_HASH or frozen.get(\"split_membership_hash\") != membership_hash or frozen.get(\"membership\") != membership:
            raise RuntimeError(\"frozen_split_identity_mismatch\")
        return frozen
    buckets = {\"train\": [], \"dev\": [], \"test\": []}
    for _, group in frame.groupby(\"stratify_key\", dropna=False):
        for i, rid in enumerate(sorted(group.ID.astype(str), key=lambda x: stable_id(x, SEED))): buckets[\"train\" if i % 10 < 7 else (\"dev\" if i % 10 < 9 else \"test\")].append(rid)
    value = {\"config_hash\": CONFIG_HASH, \"seed\": SEED, \"stratify_fields\": [\"Category\", \"Target\", \"script_bucket\"], \"split_membership_hash\": membership_hash, \"membership\": membership, \"splits\": buckets}; write_json(path, value); return value
all_dataset = audit_dataset(CONFIG[\"dataset_xlsx\"]); frozen_split = make_frozen_split(all_dataset); SPLIT_MEMBERSHIP_HASH = frozen_split[\"split_membership_hash\"]; dataset = all_dataset[all_dataset.ID.astype(str).isin(frozen_split[\"splits\"][CONFIG[\"split_name\"]])].copy(); dataset = dataset.head(CONFIG[\"smoke_rows\"] if CONFIG[\"smoke_test\"] else CONFIG[\"max_experiment_rows\"]).reset_index(drop=True)
if CONFIG[\"shard_count\"] > 1: dataset = dataset.iloc[CONFIG[\"shard_index\"]::CONFIG[\"shard_count\"]].reset_index(drop=True)
if len(dataset) > int(CONFIG[\"cache_max_records\"]): raise RuntimeError(\"cache_capacity_underprovisioned_before_work\")
log_event(\"runtime\", \"cache_capacity_bound\", cache_capacity=CONFIG[\"cache_max_records\"], selected_dataset_rows=len(dataset), capacity_hashed=True)
source_registry_list = load_source_registry(CONFIG[\"corpus_root\"]); source_registry = pd.DataFrame(list(source_registry_list));
if source_registry.empty: raise RuntimeError(\"no_corpus_documents\")
source_registry[\"source_id\"] = source_registry[\"legacy_source_id\"]; source_registry[\"legacy_source_id\"] = source_registry[\"legacy_source_id\"].fillna(\"UNKNOWN\");
if \"review_status\" not in source_registry: source_registry[\"review_status\"] = \"unreviewed\"
else: source_registry[\"review_status\"] = source_registry[\"review_status\"].fillna(\"unreviewed\")
assert source_registry.document_uid.is_unique
def canonical_manifest_rows(frame):
    fields = [\"document_uid\", \"relative_path\", \"content_sha256\", \"source_type\", \"authority_score\", \"factual_index_allowed\", \"status\", \"status_reason\", \"quarantine_reasons\", \"review_status\"]
    rows = []
    for row in frame.to_dict(\"records\"):
        rows.append({key: row.get(key) for key in fields})
    return sorted(rows, key=lambda row: (str(row[\"document_uid\"]), str(row[\"relative_path\"]), str(row[\"content_sha256\"])))
canonical_manifest_rows = canonical_manifest_rows(source_registry); document_uid_manifest_hash = stable_id(json.dumps(canonical_manifest_rows, ensure_ascii=False, sort_keys=True, default=str)); CORPUS_MANIFEST_HASH = document_uid_manifest_hash
def parquet_safe_frame(frame):
    safe = frame.copy()
    for column in safe.columns:
        if safe[column].dtype != object: continue
        if any(isinstance(value, (dict, list, tuple, set)) for value in safe[column] if value is not None):
            safe[column] = safe[column].map(lambda value: json.dumps(value, ensure_ascii=False, sort_keys=True, default=str) if isinstance(value, (dict, list, tuple, set)) else value)
    return safe
write_json(RUN / \"artifacts\" / \"source_manifest.json\", {\"rows\": source_registry.to_dict(\"records\"), \"audit_events\": list(getattr(source_registry_list, \"audit_events\", [])), \"validation_errors\": list(getattr(source_registry_list, \"validation_errors\", [])), \"file_records_before_deduplication\": int(getattr(source_registry_list, \"file_records_before_deduplication\", len(source_registry)))}); parquet_safe_frame(source_registry).to_parquet(RUN / \"artifacts\" / \"source_registry.parquet\", index=False); parquet_safe_frame(dataset).to_parquet(RUN / \"artifacts\" / \"dataset_scope.parquet\", index=False)
print(\"Corpus files before dedup:\", source_registry_list.file_records_before_deduplication, \"| documents:\", len(source_registry), \"| hidden_metadata_file audit retained by registry\")
reviewed_orgs = sorted({str(row.get(\"organisation\")) for row in source_registry.to_dict(\"records\") if row.get(\"organisation\") and row.get(\"status\") == \"accepted\"})
BASE_ENTITY_CATALOG = build_entity_catalog(source_registry.to_dict(\"records\"), dataset[\"Target\"].astype(str).tolist(), reviewed_orgs)
write_json(RUN / \"artifacts\" / \"base_entity_catalog.json\", BASE_ENTITY_CATALOG)
"""),
code("""# 05 - Format-aware PDF/HTML extraction, source-level audit, and document_uid chunks.
EXTRACTOR_REVISION = \"format_aware_pdf_html_audit_v4_fail_closed\"
COOKIE_MARKERS = (\"enable cookies\", \"please enable cookies\", \"javascript is disabled\", \"checking your browser\", \"access denied\", \"captcha\")
RECOVERY_EXPECTED_SOURCES = {\"SRC029\": \"nested-div\", \"SRC059\": \"long-heading\"}
PARENT_DOCUMENT_UID_FIELD = \"parent_document_uid\"
def normalize_text(text): return re.sub(r\"\\s+\", \" \", str(text or \"\")).strip()
def canonical_page(value): return None if value is None or (isinstance(value, float) and math.isnan(value)) or pd.isna(value) else int(value)
def split_windows(text, max_chars, overlap):
    # Exact, non-overlapping sentence windows come from the tested evaluation core.
    return sentence_aligned_windows(text, max_chars, overlap)
def content_based_interstitial(text, title=\"\"):
    lowered = normalize_text(text).casefold()
    title_lowered = normalize_text(title).casefold()
    return (any(marker in lowered for marker in COOKIE_MARKERS) and len(lowered) < 700) or (\"pubmed\" in title_lowered and \"cookies-required\" in lowered)
def _validated_fallback_path(source):
    metadata = source.get(\"manifest_metadata\") if isinstance(source.get(\"manifest_metadata\"), dict) else {}
    fallback_path = source.get(\"fallback_path\") or metadata.get(\"fallback_path\")
    fallback_sha256 = source.get(\"fallback_sha256\") or metadata.get(\"fallback_sha256\")
    if not fallback_path or not isinstance(fallback_sha256, str): return None
    candidate = Path(str(fallback_path)); candidate = candidate if candidate.is_absolute() else Path(source[\"path\"]).parent / candidate
    if not candidate.is_file(): return None
    actual = hashlib.sha256(candidate.read_bytes()).hexdigest()
    return (candidate, actual) if actual == fallback_sha256.casefold() else None
def _html_root(soup):
    # Prefer semantic content containers, including role=main, before body fallback.
    return soup.find(\"main\") or soup.find(\"article\") or soup.select_one('[role=\"main\"]') or soup.body or soup
def deduplicate_contained_text_rows(rows, return_audit=False):
    kept, seen_hashes, audit = [], set(), []
    for row in rows:
        text = normalize_text(row.get(\"text\"))
        if not text: continue
        identity = {\"document_uid\": row.get(\"document_uid\"), \"relative_path\": row.get(\"relative_path\"), \"locator\": {\"paragraph\": row.get(\"paragraph\"), \"page\": row.get(\"page\")}}
        if any(normalize_text(existing.get(\"text\")) == text for existing in kept):
            audit.append({\"filter\": \"html_element\", \"kept\": False, \"reason\": \"duplicate_text\", **identity}); continue
        contained_text = len(text) >= 40 and any(text.casefold() in normalize_text(existing.get(\"text\")).casefold() for existing in kept)
        if contained_text:
            audit.append({\"filter\": \"html_element\", \"kept\": False, \"reason\": \"contained_text\", **identity}); continue
        removed = [existing for existing in kept if len(normalize_text(existing.get(\"text\"))) >= 40 and normalize_text(existing.get(\"text\")).casefold() in text.casefold()]
        for existing in removed:
            audit.append({\"filter\": \"html_element\", \"kept\": False, \"reason\": \"contained_by_larger_text\", \"document_uid\": existing.get(\"document_uid\"), \"relative_path\": existing.get(\"relative_path\"), \"locator\": {\"paragraph\": existing.get(\"paragraph\"), \"page\": existing.get(\"page\")}})
        kept = [existing for existing in kept if existing not in removed]
        kept.append(row)
    return (kept, audit) if return_audit else kept
def extract_html_document(source):
    html_path = Path(source[\"path\"]); extraction_method = \"html_dom_content_container\"; fallback = None; fallback_attempted = False; fallback_path = None; fallback_sha256 = None; fallback_source_identity = None
    while True:
        soup = BeautifulSoup(html_path.read_text(encoding=\"utf-8\", errors=\"ignore\"), \"html.parser\")
        for element in soup([\"script\", \"style\", \"nav\", \"header\", \"footer\", \"aside\", \"form\", \"noscript\"]): element.decompose()
        root = _html_root(soup); title = normalize_text(soup.title.get_text(\" \", strip=True) if soup.title else source[\"relative_path\"])
        if not content_based_interstitial(root.get_text(\" \", strip=True), title): break
        fallback = _validated_fallback_path(source)
        if fallback_attempted or fallback is None: return [], {\"status\": \"blocked\", \"method\": \"html_cookie_interstitial\", \"reason\": \"content_based_interstitial\", \"fallback_path\": fallback_path, \"fallback_sha256\": fallback_sha256, \"fallback_source_identity\": fallback_source_identity}
        fallback_attempted = True
        html_path, fallback_hash = fallback; extraction_method = \"html_hashed_fallback\"; fallback_path = str(html_path); fallback_sha256 = fallback_hash; fallback_source_identity = stable_id(\"html-fallback-source\", fallback_path, fallback_hash); source = {**source, \"path\": str(html_path), \"fallback_path\": fallback_path, \"fallback_sha256\": fallback_hash, \"fallback_source_identity\": fallback_source_identity}
    rows, seen_hashes, section, short_element_events, empty_element_events = [], set(), title or \"document\", [], []
    for paragraph_index, element in enumerate(root.find_all([\"p\", \"li\", \"blockquote\", \"td\", \"th\", \"dt\", \"dd\", \"div\", \"section\", \"h1\", \"h2\", \"h3\", \"h4\"])):
        text = normalize_text(element.get_text(\" \", strip=True)); is_heading = element.name.startswith(\"h\")
        if not text:
            empty_element_events.append({\"filter\": \"html_element\", \"kept\": False, \"reason\": \"html_empty_element\", \"document_uid\": source.get(\"document_uid\"), \"relative_path\": source.get(\"relative_path\"), \"locator\": {\"paragraph\": paragraph_index, \"tag\": element.name}})
            continue
        if len(text) < 20:
            short_element_events.append({\"filter\": \"html_element\", \"kept\": False, \"reason\": \"html_short_element\", \"document_uid\": source.get(\"document_uid\"), \"relative_path\": source.get(\"relative_path\"), \"locator\": {\"paragraph\": paragraph_index, \"tag\": element.name}, \"text_length\": len(text)})
            if is_heading and not re.search(r\"[.!?]\", text): section = text
            continue
        if is_heading: section = text
        rows.append({**source, \"title\": title, \"page\": None, \"section\": section, \"paragraph\": paragraph_index, \"text\": text, \"extraction_method\": extraction_method, \"quality_label\": \"native\", \"parent_document_uid\": None})
    rows, dedup_events = deduplicate_contained_text_rows(rows, return_audit=True)
    html_filter_events = empty_element_events + short_element_events + dedup_events
    linked_pdf_rows = []
    for link in root.find_all(\"a\", href=True):
        href = str(link.get(\"href\", \"\")); candidate = (Path(source[\"path\"]).parent / href.split(\"?\", 1)[0]).resolve()
        if not href.casefold().endswith(\".pdf\") or not candidate.is_file(): continue
        child_bytes = candidate.read_bytes(); child_hash = hashlib.sha256(child_bytes).hexdigest(); child_source = {**source, \"document_uid\": stable_id(\"linked-pdf\", source[\"document_uid\"], child_hash), \"content_sha256\": child_hash, \"path\": str(candidate), \"relative_path\": str(candidate), \"parent_document_uid\": source[\"document_uid\"]}
        child_rows, _ = extract_pdf_document(child_source); linked_pdf_rows.extend([{**row, \"parent_document_uid\": source[\"document_uid\"], \"extraction_method\": \"linked_pdf_native\"} for row in child_rows])
    return rows + linked_pdf_rows, {\"status\": \"extracted\" if rows or linked_pdf_rows else \"empty\", \"method\": extraction_method, \"reason\": None, \"linked_pdf_count\": len(linked_pdf_rows), \"fallback_path\": fallback_path, \"fallback_sha256\": fallback_sha256, \"fallback_source_identity\": fallback_source_identity, \"html_candidate_count\": len(rows) + len(html_filter_events), \"html_empty_element_count\": len(empty_element_events), \"html_short_element_events\": short_element_events, \"html_short_element_count\": len(short_element_events), \"html_filter_events\": html_filter_events, \"html_filter_kept_count\": len(rows)}
def extract_pdf_document(source):
    rows = []; empty_page_events = []
    with fitz.open(source[\"path\"]) as pdf:
        pdf_candidate_count = len(pdf)
        for page_no, page in enumerate(pdf, 1):
            text = normalize_text(page.get_text(\"text\"))
            if text: rows.append({**source, \"page\": page_no, \"section\": \"page\", \"paragraph\": None, \"text\": text, \"extraction_method\": \"pdf_native\", \"quality_label\": \"native\"})
            else: empty_page_events.append({\"filter\": \"pdf_page\", \"kept\": False, \"reason\": \"pdf_empty_page\", \"document_uid\": source.get(\"document_uid\"), \"relative_path\": source.get(\"relative_path\"), \"locator\": {\"page\": page_no}})
    return rows, {\"status\": \"extracted\" if rows else \"empty\", \"method\": \"pdf_native\", \"reason\": None, \"pdf_candidate_count\": pdf_candidate_count, \"pdf_empty_page_count\": len(empty_page_events), \"pdf_filter_events\": empty_page_events, \"pdf_filter_kept_count\": len(rows)}
pages_rows = []
extraction_audit = []
for source in tqdm(source_registry.to_dict(\"records\"), desc=\"Extract PDF/HTML corpus\"):
    try:
        source_rows, outcome = extract_html_document(source) if source[\"document_type\"] in {\"html\", \"htm\"} else extract_pdf_document(source)
        pages_rows.extend(source_rows); extraction_audit.append({\"document_uid\": source[\"document_uid\"], \"content_sha256\": source[\"content_sha256\"], \"relative_path\": source[\"relative_path\"], \"method\": outcome[\"method\"], \"status\": outcome[\"status\"], \"text_length\": sum(len(row[\"text\"]) for row in source_rows), \"row_count\": len(source_rows), \"chunk_count\": 0, \"reason\": outcome.get(\"reason\"), \"fallback_path\": outcome.get(\"fallback_path\"), \"fallback_sha256\": outcome.get(\"fallback_sha256\"), \"html_candidate_count\": int(outcome.get(\"html_candidate_count\", 0)), \"html_empty_element_count\": int(outcome.get(\"html_empty_element_count\", 0)), \"html_short_element_count\": int(outcome.get(\"html_short_element_count\", 0)), \"html_filter_kept_count\": int(outcome.get(\"html_filter_kept_count\", 0)), \"filter_events\": outcome.get(\"html_filter_events\", outcome.get(\"pdf_filter_events\", [])), \"pdf_candidate_count\": int(outcome.get(\"pdf_candidate_count\", 0)), \"pdf_empty_page_count\": int(outcome.get(\"pdf_empty_page_count\", 0)), \"pdf_filter_kept_count\": int(outcome.get(\"pdf_filter_kept_count\", 0)), \"extractor_revision\": EXTRACTOR_REVISION})
        extraction_audit[-1].update({"fallback_source_identity": outcome.get("fallback_source_identity"), "source_identity": stable_id("source-identity", source["document_uid"], source["content_sha256"], outcome.get("fallback_source_identity"))})
    except Exception as exc:
        log_event(\"extract\", \"source_failed\", document_uid=source[\"document_uid\"], error=str(exc)); extraction_audit.append({\"document_uid\": source[\"document_uid\"], \"content_sha256\": source[\"content_sha256\"], \"relative_path\": source[\"relative_path\"], \"method\": \"failed\", \"status\": \"failed\", \"text_length\": 0, \"row_count\": 0, \"chunk_count\": 0, \"reason\": str(exc), \"filter_events\": [], \"extractor_revision\": EXTRACTOR_REVISION})
pages = pd.DataFrame(pages_rows)
chunk_rows = []; corpus_filter_rows = []
SOURCE_TEXT_BY_KEY = {}
for page in pages.to_dict(\"records\"):
    if len(page[\"text\"]) < 80:
        corpus_filter_rows.append({\"filter\": \"page\", \"kept\": False, \"reason\": \"page_short_text\", \"document_uid\": page.get(\"document_uid\"), \"relative_path\": page.get(\"relative_path\"), \"locator\": {\"page\": page.get(\"page\"), \"paragraph\": page.get(\"paragraph\")}, \"text_length\": len(page[\"text\"])})
        continue
    corpus_filter_rows.append({\"filter\": \"page\", \"kept\": True, \"reason\": None, \"document_uid\": page.get(\"document_uid\"), \"relative_path\": page.get(\"relative_path\"), \"locator\": {\"page\": page.get(\"page\"), \"paragraph\": page.get(\"paragraph\")}, \"text_length\": len(page[\"text\"])})
    page_value = canonical_page(page.get(\"page\")); paragraph = page.get(\"paragraph\")
    source_text_key = stable_id(\"source-text\", page[\"document_uid\"], page_value, paragraph); SOURCE_TEXT_BY_KEY[source_text_key] = str(page[\"text\"]); source_text_sha256 = hashlib.sha256(str(page[\"text\"]).encode(\"utf-8\")).hexdigest()
    for index, window in enumerate(split_windows(page[\"text\"], CONFIG[\"max_chunk_chars\"], CONFIG[\"chunk_overlap_chars\"])):
        text = window[\"text\"]; text_sha256 = hashlib.sha256(text.encode()).hexdigest(); chunk_id = stable_id(\"chunk\", page[\"document_uid\"], page_value, paragraph, index, text_sha256)
        chunk_rows.append({**page, \"page\": page_value, \"chunk_index\": index, \"chunk_id\": chunk_id, \"chunk_document_uid\": page[\"document_uid\"], \"document_sha256\": page[\"content_sha256\"], \"text_sha256\": text_sha256, \"text\": text, \"source_text_key\": source_text_key, \"source_text_sha256\": source_text_sha256, \"span_start\": window[\"start_char\"], \"span_end\": window[\"end_char\"], \"sentence_start\": window[\"sentence_start\"], \"sentence_end\": window[\"sentence_end\"], \"sentence_aligned\": window[\"sentence_aligned\"], \"split_reason\": window[\"split_reason\"], \"content_warning\": page[\"source_type\"] == \"harmful_examples\"})
chunks = pd.DataFrame(chunk_rows)
chunk_counts = chunks.groupby(\"document_uid\").size().to_dict() if not chunks.empty else {}
for audit in extraction_audit: audit[\"chunk_count\"] = int(chunk_counts.get(audit[\"document_uid\"], 0)); audit[\"indexable\"] = audit[\"chunk_count\"] > 0
html_events = [event for audit in extraction_audit for event in audit.get(\"filter_events\", []) if event.get(\"filter\") == \"html_element\"]
pdf_events = [event for audit in extraction_audit for event in audit.get(\"filter_events\", []) if event.get(\"filter\") == \"pdf_page\"]
def filter_manifest_row(name, input_count, kept_count, events, kept_rows=None):
    all_rows = list(events) + list(kept_rows or [])
    reasons = {}
    for event in events:
        reasons[event[\"reason\"]] = reasons.get(event[\"reason\"], 0) + 1
    return {\"filter\": name, \"input\": int(input_count), \"kept\": int(kept_count), \"dropped\": int(input_count - kept_count), \"reason_counts\": reasons, \"rows\": all_rows}
corpus_filter_manifest = {\"filters\": [filter_manifest_row(\"html_element\", sum(int(audit.get(\"html_candidate_count\", 0)) for audit in extraction_audit), sum(int(audit.get(\"html_filter_kept_count\", 0)) for audit in extraction_audit), html_events), filter_manifest_row(\"pdf_page\", sum(int(audit.get(\"pdf_candidate_count\", 0)) for audit in extraction_audit), sum(int(audit.get(\"pdf_filter_kept_count\", 0)) for audit in extraction_audit), pdf_events), filter_manifest_row(\"page\", len(corpus_filter_rows), sum(bool(row[\"kept\"]) for row in corpus_filter_rows), [row for row in corpus_filter_rows if not row[\"kept\"]], [row for row in corpus_filter_rows if row[\"kept\"]])], \"source_registry_audit_events\": list(getattr(source_registry_list, \"audit_events\", [])), \"source_registry_validation_errors\": list(getattr(source_registry_list, \"validation_errors\", [])), \"retrieval_corpus_identity_excludes_ignored_files\": True}
AUDIT_MANIFEST_HASH = stable_id(\"audit-manifest.v1\", json.dumps(corpus_filter_manifest, ensure_ascii=False, sort_keys=True, default=str))
write_json(RUN / \"artifacts\" / \"corpus_filter_manifest.json\", {**corpus_filter_manifest, \"audit_manifest_hash\": AUDIT_MANIFEST_HASH})
factual_documents = {str(row[\"document_uid\"]): row for row in source_registry[source_registry.factual_index_allowed == True].to_dict(\"records\")}
missing_factual_documents = [dict(audit) for audit in extraction_audit if audit[\"document_uid\"] in factual_documents and audit[\"chunk_count\"] == 0]
extraction_audit_hash = stable_id(EXTRACTOR_REVISION, AUDIT_MANIFEST_HASH, json.dumps(extraction_audit, ensure_ascii=False, sort_keys=True, default=str))
healthy_factual_documents = sorted(set(factual_documents) - {row[\"document_uid\"] for row in missing_factual_documents})
factual_document_coverage = len(healthy_factual_documents) / max(1, len(factual_documents))
factual_availability_gate = {\"healthy_count\": len(healthy_factual_documents), \"expected_count\": len(factual_documents), \"coverage\": factual_document_coverage, \"minimum_healthy_count\": int(CONFIG[\"minimum_healthy_factual_documents\"]), \"minimum_coverage\": float(CONFIG[\"minimum_factual_document_coverage\"]), \"pass\": len(healthy_factual_documents) >= int(CONFIG[\"minimum_healthy_factual_documents\"]) and factual_document_coverage >= float(CONFIG[\"minimum_factual_document_coverage\"])}
write_json(RUN / \"artifacts\" / \"extraction_audit.json\", {\"extractor_revision\": EXTRACTOR_REVISION, \"extraction_audit_hash\": extraction_audit_hash, \"audit_manifest_hash\": AUDIT_MANIFEST_HASH, \"rows\": extraction_audit, \"expected_factual_documents\": sorted(factual_documents), \"extracted_factual_documents\": healthy_factual_documents, \"missing_factual_documents\": missing_factual_documents, \"blocked_documents\": [row for row in extraction_audit if row[\"status\"] == \"blocked\"], \"factual_availability_gate\": factual_availability_gate})
if not factual_availability_gate[\"pass\"]:
    raise RuntimeError(\"insufficient_factual_corpus:\" + json.dumps({\"factual_availability_gate\": factual_availability_gate, \"missing_factual_documents\": missing_factual_documents, \"message\": \"Too few locally verified factual documents remain after fail-closed extraction.\"}, ensure_ascii=False, sort_keys=True))
if chunks.empty: raise RuntimeError(\"no_indexable_chunks\")
assert chunks.chunk_id.is_unique and chunks.chunk_document_uid.notna().all(); CHUNK_MANIFEST_HASH = stable_id(json.dumps(chunks[[\"document_uid\", \"chunk_id\", \"text_sha256\"]].to_dict(\"records\"), sort_keys=True))
EXPECTED_EXTRACTION_UNIVERSE = sorted(
    [
        {\"document_uid\": str(row[\"document_uid\"]), \"chunk_id\": str(row[\"chunk_id\"]), \"text_sha256\": str(row[\"text_sha256\"])}
        for row in chunks[chunks.factual_index_allowed].to_dict(\"records\")
    ],
    key=lambda row: (row[\"document_uid\"], row[\"chunk_id\"], row[\"text_sha256\"]),
)
EXPECTED_EXTRACTION_UNIVERSE_HASH = stable_id(json.dumps(EXPECTED_EXTRACTION_UNIVERSE, ensure_ascii=False, sort_keys=True))
EXPECTED_MENTION_UNIVERSE = sorted(
    [
        {\"document_uid\": str(row[\"document_uid\"]), \"chunk_id\": str(row[\"chunk_id\"]), \"text_sha256\": str(row[\"text_sha256\"]), \"text\": str(row[\"text\"])}
        for row in chunks[chunks.factual_index_allowed].to_dict(\"records\")
    ],
    key=lambda row: (row[\"document_uid\"], row[\"chunk_id\"], row[\"text_sha256\"]),
)
EXPECTED_MENTION_UNIVERSE_HASH = stable_id(json.dumps([{key: row[key] for key in (\"document_uid\", \"chunk_id\", \"text_sha256\")} for row in EXPECTED_MENTION_UNIVERSE], ensure_ascii=False, sort_keys=True))
parquet_safe_frame(pages).to_parquet(RUN / \"artifacts\" / \"pages.parquet\", index=False); parquet_safe_frame(chunks).to_parquet(RUN / \"artifacts\" / \"chunks.parquet\", index=False)
write_json(RUN / \"artifacts\" / \"extraction_manifest.json\", {\"identity\": {\"corpus_manifest_hash\": CORPUS_MANIFEST_HASH, \"chunk_manifest_hash\": CHUNK_MANIFEST_HASH, \"extraction_audit_hash\": extraction_audit_hash, \"extractor_revision\": EXTRACTOR_REVISION, \"core_source_sha256\": CORE_SOURCE_SHA256, \"schema_revision\": SEMANTIC_EXTRACTION_SCHEMA[\"version\"]}})
"""),
code("""# 06 - One Qwen lifecycle for validated semantic extraction and full query signatures.
from unsloth import FastModel
import transformers
import transformers.tokenization_utils as transformers_tokenization_utils
if not hasattr(transformers_tokenization_utils, "PreTrainedTokenizerBase"):
    transformers_tokenization_utils.PreTrainedTokenizerBase = transformers.PreTrainedTokenizerBase
from lmformatenforcer import JsonSchemaParser
from lmformatenforcer.integrations.transformers import build_transformers_prefix_allowed_tokens_fn
def verify_effective_4bit(model):
    diagnostics = {\"requested\": bool(CONFIG[\"load_in_4bit\"]), \"is_loaded_in_4bit\": bool(getattr(model, \"is_loaded_in_4bit\", False)), \"quantization_config\": str(getattr(getattr(model, \"config\", None), \"quantization_config\", None)), \"quantized_module_count\": 0, \"quantized_parameter_count\": 0, \"float32_parameter_count\": 0, \"total_parameter_count\": 0, \"float32_parameter_fraction\": 0.0}
    for module in model.modules():
        name = type(module).__name__.casefold()
        if \"4bit\" in name or \"bitsandbytes\" in name or \"bnb\" in name or \"params4bit\" in name: diagnostics[\"quantized_module_count\"] += 1
    for parameter in model.parameters():
        count = int(parameter.numel()) if hasattr(parameter, \"numel\") else 1; diagnostics[\"total_parameter_count\"] += count
        parameter_type = type(parameter).__name__.casefold()
        if \"params4bit\" in parameter_type or \"4bit\" in parameter_type or \"int8params\" in parameter_type: diagnostics[\"quantized_parameter_count\"] += count
        elif str(getattr(parameter, \"dtype\", \"\")).casefold().endswith(\"float32\"): diagnostics[\"float32_parameter_count\"] += count
    if diagnostics[\"total_parameter_count\"]: diagnostics[\"float32_parameter_fraction\"] = diagnostics[\"float32_parameter_count\"] / diagnostics[\"total_parameter_count\"]
    # Metadata flags/config strings are diagnostics only. Require inspected
    # quantized modules and reject a meaningful all-fp32 fallback; small fp32
    # norms/heads are allowed.
    diagnostics[\"effective\"] = bool(diagnostics[\"quantized_module_count\"] > 0 and diagnostics[\"float32_parameter_fraction\"] <= 0.25)
    if diagnostics[\"requested\"] and not diagnostics[\"effective\"]: raise RuntimeError(\"effective_4bit_verification_failed:\" + json.dumps(diagnostics, sort_keys=True))
    return diagnostics
def load_qwen_for_extraction():
    global model, tokenizer
    model, tokenizer = FastModel.from_pretrained(model_name=CONFIG[\"generator_model\"], revision=CONFIG[\"generator_model_revision\"], max_seq_length=CONFIG[\"max_seq_length\"], load_in_4bit=CONFIG[\"load_in_4bit\"], dtype=None, full_finetuning=False); verify_effective_4bit(model); tokenizer.pad_token = tokenizer.pad_token or tokenizer.eos_token; tokenizer.padding_side = \"left\"; model.eval(); MEMORY_SNAPSHOTS.append(gpu_snapshot(\"qwen_extraction_loaded\")); return model, tokenizer
def unload_generator():
    global model, tokenizer
    if \"model\" in globals(): del model
    if \"tokenizer\" in globals(): del tokenizer
    gc.collect(); torch.cuda.empty_cache(); MEMORY_SNAPSHOTS.append(gpu_snapshot(\"qwen_extraction_unloaded\"))
def split_visible_thinking(raw):
    splitter = globals().get(\"split_qwen_thinking\")
    if callable(splitter): return splitter(raw)[\"final_content\"]
    value = str(raw or \"\"); match = re.search(r\"<think>.*?</think>\", value, flags=re.S | re.I)
    return (value[0 : match.start()] + value[match.end() :]).strip() if match else value.strip()
def qwen_generation_trace(raw):
    trace = dict(split_qwen_thinking(raw)); base_tokenizer = getattr(globals().get(\"tokenizer\"), \"tokenizer\", globals().get(\"tokenizer\"))
    def token_count(value):
        if base_tokenizer is None or not hasattr(base_tokenizer, \"encode\"): return None
        return len(base_tokenizer.encode(str(value or \"\"), add_special_tokens=False))
    trace[\"reasoning_token_count\"] = token_count(trace[\"reasoning_content\"]); trace[\"answer_token_count\"] = token_count(trace[\"final_content\"]); trace[\"raw_generation\"] = str(raw or \"\"); return trace
def _apply_chat_template(messages, enable_thinking):
    try:
        return tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True, enable_thinking=enable_thinking)
    except TypeError as exc:
        log_event(\"generation\", \"chat_template_compatibility_fallback\", error=str(exc))
        return tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
class GenerationRowQuarantine(RuntimeError):
    def __init__(self, audit):
        self.audit = dict(audit or {})
        self.reason = str(self.audit.get('reason', 'row_quarantine'))
        super().__init__(self.reason)
class PromptBudgetQuarantine(GenerationRowQuarantine):
    def __init__(self, audit):
        super().__init__(audit)
        self.args = (f\"prompt_token_budget_exceeded:prompt_budget_irreducible:{self.reason}\",)
class ParseRateQuarantine(GenerationRowQuarantine):
    pass
class SchemaValidationQuarantine(GenerationRowQuarantine):
    pass
def preflight_prompt_token_budget(prompts, max_new_tokens, schema_tail=None):
    if callable(globals().get(\"fit_prompt_to_budget\")):
        for prompt in prompts:
            fit_prompt_to_budget(prompt, tokenizer, int(CONFIG[\"max_seq_length\"]), schema_tail=schema_tail, reserve_output_tokens=int(max_new_tokens))
    encoded = tokenizer(text=prompts, return_tensors=\"pt\", padding=True, truncation=False)
    mask = encoded[\"attention_mask\"]; lengths = mask.sum(dim=1).tolist() if hasattr(mask, \"sum\") else [sum(int(value) for value in row) for row in mask]
    violations = [{\"index\": i, \"prompt_tokens\": int(length), \"max_new_tokens\": int(max_new_tokens), \"total_tokens\": int(length) + int(max_new_tokens), \"max_seq_length\": int(CONFIG[\"max_seq_length\"])} for i, length in enumerate(lengths) if int(length) + int(max_new_tokens) > int(CONFIG[\"max_seq_length\"])]
    if violations:
        audit = {\"reason\": \"prompt_budget_irreducible\", \"violations\": violations, \"dropped_evidence_ids\": [], \"attempts\": []}
        raise PromptBudgetQuarantine(audit) if \"PromptBudgetQuarantine\" in globals() else RuntimeError(\"prompt_token_budget_exceeded:\" + json.dumps(violations, sort_keys=True))
    return encoded, [int(length) for length in lengths]
def _generate_prompt_batch(prompts, max_new_tokens, temperature=0.0, output_schema=None):
    schema_tail = json.dumps(output_schema, ensure_ascii=False, sort_keys=True) if output_schema is not None else None
    if schema_tail and all(schema_tail in prompt for prompt in prompts): schema_tail = None
    encoded, prompt_lengths = preflight_prompt_token_budget(prompts, max_new_tokens, schema_tail=schema_tail); encoded = encoded.to(model.device); base_tokenizer = getattr(tokenizer, \"tokenizer\", tokenizer)
    generation_kwargs = {\"max_new_tokens\": max_new_tokens, \"do_sample\": temperature > 0, \"temperature\": max(temperature, 1e-5), \"pad_token_id\": getattr(base_tokenizer, \"pad_token_id\", getattr(tokenizer, \"pad_token_id\", None))}
    if output_schema is not None:
        parser = JsonSchemaParser(output_schema); generation_kwargs[\"prefix_allowed_tokens_fn\"] = build_transformers_prefix_allowed_tokens_fn(base_tokenizer, parser)
    with torch.inference_mode(): output = model.generate(**encoded, **generation_kwargs)
    n = encoded[\"input_ids\"].shape[1]; return [base_tokenizer.decode(row[n:], skip_special_tokens=True) for row in output]
def generate_batch(messages_batch, max_new_tokens, temperature=0.0, enable_thinking=False, output_schema=None):
    if not enable_thinking:
        prompts = [_apply_chat_template(x, enable_thinking=False) for x in messages_batch]
        return _generate_prompt_batch(prompts, max_new_tokens, temperature=temperature, output_schema=output_schema)
    reasoning_prompts = [_apply_chat_template(x, enable_thinking=True) for x in messages_batch]
    reasoning_outputs = _generate_prompt_batch(reasoning_prompts, int(CONFIG[\"reasoning_max_new_tokens\"]), temperature=temperature)
    final_messages = []
    for messages, reasoning in zip(messages_batch, reasoning_outputs):
        final_messages.append(list(messages) + [{\"role\": \"user\", \"content\": f\"Prior model deliberation (use as private working context):\\n{reasoning.strip()}\\nNow return only the requested final JSON object.\"}])
    final_prompts = [_apply_chat_template(x, enable_thinking=False) for x in final_messages]
    final_outputs = _generate_prompt_batch(final_prompts, max_new_tokens, temperature=temperature, output_schema=output_schema)
    return [f\"<think>\\n{reasoning.strip()}\\n</think>\\n\\n{final.strip()}\" for reasoning, final in zip(reasoning_outputs, final_outputs)]
def _repair_payload_candidates(raw, prompt):
    raw_text = split_visible_thinking(raw)
    try:
        semantic_raw = json.dumps(json.loads(raw_text), ensure_ascii=False, sort_keys=True)
    except Exception:
        semantic_raw = raw_text
    full = json.dumps({\"raw_output\": semantic_raw, \"schema_task\": str(prompt)}, ensure_ascii=False, sort_keys=True)
    # The compact candidate must retain the original model output. A status-only
    # payload would permit the repair model to invent an unrelated object.
    compact = json.dumps({\"raw_output\": semantic_raw, \"schema_task\": \"Repair this prior output to the requested JSON schema without adding commentary.\"}, ensure_ascii=False, sort_keys=True)
    return (full, compact)
def _repair_prompt_token_count(prompt):
    encoded = tokenizer(text=[str(prompt)], return_tensors=\"pt\", padding=False, truncation=False)
    mask = encoded[\"attention_mask\"]
    return int(mask.sum().item()) if hasattr(mask.sum(), \"item\") else int(mask.sum())
def _adaptive_repair_messages(raw, prompt, output_schema):
    payload_candidates = _repair_payload_candidates(raw, prompt)
    schema_tail = \"JSON schema (must be followed exactly): \" + json.dumps(output_schema, ensure_ascii=False, sort_keys=True)
    def builder(selected, payload):
        return [{\"role\": \"system\", \"content\": \"Return only valid JSON matching the requested schema.\"}, {\"role\": \"user\", \"content\": f\"Repair context: {payload}\\n{schema_tail}\"}]
    def rendered_builder(selected, payload):
        return _apply_chat_template(builder(selected, payload), enable_thinking=False)
    fit = fit_adaptive_prompt_with_evidence([], prompt_builder=rendered_builder, payload_candidates=payload_candidates, token_counter=_repair_prompt_token_count, max_input_tokens=int(CONFIG[\"max_seq_length\"]), reserve_output_tokens=512)
    if fit.get(\"status\") != \"fit\":
        audit = dict(fit); audit[\"reason\"] = \"repair_prompt_budget\"; raise PromptBudgetQuarantine(audit)
    return builder([], fit.get(\"payload\", \"\")), fit
def parse_json_object(raw):
    value = split_visible_thinking(raw); start, end = value.find(\"{\"), value.rfind(\"}\")
    if start < 0 or end <= start: raise ValueError(\"invalid_json\")
    return json.loads(value[start:end + 1])
def repair_json_output(raw, prompt, output_schema):
    adaptive_builder = globals().get("_adaptive_repair_messages")
    if callable(adaptive_builder):
        messages, _ = adaptive_builder(raw, prompt, output_schema)
    else:
        # Keep lightweight extraction/probe namespaces compatible while still
        # retaining the invalid output and immutable schema tail.
        schema_tail = "JSON schema (must be followed exactly): " + json.dumps(output_schema, ensure_ascii=False, sort_keys=True)
        repair_context = json.dumps({"raw_output": split_visible_thinking(raw), "schema_task": str(prompt)}, ensure_ascii=False, sort_keys=True)
        messages = [{"role": "system", "content": "Return only valid JSON matching the requested schema."}, {"role": "user", "content": f"Repair context: {repair_context}\\n{schema_tail}"}]
    return generate_batch([messages], 512, enable_thinking=False, output_schema=output_schema)[0]
def parse_with_one_repair(raw, prompt, validator=None, output_schema=None):
    try:
        parsed = parse_json_object(raw)
        if validator is not None and not validator(parsed)[\"valid\"]: raise ValueError(\"schema_invalid\")
        return parsed, raw, \"initial\"
    except Exception:
        repaired = repair_json_output(raw, prompt, output_schema)
        try:
            parsed = parse_json_object(repaired)
            if validator is not None and not validator(parsed)[\"valid\"]: raise ValueError(\"schema_invalid\")
            original_trace = split_qwen_thinking(raw); repaired_final = split_qwen_thinking(repaired)[\"final_content\"]
            preserved = (f\"<think>\\n{original_trace['reasoning_content']}\\n</think>\\n\\n\" + repaired_final) if original_trace[\"reasoning_content\"] else repaired_final
            return parsed, preserved, \"repair\"
        except Exception: return None, repaired, \"schema_invalid\"
QUERY_SCHEMA_REVISION = \"query-signature.v2\"
# The derived cache remains validated against QUERY_SIGNATURE_SCHEMA; model decoding uses QUERY_MODEL_SCHEMA (output_schema=QUERY_SIGNATURE_SCHEMA is reserved for the derived cache contract).
ACCEPTED_PARSE_STATUSES = frozenset({\"initial\", \"repair\"})
def canonical_extraction_chunks(rows, expected_universe):
    expected_order = sorted(
        [(str(item[\"document_uid\"]), str(item[\"chunk_id\"]), str(item[\"text_sha256\"])) for item in expected_universe],
        key=lambda key: key,
    )
    ordered = sorted(
        rows,
        key=lambda row: (str(row[\"document_uid\"]), str(row[\"chunk_id\"]), str(row[\"text_sha256\"])),
    )
    actual_order = [(str(row.get(\"document_uid\")), str(row.get(\"chunk_id\")), str(row.get(\"text_sha256\"))) for row in ordered]
    if actual_order != expected_order or len(actual_order) != len(set(actual_order)): raise RuntimeError(\"extraction_universe_mismatch\")
    return ordered
def parse_extraction_with_one_repair(raw, prompt, source_text, context):
    \"\"\"Parse and semantically validate extraction with one total repair budget.\"\"\"
    try:
        payload = parse_json_object(raw)
    except Exception:
        repaired = repair_json_output(raw, prompt, SEMANTIC_EXTRACTION_SCHEMA)
        try:
            repaired_payload = parse_json_object(repaired)
        except Exception:
            return None, repaired, \"schema_invalid\", validate_extraction(None, source_text, context)
        repaired_validation = validate_extraction(repaired_payload, source_text, context)
        if repaired_validation.get(\"status\") == \"quarantined\" or repaired_validation.get(\"quarantined\"):
            return repaired_payload, repaired, \"semantic_invalid\", repaired_validation
        return repaired_payload, repaired, \"repair\", repaired_validation
    validation = validate_extraction(payload, source_text, context)
    if validation.get(\"status\") == \"quarantined\" or validation.get(\"quarantined\"):
        reasons = sorted({reason for item in validation.get(\"quarantined\", []) for reason in item.get(\"reasons\", [])})
        semantic_repair_prompt = f\"{prompt}\\nSemantic validation reasons: {json.dumps(reasons, ensure_ascii=False)}\\nExact source text: {source_text}\\nSource context: {json.dumps(context, ensure_ascii=False, sort_keys=True)}\"
        repaired = repair_json_output(raw, semantic_repair_prompt, SEMANTIC_EXTRACTION_SCHEMA)
        try:
            repaired_payload = parse_json_object(repaired)
        except Exception:
            return None, repaired, \"semantic_invalid\", validate_extraction(None, source_text, context)
        repaired_validation = validate_extraction(repaired_payload, source_text, context)
        if repaired_validation.get(\"status\") == \"quarantined\" or repaired_validation.get(\"quarantined\"):
            return repaired_payload, repaired, \"semantic_invalid\", repaired_validation
        return repaired_payload, repaired, \"repair\", repaired_validation
    return payload, raw, \"initial\", validation
def parse_mentions_with_one_repair(raw, prompt, source_text, context):
    parsed, final_raw, status = parse_with_one_repair(raw, prompt, None, MENTION_DISCOVERY_SCHEMA)
    validation = validate_mentions(parsed, source_text, context) if parsed is not None else validate_mentions(None, source_text, context)
    if validation.get(\"status\") != \"accepted\" or validation.get(\"quarantined\"): return [], final_raw, \"semantic_invalid\"
    return validation.get(\"accepted\", []), final_raw, status
def extraction_parse_status(parse_status, validation):
    if parse_status in ACCEPTED_PARSE_STATUSES and isinstance(validation, dict) and (validation.get(\"status\") == \"quarantined\" or validation.get(\"quarantined\")):
        return \"semantic_invalid\"
    return parse_status
def extraction_parse_accepted(parse_status, validation):
    return parse_status in ACCEPTED_PARSE_STATUSES and isinstance(validation, dict) and validation.get(\"status\") != \"quarantined\" and not validation.get(\"quarantined\")
PERSPECTIVE_SCHEMA = {\"type\": \"object\", \"additionalProperties\": False, \"required\": [\"perspective\", \"rationale\", \"claims_to_address\", \"supported_evidence_ids\", \"response_guidance\", \"risk_flags\", \"confidence\"], \"properties\": {\"perspective\": {\"type\": \"string\"}, \"rationale\": {\"type\": \"string\"}, \"claims_to_address\": {\"type\": \"array\", \"items\": {\"type\": \"string\"}}, \"supported_evidence_ids\": {\"type\": \"array\", \"items\": {\"type\": \"string\", \"pattern\": \"^E[1-9][0-9]*$\"}}, \"response_guidance\": {\"type\": \"array\", \"items\": {\"type\": \"string\"}}, \"risk_flags\": {\"type\": \"array\", \"items\": {\"type\": \"string\"}}, \"confidence\": {\"type\": \"number\", \"minimum\": 0, \"maximum\": 1}}}
QUERY_SIGNATURE_SCHEMA = {\"type\": \"object\", \"additionalProperties\": False, \"required\": [\"entity_ids\", \"predicates\", \"polarities\", \"modalities\", \"desired_stances\"], \"properties\": {key: {\"type\": \"array\", \"items\": {\"type\": \"string\"}} for key in [\"entity_ids\", \"predicates\", \"polarities\", \"modalities\", \"desired_stances\"]}}
QUERY_MODEL_SCHEMA = {\"type\": \"object\", \"additionalProperties\": False, \"required\": [\"entity_candidate_indices\", \"predicates\", \"polarities\", \"modalities\", \"desired_stances\"], \"properties\": {\"entity_candidate_indices\": {\"type\": \"array\", \"items\": {\"type\": \"object\", \"additionalProperties\": False, \"required\": [\"mention_id\", \"candidate_index\"], \"properties\": {\"mention_id\": {\"type\": \"string\"}, \"candidate_index\": {\"type\": [\"integer\", \"null\"], \"minimum\": 0}}}}, **{key: {\"type\": \"array\", \"items\": {\"type\": \"string\"}} for key in [\"predicates\", \"polarities\", \"modalities\", \"desired_stances\"]}}}
PLAN_SCHEMA = {\"type\": \"object\", \"additionalProperties\": False, \"required\": [\"claim_focus\", \"selected_evidence_ids\", \"response_steps\", \"tone\", \"factual_constraints\", \"safety_constraints\"], \"properties\": {\"claim_focus\": {\"type\": \"string\"}, \"selected_evidence_ids\": {\"type\": \"array\", \"items\": {\"type\": \"string\", \"pattern\": \"^E[1-9][0-9]*$\"}}, \"response_steps\": {\"type\": \"array\", \"items\": {\"type\": \"string\"}}, \"tone\": {\"type\": \"string\"}, \"factual_constraints\": {\"type\": \"array\", \"items\": {\"type\": \"string\"}}, \"safety_constraints\": {\"type\": \"array\", \"items\": {\"type\": \"string\"}}}}
FINAL_RESPONSE_SCHEMA = {\"type\": \"object\", \"additionalProperties\": False, \"required\": [\"counter_narrative\", \"cited_evidence_ids\", \"factual_claims\", \"safety_notes\"], \"properties\": {\"counter_narrative\": {\"type\": \"string\"}, \"cited_evidence_ids\": {\"type\": \"array\", \"items\": {\"type\": \"string\", \"pattern\": \"^E[1-9][0-9]*$\"}}, \"factual_claims\": {\"type\": \"array\", \"items\": {\"type\": \"string\"}}, \"safety_notes\": {\"type\": \"array\", \"items\": {\"type\": \"string\"}}}}
INTERNAL_RESPONSE_METADATA_KEYS = frozenset({\"parse_status\", \"quarantine\", \"schema_errors\", \"few_shot\", \"few_shot_prompt_revision\"})
def _strict_payload(payload, schema, label, ledger_ids=None):
    reasons = []
    if not isinstance(payload, dict): return {\"valid\": False, \"reasons\": [f\"{label}_not_object\"], \"quarantine\": True}
    required, allowed = set(schema[\"required\"]), set(schema[\"properties\"]); reasons.extend(f\"missing_{key}\" for key in sorted(required - set(payload))); reasons.extend([\"additionalProperties\"] if set(payload) - allowed else [])
    for key, spec in schema[\"properties\"].items():
        if key not in payload: continue
        value = payload[key]; kind = spec[\"type\"]
        if kind == \"string\" and not isinstance(value, str): reasons.append(f\"{key}_not_string\")
        if kind == \"array\" and (not isinstance(value, list) or not all(isinstance(item, str) for item in value)): reasons.append(f\"{key}_not_string_array\")
        if kind == \"number\" and (isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(value) or value < spec.get(\"minimum\", -math.inf) or value > spec.get(\"maximum\", math.inf)): reasons.append(f\"{key}_out_of_range\")
    for key in [\"supported_evidence_ids\", \"selected_evidence_ids\", \"cited_evidence_ids\"]:
        values = payload.get(key)
        if not isinstance(values, list): continue
        if any(not isinstance(value, str) for value in values):
            reasons.append(f\"invalid_{key}\")
            continue
        if len(values) != len(set(values)): reasons.append(f\"duplicate_{key}\")
        if any(not re.fullmatch(r\"E[1-9][0-9]*\", value) for value in values): reasons.append(f\"invalid_{key}\")
        if ledger_ids is not None and set(values) - set(ledger_ids): reasons.append(\"unknown_evidence_ids\")
    return {\"valid\": not reasons, \"reasons\": sorted(set(reasons)), \"quarantine\": bool(reasons)}
def validate_query_signature(payload): return _strict_payload(payload, QUERY_SIGNATURE_SCHEMA, \"query_signature\")
def validate_query_model_signature(payload):
    if not isinstance(payload, dict) or set(payload) != set(QUERY_MODEL_SCHEMA[\"properties\"]): return {\"valid\": False, \"reasons\": [\"query_model_schema_invalid\"]}
    if not isinstance(payload.get(\"entity_candidate_indices\"), list) or any(not isinstance(item, dict) or set(item) != {\"mention_id\", \"candidate_index\"} or not isinstance(item.get(\"mention_id\"), str) or (item.get(\"candidate_index\") is not None and (isinstance(item.get(\"candidate_index\"), bool) or not isinstance(item.get(\"candidate_index\"), int) or item.get(\"candidate_index\") < 0)) for item in payload[\"entity_candidate_indices\"]): return {\"valid\": False, \"reasons\": [\"query_candidate_schema_invalid\"]}
    return _strict_payload({key: payload[key] for key in [\"predicates\", \"polarities\", \"modalities\", \"desired_stances\"]}, {\"type\": \"object\", \"additionalProperties\": False, \"required\": [\"predicates\", \"polarities\", \"modalities\", \"desired_stances\"], \"properties\": {key: {\"type\": \"array\", \"items\": {\"type\": \"string\"}} for key in [\"predicates\", \"polarities\", \"modalities\", \"desired_stances\"]}}, \"query_model_signature\")
def validate_perspective(payload, ledger_ids=None, expected_perspective=None):
    result = _strict_payload(payload, PERSPECTIVE_SCHEMA, \"perspective\", ledger_ids)
    if result[\"valid\"] and expected_perspective is not None and payload.get(\"perspective\") != expected_perspective:
        result[\"valid\"] = False; result[\"quarantine\"] = True; result[\"reasons\"].append(\"perspective_name_mismatch\")
    return result
def validate_plan(payload, ledger_ids=None): return _strict_payload(payload, PLAN_SCHEMA, \"plan\", ledger_ids)
def validate_final_response(payload, ledger_ids=None): return _strict_payload(payload, FINAL_RESPONSE_SCHEMA, \"final_response\", ledger_ids)
def validate_structured_output(payload, schema=PERSPECTIVE_SCHEMA): return _strict_payload(payload, schema, \"structured_output\")[\"valid\"]
def build_query_signature(record):
    target_candidates = build_entity_candidates(str(record.get(\"Target\", \"\")), ENTITY_CATALOG, namespace_preference=\"corpus\", allow_target_fallback=True)
    candidate_context = {\"target\": target_candidates, \"candidate_set_hash\": target_candidates[\"candidate_set_hash\"], \"catalog_manifest_hash\": CATALOG_MANIFEST_HASH}
    return [{\"role\": \"system\", \"content\": \"Return JSON only with entity_candidate_indices, predicates, polarities, modalities, desired_stances. Never emit entity IDs; select candidate indices or null.\"}, {\"role\": \"user\", \"content\": f\"Target: {record['Target']}\\nPost: {record['Text']}\\nCandidate context: {json.dumps(candidate_context, ensure_ascii=False, sort_keys=True)}\"}]
MENTION_DISCOVERY_CACHE = RUN / \"artifacts\" / \"mention_discovery.jsonl\"; MENTION_DISCOVERY_IDENTITY_PATH = RUN / \"artifacts\" / \"mention_discovery_identity.json\"; EXTRACTION_CACHE = RUN / \"artifacts\" / \"validated_extractions.jsonl\"; QUERY_SIGNATURE_CACHE = RUN / \"artifacts\" / \"query_signatures.jsonl\"
MENTION_DISCOVERY_PROMPT_REVISION = \"mention-discovery.v1-exact-spans\"
def mention_discovery_identity(expected_universe):
    ordered = [{key: row[key] for key in (\"document_uid\", \"chunk_id\", \"text_sha256\")} for row in expected_universe]
    return {\"stage\": \"mention-discovery\", \"prompt_revision\": MENTION_DISCOVERY_PROMPT_REVISION, \"schema_revision\": MENTION_DISCOVERY_SCHEMA[\"version\"], \"model\": EXTRACTION_MODEL, \"corpus_manifest_hash\": CORPUS_MANIFEST_HASH, \"expected_mention_universe\": ordered, \"expected_mention_universe_hash\": stable_id(\"mention-universe.v1\", json.dumps(ordered, ensure_ascii=False, sort_keys=True)), \"expected_mention_universe_size\": len(ordered), \"core_source_sha256\": CORE_SOURCE_SHA256}
def verify_mention_cache(rows, expected_universe, expected_identity):
    expected_identity = dict(expected_identity or {}); expected = mention_discovery_identity(expected_universe)
    if expected_identity != expected: return False
    expected_order = [(str(row[\"document_uid\"]), str(row[\"chunk_id\"]), str(row[\"text_sha256\"])) for row in expected_universe]
    required = {\"document_uid\", \"chunk_id\", \"text_sha256\", \"text\", \"mentions\", \"parse_status\", \"raw_output\"}
    if not isinstance(rows, list) or len(rows) != len(expected_order): return False
    for index, row in enumerate(rows):
        if not isinstance(row, dict) or set(row) != required: return False
        source = expected_universe[index]; identity = (str(row.get(\"document_uid\")), str(row.get(\"chunk_id\")), str(row.get(\"text_sha256\")))
        if identity != expected_order[index] or row.get(\"text\") != source.get(\"text\") or hashlib.sha256(row[\"text\"].encode()).hexdigest() != row[\"text_sha256\"]: return False
        if row.get(\"parse_status\") not in {\"initial\", \"repair\"} or not isinstance(row.get(\"raw_output\"), str) or not isinstance(row.get(\"mentions\"), list): return False
        validation = validate_mentions({\"schema_version\": MENTION_DISCOVERY_SCHEMA[\"version\"], \"mentions\": row[\"mentions\"]}, row[\"text\"], {\"document_uid\": row[\"document_uid\"], \"chunk_id\": row[\"chunk_id\"], \"text_sha256\": row[\"text_sha256\"]})
        if validation.get(\"status\") != \"accepted\" or validation.get(\"quarantined\"): return False
    return True
def mention_records_from_cache(rows):
    return [{**mention, \"document_uid\": row[\"document_uid\"], \"chunk_id\": row[\"chunk_id\"]} for row in rows for mention in row[\"mentions\"]]
def verify_extraction_cache(rows, expected_universe):
    expected_order = [
        (str(item.get(\"document_uid\")), str(item.get(\"chunk_id\")), str(item.get(\"text_sha256\")))
        for item in expected_universe
        if isinstance(item, dict)
    ]
    expected = set(expected_order)
    if not isinstance(rows, list) or len(rows) != len(expected) or len(expected) != len(expected_order): return False
    seen = set()
    allowed_statuses = {\"initial\", \"repair\", \"semantic_invalid\", \"schema_invalid\"}
    allowed_validation_statuses = {\"accepted\", \"partial\", \"reviewed\", \"quarantined\"}
    catalog_bound = \"CATALOG_MANIFEST_HASH\" in globals()
    required_row_keys = {\"document_uid\", \"chunk_id\", \"text_sha256\", \"validation\", \"parse_status\", \"raw_output\"} | ({\"catalog_manifest_hash\"} if catalog_bound else set())
    for index, row in enumerate(rows):
        if not isinstance(row, dict) or set(row) != required_row_keys: return False
        if not all(isinstance(row.get(key), str) and row.get(key) for key in (\"document_uid\", \"chunk_id\", \"text_sha256\", \"raw_output\")): return False
        if catalog_bound and row.get(\"catalog_manifest_hash\") != CATALOG_MANIFEST_HASH: return False
        identity = (row[\"document_uid\"], row[\"chunk_id\"], row[\"text_sha256\"])
        if identity in seen or identity not in expected or identity != expected_order[index] or row.get(\"parse_status\") not in allowed_statuses: return False
        validation = row.get(\"validation\")
        if not isinstance(validation, dict) or validation.get(\"status\") not in allowed_validation_statuses or not isinstance(validation.get(\"quarantined\", []), list): return False
        source_context = validation.get(\"source_context\")
        if not isinstance(source_context, dict) or any(source_context.get(key) != row[key] for key in (\"document_uid\", \"chunk_id\", \"text_sha256\")): return False
        if not isinstance(validation.get(\"text\"), str) or hashlib.sha256(validation[\"text\"].encode()).hexdigest() != row[\"text_sha256\"]: return False
        if validation.get(\"validation_marker\") != \"mpkg-rag.validated-extraction.v1\" or not isinstance(validation.get(\"validation_fingerprint\"), str) or not validation.get(\"validation_fingerprint\") or validation.get(\"validation_fingerprint\") != _validation_fingerprint(validation): return False
        seen.add(identity)
    return seen == expected
mention_identity = mention_discovery_identity(EXPECTED_MENTION_UNIVERSE); saved_mention_identity = json.loads(MENTION_DISCOVERY_IDENTITY_PATH.read_text()) if MENTION_DISCOVERY_IDENTITY_PATH.exists() else {}; mention_rows = load_jsonl(MENTION_DISCOVERY_CACHE); mention_cache_complete = verify_mention_cache(mention_rows, EXPECTED_MENTION_UNIVERSE, saved_mention_identity.get(\"identity\")); need_mentions = not (saved_mention_identity.get(\"identity\") == mention_identity and mention_cache_complete); model_loaded = False
if need_mentions:
    load_qwen_for_extraction(); model_loaded = True
    extraction_chunks = canonical_extraction_chunks(chunks[chunks.factual_index_allowed].to_dict(\"records\"), EXPECTED_EXTRACTION_UNIVERSE)
    with MENTION_DISCOVERY_CACHE.open(\"w\", encoding=\"utf-8\") as stream:
        for start in tqdm(range(0, len(extraction_chunks), CONFIG[\"extraction_batch_size\"]), desc=\"Qwen mention discovery\"):
            batch = extraction_chunks[start:start + CONFIG[\"extraction_batch_size\"]]; base_contexts = [{k: chunk.get(k) for k in [\"document_uid\", \"chunk_id\", \"source_id\", \"source_type\", \"authority_score\", \"factual_index_allowed\", \"status\", \"content_sha256\", \"document_sha256\", \"text_sha256\"]} for chunk in batch]; mention_prompts = [build_mention_prompt(chunk[\"text\"], context) for chunk, context in zip(batch, base_contexts)]; mention_raw = generate_batch(mention_prompts, 384, enable_thinking=False, output_schema=MENTION_DISCOVERY_SCHEMA)
            for chunk, context, mention_prompt, raw_mention in zip(batch, base_contexts, mention_prompts, mention_raw):
                mentions, mention_final_raw, mention_status = parse_mentions_with_one_repair(raw_mention, mention_prompt[1][\"content\"], chunk[\"text\"], context); stream.write(json.dumps({\"document_uid\": chunk[\"document_uid\"], \"chunk_id\": chunk[\"chunk_id\"], \"text_sha256\": chunk[\"text_sha256\"], \"text\": chunk[\"text\"], \"mentions\": mentions, \"parse_status\": mention_status, \"raw_output\": mention_final_raw}, ensure_ascii=False) + \"\\n\")
    mention_rows = load_jsonl(MENTION_DISCOVERY_CACHE); mention_cache_complete = verify_mention_cache(mention_rows, EXPECTED_MENTION_UNIVERSE, mention_identity); write_json_atomic(MENTION_DISCOVERY_IDENTITY_PATH, {\"identity\": mention_identity, \"cache_state\": \"complete_ready\" if mention_cache_complete else \"complete_validation_failed\"})
if not mention_cache_complete: raise RuntimeError(\"mention_cache_identity_mismatch\")
mention_records = mention_records_from_cache(mention_rows)
ENTITY_CATALOG = build_entity_catalog(source_registry.to_dict(\"records\"), dataset[\"Target\"].astype(str).tolist(), reviewed_orgs, mention_records); CATALOG_MANIFEST_HASH = ENTITY_CATALOG[\"catalog_hash\"]
write_json(RUN / \"artifacts\" / \"entity_catalog.json\", ENTITY_CATALOG); log_event(\"catalog\", \"final_catalog_frozen\", catalog_manifest_hash=CATALOG_MANIFEST_HASH, mention_identity=mention_identity)
identity = {\"corpus_manifest_hash\": CORPUS_MANIFEST_HASH, \"chunk_manifest_hash\": CHUNK_MANIFEST_HASH, \"extraction_audit_hash\": extraction_audit_hash, \"catalog_manifest_hash\": CATALOG_MANIFEST_HASH, \"mention_identity\": mention_identity, \"expected_extraction_universe_hash\": EXPECTED_EXTRACTION_UNIVERSE_HASH, \"expected_extraction_universe_size\": len(EXPECTED_EXTRACTION_UNIVERSE), \"extraction_model\": EXTRACTION_MODEL, \"extraction_prompt_revision\": EXTRACTION_PROMPT_REVISION, \"core_source_sha256\": CORE_SOURCE_SHA256, \"schema_revision\": SEMANTIC_EXTRACTION_SCHEMA[\"version\"]}; identity_path = RUN / \"artifacts\" / \"extraction_identity.json\"; saved = json.loads(identity_path.read_text()) if identity_path.exists() else {}
def record_identity(record):
    return {\"record_id\": str(record[\"ID\"]), \"input_text_sha256\": str(record[\"input_text_sha256\"]), \"target_sha256\": hashlib.sha256(str(record.get(\"Target\", \"\")).encode()).hexdigest(), \"category_sha256\": hashlib.sha256(str(record.get(\"Category\", \"\")).encode()).hexdigest()}
record_identities = sorted([record_identity(record) for record in dataset.to_dict(\"records\")], key=lambda row: row[\"record_id\"])
query_identity = {\"record_identities\": record_identities, \"model\": EXTRACTION_MODEL, \"prompt_revision\": EXTRACTION_PROMPT_REVISION, \"core_source_sha256\": CORE_SOURCE_SHA256, \"schema_revision\": QUERY_SCHEMA_REVISION, \"catalog_manifest_hash\": CATALOG_MANIFEST_HASH, \"mention_identity\": mention_identity}; extraction_rows = load_jsonl(EXTRACTION_CACHE); extraction_cache_complete = verify_extraction_cache(extraction_rows, EXPECTED_EXTRACTION_UNIVERSE)
def cache_reuse_state(saved, expected_identity, expected_query_identity, expected_universe, extraction_cache_complete, signature_cache_complete, mention_cache_complete=True, mention_identity_matches=True):
    identity_matches = saved.get(\"identity\") == expected_identity and saved.get(\"query_identity\") == expected_query_identity and saved.get(\"expected_extraction_universe\") == expected_universe and saved.get(\"expected_extraction_universe_hash\") == EXPECTED_EXTRACTION_UNIVERSE_HASH and mention_cache_complete and mention_identity_matches
    return (not (identity_matches and extraction_cache_complete), not (identity_matches and signature_cache_complete))
def extraction_identity_payload(cache_state, parse_rate, parse_status_counts):
    return {\"identity\": identity, \"query_identity\": query_identity, \"mention_identity\": mention_identity, \"expected_extraction_universe\": EXPECTED_EXTRACTION_UNIVERSE, \"expected_extraction_universe_hash\": EXPECTED_EXTRACTION_UNIVERSE_HASH, \"cache_state\": cache_state, \"parse_rate\": parse_rate, \"parse_status_counts\": parse_status_counts}
def persist_extraction_identity_state(cache_state, parse_rate, parse_status_counts):
    write_json_atomic(identity_path, extraction_identity_payload(cache_state, parse_rate, parse_status_counts))
reuse_state = cache_reuse_state(saved, identity, query_identity, EXPECTED_EXTRACTION_UNIVERSE, extraction_cache_complete, QUERY_SIGNATURE_CACHE.exists() and saved.get(\"query_identity\") == query_identity, mention_cache_complete=mention_cache_complete, mention_identity_matches=saved.get(\"mention_identity\", saved_mention_identity.get(\"identity\")) == mention_identity); need_extraction = reuse_state[0]; need_signatures = reuse_state[1]
def query_signature_entity_ids_usable(signature, catalog):
    if not isinstance(signature, dict) or not isinstance(catalog, dict): return {\"valid\": False, \"linked_entity_ids\": [], \"target_anchor_ids\": [], \"invalid_entity_ids\": [\"<invalid>\"]}
    by_id = {str(row.get(\"entity_id\")): row for row in catalog.get(\"entities\", []) if isinstance(row, dict) and row.get(\"entity_id\")}
    entity_ids = signature.get(\"entity_ids\")
    if not isinstance(entity_ids, list) or any(not isinstance(value, str) for value in entity_ids) or len(entity_ids) != len(set(entity_ids)): return {\"valid\": False, \"linked_entity_ids\": [], \"target_anchor_ids\": [], \"invalid_entity_ids\": [\"<invalid_entity_ids>\"]}
    invalid, linked, target = [], [], []
    for entity_id in entity_ids:
        row = by_id.get(entity_id)
        if row is None or row.get(\"retrieval_allowed\") is not True or row.get(\"link_status\") != \"linked\": invalid.append(entity_id); continue
        if row.get(\"namespace\") == \"target\" or row.get(\"factual_identity_allowed\") is not True: target.append(entity_id)
        else: linked.append(entity_id)
    return {\"valid\": not invalid, \"linked_entity_ids\": linked, \"target_anchor_ids\": target, \"invalid_entity_ids\": sorted(invalid)}
def verify_query_signature_cache(rows, catalog=None):
    expected_by_id = {identity[\"record_id\"]: identity for identity in record_identities}; seen = set()
    if len(rows) != len(expected_by_id): return False
    for row in rows:
        row_id = str(row.get(\"ID\", \"\")); embedded = row.get(\"record_identity\"); expected = expected_by_id.get(row_id); catalog_bound = \"CATALOG_MANIFEST_HASH\" in globals()
        if row_id in seen or expected is None or embedded != expected or row.get(\"input_text_sha256\") != expected[\"input_text_sha256\"] or (catalog_bound and row.get(\"catalog_manifest_hash\") != CATALOG_MANIFEST_HASH): return False
        if row.get(\"parse_status\") not in ACCEPTED_PARSE_STATUSES or not validate_query_signature(row.get(\"query_signature\"))[\"valid\"]: return False
        if catalog is not None and not query_signature_entity_ids_usable(row.get(\"query_signature\"), catalog)[\"valid\"]: return False
        seen.add(row_id)
    return seen == set(expected_by_id)
def semantic_linkage_quality_gate(graph_tables, catalog, query_signatures, thresholds=None, smoke_test=False):
    limits = dict(thresholds or QUALITY_THRESHOLDS)
    if smoke_test:
        limits[\"minimum_graph_linked_claim_rate\"] = min(float(limits.get(\"minimum_graph_linked_claim_rate\", 0.02)), 0.10)
        limits[\"minimum_query_linked_entity_rate\"] = min(float(limits.get(\"minimum_query_linked_entity_rate\", 0.02)), 0.10)
    factual_ids = {str(row.get(\"entity_id\")) for row in catalog.get(\"entities\", []) if isinstance(row, dict) and row.get(\"namespace\") == \"corpus\" and row.get(\"factual_identity_allowed\") is True and row.get(\"retrieval_allowed\") is True and row.get(\"link_status\") == \"linked\"}
    accepted_claims = [claim for claim in graph_tables.get(\"Claim\", []) if isinstance(claim, dict) and claim.get(\"review_status\") == \"accepted\"]
    linked_claims = [claim for claim in accepted_claims if any(str(claim.get(key)) in factual_ids for key in (\"subject_entity_id\", \"object_entity_id\"))]
    linked_queries = 0; target_only = 0
    for signature in query_signatures or []:
        usable = query_signature_entity_ids_usable(signature, catalog)
        if usable[\"linked_entity_ids\"]: linked_queries += 1
        elif usable[\"target_anchor_ids\"] and usable[\"valid\"]: target_only += 1
    graph_rate = len(linked_claims) / max(1, len(accepted_claims)); query_rate = linked_queries / max(1, len(query_signatures or []))
    result = {\"pass\": len(linked_claims) >= int(limits.get(\"minimum_linked_claims\", 1)) and graph_rate >= float(limits.get(\"minimum_graph_linked_claim_rate\", 0.02)) and linked_queries >= int(limits.get(\"minimum_linked_queries\", 1)) and query_rate >= float(limits.get(\"minimum_query_linked_entity_rate\", 0.02)), \"accepted_claims\": len(accepted_claims), \"linked_claims\": len(linked_claims), \"graph_linked_claim_rate\": graph_rate, \"total_queries\": len(query_signatures or []), \"linked_queries\": linked_queries, \"linked_query_rate\": query_rate, \"target_anchor_only_signatures\": target_only, \"thresholds\": limits, \"scoring_calibration_status\": SCORING_CALIBRATION_STATUS, \"self_confidence_status\": SELF_CONFIDENCE_STATUS}
    return result
if need_extraction or need_signatures:
    if not model_loaded: load_qwen_for_extraction(); model_loaded = True
    if need_extraction:
        with EXTRACTION_CACHE.open(\"w\", encoding=\"utf-8\") as stream:
            extraction_chunks = canonical_extraction_chunks(chunks[chunks.factual_index_allowed].to_dict(\"records\"), EXPECTED_EXTRACTION_UNIVERSE)
            for start in tqdm(range(0, len(extraction_chunks), CONFIG[\"extraction_batch_size\"]), desc=\"Qwen semantic extraction\"):
                batch = extraction_chunks[start:start + CONFIG[\"extraction_batch_size\"]]
                base_contexts = [{k: chunk.get(k) for k in [\"document_uid\", \"chunk_id\", \"source_id\", \"source_type\", \"authority_score\", \"factual_index_allowed\", \"status\", \"content_sha256\", \"document_sha256\", \"text_sha256\"]} for chunk in batch]
                candidate_contexts = []
                for chunk, base_context in zip(batch, base_contexts):
                    candidates = {mention[\"mention_id\"]: build_entity_candidates(mention[\"text\"], ENTITY_CATALOG, namespace_filter=\"corpus\", factual_only=True) for mention in mention_records if mention.get(\"document_uid\") == chunk[\"document_uid\"] and mention.get(\"chunk_id\") == chunk[\"chunk_id\"]}
                    candidate_contexts.append({**base_context, \"entity_catalog\": ENTITY_CATALOG, \"candidate_sets\": candidates})
                prompts = [build_extraction_prompt(chunk[\"text\"], context, \"corpus\") for chunk, context in zip(batch, candidate_contexts)]
                raw_outputs = generate_batch(prompts, CONFIG[\"perspective_max_new_tokens\"], enable_thinking=False, output_schema=SEMANTIC_EXTRACTION_SCHEMA)
                for chunk, context, prompt, raw in zip(batch, candidate_contexts, prompts, raw_outputs):
                    payload, final_raw, parse_status, validation = parse_extraction_with_one_repair(raw, prompt[1][\"content\"], chunk[\"text\"], context)
                    stream.write(json.dumps({\"document_uid\": chunk[\"document_uid\"], \"chunk_id\": chunk[\"chunk_id\"], \"text_sha256\": chunk[\"text_sha256\"], \"catalog_manifest_hash\": CATALOG_MANIFEST_HASH, \"validation\": validation, \"parse_status\": parse_status, \"raw_output\": final_raw}, ensure_ascii=False) + \"\\n\")
    if need_signatures:
        with QUERY_SIGNATURE_CACHE.open(\"w\", encoding=\"utf-8\") as stream:
            for record in tqdm(dataset.to_dict(\"records\"), desc=\"Qwen query signatures\"):
                prompt = build_query_signature(record); raw = generate_batch([prompt], 256, enable_thinking=False, output_schema=QUERY_MODEL_SCHEMA)[0]; parsed, final_raw, parse_status = parse_with_one_repair(raw, prompt[1][\"content\"], validate_query_model_signature, QUERY_MODEL_SCHEMA); target_candidates = build_entity_candidates(str(record.get(\"Target\", \"\")), ENTITY_CATALOG, namespace_preference=\"corpus\", allow_target_fallback=True); resolved = resolve_query_signature_entities(parsed, {\"target\": target_candidates}, ENTITY_CATALOG) if parsed is not None else {\"valid\": False}; signature = ({\"entity_ids\": resolved[\"entity_ids\"], \"predicates\": parsed[\"predicates\"], \"polarities\": parsed[\"polarities\"], \"modalities\": parsed[\"modalities\"], \"desired_stances\": parsed[\"desired_stances\"]} if parse_status in ACCEPTED_PARSE_STATUSES and resolved.get(\"valid\") else None); stream.write(json.dumps({\"ID\": str(record[\"ID\"]), \"record_identity\": record_identity(record), \"input_text_sha256\": record[\"input_text_sha256\"], \"catalog_manifest_hash\": CATALOG_MANIFEST_HASH, \"query_signature\": signature, \"parse_status\": parse_status if signature is not None else \"semantic_invalid\"}, ensure_ascii=False) + \"\\n\")
    unload_generator()
else: print(\"reusing validated extraction and query-signature caches\")
extraction_rows = load_jsonl(EXTRACTION_CACHE); signature_rows = load_jsonl(QUERY_SIGNATURE_CACHE);
if not verify_extraction_cache(extraction_rows, EXPECTED_EXTRACTION_UNIVERSE): raise RuntimeError(\"cache_identity_mismatch: extraction rows do not match expected factual extraction universe\")
if not verify_query_signature_cache(signature_rows, catalog=ENTITY_CATALOG): raise RuntimeError(\"cache_identity_mismatch: query signature rows do not match frozen record identities or frozen catalog\")
if any(row.get(\"query_signature\") is None for row in signature_rows): raise RuntimeError(\"schema_invalid: query signature cache contains rejected payloads\")
QUERY_SIGNATURES = {row[\"ID\"]: row[\"query_signature\"] for row in signature_rows}; assert set(QUERY_SIGNATURES) == set(dataset.ID.astype(str)); query_quality = {\"total_queries\": len(signature_rows), \"valid_signatures\": sum(row.get(\"query_signature\") is not None for row in signature_rows), \"linked_entity_signatures\": sum(bool(query_signature_entity_ids_usable(row.get(\"query_signature\"), ENTITY_CATALOG)[\"linked_entity_ids\"]) for row in signature_rows if isinstance(row.get(\"query_signature\"), dict)), \"target_anchor_only_signatures\": sum(bool(query_signature_entity_ids_usable(row.get(\"query_signature\"), ENTITY_CATALOG)[\"target_anchor_ids\"]) and not bool(query_signature_entity_ids_usable(row.get(\"query_signature\"), ENTITY_CATALOG)[\"linked_entity_ids\"]) for row in signature_rows if isinstance(row.get(\"query_signature\"), dict)), \"coverage_gate\": \"enforced_by_semantic_linkage_quality_gate\"}; query_quality[\"linked_entity_rate\"] = query_quality[\"linked_entity_signatures\"] / max(1, query_quality[\"total_queries\"]); write_json(RUN / \"artifacts\" / \"query_quality_gate.json\", query_quality); parse_rate = sum(extraction_parse_accepted(row.get(\"parse_status\"), row.get(\"validation\", {})) for row in extraction_rows) / max(1, len(EXPECTED_EXTRACTION_UNIVERSE)); parse_status_counts = {status: sum(1 for row in extraction_rows if row.get(\"parse_status\") == status) for status in sorted({row.get(\"parse_status\") for row in extraction_rows})}; cache_state = \"complete_ready\" if parse_rate >= CONFIG[\"minimum_parse_rate\"] else \"complete_validation_failed\"; persist_extraction_identity_state(cache_state, parse_rate, parse_status_counts); log_event(\"extraction\", \"cache_state_persisted\", cache_state=cache_state, parse_rate=parse_rate, parse_status_counts=parse_status_counts); accepted_extractions = [row[\"validation\"] for row in extraction_rows]
if cache_state != \"complete_ready\": log_event(\"extraction\", \"parse_rate_gate_failed\", cache_state=cache_state, parse_rate=parse_rate, minimum_parse_rate=CONFIG[\"minimum_parse_rate\"], parse_status_counts=parse_status_counts); raise RuntimeError(f\"parse_rate gate failed: {parse_rate:.3f}\")
"""),
        code("""# 07 - Build the validated semantic graph and persist graph manifests.
graph_tables = build_semantic_graph(chunks.to_dict(\"records\"), accepted_extractions)
graph_yield = semantic_linkage_quality_gate(graph_tables, ENTITY_CATALOG, list(QUERY_SIGNATURES.values()), QUALITY_THRESHOLDS, smoke_test=CONFIG[\"smoke_test\"])
graph_yield.update({\"accepted_linked_claims\": graph_yield[\"linked_claims\"], \"graph_claims\": len(graph_tables[\"Claim\"]), \"graph_evidence\": len(graph_tables[\"EvidenceChunk\"]), \"entity_count\": len(graph_tables[\"Entity\"]), \"quarantined\": len(graph_tables[\"quarantined\"]), \"reviewed\": len(graph_tables[\"reviewed\"]), \"scoring_calibration_status\": SCORING_CALIBRATION_STATUS, \"self_confidence_status\": SELF_CONFIDENCE_STATUS})
write_json(RUN / \"artifacts\" / \"graph_quality_gate.json\", graph_yield)
if not graph_yield[\"pass\"] or not graph_tables[\"Claim\"] or not graph_tables[\"EvidenceChunk\"]: raise RuntimeError(\"semantic_graph_quality_gate_failed:\" + json.dumps(graph_yield, sort_keys=True))
GRAPH_MANIFEST_HASH = stable_id(json.dumps({\"graph\": graph_tables, \"catalog_manifest_hash\": CATALOG_MANIFEST_HASH}, ensure_ascii=False, sort_keys=True, default=str))
semantic_kg_nodes = pd.DataFrame([{\"node_id\": x.get(\"document_uid\"), \"node_type\": \"Document\", **x} for x in graph_tables[\"Document\"]] + [{\"node_id\": x.get(\"evidence_chunk_id\"), \"node_type\": \"EvidenceChunk\", **x} for x in graph_tables[\"EvidenceChunk\"]] + [{\"node_id\": x.get(\"claim_id\"), \"node_type\": \"Claim\", **x} for x in graph_tables[\"Claim\"]] + [{\"node_id\": x.get(\"entity_id\"), \"node_type\": \"Entity\", **x} for x in graph_tables[\"Entity\"]]); semantic_kg_edges = pd.DataFrame(graph_tables[\"edges\"])
semantic_kg_nodes.to_parquet(RUN / \"artifacts\" / \"semantic_kg_nodes.parquet\", index=False); semantic_kg_edges.to_parquet(RUN / \"artifacts\" / \"semantic_kg_edges.parquet\", index=False); write_json(RUN / \"artifacts\" / \"semantic_kg_manifest.json\", {\"graph_manifest_hash\": GRAPH_MANIFEST_HASH, \"catalog_manifest_hash\": CATALOG_MANIFEST_HASH, \"corpus_manifest_hash\": CORPUS_MANIFEST_HASH, \"chunk_manifest_hash\": CHUNK_MANIFEST_HASH, \"node_count\": len(semantic_kg_nodes), \"edge_count\": len(semantic_kg_edges)})
"""),
        code("""# 08 - Retrieval models, semantic seeds, full-signature graph retrieval, RRF, and reranking.
from sentence_transformers import SentenceTransformer, CrossEncoder
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct
from rank_bm25 import BM25Okapi
def load_retrieval_models():
    global EMBEDDER, RERANKER, BM25, QDRANT, factual_chunks
    factual_chunks = chunks[(chunks.factual_index_allowed == True) & (chunks.authority_score >= CONFIG[\"minimum_authority\"])].reset_index(drop=True);
    if factual_chunks.empty: raise RuntimeError(\"no_indexable_chunks: no factual chunks met the authority threshold\")
    EMBEDDER = SentenceTransformer(CONFIG[\"embedding_model\"], revision=CONFIG[\"embedding_model_revision\"], device=\"cuda\"); RERANKER = CrossEncoder(CONFIG[\"reranker_model\"], revision=CONFIG[\"reranker_model_revision\"], device=\"cuda\", max_length=512); BM25 = BM25Okapi([re.findall(r\"[\\w']+\", str(x).casefold()) for x in factual_chunks.text]); QDRANT = QdrantClient(path=str(RUN / \"artifacts\" / \"qdrant\")); return \"factual_chunks\"
INDEX_MANIFEST_PATH = RUN / \"artifacts\" / \"index_manifest.json\"; INDEX_IDENTITY = {\"document_uid_manifest_hash\": CORPUS_MANIFEST_HASH, \"chunk_manifest_hash\": CHUNK_MANIFEST_HASH, \"graph_manifest_hash\": GRAPH_MANIFEST_HASH, \"catalog_manifest_hash\": CATALOG_MANIFEST_HASH, \"extraction_audit_hash\": extraction_audit_hash, \"embedding_model\": CONFIG[\"embedding_model\"], \"minimum_authority\": CONFIG[\"minimum_authority\"]}
COLLECTION = load_retrieval_models(); old_index_identity = json.loads(INDEX_MANIFEST_PATH.read_text(encoding=\"utf-8\")).get(\"identity\", {}) if INDEX_MANIFEST_PATH.exists() else {}
if old_index_identity != INDEX_IDENTITY and QDRANT.collection_exists(COLLECTION): print(\"index_identity_mismatch: rebuilding Qdrant collection\"); QDRANT.delete_collection(COLLECTION)
if not QDRANT.collection_exists(COLLECTION):
    vectors = EMBEDDER.encode(factual_chunks.text.tolist(), batch_size=8, normalize_embeddings=True, convert_to_numpy=True).astype(\"float32\")
    QDRANT.create_collection(COLLECTION, vectors_config=VectorParams(size=vectors.shape[1], distance=Distance.COSINE))
    points = [PointStruct(id=i, vector=vector.tolist(), payload={**row, \"evidence_chunk_id\": stable_id(\"evidence\", row[\"document_uid\"], row[\"chunk_id\"]), \"page\": canonical_page(row.get(\"page\"))}) for i, (row, vector) in enumerate(zip(factual_chunks.to_dict(\"records\"), vectors))]
    for start in range(0, len(points), 128): QDRANT.upsert(COLLECTION, points[start:start + 128], wait=True)
write_json(INDEX_MANIFEST_PATH, {\"identity\": INDEX_IDENTITY, \"document_uid_manifest_hash\": CORPUS_MANIFEST_HASH, \"chunk_manifest_hash\": CHUNK_MANIFEST_HASH})
EVIDENCE_BY_ID = {x[\"evidence_chunk_id\"]: x for x in graph_tables[\"EvidenceChunk\"]}
def dense_search(query, k):
    vector = EMBEDDER.encode([query], normalize_embeddings=True, convert_to_numpy=True)[0].astype(\"float32\").tolist(); points = QDRANT.query_points(COLLECTION, query=vector, limit=k, with_payload=True).points; return [{**dict(p.payload), \"branch\": \"dense\", \"rank\": i, \"raw_score\": float(p.score), \"evidence_chunk_id\": dict(p.payload).get(\"evidence_chunk_id\")} for i, p in enumerate(points, 1) if float(p.score) >= CONFIG[\"minimum_dense_score\"]]
def bm25_search(query, k):
    scores = BM25.get_scores(re.findall(r\"[\\w']+\", str(query).casefold())); rows = []
    for rank, index in enumerate(np.argsort(scores)[::-1][:k], 1):
        if float(scores[index]) <= 0: continue
        row = factual_chunks.iloc[int(index)].to_dict(); row.update({\"branch\": \"bm25\", \"rank\": rank, \"raw_score\": float(scores[index]), \"evidence_chunk_id\": stable_id(\"evidence\", row[\"document_uid\"], row[\"chunk_id\"]) }); rows.append(row)
    return rows
def graph_search(target, k, post=\"\", category=\"\", query_signature=None, graph_enabled=True):
    if not graph_enabled: return []
    query_signature = query_signature or {\"entity_ids\": [], \"predicates\": [], \"polarities\": [], \"modalities\": [], \"desired_stances\": []}; seeds = dense_search(f\"evidence addressing {target} {post}\", CONFIG[\"dense_top_k\"]) + bm25_search(f\"{target} {post}\", CONFIG[\"bm25_top_k\"]); return expand_graph_from_seeds(seeds, query_signature=query_signature, graph_tables=graph_tables, config=GRAPH_CONFIG)[:k]
def retrieve_evidence(post, target, category, query_signature, graph_enabled):
    if not query_signature.get(\"entity_ids\"):
        return {\"evidence\": [], \"candidate_evidence_ids\": [], \"graph_enabled\": graph_enabled, \"abstention_reason\": \"no_linked_entities\"}
    dense_hits, bm25_hits = dense_search(f\"evidence addressing harmful or misleading claims about {target}\", CONFIG[\"dense_top_k\"]), bm25_search(f\"{target} {post}\", CONFIG[\"bm25_top_k\"]); graph_hits = graph_search(target, CONFIG[\"graph_top_k\"], post, category, query_signature, graph_enabled); candidates = reciprocal_rank_fusion({\"dense\": dense_hits, \"bm25\": bm25_hits, \"graph\": graph_hits}, {\"dense\": 1.1, \"bm25\": 1.0, \"graph\": 0.8}, CONFIG[\"rrf_constant\"])
    for item in candidates: item.update(EVIDENCE_BY_ID.get(item[\"evidence_chunk_id\"], {}))
    if not candidates: return {\"evidence\": [], \"candidate_evidence_ids\": [], \"graph_enabled\": graph_enabled, \"abstention_reason\": \"no_positive_candidates\"}
    logits = RERANKER.predict([(post, x[\"text\"]) for x in candidates], batch_size=16).tolist(); selection = select_evidence(candidates, {x[\"evidence_chunk_id\"]: float(y) for x, y in zip(candidates, logits)}, {\"minimum_rerank_probability\": CONFIG[\"minimum_rerank_probability\"], \"max_evidence\": CONFIG[\"rerank_top_k\"]}); return {\"evidence\": selection[\"selected\"], \"candidate_evidence_ids\": [x[\"evidence_chunk_id\"] for x in candidates], \"graph_enabled\": graph_enabled, \"graph_hit_count\": len(graph_hits), \"retrieval_trace\": {\"dense_hits\": len(dense_hits), \"bm25_hits\": len(bm25_hits), \"graph_hits\": len(graph_hits)}, \"abstention_reason\": selection[\"reason\"]}
"""),
        code("""# 09 - Paired graph-on/off retrieval over identical frozen IDs and retrieval-model release.
EVIDENCE_CACHE = BoundedLRU(maxsize=int(CONFIG[\"cache_max_records\"])); PAIRED_RETRIEVAL_CACHE = BoundedLRU(maxsize=int(CONFIG[\"cache_max_records\"]))
def paired_permutation(a, b, rounds=10000, seed=SEED):
    delta = np.asarray(a, dtype=float) - np.asarray(b, dtype=float); rng = np.random.default_rng(seed); observed = abs(delta.mean()); null = [abs((delta * rng.choice([-1, 1], len(delta))).mean()) for _ in range(rounds)]; return {\"mean_difference\": float(delta.mean()), \"p_value\": float((np.asarray(null) >= observed).mean())}
def paired_retrieval_metrics(graph_on, graph_off, shared_frozen_universe):
    on = graph_on[\"evidence\"]; off = graph_off[\"evidence\"]; on_by_id = {x[\"evidence_chunk_id\"]: x for x in on}; off_by_id = {x[\"evidence_chunk_id\"]: x for x in off}; universe = list(dict.fromkeys(shared_frozen_universe)); domain_size = len(universe)
    on_ids, off_ids = set(on_by_id), set(off_by_id)
    if on_ids - set(universe) or off_ids - set(universe): raise ValueError(\"selected evidence outside shared frozen universe\")
    on_selection = [1 if evidence_id in on_by_id else 0 for evidence_id in universe]; off_selection = [1 if evidence_id in off_by_id else 0 for evidence_id in universe]
    on_authority = [float(on_by_id[evidence_id].get(\"authority_score\") or 0.0) if evidence_id in on_by_id else 0.0 for evidence_id in universe]; off_authority = [float(off_by_id[evidence_id].get(\"authority_score\") or 0.0) if evidence_id in off_by_id else 0.0 for evidence_id in universe]
    on_accepted = [1 if evidence_id in on_by_id and str(on_by_id[evidence_id].get(\"status\", \"\")) == \"accepted\" else 0 for evidence_id in universe]; off_accepted = [1 if evidence_id in off_by_id and str(off_by_id[evidence_id].get(\"status\", \"\")) == \"accepted\" else 0 for evidence_id in universe]
    on_score = [float(on_by_id[evidence_id].get(\"rerank_probability\", on_by_id[evidence_id].get(\"rrf_score\", 0.0)) or 0.0) if evidence_id in on_by_id else 0.0 for evidence_id in universe]; off_score = [float(off_by_id[evidence_id].get(\"rerank_probability\", off_by_id[evidence_id].get(\"rrf_score\", 0.0)) or 0.0) if evidence_id in off_by_id else 0.0 for evidence_id in universe]
    denominator = max(1, domain_size); mean = lambda values: float(sum(values) / denominator); traces = [{\"evidence_id\": evidence_id, \"graph_on_selected\": on_selection[i], \"graph_off_selected\": off_selection[i], \"graph_on_authority\": on_authority[i], \"graph_off_authority\": off_authority[i], \"graph_on_accepted\": on_accepted[i], \"graph_off_accepted\": off_accepted[i], \"graph_on_score\": on_score[i], \"graph_off_score\": off_score[i]} for i, evidence_id in enumerate(universe)]
    return {\"shared_frozen_universe\": universe, \"universe_size\": domain_size, \"universe_traces\": traces, \"graph_on_selection_vector\": on_selection, \"graph_off_selection_vector\": off_selection, \"graph_on_authority_vector\": on_authority, \"graph_off_authority_vector\": off_authority, \"graph_on_accepted_vector\": on_accepted, \"graph_off_accepted_vector\": off_accepted, \"graph_on_score_vector\": on_score, \"graph_off_score_vector\": off_score, \"overlap\": mean([a * b for a, b in zip(on_selection, off_selection)]), \"graph_only_gain\": mean([a - b for a, b in zip(on_selection, off_selection)]), \"authority_rate\": mean(on_authority), \"graph_off_authority_rate\": mean(off_authority), \"accepted_evidence_rate\": mean(on_accepted), \"graph_off_accepted_evidence_rate\": mean(off_accepted), \"selected_score_mean\": mean(on_score), \"graph_off_selected_score_mean\": mean(off_score), \"graph_on_selected_count\": sum(on_selection), \"graph_off_selected_count\": sum(off_selection)}
def paired_statistical_comparison(metric_rows):
    on = [float(sum(row.get(\"graph_on_score_vector\", [])) / max(1, row.get(\"universe_size\", len(row.get(\"graph_on_score_vector\", []))))) for row in metric_rows]; off = [float(sum(row.get(\"graph_off_score_vector\", [])) / max(1, row.get(\"universe_size\", len(row.get(\"graph_off_score_vector\", []))))) for row in metric_rows]
    return {\"valid\": len(on) >= 2, **(paired_permutation(on, off) if len(on) >= 2 else {\"mean_difference\": None, \"p_value\": None})}
for record in tqdm(dataset.to_dict(\"records\"), desc=\"Paired retrieval graph-on/off\"):
    rid, signature = str(record[\"ID\"]), QUERY_SIGNATURES[str(record[\"ID\"])]
    graph_on, graph_off = retrieve_evidence(record[\"Text\"], record[\"Target\"], record[\"Category\"], signature, True), retrieve_evidence(record[\"Text\"], record[\"Target\"], record[\"Category\"], signature, False); shared_frozen_universe = sorted(set(graph_on.get(\"candidate_evidence_ids\", [])) | set(graph_off.get(\"candidate_evidence_ids\", []))); graph_on[\"shared_frozen_universe\"] = shared_frozen_universe; graph_off[\"shared_frozen_universe\"] = shared_frozen_universe; metrics_row = paired_retrieval_metrics(graph_on, graph_off, shared_frozen_universe); PAIRED_RETRIEVAL_CACHE[rid] = {\"record_identity\": record_identity(record), \"query_signature\": signature, \"graph_on\": graph_on, \"graph_off\": graph_off, \"frozen_evidence_ids\": shared_frozen_universe, \"shared_frozen_universe\": shared_frozen_universe, \"metrics\": metrics_row}; EVIDENCE_CACHE[rid] = graph_on if not CONFIG[\"graph_ablation\"] else graph_off
paired_metrics = [row[\"metrics\"] for row in PAIRED_RETRIEVAL_CACHE.values()]; paired_comparison = paired_statistical_comparison(paired_metrics) if paired_metrics else {\"valid\": False, \"mean_difference\": None, \"p_value\": None}; cache_snapshot = {\"evidence\": EVIDENCE_CACHE.snapshot(), \"paired_retrieval\": PAIRED_RETRIEVAL_CACHE.snapshot(), \"cache_capacity\": CONFIG[\"cache_max_records\"], \"evictions\": {\"evidence\": EVIDENCE_CACHE.evictions, \"paired_retrieval\": PAIRED_RETRIEVAL_CACHE.evictions}}; write_json(RUN / \"artifacts\" / \"paired_retrieval_evaluation.json\", {\"records\": dict(PAIRED_RETRIEVAL_CACHE), \"paired_statistical_comparison\": paired_comparison, \"cache_snapshot\": cache_snapshot}); write_json(RUN / \"artifacts\" / \"cache_diagnostics.json\", cache_snapshot); log_event(\"retrieval\", \"cache_snapshot\", cache_snapshot=cache_snapshot)
def unload_retrieval_models():
    global EMBEDDER, RERANKER, BM25, QDRANT
    for name in [\"EMBEDDER\", \"RERANKER\", \"BM25\"]:
        if name in globals(): del globals()[name]
    if \"QDRANT\" in globals():
        try: QDRANT.close()
        except Exception: pass
        del QDRANT
    gc.collect(); torch.cuda.empty_cache(); MEMORY_SNAPSHOTS.append(gpu_snapshot(\"retrieval_models_unloaded\")); print(\"retrieval_models_unloaded\")
unload_retrieval_models()
"""),
code("""# 10 - Reload the same Qwen generator only after retrieval models are unloaded.
print(\"Qwen generation load deferred until after the shared NLI lifecycle gate.\")
"""),
        code("""# 11 - Canonical E1...En ledgers for perspectives, synthesis, generation, and metrics.
PERSPECTIVES = [(\"fact_checking\", \"Correct only directly supported claims.\"), (\"cultural_context\", \"Give respectful context.\"), (\"harm_reduction\", \"Avoid amplifying harmful language.\"), (\"legal_rights\", \"Use only source-grounded rights claims.\"), (\"persuasion\", \"Suggest an empathetic response.\")]
def build_evidence_ledger(evidence):
    source_lookup = globals().get("SOURCE_TEXT_BY_KEY", {}); rows = []
    for i, x in enumerate(evidence):
        displayed = str(x.get("text", "")); source_key = x.get("source_text_key"); source_text = source_lookup.get(source_key) if source_key else displayed
        if source_key and source_text is None: raise RuntimeError("source_text_lookup_missing")
        start = int(x.get("span_start", 0) or 0); end = int(x.get("span_end", len(displayed)) or 0)
        if not isinstance(source_text, str) or start < 0 or start >= end or end > len(source_text) or source_text[start:end] != displayed: raise RuntimeError("evidence_span_invariant_failed")
        rows.append({"evidence_id": f"E{i + 1}", "ledger_id": f"E{i + 1}", "rank": x.get("rank", i), "evidence_chunk_id": x["evidence_chunk_id"], "chunk_id": x.get("chunk_id"), "document_uid": x.get("document_uid"), "source_id": x.get("source_id"), "locator": f"p. {canonical_page(x.get('page'))}" if canonical_page(x.get("page")) is not None else f"section: {x.get('section', 'document')}", "displayed_text": displayed, "evidence_text": displayed, "source_text_key": source_key or f"inline-{i}", "source_text_sha256": __import__("hashlib").sha256(source_text.encode("utf-8")).hexdigest(), "span_start": start, "span_end": end, "displayed_text_sha256": __import__("hashlib").sha256(displayed.encode("utf-8")).hexdigest()})
    return validate_evidence_ledger(rows, source_lookup) if "validate_evidence_ledger" in globals() else rows
def validate_evidence_ids(ids, ledger):
    if not isinstance(ids, list): return {\"valid\": False, \"unknown\": [str(ids)] if ids is not None else [], \"evidence_ids\": [], \"quarantine\": True, \"reason\": \"evidence_ids_not_list\"}
    values = list(ids); allowed = {x[\"evidence_id\"] for x in ledger}; unknown = [str(value) for value in values if not isinstance(value, str) or value not in allowed]; valid = all(isinstance(value, str) for value in values) and len(values) == len(set(values)) and not unknown and all(re.fullmatch(r\"E[1-9][0-9]*\", value) for value in values); return {\"valid\": valid, \"unknown\": unknown, \"evidence_ids\": values, \"quarantine\": not valid}
def evidence_for_prompt(evidence, char_budget=None):
    ledger = build_evidence_ledger(evidence)
    if char_budget is None: return evidence, ledger
    selected = select_evidence_within_budget(ledger, int(char_budget)); selected_ids = {x["evidence_id"] for x in selected}
    return [item for item, entry in zip(evidence, ledger) if entry["evidence_id"] in selected_ids], selected
def evidence_block(evidence, limit=None):
    # Whole spans only: an explicit budget selects a ledger-consistent prefix;
    # it never slices evidence text.
    _, ledger = evidence_for_prompt(evidence, limit)
    return \"\\n\\n\".join(f\"[{x['evidence_id']}] source={x['source_id']} {x['locator']} document_uid={x['document_uid']} span={x['span_start']}:{x['span_end']}\\n{x['displayed_text']}\" for x in ledger)
def _ledger_block(ledger):
    return "\\n\\n".join(f"[{x['evidence_id']}] source={x['source_id']} {x['locator']} document_uid={x['document_uid']} span={x['span_start']}:{x['span_end']}\\n{x['displayed_text']}" for x in ledger)
def _prompt_token_count(prompt):
    encoded = tokenizer(text=[str(prompt)], return_tensors="pt", padding=False, truncation=False)
    mask = encoded["attention_mask"]
    return int(mask.sum().item()) if hasattr(mask.sum(), "item") else int(mask.sum())
def _adaptive_messages(evidence, message_builder, payload_candidates, max_new_tokens):
    # Deterministic token fitting drops whole ledger spans, never text slices.
    ledger = build_evidence_ledger(evidence)
    def rendered_builder(selected, payload):
        return _apply_chat_template(message_builder(selected, payload), enable_thinking=False)
    fit = fit_adaptive_prompt_with_evidence(ledger, prompt_builder=rendered_builder, payload_candidates=payload_candidates, token_counter=_prompt_token_count, max_input_tokens=int(CONFIG["max_seq_length"]), reserve_output_tokens=int(max_new_tokens))
    if fit.get("status") != "fit": raise PromptBudgetQuarantine(fit)
    selected_ids = set(fit.get("selected_evidence_ids", [])); selected = [item for item in ledger if item["evidence_id"] in selected_ids]
    return message_builder(selected, fit.get("payload", "")), fit
def adaptive_perspective_messages(post, target, category, evidence):
    prompts = []; schema_tail = "JSON schema (must be followed exactly): " + json.dumps(PERSPECTIVE_SCHEMA, ensure_ascii=False, sort_keys=True)
    for name, goal in PERSPECTIVES:
        def builder(selected, payload, name=name, goal=goal):
            return [{"role": "system", "content": "Think carefully, then return exactly one JSON object after the thinking block."}, {"role": "user", "content": f"Perspective: {name}\\nGoal: {goal}\\nPost: {post}\\nTarget: {target}\\nCategory: {category}\\nEvidence (complete ledger spans only):\\n{_ledger_block(selected)}\\nReturn perspective, rationale, claims_to_address, supported_evidence_ids, response_guidance, risk_flags, confidence. Use only E1...En IDs.\\n{schema_tail}"}]
        prompt, fit = _adaptive_messages(evidence, builder, ("",), CONFIG["perspective_max_new_tokens"]); prompts.append((prompt, fit))
    return prompts
def enforce_perspective_parse_rate(rows):
    total_count = len(rows)
    parse_status_counts = {}
    for row in rows:
        status = str(row.get("structured_output", {}).get("parse_status", "schema_invalid"))
        parse_status_counts[status] = parse_status_counts.get(status, 0) + 1
    parsed_count = sum(parse_status_counts.get(status, 0) for status in ACCEPTED_PARSE_STATUSES)
    parse_rate = parsed_count / max(1, total_count)
    audit = {"reason": "minimum_parse_rate_failed", "parse_rate": parse_rate, "minimum_parse_rate": float(CONFIG["minimum_parse_rate"]), "parsed_count": parsed_count, "total_count": total_count, "parse_status_counts": parse_status_counts}
    if parse_rate < audit["minimum_parse_rate"]: raise ParseRateQuarantine(audit)
    return audit
def perspective_messages(post, target, category, evidence):
    prompt_evidence, _ = evidence_for_prompt(evidence, CONFIG.get("evidence_char_budget")); return [[{\"role\": \"system\", \"content\": \"Return exactly one JSON object and no reasoning trace.\"}, {\"role\": \"user\", \"content\": f\"Perspective: {name}\\nGoal: {goal}\\nPost: {post}\\nTarget: {target}\\nCategory: {category}\\nEvidence (full exact displayed spans):\\n{evidence_block(prompt_evidence)}\\nReturn perspective, rationale, claims_to_address, supported_evidence_ids, response_guidance, risk_flags, confidence. Use only E1...En IDs.\"}] for name, goal in PERSPECTIVES]
def run_perspectives(post, target, category, evidence):
    rows = []; prompt_entries = adaptive_perspective_messages(post, target, category, evidence); prompts = [entry[0] for entry in prompt_entries]; raw_outputs = generate_batch(prompts, CONFIG[\"perspective_max_new_tokens\"], temperature=0.6, enable_thinking=CONFIG[\"thinking_enabled\"], output_schema=PERSPECTIVE_SCHEMA)
    for (name, goal), (prompt, fit), raw in zip(PERSPECTIVES, prompt_entries, raw_outputs):
        allowed = set(fit.get(\"selected_evidence_ids\", []))
        parsed, final_raw, status = parse_with_one_repair(raw, prompt[1][\"content\"], lambda payload: validate_perspective(payload, allowed, expected_perspective=name), PERSPECTIVE_SCHEMA);
        if parsed is None: parsed = {\"perspective\": name, \"rationale\": \"\", \"claims_to_address\": [], \"supported_evidence_ids\": [], \"response_guidance\": [], \"risk_flags\": [status], \"confidence\": 0.0, \"parse_status\": status, \"quarantine\": True}
        else: parsed[\"parse_status\"] = status
        rows.append({\"perspective\": name, \"goal\": goal, \"perspective_rationale\": str(parsed.get(\"rationale\", \"\")), \"structured_output\": parsed, \"supported_evidence_ids\": parsed.get(\"supported_evidence_ids\", []), \"reasoning_trace\": qwen_generation_trace(final_raw), \"raw_output\": final_raw})
    enforce_perspective_parse_rate(rows)
    return rows
"""),
        code("""# 12 - Synthesis, final generation, canonical citation validation, and checkpoint identity.
VARIANT_COLUMNS = {\"qwen_zero_shot\": \"qwen-zero-shot-counter-narrative\", \"qwen_few_shot\": \"qwen-few-shot-counter-narrative\", \"kg_rag\": \"kg-rag-generated-source-grounded-counter-narrative\", \"mp_kg_rag\": \"mp-kg-rag-generated-source-grounded-counter-narrative\"}
def resolve_citation_tokens(text, evidence_ledger):
    by_id = {x[\"evidence_id\"]: x for x in evidence_ledger}; tokens = re.findall(r\"\\[(E\\d+)\\]\", str(text)); return tokens, [by_id[x] for x in tokens if x in by_id], sorted(set(tokens) - set(by_id))
def claim_level_citations(text, evidence_ledger, factual_claims=None):
    if "build_claim_citation_records" in globals():
        return build_claim_citation_records(text, evidence_ledger, factual_claims=factual_claims)
    return [{"claim": sentence, "claim_text": sentence, "evidence_ids": resolve_citation_tokens(sentence, evidence_ledger)[0], "is_factual": bool(factual_claims), "citation_format_valid": True, "unknown_evidence_ids": []} for sentence in re.split(r"(?<=[.!?])\\s+", str(text).strip()) if sentence]
def synthesize_plan(post, target, category, evidence, perspective_outputs):
    compact = [{k: p[\"structured_output\"].get(k, []) for k in [\"perspective\", \"rationale\", \"claims_to_address\", \"supported_evidence_ids\", \"response_guidance\", \"risk_flags\"]} for p in perspective_outputs if p[\"structured_output\"].get(\"parse_status\") in ACCEPTED_PARSE_STATUSES]
    compact_minimal = [{\"perspective\": p.get(\"perspective\"), \"supported_evidence_ids\": p.get(\"supported_evidence_ids\", [])} for p in compact]
    payload_candidates = (json.dumps(compact, ensure_ascii=False, sort_keys=True), json.dumps(compact_minimal, ensure_ascii=False, sort_keys=True))
    schema_tail = \"JSON schema (must be followed exactly): \" + json.dumps(PLAN_SCHEMA, ensure_ascii=False, sort_keys=True)
    def builder(selected, payload):
        return [{\"role\": \"system\", \"content\": \"Return one JSON evidence plan only.\"}, {\"role\": \"user\", \"content\": f\"Post: {post}\\nTarget: {target}\\nEvidence (complete ledger spans only):\\n{_ledger_block(selected)}\\nPerspectives: {payload}\\nReturn claim_focus, selected_evidence_ids, response_steps, tone, factual_constraints, safety_constraints. Use only E IDs.\\n{schema_tail}\"}]
    prompt, fit = _adaptive_messages(evidence, builder, payload_candidates, CONFIG[\"plan_max_new_tokens\"]); ledger_ids = set(fit.get(\"selected_evidence_ids\", [])); raw = generate_batch([prompt], CONFIG[\"plan_max_new_tokens\"], temperature=0.6, enable_thinking=CONFIG[\"thinking_enabled\"], output_schema=PLAN_SCHEMA)[0]; parsed, final_raw, status = parse_with_one_repair(raw, prompt[1][\"content\"], lambda payload: validate_plan(payload, ledger_ids), PLAN_SCHEMA);
    if parsed is None: raise SchemaValidationQuarantine({'reason': 'plan_schema_invalid', 'stage': 'plan', 'parse_status': status})
    parsed[\"parse_status\"] = status; return parsed, final_raw
def generate_final_counter_narrative(post, target, evidence, plan):
    compact_plan = {key: plan.get(key) for key in [\"claim_focus\", \"selected_evidence_ids\", \"response_steps\", \"tone\"] if key in plan}
    payload_candidates = (json.dumps(plan, ensure_ascii=False, sort_keys=True), json.dumps(compact_plan, ensure_ascii=False, sort_keys=True))
    schema_tail = \"JSON schema (must be followed exactly): \" + json.dumps(FINAL_RESPONSE_SCHEMA, ensure_ascii=False, sort_keys=True)
    def builder(selected, payload):
        return [{\"role\": \"system\", \"content\": \"Write one safe evidence-grounded counter-narrative as JSON only.\"}, {\"role\": \"user\", \"content\": f\"Post: {post}\\nTarget: {target}\\nPlan: {payload}\\nEvidence (complete ledger spans only):\\n{_ledger_block(selected)}\\nFactual claims require inline [E1] citations. If there are no factual claims, return exactly one approved safe-abstention template and set factual_claims to []. Return counter_narrative, cited_evidence_ids, factual_claims, safety_notes. Use only E IDs; chunk IDs are metadata only.\\n{schema_tail}\"}]
    prompt, fit = _adaptive_messages(evidence, builder, payload_candidates, CONFIG[\"answer_max_new_tokens\"]); ledger_ids = set(fit.get(\"selected_evidence_ids\", [])); raw = generate_batch([prompt], CONFIG[\"answer_max_new_tokens\"], temperature=0.6, enable_thinking=CONFIG[\"thinking_enabled\"], output_schema=FINAL_RESPONSE_SCHEMA)[0]; parsed, final_raw, status = parse_with_one_repair(raw, prompt[1][\"content\"], lambda payload: validate_final_response(payload, ledger_ids), FINAL_RESPONSE_SCHEMA);
    if parsed is None: raise SchemaValidationQuarantine({'reason': 'final_schema_invalid', 'stage': 'final', 'parse_status': status})
    parsed[\"parse_status\"] = status; return parsed, final_raw
def grounding_repair_response(post, target, evidence, plan, failure_reasons):
    payload_candidates = (json.dumps({\"plan\": plan, \"failures\": failure_reasons}, ensure_ascii=False, sort_keys=True), json.dumps({\"failures\": [str(item) for item in failure_reasons[:3]]}, ensure_ascii=False, sort_keys=True))
    schema_tail = \"JSON schema (must be followed exactly): \" + json.dumps(FINAL_RESPONSE_SCHEMA, ensure_ascii=False, sort_keys=True)
    def builder(selected, payload):
        return [{\"role\": \"system\", \"content\": \"Repair one failed evidence-grounded response as strict JSON.\"}, {\"role\": \"user\", \"content\": f\"Post: {post}\\nTarget: {target}\\nRepair context: {payload}\\nEvidence (complete ledger spans only):\\n{_ledger_block(selected)}\\nReturn counter_narrative, cited_evidence_ids, factual_claims, safety_notes. Cite only supported E IDs.\\n{schema_tail}\"}]
    prompt, fit = _adaptive_messages(evidence, builder, payload_candidates, CONFIG[\"answer_max_new_tokens\"]); ledger_ids = set(fit.get(\"selected_evidence_ids\", [])); raw = generate_batch([prompt], CONFIG[\"answer_max_new_tokens\"], temperature=0.6, enable_thinking=CONFIG[\"thinking_enabled\"], output_schema=FINAL_RESPONSE_SCHEMA)[0]; parsed, final_raw, status = parse_with_one_repair(raw, prompt[1][\"content\"], lambda payload: validate_final_response(payload, ledger_ids), FINAL_RESPONSE_SCHEMA)
    if parsed is None: raise SchemaValidationQuarantine({'reason': 'grounding_repair_schema_invalid', 'stage': 'grounding_repair', 'parse_status': status})
    parsed['parse_status'] = status; parsed['grounding_repair_reasons'] = list(failure_reasons); return parsed, final_raw
SAFE_ABSTENTION_TEMPLATES = {"I cannot verify this from the available evidence.", "I can’t verify this from the available evidence.", "இந்தக் கூற்றை கிடைக்கும் ஆதாரங்களிலிருந்து சரிபார்க்க முடியவில்லை.", "उपलब्ध साक्ष्यों से इस दावे की पुष्टि नहीं कर सकता।"}
def safe_abstention_validator(text):
    raw = unicodedata.normalize("NFKC", str(text or "")).strip()
    if re.search(r"\\[(?:E[1-9][0-9]*)\\]", raw): return False
    normalized = re.sub(r"\\s+", " ", raw).strip().rstrip(" .!?！？。॥")
    return any(normalized == re.sub(r"\\s+", " ", unicodedata.normalize("NFKC", item).strip()).rstrip(" .!?！？。॥") for item in SAFE_ABSTENTION_TEMPLATES)
NLI_CALIBRATION_ARTIFACT = {}
LANGUAGE_DETECTOR = load_language_detector()
if LANGUAGE_DETECTOR is None: raise RuntimeError('language_evaluation_detector_unavailable')
if CONFIG.get('nli_model_label_mapping') != MODEL_LABEL_MAPPING or CONFIG.get('nli_dataset_label_mapping') != DATASET_LABEL_MAPPING:
    raise RuntimeError('nli_label_mapping_config_mismatch')
NLI_CALIBRATION_PATH = RUN / 'artifacts' / 'nli_calibration.json'
if NLI_CALIBRATION_PATH.exists():
    NLI_CALIBRATION_ARTIFACT = json.loads(NLI_CALIBRATION_PATH.read_text(encoding='utf-8'))
def create_nli_calibration_artifact(language, heldout_rows, predictor, *, dataset_id, dataset_revision, split='validation', label_mapping=None):
    rows = list(heldout_rows)
    if not rows: raise RuntimeError('nli_calibration_dataset_empty')
    example_ids = [str(row['id']) for row in rows]
    example_digest = stable_identity_hash([{key: row.get(key) for key in ['id', 'premise', 'hypothesis', 'label']} for row in rows])
    raw_predictions = [predictor(row['premise'], row['hypothesis']) for row in rows]
    predictions = [int(value.get('label')) if isinstance(value, dict) else int(value) for value in raw_predictions]
    labels = [int(row['label']) for row in rows]; accuracy = sum(prediction == label for prediction, label in zip(predictions, labels)) / len(labels)
    per_label = {}
    for label in sorted(set(labels)):
        support = sum(value == label for value in labels); true_positive = sum(value == label and prediction == label for value, prediction in zip(labels, predictions)); predicted = sum(prediction == label for prediction in predictions)
        per_label[str(label)] = {'n': support, 'correct': true_positive, 'precision': true_positive / predicted if predicted else 0.0, 'recall': true_positive / support if support else 0.0}
    selected_label_mapping = label_mapping or CONFIG.get('nli_dataset_label_mapping', DATASET_LABEL_MAPPING)
    entailment_label = int(selected_label_mapping.get('entailment')) if 'entailment' in selected_label_mapping else int(next(key for key, value in selected_label_mapping.items() if str(value).casefold() == 'entailment'))
    per_label['entailment'] = per_label.get(str(entailment_label), {'n': 0, 'correct': 0, 'precision': 0.0, 'recall': 0.0})
    dataset_hash = normalized_dataset_content_hash(rows)
    scores = [float(value.get('entailment_probability', 0.0)) if isinstance(value, dict) else float(value == 0) for value in raw_predictions]
    indexed_predictions = {str(row['id']): {'label': prediction, 'entailment_probability': score} for row, prediction, score in zip(rows, predictions, scores)}
    calibration = calibrate_nli_threshold(rows, lambda row: indexed_predictions[str(row['id'])], seed=SEED, min_support=int(CONFIG['nli_calibration_min_support']), n_bootstrap=int(CONFIG['nli_calibration_bootstrap']), entailment_label=entailment_label)
    calibration_rows = [row for row in rows if str(row['id']) in calibration['calibration_ids']]
    audit_rows = [row for row in rows if str(row['id']) in calibration['audit_ids']]
    audit_stats = {}
    audit_predictions = [indexed_predictions[str(item['id'])]['label'] for item in audit_rows]
    for raw_label in sorted({int(row['label']) for row in audit_rows}):
        support = sum(int(row['label']) == raw_label for row in audit_rows)
        correct = sum(int(row['label']) == raw_label and prediction == raw_label for row, prediction in zip(audit_rows, audit_predictions))
        predicted = sum(prediction == raw_label for prediction in audit_predictions)
        audit_stats[str(raw_label)] = {'n_total': support, 'positive_support': support, 'correct': correct, 'precision': correct / predicted if predicted else 0.0, 'recall': correct / support if support else 0.0}
    audit_stats['entailment'] = {**calibration['metrics']['audit']['entailment'], 'n_total': len(audit_rows), 'positive_support': calibration['metrics']['audit']['entailment']['support']}
    artifact = build_nli_calibration_artifact(model_id=CONFIG['nli_model_id'], model_revision=CONFIG['nli_model_revision'], dataset_id=dataset_id, dataset_revision=dataset_revision, dataset_content_hash=dataset_hash, language=language, split=split, label_mapping=selected_label_mapping, threshold=calibration['threshold'], n_examples=len(rows), example_ids=example_ids, example_content_digest=example_digest, accuracy=calibration['metrics']['audit']['accuracy'], per_label_stats=audit_stats, calibration_metadata={'method': calibration['method'], 'seed': calibration['seed'], 'criterion': calibration['criterion'], 'entailment_label': calibration['entailment_label'], 'calibration_ids': sorted(calibration['calibration_ids']), 'audit_ids': sorted(calibration['audit_ids']), 'calibration_content_digest': normalized_dataset_content_hash(calibration_rows), 'audit_content_digest': normalized_dataset_content_hash(audit_rows), 'support': calibration['support'], 'metrics': calibration['metrics'], 'dataset_content_hash': dataset_hash}, code_hash=CORE_SOURCE_SHA256, eval_core_hash=EVAL_CORE_SOURCE_SHA256)
    existing = NLI_CALIBRATION_ARTIFACT.get(language) if isinstance(NLI_CALIBRATION_ARTIFACT, dict) else None
    payload = dict(NLI_CALIBRATION_ARTIFACT) if isinstance(NLI_CALIBRATION_ARTIFACT, dict) else {}; payload[language] = artifact; write_json_atomic(NLI_CALIBRATION_PATH, payload)
    verified = verify_nli_calibration_artifact(artifact, model_id=CONFIG['nli_model_id'], model_revision=CONFIG['nli_model_revision'], dataset_id=dataset_id, dataset_revision=dataset_revision, dataset_content_hash=dataset_hash, language=language, split=split, code_hash=CORE_SOURCE_SHA256, eval_core_hash=EVAL_CORE_SOURCE_SHA256, dataset_examples=rows, label_mapping=label_mapping or CONFIG.get('nli_dataset_label_mapping'))
    if not verified.get('enabled'): raise RuntimeError('nli_calibration_self_verification_failed')
    quality = nli_calibration_quality(artifact, min_accuracy=CONFIG['nli_min_accuracy'], min_entailment_precision=CONFIG['nli_min_entailment_precision'], min_entailment_recall=CONFIG['nli_min_entailment_recall'], min_per_label_support=CONFIG['nli_min_per_label_support'], min_accuracy_lower=CONFIG['nli_min_accuracy_lower'], min_entailment_precision_lower=CONFIG['nli_min_entailment_precision_lower'], min_entailment_recall_lower=CONFIG['nli_min_entailment_recall_lower'])
    if not quality.get('enabled'): raise RuntimeError(f"nli_calibration_quality_failed:{language}")
    NLI_CALIBRATION_ARTIFACT[language] = artifact
    return artifact
def _verified_nli_model_indices(classifier):
    raw = getattr(getattr(getattr(classifier, 'model', None), 'config', None), 'id2label', None)
    if not isinstance(raw, dict): raise RuntimeError('nli_model_id2label_unavailable')
    normalized = {}
    for key, value in raw.items():
        key_text = str(key); key_text = f'LABEL_{key_text}' if key_text.isdigit() else key_text
        normalized[key_text] = str(value).casefold()
    if validate_nli_label_mapping(normalized, kind='model')['valid'] is False: raise RuntimeError('nli_model_id2label_mapping_mismatch')
    indices = {}
    for key, value in normalized.items():
        index = int(key.rsplit('_', 1)[1]); indices[key] = index; indices[key.casefold()] = index; indices[value] = index
    return indices
NLI_SHARED_PIPELINE = None
NLI_SHARED_MODEL_INDICES = None
NLI_SHARED_DEVICE = -1  # CPU offload permits the one evaluator to coexist with Qwen without GPU residency.
def _nli_memory_gate():
    if NLI_SHARED_DEVICE != -1 and globals().get("model") is not None and torch.cuda.is_available():
        raise RuntimeError("nli_qwen_gpu_coexistence")
    return {"device": NLI_SHARED_DEVICE, "qwen_gpu_resident": bool(globals().get("model") is not None and torch.cuda.is_available()), "offloaded": NLI_SHARED_DEVICE == -1}
def load_shared_nli_pipeline():
    global NLI_SHARED_PIPELINE, NLI_SHARED_MODEL_INDICES
    _nli_memory_gate()
    if NLI_SHARED_PIPELINE is None:
        from transformers import pipeline
        NLI_SHARED_PIPELINE = pipeline("text-classification", model=CONFIG["nli_model_id"], revision=CONFIG["nli_model_revision"], tokenizer=CONFIG["nli_model_id"], device=NLI_SHARED_DEVICE)
        NLI_SHARED_MODEL_INDICES = _verified_nli_model_indices(NLI_SHARED_PIPELINE)
    return NLI_SHARED_PIPELINE, NLI_SHARED_MODEL_INDICES
def release_shared_nli_pipeline():
    global NLI_SHARED_PIPELINE, NLI_SHARED_MODEL_INDICES
    NLI_SHARED_PIPELINE = None; NLI_SHARED_MODEL_INDICES = None; gc.collect()
    if torch.cuda.is_available(): torch.cuda.empty_cache()
    MEMORY_SNAPSHOTS.append(gpu_snapshot("shared_nli_pipeline_released"))
def load_or_create_nli_calibration(language, classifier=None, model_indices=None):
    provenance = CONFIG['nli_calibration_provenance'][language]
    from datasets import load_dataset
    all_rows, provenance = load_nli_dataset_rows(load_dataset, provenance)
    rows = list(select_stratified_nli_rows(all_rows, max_rows=int(CONFIG.get('nli_calibration_examples', 256)), seed=SEED))
    if len(rows) < 6 or len({int(row['label']) for row in rows}) < 3:
        raise RuntimeError('nli_calibration_stratified_sample_incomplete')
    selected_hash = normalized_dataset_content_hash(rows)
    provenance = {**provenance, 'dataset_content_hash': selected_hash}
    entry = NLI_CALIBRATION_ARTIFACT.get(language) if isinstance(NLI_CALIBRATION_ARTIFACT, dict) else None
    if classifier is None or model_indices is None:
        classifier, model_indices = load_shared_nli_pipeline()
    label_map = DATASET_LABEL_MAPPING
    def predictor(premise, hypothesis):
        scores = classifier({'text': premise, 'text_pair': hypothesis}, top_k=None); values = scores[0] if scores and isinstance(scores[0], list) else scores; best = max(values, key=lambda item: float(item['score']))
        if CONFIG.get('nli_model_label_mapping') != MODEL_LABEL_MAPPING:
            raise RuntimeError('nli_model_label_mapping_mismatch')
        indexed = {model_indices.get(str(item['label']).casefold(), model_indices.get(str(item['label']), -1)): float(item['score']) for item in values}
        return {'label': {0: 2, 1: 1, 2: 0}.get(model_indices.get(str(best['label']).casefold(), model_indices.get(str(best['label']), -1)), -1), 'entailment_probability': indexed.get(2, 0.0)}
    expected = verify_nli_calibration_artifact(entry, model_id=provenance['model_id'], model_revision=provenance['model_revision'], dataset_id=provenance['dataset_id'], dataset_revision=provenance['dataset_revision'], dataset_content_hash=provenance['dataset_content_hash'], language=language, split=provenance['split'], code_hash=provenance['code_hash'], eval_core_hash=provenance['eval_core_hash'], dataset_examples=rows, label_mapping=provenance.get('dataset_label_mapping'), fresh_predictor=lambda row: predictor(row['premise'], row['hypothesis']), calibration_seed=SEED, calibration_min_support=int(CONFIG['nli_calibration_min_support']), calibration_bootstrap=int(CONFIG['nli_calibration_bootstrap']), calibration_criterion='f1', calibration_entailment_label=0) if entry else {'enabled': False}
    if expected.get('enabled'):
        quality = nli_calibration_quality(entry, min_accuracy=CONFIG['nli_min_accuracy'], min_entailment_precision=CONFIG['nli_min_entailment_precision'], min_entailment_recall=CONFIG['nli_min_entailment_recall'], min_per_label_support=CONFIG['nli_min_per_label_support'], min_accuracy_lower=CONFIG['nli_min_accuracy_lower'], min_entailment_precision_lower=CONFIG['nli_min_entailment_precision_lower'], min_entailment_recall_lower=CONFIG['nli_min_entailment_recall_lower'])
        if not quality.get('enabled'): raise RuntimeError(f"nli_calibration_quality_failed:{language}")
        return entry
    return create_nli_calibration_artifact(language, rows, predictor, dataset_id=provenance['dataset_id'], dataset_revision=provenance['dataset_revision'], split=provenance['split'], label_mapping=label_map)
def _load_calibrated_nli_model(**kwargs):
    if MODEL_LABEL_MAPPING != {'LABEL_0': 'contradiction', 'LABEL_1': 'neutral', 'LABEL_2': 'entailment'}: raise RuntimeError('nli_model_label_mapping_mismatch')
    load_shared_nli_pipeline()  # validate the shared lifecycle before returning a non-retaining evaluator
    def score(premise, hypothesis):
        classifier, model_indices = load_shared_nli_pipeline()
        output = classifier({'text': premise, 'text_pair': hypothesis}, top_k=None)
        labels = {model_indices.get(str(item['label']).casefold(), model_indices.get(str(item['label']), -1)): float(item['score']) for item in (output[0] if output and isinstance(output[0], list) else output)}
        return labels.get(2, 0.0)
    return score
NLI_RUNTIME_BY_LANGUAGE = {}
CITATION_NLI_EVALUATORS = {}
NLI_THRESHOLDS = {}
def ensure_nli_calibration_artifacts():
    classifier, model_indices = load_shared_nli_pipeline()
    write_json(RUN / 'artifacts' / 'nli_memory_gate.json', _nli_memory_gate())
    for _language in CONFIG['evaluation_languages']:
        try:
            load_or_create_nli_calibration(_language, classifier, model_indices)
        except Exception as _calibration_error:
            write_json(RUN / 'artifacts' / 'nli_calibration_gate.json', {'status': 'blocked', 'language': _language, 'reason': str(_calibration_error)})
            raise RuntimeError(f'nli_calibration_gate_failed:{_language}')
ensure_nli_calibration_artifacts()
for _language in CONFIG['evaluation_languages']:
    _runtime_provenance = dict(CONFIG['nli_calibration_provenance'][_language])
    _runtime_provenance['dataset_content_hash'] = NLI_CALIBRATION_ARTIFACT[_language].get('dataset_content_hash')
    _runtime = load_nli_runtime_evaluator(_language, calibration_artifact=NLI_CALIBRATION_ARTIFACT, calibration_provenance={_language: _runtime_provenance}, model_loader=_load_calibrated_nli_model, model_id=CONFIG['nli_model_id'], revision=CONFIG['nli_model_revision'])
    NLI_RUNTIME_BY_LANGUAGE[_language] = _runtime
    CITATION_NLI_EVALUATORS[_language] = _runtime.get('evaluator') if _runtime.get('enabled') else None
    NLI_THRESHOLDS[_language] = float(_runtime.get('threshold', 0.5))
if CONFIG.get('require_citation_nli') and any(not NLI_RUNTIME_BY_LANGUAGE[_language].get('enabled') for _language in CONFIG['evaluation_languages']):
    write_json(RUN / 'artifacts' / 'nli_calibration_gate.json', {'status': 'blocked', 'reason': 'missing_or_unverified_language_calibration', 'languages': NLI_RUNTIME_BY_LANGUAGE})
    raise RuntimeError('nli_calibration_gate_failed_before_generation')
NLI_RUNTIME_BY_LANGUAGE = {language: {key: value for key, value in runtime.items() if key != 'evaluator'} for language, runtime in NLI_RUNTIME_BY_LANGUAGE.items()}
def validate_response(parsed, evidence, language=None):
    ledger = build_evidence_ledger(evidence); ledger_ids = {x["evidence_id"] for x in ledger}; clean_payload = {key: value for key, value in parsed.items() if key not in INTERNAL_RESPONSE_METADATA_KEYS} if isinstance(parsed, dict) else parsed; schema = validate_final_response(clean_payload, ledger_ids) if isinstance(parsed, dict) and parsed.get("parse_status") in ACCEPTED_PARSE_STATUSES else {"valid": False, "reasons": ["schema_invalid"], "quarantine": True}; narrative = str(parsed.get("counter_narrative", "")).strip() if isinstance(parsed, dict) else ""; cited_ids = parsed.get("cited_evidence_ids", []) if isinstance(parsed, dict) else []; ids = validate_evidence_ids(cited_ids, ledger); tokens, resolved, unknown = resolve_citation_tokens(narrative, ledger); reasons = list(schema.get("reasons", [])); factual_claims = parsed.get("factual_claims", []) if isinstance(parsed, dict) and isinstance(parsed.get("factual_claims", []), list) else []
    if not narrative: reasons.append("empty_narrative")
    declared_tokens = {value for value in cited_ids if isinstance(value, str)} if isinstance(cited_ids, list) else set()
    if not isinstance(cited_ids, list) or any(not isinstance(value, str) for value in cited_ids) or set(tokens) != declared_tokens: reasons.append("inline_and_declared_evidence_mismatch")
    if unknown: reasons.append("unknown_inline_evidence_ids")
    if "build_claim_citation_records" not in globals() or "aggregate_citation_support" not in globals():
        claim_records = []; support = {"entailment_status": "helper_unavailable", "pass": False, "abstention": False, "format_compliance": {"valid": False}, "citation_precision": None, "syntactic_citation_precision": None, "claim_citation_recall": None, "overcitation_count": 0, "citation_entailment": None}
    else:
        claim_records = build_claim_citation_records(narrative, ledger, factual_claims=factual_claims, safe_non_factual_validator=safe_abstention_validator)
        evaluator = globals().get('CITATION_NLI_EVALUATORS', {}).get(language)
        support = aggregate_citation_support(claim_records, evaluator=evaluator, threshold=float(globals().get('NLI_THRESHOLDS', {}).get(language, 0.5)), require_nli=bool(globals().get("CONFIG", {}).get("require_citation_nli", True)))
    if support["entailment_status"] == "helper_unavailable": reasons.append("citation_helpers_unavailable")
    if support["entailment_status"] == "unavailable": reasons.append("citation_entailment_unavailable")
    if support["entailment_status"] == "evaluator_error": reasons.append("citation_entailment_evaluator_error")
    if support["entailment_status"] == "scored_incomplete": reasons.append("citation_entailment_incomplete")
    if not support["pass"] and not support["abstention"] and support["entailment_status"] == "scored": reasons.append("citation_support_failed")
    status = support.get("entailment_status"); mode = "format_and_entailment" if status == "scored" else "format_only_abstention" if support.get("abstention") else "evaluator_error" if status == "evaluator_error" else "format_only_nli_unavailable"
    return {"citation_count": len(tokens), "valid_citation_count": len(resolved), "unknown_evidence_ids": sorted(set(ids["unknown"]) | set(unknown)), "claim_level_citations": claim_records, "citation_entailment": support.get("citation_entailment"), "entailment_mean": support.get("entailment_mean"), "evaluated_claim_entailment_mean": support.get("evaluated_claim_entailment_mean"), "evaluated_claim_count": support.get("evaluated_claim_count"), "incomplete_claim_count": support.get("incomplete_claim_count"), "citation_entailment_status": status, "citation_validation_mode": mode, "citation_format_compliance": support.get("format_compliance"), "citation_precision": support.get("citation_precision"), "syntactic_citation_precision": support.get("syntactic_citation_precision"), "claim_citation_recall": support.get("claim_citation_recall"), "citation_recall": support.get("citation_recall"), "citation_necessity": support.get("citation_necessity"), "necessary_citation_count": support.get("necessary_citation_count"), "overcitation_count": support.get("overcitation_count"), "schema_valid": schema["valid"], "validation_reasons": sorted(set(reasons)), "pass": schema["valid"] and ids["valid"] and not reasons and support["pass"]}
def authenticated_checkpoint_evidence(record, variant):
    if variant not in {\"kg_rag\", \"mp_kg_rag\"}: return []
    bundle = EVIDENCE_CACHE.get(str(record[\"ID\"]))
    if not isinstance(bundle, dict) or not isinstance(bundle.get(\"evidence\"), list): raise RuntimeError(\"resume_identity_mismatch\")
    return bundle[\"evidence\"]
def checkpoint_raw_envelope(final_raw, variant, mp_fields):
    if not isinstance(final_raw, str) or variant not in VARIANT_COLUMNS: raise RuntimeError(\"resume_identity_mismatch\")
    expected_mp_keys = {\"perspective_rationale\", \"perspective_parse_rate\", \"mp_perspective_outputs\", \"mp_response_plan\", \"mp_plan_raw_output\"}
    if variant == \"mp_kg_rag\" and set(mp_fields) != expected_mp_keys: raise RuntimeError(\"resume_identity_mismatch\")
    if variant != \"mp_kg_rag\" and mp_fields: raise RuntimeError(\"resume_identity_mismatch\")
    return json.dumps({\"schema_version\": \"mpkg-rag.checkpoint-envelope.v2\", \"variant\": variant, \"final_raw_output\": final_raw, \"mp_fields\": mp_fields}, ensure_ascii=False, sort_keys=True, separators=(\",\", \":\"))
def checkpoint_materialization(payload, evidence, variant, language=None):
    if variant not in VARIANT_COLUMNS: raise RuntimeError(\"resume_identity_mismatch\")
    narrative = str(payload.get(\"counter_narrative\", \"\")).strip()
    verification = validate_response(payload, evidence, language=language) if language is not None else validate_response(payload, evidence)
    output_language = LANGUAGE_DETECTOR(narrative) if globals().get('LANGUAGE_DETECTOR') is not None else None
    verification['output_language'] = output_language
    if language is not None and output_language is not None and output_language != language:
        verification.setdefault('validation_reasons', []).append('output_language_mismatch'); verification['pass'] = False
    if not verification.get('schema_valid'): raise SchemaValidationQuarantine({'reason': 'final_schema_invalid', 'stage': 'checkpoint_materialization', 'variant': variant, 'validation_reasons': verification.get('validation_reasons', [])})
    if variant in {\"kg_rag\", \"mp_kg_rag\"} and not verification.get('pass'): raise SchemaValidationQuarantine({'reason': 'citation_support_failed', 'stage': 'checkpoint_materialization', 'variant': variant, 'validation_reasons': verification.get('validation_reasons', [])})
    return {\"response\": narrative, \"parsed_counter_narrative\": narrative, VARIANT_COLUMNS[variant]: narrative, \"evidence\": evidence, \"evidence_ledger\": build_evidence_ledger(evidence), \"verification\": verification, \"input_language\": language, \"output_language\": output_language}
def checkpoint_reasoning_trace(final_raw):
    trace_builder = globals().get(\"qwen_generation_trace\")
    if callable(trace_builder): return trace_builder(final_raw)
    raw = str(final_raw or \"\"); match = re.search(r\"<think>\\s*(.*?)\\s*</think>\", raw, flags=re.S | re.I)
    return {\"reasoning_content\": match.group(1).strip() if match else \"\", \"final_content\": (raw[match.end():].strip() if match else raw.strip()), \"thinking_status\": \"complete\" if match else \"not_emitted\", \"reasoning_truncated\": False, \"reasoning_token_count\": None, \"answer_token_count\": None, \"raw_generation\": raw}
def canonical_checkpoint_row(record, variant, final_raw, parsed, evidence, mp_fields):
    detector = globals().get('LANGUAGE_DETECTOR')
    language = record.get('language') or record.get('Language') or (detector(record.get('Text', '')) if detector is not None else None)
    materialized = checkpoint_materialization(parsed, evidence, variant, language=language)
    trace_builder = globals().get(\"qwen_generation_trace\"); reasoning_trace = trace_builder(final_raw) if callable(trace_builder) else {\"reasoning_content\": \"\", \"final_content\": str(final_raw or \"\"), \"thinking_status\": \"not_emitted\", \"reasoning_truncated\": False, \"reasoning_token_count\": None, \"answer_token_count\": None, \"raw_generation\": str(final_raw or \"\")}
    return {\"ID\": str(record[\"ID\"]), \"Text\": record[\"Text\"], \"Category\": record[\"Category\"], \"Target\": record[\"Target\"], \"Counter Narrative\": record[\"Counter Narrative\"], \"variant\": variant, \"split_name\": CONFIG[\"split_name\"], \"input_text_sha256\": record[\"input_text_sha256\"], \"config_hash\": CONFIG_HASH, \"prompt_template_hash\": PROMPT_TEMPLATE_HASH, \"checkpoint_identity\": checkpoint_identity(record, variant), \"parse_status\": parsed.get(\"parse_status\", \"initial\"), \"reasoning_trace\": reasoning_trace, \"raw_output\": checkpoint_raw_envelope(final_raw, variant, mp_fields), **materialized, **mp_fields}
def baseline_response(post, target, few_shot=False):
    examples = \"\" if not few_shot else \"\\nFrozen examples (do not copy identities or answers):\\n\" + \"\\n\".join(f\"Example post: {item['post']}\\nTarget: {item['target']}\\nSafe response: {item['response']}\" for item in FEW_SHOT_EXAMPLES)
    prompt_revision = FEW_SHOT_PROMPT_REVISION if few_shot else \"zero-shot.v1\"; schema_tail = \"JSON schema (must be followed exactly): \" + json.dumps(FINAL_RESPONSE_SCHEMA, ensure_ascii=False, sort_keys=True)
    def builder(selected, payload):
        return [{\"role\": \"system\", \"content\": f\"Return exactly one JSON object matching the final response schema. Prompt variant: {prompt_revision}.\"}, {\"role\": \"user\", \"content\": f\"Post: {post}\\nTarget: {target}{examples}\\nReturn counter_narrative, cited_evidence_ids, factual_claims, safety_notes. Do not add wrappers or extra keys.\\n{schema_tail}\"}]
    if callable(globals().get(\"_adaptive_messages\")):
        prompt, _ = _adaptive_messages([], builder, (\"\",), CONFIG[\"answer_max_new_tokens\"])
    else:
        prompt = builder([], \"\")
    raw = generate_batch([prompt], CONFIG[\"answer_max_new_tokens\"], temperature=0.6, enable_thinking=CONFIG.get(\"thinking_enabled\", True), output_schema=FINAL_RESPONSE_SCHEMA)[0]; parsed, final_raw, status = parse_with_one_repair(raw, prompt[1][\"content\"], validate_final_response, FINAL_RESPONSE_SCHEMA); result = parsed if parsed is not None else {\"counter_narrative\": \"\", \"cited_evidence_ids\": [], \"factual_claims\": [], \"safety_notes\": [], \"quarantine\": True}; result[\"parse_status\"] = status; result[\"few_shot\"] = bool(few_shot); result[\"few_shot_prompt_revision\"] = prompt_revision; return result, final_raw
def generate_variant(record, variant):
    rid, evidence = str(record[\"ID\"]), (EVIDENCE_CACHE[str(record[\"ID\"])][\"evidence\"] if variant in {\"kg_rag\", \"mp_kg_rag\"} else []); post, target, category = record[\"Text\"], record[\"Target\"], record[\"Category\"]
    if variant == \"qwen_zero_shot\": parsed, raw = baseline_response(post, target)
    elif variant == \"qwen_few_shot\": parsed, raw = baseline_response(post, target, True)
    elif variant == \"kg_rag\": parsed, raw = generate_final_counter_narrative(post, target, evidence, {\"claim_focus\": \"Address the claim respectfully.\", \"selected_evidence_ids\": [x[\"evidence_id\"] for x in build_evidence_ledger(evidence)], \"response_steps\": [], \"tone\": \"empathetic\", \"factual_constraints\": [], \"safety_constraints\": [\"avoid amplification\"]})
    elif variant == \"mp_kg_rag\":
        agents = run_perspectives(post, target, category, evidence); plan, plan_raw = synthesize_plan(post, target, category, evidence, agents)
        parsed, raw = generate_final_counter_narrative(post, target, evidence, plan)
    else: raise ValueError(variant)
    if not isinstance(parsed, dict) or parsed.get(\"parse_status\") not in ACCEPTED_PARSE_STATUSES or parsed.get(\"quarantine\"):
        raise SchemaValidationQuarantine({\"reason\": \"generation_schema_invalid\", \"stage\": variant, \"parse_status\": parsed.get(\"parse_status\") if isinstance(parsed, dict) else None})
    narrative = str(parsed.get(\"counter_narrative\", \"\")); ledger = build_evidence_ledger(evidence); extra = {\"perspective_rationale\": {p[\"perspective\"]: p[\"perspective_rationale\"] for p in agents}, \"perspective_parse_rate\": sum(p[\"structured_output\"].get(\"parse_status\") in ACCEPTED_PARSE_STATUSES for p in agents) / max(1, len(agents)), \"mp_perspective_outputs\": agents, \"mp_response_plan\": {**plan, \"reasoning_trace\": qwen_generation_trace(plan_raw)}, \"mp_plan_raw_output\": plan_raw} if variant == \"mp_kg_rag\" else {}
    try:
        return canonical_checkpoint_row(record, variant, raw, parsed, evidence, extra)
    except SchemaValidationQuarantine as failure:
        if variant not in {\"kg_rag\", \"mp_kg_rag\"}: raise
        repaired, repaired_raw = grounding_repair_response(post, target, evidence, plan if variant == \"mp_kg_rag\" else {\"claim_focus\": \"Use only evidence.\"}, [str(failure)])
        return canonical_checkpoint_row(record, variant, repaired_raw, repaired, evidence, extra)
def revalidate_checkpoint_row(row, record, variant):
    if not isinstance(row, dict) or not isinstance(row.get(\"ID\"), (str, int)): raise RuntimeError(\"resume_identity_mismatch\")
    validate_checkpoint_identity(row.get(\"checkpoint_identity\"), checkpoint_identity(record, variant))
    if not isinstance(row.get(\"raw_output\"), str) or row.get(\"parse_status\") not in ACCEPTED_PARSE_STATUSES: raise RuntimeError(\"resume_identity_mismatch\")
    try:
        envelope = parse_json_object(row[\"raw_output\"])
        if set(envelope) != {\"schema_version\", \"variant\", \"final_raw_output\", \"mp_fields\"} or envelope.get(\"schema_version\") != \"mpkg-rag.checkpoint-envelope.v2\" or envelope.get(\"variant\") != variant or not isinstance(envelope.get(\"final_raw_output\"), str) or not isinstance(envelope.get(\"mp_fields\"), dict): raise RuntimeError(\"resume_identity_mismatch\")
        payload = parse_json_object(envelope[\"final_raw_output\"]); payload[\"parse_status\"] = row[\"parse_status\"]
        authenticated_evidence = authenticated_checkpoint_evidence(record, variant)
        canonical = canonical_checkpoint_row(record, variant, envelope[\"final_raw_output\"], payload, authenticated_evidence, envelope[\"mp_fields\"])
    except Exception as exc:
        if isinstance(exc, RuntimeError) and str(exc) == \"resume_identity_mismatch\": raise
        raise RuntimeError(\"resume_identity_mismatch\")
    if row != canonical: raise RuntimeError(\"resume_identity_mismatch\")
    return row
def checkpoint_lock(path):
    return Path(path).with_name(Path(path).name + \".lock\")
def read_checkpoint_rows_locked(path):
    path = Path(path); lock_path = checkpoint_lock(path); lock_path.parent.mkdir(parents=True, exist_ok=True)
    with lock_path.open(\"a+\", encoding=\"utf-8\") as lock:
        fcntl.flock(lock.fileno(), fcntl.LOCK_EX)
        try: return load_jsonl(path)
        finally: fcntl.flock(lock.fileno(), fcntl.LOCK_UN)
def append_checkpoint_row(path, row):
    path = Path(path); lock_path = checkpoint_lock(path); lock_path.parent.mkdir(parents=True, exist_ok=True)
    with lock_path.open(\"a+\", encoding=\"utf-8\") as lock:
        fcntl.flock(lock.fileno(), fcntl.LOCK_EX)
        try:
            existing_rows = load_jsonl(path); record_id = str(row.get(\"ID\", \"\"))
            for existing in existing_rows:
                if str(existing.get(\"ID\", \"\")) != record_id: continue
                if existing == row: return False
                raise RuntimeError(\"duplicate_checkpoint_conflict\")
            with path.open(\"a\", encoding=\"utf-8\") as stream: stream.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + \"\\n\"); stream.flush(); os.fsync(stream.fileno())
            return True
        finally: fcntl.flock(lock.fileno(), fcntl.LOCK_UN)
def load_checkpoint_rows(path, data, variant):
    previous = read_checkpoint_rows_locked(path); expected = {str(row[\"ID\"]): row for row in data.to_dict(\"records\")}; seen = set()
    for row in previous:
        rid = str(row.get(\"ID\", \"\"))
        if rid in seen: raise RuntimeError(\"duplicate_checkpoint_ids\")
        if rid not in expected: raise RuntimeError(\"unknown_checkpoint_id\")
        revalidate_checkpoint_row(row, expected[rid], variant); seen.add(rid)
    return previous, {str(row[\"ID\"]): row for row in previous}
def row_quarantine_row(record, variant, audit):
    audit = dict(audit or {}); reason = str(audit.get(\"reason\", \"row_quarantine\")); prompt_quarantine = reason.startswith(\"prompt_budget\") or reason == \"repair_prompt_budget\"
    return {\"ID\": str(record[\"ID\"]), \"Text\": record.get(\"Text\"), \"Category\": record.get(\"Category\"), \"Target\": record.get(\"Target\"), \"Counter Narrative\": record.get(\"Counter Narrative\"), \"variant\": variant, \"split_name\": CONFIG[\"split_name\"], \"parsed_counter_narrative\": None, \"response\": None, \"raw_output\": None, \"parse_status\": \"quarantined\", \"prompt_quarantine\": prompt_quarantine, \"prompt_quarantine_reason\": reason if prompt_quarantine else None, \"prompt_budget_audit\": audit, \"generation_quarantine_reason\": reason, \"verification\": {\"pass\": False, \"validation_reasons\": [reason]}, \"evidence\": [], \"evidence_ledger\": []}
def prompt_budget_quarantine_row(record, variant, audit):
    return row_quarantine_row(record, variant, audit)
def _excel_cell(value):
    if isinstance(value, (dict, list, tuple)): value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    if value is None: return None
    return str(value)[:32767]
def _style_production_workbook(path):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    workbook = load_workbook(path); header_fill = PatternFill(\"solid\", fgColor=\"1F4E78\"); header_font = Font(color=\"FFFFFF\", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = \"A2\"; sheet.auto_filter.ref = sheet.dimensions; sheet.sheet_view.showGridLines = False
        for cell in sheet[1]: cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal=\"center\", vertical=\"center\", wrap_text=True)
        sheet.row_dimensions[1].height = 32
        for column_cells in sheet.columns:
            letter = column_cells[0].column_letter; header = str(column_cells[0].value or \"\").casefold(); width = 14
            if any(key in header for key in [\"text\", \"narrative\", \"reasoning\", \"evidence\", \"perspective\", \"plan\", \"verification\", \"raw\"]): width = 48
            elif header in {\"id\", \"status\", \"variant\", \"category\", \"target\"}: width = 18
            sheet.column_dimensions[letter].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row: cell.alignment = Alignment(vertical=\"top\", wrap_text=True)
    workbook.save(path)
def export_production_workbook(frame):
    output_path = RUN / \"exports\" / \"dataset_with_all_rag_counter_narratives.xlsx\"
    base_columns = [\"ID\", \"Text\", \"Category\", \"Target\", \"Counter Narrative\"]; outputs = dataset[base_columns].copy(); outputs[\"ID\"] = outputs[\"ID\"].astype(str)
    trace_rows = {}; variant_output_columns = {\"qwen_zero_shot\": \"zero-shot-counter-narrative\", \"qwen_few_shot\": \"few-shot-counter-narrative\", \"kg_rag\": \"kg-rag-counter-narrative\", \"mp_kg_rag\": \"mp-kg-rag-counter-narrative\"}
    for variant, output_column in variant_output_columns.items():
        selected = frame[frame.variant.eq(variant)].copy(); selected[\"ID\"] = selected[\"ID\"].astype(str); indexed = selected.drop_duplicates(\"ID\").set_index(\"ID\")
        outputs[output_column] = outputs[\"ID\"].map(indexed[\"parsed_counter_narrative\"]); outputs[variant + \"-status\"] = outputs[\"ID\"].map(indexed[\"parse_status\"])
        for row in selected.to_dict(\"records\"):
            trace = row.get(\"reasoning_trace\") if isinstance(row.get(\"reasoning_trace\"), dict) else {}
            trace_rows.setdefault(variant, []).append({\"ID\": str(row[\"ID\"]), \"reasoning_content\": trace.get(\"reasoning_content\"), \"thinking_status\": trace.get(\"thinking_status\"), \"reasoning_truncated\": trace.get(\"reasoning_truncated\"), \"reasoning_token_count\": trace.get(\"reasoning_token_count\"), \"answer_token_count\": trace.get(\"answer_token_count\"), \"parse_status\": row.get(\"parse_status\"), \"final_counter_narrative\": row.get(\"parsed_counter_narrative\"), \"evidence_ledger\": _excel_cell(row.get(\"evidence_ledger\")), \"verification\": _excel_cell(row.get(\"verification\")), \"mp_perspective_outputs\": _excel_cell(row.get(\"mp_perspective_outputs\")), \"mp_response_plan\": _excel_cell(row.get(\"mp_response_plan\")), \"raw_generation\": _excel_cell(trace.get(\"raw_generation\"))})
    evidence_rows = []
    for row in frame[frame.variant.isin([\"kg_rag\", \"mp_kg_rag\"])].to_dict(\"records\"):
        for evidence in row.get(\"evidence_ledger\", []) if isinstance(row.get(\"evidence_ledger\"), list) else []: evidence_rows.append({\"ID\": str(row[\"ID\"]), \"variant\": row[\"variant\"], **{key: _excel_cell(value) for key, value in evidence.items()}})
    quality = frame.groupby(\"variant\", dropna=False).agg(rows=(\"ID\", \"size\"), unique_ids=(\"ID\", \"nunique\"), accepted=(\"parse_status\", lambda values: int(sum(value in ACCEPTED_PARSE_STATUSES for value in values)))).reset_index(); quality[\"parse_rate\"] = quality[\"accepted\"] / quality[\"rows\"].clip(lower=1)
    manifest = pd.DataFrame([{\"run_name\": RUN_NAME, \"model\": CONFIG[\"generator_model\"], \"model_revision\": CONFIG[\"generator_model_revision\"], \"load_in_4bit\": CONFIG[\"load_in_4bit\"], \"thinking_enabled\": CONFIG[\"thinking_enabled\"], \"input_rows\": len(dataset), \"input_sha256\": hashlib.sha256(Path(CONFIG[\"dataset_xlsx\"]).read_bytes()).hexdigest(), \"config_hash\": CONFIG_HASH, \"prompt_template_hash\": PROMPT_TEMPLATE_HASH}])
    with pd.ExcelWriter(output_path, engine=\"openpyxl\") as writer:
        outputs.to_excel(writer, sheet_name=\"Outputs\", index=False)
        sheet_names = {\"qwen_zero_shot\": \"Zero-Shot Trace\", \"qwen_few_shot\": \"Few-Shot Trace\", \"kg_rag\": \"KG-RAG Trace\", \"mp_kg_rag\": \"MP-KG-RAG Trace\"}
        for variant, sheet_name in sheet_names.items(): pd.DataFrame(trace_rows.get(variant, [])).to_excel(writer, sheet_name=sheet_name, index=False)
        pd.DataFrame(evidence_rows).to_excel(writer, sheet_name=\"Evidence Ledger\", index=False); manifest.to_excel(writer, sheet_name=\"Run Manifest\", index=False); quality.to_excel(writer, sheet_name=\"Quality Summary\", index=False)
    _style_production_workbook(output_path); return output_path
def generate_all_variants(data):
    global GENERATION_QUARANTINE_SUMMARY
    outputs = []; generation_quarantines = []
    for variant in CONFIG[\"generation_variants\"]:
        checkpoint = RUN / \"checkpoints\" / (\"mp_kg_rag_rows.jsonl\" if variant == \"mp_kg_rag\" else f\"{variant}_rows.jsonl\"); previous, by_id = load_checkpoint_rows(checkpoint, data, variant)
        for record in tqdm(data.to_dict(\"records\"), desc=f\"Generate {variant}\"):
            if str(record[\"ID\"]) in by_id: outputs.append(by_id[str(record[\"ID\"])]); continue
            try:
                row = generate_variant(record, variant)
            except GenerationRowQuarantine as failure:
                audit = {\"record_id\": str(record[\"ID\"]), \"variant\": variant, \"quarantine_type\": type(failure).__name__, **failure.audit}; generation_quarantines.append(audit); row = row_quarantine_row(record, variant, audit); outputs.append(row); continue
            append_checkpoint_row(checkpoint, row); outputs.append(row)
    type_counts = {}; reason_counts = {}; variant_counts = {}
    for audit in generation_quarantines:
        quarantine_type = str(audit.get(\"quarantine_type\", \"GenerationRowQuarantine\")); reason = str(audit.get(\"reason\", \"row_quarantine\")); variant_name = str(audit.get(\"variant\", \"unknown\"))
        type_counts[quarantine_type] = type_counts.get(quarantine_type, 0) + 1
        reason_counts[reason] = reason_counts.get(reason, 0) + 1
        variant_counts[variant_name] = variant_counts.get(variant_name, 0) + 1
    GENERATION_QUARANTINE_SUMMARY = {\"count\": len(generation_quarantines), \"counts_by_quarantine_type\": type_counts, \"counts_by_reason\": reason_counts, \"counts_by_variant\": variant_counts}
    write_json(RUN / \"artifacts\" / \"generation_quarantine.json\", {**GENERATION_QUARANTINE_SUMMARY, \"rows\": generation_quarantines})
    prompt_rows = [row for row in generation_quarantines if str(row.get(\"reason\", \"\")).startswith(\"prompt_budget\") or row.get(\"reason\") == \"repair_prompt_budget\"]
    write_json(RUN / \"artifacts\" / \"prompt_budget_quarantine.json\", {\"count\": len(prompt_rows), \"rows\": prompt_rows, \"reason\": \"prompt_budget_irreducible\"})
    frame = pd.DataFrame(outputs); frame.to_json(RUN / \"exports\" / \"all_variant_outputs.jsonl\", orient=\"records\", lines=True, force_ascii=False); frame.to_excel(RUN / \"exports\" / \"all_variant_outputs.xlsx\", index=False); frame[frame.variant.eq(\"mp_kg_rag\")].to_excel(RUN / \"exports\" / \"dataset_with_mp_kg_rag_counter_narratives.xlsx\", index=False); export_production_workbook(frame); return frame
def load_qwen_for_generation(): return load_qwen_for_extraction()
release_shared_nli_pipeline()
load_qwen_for_generation(); print("Generator model:", CONFIG["generator_model"], "is_loaded_in_4bit:", getattr(model, "is_loaded_in_4bit", None), "quantization_config:", getattr(getattr(model, "config", None), "quantization_config", None), "first_parameter_dtype:", next(model.parameters()).dtype)
generation_frame = generate_all_variants(dataset)
GENERATION_FRAME_MATERIALIZED = generation_frame.copy(deep=True)
unload_generator()
if \"model\" in globals() or \"tokenizer\" in globals(): raise RuntimeError(\"qwen_generation_unload_failed\")
QWEN_METRICS_MEMORY_GATE_GB = float(os.environ.get(\"QWEN_METRICS_MEMORY_GATE_GB\", \"1.0\"))
QWEN_METRICS_GPU_SNAPSHOT = gpu_snapshot(\"qwen_generation_unloaded_before_metrics\")
if float(QWEN_METRICS_GPU_SNAPSHOT.get(\"allocated_gb\", 0.0)) > QWEN_METRICS_MEMORY_GATE_GB: raise RuntimeError(\"qwen_metrics_memory_gate\")
QWEN_METRICS_GPU_GATE_PASSED = True
"""),
        code("""# 13 - Language-aware automatic metrics and corrected paired comparisons.
from rouge_score import rouge_scorer
from nltk.translate.meteor_score import meteor_score
ROUGE = rouge_scorer.RougeScorer(['rougeL'], use_stemmer=True)
LANGUAGE_DETECTOR = load_language_detector()
if LANGUAGE_DETECTOR is None: raise RuntimeError('language_evaluation_detector_unavailable')
BERTSCORE_DEVICE = \"cuda\" if QWEN_METRICS_GPU_GATE_PASSED else \"cpu\"
if \"model\" in globals() or \"tokenizer\" in globals(): raise RuntimeError(\"detoxify_model_unloaded_before_metrics\")
def _rouge(reference, hypothesis): return ROUGE.score(reference, hypothesis)['rougeL'].fmeasure
def _meteor(reference, hypothesis): return meteor_score([reference.split()], hypothesis.split())
def reference_metric_row(row):
    reference = normalize_optional_text(row["Counter Narrative"] if "Counter Narrative" in row else None)
    response = normalize_optional_text(row.get("parsed_counter_narrative"))
    if reference is None: return {'rouge_l': None, 'reference_metric_status': 'excluded_missing_reference', 'reference_metrics_excluded': True}
    if response is None: return {'rouge_l': None, 'reference_metric_status': 'missing_hypothesis', 'reference_metrics_excluded': False}
    return {'rouge_l': ROUGE.score(reference, response)['rougeL'].fmeasure, 'reference_metric_status': 'scored', 'reference_metrics_excluded': False}
def citation_metrics(row):
    verification = row.get('verification') if isinstance(row.get('verification'), dict) else {}
    if str(row.get('variant', '')).casefold() not in {'kg_rag', 'mp_kg_rag'}:
        return {name: None for name in ('citation_precision', 'citation_recall', 'citation_necessity', 'citation_entailment')} | {name: 'not_applicable_non_rag' for name in ('citation_precision_status', 'citation_recall_status', 'citation_necessity_status', 'citation_entailment_status')}
    return {'citation_precision': verification.get('citation_precision'), 'citation_recall': verification.get('citation_recall'), 'citation_necessity': verification.get('citation_necessity'), 'citation_entailment': verification.get('citation_entailment'), 'citation_precision_status': verification.get('citation_entailment_status'), 'citation_recall_status': verification.get('citation_entailment_status'), 'citation_necessity_status': verification.get('citation_entailment_status'), 'citation_entailment_status': verification.get('citation_entailment_status')}
def _detoxify(hypothesis): return float(DETOXIFY_MODEL.predict(hypothesis).get('toxicity'))
DETOXIFY_MODEL = None
DETOXIFY_DEVICE = 'cpu'
DETOXIFY_STATUS = {'status': 'unavailable', 'device': DETOXIFY_DEVICE, 'required': bool(CONFIG.get('require_detoxify', True))}
try:
    from detoxify import Detoxify
    if \"model\" in globals() or \"tokenizer\" in globals(): raise RuntimeError(\"detoxify_model_unloaded_before_metrics\")
    DETOXIFY_MODEL = Detoxify('original', device='cpu')
    DETOXIFY_STATUS = {'status': 'loaded', 'device': DETOXIFY_DEVICE, 'required': bool(CONFIG.get('require_detoxify', True)), 'model': 'original'}
except Exception as error:
    DETOXIFY_MODEL = None
    DETOXIFY_STATUS = {'status': 'unavailable', 'device': DETOXIFY_DEVICE, 'required': bool(CONFIG.get('require_detoxify', True)), 'reason': f'{type(error).__name__}:{error}'}
    # Persist the unavailable contract before the required-model fail-closed
    # raise, so an aborted run still retains an explicit audit record.
    write_json(RUN / 'artifacts' / 'detoxify_metric_contract.json', DETOXIFY_STATUS)
    if CONFIG.get('require_detoxify', True): raise RuntimeError(f'detoxify_required_unavailable:{DETOXIFY_STATUS[\"reason\"]}') from error
write_json(RUN / 'artifacts' / 'detoxify_metric_contract.json', DETOXIFY_STATUS)
_BERTSCORE_LOCAL_MODELS = {}
_BERTSCORE_CONFIG_HASH = stable_identity_hash(multilingual_bertscore_config({'model_id': CONFIG['bertscore_model_id'], 'revision': CONFIG['bertscore_model_revision']}))
def _bertscore(**kwargs):
    from huggingface_hub import snapshot_download
    from bert_score import score
    cache_key = (kwargs['model_id'], kwargs['revision'])
    if cache_key not in _BERTSCORE_LOCAL_MODELS:
        _BERTSCORE_LOCAL_MODELS[cache_key] = snapshot_download(repo_id=kwargs['model_id'], revision=kwargs['revision'])
    local_model = _BERTSCORE_LOCAL_MODELS[cache_key]
    _, _, f1 = score([kwargs['hypothesis']], [kwargs['reference']], lang=kwargs['language'], model_type=local_model, num_layers=BERTSCORE_NUM_LAYERS, device=BERTSCORE_DEVICE, verbose=False)
    return {'f1': float(f1[0].item())}
metric_rows = []
for _, row in generation_frame.iterrows():
    evaluated = evaluate_multilingual_record(row.to_dict(), language_detector=LANGUAGE_DETECTOR, rouge_scorer=_rouge, meteor_scorer=_meteor, bertscore_scorer=_bertscore, detoxify_scorer=_detoxify if DETOXIFY_MODEL is not None else None, bertscore_config={'model_id': CONFIG['bertscore_model_id'], 'revision': CONFIG['bertscore_model_revision']})
    output = {'ID': str(row['ID']), 'variant': str(row['variant']), 'response': normalize_optional_text(row.get('parsed_counter_narrative')), **{key: evaluated.get(key) for key in ['language', 'language_status', 'expected_language', 'input_language', 'reference_language', 'output_language', 'language_match', 'input_language_match', 'reference_language_match', 'script_match', 'reference_available']}}
    for metric_name, metric_value in evaluated['metrics'].items():
        output[metric_name] = metric_value.get('value'); output[f'{metric_name}_status'] = metric_value.get('status'); output[f'{metric_name}_reason'] = metric_value.get('reason')
    output.update(citation_metrics(row.to_dict()))
    metric_rows.append(output)
metrics = pd.DataFrame(metric_rows)
validate_metric_rows_unique(metric_rows)
language_summary = summarize_metric_records([{'language': row.get('language'), 'metrics': {name: {'value': row.get(name), 'status': row.get(f'{name}_status'), 'reason': row.get(f'{name}_reason')} for name in ['chrf', 'bertscore', 'rouge_l', 'meteor', 'detoxify']}} for row in metric_rows])
language_diagnostics = summarize_language_records(metric_rows)
coverage = {language: sum(row.get('bertscore_status') == 'scored' for row in metric_rows if row.get('language') == language) for language in CONFIG['evaluation_languages']}
observed_languages = {row.get('language') for row in metric_rows if row.get('language') in CONFIG['evaluation_languages']}
if any(coverage[language] < 1 for language in observed_languages): raise RuntimeError('bertscore_minimum_language_coverage_failed')
DETERMINISTIC_PLOT_KWARGS = dict(errorbar=None, bootstrap_seed=BOOTSTRAP_SEED)
write_json(RUN / 'artifacts' / 'language_metric_summary.json', {'rows': language_summary, 'language_diagnostics': language_diagnostics, 'bertscore_coverage': coverage, 'languages': CONFIG['evaluation_languages'], 'bertscore_config': multilingual_bertscore_config({'model_id': CONFIG['bertscore_model_id'], 'revision': CONFIG['bertscore_model_revision']}), 'deterministic_plot_kwargs': DETERMINISTIC_PLOT_KWARGS})
metrics.to_parquet(RUN / 'exports' / 'per_record_automatic_metrics.parquet', index=False); metrics.to_excel(RUN / 'exports' / 'per_record_automatic_metrics.xlsx', index=False)
pairwise = pairwise_metric_family(metric_rows, metrics=('chrf', 'bertscore', 'rouge_l', 'meteor', 'detoxify', 'citation_precision', 'citation_recall', 'citation_necessity', 'citation_entailment'), directions={'detoxify': 'lower'}, seed=SEED, permutations=10000)
write_json(RUN / 'artifacts' / 'pairwise_variant_comparisons.json', pairwise)
write_json(RUN / 'artifacts' / 'reference_metric_audit.json', build_reference_metric_audit(metric_rows, metrics=('chrf', 'bertscore', 'rouge_l', 'meteor', 'detoxify')))
"""),
        code("""# 14 - Bounded stratified human annotation workbook and reliability gate.
HUMAN_METRICS = ['politeness_respect', 'helpfulness', 'factuality', 'safety', 'empathy', 'fluency', 'inclusivity', 'non_confrontational', 'contextual_coherence', 'persuasiveness']
def build_human_annotation_workbook(frame, output_path):
    dataset_keys = dataset.set_index(dataset['ID'].astype(str))[['script_bucket', 'stratify_key']].to_dict('index') if 'script_bucket' in dataset.columns else {}
    source = []
    for row in frame.to_dict('records'):
        key = dataset_keys.get(str(row['ID']), {})
        source.append({**row, 'script_bucket': key.get('script_bucket', 'unknown'), 'stratify_key': key.get('stratify_key', 'unknown')})
    sample = sample_annotation_records(source, max_ids=int(CONFIG['annotation_max_ids']), seed=SEED, stratify_key='stratify_key')
    ratings_source = pd.DataFrame(sample['rows']).copy()
    key_rows = ratings_source[['ID', 'variant']].copy(); key_rows.insert(0, 'blind_id', [f'R{n:04d}' for n in range(1, len(key_rows) + 1)])
    key_rows['evidence_ids'] = ratings_source.get('evidence_ledger', pd.Series([None] * len(key_rows))).map(lambda value: [item.get('evidence_id') for item in value] if isinstance(value, list) else None)
    key_rows.to_json(RUN / 'artifacts' / 'annotation_blind_key.jsonl', orient='records', lines=True, force_ascii=False)
    ratings = ratings_source.rename(columns={'Text': 'post', 'Target': 'target', 'parsed_counter_narrative': 'response'}).copy()
    ratings.insert(0, 'blind_id', key_rows['blind_id'])
    ratings = ratings[['blind_id', 'post', 'target', 'response'] + [column for column in HUMAN_METRICS if column in ratings.columns]]
    if ratings.empty: raise RuntimeError('annotation_sample_empty')
    for metric in HUMAN_METRICS: ratings[metric] = pd.NA
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        ratings.to_excel(writer, sheet_name='Ratings', index=False)
        pd.DataFrame({'metric': HUMAN_METRICS}).to_excel(writer, sheet_name='Rubric', index=False)
        pd.DataFrame([sample]).to_json(RUN / 'artifacts' / 'annotation_sampling_manifest.json', orient='records')
    return sample
annotation_sample = build_human_annotation_workbook(generation_frame, RUN / 'exports' / 'human_annotation_blinded.xlsx'); print('human ratings are required; at least two distinct raters must complete the workbook')
def export_human_agreement(completed_rows):
    report = human_agreement_report(completed_rows)
    write_json(RUN / 'artifacts' / 'human_human_agreement.json', report)
    return report
COMPLETED_ANNOTATIONS = RUN / 'input' / 'human_annotations_completed.jsonl'
if COMPLETED_ANNOTATIONS.exists():
    completed_rows = load_jsonl(COMPLETED_ANNOTATIONS)
    if len({str(row.get('rater_id')) for row in completed_rows if row.get('rater_id')}) < 2: raise RuntimeError('human_agreement_requires_two_distinct_raters')
    export_human_agreement(completed_rows)
else:
    write_json(RUN / 'artifacts' / 'human_human_agreement.json', {'status': 'pending', 'reason': 'completed_annotations_not_supplied', 'required_distinct_raters': 2, 'weighted_kappa': None, 'krippendorff_alpha_ordinal': None})
"""),
        code("""# 15 - Final run manifest.
NLI_CALIBRATION_MANIFEST = {language: {key: value for key, value in runtime.items() if key != 'evaluator'} for language, runtime in NLI_RUNTIME_BY_LANGUAGE.items()}
write_json(RUN / \"exports\" / \"run_manifest.json\", {\"created_at\": now(), \"run_name\": RUN_NAME, \"run_identity_hash\": RUN_IDENTITY_HASH, \"config\": {k: str(v) if isinstance(v, Path) else v for k, v in CONFIG.items()}, \"config_hash\": CONFIG_HASH, \"lockfile_sha256\": LOCKFILE_SHA256, \"environment_fingerprint_hash\": ENVIRONMENT_FINGERPRINT_HASH, \"managed_accelerator_contract\": MANAGED_ACCELERATOR_CONTRACT, \"managed_accelerator_contract_hash\": MANAGED_ACCELERATOR_CONTRACT_HASH, \"qwen35_transformers_compatibility\": QWEN35_TRANSFORMERS_COMPATIBILITY, \"prompt_template_hash\": PROMPT_TEMPLATE_HASH, \"core_source_sha256\": CORE_SOURCE_SHA256, \"eval_core_source_sha256\": EVAL_CORE_SOURCE_SHA256, \"corpus_manifest_hash\": CORPUS_MANIFEST_HASH, \"audit_manifest_hash\": AUDIT_MANIFEST_HASH, \"chunk_manifest_hash\": CHUNK_MANIFEST_HASH, \"graph_manifest_hash\": GRAPH_MANIFEST_HASH, \"checkpoint_identity\": [\"input_text_sha256\", \"corpus_manifest_hash\", \"audit_manifest_hash\", \"chunk_manifest_hash\", \"graph_manifest_hash\", \"run_identity_hash\", \"lockfile_sha256\", \"environment_fingerprint_hash\", \"input_language\", \"output_language\"], \"paired_retrieval_evaluation\": True, \"memory_snapshots\": MEMORY_SNAPSHOTS, \"environment_fingerprint\": ENVIRONMENT_FINGERPRINT, \"evaluator_models\": {\"bertscore\": {\"model_id\": BERTSCORE_MODEL_ID, \"revision\": BERTSCORE_MODEL_REVISION, \"local_cache_keys\": [list(key) for key in _BERTSCORE_LOCAL_MODELS]}, \"nli\": {\"model_id\": CONFIG[\"nli_model_id\"], \"revision\": CONFIG[\"nli_model_revision\"]}}, \"nli_calibration_status\": NLI_RUNTIME_BY_LANGUAGE, \"dataset_limitations\": {\"en\": \"XNLI English validation\", \"hi\": \"IndicXNLI Hindi validation\", \"ta\": \"IndicXNLI Tamil validation; no evaluator without passing held-out gate\"}})
_run_manifest_path = RUN / \"exports\" / \"run_manifest.json\"
_run_manifest = json.loads(_run_manifest_path.read_text(encoding=\"utf-8\"))
_run_manifest[\"lock_package_counts\"] = LOCK_PACKAGE_COUNTS
_run_manifest[\"detoxify_metric\"] = DETOXIFY_STATUS
_run_manifest[\"detoxify_device\"] = DETOXIFY_DEVICE
_run_manifest[\"detoxify_status\"] = DETOXIFY_STATUS[\"status\"]
_run_manifest[\"metric_devices\"] = {\"bertscore\": BERTSCORE_DEVICE, \"detoxify\": DETOXIFY_DEVICE}
_run_manifest[\"generation_quarantine\"] = GENERATION_QUARANTINE_SUMMARY
write_json(_run_manifest_path, _run_manifest)
"""),
    ]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"cells": cells, "metadata": {"kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"}, "language_info": {"name": "python", "version": "3.11"}}, "nbformat": 4, "nbformat_minor": 5}
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    destination = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("outputs/remote_vm_qwen35_mpkg_rag.ipynb")
    build_notebook(destination)
    print(destination.resolve())
