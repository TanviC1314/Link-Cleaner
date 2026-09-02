# Generated from remote_vm_qwen35_mpkg_rag.ipynb; do not edit directly.

_LOCKFILE_SHA256_EXPECTED = "19b68a84617c002acd47042274904de7e98f59e5b097c8fcd350bc0cab4c0fb1"


# %% [notebook cell 2]

"""Manifest-driven corpus identity and provenance registry."""


import hashlib
import json
import math
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable


SUPPORTED_SUFFIXES = {".pdf", ".html", ".htm"}
HIDDEN_NAMES = {".ds_store", "thumbs.db"}
FACTUAL_SOURCE_TYPES = {
    "official",
    "peer_reviewed",
    "court",
    "government",
    "un_agency",
    "ngo_report",
    "survey",
    "education",
}
AUTHORITY_SCORES = {
    "official": 1.00,
    "peer_reviewed": 0.95,
    "court": 0.95,
    "government": 0.92,
    "un_agency": 0.92,
    "ngo_report": 0.75,
    "survey": 0.75,
    "education": 0.72,
    "oral_history": 0.55,
    "news_archive": 0.50,
    "public_debate": 0.35,
    "harmful_examples": 0.20,
}
SOURCE_ID_PATTERN = re.compile(r"\bSRC\d+\b", re.IGNORECASE)
TYPE_SOURCE_LABELS = {
    "peer reviewed full text article": "peer_reviewed",
    "peer reviewed research": "peer_reviewed",
    "peer reviewed review": "peer_reviewed",
    "systematic review source page": "peer_reviewed",
    "narrative review source page": "peer_reviewed",
    "european journal of public health pmc": "peer_reviewed",
    "official guidance": "official",
    "policy guidance": "official",
    "global legal report page": "court",
    "un fact sheet": "un_agency",
    "un hate speech framework": "un_agency",
    "un human rights booklet": "un_agency",
    "un report": "un_agency",
    "education guidance": "education",
    "education survey report": "survey",
    "human rights report": "ngo_report",
    "population survey report": "survey",
}
ORGANISATION_SOURCE_LABELS = {
    "ohchr": "un_agency",
    "unesco": "un_agency",
    "who": "un_agency",
    "unicef": "un_agency",
    "ilo": "un_agency",
    "united nations": "un_agency",
    "united nations free equal": "un_agency",
    "human rights watch": "ngo_report",
    "amnesty international": "ngo_report",
    "ilga world": "ngo_report",
    "pubmed": "peer_reviewed",
    "pubmed central": "peer_reviewed",
}
SOURCE_TYPE_PRIORITY = (
    "court",
    "official",
    "government",
    "un_agency",
    "peer_reviewed",
    "ngo_report",
    "survey",
    "education",
    "oral_history",
    "news_archive",
    "public_debate",
    "harmful_examples",
)
RESERVED_DERIVED_FIELDS = frozenset(
    {
        "document_uid",
        "legacy_source_id",
        "legacy_source_ids",
        "relative_path",
        "relative_paths",
        "path",
        "content_sha256",
        "document_type",
        "source_type",
        "authority_score",
        "factual_index_allowed",
        "status",
        "status_reason",
        "quarantine_reasons",
        "provenance",
        "manifest_metadata",
        "manifest_sources",
        "validation_errors",
        "source_id",
        "sha256",
        "local_file",
        "filename",
    }
)


class CorpusValidationError(ValueError):
    """Raised when strict registry loading rejects validation errors."""


class SourceRegistry(list):
    """List of deduplicated source rows with non-row audit information."""

    file_records_before_deduplication: int
    audit_events: list[dict[str, Any]]
    validation_errors: list[dict[str, Any]]

    def __init__(
        self,
        rows: Iterable[dict[str, Any]] = (),
        *,
        file_records_before_deduplication: int = 0,
        audit_events: list[dict[str, Any]] | None = None,
        validation_errors: list[dict[str, Any]] | None = None,
    ) -> None:
        super().__init__(rows)
        self.file_records_before_deduplication = file_records_before_deduplication
        self.audit_events = audit_events or []
        self.validation_errors = validation_errors or []


def stable_id(*parts: object) -> str:
    """Return a deterministic content-safe identifier for ordered parts."""

    payload = json.dumps(
        [str(part) for part in parts],
        ensure_ascii=False,
        separators=(",", ":"),
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:20]


def _entity_normalize(value: object) -> str:
    """Normalize a surface form for deterministic catalog matching only."""

    return " ".join(unicodedata.normalize("NFKC", str(value or "")).casefold().split())


def _catalog_entity_id(namespace: str, canonical_name: str, entity_type: str) -> str:
    prefix = "TGT" if namespace == "target" else "ENT"
    return f"{prefix}:{stable_id('catalog-entity', namespace, entity_type, _entity_normalize(canonical_name))}"


ENTITY_TYPE_ALIASES = {
    "org": "organization",
    "organisation": "organization",
    "institution": "organization",
    "company": "organization",
    "agency": "organization",
    "person": "person",
    "human": "person",
    "individual": "person",
    "people": "group",
    "population": "group",
    "community": "group",
    "collective": "group",
    "location": "place",
    "country": "place",
    "city": "place",
    "law": "policy",
    "legislation": "policy",
    "regulation": "policy",
    "event": "event",
    "concept": "concept",
    "issue": "concept",
    "dataset_target": "dataset_target",
}


def _normalize_entity_type(value: object) -> str:
    normalized = _entity_normalize(value or "unknown") or "unknown"
    return ENTITY_TYPE_ALIASES.get(normalized, normalized if normalized in {"organization", "person", "group", "place", "policy", "event", "concept", "dataset_target", "unknown"} else "unknown")


def _entity_tokens(value: object) -> list[str]:
    return re.findall(r"[\w]+", _entity_normalize(value))


def _acronym_key(value: object) -> str | None:
    surface = unicodedata.normalize("NFKC", str(value or "")).strip()
    tokens = _entity_tokens(surface)
    if not tokens:
        return None
    if len(tokens) == 1 and 2 <= len(tokens[0]) <= 8 and (surface.isupper() or len(tokens[0]) <= 4):
        return tokens[0].casefold()
    if 2 <= len(tokens) <= 8 and all(len(token) == 1 for token in tokens) and any(not char.isalnum() and not char.isspace() for char in surface):
        return "".join(tokens).casefold()
    return None


def _is_conservative_acronym(left: str, right: str) -> bool:
    acronym, right_tokens = _acronym_key(left), _entity_tokens(right)
    if acronym is None or len(right_tokens) < 2:
        return False
    stopwords = {"a", "an", "and", "for", "of", "the", "to"}
    initials = "".join(token[0] for token in right_tokens if token not in stopwords)
    return acronym.casefold() == initials.casefold()


def build_entity_catalog(
    corpus_rows: Iterable[dict[str, Any]] = (),
    dataset_targets: Iterable[str] = (),
    reviewed_manifest_organizations: Iterable[str] = (),
    mention_spans: Iterable[dict[str, Any]] = (),
) -> dict[str, Any]:
    """Build a closed, deterministic corpus-local entity catalog.

    Model output never supplies IDs to this function.  IDs are derived from a
    normalized label, entity type, and namespace; aliases are retained for
    candidate generation and are never used as graph identity.
    """

    candidates: dict[tuple[str, str, str], dict[str, Any]] = {}

    def add(label: object, entity_type: object, namespace: str, provenance: dict[str, Any] | None = None) -> None:
        canonical = " ".join(str(label or "").split())
        normalized = _entity_normalize(canonical)
        if not normalized:
            return
        kind = _normalize_entity_type(entity_type)
        key = (namespace, kind, normalized)
        row = candidates.setdefault(key, {"namespace": namespace, "entity_type": kind, "labels": set(), "provenance": []})
        row["labels"].add(canonical)
        if provenance and provenance not in row["provenance"]:
            row["provenance"].append(provenance)

    for target in dataset_targets or ():
        add(target, "dataset_target", "target", {"kind": "dataset_target"})
    for organization in reviewed_manifest_organizations or ():
        add(organization, "organization", "corpus", {"kind": "reviewed_manifest_organization"})
    for row in corpus_rows or ():
        if not isinstance(row, dict):
            continue
        metadata = row.get("manifest_metadata") if isinstance(row.get("manifest_metadata"), dict) else row
        organization = metadata.get("organisation") or metadata.get("organization")
        if organization:
            add(organization, "organization", "corpus", {"kind": "manifest", "document_uid": row.get("document_uid")})
    for mention in mention_spans or ():
        if not isinstance(mention, dict):
            continue
        text = mention.get("text")
        if isinstance(text, str) and text.strip():
            add(text, mention.get("entity_type", mention.get("mention_type", "unknown")), "corpus", {"kind": "validated_mention", "mention_id": mention.get("mention_id"), "document_uid": mention.get("document_uid"), "chunk_id": mention.get("chunk_id"), "start": mention.get("start"), "end": mention.get("end"), "text": text})

    nodes = sorted(candidates)
    parent = {node: node for node in nodes}
    def find(node):
        while parent[node] != node:
            parent[node] = parent[parent[node]]
            node = parent[node]
        return node
    def union(left, right):
        left_root, right_root = find(left), find(right)
        if left_root != right_root:
            parent[max(left_root, right_root)] = min(left_root, right_root)
    # Exact aliases are safe within a namespace/type-compatible catalog. An
    # acronym expansion is deliberately stricter: only a one-to-one pairing
    # may be merged, so ABC cannot absorb two organizations sharing initials
    # and cannot create a transitive alias chain.
    for index, left in enumerate(nodes):
        for right in nodes[index + 1:]:
            if left[0] == right[0] and (left[1] == right[1] or "unknown" in {left[1], right[1]}):
                left_labels, right_labels = candidates[left]["labels"], candidates[right]["labels"]
                if any(_entity_normalize(a) == _entity_normalize(b) for a in left_labels for b in right_labels):
                    union(left, right)
    ambiguous_acronym_nodes: set[tuple[str, str, str]] = set()
    for acronym in nodes:
        acronym_labels = candidates[acronym]["labels"]
        if not any(_acronym_key(label) for label in acronym_labels):
            continue
        compatible = []
        compatible_roots = set()
        for expansion in nodes:
            if expansion == acronym or expansion[0] != acronym[0] or not (expansion[1] == acronym[1] or "unknown" in {expansion[1], acronym[1]}) or not any(len(_entity_tokens(label)) >= 2 for label in candidates[expansion]["labels"]):
                continue
            if any(_is_conservative_acronym(a, b) for a in acronym_labels for b in candidates[expansion]["labels"]):
                expansion_root = find(expansion)
                if expansion_root not in compatible_roots:
                    compatible.append(expansion)
                    compatible_roots.add(expansion_root)
        if len(compatible) != 1:
            if compatible:
                ambiguous_acronym_nodes.add(acronym)
            continue
        expansion = compatible[0]
        reverse_roots = set()
        for other in nodes:
            if other == expansion or other[0] != expansion[0] or not (other[1] == expansion[1] or "unknown" in {other[1], expansion[1]}) or not any(_acronym_key(label) for label in candidates[other]["labels"]):
                continue
            if any(_is_conservative_acronym(a, b) for a in candidates[other]["labels"] for b in candidates[expansion]["labels"]):
                reverse_roots.add(find(other))
        if len(reverse_roots) == 1:
            union(acronym, expansion)
        else:
            ambiguous_acronym_nodes.add(acronym)

    grouped: dict[Any, list[dict[str, Any]]] = defaultdict(list)
    for node in nodes:
        grouped[find(node)].append(candidates[node])
    rows = []
    ambiguous_roots = {find(node) for node in ambiguous_acronym_nodes}
    for root, group in grouped.items():
        labels = sorted({label for item in group for label in item["labels"]}, key=lambda value: (-len(_entity_tokens(value)), -len(value), _entity_normalize(value), value))
        canonical = labels[0]
        namespace = group[0]["namespace"]
        known_types = sorted({item["entity_type"] for item in group if item["entity_type"] != "unknown"})
        kind = known_types[0] if known_types else "unknown"
        ambiguous = root in ambiguous_roots
        rows.append({"entity_id": _catalog_entity_id(namespace, canonical, kind), "canonical_name": canonical, "aliases": sorted(labels, key=lambda value: (_entity_normalize(value), value)), "entity_type": kind, "namespace": namespace, "external_refs": [], "catalog_status": "ambiguous" if ambiguous else "active", "link_status": "ambiguous" if ambiguous else "linked", "retrieval_allowed": not ambiguous, "factual_identity_allowed": namespace != "target", "provenance": sorted({json.dumps(value, ensure_ascii=False, sort_keys=True, default=str): value for item in group for value in item["provenance"]}.values(), key=lambda value: json.dumps(value, sort_keys=True, default=str))})
    alias_to_ids: dict[tuple[str, str], set[str]] = defaultdict(set)
    for row in rows:
        for alias in row["aliases"]:
            alias_to_ids[(row["namespace"], _entity_normalize(alias))].add(row["entity_id"])
    for row in rows:
        collisions = [alias for alias in row["aliases"] if len(alias_to_ids[(row["namespace"], _entity_normalize(alias))]) > 1]
        if collisions:
            row["catalog_status"] = "ambiguous"; row["link_status"] = "ambiguous"; row["retrieval_allowed"] = False; row["ambiguity_aliases"] = sorted(collisions)
        elif row.get("link_status") != "ambiguous":
            row["link_status"] = "linked"
    rows = sorted(rows, key=lambda row: row["entity_id"])
    revision = "entity-catalog.v4.namespace-aware-acronym-normalization"
    catalog_hash = stable_id(revision, json.dumps(rows, ensure_ascii=False, sort_keys=True, default=str))
    return {"catalog_revision": revision, "catalog_hash": catalog_hash, "entities": rows}


def build_entity_candidates(
    mention_text: str,
    catalog: dict[str, Any],
    max_candidates: int = 8,
    *,
    namespace_filter: str | Iterable[str] | None = None,
    namespace_preference: str | None = None,
    allow_target_fallback: bool = False,
    factual_only: bool = False,
) -> dict[str, Any]:
    """Return a deterministic, hashed candidate list for one validated mention."""

    if not isinstance(catalog, dict) or not isinstance(catalog.get("entities"), list):
        raise TypeError("entity catalog must contain an entities list")
    needle = _entity_normalize(mention_text)
    if namespace_filter is None:
        allowed_namespaces = None
        namespace_filter_key: list[str] | None = None
    elif isinstance(namespace_filter, str):
        allowed_namespaces = {namespace_filter}
        namespace_filter_key = [namespace_filter]
    else:
        namespace_filter_key = sorted({str(value) for value in namespace_filter})
        allowed_namespaces = set(namespace_filter_key)
    rows = []
    for entity in catalog["entities"]:
        if not isinstance(entity, dict) or entity.get("retrieval_allowed") is not True:
            continue
        if allowed_namespaces is not None and entity.get("namespace") not in allowed_namespaces:
            continue
        if factual_only and not (entity.get("namespace") == "corpus" and entity.get("factual_identity_allowed") is True):
            continue
        aliases = entity.get("aliases", [])
        if needle and any(_entity_normalize(alias) == needle for alias in aliases if isinstance(alias, str)):
            rows.append({
                "entity_id": entity["entity_id"],
                "canonical_name": entity["canonical_name"],
                "entity_type": entity["entity_type"],
                "namespace": entity["namespace"],
                "catalog_revision": catalog.get("catalog_revision"),
            })
    if namespace_preference is not None:
        preferred = [row for row in rows if row["namespace"] == namespace_preference]
        if preferred:
            rows = preferred
        elif allow_target_fallback:
            rows = [row for row in rows if row["namespace"] == "target"]
        else:
            rows = []
    rows.sort(key=lambda row: (row["namespace"] != "corpus", row["namespace"] != "target", row["canonical_name"].casefold(), row["entity_id"]))
    rows = rows[:max_candidates]
    policy = {"max_candidates": int(max_candidates), "namespace_filter": namespace_filter_key, "namespace_preference": namespace_preference, "allow_target_fallback": bool(allow_target_fallback), "factual_only": bool(factual_only)}
    candidate_hash = stable_id("entity-candidates.v3", catalog.get("catalog_hash"), needle, json.dumps(policy, sort_keys=True, ensure_ascii=False), json.dumps(rows, sort_keys=True, ensure_ascii=False))
    return {
        "mention_text": mention_text,
        "normalized_mention": needle,
        "catalog_revision": catalog.get("catalog_revision"),
        "catalog_hash": catalog.get("catalog_hash"),
        "max_candidates": int(max_candidates),
        "namespace_filter": namespace_filter_key,
        "namespace_preference": namespace_preference,
        "allow_target_fallback": bool(allow_target_fallback),
        "factual_only": bool(factual_only),
        "candidates": rows,
        "candidate_set_hash": candidate_hash,
    }


def _load_records(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, list) or not all(isinstance(item, dict) for item in value):
        raise ValueError(f"Expected a JSON array of objects: {path}")
    return value


def _normalized(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value)).replace("\\", "/")
    return re.sub(r"/+", "/", text).strip().casefold()


def _normalized_path(value: object) -> str:
    text = _normalized(value)
    while text.startswith("./"):
        text = text[2:]
    return text.rstrip("/")


def _basename(value: object) -> str:
    return _normalized_path(value).rsplit("/", 1)[-1]


def _normalized_label(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).casefold()
    return " ".join(re.findall(r"[a-z0-9]+", text))


def _source_id(value: object) -> str | None:
    match = SOURCE_ID_PATTERN.search(str(value or ""))
    return match.group(0).upper() if match else None


def _record_paths(record: dict[str, Any]) -> tuple[set[str], str | None]:
    paths: set[str] = set()
    for key in ("local_file", "path", "filename", "relative_path"):
        value = record.get(key)
        if value:
            normalized = _normalized_path(value)
            paths.add(normalized)
    basename = next(iter({_basename(path) for path in paths}), None)
    return paths, basename


def _path_matches_file(record: dict[str, Any], relative_path: str) -> bool:
    paths, _ = _record_paths(record)
    relative = _normalized_path(relative_path)
    return relative in paths or any(path.endswith("/" + relative) for path in paths)


def _hash_matches_file(record: dict[str, Any], content_sha256: str) -> bool:
    expected_hash = str(record.get("sha256") or record.get("content_sha256") or "").casefold()
    return bool(expected_hash) and expected_hash == content_sha256


def _unique_values(records: Iterable[dict[str, Any]], key: str) -> list[Any]:
    values: list[Any] = []
    seen: set[str] = set()
    for record in records:
        value = record.get(key)
        if value in (None, "", []):
            continue
        marker = json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
        if marker not in seen:
            values.append(value)
            seen.add(marker)
    return values


def _field_values(records: Iterable[dict[str, Any]], key: str) -> list[Any]:
    values: list[Any] = []
    for record in records:
        value = record.get(key)
        if isinstance(value, list):
            values.extend(value)
        elif value not in (None, ""):
            values.append(value)
    return values


def _merged_metadata(records: list[tuple[str, dict[str, Any]]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key in sorted({key for _, record in records for key in record}):
        values = _unique_values((record for _, record in records), key)
        if not values:
            continue
        result[key] = values[0] if len(values) == 1 else values
    return result


def _classify_source_type(records: list[tuple[str, dict[str, Any]]]) -> str:
    record_values = [record for _, record in records]
    type_labels = {
        _normalized_label(value)
        for value in _field_values(record_values, "type")
        if _normalized_label(value)
    }
    organisation_labels = {
        _normalized_label(value)
        for value in _field_values(record_values, "organisation")
        if _normalized_label(value)
    }
    has_doi = any(record.get("doi") or record.get("openalex") for _, record in records)

    candidates = {
        TYPE_SOURCE_LABELS[label]
        for label in type_labels
        if label in TYPE_SOURCE_LABELS
    }
    candidates.update(
        ORGANISATION_SOURCE_LABELS[label]
        for label in organisation_labels
        if label in ORGANISATION_SOURCE_LABELS
    )
    if has_doi:
        candidates.add("peer_reviewed")
    for source_type in SOURCE_TYPE_PRIORITY:
        if source_type in candidates:
            return source_type
    return "unknown"


def _source_ids(
    relative_paths: Iterable[str], records: list[tuple[str, dict[str, Any]]]
) -> list[str]:
    values: set[str] = set()
    for _, record in records:
        value = _source_id(record.get("source_id"))
        if value:
            values.add(value)
    for relative_path in relative_paths:
        value = _source_id(relative_path)
        if value:
            values.add(value)
    return sorted(values)


def _record_validation_errors(
    relative_path: str,
    actual_sha256: str,
    records: list[tuple[str, dict[str, Any]]],
) -> list[dict[str, Any]]:
    errors: list[dict[str, Any]] = []
    for origin, record in records:
        expected = str(record.get("sha256") or record.get("content_sha256") or "").casefold()
        if expected and expected != actual_sha256:
            errors.append(
                {
                    "event": "hash_mismatch",
                    "relative_path": relative_path,
                    "manifest_source": origin,
                    "source_id": _source_id(record.get("source_id")),
                    "expected_sha256": expected,
                    "actual_sha256": actual_sha256,
                }
            )
    return errors


def load_source_registry(corpus_root: Path, *, strict: bool = False) -> SourceRegistry:
    """Load, validate, merge, and content-deduplicate a corpus registry.

    Supported files on disk define the accounting baseline. The three JSON
    inventories are metadata overlays; no source ID is used as a document
    identity. By default validation errors are retained on the result so a
    corpus audit can complete. ``strict=True`` raises after the full audit.
    """

    corpus_root = Path(corpus_root)
    documents_root = corpus_root / "documents"
    if not documents_root.is_dir():
        raise FileNotFoundError(f"Expected corpus documents directory: {documents_root}")

    sources = [
        ("source_manifest", _load_records(corpus_root / "source_manifest.json")),
        ("added_openalex_sources", _load_records(corpus_root / "added_openalex_sources.json")),
        ("local_file_inventory", _load_records(corpus_root / "local_file_inventory.json")),
    ]
    overlays = [(origin, record) for origin, records in sources for record in records]

    audit_events: list[dict[str, Any]] = []
    file_records: list[dict[str, Any]] = []
    for path in sorted(documents_root.rglob("*")):
        if not path.is_file():
            continue
        relative_path = path.relative_to(corpus_root).as_posix()
        if path.name.casefold() in HIDDEN_NAMES or path.name.startswith("."):
            audit_events.append({"event": "hidden_metadata_file", "relative_path": relative_path})
            continue
        if path.suffix.casefold() not in SUPPORTED_SUFFIXES:
            audit_events.append({"event": "unsupported_file", "relative_path": relative_path})
            continue

        actual_sha256 = hashlib.sha256(path.read_bytes()).hexdigest()
        file_records.append(
            {
                "relative_path": relative_path,
                "path": str(path),
                "content_sha256": actual_sha256,
                "document_type": "html" if path.suffix.casefold() in {".html", ".htm"} else "pdf",
                "matched_manifest_records": [],
                "validation_errors": [],
            }
        )

    supported_by_basename: dict[str, list[str]] = defaultdict(list)
    for file_record in file_records:
        supported_by_basename[_basename(file_record["relative_path"])].append(
            file_record["relative_path"]
        )

    for file_record in file_records:
        relative_path = file_record["relative_path"]
        content_sha256 = file_record["content_sha256"]
        basename = _basename(relative_path)
        matched: list[tuple[str, dict[str, Any]]] = []
        basename_candidates: list[tuple[str, dict[str, Any]]] = []
        for origin, record in overlays:
            if _path_matches_file(record, relative_path) or _hash_matches_file(record, content_sha256):
                matched.append((origin, record))
            elif _record_paths(record)[1] == basename:
                basename_candidates.append((origin, record))

        if basename_candidates:
            basename_paths = supported_by_basename[basename]
            if len(basename_paths) == 1 and len(basename_candidates) == 1:
                matched.extend(basename_candidates)
            else:
                ambiguity = {
                    "event": "ambiguous_basename",
                    "relative_path": relative_path,
                    "basename": basename,
                    "affected_paths": sorted(basename_paths),
                    "candidate_manifest_sources": sorted(
                        {origin for origin, _ in basename_candidates}
                    ),
                    "reason": "basename-only overlay is not unique",
                }
                file_record["validation_errors"].append(ambiguity)
                audit_events.append(ambiguity.copy())

        file_record["matched_manifest_records"] = matched
        file_record["validation_errors"].extend(
            _record_validation_errors(relative_path, content_sha256, matched)
        )

    validation_errors = [
        error for file_record in file_records for error in file_record["validation_errors"]
    ]
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for file_record in file_records:
        groups[file_record["content_sha256"]].append(file_record)

    rows: list[dict[str, Any]] = []
    for content_sha256, group in sorted(groups.items()):
        relative_paths = sorted(file_record["relative_path"] for file_record in group)
        matched = [
            item
            for file_record in group
            for item in file_record["matched_manifest_records"]
        ]
        metadata = _merged_metadata(matched)
        source_type = _classify_source_type(matched)
        source_ids = _source_ids(relative_paths, matched)
        row_errors = [
            error
            for file_record in group
            for error in file_record["validation_errors"]
        ]
        quarantine_reasons = sorted(
            {
                "manifest_hash_mismatch"
                if error["event"] == "hash_mismatch"
                else "ambiguous_basename"
                if error["event"] == "ambiguous_basename"
                else str(error["event"])
                for error in row_errors
            }
        )
        quarantined = bool(quarantine_reasons)
        provenance = [
            {
                "relative_path": file_record["relative_path"],
                "legacy_source_ids": _source_ids(
                    [file_record["relative_path"]], file_record["matched_manifest_records"]
                ),
                "manifest_sources": sorted({origin for origin, _ in file_record["matched_manifest_records"]}),
            }
            for file_record in group
        ]
        row: dict[str, Any] = {
            "document_uid": stable_id("document", content_sha256),
            "legacy_source_id": source_ids[0] if source_ids else None,
            "legacy_source_ids": source_ids,
            "relative_path": relative_paths[0],
            "relative_paths": relative_paths,
            "path": str(corpus_root / relative_paths[0]),
            "content_sha256": content_sha256,
            "document_type": group[0]["document_type"],
            "source_type": source_type,
            "authority_score": AUTHORITY_SCORES.get(source_type),
            "factual_index_allowed": source_type in FACTUAL_SOURCE_TYPES and not quarantined,
            "status": "quarantined" if quarantined else "accepted",
            "status_reason": quarantine_reasons[0] if len(quarantine_reasons) == 1 else (
                "multiple_validation_errors" if quarantine_reasons else "validated"
            ),
            "quarantine_reasons": quarantine_reasons,
            "provenance": provenance,
            "manifest_metadata": metadata,
            "manifest_sources": sorted({origin for origin, _ in matched}),
            "validation_errors": row_errors,
        }
        for key, value in metadata.items():
            if key not in RESERVED_DERIVED_FIELDS:
                row[key] = value
        rows.append(row)

    result = SourceRegistry(
        rows,
        file_records_before_deduplication=len(file_records),
        audit_events=audit_events,
        validation_errors=validation_errors,
    )
    if strict and validation_errors:
        raise CorpusValidationError(
            f"{len(validation_errors)} corpus validation error(s); first: {validation_errors[0]}"
        )
    return result


SEMANTIC_PREDICATES = frozenset(
    {
        "associated_with",
        "calls_for",
        "causes",
        "contains",
        "describes",
        "denies",
        "excludes",
        "has_property",
        "includes",
        "is",
        "opposes",
        "supports",
        "targets",
    }
)
SEMANTIC_POLARITIES = frozenset({"affirmed", "negated", "uncertain", "mixed"})
SEMANTIC_MODALITIES = frozenset(
    {"asserted", "conditional", "necessary", "possible", "prohibited", "recommended", "reported"}
)
EVIDENCE_STANCES = frozenset({"supports", "refutes", "quotes", "contextualizes"})
ENTITY_STATUSES = frozenset({"canonical", "ambiguous", "nil"})
VALIDATED_EXTRACTION_MARKER = "mpkg-rag.validated-extraction.v1"
_ENTITY_ID_PATTERN = re.compile(r"^[A-Za-z][A-Za-z0-9+._:/?#%=-]{1,255}$")
_TOP_LEVEL_KEYS = frozenset({"schema_version", "claims"})
_CLAIM_KEYS = frozenset(
    {
        "claim_id",
        "mentions",
        "subject",
        "predicate",
        "object",
        "polarity",
        "polarity_scope",
        "modality",
        "attribution",
        "evidence_stance",
        "model_confidence",
    }
)
_MENTION_KEYS = frozenset(
    {"mention_id", "text", "start", "end", "entity_id", "candidate_index", "entity_status", "canonical_name"}
)


class ValidatedExtractionResult(dict):
    """JSON-persistable result marker for the validation-to-graph boundary."""

    validation_marker = VALIDATED_EXTRACTION_MARKER


SEMANTIC_EXTRACTION_SCHEMA = {
    "version": "semantic-claims.v1",
    "type": "object",
    "additionalProperties": False,
    "required": ["schema_version", "claims"],
    "properties": {
        "schema_version": {"const": "semantic-claims.v1"},
        "claims": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "required": [
                    "mentions",
                    "subject",
                    "predicate",
                    "object",
                    "polarity",
                    "modality",
                    "attribution",
                    "evidence_stance",
                    "model_confidence",
                ],
                "properties": {
                    "claim_id": {"type": "string"},
                    "mentions": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "additionalProperties": False,
                            "required": [
                                "mention_id",
                                "text",
                                "start",
                                "end",
                                "candidate_index",
                                "canonical_name",
                            ],
                            "properties": {
                                "mention_id": {"type": "string"},
                                "text": {"type": "string"},
                                "start": {"type": "integer", "minimum": 0},
                                "end": {"type": "integer", "minimum": 1},
                                "candidate_index": {"type": ["integer", "null"], "minimum": 0},
                                "canonical_name": {"type": "string", "minLength": 1},
                            },
                        },
                    },
                    "subject": {
                        "type": "object",
                        "additionalProperties": False,
                        "required": ["mention_id"],
                        "properties": {"mention_id": {"type": "string", "minLength": 1}},
                    },
                    "predicate": {"type": "string", "enum": sorted(SEMANTIC_PREDICATES)},
                    "object": {"type": "object"},
                    "polarity": {"type": "string", "enum": sorted(SEMANTIC_POLARITIES)},
                    "polarity_scope": {
                        "type": "object",
                        "additionalProperties": False,
                        "required": ["text", "start", "end"],
                        "properties": {
                            "text": {"type": "string"},
                            "start": {"type": "integer", "minimum": 0},
                            "end": {"type": "integer", "minimum": 1},
                        },
                    },
                    "modality": {"type": "string", "enum": sorted(SEMANTIC_MODALITIES)},
                    "attribution": {"type": ["object", "string", "null"]},
                    "evidence_stance": {"type": "string", "enum": sorted(EVIDENCE_STANCES)},
                    "model_confidence": {"type": "number", "minimum": 0, "maximum": 1},
                },
            },
        },
    },
}


def build_extraction_prompt(
    text: str, source_context: dict[str, Any], mode: str
) -> list[dict[str, str]]:
    """Build a model-agnostic extraction prompt for the versioned payload."""

    schema = json.dumps(SEMANTIC_EXTRACTION_SCHEMA, ensure_ascii=False, sort_keys=True)
    context = json.dumps(source_context, ensure_ascii=False, sort_keys=True)
    system = (
        "Extract only claims explicitly supported by the supplied text. Return one JSON object "
        "with schema_version=semantic-claims.v1 and a claims list. Preserve subject/object roles, "
        "use exact Python string offsets and mention text, and never infer a relation from keywords. "
        "Never emit entity IDs. For every mention, emit candidate_index only; the validator assigns the "
        "closed-catalog entity ID. Use null when no supplied candidate is supported by the text. "
        f"Mode: {mode}. Schema: {schema}"
    )
    user = f"Source context: {context}\nText (do not normalize):\n{text}"
    return [
        {"role": "system", "content": system},
        {"role": "user", "content": user},
    ]


MENTION_DISCOVERY_SCHEMA = {
    "version": "mention-discovery.v1",
    "type": "object",
    "additionalProperties": False,
    "required": ["schema_version", "mentions"],
    "properties": {
        "schema_version": {"const": "mention-discovery.v1"},
        "mentions": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "required": ["mention_id", "text", "start", "end", "mention_type"],
                "properties": {
                    "mention_id": {"type": "string", "minLength": 1},
                    "text": {"type": "string", "minLength": 1},
                    "start": {"type": "integer", "minimum": 0},
                    "end": {"type": "integer", "minimum": 1},
                    "mention_type": {"type": "string", "minLength": 1},
                },
            },
        },
    },
}


def build_mention_prompt(text: str, source_context: dict[str, Any]) -> list[dict[str, str]]:
    schema = json.dumps(MENTION_DISCOVERY_SCHEMA, ensure_ascii=False, sort_keys=True)
    return [
        {"role": "system", "content": "Discover only exact entity mention spans. Never emit entity IDs. Return JSON matching this schema: " + schema},
        {"role": "user", "content": f"Source context: {json.dumps(source_context, ensure_ascii=False, sort_keys=True)}\nText (do not normalize):\n{text}"},
    ]


def validate_mentions(payload: Any, text: str, source_context: dict[str, Any] | None = None) -> dict[str, Any]:
    """Validate mention spans before they can seed catalog construction."""

    result = {"schema_version": MENTION_DISCOVERY_SCHEMA["version"], "accepted": [], "quarantined": [], "status": "quarantined"}
    if not isinstance(text, str) or not isinstance(payload, dict) or set(payload) != {"schema_version", "mentions"} or payload.get("schema_version") != MENTION_DISCOVERY_SCHEMA["version"] or not isinstance(payload.get("mentions"), list):
        result["quarantined"].append({"record": payload, "reasons": ["mention_payload_schema_invalid"], "source_context": source_context, "status": "quarantined"})
        return result
    seen: set[str] = set()
    for mention in payload["mentions"]:
        reasons: list[str] = []
        if not isinstance(mention, dict):
            reasons.append("mention_not_object")
        else:
            if set(mention) != {"mention_id", "text", "start", "end", "mention_type"}:
                reasons.append("mention_schema_invalid")
            mention_id, surface, start, end = (mention.get(key) for key in ("mention_id", "text", "start", "end"))
            if not isinstance(mention_id, str) or not mention_id or mention_id in seen:
                reasons.append("mention_id_invalid_or_duplicate")
            if not isinstance(surface, str) or not surface:
                reasons.append("mention_text_invalid")
            if isinstance(start, bool) or not isinstance(start, int) or isinstance(end, bool) or not isinstance(end, int) or start < 0 or end <= start or end > len(text) or text[start:end] != surface:
                reasons.append("mention_span_mismatch")
            if not isinstance(mention.get("mention_type"), str) or not mention["mention_type"]:
                reasons.append("mention_type_invalid")
        if reasons:
            result["quarantined"].append({"record": mention, "reasons": sorted(set(reasons)), "source_context": source_context, "status": "quarantined"})
        else:
            seen.add(mention["mention_id"])
            result["accepted"].append(dict(mention))
    result["status"] = "accepted" if not result["quarantined"] else ("partial" if result["accepted"] else "quarantined")
    return result


def resolve_query_signature_entities(payload: Any, candidate_sets: dict[str, Any], catalog: dict[str, Any]) -> dict[str, Any]:
    """Map query candidate indices to catalog IDs; raw model IDs are forbidden."""

    if not isinstance(payload, dict):
        return {"valid": False, "reasons": ["query_signature_not_object"], "entity_ids": []}
    if payload.get("entity_ids"):
        return {"valid": False, "reasons": ["raw_query_entity_ids_forbidden"], "entity_ids": []}
    indices = payload.get("entity_candidate_indices", [])
    if not isinstance(indices, list):
        return {"valid": False, "reasons": ["query_candidate_indices_not_list"], "entity_ids": []}
    entity_ids: list[str] = []
    reasons: list[str] = []
    for item in indices:
        if not isinstance(item, dict) or not isinstance(item.get("mention_id"), str):
            reasons.append("query_candidate_schema_invalid")
            continue
        candidate_set = candidate_sets.get(item["mention_id"])
        if not isinstance(candidate_set, dict):
            reasons.append("query_candidate_set_missing")
            continue
        try:
            expected = build_entity_candidates(
                candidate_set.get("mention_text", ""),
                catalog,
                max_candidates=int(candidate_set.get("max_candidates", 8)),
                namespace_filter=candidate_set.get("namespace_filter"),
                namespace_preference=candidate_set.get("namespace_preference"),
                allow_target_fallback=bool(candidate_set.get("allow_target_fallback", False)),
                factual_only=bool(candidate_set.get("factual_only", False)),
            )
        except (TypeError, ValueError):
            reasons.append("query_candidate_set_authentication_failed")
            continue
        if candidate_set.get("candidate_set_hash") != expected.get("candidate_set_hash"):
            reasons.append("query_candidate_set_hash_mismatch")
        if candidate_set != expected:
            reasons.append("query_candidate_set_authentication_failed")
            continue
        index = item.get("candidate_index")
        if index is None:
            continue
        if isinstance(index, bool) or not isinstance(index, int) or not 0 <= index < len(candidate_set.get("candidates", [])):
            reasons.append("query_candidate_index_unknown")
            continue
        entity_ids.append(expected["candidates"][index]["entity_id"])
    return {"valid": not reasons and len(entity_ids) == len(set(entity_ids)), "reasons": sorted(set(reasons)), "entity_ids": sorted(set(entity_ids))}


def _semantic_context(source_context: Any) -> tuple[dict[str, Any] | None, list[str]]:
    if not isinstance(source_context, dict):
        return None, ["source_context_not_object"]
    context = dict(source_context)
    reasons = []
    for key in ("document_uid", "chunk_id"):
        if not isinstance(context.get(key), str) or not context[key]:
            reasons.append(f"missing_{key}")
    return (context if not reasons else None), reasons


def _claim_ref(value: Any) -> str | None:
    if isinstance(value, dict) and set(value) == {"mention_id"} and isinstance(value["mention_id"], str) and value["mention_id"]:
        return value["mention_id"]
    return None


def _nil_entity_id(context: dict[str, Any], mention: dict[str, Any], surface: str) -> str:
    return "NIL:" + stable_id(
        "entity",
        context["document_uid"],
        context["chunk_id"],
        surface,
        mention["start"],
        mention["end"],
    )


def _strict_catalog_context(context: dict[str, Any]) -> bool:
    return isinstance(context.get("entity_catalog"), dict)


def _validated_candidate_for_mention(
    mention: dict[str, Any], context: dict[str, Any], mention_id: str
) -> tuple[str | None, str | None, list[str]]:
    """Resolve a model candidate index against the pinned catalog, never a raw ID."""

    if not _strict_catalog_context(context):
        return mention.get("entity_id"), mention.get("entity_status"), []
    reasons: list[str] = []
    if mention.get("entity_id") is not None:
        reasons.append("raw_entity_id_forbidden")
    candidate_sets = context.get("candidate_sets")
    candidate_set = candidate_sets.get(mention_id) if isinstance(candidate_sets, dict) else None
    if not isinstance(candidate_set, dict):
        reasons.append("missing_candidate_set")
        return None, "nil", reasons
    catalog = context["entity_catalog"]
    catalog_rows = {
        row.get("entity_id"): row
        for row in catalog.get("entities", [])
        if isinstance(row, dict) and isinstance(row.get("entity_id"), str)
    }
    try:
        expected_candidate_set = build_entity_candidates(
            candidate_set.get("mention_text", ""),
            catalog,
            max_candidates=int(candidate_set.get("max_candidates", 8)),
            namespace_filter=candidate_set.get("namespace_filter"),
            namespace_preference=candidate_set.get("namespace_preference"),
            allow_target_fallback=bool(candidate_set.get("allow_target_fallback", False)),
            factual_only=bool(candidate_set.get("factual_only", False)),
        )
    except (TypeError, ValueError):
        reasons.append("candidate_set_authentication_failed")
        return None, "nil", reasons
    candidates = expected_candidate_set.get("candidates")
    if candidate_set.get("candidate_set_hash") != expected_candidate_set.get("candidate_set_hash"):
        reasons.append("candidate_set_hash_mismatch")
    if candidate_set != expected_candidate_set:
        reasons.append("candidate_set_authentication_failed")
    if candidate_set.get("catalog_revision") != catalog.get("catalog_revision"):
        reasons.append("catalog_revision_mismatch")
    if not isinstance(candidates, list):
        reasons.append("candidate_list_invalid")
        return None, "nil", reasons
    for candidate in candidates:
        if not isinstance(candidate, dict) or candidate.get("entity_id") not in catalog_rows:
            reasons.append("candidate_not_in_catalog")
            break
        if candidate_set.get("factual_only") and not (candidate.get("namespace") == "corpus" and catalog_rows[candidate["entity_id"]].get("factual_identity_allowed") is True):
            reasons.append("candidate_namespace_policy_violation")
            break
    candidate_index = mention.get("candidate_index")
    if candidate_index is None:
        return None, "nil", reasons
    if isinstance(candidate_index, bool) or not isinstance(candidate_index, int) or not 0 <= candidate_index < len(candidates):
        reasons.append("candidate_index_unknown")
        return None, "nil", reasons
    candidate = expected_candidate_set["candidates"][candidate_index]
    entity_id = candidate.get("entity_id")
    entity = catalog_rows.get(entity_id)
    if not isinstance(entity_id, str) or entity is None or entity.get("retrieval_allowed") is not True:
        reasons.append("candidate_not_retrieval_allowed")
        return None, "nil", reasons
    return entity_id, "canonical", reasons


def _quarantined_record(record: Any, reasons: list[str], context: dict[str, Any] | None) -> dict[str, Any]:
    return {
        "record": record,
        "reasons": sorted(set(reasons)),
        "source_context": context,
        "status": "quarantined",
    }


def _reviewed_record(record: Any, reasons: list[str], context: dict[str, Any] | None) -> dict[str, Any]:
    return {
        "record": record,
        "reasons": sorted(set(reasons)),
        "source_context": context,
        "status": "reviewed",
    }


def _normalized_json(value: Any) -> Any:
    return json.loads(json.dumps(value, ensure_ascii=False, sort_keys=True, default=str))


def _validation_fingerprint(result: dict[str, Any]) -> str:
    digest_material = {
        key: value for key, value in result.items() if key != "validation_fingerprint"
    }
    serialized = json.dumps(digest_material, ensure_ascii=False, sort_keys=True, default=str)
    return stable_id("validated-extraction", serialized)


def _new_validation_result(payload: Any, text: Any, source_context: Any) -> ValidatedExtractionResult:
    return ValidatedExtractionResult(
        {
            "validation_marker": VALIDATED_EXTRACTION_MARKER,
            "validation_fingerprint": "",
            "schema_version": SEMANTIC_EXTRACTION_SCHEMA["version"],
            "payload": payload,
            "normalized_payload": None,
            "source_context": dict(source_context) if isinstance(source_context, dict) else source_context,
            "text": text,
            "accepted": [],
            "quarantined": [],
            "reviewed": [],
        }
    )


def _finish_validation_result(result: ValidatedExtractionResult) -> ValidatedExtractionResult:
    normalized_claims = list(result.get("accepted", []))
    normalized_claims.extend(
        item["record"]
        for item in result.get("reviewed", [])
        if isinstance(item, dict) and isinstance(item.get("record"), dict)
    )
    result["normalized_payload"] = {
        "schema_version": SEMANTIC_EXTRACTION_SCHEMA["version"],
        "claims": _normalized_json(normalized_claims),
    }
    result["validation_fingerprint"] = _validation_fingerprint(result)
    return result


def validate_extraction(
    payload: Any, text: str, source_context: dict[str, Any]
) -> dict[str, Any]:
    """Validate a Qwen extraction payload without loading or invoking a model."""

    context, context_reasons = _semantic_context(source_context)
    result = _new_validation_result(payload, text, source_context)
    if not isinstance(text, str):
        result["quarantined"].append(_quarantined_record(payload, ["source_text_not_string"], context))
        result["status"] = "quarantined"
        return _finish_validation_result(result)
    if not isinstance(payload, dict):
        result["quarantined"].append(_quarantined_record(payload, ["payload_not_object"], context))
        result["status"] = "quarantined"
        return _finish_validation_result(result)
    top_level_reasons = []
    if set(payload) - _TOP_LEVEL_KEYS:
        top_level_reasons.append("extra_payload_fields")
    if "schema_version" not in payload:
        top_level_reasons.append("missing_schema_version")
    if "claims" not in payload:
        top_level_reasons.append("missing_claims")
    if top_level_reasons:
        result["quarantined"].append(_quarantined_record(payload, top_level_reasons, context))
        result["status"] = "quarantined"
        return _finish_validation_result(result)
    if payload.get("schema_version") != SEMANTIC_EXTRACTION_SCHEMA["version"]:
        result["quarantined"].append(
            _quarantined_record(payload, ["unsupported_schema_version"], context)
        )
        result["status"] = "quarantined"
        return _finish_validation_result(result)
    if not isinstance(payload.get("claims"), list):
        result["quarantined"].append(_quarantined_record(payload, ["claims_not_list"], context))
        result["status"] = "quarantined"
        return _finish_validation_result(result)
    if context_reasons:
        result["quarantined"].extend(
            _quarantined_record(claim, context_reasons, context) for claim in payload["claims"]
        )
        result["status"] = "quarantined"
        return _finish_validation_result(result)

    for claim_index, claim in enumerate(payload["claims"]):
        reasons: list[str] = []
        if not isinstance(claim, dict):
            result["quarantined"].append(
                _quarantined_record(claim, ["claim_not_object"], context)
            )
            continue
        if set(claim) - _CLAIM_KEYS:
            reasons.append("extra_claim_fields")
        if "claim_id" in claim and (not isinstance(claim["claim_id"], str) or not claim["claim_id"]):
            reasons.append("invalid_claim_id")
        required = {
            "mentions",
            "subject",
            "predicate",
            "object",
            "polarity",
            "modality",
            "attribution",
            "evidence_stance",
            "model_confidence",
        }
        reasons.extend(f"missing_{key}" for key in sorted(required - set(claim)))
        mentions = claim.get("mentions")
        normalized_mentions: list[dict[str, Any]] = []
        mention_by_id: dict[str, dict[str, Any]] = {}
        if not isinstance(mentions, list) or not mentions:
            reasons.append("mentions_not_nonempty_list")
        else:
            for mention_index, mention in enumerate(mentions):
                if not isinstance(mention, dict):
                    reasons.append("mention_not_object")
                    continue
                mention_reasons: list[str] = []
                if set(mention) - _MENTION_KEYS:
                    mention_reasons.append("extra_mention_fields")
                required_mention_keys = _MENTION_KEYS - ({"entity_id", "candidate_index", "entity_status"} if _strict_catalog_context(context or {}) else {"candidate_index"})
                missing_mention_keys = required_mention_keys - set(mention)
                mention_reasons.extend(f"missing_{key}" for key in sorted(missing_mention_keys))
                mention_id = mention.get("mention_id")
                if not isinstance(mention_id, str) or not mention_id:
                    mention_reasons.append("invalid_mention_id")
                    reasons.extend(mention_reasons)
                    continue
                if mention_id in mention_by_id:
                    mention_reasons.append("duplicate_mention_id")
                surface = mention.get("text")
                start = mention.get("start")
                end = mention.get("end")
                if not isinstance(surface, str) or not surface:
                    mention_reasons.append("mention_text_not_string")
                    reasons.extend(mention_reasons)
                    continue
                if isinstance(start, bool) or not isinstance(start, int) or isinstance(end, bool) or not isinstance(end, int):
                    mention_reasons.append("mention_offsets_not_integer")
                    reasons.extend(mention_reasons)
                    continue
                if start < 0 or end <= start or end > len(text) or text[start:end] != surface:
                    mention_reasons.append("mention_span_mismatch")
                    reasons.extend(mention_reasons)
                    continue
                supplied_entity_id = mention.get("entity_id")
                entity_status = mention.get("entity_status")
                entity_id, resolved_status, candidate_reasons = _validated_candidate_for_mention(
                    mention, context or {}, mention_id
                )
                if _strict_catalog_context(context or {}):
                    entity_status = resolved_status
                    mention_reasons.extend(candidate_reasons)
                else:
                    entity_id = supplied_entity_id
                if not isinstance(entity_status, str) or entity_status not in ENTITY_STATUSES:
                    mention_reasons.append("unknown_entity_status")
                if entity_id is not None and not isinstance(entity_id, str):
                    mention_reasons.append("invalid_entity_id_type")
                    reasons.extend(mention_reasons)
                    continue
                if isinstance(entity_id, str) and (
                    not _ENTITY_ID_PATTERN.fullmatch(entity_id)
                    or (entity_id.startswith(("NIL-", "NIL:")) and entity_status == "canonical")
                ):
                    mention_reasons.append("invalid_entity_id")
                if entity_status == "canonical" and not entity_id:
                    mention_reasons.append("missing_canonical_entity_id")
                if entity_status in {"ambiguous", "nil"}:
                    if entity_id is not None and not entity_id.startswith(("NIL-", "NIL:")):
                        mention_reasons.append("unresolved_entity_id_must_be_nil")
                    if entity_id is None and context is not None:
                        entity_id = _nil_entity_id(context, mention, surface)
                canonical_name = mention.get("canonical_name")
                if not isinstance(canonical_name, str) or not canonical_name:
                    mention_reasons.append("invalid_canonical_name")
                normalized_mention = {
                    "mention_id": mention_id,
                    "text": surface,
                    "start": start,
                    "end": end,
                    "entity_id": entity_id,
                    "candidate_index": mention.get("candidate_index"),
                    "entity_status": entity_status,
                    "canonical_name": canonical_name,
                }
                if mention_reasons:
                    reasons.extend(mention_reasons)
                else:
                    mention_by_id[mention_id] = normalized_mention
                    normalized_mentions.append(normalized_mention)

        predicate = claim.get("predicate")
        if predicate not in SEMANTIC_PREDICATES:
            reasons.append("unknown_predicate")
        if claim.get("polarity") not in SEMANTIC_POLARITIES:
            reasons.append("unknown_polarity")
        if claim.get("modality") not in SEMANTIC_MODALITIES:
            reasons.append("unknown_modality")
        if claim.get("evidence_stance") not in EVIDENCE_STANCES:
            reasons.append("unknown_evidence_stance")
        if not isinstance(claim.get("attribution"), (dict, str, type(None))):
            reasons.append("invalid_attribution")
        confidence = claim.get("model_confidence")
        if isinstance(confidence, bool) or not isinstance(confidence, (int, float)) or not 0 <= confidence <= 1:
            reasons.append("invalid_model_confidence")

        if not isinstance(claim.get("subject"), dict):
            reasons.append("subject_not_object")
        elif set(claim["subject"]) != {"mention_id"}:
            reasons.append("subject_schema_invalid")
        subject_ref = _claim_ref(claim.get("subject"))
        if subject_ref is None or subject_ref not in mention_by_id:
            reasons.append("subject_endpoint_missing")
        object_value: dict[str, Any] | None = None
        object_ref = _claim_ref(claim.get("object"))
        if not isinstance(claim.get("object"), dict):
            reasons.append("object_not_object")
        elif object_ref is not None:
            if object_ref not in mention_by_id:
                reasons.append("object_endpoint_missing")
        else:
            raw_object = claim.get("object")
            if isinstance(raw_object, dict) and set(raw_object) == {"value", "value_type"}:
                value = raw_object["value"]
                if not isinstance(raw_object["value_type"], str) or not isinstance(value, (str, int, float, bool)) or value == "":
                    reasons.append("object_value_invalid")
                else:
                    object_value = {"value": value, "value_type": raw_object["value_type"]}
            else:
                reasons.append("object_schema_invalid")

        polarity_scope = claim.get("polarity_scope")
        if claim.get("polarity") in {"negated", "uncertain", "mixed"} and polarity_scope is None:
            reasons.append("missing_polarity_scope")
        if polarity_scope is not None:
            if not isinstance(polarity_scope, dict) or set(polarity_scope) != {"text", "start", "end"}:
                reasons.append("polarity_scope_schema_invalid")
            else:
                scope_text = polarity_scope["text"]
                scope_start = polarity_scope["start"]
                scope_end = polarity_scope["end"]
                if (
                    not isinstance(scope_text, str)
                    or isinstance(scope_start, bool)
                    or not isinstance(scope_start, int)
                    or isinstance(scope_end, bool)
                    or not isinstance(scope_end, int)
                    or scope_start < 0
                    or scope_end <= scope_start
                    or scope_end > len(text)
                    or text[scope_start:scope_end] != scope_text
                ):
                    reasons.append("polarity_scope_mismatch")

        if reasons:
            result["quarantined"].append(_quarantined_record(claim, reasons, context))
            continue
        normalized_claim = {
            "claim_id": claim.get("claim_id") or "claim-" + stable_id("claim", claim_index, context["chunk_id"]),
            "mentions": normalized_mentions,
            "subject": {"mention_id": subject_ref},
            "predicate": predicate,
            "object": {"mention_id": object_ref} if object_ref is not None else object_value,
            "polarity": claim["polarity"],
            "modality": claim["modality"],
            "attribution": claim["attribution"],
            "evidence_stance": claim["evidence_stance"],
            "model_confidence": confidence,
            "source_context": context,
        }
        if polarity_scope is not None:
            normalized_claim["polarity_scope"] = dict(polarity_scope)
        unresolved = any(
            mention["entity_status"] in {"ambiguous", "nil"}
            for mention in normalized_mentions
        )
        if unresolved or claim["evidence_stance"] == "quotes":
            review_reasons = ["unresolved_entity"] if unresolved else []
            if claim["evidence_stance"] == "quotes":
                review_reasons.append("quoted_only")
            result["reviewed"].append(_reviewed_record(normalized_claim, review_reasons, context))
        else:
            result["accepted"].append(normalized_claim)
    if result["accepted"] and not result["quarantined"] and not result["reviewed"]:
        result["status"] = "accepted"
    elif result["reviewed"] and not result["accepted"] and not result["quarantined"]:
        result["status"] = "reviewed"
    elif result["quarantined"] and not result["accepted"] and not result["reviewed"]:
        result["status"] = "quarantined"
    else:
        result["status"] = "partial"
    return _finish_validation_result(result)


def _semantic_key(value: Any) -> str:
    if isinstance(value, str):
        return " ".join(value.casefold().split())
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


_VALIDATED_RESULT_KEYS = frozenset(
    {
        "validation_marker",
        "validation_fingerprint",
        "schema_version",
        "payload",
        "normalized_payload",
        "source_context",
        "text",
        "accepted",
        "quarantined",
        "reviewed",
        "status",
    }
)


def _iter_accepted_extractions(
    accepted_extractions: Any,
) -> tuple[list[tuple[dict[str, Any], dict[str, Any], str]], list[Any], list[Any]]:
    accepted: list[tuple[dict[str, Any], dict[str, Any], str]] = []
    quarantined: list[Any] = []
    reviewed: list[Any] = []
    records = (
        [(None, accepted_extractions)]
        if isinstance(accepted_extractions, dict)
        else list(enumerate(accepted_extractions or []))
    )
    for _, record in records:
        if not isinstance(record, dict) or record.get("validation_marker") != VALIDATED_EXTRACTION_MARKER:
            quarantined.append(_quarantined_record(record, ["unvalidated_extraction_result"], None))
            continue
        if not _VALIDATED_RESULT_KEYS.issubset(record) or set(record) - _VALIDATED_RESULT_KEYS:
            quarantined.append(_quarantined_record(record, ["validated_result_schema_invalid"], None))
            continue
        if record.get("status") not in {"accepted", "partial", "reviewed", "quarantined"}:
            quarantined.append(_quarantined_record(record, ["validation_result_status_invalid"], record.get("source_context")))
            continue
        source_context = record.get("source_context")
        text = record.get("text")
        if not isinstance(source_context, dict) or not isinstance(text, str):
            quarantined.append(_quarantined_record(record, ["validated_result_provenance_invalid"], source_context))
            continue
        if record.get("schema_version") != SEMANTIC_EXTRACTION_SCHEMA["version"]:
            quarantined.append(_quarantined_record(record, ["validated_result_schema_invalid"], source_context))
            continue
        revalidated = validate_extraction(record.get("payload"), text, source_context)
        comparable_fields = (
            "normalized_payload",
            "text",
            "source_context",
            "status",
            "accepted",
            "reviewed",
            "quarantined",
        )
        if any(record.get(field) != revalidated.get(field) for field in comparable_fields):
            quarantined.append(_quarantined_record(record, ["validation_result_stale"], source_context))
            continue
        if record.get("validation_fingerprint") != revalidated.get("validation_fingerprint"):
            quarantined.append(_quarantined_record(record, ["validation_result_tampered"], source_context))
            continue
        quarantined.extend(revalidated["quarantined"])
        reviewed.extend(revalidated["reviewed"])
        if record.get("status") == "quarantined":
            quarantined.extend(
                _quarantined_record(claim, ["quarantined_extraction_result"], source_context)
                for claim in revalidated["accepted"]
            )
            continue
        if record.get("status") == "reviewed":
            reviewed.extend(
                _reviewed_record(claim, ["reviewed_only"], source_context)
                for claim in revalidated["accepted"]
            )
            continue
        for claim in revalidated["accepted"]:
            accepted.append((claim, source_context, text))
    return accepted, quarantined, reviewed


def _chunk_provenance_reasons(
    context: dict[str, Any], extraction_text: str, chunk: dict[str, Any] | None
) -> list[str]:
    if chunk is None:
        return ["missing_chunk_provenance"]
    reasons: list[str] = []
    chunk_text = chunk.get("text")
    if not isinstance(chunk_text, str):
        reasons.append("missing_chunk_text")
    elif chunk_text != extraction_text:
        reasons.append("chunk_text_mismatch")
    if chunk.get("factual_index_allowed") is not True:
        reasons.append("factual_index_policy_missing_or_denied")
    if chunk.get("status") not in {"accepted", "validated"}:
        reasons.append("chunk_status_not_accepted")
    if context.get("factual_index_allowed") is not None and context.get("factual_index_allowed") is not True:
        reasons.append("factual_index_policy_mismatch")
    if context.get("status") is not None and context.get("status") not in {"accepted", "validated"}:
        reasons.append("chunk_status_not_accepted")
    if "text" in context and context["text"] != extraction_text:
        reasons.append("source_context_mismatch")
    for key in ("source_id", "source_type", "authority_score", "factual_index_allowed", "status"):
        if key in context and key in chunk and context[key] != chunk[key]:
            reasons.append("source_context_mismatch")
    if chunk.get("source_type") == "harmful_examples" or context.get("source_type") == "harmful_examples":
        reasons.append("harmful_source_type")
    if isinstance(chunk_text, str):
        actual_hash = hashlib.sha256(chunk_text.encode("utf-8")).hexdigest()
        text_hashes = [
            owner["text_sha256"]
            for owner in (context, chunk)
            if owner.get("text_sha256") is not None
        ]
        if any(expected != actual_hash for expected in text_hashes):
            reasons.append("chunk_hash_mismatch")
        if len(set(text_hashes)) > 1:
            reasons.append("source_context_hash_mismatch")
    document_hashes = [
        owner[key]
        for owner in (context, chunk)
        for key in ("content_sha256", "document_sha256")
        if owner.get(key) is not None
    ]
    if len(set(document_hashes)) > 1:
        reasons.append("document_hash_mismatch")
    return sorted(set(reasons))


def _audit_field(chunk: dict[str, Any], context: dict[str, Any], key: str) -> Any:
    return chunk[key] if key in chunk else context.get(key)


def build_semantic_graph(chunks: Iterable[dict[str, Any]], accepted_extractions: Any) -> dict[str, Any]:
    """Build accepted semantic tables and typed edges from validated extractions."""

    chunk_rows = list(chunks or [])
    chunk_by_key = {
        (row.get("document_uid"), row.get("chunk_id")): row
        for row in chunk_rows
        if isinstance(row, dict)
    }
    accepted, quarantined, reviewed = _iter_accepted_extractions(accepted_extractions)
    graph: dict[str, Any] = {
        "schema_version": SEMANTIC_EXTRACTION_SCHEMA["version"],
        "Document": [],
        "EvidenceChunk": [],
        "Mention": [],
        "Entity": [],
        "Claim": [],
        "edges": [],
        "quarantined": quarantined,
        "reviewed": reviewed,
    }
    documents: dict[str, dict[str, Any]] = {}
    evidence_chunks: dict[str, dict[str, Any]] = {}
    mentions: dict[str, dict[str, Any]] = {}
    entities: dict[str, dict[str, Any]] = {}
    claims: dict[str, dict[str, Any]] = {}
    edges: list[dict[str, Any]] = []

    def add_edge(source_id: str, edge_type: str, target_id: str, **extra: Any) -> None:
        edges.append({"source_id": source_id, "type": edge_type, "target_id": target_id, **extra})

    for extraction_index, (claim, context, extraction_text) in enumerate(accepted):
        document_uid = context.get("document_uid")
        chunk_id = context.get("chunk_id")
        if not isinstance(document_uid, str) or not isinstance(chunk_id, str):
            graph["quarantined"].append(_quarantined_record(claim, ["missing_source_context"], context))
            continue
        chunk = chunk_by_key.get((document_uid, chunk_id))
        provenance_reasons = _chunk_provenance_reasons(context, extraction_text, chunk)
        if provenance_reasons:
            graph["quarantined"].append(_quarantined_record(claim, provenance_reasons, context))
            continue
        assert chunk is not None
        evidence_id = stable_id("evidence", document_uid, chunk_id)
        documents.setdefault(
            document_uid,
            {
                "document_uid": document_uid,
                "source_id": _audit_field(chunk, context, "source_id"),
                "source_type": _audit_field(chunk, context, "source_type"),
                "authority_score": _audit_field(chunk, context, "authority_score"),
                "factual_index_allowed": _audit_field(chunk, context, "factual_index_allowed"),
                "status": _audit_field(chunk, context, "status") or "accepted",
                "content_sha256": _audit_field(chunk, context, "content_sha256"),
                "document_sha256": _audit_field(chunk, context, "document_sha256"),
                "text_sha256": _audit_field(chunk, context, "text_sha256"),
                "sha256": _audit_field(chunk, context, "sha256"),
            },
        )
        evidence_chunks.setdefault(
            evidence_id,
            {
                "evidence_chunk_id": evidence_id,
                "document_uid": document_uid,
                "chunk_id": chunk_id,
                "text": chunk["text"],
                "source_id": _audit_field(chunk, context, "source_id"),
                "source_type": _audit_field(chunk, context, "source_type"),
                "authority_score": _audit_field(chunk, context, "authority_score"),
                "factual_index_allowed": _audit_field(chunk, context, "factual_index_allowed"),
                "status": _audit_field(chunk, context, "status") or "accepted",
                "content_sha256": _audit_field(chunk, context, "content_sha256"),
                "document_sha256": _audit_field(chunk, context, "document_sha256"),
                "text_sha256": _audit_field(chunk, context, "text_sha256"),
                "sha256": _audit_field(chunk, context, "sha256"),
            },
        )
        mention_by_id = {mention["mention_id"]: mention for mention in claim.get("mentions", [])}
        subject_mention = mention_by_id.get(claim.get("subject", {}).get("mention_id"))
        object_ref = claim.get("object", {}).get("mention_id") if isinstance(claim.get("object"), dict) else None
        object_mention = mention_by_id.get(object_ref) if object_ref else None
        if subject_mention is None or (object_ref and object_mention is None):
            graph["quarantined"].append(_quarantined_record(claim, ["graph_endpoint_missing"], context))
            continue
        occurrence_id = stable_id("occurrence", document_uid, chunk_id, extraction_index, claim.get("claim_id"))

        def entity_semantic_id(mention: dict[str, Any]) -> str:
            if not str(mention["entity_id"]).startswith("NIL-"):
                return "entity:" + str(mention["entity_id"])
            return "entity-name:" + _semantic_key(mention.get("canonical_name") or mention["text"])

        subject_key = entity_semantic_id(subject_mention)
        if object_mention is not None:
            object_key = entity_semantic_id(object_mention)
        else:
            object_value = claim.get("object", {}).get("value")
            object_key = "value:" + _semantic_key(object_value)
        semantic_signature = json.dumps(
            [
                subject_key,
                claim.get("predicate"),
                object_key,
                claim.get("polarity"),
                claim.get("modality"),
                claim.get("attribution"),
            ],
            ensure_ascii=False,
            sort_keys=True,
            default=str,
        )
        claim_id = "CLM-" + stable_id("claim", document_uid, semantic_signature)
        claim_row = claims.setdefault(
            claim_id,
            {
                "claim_id": claim_id,
                "document_uid": document_uid,
                "subject_entity_id": subject_mention["entity_id"],
                "predicate": claim["predicate"],
                "object_entity_id": object_mention["entity_id"] if object_mention else None,
                "object_value": claim.get("object", {}).get("value") if not object_mention else None,
                "model_confidence": claim["model_confidence"],
                "model_confidences": [claim["model_confidence"]],
                "polarity": claim["polarity"],
                "modality": claim["modality"],
                "attribution": claim["attribution"],
                "stance": claim["evidence_stance"],
                "evidence_stances": [claim["evidence_stance"]],
                "review_status": "accepted",
                "status": "accepted",
                "evidence_occurrence_ids": [],
            },
        )
        if claim_row["model_confidences"][-1] != claim["model_confidence"]:
            claim_row["model_confidences"].append(claim["model_confidence"])
        if claim["evidence_stance"] not in claim_row["evidence_stances"]:
            claim_row["evidence_stances"].append(claim["evidence_stance"])
        for mention in claim.get("mentions", []):
            mention_id = stable_id("mention", occurrence_id, mention["mention_id"])
            mentions[mention_id] = {
                "mention_id": mention_id,
                "document_uid": document_uid,
                "evidence_chunk_id": evidence_id,
                "claim_occurrence_id": occurrence_id,
                "text": mention["text"],
                "start": mention["start"],
                "end": mention["end"],
                "entity_id": mention["entity_id"],
                "entity_status": mention["entity_status"],
            }
            entities.setdefault(
                mention["entity_id"],
                {
                    "entity_id": mention["entity_id"],
                    "canonical_name": mention["canonical_name"],
                    "entity_status": mention["entity_status"],
                    "status": "accepted",
                },
            )
        if not any(edge["source_id"] == claim_id and edge["type"] == "has_subject" for edge in edges):
            add_edge(claim_id, "has_subject", subject_mention["entity_id"], target_type="entity")
            if object_mention is not None:
                add_edge(claim_id, "has_object", object_mention["entity_id"], target_type="entity")
            else:
                value_id = "VAL-" + stable_id("value", document_uid, object_key)
                add_edge(
                    claim_id,
                    "has_object",
                    value_id,
                    target_type="value",
                    value=claim.get("object", {}).get("value"),
                )
        add_edge(evidence_id, claim["evidence_stance"], claim_id, stance=claim["evidence_stance"], occurrence_id=occurrence_id)
        add_edge(evidence_id, "from_document", document_uid, occurrence_id=occurrence_id)
        add_edge(claim_id, "evidenced_by", evidence_id, occurrence_id=occurrence_id)
        claim_row["evidence_occurrence_ids"].append(occurrence_id)

    graph["Document"] = list(documents.values())
    graph["EvidenceChunk"] = list(evidence_chunks.values())
    graph["Mention"] = list(mentions.values())
    graph["Entity"] = list(entities.values())
    graph["Claim"] = list(claims.values())
    graph["edges"] = edges
    graph["Edges"] = edges
    return graph


_RETRIEVAL_WEIGHT_KEYS = (
    "query_entity",
    "predicate",
    "polarity",
    "modality",
    "stance",
    "review_state",
    "authority",
    "extraction_confidence",
    "seed_score",
    "hop_decay",
)
_RRF_BRANCH_NAMES = frozenset({"dense", "bm25", "graph"})


def _finite_number(value: Any, name: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(value):
        raise ValueError(f"{name} must be a finite number")
    return float(value)


def _validate_retrieval_config(config: Any, *, expansion: bool = True) -> dict[str, Any]:
    if not isinstance(config, dict):
        raise TypeError("retrieval config must be an object")
    required = {"minimum_rerank_probability", "max_evidence"}
    if expansion:
        required |= {
            "max_hops",
            "minimum_dense_score",
            "minimum_graph_score",
            "hop_decay",
            "weights",
            "review_state_scores",
        }
    missing = sorted(key for key in required if key not in config)
    if missing:
        raise ValueError(f"retrieval config missing required fields: {', '.join(missing)}")

    normalized = dict(config)
    if "max_hops" in config:
        max_hops = config["max_hops"]
        if isinstance(max_hops, bool) or not isinstance(max_hops, int) or not 0 <= max_hops <= 8:
            raise ValueError("max_hops must be an integer from 0 through 8")
        normalized["max_hops"] = max_hops
    for key in ("minimum_dense_score", "minimum_graph_score", "hop_decay"):
        if key in config:
            normalized[key] = _finite_number(config[key], key)
    for key in ("minimum_dense_score", "minimum_graph_score"):
        if key in normalized and normalized[key] < 0:
            raise ValueError(f"{key} must be non-negative")
    if "hop_decay" in config and not 0 < normalized["hop_decay"] <= 1:
        raise ValueError("hop_decay must be greater than 0 and at most 1")
    minimum_probability = _finite_number(
        config["minimum_rerank_probability"], "minimum_rerank_probability"
    )
    if not 0 <= minimum_probability <= 1:
        raise ValueError("minimum_rerank_probability must be between 0 and 1")
    normalized["minimum_rerank_probability"] = minimum_probability
    max_evidence = config["max_evidence"]
    if isinstance(max_evidence, bool) or not isinstance(max_evidence, int) or max_evidence < 1:
        raise ValueError("max_evidence must be a positive integer")
    normalized["max_evidence"] = max_evidence

    if expansion:
        weights = config["weights"]
        if not isinstance(weights, dict):
            raise TypeError("retrieval config weights must be an object")
        missing_weights = [key for key in _RETRIEVAL_WEIGHT_KEYS if key not in weights]
        if missing_weights:
            raise ValueError(
                "retrieval config weights missing required fields: "
                + ", ".join(missing_weights)
            )
        normalized_weights = {
            key: _finite_number(weights[key], f"weights.{key}")
            for key in _RETRIEVAL_WEIGHT_KEYS
        }
        if any(value < 0 for value in normalized_weights.values()):
            raise ValueError("retrieval config weights cannot be negative")
        if not any(value > 0 for value in normalized_weights.values()):
            raise ValueError("retrieval config requires at least one positive weight")
        normalized["weights"] = normalized_weights
        review_scores = config["review_state_scores"]
        if not isinstance(review_scores, dict):
            raise TypeError("review_state_scores must be an object")
        if "accepted" not in review_scores or "reviewed" not in review_scores:
            raise ValueError("review_state_scores must include accepted and reviewed")
        normalized["review_state_scores"] = {
            str(key): _finite_number(value, f"review_state_scores.{key}")
            for key, value in review_scores.items()
        }
    return normalized


def _query_values(query_signature: dict[str, Any], *names: str) -> tuple[str, ...]:
    value: Any = None
    for name in names:
        if name in query_signature:
            value = query_signature[name]
            break
    if value is None:
        return ()
    if isinstance(value, str):
        value = [value]
    if not isinstance(value, (list, tuple, set, frozenset)):
        raise TypeError(f"query field {names[0]} must be a string or sequence of strings")
    normalized: list[str] = []
    for item in value:
        if not isinstance(item, str) or not item:
            raise ValueError(f"query field {names[0]} contains an invalid value")
        normalized.append(item)
    return tuple(dict.fromkeys(normalized))


def _normalize_query_signature(query_signature: Any) -> dict[str, tuple[str, ...]]:
    if not isinstance(query_signature, dict):
        raise TypeError("query_signature must be an object")
    return {
        "entity_ids": _query_values(query_signature, "entity_ids", "canonical_entity_ids", "entity_id"),
        "predicates": _query_values(query_signature, "predicates", "predicate"),
        "polarities": _query_values(query_signature, "polarities", "polarity"),
        "modalities": _query_values(query_signature, "modalities", "modality"),
        "stances": _query_values(
            query_signature,
            "desired_stances",
            "desired_stance",
            "useful_stances",
            "useful_stance",
            "stances",
            "stance",
        ),
    }


def _table_index(graph_tables: dict[str, Any], table_name: str, key: str) -> dict[str, dict[str, Any]]:
    rows = graph_tables.get(table_name, [])
    if not isinstance(rows, list):
        raise TypeError(f"graph table {table_name} must be a list")
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        if isinstance(row, dict) and isinstance(row.get(key), str):
            result[row[key]] = row
    return result


def _edge_indexes(
    graph_tables: dict[str, Any],
    node_ids: set[str],
    evidence_ids: set[str],
    claim_ids: set[str],
) -> tuple[dict[str, list[tuple[str, str, dict[str, Any]]]], dict[str, set[str]]]:
    edges = graph_tables.get("edges", graph_tables.get("Edges", []))
    if not isinstance(edges, list):
        raise TypeError("graph edges must be a list")
    adjacency: dict[str, list[tuple[str, str, dict[str, Any]]]] = defaultdict(list)
    claims_by_evidence: dict[str, set[str]] = defaultdict(set)
    for edge in edges:
        if not isinstance(edge, dict):
            continue
        source_id = edge.get("source_id")
        target_id = edge.get("target_id")
        edge_type = edge.get("type")
        if not all(isinstance(value, str) for value in (source_id, target_id, edge_type)):
            continue
        if source_id not in node_ids or target_id not in node_ids:
            continue
        if edge_type not in EVIDENCE_STANCES | {"has_subject", "has_object", "evidenced_by"}:
            continue
        adjacency[source_id].append((target_id, edge_type, edge))
        adjacency[target_id].append((source_id, f"reverse:{edge_type}", edge))
        if edge_type in EVIDENCE_STANCES and source_id in evidence_ids and target_id in claim_ids:
            claims_by_evidence[source_id].add(target_id)
        elif edge_type == "evidenced_by" and target_id in evidence_ids and source_id in claim_ids:
            claims_by_evidence[target_id].add(source_id)
    for neighbors in adjacency.values():
        neighbors.sort(key=lambda item: (item[0], item[1]))
    return adjacency, claims_by_evidence


def _seed_evidence_id(
    hit: dict[str, Any], evidence_by_id: dict[str, dict[str, Any]], evidence_by_chunk: dict[tuple[str, str], str]
) -> str | None:
    if "evidence_chunk_id" in hit:
        evidence_id = hit["evidence_chunk_id"]
        if isinstance(evidence_id, str) and evidence_id in evidence_by_id:
            return evidence_id
        return None
    key = (hit.get("document_uid"), hit.get("chunk_id"))
    return evidence_by_chunk.get(key)


def _claim_entity_ids(claim: dict[str, Any]) -> set[str]:
    return {
        value
        for key in ("subject_entity_id", "object_entity_id")
        if isinstance((value := claim.get(key)), str) and value
    }


def _component(value: float, weight: float, role: str) -> dict[str, Any]:
    return {
        "value": float(value),
        "weight": float(weight),
        "contribution": float(value * weight),
        "role": role,
    }


def expand_graph_from_seeds(
    seed_hits: Iterable[dict[str, Any]],
    query_signature: dict[str, Any],
    graph_tables: dict[str, Any],
    config: dict[str, Any],
) -> list[dict[str, Any]]:
    """Expand accepted semantic evidence from dense/BM25 seeds with auditable traces.

    Query fields are canonical entity IDs, predicates, polarity, modality, and desired
    or useful stance. Seeds must already be produced by semantic dense/BM25 retrieval;
    this function only filters them and traverses explicit typed graph edges.
    """

    normalized_config = _validate_retrieval_config(config, expansion=True)
    query = _normalize_query_signature(query_signature)
    if not isinstance(graph_tables, dict):
        raise TypeError("graph_tables must be an object")
    evidence_by_id = _table_index(graph_tables, "EvidenceChunk", "evidence_chunk_id")
    document_by_id = _table_index(graph_tables, "Document", "document_uid")
    claim_by_id = _table_index(graph_tables, "Claim", "claim_id")
    evidence_by_chunk = {
        (row.get("document_uid"), row.get("chunk_id")): evidence_id
        for evidence_id, row in evidence_by_id.items()
        if isinstance(row.get("document_uid"), str) and isinstance(row.get("chunk_id"), str)
    }
    node_ids = set(evidence_by_id) | set(document_by_id) | set(claim_by_id)
    node_ids |= {
        str(row["entity_id"])
        for row in graph_tables.get("Entity", [])
        if isinstance(row, dict) and isinstance(row.get("entity_id"), str)
    }
    adjacency, claims_by_evidence = _edge_indexes(
        graph_tables, node_ids, set(evidence_by_id), set(claim_by_id)
    )

    retained_seeds_by_branch: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for hit in seed_hits or []:
        if not isinstance(hit, dict):
            raise TypeError("each seed hit must be an object")
        branch = hit.get("branch")
        if branch not in {"dense", "bm25"}:
            raise ValueError("seed branch must be exactly dense or bm25")
        rank = hit.get("rank")
        if isinstance(rank, bool) or not isinstance(rank, int) or rank < 1:
            raise ValueError("seed rank must be a positive integer")
        raw_score = _finite_number(hit.get("raw_score"), "seed raw_score")
        if raw_score <= 0:
            continue
        if branch == "dense" and raw_score < normalized_config["minimum_dense_score"]:
            continue
        evidence_id = _seed_evidence_id(hit, evidence_by_id, evidence_by_chunk)
        if evidence_id is None:
            continue
        retained_seeds_by_branch[branch].append(
            {
                "branch": branch,
                "rank": rank,
                "raw_score": raw_score,
                "chunk_id": hit.get("chunk_id"),
                "document_uid": hit.get("document_uid"),
                "evidence_chunk_id": evidence_id,
                "source_trace": hit.get("source_trace", hit.get("trace")),
            }
        )
    seeds_by_evidence: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for branch in sorted(retained_seeds_by_branch):
        branch_seeds = retained_seeds_by_branch[branch]
        raw_scores = [seed["raw_score"] for seed in branch_seeds]
        minimum = min(raw_scores)
        maximum = max(raw_scores)
        spread = maximum - minimum
        for seed in branch_seeds:
            seed["normalized_score"] = (
                1.0 if spread == 0 else (seed["raw_score"] - minimum) / spread
            )
            seeds_by_evidence[seed["evidence_chunk_id"]].append(seed)
    if not seeds_by_evidence:
        return []

    paths_by_evidence: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for seed_id in sorted(seeds_by_evidence):
        paths_by_evidence[seed_id].append({"nodes": [seed_id], "relations": [], "hop": 0})
        frontier = [(seed_id, (seed_id,), ())]
        while frontier:
            node_id, path, relations = frontier.pop(0)
            hop = len(relations)
            if hop >= normalized_config["max_hops"]:
                continue
            for neighbor, relation, _edge in adjacency.get(node_id, []):
                if neighbor in path:
                    continue
                next_path = path + (neighbor,)
                next_relations = relations + (relation,)
                if neighbor in evidence_by_id:
                    paths_by_evidence[neighbor].append(
                        {
                            "nodes": list(next_path),
                            "relations": list(next_relations),
                            "hop": hop + 1,
                        }
                    )
                frontier.append((neighbor, next_path, next_relations))

    results: list[dict[str, Any]] = []
    weights = normalized_config["weights"]
    review_scores = normalized_config["review_state_scores"]
    for evidence_id in sorted(paths_by_evidence):
        evidence = evidence_by_id[evidence_id]
        paths = paths_by_evidence[evidence_id]
        unique_paths: dict[str, dict[str, Any]] = {}
        for path in paths:
            marker = json.dumps(path, sort_keys=True, separators=(",", ":"))
            unique_paths[marker] = path
        graph_paths = sorted(unique_paths.values(), key=lambda path: (path["hop"], path["nodes"], path["relations"]))
        hop = min(path["hop"] for path in graph_paths)
        connected_claim_ids = sorted(claims_by_evidence.get(evidence_id, set()))
        claims = [claim_by_id[claim_id] for claim_id in connected_claim_ids if claim_id in claim_by_id]
        # A dense/BM25 seed is not a graph result by itself.  In particular,
        # claimless chunks must never acquire a positive score from authority,
        # review state, or hop-0 decay alone.
        if hop == 0 and not claims:
            continue
        entity_matches = any(_claim_entity_ids(claim) & set(query["entity_ids"]) for claim in claims)
        predicate_matches = any(claim.get("predicate") in query["predicates"] for claim in claims)
        polarity_matches = any(claim.get("polarity") in query["polarities"] for claim in claims)
        modality_matches = any(claim.get("modality") in query["modalities"] for claim in claims)
        # A hop-0 seed is not a semantic graph hit merely because an accepted
        # claim has authority/confidence. Require at least one requested
        # entity, predicate, polarity, or modality match.
        if hop == 0 and not (entity_matches or predicate_matches):
            continue
        stance_matches = any(
            claim.get("stance") in query["stances"] or claim.get("evidence_stance") in query["stances"]
            for claim in claims
        )
        review_state = str(evidence.get("status") or evidence.get("review_status") or "unknown")
        review_value = review_scores.get(review_state, review_scores.get("unknown", 0.0))
        document = document_by_id.get(evidence.get("document_uid"), {})
        authority = evidence.get("authority_score", document.get("authority_score", 0.0))
        authority_value = min(1.0, max(0.0, _finite_number(authority or 0.0, "authority_score")))
        confidence_value = max(
            [_finite_number(claim.get("model_confidence", 0.0), "model_confidence") for claim in claims] or [0.0]
        )
        seed_hits_for_evidence = sorted(
            seeds_by_evidence.get(evidence_id, []),
            key=lambda hit: (
                -hit["normalized_score"],
                -hit["raw_score"],
                str(hit.get("branch")),
                hit["rank"],
            ),
        )
        seed_value = seed_hits_for_evidence[0]["normalized_score"] if seed_hits_for_evidence else 0.0
        hop_value = normalized_config["hop_decay"] ** hop
        values = {
            "query_entity": 1.0 if entity_matches else 0.0,
            "predicate": 1.0 if predicate_matches else 0.0,
            "polarity": 1.0 if polarity_matches else 0.0,
            "modality": 1.0 if modality_matches else 0.0,
            "stance": 1.0 if stance_matches else 0.0,
            "review_state": review_value,
            "authority": authority_value,
            "extraction_confidence": confidence_value,
            "seed_score": seed_value,
            "hop_decay": hop_value,
        }
        roles = {
            "query_entity": "canonical entity ID match on a connected claim",
            "predicate": "predicate match on a connected claim",
            "polarity": "polarity qualifier match on a connected claim",
            "modality": "modality qualifier match on a connected claim",
            "stance": "desired/useful stance match on a connected claim",
            "review_state": "configured accepted/reviewed state score",
            "authority": "bounded source authority score",
            "extraction_confidence": "maximum validated claim confidence",
            "seed_score": "maximum retained per-branch normalized dense/BM25 seed score",
            "hop_decay": "configured decay raised to shortest graph hop",
        }
        components = {
            key: _component(values[key], weights[key], roles[key]) for key in _RETRIEVAL_WEIGHT_KEYS
        }
        graph_score = sum(item["contribution"] for item in components.values())
        if graph_score <= max(0.0, normalized_config["minimum_graph_score"]):
            continue
        relations = sorted({relation for path in graph_paths for relation in path["relations"]})
        stances = sorted(
            {
                str(claim.get("stance"))
                for claim in claims
                if isinstance(claim.get("stance"), str)
            }
        )
        trace = {
            "components": components,
            "total": graph_score,
            "relations": relations,
            "stances": stances,
            "seed_hits": seed_hits_for_evidence,
            "query_signature": query,
        }
        results.append(
            {
                **evidence,
                "candidate_id": evidence_id,
                "evidence_chunk_id": evidence_id,
                "graph_score": graph_score,
                "score": graph_score,
                "hop": hop,
                "graph_paths": graph_paths,
                "relations": relations,
                "stances": stances,
                "seed_hits": seed_hits_for_evidence,
                "trace": trace,
                "source_trace": trace,
            }
        )
    results.sort(key=lambda row: (-row["graph_score"], row["evidence_chunk_id"]))
    for rank, row in enumerate(results, start=1):
        row["branch"] = "graph"
        row["rank"] = rank
        row["raw_score"] = row["graph_score"]
    return results


def _fusion_candidate_key(hit: dict[str, Any]) -> str:
    for key in ("evidence_chunk_id", "candidate_id"):
        if isinstance(hit.get(key), str) and hit[key]:
            return hit[key]
    document_uid = hit.get("document_uid")
    chunk_id = hit.get("chunk_id")
    if isinstance(document_uid, str) and isinstance(chunk_id, str):
        return f"{document_uid}:{chunk_id}"
    if isinstance(chunk_id, str) and chunk_id:
        return chunk_id
    raise ValueError("fusion hit requires an evidence_chunk_id, candidate_id, or chunk_id")


def reciprocal_rank_fusion(
    branch_hits: dict[str, Iterable[dict[str, Any]]] | Iterable[dict[str, Any]],
    weights: dict[str, float],
    constant: float,
) -> list[dict[str, Any]]:
    """Fuse branches while retaining every rank, score, contribution, path, and stance."""

    if isinstance(branch_hits, dict):
        grouped = {str(branch): list(hits or []) for branch, hits in branch_hits.items()}
    else:
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for hit in branch_hits or []:
            if not isinstance(hit, dict) or not isinstance(hit.get("branch"), str):
                raise TypeError("sequence fusion hits must include a branch")
            grouped[hit["branch"]].append(hit)
    if any(branch not in _RRF_BRANCH_NAMES for branch in grouped):
        raise ValueError("RRF branch names must be exactly dense, bm25, or graph")
    if not isinstance(weights, dict) or not weights:
        raise ValueError("fusion weights must be a non-empty object")
    if any(str(branch) not in _RRF_BRANCH_NAMES for branch in weights):
        raise ValueError("RRF weight names must be exactly dense, bm25, or graph")
    constant_value = _finite_number(constant, "rrf constant")
    if constant_value <= 0:
        raise ValueError("rrf constant must be greater than 0")
    normalized_weights = {}
    for branch, weight in weights.items():
        normalized_weights[str(branch)] = _finite_number(weight, f"weights.{branch}")
        if normalized_weights[str(branch)] < 0:
            raise ValueError("fusion weights cannot be negative")
    candidates: dict[str, dict[str, Any]] = {}
    for branch in sorted(grouped):
        if branch not in normalized_weights:
            raise ValueError(f"missing fusion weight for branch: {branch}")
        best_hits: dict[str, dict[str, Any]] = {}
        for hit in grouped[branch]:
            if not isinstance(hit, dict):
                raise TypeError("each fusion hit must be an object")
            raw_score = _finite_number(hit.get("raw_score"), "fusion raw_score")
            if raw_score <= 0:
                continue
            graph_score = hit.get("graph_score")
            if branch == "graph":
                if (
                    isinstance(graph_score, bool)
                    or not isinstance(graph_score, (int, float))
                    or not math.isfinite(graph_score)
                    or graph_score <= 0
                ):
                    continue
            elif graph_score is not None and _finite_number(graph_score, "fusion graph_score") <= 0:
                continue
            rank = hit.get("rank")
            if isinstance(rank, bool) or not isinstance(rank, int) or rank < 1:
                raise ValueError("fusion rank must be a positive integer")
            key = _fusion_candidate_key(hit)
            previous = best_hits.get(key)
            ordering = (rank, -raw_score, json.dumps(hit, sort_keys=True, default=str))
            previous_ordering = (
                (previous["rank"], -previous["raw_score"], json.dumps(previous, sort_keys=True, default=str))
                if previous is not None
                else None
            )
            if previous is None or ordering < previous_ordering:
                best_hits[key] = {"hit": hit, "rank": rank, "raw_score": raw_score}
        for key in sorted(best_hits):
            hit = best_hits[key]["hit"]
            raw_score = best_hits[key]["raw_score"]
            rank = best_hits[key]["rank"]
            row = candidates.setdefault(
                key,
                {
                    "candidate_id": key,
                    "evidence_chunk_id": hit.get("evidence_chunk_id", key),
                    "chunk_id": hit.get("chunk_id"),
                    "document_uid": hit.get("document_uid"),
                    "rrf_score": 0.0,
                    "branch_traces": [],
                    "graph_paths": [],
                    "relations": [],
                    "stances": [],
                },
            )
            contribution = normalized_weights[branch] / (constant_value + rank)
            row["rrf_score"] += contribution
            graph_paths = hit.get("graph_paths", []) if isinstance(hit.get("graph_paths"), list) else []
            relations = hit.get("relations") if isinstance(hit.get("relations"), list) else [hit.get("relation")]
            stances = hit.get("stances") if isinstance(hit.get("stances"), list) else [hit.get("stance")]
            source_trace = hit.get("source_trace", hit.get("trace"))
            row["branch_traces"].append(
                {
                    "branch": branch,
                    "rank": rank,
                    "raw_score": raw_score,
                    "weight": normalized_weights[branch],
                    "weighted_contribution": contribution,
                    "score": contribution,
                    "graph_paths": list(graph_paths),
                    "relations": [relation for relation in relations if relation is not None],
                    "stances": [stance for stance in stances if stance is not None],
                    "source_trace": source_trace,
                }
            )
            for path in graph_paths:
                if path not in row["graph_paths"]:
                    row["graph_paths"].append(path)
            for relation in relations:
                if relation is not None and relation not in row["relations"]:
                    row["relations"].append(relation)
            for stance in stances:
                if stance is not None and stance not in row["stances"]:
                    row["stances"].append(stance)
    for row in candidates.values():
        row["branch_traces"].sort(key=lambda item: (item["branch"], item["rank"], item["raw_score"]))
        row["graph_paths"].sort(key=lambda path: json.dumps(path, sort_keys=True, default=str))
        row["relations"].sort()
        row["stances"].sort()
    return sorted(candidates.values(), key=lambda row: (-row["rrf_score"], row["candidate_id"]))


def _stable_sigmoid(logit: float) -> float:
    if logit >= 0:
        z = math.exp(-logit)
        return 1.0 / (1.0 + z)
    z = math.exp(logit)
    return z / (1.0 + z)


def select_evidence(
    candidates: Iterable[dict[str, Any]],
    rerank_scores: dict[str, float] | Iterable[float] | None,
    config: dict[str, Any],
) -> dict[str, Any]:
    """Convert reranker logits to probabilities and abstain below the configured floor."""

    normalized_config = _validate_retrieval_config(config, expansion=False)
    candidate_rows = list(candidates or [])
    for candidate in candidate_rows:
        if not isinstance(candidate, dict):
            raise TypeError("each evidence candidate must be an object")
    score_by_id: dict[str, float] = {}
    score_sequence: list[float] | None = None
    if isinstance(rerank_scores, dict):
        for key, value in rerank_scores.items():
            score_by_id[str(key)] = _finite_number(value, f"rerank_scores.{key}")
    elif rerank_scores is not None:
        score_sequence = [
            _finite_number(value, "rerank score") for value in list(rerank_scores)
        ]
        if len(score_sequence) != len(candidate_rows):
            raise ValueError("rerank score sequence must align with candidates")

    selected: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    for index, candidate in enumerate(candidate_rows):
        candidate_id = candidate.get("evidence_chunk_id", candidate.get("candidate_id", candidate.get("chunk_id")))
        if score_sequence is not None:
            logit = score_sequence[index]
        elif isinstance(candidate.get("rerank_logit"), (int, float)) and not isinstance(candidate.get("rerank_logit"), bool):
            logit = _finite_number(candidate["rerank_logit"], "candidate rerank_logit")
        elif candidate_id in score_by_id:
            logit = score_by_id[candidate_id]
        else:
            raise ValueError(f"missing rerank score for candidate: {candidate_id}")
        probability = _stable_sigmoid(logit)
        row = dict(candidate)
        row["rerank_logit"] = logit
        row["rerank_probability"] = probability
        if probability >= normalized_config["minimum_rerank_probability"]:
            selected.append(row)
        else:
            rejected.append(row)
    selected.sort(
        key=lambda row: (-row["rerank_probability"], str(row.get("evidence_chunk_id", row.get("candidate_id", ""))))
    )
    selected = selected[: normalized_config["max_evidence"]]
    if not candidate_rows:
        return {"selected": [], "rejected": [], "abstained": True, "reason": "no_candidates"}
    if not selected:
        return {
            "selected": [],
            "rejected": rejected,
            "abstained": True,
            "reason": "all_candidates_below_minimum_rerank_probability",
        }
    return {"selected": selected, "rejected": rejected, "abstained": False, "reason": None}


def split_qwen_thinking(raw_output: Any) -> dict[str, Any]:
    """Separate Qwen-emitted thinking from the user-facing final content.

    Qwen-family generation can contain a complete ``<think>`` block, or only
    the closing tag when the opening tag was supplied by the chat template.
    An opening tag without a close is preserved and explicitly marked as a
    truncated generation instead of being mistaken for a final answer.
    """

    raw = str(raw_output or "")
    complete = re.search(r"<think>\s*(.*?)\s*</think>", raw, flags=re.DOTALL | re.IGNORECASE)
    if complete:
        return {
            "reasoning_content": complete.group(1).strip(),
            "final_content": (raw[0 : complete.start()] + raw[complete.end() :]).strip(),
            "thinking_status": "complete",
            "reasoning_truncated": False,
        }
    closing = re.search(r"</think>", raw, flags=re.IGNORECASE)
    if closing:
        return {
            "reasoning_content": raw[0 : closing.start()].removeprefix("<think>").strip(),
            "final_content": raw[closing.end() :].strip(),
            "thinking_status": "complete_closing_tag_only",
            "reasoning_truncated": False,
        }
    opening = re.search(r"<think>", raw, flags=re.IGNORECASE)
    if opening:
        return {
            "reasoning_content": raw[opening.end() :].strip(),
            "final_content": "",
            "thinking_status": "truncated",
            "reasoning_truncated": True,
        }
    return {
        "reasoning_content": "",
        "final_content": raw.strip(),
        "thinking_status": "not_emitted",
        "reasoning_truncated": False,
    }

CORE_SOURCE_SHA256 = '14271e34f998c2a5448094ba72f1f27dce4f9225c7e239fb9f5bc213d5f6051c'



# %% [notebook cell 3]

"""Pure, dependency-light evaluation and audit helpers.

The functions in this module deliberately return nullable values for missing
inputs and validate statistical inputs instead of silently coercing them.
They are also suitable for injecting verbatim into the generated notebook.
"""


import math
import numbers
import random
import unicodedata
from collections import Counter, OrderedDict
from collections.abc import Callable, Mapping, MutableMapping, Sequence
import hashlib
import re
from typing import Any


_CITATION_TOKEN_RE = __import__("re").compile(r"\[(E[1-9][0-9]*)\]")


class BoundedLRU(MutableMapping):
    """Small deterministic LRU cache with a hard memory bound."""

    def __init__(self, maxsize: int = 128):
        if isinstance(maxsize, bool) or not isinstance(maxsize, int) or maxsize < 1:
            raise ValueError("maxsize must be a positive integer")
        self.maxsize = maxsize
        self._data = OrderedDict()
        self.evictions = 0

    def __getitem__(self, key):
        value = self._data.pop(key)
        self._data[key] = value
        return value

    def __setitem__(self, key, value):
        if key in self._data:
            del self._data[key]
        self._data[key] = value
        while len(self._data) > self.maxsize:
            self._data.popitem(last=False)
            self.evictions += 1

    def __delitem__(self, key):
        del self._data[key]

    def __iter__(self):
        return iter(self._data)

    def __len__(self):
        return len(self._data)

    def __contains__(self, key):
        return key in self._data

    def get(self, key, default=None):
        try:
            return self[key]
        except KeyError:
            return default

    def peek(self, key, default=None):
        """Read without changing recency, useful for diagnostics."""
        return self._data.get(key, default)

    def snapshot(self) -> dict[str, Any]:
        """Return a detached JSON-safe cache diagnostic snapshot."""
        return {
            "capacity": self.maxsize,
            "size": len(self._data),
            "evictions": self.evictions,
            "items": [{"key": _json_safe_cache_value(key), "value": _json_safe_cache_value(value)} for key, value in self._data.items()],
        }


def _json_safe_cache_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Mapping):
        return {str(key): _json_safe_cache_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe_cache_value(item) for item in value]
    return str(value)


def derive_shard_run_name(base_name: Any, shard_index: Any = 0, shard_count: Any = 1) -> str:
    """Return a unique, validated run directory name for one shard."""

    base = normalize_optional_text(base_name)
    if base is None or "/" in base or "\\" in base or base in {".", ".."}:
        raise ValueError("shard run base name must be a non-empty path component")
    if isinstance(shard_index, bool) or not isinstance(shard_index, int):
        raise ValueError("shard index must be an integer")
    if isinstance(shard_count, bool) or not isinstance(shard_count, int):
        raise ValueError("shard count must be an integer")
    if shard_count < 1 or shard_index < 0 or shard_index >= shard_count:
        raise ValueError("invalid shard range")
    return f"{base}__shard-{shard_index}-of-{shard_count}"


def derive_effective_cache_capacity(base_capacity: Any, *, row_limit: Any, shard_count: Any = 1) -> int:
    """Derive a bounded cache capacity before config/run identity is hashed."""
    if isinstance(base_capacity, bool) or not isinstance(base_capacity, int) or base_capacity < 1:
        raise ValueError("cache capacity must be a positive integer")
    if isinstance(row_limit, bool) or not isinstance(row_limit, int) or row_limit < 0:
        raise ValueError("row limit must be a non-negative integer")
    if isinstance(shard_count, bool) or not isinstance(shard_count, int) or shard_count < 1:
        raise ValueError("shard count must be a positive integer")
    return max(base_capacity, (row_limit + shard_count - 1) // shard_count)


def identity_config(config: Mapping[str, Any], *, shard_index: Any = None, shard_count: Any = None) -> dict[str, Any]:
    """Build shard-aware identity while keeping split identity shard-independent.

    ``config_hash`` and ``split_identity_hash`` intentionally exclude shard
    assignment. ``run_identity_hash`` includes the validated shard identity so
    checkpoints and logs cannot be reused across workers.
    """

    if not isinstance(config, Mapping):
        raise TypeError("config must be a mapping")
    values = dict(config)
    configured_index = values.pop("shard_index", 0)
    configured_count = values.pop("shard_count", 1)
    index = configured_index if shard_index is None else shard_index
    count = configured_count if shard_count is None else shard_count
    if isinstance(index, bool) or not isinstance(index, int) or isinstance(count, bool) or not isinstance(count, int):
        raise ValueError("shard index/count must be integers")
    if count < 1 or index < 0 or index >= count:
        raise ValueError("invalid shard range")
    canonical = jsonable_identity(values)
    split_keys = {key: canonical[key] for key in sorted(canonical) if key not in {"shard_index", "shard_count", "run_name", "run_dir"}}
    shard = {"index": index, "count": count}
    config_hash = stable_identity_hash(canonical)
    split_identity_hash = stable_identity_hash(split_keys)
    run_identity_hash = stable_identity_hash({"config": canonical, "shard": shard})
    return {"config": canonical, "config_hash": config_hash, "split_identity_hash": split_identity_hash, "run_identity_hash": run_identity_hash, "shard": shard}


def jsonable_identity(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): jsonable_identity(value[key]) for key in sorted(value, key=str)}
    if isinstance(value, (list, tuple)):
        return [jsonable_identity(item) for item in value]
    if hasattr(value, "as_posix"):
        return str(value)
    return value


def stable_identity_hash(value: Any) -> str:
    payload = __import__("json").dumps(jsonable_identity(value), ensure_ascii=False, sort_keys=True, default=str, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def normalize_category(value: Any) -> str | None:
    """Normalize category labels before matching or stratification."""

    text = normalize_optional_text(value)
    return " ".join(text.split()).casefold() if text is not None else None


def filter_rows_by_category(rows: Sequence[Mapping[str, Any]], allowed_categories: Sequence[Any], *, category_key: str = "Category", id_key: str = "ID") -> dict[str, Any]:
    allowed = {normalize_category(value) for value in allowed_categories}
    allowed.discard(None)
    decisions = []
    reasons = []
    for row in rows:
        category = normalize_category(row.get(category_key) if isinstance(row, Mapping) else None)
        keep = category in allowed
        decisions.append(keep)
        reasons.append(None if keep else "category_not_allowed")
    result = filter_audit(rows, decisions, filter_name="category", reasons=reasons, id_key=id_key)
    result["normalized_allowed_categories"] = sorted(allowed)
    return {"rows": result.pop("kept_rows"), "manifest": result}


def quarantine_missing_references(rows: Sequence[Mapping[str, Any]], *, reference_key: str = "Counter Narrative", id_key: str = "ID") -> dict[str, Any]:
    """Keep all rows for generation while excluding missing references from scoring."""

    missing = [normalize_optional_text(row.get(reference_key) if isinstance(row, Mapping) else None) is None for row in rows]
    reasons = ["missing_reference" if flag else None for flag in missing]
    audit = filter_audit(rows, [not flag for flag in missing], filter_name="reference", reasons=reasons, id_key=id_key)
    quarantined = [{"record_id": str(row.get(id_key, index)), "row_index": index, "reason": "missing_reference", "row": row} for index, (row, flag) in enumerate(zip(rows, missing)) if flag]
    audit_rows = audit.pop("manifest_rows")
    return {"generation_rows": list(rows), "scorable_rows": audit.pop("kept_rows"), "quarantined_rows": quarantined, "manifest": {**audit, "rows": audit_rows, "purpose": "reference_metrics_only"}}


def validate_evidence_ledger(
    evidence_ledger: Sequence[Mapping[str, Any]],
    source_text_lookup: Mapping[str, str] | None = None,
) -> list[dict[str, Any]]:
    """Validate canonical evidence spans and return a detached copy.

    A ledger is an audit boundary: offsets, displayed text, and its digest
    must agree before any claim-level scoring can consume it.
    """

    if not isinstance(evidence_ledger, Sequence) or isinstance(evidence_ledger, (str, bytes)):
        raise ValueError("ledger must be a sequence")
    validated: list[dict[str, Any]] = []
    seen_evidence_ids: set[str] = set()
    for index, raw in enumerate(evidence_ledger):
        if not isinstance(raw, Mapping):
            raise ValueError(f"ledger row {index} is invalid")
        item = dict(raw)
        evidence_id = item.get("evidence_id")
        if evidence_id in seen_evidence_ids:
            raise ValueError(f"duplicate evidence_id: {evidence_id}")
        seen_evidence_ids.add(evidence_id)
        displayed = item.get("displayed_text")
        start = item.get("span_start")
        end = item.get("span_end")
        digest = item.get("displayed_text_sha256", item.get("text_sha256"))
        if not isinstance(evidence_id, str) or not _CITATION_TOKEN_RE.fullmatch(f"[{evidence_id}]"):
            raise ValueError(f"ledger evidence_id invalid: {evidence_id!r}")
        if not isinstance(displayed, str) or not displayed:
            raise ValueError(f"ledger displayed_text invalid for {evidence_id}")
        if isinstance(start, bool) or isinstance(end, bool) or not isinstance(start, int) or not isinstance(end, int) or start < 0 or end < start:
            raise ValueError(f"ledger offsets invalid for {evidence_id}")
        expected_digest = __import__("hashlib").sha256(displayed.encode("utf-8")).hexdigest()
        if digest != expected_digest:
            raise ValueError(f"ledger displayed_text hash mismatch for {evidence_id}")
        source_text = item.get("source_text")
        source_key = item.get("source_text_key")
        if source_text is None and source_text_lookup is not None:
            if not isinstance(source_key, str) or source_key not in source_text_lookup:
                raise ValueError(f"ledger source text lookup missing for {evidence_id}")
            source_text = source_text_lookup[source_key]
        if source_text is not None and (not isinstance(source_text, str) or end > len(source_text)):
            raise ValueError(f"ledger offsets exceed source text for {evidence_id}")
        if source_text is not None and start >= end:
            raise ValueError(f"ledger offsets empty for {evidence_id}")
        if source_text is not None and item.get("source_text_sha256") != __import__("hashlib").sha256(source_text.encode("utf-8")).hexdigest():
            raise ValueError(f"ledger source text hash mismatch for {evidence_id}")
        evidence_text = item.get("evidence_text")
        if source_text is not None and (not isinstance(source_text, str) or source_text[start:end] != displayed):
            raise ValueError(f"ledger source span mismatch for {evidence_id}")
        if source_text is None and evidence_text is not None and evidence_text != displayed:
            raise ValueError(f"ledger evidence text mismatch for {evidence_id}")
        validated.append(item)
    return validated


def select_evidence_within_budget(
    evidence_ledger: Sequence[Mapping[str, Any]],
    char_budget: int,
) -> list[dict[str, Any]]:
    """Select a prefix of whole evidence spans without slicing any span."""

    if not isinstance(char_budget, int) or char_budget < 1:
        raise ValueError("evidence budget must be a positive integer")
    ledger = validate_evidence_ledger(evidence_ledger)
    selected: list[dict[str, Any]] = []
    used = 0
    for item in ledger:
        cost = len(item["displayed_text"])
        if used + cost > char_budget:
            if not selected:
                raise ValueError("evidence budget overflow: first span does not fit")
            break
        selected.append(item)
        used += cost
    return selected


def _sentence_spans(text: str) -> list[dict[str, Any]]:
    """Split text without losing source offsets or inter-sentence whitespace."""
    import re

    if not isinstance(text, str):
        text = str(text)
    if not text:
        return []
    spans: list[dict[str, Any]] = []
    start = 0
    # A sentence ends at terminal punctuation followed by whitespace or EOF.
    # Newlines also delimit prose when no terminal punctuation is present.
    boundary = re.compile(r"[.!?。！？।॥](?:[\"'»”’)]*)?(?=\s|$)|\n+")
    for match in boundary.finditer(text):
        end = match.end()
        if text[start:end].strip():
            spans.append({"text": text[start:end], "start_char": start, "end_char": end})
        start = end
    if start < len(text) and text[start:].strip():
        spans.append({"text": text[start:], "start_char": start, "end_char": len(text)})
    return spans


def sentence_aligned_windows(
    text: Any,
    max_chars: int,
    overlap: int = 0,
) -> list[dict[str, Any]]:
    """Return non-overlapping windows packed at sentence boundaries.

    Every returned ``text`` is an exact slice of the input, and therefore its
    ``start_char``/``end_char`` offsets can be used to audit what was shown to
    a model. ``overlap`` is accepted for call-site compatibility but is
    intentionally not applied: overlapping windows duplicate evidence and
    make citation necessity impossible to audit. A sentence longer than the
    limit is split only at whitespace (with an explicit ``oversized_sentence``
    marker), so the hard character bound is never violated.
    """

    if not isinstance(max_chars, int) or max_chars < 1:
        raise ValueError("max_chars must be a positive integer")
    if not isinstance(overlap, int) or overlap < 0:
        raise ValueError("overlap must be a non-negative integer")
    source = "" if text is None else str(text)
    sentences = _sentence_spans(source)
    if not sentences:
        return []
    windows: list[dict[str, Any]] = []
    index = 0
    while index < len(sentences):
        sentence = sentences[index]
        if len(sentence["text"]) > max_chars:
            # Preserve the original span while making bounded word pieces.
            cursor = sentence["start_char"]
            limit = sentence["end_char"]
            while cursor < limit:
                proposed = min(limit, cursor + max_chars)
                if proposed < limit:
                    boundary = source.rfind(" ", cursor + 1, proposed + 1)
                    if boundary > cursor:
                        proposed = boundary
                if proposed <= cursor:
                    proposed = min(limit, cursor + max_chars)
                windows.append({
                    "text": source[cursor:proposed],
                    "start_char": cursor,
                    "end_char": proposed,
                    "sentence_start": sentence["start_char"],
                    "sentence_end": sentence["end_char"],
                    "sentence_aligned": False,
                    "split_reason": "oversized_sentence",
                })
                cursor = proposed
            index += 1
            continue
        end_index = index + 1
        while end_index < len(sentences):
            candidate_end = sentences[end_index]["end_char"]
            if candidate_end - sentence["start_char"] > max_chars:
                break
            end_index += 1
        final_end = sentences[end_index - 1]["end_char"]
        windows.append({
            "text": source[sentence["start_char"]:final_end],
            "start_char": sentence["start_char"],
            "end_char": final_end,
            "sentence_start": sentence["start_char"],
            "sentence_end": final_end,
            "sentence_aligned": True,
            "split_reason": None,
        })
        index = end_index
    return windows


def _claim_text(value: Any) -> str:
    import re

    value = _CITATION_TOKEN_RE.sub("", str(value or ""))
    value = re.sub(r"\s+([,.;:!?])", r"\1", value)
    value = re.sub(r"\s+", " ", value).strip().casefold()
    return value.rstrip(" .!?。！？।॥")


def build_claim_citation_records(
    narrative: Any,
    evidence_ledger: Sequence[Mapping[str, Any]],
    *,
    factual_claims: Sequence[Any] | None = None,
    abstention_validator: Callable[[str], bool] | None = None,
    safe_non_factual_validator: Callable[[str], bool] | None = None,
) -> list[dict[str, Any]]:
    """Create auditable sentence-level claim/citation records.

    An empty ``factual_claims`` declaration is not trusted on its own. It is a
    safe abstention only when the caller supplies and passes a deterministic
    ``abstention_validator``; otherwise every non-empty narrative sentence is
    treated as factual. Unmatched factual claims are added as uncited records
    so omissions cannot be hidden by sentence segmentation.
    """

    validated_ledger = validate_evidence_ledger(evidence_ledger)
    by_id = {str(item.get("evidence_id")): item for item in validated_ledger}
    supplied = None if factual_claims is None else [_claim_text(value) for value in factual_claims if _claim_text(value)]
    narrative_text = "" if narrative is None else str(narrative)
    safe_validator = safe_non_factual_validator or abstention_validator
    sentences = _sentence_spans(narrative_text)
    records: list[dict[str, Any]] = []
    matched: set[str] = set()
    for index, span in enumerate(sentences):
        raw_start = span["start_char"]
        raw_end = span["end_char"]
        while raw_start < raw_end and narrative_text[raw_start].isspace(): raw_start += 1
        while raw_end > raw_start and narrative_text[raw_end - 1].isspace(): raw_end -= 1
        claim = narrative_text[raw_start:raw_end]
        normalized = _claim_text(claim)
        declared = supplied is not None and normalized in supplied
        ids = _CITATION_TOKEN_RE.findall(claim)
        safe_non_factual = False
        if safe_validator is not None:
            try: safe_non_factual = bool(safe_validator(claim))
            except Exception: safe_non_factual = False
        factual = bool(normalized) and (declared or bool(ids) or not safe_non_factual)
        abstention_authorized = factual_claims == [] and safe_non_factual and safe_validator is not None
        if factual:
            matched.add(normalized)
        ids = _CITATION_TOKEN_RE.findall(claim)
        unknown = sorted(set(ids) - set(by_id))
        records.append({
            "claim_index": index,
            "claim": claim,
            "claim_text": _CITATION_TOKEN_RE.sub("", claim).strip(),
            "start_char": raw_start,
            "end_char": raw_end,
            "is_factual": factual,
            "factual_claim": factual,
            "abstention_authorized": abstention_authorized,
            "evidence_ids": ids,
            "citation_ids": ids,
            "unknown_evidence_ids": unknown,
            "citation_format_valid": not unknown and len(ids) == len(set(ids)),
            "evidence_spans": [str(by_id[item].get("displayed_text", by_id[item].get("text", ""))) for item in ids if item in by_id],
        })
    if supplied is not None:
        for claim in factual_claims or []:
            normalized = _claim_text(claim)
            if normalized and normalized not in matched:
                records.append({
                    "claim_index": len(records), "claim": str(claim).strip(), "claim_text": str(claim).strip(),
                    "start_char": None, "end_char": None, "is_factual": True, "factual_claim": True,
                    "abstention_authorized": False,
                    "evidence_ids": [], "citation_ids": [], "unknown_evidence_ids": [],
                    "citation_format_valid": True, "evidence_spans": [],
                })
                matched.add(normalized)
    return records


def aggregate_citation_support(
    records: Sequence[Mapping[str, Any]],
    *,
    evaluator: Callable[[str, str], float] | None,
    require_nli: bool = False,
    threshold: float = 0.5,
) -> dict[str, Any]:
    """Aggregate claim-level support with joint and leave-one-out entailment.

    The callback is deliberately tiny and injectable: ``(premise, hypothesis)
    -> score``. Format compliance is reported independently from entailment;
    resolving ``[E#]`` never counts as factual support.
    """

    if not 0.0 <= float(threshold) <= 1.0:
        raise ValueError("threshold must be in [0, 1]")
    factual = [row for row in records if bool(row.get("is_factual", row.get("factual_claim", False)))]
    all_ids = [item for row in records for item in row.get("evidence_ids", [])]
    valid_ids = [item for row in records for item in row.get("evidence_ids", []) if item not in set(row.get("unknown_evidence_ids", []))]
    format_result = {
        "valid": all(bool(row.get("citation_format_valid", False)) for row in records),
        "citation_count": len(all_ids),
        "valid_citation_count": len(valid_ids),
        "unknown_evidence_ids": sorted({str(item) for row in records for item in row.get("unknown_evidence_ids", [])}),
        "duplicate_citation_count": sum(max(0, len(row.get("evidence_ids", [])) - len(set(row.get("evidence_ids", [])))) for row in records),
    }
    abstention_authorized = bool(records) and all(bool(row.get("abstention_authorized", False)) for row in records)
    abstention = abstention_authorized and not all_ids and not factual
    base = {
        "format_compliance": format_result,
        "syntactic_citation_precision": (len(valid_ids) / len(all_ids)) if all_ids else None,
        "factual_claim_count": len(factual), "supported_claim_count": 0,
        "claim_citation_recall": 0.0 if factual else None, "necessary_citation_count": 0,
        "citation_recall": 0.0 if factual else None,
        "citation_necessity": None,
        "overcitation_count": 0, "citation_precision": None, "claim_results": [],
        "abstention": abstention, "entailment_mean": None, "citation_entailment": None,
        "evaluated_claim_count": 0, "incomplete_claim_count": 0, "evaluated_claim_entailment_mean": None,
    }
    if not factual:
        base.update({"entailment_status": "not_required" if abstention else "invalid_abstention", "pass": bool(format_result["valid"] and abstention)})
        return base
    if evaluator is None:
        if require_nli:
            base.update({"entailment_status": "unavailable", "incomplete_claim_count": len(factual), "pass": False})
        else:
            base.update({"entailment_status": "not_requested", "pass": bool(format_result["valid"])})
        return base
    scores: list[float] = []
    for row in factual:
        ids = list(row.get("evidence_ids", []))
        spans = list(row.get("evidence_spans", []))
        claim = str(row.get("claim_text", row.get("claim", ""))).strip()
        if not ids or len(spans) != len(ids) or row.get("unknown_evidence_ids"):
            base["claim_results"].append({"claim": claim, "evidence_ids": ids, "joint_score": None, "supported": False, "necessary_evidence_ids": [], "leave_one_out": {}})
            base["incomplete_claim_count"] += 1
            continue
        try:
            def checked_score(premise: str) -> float:
                score = float(evaluator(premise, claim))
                if not math.isfinite(score) or not 0.0 <= score <= 1.0:
                    raise ValueError("evaluator score must be finite and in [0, 1]")
                return score
            joint = checked_score("\n\n".join(spans))
            individual = [checked_score(span) for span in spans]
            leave_one_out = {}
            necessary_ids = []
            for index, evidence_id in enumerate(ids):
                remaining = "\n\n".join(spans[:index] + spans[index + 1:])
                leave_score = checked_score(remaining) if remaining else 0.0
                leave_one_out[evidence_id] = leave_score
                if joint >= threshold and leave_score < threshold:
                    necessary_ids.append(evidence_id)
        except Exception:
            base["evaluated_claim_entailment_mean"] = sum(scores) / len(scores) if scores else None
            base["incomplete_claim_count"] = len(factual) - len(scores)
            base.update({"entailment_status": "evaluator_error", "pass": False})
            return base
        supported = joint >= threshold
        scores.append(joint)
        base["evaluated_claim_count"] += 1
        base["claim_results"].append({"claim": claim, "evidence_ids": ids, "joint_score": joint, "individual_scores": individual, "supported": supported, "necessary_evidence_ids": necessary_ids, "leave_one_out": leave_one_out})
        base["supported_claim_count"] += int(supported)
        base["necessary_citation_count"] += len(necessary_ids)
        if supported:
            base["overcitation_count"] += max(0, len(ids) - len(necessary_ids))
        else:
            # With no joint entailment, none of the cited passages supports the
            # claim; every cited passage is therefore unnecessary.
            base["overcitation_count"] += len(ids)
    base["claim_citation_recall"] = base["supported_claim_count"] / len(factual)
    base["citation_recall"] = base["claim_citation_recall"]
    base["citation_precision"] = base["necessary_citation_count"] / len(all_ids) if all_ids else 0.0
    base["citation_necessity"] = base["necessary_citation_count"] / len(valid_ids) if valid_ids else 0.0
    base["evaluated_claim_entailment_mean"] = sum(scores) / len(scores) if scores else None
    if base["incomplete_claim_count"]:
        base["entailment_status"] = "scored_incomplete"
        base["pass"] = False
    else:
        base["entailment_mean"] = base["evaluated_claim_entailment_mean"]
        base["citation_entailment"] = base["entailment_mean"]
        base["entailment_status"] = "scored"
        base["pass"] = bool(format_result["valid"] and base["supported_claim_count"] == len(factual) and base["overcitation_count"] == 0)
    return base


def normalize_optional_text(value: Any) -> str | None:
    """Return normalized non-empty text, or ``None`` for a missing value.

    Only actual missing values are removed.  The literal string ``"nan"`` is
    valid user text and therefore is not treated as a missing reference.
    """

    if value is None:
        return None
    if isinstance(value, numbers.Real):
        try:
            if math.isnan(value):
                return None
        except (TypeError, ValueError):
            pass
    is_nan = getattr(value, "is_nan", None)
    if callable(is_nan) and is_nan():
        return None
    # pandas.NA and similar scalar sentinels cannot be used in a boolean test.
    try:
        if value is not value:  # NaN-like objects whose identity is stable.
            return None
    except Exception:
        pass
    if type(value).__name__ in {"NAType", "NaTType"}:
        return None
    # numpy datetime64/timedelta64 NaT values are not Real and expose no
    # pandas-style sentinel API; their canonical string representation is
    # stable and can be checked without importing numpy.
    if type(value).__name__ in {"datetime64", "timedelta64"} and str(value) == "NaT":
        return None
    text = value.decode("utf-8", errors="replace") if isinstance(value, bytes) else str(value)
    text = unicodedata.normalize("NFKC", text).strip()
    return text or None


def fit_prompt_to_budget(
    prompt: Any,
    tokenizer: Any,
    max_input_tokens: int,
    *,
    schema_tail: Any = None,
    reserve_output_tokens: int = 0,
) -> str:
    """Preflight a rendered prompt without implicit tokenizer truncation.

    The helper never silently cuts prompt content. Callers must bound evidence
    explicitly before rendering; a schema tail, when supplied, is appended to
    the exact text being measured and therefore cannot disappear at the right
    edge of a tokenizer window.
    """

    if not isinstance(max_input_tokens, int) or isinstance(max_input_tokens, bool) or max_input_tokens < 1:
        raise ValueError("max_input_tokens must be a positive integer")
    if not isinstance(reserve_output_tokens, int) or isinstance(reserve_output_tokens, bool) or reserve_output_tokens < 0:
        raise ValueError("reserve_output_tokens must be a non-negative integer")
    if reserve_output_tokens >= max_input_tokens:
        raise ValueError("reserved output leaves no input budget")
    text = "" if prompt is None else str(prompt)
    tail = normalize_optional_text(schema_tail)
    if tail:
        text = text.rstrip() + "\n" + tail
    encoded = tokenizer(text=[text], return_tensors="pt", padding=False, truncation=False)
    mask = encoded.get("attention_mask") if isinstance(encoded, Mapping) else encoded["attention_mask"]
    if hasattr(mask, "sum") and not isinstance(mask, (list, tuple)):
        try:
            length = int(mask.sum().item())
        except AttributeError:
            length = int(mask.sum())
    else:
        length = sum(int(value) for value in mask[0])
    available = max_input_tokens - reserve_output_tokens
    if length > available:
        raise RuntimeError(f"prompt_token_budget_exceeded:{jsonable_identity({'prompt_tokens': length, 'input_budget': available, 'reserve_output_tokens': reserve_output_tokens, 'schema_tail_preserved': bool(tail)})}")
    return text


def fit_adaptive_prompt_with_evidence(
    evidence_ledger: Sequence[Mapping[str, Any]],
    *,
    prompt_builder: Callable[[Sequence[Mapping[str, Any]], str], str],
    payload_candidates: Sequence[str] = ("",),
    token_counter: Callable[[str], int],
    max_input_tokens: int,
    reserve_output_tokens: int = 0,
) -> dict[str, Any]:
    """Fit a prompt by dropping only whole ranked evidence spans.

    ``prompt_builder`` owns the immutable post/target/schema-tail text. The
    helper only varies the ranked evidence prefix and deterministic payload
    candidates, so it can never slice an evidence span or schema tail. A base
    prompt that cannot fit with no evidence and the smallest payload is
    returned as an explicit quarantine result instead of raising a whole-run
    overflow error.
    """

    if not isinstance(max_input_tokens, int) or isinstance(max_input_tokens, bool) or max_input_tokens < 1:
        raise ValueError("max_input_tokens must be a positive integer")
    if not isinstance(reserve_output_tokens, int) or isinstance(reserve_output_tokens, bool) or reserve_output_tokens < 0:
        raise ValueError("reserve_output_tokens must be a non-negative integer")
    if reserve_output_tokens >= max_input_tokens:
        raise ValueError("reserved output leaves no input budget")
    if not callable(prompt_builder) or not callable(token_counter):
        raise TypeError("prompt_builder and token_counter must be callable")
    ledger = validate_evidence_ledger(evidence_ledger)
    ranked = [
        item
        for _, item in sorted(
            enumerate(ledger),
            key=lambda pair: (
                float(pair[1].get("rank", pair[0])) if str(pair[1].get("rank", pair[0])).replace(".", "", 1).isdigit() else float(pair[0]),
                pair[0],
            ),
        )
    ]
    candidates = [str(payload) for payload in payload_candidates] or [""]
    available = max_input_tokens - reserve_output_tokens
    attempts: list[dict[str, Any]] = []
    for payload_index, payload in enumerate(candidates):
        for count in range(len(ranked), -1, -1):
            selected = ranked[:count]
            prompt = str(prompt_builder(selected, payload))
            prompt_tokens = int(token_counter(prompt))
            attempt = {"payload_index": payload_index, "selected_evidence_count": count, "prompt_tokens": prompt_tokens}
            attempts.append(attempt)
            if prompt_tokens <= available:
                selected_ids = [str(item["evidence_id"]) for item in selected]
                return {
                    "status": "fit",
                    "prompt": prompt,
                    "payload": payload,
                    "payload_index": payload_index,
                    "prompt_tokens": prompt_tokens,
                    "available_input_tokens": available,
                    "selected_evidence_ids": selected_ids,
                    "dropped_evidence_ids": [str(item["evidence_id"]) for item in ranked[count:]],
                    "attempts": attempts,
                }
    return {
        "status": "quarantine",
        "reason": "prompt_budget_irreducible",
        "prompt": None,
        "payload": candidates[-1],
        "payload_index": len(candidates) - 1,
        "prompt_tokens": attempts[-1]["prompt_tokens"] if attempts else None,
        "available_input_tokens": available,
        "selected_evidence_ids": [],
        "dropped_evidence_ids": [str(item["evidence_id"]) for item in ranked],
        "attempts": attempts,
    }


_SCRIPT_RANGES = {
    "tamil": ((0x0B80, 0x0BFF),),
    "devanagari": ((0x0900, 0x097F),),
    "bengali": ((0x0980, 0x09FF),),
    "telugu": ((0x0C00, 0x0C7F),),
    "kannada": ((0x0C80, 0x0CFF),),
    "malayalam": ((0x0D00, 0x0D7F),),
    "arabic": ((0x0600, 0x06FF), (0x0750, 0x077F), (0x08A0, 0x08FF)),
    "cyrillic": ((0x0400, 0x04FF),),
    "greek": ((0x0370, 0x03FF),),
    "hebrew": ((0x0590, 0x05FF),),
    "thai": ((0x0E00, 0x0E7F),),
    "hangul": ((0x1100, 0x11FF), (0xAC00, 0xD7AF)),
    "han": ((0x3400, 0x4DBF), (0x4E00, 0x9FFF)),
    "japanese": ((0x3040, 0x30FF),),
}


def _script_for_char(char: str) -> str | None:
    codepoint = ord(char)
    if (0x41 <= codepoint <= 0x5A) or (0x61 <= codepoint <= 0x7A) or (
        0x00C0 <= codepoint <= 0x024F
    ):
        return "latin"
    for script, ranges in _SCRIPT_RANGES.items():
        if any(start <= codepoint <= end for start, end in ranges):
            return script
    return None


def script_bucket(text: Any) -> str:
    """Classify the scripts represented by alphabetic Unicode characters.

    ``mixed`` is intentional when a response contains more than one script;
    punctuation, digits, and whitespace do not make a script mixed.
    """

    normalized = normalize_optional_text(text)
    if normalized is None:
        return "unknown"
    scripts = {_script_for_char(char) for char in normalized}
    scripts.discard(None)
    if not scripts:
        return "unknown"
    if len(scripts) == 1:
        return next(iter(scripts))
    return "mixed"


_LANGUAGE_LABELS = {
    "en": "en", "eng": "en", "english": "en",
    "ta": "ta", "tam": "ta", "tamil": "ta",
    "hi": "hi", "hin": "hi", "hindi": "hi",
    "fr": "fr", "fra": "fr", "fre": "fr", "french": "fr",
    "es": "es", "spa": "es", "spanish": "es",
    "de": "de", "deu": "de", "ger": "de", "german": "de",
    "pt": "pt", "por": "pt", "portuguese": "pt",
    "it": "it", "ita": "it", "italian": "it",
    "ru": "ru", "rus": "ru", "russian": "ru",
    "zh": "zh", "zho": "zh", "chi": "zh", "chinese": "zh",
    "ja": "ja", "jpn": "ja", "japanese": "ja",
    "ko": "ko", "kor": "ko", "korean": "ko",
}


def _language_label(value: Any) -> str | None:
    normalized = normalize_optional_text(value)
    if normalized is None:
        return None
    return _LANGUAGE_LABELS.get(normalized.casefold())


def language_match(expected: Any, observed: Any) -> bool | None:
    """Compare trusted language labels, returning ``None`` for unknown labels.

    This deliberately does not infer language from script: many languages
    share Latin or Han scripts. Use :func:`script_match` for a script-only
    diagnostic, never as a language claim.
    """

    expected_label = _language_label(expected)
    observed_label = _language_label(observed)
    if expected_label is None or observed_label is None:
        return None
    return expected_label == observed_label


def script_match(expected: Any, observed: Any) -> bool | None:
    """Compare detected scripts for diagnostic reporting only."""

    expected_script = script_bucket(expected)
    observed_script = script_bucket(observed)
    if "unknown" in {expected_script, observed_script}:
        return None
    return expected_script == observed_script and observed_script != "mixed"


def filter_audit(
    rows: Sequence[Any],
    keep: Callable[[Any], bool] | Sequence[bool],
    *,
    filter_name: str = "filter",
    reasons: Sequence[str | None] | None = None,
    id_key: str = "ID",
) -> dict[str, Any]:
    """Apply a filter while returning rows and an explicit audit manifest."""

    if callable(keep):
        decisions = [bool(keep(row)) for row in rows]
    else:
        decisions = [bool(value) for value in keep]
        if len(decisions) != len(rows):
            raise ValueError("filter decisions must align with rows")
    if reasons is None:
        reasons = [None if decision else filter_name for decision in decisions]
    else:
        reasons = list(reasons)
        if len(reasons) != len(rows):
            raise ValueError("filter reasons must align with rows")
    kept_rows = [row for row, decision in zip(rows, decisions) if decision]
    dropped_rows = [row for row, decision in zip(rows, decisions) if not decision]
    manifest_rows = []
    for index, (row, decision, reason) in enumerate(zip(rows, decisions, reasons)):
        row_id = row.get(id_key) if isinstance(row, Mapping) else None
        manifest_rows.append({"index": index, "record_id": None if row_id is None else str(row_id), "filter": filter_name, "kept": bool(decision), "reason": None if decision else (reason or filter_name)})
    events = [item for item in manifest_rows if not item["kept"]]
    reason_counts = Counter(item["reason"] for item in events)
    return {
        "filter": filter_name,
        "input": len(rows),
        "kept": len(kept_rows),
        "dropped": len(dropped_rows),
        "kept_rows": kept_rows,
        "dropped_rows": dropped_rows,
        "events": events,
        "rows": manifest_rows,
        "manifest_rows": manifest_rows,
        "reason_counts": dict(sorted(reason_counts.items())),
    }


def _clean_metric_text(value: Any) -> str | None:
    text = normalize_optional_text(value)
    return text.casefold() if text is not None else None


def unicode_chrf(reference: Any, hypothesis: Any, *, max_order: int = 6, beta: float = 2.0) -> float | None:
    """Compute a Unicode character n-gram F-score without an ASCII tokenizer."""

    if max_order < 1 or beta <= 0:
        raise ValueError("max_order must be positive and beta must be positive")
    ref = _clean_metric_text(reference)
    hyp = _clean_metric_text(hypothesis)
    if ref is None or hyp is None:
        return None
    # Whitespace is a boundary artifact, not a language character.  Keeping
    # all other code points makes this valid for Tamil, Devanagari, and emoji.
    ref = "".join(ref.split())
    hyp = "".join(hyp.split())
    if not ref or not hyp:
        return 0.0
    scores: list[float] = []
    beta_sq = beta * beta
    available_order = min(max_order, len(ref), len(hyp))
    for order in range(1, available_order + 1):
        ref_ngrams = Counter(ref[index : index + order] for index in range(len(ref) - order + 1))
        hyp_ngrams = Counter(hyp[index : index + order] for index in range(len(hyp) - order + 1))
        overlap = sum((ref_ngrams & hyp_ngrams).values())
        precision = overlap / sum(hyp_ngrams.values()) if hyp_ngrams else 0.0
        recall = overlap / sum(ref_ngrams.values()) if ref_ngrams else 0.0
        score = (1 + beta_sq) * precision * recall / (beta_sq * precision + recall) if precision and recall else 0.0
        scores.append(score)
    return sum(scores) / len(scores) if scores else 0.0


# Scientific-evaluation runtime contracts.  These helpers intentionally accept
# small injected callables so unit tests never need a model, a network, or a
# GPU.  Every metric is represented by ``{value, status, reason}``; a missing
# value is never converted to zero for an aggregate.
SUPPORTED_EVAL_LANGUAGES = frozenset({"en", "hi", "ta"})
ENGLISH_ONLY_METRICS = frozenset({"rouge_l", "meteor", "detoxify"})
BERTSCORE_MODEL_ID = "FacebookAI/xlm-roberta-large"
# A commit/revision, rather than ``main`` or an unpinned tag, is required for
# reproducible BERTScore.  Deployments may replace this with their audited
# revision through the explicit config argument.
BERTSCORE_MODEL_REVISION = "91703fe22dd9e5054634ccfb5b875f12f69158ec"
BERTSCORE_NUM_LAYERS = 24
NLI_MODEL_ID = "joeddav/xlm-roberta-large-xnli"
NLI_MODEL_REVISION = "07f8772bf0306314a97e4913cafde2cabf9814a9"
NLI_DATASET_PROVENANCE = {
    "en": {"dataset_id": "facebook/xnli", "dataset_revision": "072e4eb2b447bd887a772a7ab826ce0a7222b782", "config": "en", "split": "validation"},
    "hi": {"dataset_id": "mteb/IndicXnliPairClassification", "dataset_revision": "027e97b9afe84ea3447b57b7705b8864bb2b3a83", "config": "hi", "split": "test", "source_format": "parquet", "columns": {"sentence1": "premise", "sentence2": "hypothesis", "labels": "label"}},
    "ta": {"dataset_id": "mteb/IndicXnliPairClassification", "dataset_revision": "027e97b9afe84ea3447b57b7705b8864bb2b3a83", "config": "ta", "split": "test", "source_format": "parquet", "columns": {"sentence1": "premise", "sentence2": "hypothesis", "labels": "label"}},
}
DEFAULT_BERTSCORE_SETTINGS = {
    "model_id": BERTSCORE_MODEL_ID,
    "revision": BERTSCORE_MODEL_REVISION,
    "num_layers": BERTSCORE_NUM_LAYERS,
    "languages": {language: {"model_id": BERTSCORE_MODEL_ID, "revision": BERTSCORE_MODEL_REVISION} for language in sorted(SUPPORTED_EVAL_LANGUAGES)},
}
globals().update({"DEFAULT_BERTSCORE_CONFIG": DEFAULT_BERTSCORE_SETTINGS})


def load_nli_dataset_rows(dataset_loader: Callable[..., Any], provenance: Mapping[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Load a pinned NLI split and normalize its data-only schema.

    IndicXNLI is intentionally consumed as the ``mteb`` parquet export.  This
    rejects the old executable dataset-script repository before any network
    loader is called, which is required for ``datasets>=4`` compatibility.
    The returned provenance includes a content digest for the exact rows used
    by calibration; callers should persist it with the calibration artifact.
    """

    if not isinstance(provenance, Mapping):
        raise ValueError("nli provenance must be a mapping")
    dataset_id = str(provenance.get("dataset_id") or "")
    if dataset_id.casefold() == "divyanshu/indicxnli":
        raise ValueError("legacy NLI dataset script is not allowed; use data-only parquet")
    revision = str(provenance.get("dataset_revision") or "")
    if not re.fullmatch(r"[0-9a-fA-F]{40}", revision):
        raise ValueError("NLI dataset revision must be an immutable commit")
    config = str(provenance.get("config") or "")
    split = str(provenance.get("split") or "")
    if dataset_id == "mteb/IndicXnliPairClassification":
        if config not in {"hi", "ta"} or split != "test" or provenance.get("source_format") != "parquet":
            raise ValueError("IndicXNLI must use the pinned data-only parquet test split")
        columns = provenance.get("columns")
        expected_columns = {"sentence1": "premise", "sentence2": "hypothesis", "labels": "label"}
        if columns != expected_columns:
            raise ValueError("IndicXNLI parquet column mapping is invalid")
    dataset = dataset_loader(dataset_id, config, split=split, revision=revision)
    raw_rows = [dict(row) for row in dataset]
    required = ("sentence1", "sentence2", "labels") if dataset_id == "mteb/IndicXnliPairClassification" else ("premise", "hypothesis", "label")
    if any(not all(column in row for column in required) for row in raw_rows):
        raise ValueError("NLI dataset parquet columns are missing")
    rows = []
    for index, row in enumerate(raw_rows):
        if dataset_id == "mteb/IndicXnliPairClassification":
            premise, hypothesis, label = row["sentence1"], row["sentence2"], row["labels"]
        else:
            premise, hypothesis, label = row["premise"], row["hypothesis"], row["label"]
        if not isinstance(premise, str) or not isinstance(hypothesis, str):
            raise ValueError("NLI sentence columns must contain strings")
        try:
            label = int(label)
        except (TypeError, ValueError) as exc:
            raise ValueError("NLI labels must be integer class IDs") from exc
        if label not in {0, 1, 2}:
            raise ValueError("NLI labels must be in {0, 1, 2}")
        row_id = row.get("pairID", row.get("id", index))
        rows.append({"id": str(row_id), "premise": premise, "hypothesis": hypothesis, "label": label})
    if len({row["id"] for row in rows}) != len(rows):
        raise ValueError("NLI example IDs must be unique")
    digest = normalized_dataset_content_hash(rows)
    enriched = dict(provenance)
    enriched["dataset_content_hash"] = digest
    enriched["selected_content_hash"] = digest
    enriched["row_count"] = len(rows)
    enriched["schema_revision"] = "nli-pair-normalization.v1"
    return rows, enriched


def multilingual_bertscore_config(overrides: Mapping[str, Any] | None = None) -> dict[str, Any]:
    """Return a detached, pinned BERTScore config for en/hi/ta."""

    config = {
        "model_id": DEFAULT_BERTSCORE_SETTINGS["model_id"],
        "revision": DEFAULT_BERTSCORE_SETTINGS["revision"],
        "num_layers": DEFAULT_BERTSCORE_SETTINGS["num_layers"],
        "languages": {key: dict(value) for key, value in DEFAULT_BERTSCORE_SETTINGS["languages"].items()},
    }
    if overrides:
        for key in ("model_id", "revision"):
            if key in overrides:
                value = normalize_optional_text(overrides[key])
                if value is None or value.casefold() in {"main", "latest", "none"}:
                    raise ValueError(f"BERTScore {key} must be pinned")
                config[key] = value
        if "languages" in overrides:
            for language, value in overrides["languages"].items():
                label = _language_label(language)
                if label not in SUPPORTED_EVAL_LANGUAGES or not isinstance(value, Mapping):
                    raise ValueError("BERTScore language config is invalid")
                model_id = normalize_optional_text(value.get("model_id", config["model_id"]))
                revision = normalize_optional_text(value.get("revision", config["revision"]))
                if model_id is None or revision is None or revision.casefold() in {"main", "latest", "none"}:
                    raise ValueError("BERTScore language config must be pinned")
                config["languages"][label] = {"model_id": model_id, "revision": revision}
        if "num_layers" in overrides:
            if isinstance(overrides["num_layers"], bool) or not isinstance(overrides["num_layers"], int) or overrides["num_layers"] < 1:
                raise ValueError("BERTScore num_layers must be positive")
            config["num_layers"] = overrides["num_layers"]
    return config


def _metric_result(value: Any = None, status: str = "unavailable", reason: str | None = None) -> dict[str, Any]:
    if status == "scored":
        try:
            if value is None or not math.isfinite(float(value)):
                return {"value": None, "status": "invalid_score", "reason": "nonfinite_metric_value"}
            value = float(value)
        except (TypeError, ValueError):
            return {"value": None, "status": "invalid_score", "reason": "non_numeric_metric_value"}
    return {"value": value, "status": status, "reason": reason}


def _detector_language(detector: Any, text: Any) -> str | None:
    if detector is None or normalize_optional_text(text) is None:
        return None
    try:
        raw = detector(text) if callable(detector) else detector.detect_language_of(str(text))
    except Exception:
        return None
    # lingua returns an enum whose name is e.g. ``TAMIL``; langdetect returns
    # a two-letter code.  Normalize both without treating script as language.
    raw_name = getattr(raw, "name", raw)
    raw_name = str(raw_name).split(".")[-1].replace("LANGUAGE", "").strip(" _-")
    return _language_label(raw_name) or _language_label(raw)


def load_language_detector(detector_factory: Callable[[], Any] | None = None) -> Callable[[str], str | None] | None:
    """Load an offline lingua detector, with no network/download fallback.

    A caller-supplied factory is used in tests and production wrappers.  If
    lingua is absent, ``None`` is returned and the notebook records language
    metrics as unavailable rather than silently guessing from script.
    """

    if detector_factory is not None:
        detector = detector_factory()
        if detector is None:
            return None
        return lambda text: _detector_language(detector, text)
    try:
        from lingua import Language, LanguageDetectorBuilder
        detector = LanguageDetectorBuilder.from_languages(
            Language.ENGLISH, Language.HINDI, Language.TAMIL
        ).build()
        return lambda text: _detector_language(detector, text)
    except Exception:
        return None


def evaluate_multilingual_record(
    record: Mapping[str, Any],
    *,
    language_detector: Callable[[str], Any] | None = None,
    rouge_scorer: Callable[..., Any] | None = None,
    meteor_scorer: Callable[..., Any] | None = None,
    bertscore_scorer: Callable[..., Any] | None = None,
    detoxify_scorer: Callable[..., Any] | None = None,
    bertscore_config: Mapping[str, Any] | None = None,
    reference_key: str = "Counter Narrative",
    response_key: str = "parsed_counter_narrative",
) -> dict[str, Any]:
    """Score one record with explicit language/status provenance.

    English-only ROUGE/METEOR/Detoxify callbacks are never invoked for Hindi,
    Tamil, or unknown languages.  BERTScore and Unicode chrF are multilingual;
    BERTScore still requires a configured scorer and a pinned model revision.
    """

    if not isinstance(record, Mapping):
        raise TypeError("record must be a mapping")
    reference = normalize_optional_text(record.get(reference_key))
    hypothesis = normalize_optional_text(record.get(response_key, record.get("response")))
    expected = _language_label(record.get("language", record.get("Language")))
    input_language = _detector_language(language_detector, record.get("Text"))
    reference_language = _detector_language(language_detector, reference)
    output_language = _detector_language(language_detector, hypothesis)
    language = expected or input_language
    detected = [value for value in (input_language, reference_language, output_language) if value is not None]
    if language is None or language not in SUPPORTED_EVAL_LANGUAGES:
        language_status = "unsupported_language" if language else "unavailable"
    elif language_detector is None or input_language is None:
        language_status = "unavailable"
    elif any(value != language for value in detected):
        language_status = "mismatch"
    else:
        language_status = "scored"
    result: dict[str, Any] = {
        "ID": str(record.get("ID", "")),
        "language": language,
        "language_status": language_status,
        "expected_language": expected,
        "input_language": input_language,
        "reference_language": reference_language,
        "output_language": output_language,
        "language_match": language_match(language, output_language),
        "input_language_match": language_match(language, input_language),
        "reference_language_match": language_match(language, reference_language),
        "script_match": script_match(record.get("Text"), hypothesis),
        "reference_available": reference is not None,
        "metrics": {},
    }
    missing_status = "excluded_missing_reference" if reference is None else None
    result["metrics"]["chrf"] = _metric_result(
        unicode_chrf(reference, hypothesis), "scored" if reference is not None and hypothesis is not None else (missing_status or "missing_hypothesis"),
        None if reference is not None and hypothesis is not None else ("missing_reference" if reference is None else "missing_hypothesis"),
    )
    config = multilingual_bertscore_config(bertscore_config)
    if language_status == "mismatch":
        result["metrics"]["bertscore"] = _metric_result(status="language_mismatch", reason="detected_language_mismatch")
    elif language_status != "scored" or language not in SUPPORTED_EVAL_LANGUAGES:
        result["metrics"]["bertscore"] = _metric_result(status="unsupported_language", reason="language_not_supported")
    elif reference is None:
        result["metrics"]["bertscore"] = _metric_result(status="excluded_missing_reference", reason="missing_reference")
    elif hypothesis is None:
        result["metrics"]["bertscore"] = _metric_result(status="missing_hypothesis", reason="missing_hypothesis")
    elif bertscore_scorer is None:
        result["metrics"]["bertscore"] = _metric_result(status="unavailable", reason="bertscore_scorer_unavailable")
    else:
        try:
            payload = bertscore_scorer(reference=reference, hypothesis=hypothesis, language=language, model_id=config["languages"][language]["model_id"], revision=config["languages"][language]["revision"])
            value = payload.get("f1") if isinstance(payload, Mapping) else payload
            result["metrics"]["bertscore"] = _metric_result(value, "scored")
        except Exception as exc:
            result["metrics"]["bertscore"] = _metric_result(status="error", reason=f"bertscore_error:{type(exc).__name__}")
    for name, scorer in (("rouge_l", rouge_scorer), ("meteor", meteor_scorer), ("detoxify", detoxify_scorer)):
        if language_status == "mismatch":
            result["metrics"][name] = _metric_result(status="language_mismatch", reason="detected_language_mismatch")
            continue
        if language_status != "scored" or language != "en":
            result["metrics"][name] = _metric_result(status="unsupported_language", reason="english_only_metric")
            continue
        if name != "detoxify" and (reference is None or hypothesis is None):
            result["metrics"][name] = _metric_result(status=missing_status or "missing_hypothesis", reason="missing_reference" if reference is None else "missing_hypothesis")
            continue
        if scorer is None:
            result["metrics"][name] = _metric_result(status="unavailable", reason=f"{name}_scorer_unavailable")
            continue
        try:
            if name == "detoxify":
                value = scorer(hypothesis)
            else:
                value = scorer(reference, hypothesis)
            if isinstance(value, Mapping):
                value = value.get("rougeL", value.get("meteor", value.get("score", value.get("value"))))
            result["metrics"][name] = _metric_result(value, "scored")
        except Exception as exc:
            result["metrics"][name] = _metric_result(status="error", reason=f"{name}_error:{type(exc).__name__}")
    return result


def evaluate_multilingual_records(records: Sequence[Mapping[str, Any]], **kwargs: Any) -> list[dict[str, Any]]:
    return [evaluate_multilingual_record(row, **kwargs) for row in records]


def summarize_metric_records(records: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    """Return language/metric aggregates using only explicitly scored values."""

    grouped: dict[tuple[str, str], list[float]] = {}
    statuses: Counter[tuple[str, str, str]] = Counter()
    for row in records:
        language = row.get("language") or "unknown"
        for metric, item in (row.get("metrics") or {}).items():
            item = item if isinstance(item, Mapping) else {"value": item, "status": "scored"}
            status = str(item.get("status", "unavailable")); statuses[(str(language), metric, status)] += 1
            try:
                value = float(item.get("value"))
                if status == "scored" and math.isfinite(value): grouped.setdefault((str(language), metric), []).append(value)
            except (TypeError, ValueError):
                pass
    output = []
    keys = sorted(set(grouped) | {(language, metric) for language, metric, _ in statuses})
    for language, metric in keys:
        values = grouped.get((language, metric), [])
        status_counts = {status: count for (lang, met, status), count in statuses.items() if (lang, met) == (language, metric)}
        output.append({"language": language, "metric": metric, "n_scored": len(values), "mean": sum(values) / len(values) if values else None, "status_counts": dict(sorted(status_counts.items()))})
    return output


def summarize_language_records(records: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    """Summarize detector outputs and match rates by expected language."""
    grouped: dict[str, dict[str, Any]] = {}
    for row in records:
        language = str(row.get("language") or row.get("expected_language") or "unknown")
        item = grouped.setdefault(language, {"language": language, "records": 0, "input_language_counts": Counter(), "reference_language_counts": Counter(), "output_language_counts": Counter(), "input_matches": 0, "reference_matches": 0, "output_matches": 0, "input_status": Counter(), "reference_status": Counter(), "output_status": Counter()})
        item["records"] += 1
        for field, count_key, status_key, match_key in (("input_language", "input_language_counts", "input_status", "input_language_match"), ("reference_language", "reference_language_counts", "reference_status", "reference_language_match"), ("output_language", "output_language_counts", "output_status", "language_match")):
            detected = row.get(field) or "unknown"; item[count_key][str(detected)] += 1
            match = row.get(match_key)
            item["%s_matches" % field.split("_")[0] if field == "input_language" else ("reference_matches" if field == "reference_language" else "output_matches")] += int(match is True)
            item[status_key]["matched" if match is True else "mismatched" if match is False else "unavailable"] += 1
    output = []
    for language, item in sorted(grouped.items()):
        records_count = item["records"]
        output.append({"language": language, "records": records_count, "input_language_counts": dict(item["input_language_counts"]), "reference_language_counts": dict(item["reference_language_counts"]), "output_language_counts": dict(item["output_language_counts"]), "input_match_rate": item["input_matches"] / records_count if records_count else None, "reference_match_rate": item["reference_matches"] / records_count if records_count else None, "output_match_rate": item["output_matches"] / records_count if records_count else None, "input_status": dict(item["input_status"]), "reference_status": dict(item["reference_status"]), "output_status": dict(item["output_status"])})
    return output


def validate_metric_rows_unique(rows: Sequence[Mapping[str, Any]], *, id_key: str = "ID", variant_key: str = "variant") -> bool:
    seen: set[tuple[str, str]] = set()
    for row in rows:
        key = (str(row.get(id_key)), str(row.get(variant_key)))
        if key in seen:
            raise ValueError(f"duplicate metric row for ID/variant: {key[0]}/{key[1]}")
        seen.add(key)
    return True


def build_reference_metric_audit(
    rows: Sequence[Mapping[str, Any]], *, metrics: Sequence[str], language_key: str = "language"
) -> dict[str, Any]:
    """Report denominators and exclusions independently for every metric."""
    validate_metric_rows_unique(rows)
    output: dict[str, Any] = {}
    for metric in metrics:
        item: dict[str, Any] = {"input": len(rows), "scorable": 0, "excluded_by_reason": Counter(), "language_counts": Counter(), "coverage": None}
        for row in rows:
            language = str(row.get(language_key) or "unknown"); item["language_counts"][language] += 1
            status = str(row.get(f"{metric}_status") or "unavailable")
            value = row.get(metric)
            try: valid = status == "scored" and value is not None and math.isfinite(float(value))
            except (TypeError, ValueError): valid = False
            if valid: item["scorable"] += 1
            else:
                reason = str(row.get(f"{metric}_reason") or ("missing_reference" if row.get("reference_available") is False else status))
                item["excluded_by_reason"][reason] += 1
        item["coverage"] = item["scorable"] / item["input"] if item["input"] else None
        output[metric] = {"input": item["input"], "scorable": item["scorable"], "excluded_by_reason": dict(item["excluded_by_reason"]), "language_counts": dict(item["language_counts"]), "coverage": item["coverage"]}
    return output


def benjamini_hochberg(p_values: Sequence[float]) -> list[float]:
    """Return BH-adjusted q-values in the original order."""

    values = [float(value) for value in p_values]
    if any(not math.isfinite(value) or not 0.0 <= value <= 1.0 for value in values):
        raise ValueError("p-values must be finite numbers in [0, 1]")
    count = len(values)
    if not count:
        return []
    ranked = sorted(enumerate(values), key=lambda item: (item[1], item[0]))
    adjusted = [0.0] * count
    running = 1.0
    for rank in range(count - 1, -1, -1):
        original_index, p_value = ranked[rank]
        running = min(running, p_value * count / (rank + 1))
        adjusted[original_index] = min(1.0, max(0.0, running))
    return adjusted


def _paired_values(
    rows: Any,
    variant_a: str,
    variant_b: str,
    id_key: str,
) -> tuple[list[str], list[float], list[float], dict[str, int]]:
    if isinstance(rows, Mapping):
        rows = [dict(row, **{id_key: key}) if isinstance(row, Mapping) else {id_key: key, variant_a: row[0], variant_b: row[1]} for key, row in rows.items()]
    if not isinstance(rows, Sequence) or isinstance(rows, (str, bytes)):
        raise TypeError("rows must be a sequence of mappings")
    aligned: dict[str, tuple[float, float]] = {}
    seen_raw: set[tuple[str, str]] = set()
    seen_canonical: dict[str, tuple[str, str]] = {}
    exclusions: Counter[str] = Counter()
    for row in rows:
        if not isinstance(row, Mapping) or id_key not in row:
            raise ValueError(f"each row must contain {id_key!r}")
        raw_id = row[id_key]
        if _invalid_pair_id(raw_id):
            raise ValueError(f"invalid paired ID: {raw_id!r}")
        raw_token = (type(raw_id).__qualname__, repr(raw_id))
        identifier = str(raw_id)
        if raw_token in seen_raw:
            raise ValueError(f"duplicate paired ID: {identifier}")
        seen_raw.add(raw_token)
        prior_raw = seen_canonical.get(identifier)
        if prior_raw is not None and prior_raw != raw_token:
            raise ValueError(f"ID collision after canonicalization: {identifier}")
        seen_canonical[identifier] = raw_token
        left, right = row.get(variant_a), row.get(variant_b)
        left_status, right_status = _score_status(left), _score_status(right)
        if left_status == "nonfinite" or right_status == "nonfinite":
            exclusions["nonfinite"] += 1
            continue
        if left_status == "missing" or right_status == "missing":
            exclusions["incomplete"] += 1
            continue
        if left_status == "invalid" or right_status == "invalid":
            exclusions["invalid"] += 1
            continue
        left, right = float(left), float(right)
        aligned[identifier] = (left, right)
    identifiers = sorted(aligned)
    left = [aligned[key][0] for key in identifiers]
    right = [aligned[key][1] for key in identifiers]
    return identifiers, left, right, dict(exclusions)


def _invalid_pair_id(value: Any) -> bool:
    if value is None or type(value).__name__ in {"NAType", "NaTType"}:
        return True
    if isinstance(value, str):
        return not value.strip()
    if isinstance(value, numbers.Number):
        try:
            if isinstance(value, numbers.Complex) and not isinstance(value, numbers.Real):
                return not (math.isfinite(float(value.real)) and math.isfinite(float(value.imag)))
            return not math.isfinite(float(value))
        except (TypeError, ValueError):
            return False
    if type(value).__name__ in {"datetime64", "timedelta64"}:
        return str(value) == "NaT"
    return False


def _score_status(value: Any) -> str:
    if value is None or (isinstance(value, str) and not value.strip()):
        return "missing"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "invalid"
    return "valid" if math.isfinite(number) else "nonfinite"


def paired_variant_tests(
    rows: Any,
    *,
    variant_a: str = "mp_kg_rag",
    variant_b: str = "kg_rag",
    id_key: str = "ID",
    seed: int = 0,
    permutations: int = 10_000,
) -> dict[str, Any]:
    """Compare two variants on the exact shared ID set using sign permutations."""

    if permutations < 1:
        raise ValueError("permutations must be positive")
    identifiers, left, right, exclusion_counts = _paired_values(rows, variant_a, variant_b, id_key)
    differences = [a - b for a, b in zip(left, right)]
    n = len(differences)
    mean_difference = sum(differences) / n if n else None
    tolerance = 1e-12
    wins = sum(value > tolerance for value in differences)
    ties = sum(abs(value) <= tolerance for value in differences)
    losses = n - wins - ties
    result: dict[str, Any] = {
        "ids": identifiers,
        "n": n,
        "mean_difference": mean_difference,
        "win_rate": wins / n if n else None,
        "tie_rate": ties / n if n else None,
        "loss_rate": losses / n if n else None,
        "p_value": None,
        "seed": seed,
        "permutations": permutations,
        "excluded_rows": sum(exclusion_counts.values()),
        "exclusion_counts": exclusion_counts,
    }
    if not n:
        result["p_value"] = None
        result["status"] = "insufficient_pairs"
        return result
    if not any(abs(value) > tolerance for value in differences):
        result["p_value"] = 1.0
        return result
    observed = abs(mean_difference)
    if n <= 16:
        signs = range(1 << n)
        exceedances = 0
        total = 0
        for mask in signs:
            signed_mean = sum(value if mask & (1 << index) else -value for index, value in enumerate(differences)) / n
            exceedances += abs(signed_mean) >= observed - 1e-15
            total += 1
        result["p_value"] = exceedances / total
    else:
        rng = random.Random(seed)
        exceedances = 0
        for _ in range(permutations):
            signed_mean = sum(value if rng.getrandbits(1) else -value for value in differences) / n
            exceedances += abs(signed_mean) >= observed - 1e-15
        result["p_value"] = (exceedances + 1) / (permutations + 1)
    return result


DEFAULT_VARIANT_COMPARISONS = (
    ("mp_kg_rag", "kg_rag"),
    ("mp_kg_rag", "qwen_zero_shot"),
    ("kg_rag", "qwen_zero_shot"),
)


def pairwise_metric_family(
    rows: Sequence[Mapping[str, Any]],
    *,
    metrics: Sequence[str],
    comparisons: Sequence[tuple[str, str]] = DEFAULT_VARIANT_COMPARISONS,
    id_key: str = "ID",
    variant_key: str = "variant",
    seed: int = 0,
    permutations: int = 10_000,
    directions: Mapping[str, str] | None = None,
) -> dict[str, Any]:
    """Run all declared paired variant/metric tests and one BH family.

    Rows may be long (``ID``, ``variant``, metric columns) or wide (``ID`` and
    variant columns).  Missing/unsupported metric values are excluded by the
    tested paired helper, never filled with zero.  Every p-value in the full
    metric-by-comparison grid participates in one correction family.
    """

    if not metrics:
        raise ValueError("at least one metric is required")
    validate_metric_rows_unique(rows, id_key=id_key, variant_key=variant_key)
    long_rows = [row for row in rows if isinstance(row, Mapping) and variant_key in row]
    outputs: list[dict[str, Any]] = []
    p_locations: list[tuple[int, str]] = []
    for left_variant, right_variant in comparisons:
        comparison_result = {"comparison": f"{left_variant}_vs_{right_variant}", "variant_a": left_variant, "variant_b": right_variant, "metrics": {}}
        for metric in metrics:
            wide: dict[str, dict[str, Any]] = {}
            if long_rows:
                for row in long_rows:
                    identifier = row.get(id_key)
                    if identifier is None or row.get(variant_key) not in {left_variant, right_variant}:
                        continue
                    value = row.get(metric)
                    status = row.get(f"{metric}_status")
                    # A metric status is authoritative when present.  In
                    # particular, non-RAG variants must remain null for
                    # citation support metrics instead of becoming artificial
                    # zero scores in a paired test.
                    if status is not None and str(status) != "scored":
                        value = None
                    wide.setdefault(str(identifier), {id_key: identifier})[str(row[variant_key])] = value
            else:
                wide = {}
                for row in rows:
                    if not isinstance(row, Mapping) or id_key not in row:
                        continue
                    detached = dict(row)
                    status = detached.get(f"{metric}_status")
                    if status is not None and str(status) != "scored":
                        for variant in (left_variant, right_variant):
                            if variant in detached:
                                detached[variant] = None
                    wide[str(row[id_key])] = detached
            paired_rows = [value for value in wide.values()]
            direction = (directions or {}).get(metric, "higher")
            if direction not in {"higher", "lower"}:
                raise ValueError(f"invalid metric direction: {metric}")
            if direction == "lower":
                paired_rows = [dict(row, **{left_variant: row.get(right_variant), right_variant: row.get(left_variant)}) for row in paired_rows]
            result = paired_variant_tests(paired_rows, variant_a=left_variant, variant_b=right_variant, id_key=id_key, seed=seed, permutations=permutations)
            result["direction"] = direction
            result["metric"] = metric
            comparison_result["metrics"][metric] = result
            if result.get("p_value") is not None:
                p_locations.append((len(outputs), metric))
        outputs.append(comparison_result)
    q_values = benjamini_hochberg([outputs[index]["metrics"][metric]["p_value"] for index, metric in p_locations]) if p_locations else []
    for q_value, (index, metric) in zip(q_values, p_locations):
        outputs[index]["metrics"][metric]["q_value"] = q_value
        outputs[index]["metrics"][metric]["bh_family"] = "declared_metric_x_comparison"
    # Make the correction auditable even when a test has no paired rows.
    for comparison in outputs:
        for metric, result in comparison["metrics"].items():
            result.setdefault("q_value", None)
            result.setdefault("bh_family", "declared_metric_x_comparison")
    return {"comparisons": outputs, "metrics": list(metrics), "declared_comparisons": [f"{a}_vs_{b}" for a, b in comparisons], "bh_family": "declared_metric_x_comparison", "bh_family_size": len(p_locations), "seed": seed, "permutations": permutations}


def bootstrap_mean_ci(
    values: Sequence[Any],
    *,
    confidence: float = 0.95,
    n_resamples: int = 10_000,
    seed: int = 0,
) -> dict[str, Any]:
    """Return a seeded percentile bootstrap CI, excluding missing values."""

    if not 0.0 < confidence < 1.0:
        raise ValueError("confidence must be in (0, 1)")
    if n_resamples < 1:
        raise ValueError("n_resamples must be positive")
    clean: list[float] = []
    for value in values:
        try:
            number = float(value)
        except (TypeError, ValueError):
            continue
        if math.isfinite(number):
            clean.append(number)
    if not clean:
        return {
            "n": 0,
            "mean": None,
            "lower": None,
            "upper": None,
            "low": None,
            "high": None,
            "confidence": confidence,
            "seed": seed,
        }
    mean = sum(clean) / len(clean)
    rng = random.Random(seed)
    samples = []
    for _ in range(n_resamples):
        samples.append(sum(clean[rng.randrange(len(clean))] for _ in clean) / len(clean))
    samples.sort()
    alpha = (1.0 - confidence) / 2.0
    lower = samples[min(len(samples) - 1, max(0, int(math.floor(alpha * len(samples)))))]
    upper_index = min(len(samples) - 1, max(0, int(math.ceil((1.0 - alpha) * len(samples))) - 1))
    upper = samples[upper_index]
    return {
        "n": len(clean),
        "mean": mean,
        "lower": lower,
        "upper": upper,
        "low": lower,
        "high": upper,
        "confidence": confidence,
        "seed": seed,
    }


def _rater_matrix(
    rows: Any,
    rater_columns: Sequence[str] | None = None,
    *,
    exact_two: bool = False,
) -> list[list[Any]]:
    if isinstance(rows, Mapping):
        if not rows:
            raise ValueError("at least two raters are required")
        columns = list(rater_columns or rows.keys())
        if len(columns) < 2:
            raise ValueError("at least two raters are required")
        if exact_two and len(columns) != 2:
            raise ValueError("exactly two raters are required")
        values = [rows[column] for column in columns]
        if not all(isinstance(value, Sequence) and not isinstance(value, (str, bytes)) for value in values):
            raise TypeError("rater mapping values must be sequences")
        if len({len(value) for value in values}) != 1:
            raise ValueError("rater columns must have equal lengths")
        return [list(row) for row in zip(*values)]
    if not isinstance(rows, Sequence) or isinstance(rows, (str, bytes)):
        raise TypeError("rows must be a sequence")
    if rater_columns is not None:
        columns = list(rater_columns)
        if len(columns) < 2:
            raise ValueError("at least two raters are required")
        matrix = []
        for row in rows:
            if not isinstance(row, Mapping):
                raise TypeError("rater_columns require mapping rows")
            missing = [column for column in columns if column not in row]
            if missing:
                raise ValueError(f"missing rater column: {missing[0]}")
            matrix.append([row[column] for column in columns])
    else:
        matrix = [list(row.values()) if isinstance(row, Mapping) else list(row) for row in rows]
    if matrix and len(matrix[0]) < 2:
        raise ValueError("at least two raters are required")
    if exact_two and matrix and len(matrix[0]) != 2:
        raise ValueError("exactly two raters are required")
    if matrix and any(len(row) != len(matrix[0]) for row in matrix):
        raise ValueError("each row must contain the same number of raters")
    if not matrix:
        raise ValueError("at least two raters are required")
    return matrix


def weighted_kappa_rows(rows: Any, *, rater_columns: Sequence[str] | None = None, weights: str = "quadratic") -> float:
    """Compute weighted Cohen kappa for exactly two rater columns."""

    matrix = _rater_matrix(rows, rater_columns, exact_two=True)
    pairs = [(row[0], row[1]) for row in matrix if row[0] is not None and row[1] is not None]
    if not pairs:
        return float("nan")
    if weights not in {"linear", "quadratic"}:
        raise ValueError("weights must be 'linear' or 'quadratic'")
    categories = sorted({value for pair in pairs for value in pair})
    if len(categories) == 1:
        return 1.0
    positions = {category: index for index, category in enumerate(categories)}
    size = len(categories) - 1
    def agreement(a: Any, b: Any) -> float:
        distance = abs(positions[a] - positions[b]) / size
        return 1.0 - (distance if weights == "linear" else distance * distance)
    total = len(pairs)
    observed = sum(agreement(a, b) for a, b in pairs) / total
    left_counts = Counter(a for a, _ in pairs)
    right_counts = Counter(b for _, b in pairs)
    expected = sum(
        (left_counts[a] / total) * (right_counts[b] / total) * agreement(a, b)
        for a in categories for b in categories
    )
    return 1.0 if math.isclose(expected, 1.0) else (observed - expected) / (1.0 - expected)


def krippendorff_alpha_ordinal(rows: Any, *, rater_columns: Sequence[str] | None = None) -> float:
    """Compute ordinal Krippendorff alpha using coincidence matrices.

    Each unit is normalized by its own valid-rater count. This is important
    when missing annotations leave different units with different numbers of
    ratings; averaging raw pair distances is not equivalent.
    """

    matrix = _rater_matrix(rows, rater_columns)
    units = [[value for value in row if value is not None] for row in matrix]
    units = [unit for unit in units if len(unit) >= 2]
    if not units:
        return float("nan")
    categories = sorted({value for unit in units for value in unit})
    positions = {category: index for index, category in enumerate(categories)}
    denominator = max(1, len(categories) - 1)

    def distance(a: Any, b: Any) -> float:
        delta = (positions[a] - positions[b]) / denominator
        return delta * delta

    # coincidence[c][k] is the expected number of ordered coincidences for
    # category c and k within units, with each unit divided by n_u - 1.
    coincidence: Counter[tuple[Any, Any]] = Counter()
    for unit in units:
        counts = Counter(unit)
        unit_size = len(unit)
        for left in categories:
            for right in categories:
                if left == right:
                    count = counts[left] * max(0, counts[left] - 1)
                else:
                    count = counts[left] * counts[right]
                coincidence[(left, right)] += count / (unit_size - 1)
    total_coincidences = sum(coincidence.values())
    if total_coincidences <= 0:
        return float("nan")

    observed = sum(value * distance(left, right) for (left, right), value in coincidence.items()) / total_coincidences
    marginal = Counter()
    for (left, _), value in coincidence.items():
        marginal[left] += value
    expected = 0.0
    for left in categories:
        for right in categories:
            if left == right:
                pair_mass = marginal[left] * max(0.0, marginal[left] - 1.0)
            else:
                pair_mass = marginal[left] * marginal[right]
            expected += pair_mass * distance(left, right)
    expected /= total_coincidences * max(1.0, total_coincidences - 1.0)
    if math.isclose(expected, 0.0):
        return 1.0 if math.isclose(observed, 0.0) else 0.0
    return 1.0 - observed / expected


def sample_annotation_records(
    rows: Sequence[Mapping[str, Any]],
    *,
    max_ids: int,
    seed: int = 0,
    id_key: str = "ID",
    variant_key: str = "variant",
    stratify_key: str = "stratify_key",
) -> dict[str, Any]:
    """Sample IDs by stratum while retaining every variant for each ID."""

    if isinstance(max_ids, bool) or not isinstance(max_ids, int) or max_ids < 1:
        raise ValueError("max_ids must be a positive integer")
    groups: dict[str, list[tuple[str, list[Mapping[str, Any]]]]] = {}
    by_id: dict[str, list[Mapping[str, Any]]] = {}
    strata: dict[str, str] = {}
    raw_ids: dict[str, Any] = {}
    for row in rows:
        if not isinstance(row, Mapping) or id_key not in row:
            raise ValueError("annotation row must contain ID")
        identifier = str(row[id_key])
        if not identifier:
            raise ValueError("annotation ID must be non-empty")
        by_id.setdefault(identifier, []).append(row)
        raw_ids.setdefault(identifier, row[id_key])
        strata.setdefault(identifier, str(row.get(stratify_key, "unknown")))
    for identifier, values in by_id.items():
        groups.setdefault(strata[identifier], []).append((identifier, values))
    rng = random.Random(seed)
    for values in groups.values():
        values.sort(key=lambda item: item[0]); rng.shuffle(values)
    all_ids = sorted(by_id)
    selected: list[str] = []
    # Round-robin over shuffled strata gives every non-empty language/script
    # stratum a chance before filling the remainder from the global pool.
    cursors = {key: 0 for key in groups}
    while len(selected) < min(max_ids, len(all_ids)):
        progressed = False
        for stratum in sorted(groups):
            values = groups[stratum]
            cursor = cursors[stratum]
            if cursor < len(values) and values[cursor][0] not in selected:
                selected.append(values[cursor][0]); cursors[stratum] += 1; progressed = True
                if len(selected) >= max_ids: break
        if not progressed: break
    selected_set = set(selected)
    sampled_rows = [row for row in rows if str(row[id_key]) in selected_set]
    return {
        "rows": sampled_rows,
        "selected_ids": selected,
        "selected_id_count": len(selected),
        "input_id_count": len(all_ids),
        "max_ids": max_ids,
        "seed": seed,
        "stratify_key": stratify_key,
        "stratum_counts": {key: sum(1 for identifier in selected if strata[identifier] == key) for key in sorted(groups)},
        "variants_per_id": {identifier: sorted({str(row.get(variant_key)) for row in by_id[identifier]}) for identifier in selected},
    }


def human_agreement_report(
    rows: Sequence[Mapping[str, Any]],
    *,
    id_key: str = "ID",
    rater_key: str = "rater",
    score_key: str = "score",
) -> dict[str, Any]:
    """Compute overlap/missingness, pairwise weighted kappa, and ordinal alpha."""

    if not rows:
        raise ValueError("at least two distinct raters are required")
    raters = sorted({str(row.get(rater_key, row.get("rater_id"))) for row in rows if row.get(rater_key, row.get("rater_id")) is not None})
    if len(raters) < 2:
        raise ValueError("at least two distinct raters are required")
    matrix: dict[str, dict[str, Any]] = {}
    for row in rows:
        if id_key not in row:
            raise ValueError("annotation row must contain ID")
        rater = row.get(rater_key, row.get("rater_id")); identifier = str(row[id_key])
        if rater is None:
            raise ValueError("annotation row must contain rater")
        rater = str(rater)
        if rater in matrix.setdefault(identifier, {}):
            raise ValueError(f"duplicate annotation for {identifier}/{rater}")
        matrix[identifier][rater] = row.get(score_key, row.get("rating"))
    ordered_rows = [[matrix[identifier].get(rater) for rater in raters] for identifier in sorted(matrix)]
    overlap_count = sum(sum(value is not None for value in row) >= 2 for row in ordered_rows)
    expected_cells = len(ordered_rows) * len(raters)
    observed_cells = sum(value is not None for row in ordered_rows for value in row)
    pairwise: dict[str, float] = {}
    for left_index, left in enumerate(raters):
        for right in raters[left_index + 1:]:
            key = f"{left}_vs_{right}"
            pairs = [{"left": row.get(left), "right": row.get(right)} for row in matrix.values()]
            value = weighted_kappa_rows(pairs, rater_columns=("left", "right"))
            pairwise[key] = None if isinstance(value, float) and math.isnan(value) else value
    alpha_value = krippendorff_alpha_ordinal(ordered_rows)
    alpha = None if isinstance(alpha_value, float) and math.isnan(alpha_value) else alpha_value
    return {
        "distinct_raters": len(raters), "raters": raters, "unit_count": len(ordered_rows),
        "overlap_count": overlap_count, "missingness_count": expected_cells - observed_cells,
        "observed_cells": observed_cells, "expected_cells": expected_cells,
        "weighted_kappa": pairwise if len(pairwise) != 1 else next(iter(pairwise.values())),
        "weighted_kappa_pairwise": pairwise, "krippendorff_alpha_ordinal": alpha,
        "status": "scored" if overlap_count else "insufficient_overlap",
    }


def build_nli_calibration_artifact(
    *, model_id: str, model_revision: str, dataset_id: str, dataset_revision: str,
    language: str, split: str, label_mapping: Mapping[str, Any], threshold: float,
    n_examples: int, example_ids: Sequence[Any], example_content_digest: str,
    dataset_content_hash: str | None = None,
    accuracy: float, per_label_stats: Mapping[str, Any], code_hash: str, eval_core_hash: str,
    calibration_metadata: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Create an authenticated held-out calibration record.

    The digest covers every provenance and quality field.  Verification must be
    rerun at startup; callers must not treat a hand-edited JSON file as a gate.
    """
    label = _language_label(language)
    required_text = (model_id, model_revision, dataset_id, dataset_revision, label, split, example_content_digest, code_hash, eval_core_hash)
    if any(normalize_optional_text(value) is None for value in required_text) or not isinstance(label_mapping, Mapping) or not label_mapping:
        raise ValueError("incomplete calibration provenance")
    if not isinstance(n_examples, int) or n_examples < 1 or len(example_ids) != n_examples:
        raise ValueError("calibration example count mismatch")
    if not 0.0 <= float(threshold) <= 1.0 or not 0.0 <= float(accuracy) <= 1.0:
        raise ValueError("calibration scores must be in [0, 1]")
    if any(str(key).startswith("LABEL_") for key in label_mapping):
        if not validate_nli_label_mapping(label_mapping, kind="model")["valid"]: raise ValueError("model label mapping mismatch")
    elif all(str(key) in {"0", "1", "2"} for key in label_mapping):
        if not validate_nli_label_mapping(label_mapping, kind="dataset")["valid"]: raise ValueError("dataset label mapping mismatch")
    payload = {
        "schema": "mpkg-nli-calibration.v2", "model_id": str(model_id), "model_revision": str(model_revision),
        "dataset_id": str(dataset_id), "dataset_revision": str(dataset_revision), "dataset_content_hash": str(dataset_content_hash or ""), "language": label, "split": str(split),
        "label_mapping": jsonable_identity(label_mapping), "threshold": float(threshold), "n_examples": n_examples,
        "example_ids": [str(value) for value in example_ids], "example_content_digest": str(example_content_digest),
        "accuracy": float(accuracy), "per_label_stats": jsonable_identity(per_label_stats), "code_hash": str(code_hash), "eval_core_hash": str(eval_core_hash),
    }
    if calibration_metadata is not None:
        payload["calibration"] = jsonable_identity(calibration_metadata)
    payload["artifact_digest"] = stable_identity_hash(payload)
    return payload


def verify_nli_calibration_artifact(
    artifact: Mapping[str, Any] | None, *, model_id: str, model_revision: str,
    dataset_id: str, dataset_revision: str, language: str, split: str | None = None, dataset_content_hash: str | None = None,
    dataset_examples: Sequence[Mapping[str, Any]] | None = None,
    label_mapping: Mapping[str, Any] | None = None,
    code_hash: str | None = None, eval_core_hash: str | None = None,
    fresh_predictor: Callable[[Mapping[str, Any]], Mapping[str, Any]] | None = None,
    calibration_seed: int | None = None, calibration_min_support: int | None = None,
    calibration_bootstrap: int | None = None, calibration_criterion: str | None = None,
    calibration_entailment_label: int | None = None,
) -> dict[str, Any]:
    """Authenticate calibration provenance and reject stale/forged records."""
    label = _language_label(language)
    if not isinstance(artifact, Mapping) or artifact.get("language") != label:
        return {"enabled": False, "status": "unverified", "reason": "language_mismatch"}
    expected = {"model_id": model_id, "model_revision": model_revision, "dataset_id": dataset_id, "dataset_revision": dataset_revision}
    if split is not None: expected["split"] = split
    if dataset_content_hash is not None: expected["dataset_content_hash"] = dataset_content_hash
    if dataset_examples is not None and dataset_content_hash != normalized_dataset_content_hash(dataset_examples):
        return {"enabled": False, "status": "unverified", "reason": "dataset_content_digest_mismatch"}
    if dataset_examples is not None:
        try:
            metadata_reason = _validate_calibration_metadata(artifact, dataset_examples)
        except (TypeError, ValueError, KeyError):
            metadata_reason = "calibration_metadata_invalid"
        if metadata_reason:
            return {"enabled": False, "status": "unverified", "reason": metadata_reason}
        expected_ids = [str(row.get("id")) for row in dataset_examples]
        observed_ids = [str(value) for value in artifact.get("example_ids", [])]
        if observed_ids != expected_ids:
            return {"enabled": False, "status": "unverified", "reason": "calibration_example_ids_mismatch"}
        expected_example_digest = stable_identity_hash([{key: row.get(key) for key in ("id", "premise", "hypothesis", "label")} for row in dataset_examples])
        if artifact.get("example_content_digest") != expected_example_digest:
            return {"enabled": False, "status": "unverified", "reason": "calibration_example_digest_mismatch"}
    if label_mapping is not None and artifact.get("label_mapping") != jsonable_identity(label_mapping):
        return {"enabled": False, "status": "unverified", "reason": "label_mapping_mismatch"}
    if fresh_predictor is not None:
        if dataset_examples is None or None in (calibration_seed, calibration_min_support, calibration_bootstrap, calibration_criterion, calibration_entailment_label):
            return {"enabled": False, "status": "unverified", "reason": "fresh_calibration_configuration_missing"}
        try:
            fresh = calibrate_nli_threshold(dataset_examples, fresh_predictor, seed=int(calibration_seed), min_support=int(calibration_min_support), n_bootstrap=int(calibration_bootstrap), criterion=str(calibration_criterion), entailment_label=int(calibration_entailment_label))
            stored = artifact.get("calibration")
            if not isinstance(stored, Mapping):
                return {"enabled": False, "status": "unverified", "reason": "fresh_calibration_metadata_missing"}
            if abs(float(artifact.get("threshold")) - float(fresh["threshold"])) > 1e-12 or int(stored.get("seed")) != int(fresh["seed"]) or stored.get("criterion") != fresh["criterion"] or int(stored.get("entailment_label")) != int(fresh["entailment_label"]):
                return {"enabled": False, "status": "unverified", "reason": "fresh_calibration_threshold_or_config_mismatch"}
            if sorted(map(str, stored.get("calibration_ids", []))) != sorted(map(str, fresh["calibration_ids"])) or sorted(map(str, stored.get("audit_ids", []))) != sorted(map(str, fresh["audit_ids"])) or stored.get("support") != fresh["support"]:
                return {"enabled": False, "status": "unverified", "reason": "fresh_calibration_split_mismatch"}
            fresh_audit = fresh["metrics"]["audit"]; stored_audit = stored.get("metrics", {}).get("audit", {})
            for key in ("accuracy",):
                if abs(float(stored_audit.get(key)) - float(fresh_audit[key])) > 1e-12 or abs(float(artifact.get(key)) - float(fresh_audit[key])) > 1e-12:
                    return {"enabled": False, "status": "unverified", "reason": "fresh_calibration_accuracy_mismatch"}
            for key in ("accuracy_ci",):
                if stored_audit.get(key) != fresh_audit.get(key):
                    return {"enabled": False, "status": "unverified", "reason": "fresh_calibration_accuracy_ci_mismatch"}
            fresh_entailment = fresh_audit["entailment"]; stored_entailment = stored_audit.get("entailment", {})
            for key in ("precision", "recall", "f1", "precision_ci", "recall_ci", "f1_ci"):
                if stored_entailment.get(key) != fresh_entailment.get(key):
                    return {"enabled": False, "status": "unverified", "reason": f"fresh_calibration_{key}_mismatch"}
        except (TypeError, ValueError, KeyError, AttributeError):
            return {"enabled": False, "status": "unverified", "reason": "fresh_calibration_recompute_failed"}
    if code_hash is not None: expected["code_hash"] = code_hash
    if eval_core_hash is not None: expected["eval_core_hash"] = eval_core_hash
    if any(artifact.get(key) != value for key, value in expected.items()):
        return {"enabled": False, "status": "unverified", "reason": "calibration_provenance_mismatch"}
    digest_payload = {key: artifact.get(key) for key in artifact if key != "artifact_digest"}
    if artifact.get("artifact_digest") != stable_identity_hash(digest_payload):
        return {"enabled": False, "status": "unverified", "reason": "calibration_digest_mismatch"}
    try:
        if int(artifact.get("n_examples")) < 1 or not 0.0 <= float(artifact.get("threshold")) <= 1.0 or not normalize_optional_text(artifact.get("example_content_digest")):
            raise ValueError
        if len(artifact.get("example_ids", [])) != int(artifact["n_examples"]):
            raise ValueError
    except (TypeError, ValueError, KeyError):
        return {"enabled": False, "status": "unverified", "reason": "calibration_quality_invalid"}
    return {"enabled": True, "status": "calibrated", "dataset_id": artifact["dataset_id"], "dataset_revision": artifact["dataset_revision"], "language": label, "threshold": float(artifact["threshold"]), "n_examples": int(artifact["n_examples"]), "artifact_digest": artifact["artifact_digest"]}


def nli_calibration_quality(artifact: Mapping[str, Any], *, min_accuracy: float, min_entailment_precision: float, min_entailment_recall: float, min_per_label_support: int, min_accuracy_lower: float | None = None, min_entailment_precision_lower: float | None = None, min_entailment_recall_lower: float | None = None) -> dict[str, Any]:
    """Apply the configured held-out quality gate after provenance auth."""
    try:
        accuracy = float(artifact.get("accuracy")); stats = artifact.get("per_label_stats")
        entailment = stats.get("entailment") if isinstance(stats, Mapping) else None
        precision = float(entailment.get("precision")); recall = float(entailment.get("recall")); support = int(entailment.get("n_total"))
        positive_support = int(entailment.get("positive_support"))
        metadata = artifact.get("calibration"); audit_metrics = metadata.get("metrics", {}).get("audit", {}) if isinstance(metadata, Mapping) else {}
        accuracy_lower = float(audit_metrics.get("accuracy_ci", {}).get("lower")); precision_lower = float(audit_metrics.get("entailment", {}).get("precision_ci", {}).get("lower")); recall_lower = float(audit_metrics.get("entailment", {}).get("recall_ci", {}).get("lower"))
    except (TypeError, ValueError, AttributeError):
        return {"enabled": False, "status": "quality_unavailable", "reason": "missing_quality_statistics"}
    numeric_support = {str(key): int(value.get("n_total")) for key, value in stats.items() if str(key).isdigit() and isinstance(value, Mapping) and value.get("n_total") is not None}
    lower_bounds = {"accuracy": accuracy_lower, "entailment_precision": precision_lower, "entailment_recall": recall_lower}
    lower_thresholds = {"accuracy": min_accuracy if min_accuracy_lower is None else min_accuracy_lower, "entailment_precision": min_entailment_precision if min_entailment_precision_lower is None else min_entailment_precision_lower, "entailment_recall": min_entailment_recall if min_entailment_recall_lower is None else min_entailment_recall_lower}
    point_failed = accuracy < min_accuracy or precision < min_entailment_precision or recall < min_entailment_recall or support < min_per_label_support or positive_support < min_per_label_support or any(value < min_per_label_support for value in numeric_support.values())
    ci_failed = any(lower_bounds[key] < lower_thresholds[key] for key in lower_bounds)
    if point_failed or ci_failed:
        return {"enabled": False, "status": "quality_failed", "reason": "calibration_quality_below_threshold", "accuracy": accuracy, "entailment_precision": precision, "entailment_recall": recall, "entailment_support": support, "positive_support": positive_support, "per_label_support": numeric_support, "lower_bounds": lower_bounds, "lower_thresholds": lower_thresholds}
    return {"enabled": True, "status": "quality_passed", "accuracy": accuracy, "entailment_precision": precision, "entailment_recall": recall, "entailment_support": support, "positive_support": positive_support, "per_label_support": numeric_support, "lower_bounds": lower_bounds, "lower_thresholds": lower_thresholds}


MODEL_LABEL_MAPPING = {"LABEL_0": "contradiction", "LABEL_1": "neutral", "LABEL_2": "entailment"}
DATASET_LABEL_MAPPING = {"0": "entailment", "1": "neutral", "2": "contradiction"}


def validate_nli_label_mapping(mapping: Mapping[str, Any], *, kind: str) -> dict[str, Any]:
    expected = MODEL_LABEL_MAPPING if kind == "model" else DATASET_LABEL_MAPPING if kind == "dataset" else None
    if expected is None or not isinstance(mapping, Mapping):
        return {"valid": False, "reason": "unknown_label_mapping_kind"}
    normalized = {str(key): str(value).casefold() for key, value in mapping.items()}
    valid = normalized == {key: value for key, value in expected.items()}
    return {"valid": valid, "reason": None if valid else "label_mapping_mismatch", "expected": dict(expected), "observed": normalized}


def normalized_dataset_content_hash(rows: Sequence[Mapping[str, Any]]) -> str:
    normalized = []
    for row in rows:
        normalized.append({str(key): unicodedata.normalize("NFKC", str(value)) if isinstance(value, str) else value for key, value in sorted(row.items(), key=lambda item: str(item[0]))})
    payload = __import__("json").dumps(normalized, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _validate_calibration_metadata(artifact: Mapping[str, Any], dataset_examples: Sequence[Mapping[str, Any]]) -> str | None:
    metadata = artifact.get("calibration")
    required = {"method", "seed", "criterion", "entailment_label", "calibration_ids", "audit_ids", "calibration_content_digest", "audit_content_digest", "support", "metrics", "dataset_content_hash"}
    if not isinstance(metadata, Mapping) or not required.issubset(metadata):
        return "calibration_metadata_missing_or_incomplete"
    expected_ids = [str(row.get("id")) for row in dataset_examples]
    if len(expected_ids) != len(set(expected_ids)):
        return "dataset_example_ids_not_unique"
    if metadata.get("method") != "stratified_calibration_audit_v1" or metadata.get("criterion") not in {"f1", "precision_recall"} or not isinstance(metadata.get("seed"), int):
        return "calibration_method_metadata_invalid"
    mapping = artifact.get("label_mapping")
    try:
        expected_entailment_label = int(next(key for key, value in mapping.items() if str(value).casefold() == "entailment"))
    except (AttributeError, StopIteration, TypeError, ValueError):
        return "calibration_label_mapping_invalid"
    if int(metadata.get("entailment_label")) != expected_entailment_label:
        return "calibration_entailment_label_mismatch"
    calibration_ids = metadata.get("calibration_ids"); audit_ids = metadata.get("audit_ids")
    if not isinstance(calibration_ids, list) or not isinstance(audit_ids, list) or calibration_ids != sorted(set(map(str, calibration_ids))) or audit_ids != sorted(set(map(str, audit_ids))):
        return "calibration_split_ids_invalid"
    calibration_ids = list(map(str, calibration_ids)); audit_ids = list(map(str, audit_ids))
    if set(calibration_ids) & set(audit_ids) or set(calibration_ids) | set(audit_ids) != set(expected_ids):
        return "calibration_split_coverage_invalid"
    if metadata.get("dataset_content_hash") != artifact.get("dataset_content_hash"):
        return "calibration_dataset_digest_mismatch"
    by_id = {str(row.get("id")): row for row in dataset_examples}
    calibration_rows = [by_id[identifier] for identifier in expected_ids if identifier in set(calibration_ids)]
    audit_rows = [by_id[identifier] for identifier in expected_ids if identifier in set(audit_ids)]
    if metadata.get("calibration_content_digest") != normalized_dataset_content_hash(calibration_rows) or metadata.get("audit_content_digest") != normalized_dataset_content_hash(audit_rows):
        return "calibration_split_content_digest_mismatch"
    support = metadata.get("support")
    if not isinstance(support, Mapping) or int(support.get("calibration", -1)) != len(calibration_rows) or int(support.get("audit", -1)) != len(audit_rows):
        return "calibration_support_mismatch"
    per_label = support.get("per_label")
    if not isinstance(per_label, Mapping):
        return "calibration_per_label_support_missing"
    for raw_label in sorted({str(row.get("label")) for row in dataset_examples}):
        expected_calibration = sum(str(row.get("label")) == raw_label for row in calibration_rows)
        expected_audit = sum(str(row.get("label")) == raw_label for row in audit_rows)
        observed = per_label.get(raw_label)
        if not isinstance(observed, Mapping) or int(observed.get("calibration", -1)) != expected_calibration or int(observed.get("audit", -1)) != expected_audit:
            return "calibration_per_label_support_mismatch"
    artifact_stats = artifact.get("per_label_stats")
    if not isinstance(artifact_stats, Mapping):
        return "calibration_per_label_metrics_missing"
    for raw_label in sorted({str(row.get("label")) for row in dataset_examples}):
        observed = artifact_stats.get(raw_label)
        if not isinstance(observed, Mapping) or int(observed.get("n_total", -1)) != sum(str(row.get("label")) == raw_label for row in audit_rows):
            return "calibration_artifact_label_support_mismatch"
    metrics = metadata.get("metrics"); audit_metrics = metrics.get("audit") if isinstance(metrics, Mapping) else None
    try:
        if float(audit_metrics.get("accuracy")) != float(artifact.get("accuracy")):
            return "calibration_accuracy_mismatch"
        if not isinstance(audit_metrics.get("accuracy_ci"), Mapping):
            return "calibration_accuracy_ci_missing"
        entailment = audit_metrics.get("entailment")
        for key in ("precision", "recall", "f1", "precision_ci", "recall_ci", "f1_ci"):
            if key not in entailment:
                return "calibration_entailment_metric_missing"
        stats = artifact.get("per_label_stats")
        if not isinstance(stats, Mapping) or not isinstance(stats.get("entailment"), Mapping):
            return "calibration_per_label_metrics_missing"
        stored_entailment = stats["entailment"]
        for key in ("precision", "recall"):
            if float(stored_entailment.get(key)) != float(entailment.get(key)):
                return "calibration_entailment_metric_mismatch"
        if float(stored_entailment.get("f1")) != float(entailment.get("f1")):
            return "calibration_entailment_metric_mismatch"
        total = sum(int(value.get("n_total")) for key, value in stats.items() if str(key).isdigit())
        correct = sum(int(value.get("correct")) for key, value in stats.items() if str(key).isdigit())
        if total < 1 or abs(correct / total - float(artifact.get("accuracy"))) > 1e-12:
            return "calibration_accuracy_not_reproducible"
        for key in ("precision", "recall", "f1"):
            value = float(entailment.get(key))
            if not 0.0 <= value <= 1.0:
                return "calibration_metric_out_of_range"
        for key in ("accuracy_ci", "precision_ci", "recall_ci", "f1_ci"):
            interval = audit_metrics.get(key) if key == "accuracy_ci" else entailment.get(key)
            if not isinstance(interval, Mapping) or not 0.0 <= float(interval.get("lower")) <= float(interval.get("upper")) <= 1.0:
                return "calibration_ci_invalid"
    except (TypeError, ValueError, AttributeError):
        return "calibration_metrics_invalid"
    return None


def _binary_metrics(gold: Sequence[int], scores: Sequence[float], threshold: float, *, entailment_label: int = 2) -> dict[str, Any]:
    truth = [int(value == entailment_label) for value in gold]; pred = [int(score >= threshold) for score in scores]
    tp = sum(a == b == 1 for a, b in zip(truth, pred)); fp = sum(a == 0 and b == 1 for a, b in zip(truth, pred)); fn = sum(a == 1 and b == 0 for a, b in zip(truth, pred))
    precision = tp / (tp + fp) if tp + fp else 0.0; recall = tp / (tp + fn) if tp + fn else 0.0
    return {"precision": precision, "recall": recall, "f1": 2 * precision * recall / (precision + recall) if precision + recall else 0.0, "support": sum(truth)}


def _bootstrap_binary_metric_ci(gold: Sequence[int], scores: Sequence[float], threshold: float, *, entailment_label: int, metric: str = "f1", n_resamples: int, seed: int) -> dict[str, float | None]:
    if not gold or n_resamples < 1:
        return {"lower": None, "upper": None}
    rng = random.Random(seed); values: list[float] = []; indices = list(range(len(gold)))
    for _ in range(n_resamples):
        sampled = [rng.choice(indices) for _ in indices]
        metric_result = _binary_metrics([gold[index] for index in sampled], [scores[index] for index in sampled], threshold, entailment_label=entailment_label)
        values.append(float(metric_result[metric]))
    values.sort()
    lower = values[min(len(values) - 1, int(0.025 * len(values)))]
    upper = values[min(len(values) - 1, int(0.975 * len(values)))]
    return {"lower": lower, "upper": upper}


def select_stratified_nli_rows(rows: Sequence[Mapping[str, Any]], *, max_rows: int, seed: int = 0) -> list[Mapping[str, Any]]:
    """Select a deterministic, label-stratified NLI sample for calibration.

    Rows are sorted by stable content before seeded shuffling, so dataset
    iteration order cannot silently change the calibration set.  The
    round-robin fill keeps every observed gold label represented.
    """
    if isinstance(max_rows, bool) or not isinstance(max_rows, int) or max_rows < 1:
        raise ValueError("max_rows must be positive")
    groups: dict[int, list[Mapping[str, Any]]] = {}
    for row in rows:
        if "label" not in row:
            raise ValueError("NLI row missing label")
        groups.setdefault(int(row["label"]), []).append(row)
    rng = random.Random(seed)
    for label, values in groups.items():
        values.sort(key=lambda row: stable_identity_hash({key: row.get(key) for key in ("id", "premise", "hypothesis", "label")}))
        rng.shuffle(values)
    target = min(max_rows, sum(len(values) for values in groups.values()))
    selected: list[Mapping[str, Any]] = []
    cursor = {label: 0 for label in sorted(groups)}
    while len(selected) < target:
        progressed = False
        for label in sorted(groups):
            if cursor[label] < len(groups[label]):
                selected.append(groups[label][cursor[label]])
                cursor[label] += 1
                progressed = True
                if len(selected) >= target:
                    break
        if not progressed:
            break
    return selected


def calibrate_nli_threshold(rows: Sequence[Mapping[str, Any]], predictor: Callable[[Mapping[str, Any]], Mapping[str, Any]], *, seed: int = 0, min_support: int = 50, n_bootstrap: int = 2000, criterion: str = "f1", entailment_label: int = 2) -> dict[str, Any]:
    """Deterministically stratify, tune on calibration, and audit untouched data."""
    if criterion not in {"f1", "precision_recall"} or min_support < 1:
        raise ValueError("invalid calibration configuration")
    by_label: dict[int, list[Mapping[str, Any]]] = {}
    for row in rows:
        label = int(row["label"]); by_label.setdefault(label, []).append(row)
    rng = random.Random(seed); calibration: list[Mapping[str, Any]] = []; audit: list[Mapping[str, Any]] = []
    for label in sorted(by_label):
        values = sorted(by_label[label], key=lambda row: str(row.get("id"))); rng.shuffle(values)
        if len(values) < 2 * min_support:
            raise ValueError(f"insufficient_label_support:{label}")
        cut = len(values) // 2; calibration.extend(values[:cut]); audit.extend(values[cut:])
    def predicted(values):
        output = [predictor(row) for row in values]
        return [int(item["label"]) for item in output], [float(item["entailment_probability"]) for item in output]
    cal_pred_labels, cal_scores = predicted(calibration); audit_pred_labels, audit_scores = predicted(audit)
    cal_gold = [int(row["label"]) for row in calibration]; audit_gold = [int(row["label"]) for row in audit]
    candidates = sorted({0.05 * index for index in range(1, 20)} | {round(score, 6) for score in cal_scores if 0.0 <= score <= 1.0})
    def objective(value: float) -> float:
        metrics = _binary_metrics(cal_gold, cal_scores, value, entailment_label=entailment_label)
        return metrics["f1"] if criterion == "f1" else metrics["precision"] * metrics["recall"]
    threshold = max(candidates, key=lambda value: (objective(value), -value))
    binary = _binary_metrics(audit_gold, audit_scores, threshold, entailment_label=entailment_label)
    accuracy = sum(gold == prediction for gold, prediction in zip(audit_gold, audit_pred_labels)) / len(audit_gold)
    truth = [int(label == entailment_label) for label in audit_gold]
    pred = [int(score >= threshold) for score in audit_scores]
    accuracy_ci = bootstrap_mean_ci([float(gold == prediction) for gold, prediction in zip(audit_gold, audit_pred_labels)], n_resamples=n_bootstrap, seed=seed)
    return {"method": "stratified_calibration_audit_v1", "seed": seed, "threshold": threshold, "criterion": criterion, "entailment_label": entailment_label, "calibration_ids": {str(row["id"]) for row in calibration}, "audit_ids": {str(row["id"]) for row in audit}, "support": {"calibration": len(calibration), "audit": len(audit), "per_label": {str(label): {"calibration": sum(int(row["label"]) == label for row in calibration), "audit": sum(int(row["label"]) == label for row in audit)} for label in by_label}}, "metrics": {"audit": {"accuracy": accuracy, "accuracy_ci": accuracy_ci, "entailment": {**binary, "precision_ci": _bootstrap_binary_metric_ci(audit_gold, audit_scores, threshold, entailment_label=entailment_label, metric="precision", n_resamples=n_bootstrap, seed=seed + 1), "recall_ci": _bootstrap_binary_metric_ci(audit_gold, audit_scores, threshold, entailment_label=entailment_label, metric="recall", n_resamples=n_bootstrap, seed=seed + 2), "f1_ci": _bootstrap_binary_metric_ci(audit_gold, audit_scores, threshold, entailment_label=entailment_label, metric="f1", n_resamples=n_bootstrap, seed=seed + 3)}}}}


def validate_nli_calibration_artifact(artifact: Mapping[str, Any] | None, language: str) -> dict[str, Any]:
    """Require an authenticated per-language calibration artifact."""
    label = _language_label(language)
    entry = artifact.get(label) if isinstance(artifact, Mapping) and label else None
    if not isinstance(entry, Mapping):
        return {"enabled": False, "status": "uncalibrated", "reason": "missing_language_calibration"}
    required = {key: entry.get(key) for key in ("model_id", "model_revision", "dataset_id", "dataset_revision", "language")}
    return verify_nli_calibration_artifact(entry, model_id=required["model_id"], model_revision=required["model_revision"], dataset_id=required["dataset_id"], dataset_revision=required["dataset_revision"], language=label, split=entry.get("split"), code_hash=entry.get("code_hash"), eval_core_hash=entry.get("eval_core_hash"))


def load_nli_runtime_evaluator(
    language: str,
    *,
    calibration_artifact: Mapping[str, Any] | None,
    model_loader: Callable[..., Callable[[str, str], float]] | None = None,
    model_id: str = NLI_MODEL_ID,
    revision: str = NLI_MODEL_REVISION,
    calibration_provenance: Mapping[str, Mapping[str, Any]] | None = None,
) -> dict[str, Any]:
    """Load multilingual NLI only after a per-language calibration gate."""

    label = _language_label(language)
    expected = calibration_provenance.get(label) if isinstance(calibration_provenance, Mapping) else None
    if expected is not None:
        entry = calibration_artifact.get(label) if isinstance(calibration_artifact, Mapping) else None
        if not isinstance(entry, Mapping):
            calibration = {"enabled": False, "status": "unverified", "reason": "missing_language_calibration"}
        else:
            calibration = verify_nli_calibration_artifact(entry, model_id=str(expected.get("model_id")), model_revision=str(expected.get("model_revision")), dataset_id=str(expected.get("dataset_id")), dataset_revision=str(expected.get("dataset_revision")), dataset_content_hash=expected.get("dataset_content_hash"), language=label, split=expected.get("split"), code_hash=expected.get("code_hash"), eval_core_hash=expected.get("eval_core_hash"), label_mapping=expected.get("dataset_label_mapping"))
    else:
        calibration = validate_nli_calibration_artifact(calibration_artifact, language)
    if not calibration["enabled"]:
        return {"evaluator": None, "model_id": model_id, "revision": revision, "language": _language_label(language), **calibration}
    if model_loader is None:
        return {"evaluator": None, "model_id": model_id, "revision": revision, "language": _language_label(language), "enabled": False, "status": "unavailable", "reason": "nli_model_loader_unavailable", **{key: value for key, value in calibration.items() if key != "enabled"}}
    try:
        evaluator = model_loader(model_id=model_id, revision=revision, language=_language_label(language), threshold=calibration["threshold"])
    except Exception as exc:
        return {"evaluator": None, "model_id": model_id, "revision": revision, "language": _language_label(language), "enabled": False, "status": "error", "reason": f"nli_model_load_error:{type(exc).__name__}"}
    if not callable(evaluator):
        return {"evaluator": None, "model_id": model_id, "revision": revision, "language": _language_label(language), "enabled": False, "status": "error", "reason": "nli_model_loader_invalid"}
    return {"evaluator": evaluator, "model_id": model_id, "revision": revision, "language": _language_label(language), **calibration}

EVAL_CORE_SOURCE_SHA256 = 'f8055e11a959bcd414f3e0878b918e778176889822e0f87bdf3fece93e2c0cf3'



# %% [notebook cell 4]

# 02 - Imports, reproducibility, run layout, and memory diagnostics.
import os, re, gc, json, time, math, random, hashlib, unicodedata, tempfile, platform, sys, importlib.metadata, fcntl, atexit
from pathlib import Path
from packaging.version import Version
import numpy as np, pandas as pd, torch
from tqdm.auto import tqdm

LOCKFILE_SHA256 = _LOCKFILE_SHA256_EXPECTED
LOCK_PACKAGE_COUNTS = {"emitted_application_package_count": 115, "managed_package_count": 23, "resolved_package_count": 138}
DEFAULT_GENERATOR_MODEL_REVISION = "3764fa359b9082ea5a1e4a5e3ac3aaf6e9671636"
DEFAULT_EMBEDDING_MODEL_REVISION = "5617a9f61b028005a4858fdac845db406aefb181"
DEFAULT_RERANKER_MODEL_REVISION = "953dc6f6f85a1b2dbfca4c34a2796e7dde08d41e"
def require_model_revision(revision, model_role):
    value = str(revision or "").strip()
    if re.fullmatch(r"[0-9a-fA-F]{40}", value) is None:
        raise RuntimeError(f"immutable_model_revision_unavailable:{model_role}")
    return value.lower()
GENERATOR_MODEL_REVISION = require_model_revision(os.environ.get("MODEL_REVISION", DEFAULT_GENERATOR_MODEL_REVISION), "generator")
EMBEDDING_MODEL_REVISION = require_model_revision(os.environ.get("EMBEDDING_MODEL_REVISION", DEFAULT_EMBEDDING_MODEL_REVISION), "embedding")
RERANKER_MODEL_REVISION = require_model_revision(os.environ.get("RERANKER_MODEL_REVISION", DEFAULT_RERANKER_MODEL_REVISION), "reranker")
QWEN35_MIN_TRANSFORMERS_VERSION = "5.2.0"
def validate_qwen35_transformers_compatibility():
    observed = importlib.metadata.version("transformers")
    if Version(observed) < Version(QWEN35_MIN_TRANSFORMERS_VERSION):
        raise RuntimeError(f"qwen35_transformers_incompatible:observed={observed}:minimum={QWEN35_MIN_TRANSFORMERS_VERSION}")
    return {"observed": observed, "minimum": QWEN35_MIN_TRANSFORMERS_VERSION, "status": "compatible"}
QWEN35_TRANSFORMERS_COMPATIBILITY = validate_qwen35_transformers_compatibility()
SEED = 3407
BOOTSTRAP_SEED = SEED
os.environ.setdefault("PYTHONHASHSEED", str(SEED))
random.seed(SEED); np.random.seed(SEED); torch.manual_seed(SEED)
if torch.cuda.is_available(): torch.cuda.manual_seed_all(SEED)
assert torch.cuda.is_available(), "An NVIDIA GPU is required."
MANAGED_ACCELERATOR_CONTRACT = {
    "torch": "2.11.0+cu128",
    "torchvision": "0.26.0+cu128",
    "triton": "3.6.0",
    "xformers": "0.0.35",
    "torchao": "0.17.0+cu128",
    "nvidia-cuda-runtime-cu12": "12.8.90",
    "nvidia-nvjitlink-cu12": "12.8.93",
    "nvidia-nvtx-cu12": "12.8.90",
    "torch_cuda": "12.8",
}
MANAGED_ACCELERATOR_CONTRACT_HASH = stable_id("managed-accelerator-contract.v1", json.dumps(MANAGED_ACCELERATOR_CONTRACT, sort_keys=True))
def validate_managed_accelerator():
    observed = {}
    for package_name, expected_version in MANAGED_ACCELERATOR_CONTRACT.items():
        if package_name == "torch_cuda": continue
        try:
            observed[package_name] = importlib.metadata.version(package_name)
        except importlib.metadata.PackageNotFoundError as error:
            raise RuntimeError(f"managed_accelerator_distribution_missing:{package_name}") from error
        if observed[package_name] != expected_version:
            raise RuntimeError(f"managed_accelerator_distribution_mismatch:{package_name}:{observed[package_name]}:{expected_version}")
    torch_cuda = str(getattr(torch.version, "cuda", "") or "")
    if torch_cuda != MANAGED_ACCELERATOR_CONTRACT["torch_cuda"]:
        raise RuntimeError(f"managed_torch_cuda_mismatch:torch_cuda={torch_cuda}:expected={MANAGED_ACCELERATOR_CONTRACT['torch_cuda']}")
    try:
        import torchvision, triton, xformers, torchao, bitsandbytes
        if not callable(getattr(bitsandbytes, "matmul_4bit", None)) or getattr(getattr(bitsandbytes, "cextension", None), "lib", None) is None:
            raise RuntimeError("bitsandbytes_kernel_unavailable")
        import xformers.ops
        import triton.language as tl
    except Exception as error:
        raise RuntimeError(f"managed_accelerator_import_failed:{type(error).__name__}:{error}") from error
    try:
        probe = torch.ones((2, 2), device="cuda", dtype=torch.float32)
        result = probe @ probe
        torch.cuda.synchronize()
        if not torch.allclose(result, torch.full((2, 2), 2.0, device="cuda")):
            raise RuntimeError("matmul_result_mismatch")
        bnb_input = torch.ones((1, 2), device="cuda", dtype=torch.float16)
        bnb_weight = torch.ones((2, 2), device="cuda", dtype=torch.float16)
        bnb_quantized, bnb_state = bitsandbytes.functional.quantize_4bit(bnb_weight, quant_type="nf4")
        bnb_result = bitsandbytes.matmul_4bit(bnb_input, bnb_quantized.t(), bnb_state)
        if tuple(bnb_result.shape) != (1, 2) or not bool(torch.isfinite(bnb_result).all()):
            raise RuntimeError("bitsandbytes_smoke_failed")
        xformers_query = torch.randn((1, 4, 1, 8), device="cuda", dtype=torch.float16)
        xformers_key = torch.randn((1, 4, 1, 8), device="cuda", dtype=torch.float16)
        xformers_value = torch.randn((1, 4, 1, 8), device="cuda", dtype=torch.float16)
        xformers_result = xformers.ops.memory_efficient_attention(xformers_query, xformers_key, xformers_value)
        torch.cuda.synchronize()
        if tuple(xformers_result.shape) != tuple(xformers_query.shape) or not bool(torch.isfinite(xformers_result).all()):
            raise RuntimeError("xformers_smoke_failed")
        @triton.jit
        def _managed_contract_add_kernel(input_ptr, output_ptr, n_elements, BLOCK_SIZE: tl.constexpr):
            pid = tl.program_id(0)
            offsets = pid * BLOCK_SIZE + tl.arange(0, BLOCK_SIZE)
            mask = offsets < n_elements
            values = tl.load(input_ptr + offsets, mask=mask)
            tl.store(output_ptr + offsets, values + 1.0, mask=mask)
        triton_input = torch.ones(128, device="cuda", dtype=torch.float32)
        triton_output = torch.empty_like(triton_input)
        _managed_contract_add_kernel[lambda meta: (triton.cdiv(triton_input.numel(), meta["BLOCK_SIZE"]),)](triton_input, triton_output, triton_input.numel(), BLOCK_SIZE=128)
        torch.cuda.synchronize()
        if not bool(torch.allclose(triton_output, torch.full_like(triton_output, 2.0))):
            raise RuntimeError("triton_kernel_smoke_failed")
        del bnb_input, bnb_weight, bnb_quantized, bnb_state, bnb_result, probe, result, xformers_query, xformers_key, xformers_value, xformers_result, triton_input, triton_output
        torch.cuda.empty_cache()
    except Exception as error:
        raise RuntimeError(f"managed_accelerator_smoke_failed:{type(error).__name__}:{error}") from error
    return {"contract": MANAGED_ACCELERATOR_CONTRACT, "contract_hash": MANAGED_ACCELERATOR_CONTRACT_HASH, "observed_distributions": observed, "torch_cuda": torch_cuda, "bitsandbytes_kernel": "matmul_4bit", "xformers_kernel": "memory_efficient_attention", "triton_kernel": "managed_contract_add_kernel", "cuda_smoke": "passed"}
MANAGED_ACCELERATOR = validate_managed_accelerator()
PROJECT_ROOT = Path(os.environ.get("PROJECT_ROOT", "/workspace/mp_kg_rag")).expanduser().resolve()
SHARD_INDEX = int(os.environ.get("SHARD_INDEX", "0")); SHARD_COUNT = int(os.environ.get("SHARD_COUNT", "1"))
MAX_EXPERIMENT_ROWS = int(os.environ.get("MAX_EXPERIMENT_ROWS", "1550")); CACHE_BASE_CAPACITY = int(os.environ.get("CACHE_MAX_RECORDS", "256")); EFFECTIVE_CACHE_CAPACITY = derive_effective_cache_capacity(CACHE_BASE_CAPACITY, row_limit=MAX_EXPERIMENT_ROWS, shard_count=SHARD_COUNT)
BASE_RUN_NAME = os.environ.get("RUN_NAME", "qwen35_mp_kg_rag_v2_semantic")
RUN_NAME = derive_shard_run_name(BASE_RUN_NAME, SHARD_INDEX, SHARD_COUNT)
RUN = PROJECT_ROOT / "runs" / RUN_NAME
RUN_LOCK_HANDLE = None
def acquire_run_lock():
    global RUN_LOCK_HANDLE
    RUN.mkdir(parents=True, exist_ok=True)
    RUN_LOCK_HANDLE = (RUN / ".run.lock").open("a+", encoding="utf-8")
    try:
        fcntl.flock(RUN_LOCK_HANDLE.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
    except BlockingIOError:
        RUN_LOCK_HANDLE.close(); RUN_LOCK_HANDLE = None
        raise RuntimeError("run_already_active")
def release_run_lock():
    global RUN_LOCK_HANDLE
    if RUN_LOCK_HANDLE is None: return
    try: fcntl.flock(RUN_LOCK_HANDLE.fileno(), fcntl.LOCK_UN)
    finally: RUN_LOCK_HANDLE.close(); RUN_LOCK_HANDLE = None
acquire_run_lock(); atexit.register(release_run_lock)
for folder in [RUN / x for x in ["artifacts", "checkpoints", "exports", "logs", "review_queue"]]: folder.mkdir(parents=True, exist_ok=True)
def now(): return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
def write_json(path, value): Path(path).write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")
def write_json_atomic(path, value):
    path = Path(path); path.parent.mkdir(parents=True, exist_ok=True); temporary = None
    try:
        with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=path.parent, prefix=f".{path.name}.", suffix=".tmp", delete=False) as stream:
            temporary = Path(stream.name); json.dump(value, stream, ensure_ascii=False, indent=2); stream.flush(); os.fsync(stream.fileno())
        os.replace(temporary, path)
    except Exception:
        if temporary is not None: temporary.unlink(missing_ok=True)
        raise
def _locked_append_line(path, line):
    path = Path(path); path.parent.mkdir(parents=True, exist_ok=True); lock_path = path.with_name(path.name + ".lock")
    with lock_path.open("a+", encoding="utf-8") as lock:
        fcntl.flock(lock.fileno(), fcntl.LOCK_EX)
        try:
            with path.open("a", encoding="utf-8") as stream: stream.write(line); stream.flush(); os.fsync(stream.fileno())
        finally: fcntl.flock(lock.fileno(), fcntl.LOCK_UN)
def append_jsonl(path, row):
    _locked_append_line(path, json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")
def load_jsonl(path):
    path = Path(path); return [json.loads(x) for x in path.read_text(encoding="utf-8").splitlines() if x.strip()] if path.exists() else []
def log_event(stage, event, **data):
    append_jsonl(RUN / "logs" / "events.jsonl", {"at": now(), "stage": stage, "event": event, "run_name": RUN_NAME, "config_hash": globals().get("CONFIG_HASH"), "run_identity_hash": globals().get("RUN_IDENTITY_HASH"), "shard": {"index": SHARD_INDEX, "count": SHARD_COUNT}, "manifest_hashes": {key: globals().get(key) for key in ["CORPUS_MANIFEST_HASH", "CHUNK_MANIFEST_HASH", "GRAPH_MANIFEST_HASH", "SPLIT_MEMBERSHIP_HASH"]}, **data})
def _resolved_identity(name):
    value = globals().get(name)
    return value if value is not None else f"pending:{name}"
def log_event(stage, event, **data):
    append_jsonl(RUN / "logs" / "events.jsonl", {"at": now(), "stage": stage, "event": event, "run_name": RUN_NAME, "config_hash": _resolved_identity("CONFIG_HASH"), "run_identity_hash": _resolved_identity("RUN_IDENTITY_HASH"), "identity_status": "ready" if globals().get("RUN_IDENTITY_HASH") else "pending", "shard": {"index": SHARD_INDEX, "count": SHARD_COUNT}, "manifest_hashes": {key: _resolved_identity(key) for key in ["CORPUS_MANIFEST_HASH", "CHUNK_MANIFEST_HASH", "GRAPH_MANIFEST_HASH", "SPLIT_MEMBERSHIP_HASH"]}, **data})
def gpu_snapshot(label):
    if not torch.cuda.is_available(): return {"label": label, "cuda": False}
    torch.cuda.synchronize(); return {"label": label, "cuda": True, "device": torch.cuda.get_device_name(0), "bf16_supported": bool(torch.cuda.is_bf16_supported()), "allocated_gb": torch.cuda.memory_allocated() / 2**30, "reserved_gb": torch.cuda.memory_reserved() / 2**30, "peak_allocated_gb": torch.cuda.max_memory_allocated() / 2**30, "peak_reserved_gb": torch.cuda.max_memory_reserved() / 2**30}
MEMORY_SNAPSHOTS = [gpu_snapshot("before_models")]
def package_version(name):
    try: return importlib.metadata.version(name)
    except importlib.metadata.PackageNotFoundError: return None
def _lock_requirements(lock_source):
    requirements = {}
    for raw_line in lock_source.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or line.startswith("--"): continue
        if "==" not in line or line.count("==") != 1: raise RuntimeError("invalid_exact_lock_entry")
        name, version = (part.strip() for part in line.split("=="))
        version = version.rstrip("\\").strip()
        if not name or not version or any(char in name for char in "<>!=~[]; "): raise RuntimeError("invalid_exact_lock_entry")
        requirements[name.lower().replace("_", "-")] = version
    if not requirements: raise RuntimeError("empty_dependency_lock")
    return dict(sorted(requirements.items()))
APPLICATION_CONSTRAINTS = _lock_requirements('# Generated with uv pip compile from requirements-remote-vm.in against the\n# audited CUDA 12.8 resolver for CPython 3.11/x86_64-manylinux_2_28.\n# This is a hashed, transitive application lock: emitted distributions carry\n# exact versions and artifact hashes. The managed PyTorch/CUDA stack is\n# intentionally omitted from installation so this lock cannot replace the\n# base image runtime. cut-cross-entropy remains emitted and hash-pinned.\n# Install with --no-deps after the managed accelerator contract is validated.\n# emitted_application_package_count=115\n# managed_package_count=23\n# resolved_package_count=138\nabsl-py==2.5.0 \\\n    --hash=sha256:0c996f25c0490700fadabe6351630f6111534fa0ae252cc6d2014ea3b141135f \\\n    --hash=sha256:0f17b89f2a4eaaedc4f28c622998aa690564b3012a396a4ffad0821007fe03ba\naccelerate==1.10.1 \\\n    --hash=sha256:3621cff60b9a27ce798857ece05e2b9f56fcc71631cfb31ccf71f0359c311f11 \\\n    --hash=sha256:3dea89e433420e4bfac0369cae7e36dcd6a56adfcfd38cdda145c6225eab5df8\naiohappyeyeballs==2.7.1 \\\n    --hash=sha256:065665c041c42a5938ed220bdcd7230f22527fbec085e1853d2402c8a3615d9d \\\n    --hash=sha256:9243213661e29250eb41368e5daa826fc017156c3b8a11440826b2e3ed376472\naiohttp==3.14.3 \\\n    --hash=sha256:03cd2bde3d7f085b64e549c985f4bb928cad7e8ecf5323bfca320db548d81b39 \\\n    --hash=sha256:041badb8f84396357c4d3ad26de6afd7a32b112f43d3c63045c0c8278cfd2043 \\\n    --hash=sha256:0a5ff2dfbb9ce645fa5b8ef3e02c6c0b9cc3f6030ff863d0c51fffc50cb5541b \\\n    --hash=sha256:0fdea2281997af69da84c77ffa6f5938a0285f21fb3887c249d67419ca865b3d \\\n    --hash=sha256:11fb37ef075669eee52ab1928fbf6e1741fada40409fa309ebde9607a962aebf \\\n    --hash=sha256:134ac5ddcf61c6fad984b9a5727d83492ada43d63471db20fb73042c13fca62f \\\n    --hash=sha256:152516815ef926786a0b6ae2b8f1fd2e0c71582dee0b435636865316fd4891b7 \\\n    --hash=sha256:1576145bdceeb92382d899751e12743a3a5b8e460a841e3e50543859e54864dc \\\n    --hash=sha256:16100ad3ab8d649fdfbee87602d9d2dcdca9df0b9eda8a1b5fdc0d41f96da559 \\\n    --hash=sha256:16ea7e24c309fb7c0bbd505d149abe4fe4dccfb8db911db7dbec0921bc889a6f \\\n    --hash=sha256:18c441d0a8fca6de8d1f546849b9f0ab20d435993e2c5b59562b2fae6be2f929 \\\n    --hash=sha256:18cb43369747b2ae007bd2655fb8e63a099c2ff1d207962943636dac989b3147 \\\n    --hash=sha256:1b59533861b70a2185c8f4f350f791f39d64358ef6944ce71c5240c9ec0982c9 \\\n    --hash=sha256:1c5281acc88b92396f88c7e1e2748f8466689df22b80170e4f51efa712fb47a8 \\\n    --hash=sha256:1c5ec8fb1bcc31a8466f74aaf26c345d5c386fa4bd08a3f0eb9c7a4a3fe8b5bf \\\n    --hash=sha256:1caa7b0d05f3e3a36f87788c59e970a7ee1cefcfcbb924a9f138c4a6551c9cb7 \\\n    --hash=sha256:21c016079415ed3fd676963e9793700a566d85dbbd6bfc564b9b2d209147dcc8 \\\n    --hash=sha256:2498f0fe69ead802f9675beca44a7c21c62fdaa4ec5145ea1c3ad6edbee29f85 \\\n    --hash=sha256:25bd2708db6bdf6a6630dd37bdcdfcb47c4434d22ac69c64665b802910140b30 \\\n    --hash=sha256:270d3dace9ca2f10f0da5d8ebe519b7a310fc6112ed916e32df5866df0888553 \\\n    --hash=sha256:2e1161602f45a54de2ce0905243a95f58cb42dcd378402f3697f5e0b21e9d2e7 \\\n    --hash=sha256:2e9878ae68e4a5f1c0abe4dd497dbc3d51946f5837b56759e2a02e78fa90ef86 \\\n    --hash=sha256:30402d03a7c0ff52bce290b57e564e9079fd9d0cb545c8aba73f86a103162d2e \\\n    --hash=sha256:33a2d7c28d33797a2e99923dffa63f83d908a19b6bf26cfe80fa790aa5e1a75a \\\n    --hash=sha256:362a3fd481769cac1a824514bcd86fda51c65e8fe6e051099e008fddde6db17c \\\n    --hash=sha256:38901a84da3ce22249f6e860bf8f90d141bcab7da090cc398f8bb58c0e44b7da \\\n    --hash=sha256:39aded8c7f3b935b54aab1d8d73c70ec0ee2d3ec3b943e0e86611bc150ba47f5 \\\n    --hash=sha256:3a26434dafe408229ff3403458ca58de24fb51936504decac49ce6755f77e59d \\\n    --hash=sha256:3ae5b3a59436d089b5395d910121a390feed4d00578eb95a0fd1a329fe963100 \\\n    --hash=sha256:3d4f72af88ac2474bb5bca640030320e3d38a0163a1d7533500e87be458eef71 \\\n    --hash=sha256:3f42e9b78301f11c8f861746175d8b9c1ccef713fcad9eab396e2f6db8ed4a22 \\\n    --hash=sha256:42a67efc36300d052fb4508a53e8b6901b9284b599ae63945c377569c5fcc1e1 \\\n    --hash=sha256:48d67b87db6279c044760787eb01f6413032c2e6f3ba1cafaa492b1c8e578479 \\\n    --hash=sha256:498c6c623134f8e09a3c4e60bcd607a0b4590dd7dbf08dd40851b27cbb520ccb \\\n    --hash=sha256:49f7325beb0f85ef4aef5f48f490269575f83e6e2acad00a1d80b807eb027062 \\\n    --hash=sha256:4e3ac92d90e92773b2362d506068e9a948192bd553e743c5b2429e28527c8661 \\\n    --hash=sha256:530125ee1163c4219af35dc3aa1206e541e7b31b6efc1a3f93b70a136f65d427 \\\n    --hash=sha256:5373dc80ad1aa2fb9ad95c83f24eef418bbda3a61375f128e5b0192e4f3f9b32 \\\n    --hash=sha256:53e5179d8abb5710f8e83ba207c41c8d1261fcffd4616500e15ca2b7a33be10a \\\n    --hash=sha256:53e7b4ce82b54a8bcc71b3b67a5cbd177ca1d7f592cbc92cd38b7349f73482db \\\n    --hash=sha256:543906c127fb1d929b95076db19b83fa2d46751006ff1e23b093aa5ac4d8db42 \\\n    --hash=sha256:54cfcdee2770dac994417cbb0ee1f3eb0e7cb6b30c79bf44f2c02ff79ec5124a \\\n    --hash=sha256:55bdcc472aafe2de4a253045cc128007a64f1e0264fb675791e132ea5edaa3bd \\\n    --hash=sha256:56f355e79f71aef2a85c80305cc915f894b170dba76de5fe84f6351939b83c06 \\\n    --hash=sha256:5895ef58c4620afe02fa16044f023dc4dafec08158f9d08874a46a7dbc0341b8 \\\n    --hash=sha256:5bcb6ff3fdab1258a192679ff1a05d44f59626430aa05cd1a9d2447423599228 \\\n    --hash=sha256:5f08ec777f35ee70720233b8b9811d3bb5d728137f30ac91b7457709c3261ac0 \\\n    --hash=sha256:614c61d478b83953e261d02bb2df750f17227cd33ef8002945bf5aebbde21919 \\\n    --hash=sha256:617105e2c3018ee38d0c8ce5ee3c84f621a6d8b9f723202aacaff28449ca91ee \\\n    --hash=sha256:6debfa7312ff9d4c124dc71d72e9a0a4b9e0879e48ba6fcb42bef5c3300289e2 \\\n    --hash=sha256:7041d52c3a7fa20c9e8c182b534704abb19502c8bdcbde7ab23bfda6f642394f \\\n    --hash=sha256:70c987b27534f9ae1a723f47ae921571d616da21d3208282bf4c52af5164ac43 \\\n    --hash=sha256:74ab5b6a9fb13e873e5a90946588baecaf488745e1db1a4a5c433f971f035098 \\\n    --hash=sha256:78253b573e6ffab5028924fc98bc281aae05445969982a10864bc360dea2016c \\\n    --hash=sha256:7a75aa63cbf9b21cfaf60dc2657e19df2c2867d91707d653fee171ffeedd1371 \\\n    --hash=sha256:8800c996b01c2772a783e3e46f3e1abd5823029adca0df54231960de9bfefa5b \\\n    --hash=sha256:89176250f686cb9853c0fb7ead90e639e915b84a6f43eedc2a4e7ec21f1037f0 \\\n    --hash=sha256:8a5fd34f7f7410d1730d5c2ba873cacb2eed3fede366feb268a70ba22581ed8f \\\n    --hash=sha256:8b3b60de05f3dcb6f6a00f818bb2ec781cee4de0645f59ccaf99b1d1823b6100 \\\n    --hash=sha256:8f2f1c4c032c7cedd7d8da6f54c97b70266c6570c3108d3fdffee7188bb70529 \\\n    --hash=sha256:9491196535a88924a60afd5b5f434b5b203b6cc616250878dbdb223a8f7844bc \\\n    --hash=sha256:9aa6e61fdf20105c4144e755bd586008ff450791d67b1c8146fdc15959c4d51c \\\n    --hash=sha256:9d9edccfe496b476db5f398d97b865e9a6752bcf8aec4eef8390ce20fb64bb41 \\\n    --hash=sha256:9fc7b5bfec6573f3ae844f457fdde5adeb713f8b8e4a81ad64fc207b49383716 \\\n    --hash=sha256:a0dc483c00da8b673abbb367eb6f8d8f4bcec30eb58529ea13cb42e7fd2dfa33 \\\n    --hash=sha256:a3a8296e7ab5c295f53f1041487cb088e1480775aafbf7fe545d93b770a0f96f \\\n    --hash=sha256:a3e22975f905b89a55a488c2a08f2fdb2186175349e917d48985cc468a3d4c6e \\\n    --hash=sha256:a4af35c443e0b1a1bd6a8af3f3485d7fda15c142751a00f3ff8090f0b93346fa \\\n    --hash=sha256:a94dbaae5ae27bd849c93570669bff91e0510f33a80805738e3de72a7be0447b \\\n    --hash=sha256:ac74facc01463f138b0da5580329cfcc82818dea5656e83ddcd11268fc12ff80 \\\n    --hash=sha256:ad4c8b7488d745d2ca4838ebd8ae5ba9b56341d30b1da43640e4ce87f9f49646 \\\n    --hash=sha256:b014a6ed7cf912e787149fdc529166d3ceabac23f26efeea3158c9aba2354e7e \\\n    --hash=sha256:b20032766aedf6261c7a566585a40867d092ac03a0d81592d5370ef9b054f99b \\\n    --hash=sha256:b2466434105a4e03113c36ec775cc2ebe6676b62eae326fa670bb607ef788c1c \\\n    --hash=sha256:b304db572b4368edd8dda8a2274f73156fe15558fca4a917cb8a09fc47af5963 \\\n    --hash=sha256:ba59d59aba08ac02fc03b0c8983ccd5ee39a199d0552ce9e6d2b4845b34d59ae \\\n    --hash=sha256:bd52f811e65f6fb634b1047159657c98f52b407f8efec907bcfc09da9a4c0a25 \\\n    --hash=sha256:bdd0e2834dce1a26c1bbe26464861e16bbe217042cbff619247c11594472518c \\\n    --hash=sha256:c23ec8ee9d5ab2f5421f9c7fffce208435607af27fd46d4a44e031954352838f \\\n    --hash=sha256:c39846c3aad97a8530c89d7a3869a8f8e9e3762c6ac0504481e5c80948f7e807 \\\n    --hash=sha256:c3c200cf9757edd785051dc699c7ecbec22110dbfcb3fefc7a9f9695eda8ea7a \\\n    --hash=sha256:c7d3a97c678d34fc5b59da671ee9cd630096ddc643e7b5a30d54a2a6f3574d3f \\\n    --hash=sha256:c8653fd547c93a61aadc612007790f5555cdd18946fa48cf45e26d8ea4ea473d \\\n    --hash=sha256:cc7cb243a68167172f48c1fd43cee91ec4b1d40cefd190edd43369d1a6bc9c82 \\\n    --hash=sha256:ccd4893707b3e2a13e39c90d43cf80edf2e4d0457935bcc103bf2346214c3f15 \\\n    --hash=sha256:cd817772b2fcf2b8c0905795318485f9ec16eae60b29feb7f4c77085311637f0 \\\n    --hash=sha256:cda5fd5c95ad7a125a2e8464acc78b98b94c475a3780d6aa0aa157c93f470f4d \\\n    --hash=sha256:cef89a58e628c4efcac3275c2d68083f82426dcdc89c1492a6f654f9f7ea6ab9 \\\n    --hash=sha256:d1558173930a5a8d3069cee5c92fc91c87c4dbcb099debbb3622053717145a19 \\\n    --hash=sha256:d6088ec9894113802bddb3c09e974929aed2c7b3a8c456219b8aab4481f1a239 \\\n    --hash=sha256:d6218d92e450824e9b4881f44e8c09f1853b490f9a64130801024a4793b1b3b0 \\\n    --hash=sha256:d77640cc618c1d99fc4f8589c0f24a730adfa54eb1e57ef7bf0c8dfb78da898c \\\n    --hash=sha256:d7d2deec16eeedf55f2c7cf75b521ea3856a5177e123844f8fd0f114ce252cb5 \\\n    --hash=sha256:db332af25642007330fca8be5c4d194caf2bea7a7fc84415aff3497af5dfee6b \\\n    --hash=sha256:dd54d0e8717de95939766febac482ac0474d8ac3b048115f9f2b1d23a16e7db4 \\\n    --hash=sha256:ddcac3c6b382e81f1dd0499199d4136b877beb4cb5ef770bbbfba56c4b8f55d2 \\\n    --hash=sha256:df82f3787c940c94986b34222d59c9e38843fba85139f36e85255a82ad5355a9 \\\n    --hash=sha256:dfa68deb2a443bdaa3ea5297b0699c1464f08aef3812b486d1348eee61b07dc0 \\\n    --hash=sha256:dff9461ec275f22135650d5ba4b4931a11f3958df7dfbb8db630000d4dee0883 \\\n    --hash=sha256:e1e74298bab6ee0d6e749ed4fd1901c7e604bdda32c03d787a2cc71c46d0433d \\\n    --hash=sha256:e2667f0bbe7eb6c74eae5e9691441ad186e5845ca3cff63230fc09c4e7514f5d \\\n    --hash=sha256:e3be98a7c30b8c25d573dafba7171d66dfb05ee6a9070fc46535464ff97700a6 \\\n    --hash=sha256:e568e14940c09955aa51f4e645b6daa18a581c5dcfcd73744dcc86a856e3ced3 \\\n    --hash=sha256:e72ee89e28d907a18f46959b4eb0bb06701cc7f8cf4366e00029e2ccfaaf5924 \\\n    --hash=sha256:e92eb8acc45eb6a9f4935071a77edf5b85cc6f8dfad5cd99e97653c26593cdde \\\n    --hash=sha256:ea05e1f97ceea523942d9b2a7d7c0359d781d683d6b043f5943a602b14da4787 \\\n    --hash=sha256:eac645b09bcfdf73df7536331f0678c1086ea250981118ddb5199e17ccef72bb \\\n    --hash=sha256:eb0495d778817619273c108784292be161a924b9f5ae5cbbc70a2caa6838250b \\\n    --hash=sha256:ebe8e504f058fe91223351cecd2d9d6946c9d241bb0250d898ffbdf584cc72b0 \\\n    --hash=sha256:ed099d105449c4f9e84f24af203cd131349d4761d8813fa7e02c32e7128cd910 \\\n    --hash=sha256:f0f177d1b195b9e06376cfd7d308d8a1b920909a609d03ac82a8c73bbb16d3b9 \\\n    --hash=sha256:f3d2669fe7dec7fc359ecdb5984b29b50d85d5d00f8c1cb61de4f4a24ee42627 \\\n    --hash=sha256:f4e05329faa0ea1a404b37de4f034fd2c2defcca06a68dc6745e4e56c88e8a48 \\\n    --hash=sha256:f53bcd52f585e1ac3e590d61434eb61f9a88c38df041b4ea126d97144344a77b \\\n    --hash=sha256:f55119f7bf25f49ed210f6096090715da24f2943c62102448915fde3c62877ce \\\n    --hash=sha256:f631fe87a6f30df5fbe6d79640b25e4cffb38c31c7fb6f10871517b84b0f8c1a \\\n    --hash=sha256:f8fb78a83c9e5f741ca3a68cfb455c1f5bb83b4e7249a3848b3cd78d0a8563b0 \\\n    --hash=sha256:fa9467a8113aa69d3d7c55a70ef0b7c636010a40993f3df9d9d0d73b3eb7ef24 \\\n    --hash=sha256:fd51ebf9d3a00c074df4ede271023f4d2dba289bcc740b88191872716014e3c5\naiosignal==1.4.0 \\\n    --hash=sha256:053243f8b92b990551949e63930a839ff0cf0b0ebbe0597b0f3fb19e1a0fe82e \\\n    --hash=sha256:f47eecd9468083c2029cc99945502cb7708b082c232f9aca65da147157b251c7\nannotated-doc==0.0.5 \\\n    --hash=sha256:117bac03a25ede5df5440e855b32d556049ca169ead221505badf432fed4b101 \\\n    --hash=sha256:c7e58ce09192557605d8bbd92836d7e1d520ac9580096042c0bfd197efacf1bb\nannotated-types==0.8.0 \\\n    --hash=sha256:13b2beaad985e05e2d6407ee4c4f35590b11f8d693a258a561055cac8f64cab7 \\\n    --hash=sha256:f072f4d804ea359e4eaf198b1af7a8b0943881a87f31bb764f8bf219bb9419e0\nanyio==4.14.2 \\\n    --hash=sha256:9f505dda5ac9f0c8309b5e8bd445a8c2bf7246f3ce950121e45ea15bc41d1494 \\\n    --hash=sha256:cfa139f3ed1a23ee8f88a145ddb5ac7605b8bbfd8592baacd7ce3d8bb4313c7f\nattrs==26.1.0 \\\n    --hash=sha256:c647aa4a12dfbad9333ca4e71fe62ddc36f4e63b2d260a37a8b83d2f043ac309 \\\n    --hash=sha256:d03ceb89cb322a8fd706d4fb91940737b6642aa36998fe130a9bc96c985eff32\nbeautifulsoup4==4.14.2 \\\n    --hash=sha256:2a98ab9f944a11acee9cc848508ec28d9228abfd522ef0fad6a02a72e0ded69e \\\n    --hash=sha256:5ef6fa3a8cbece8488d66985560f97ed091e22bbc4e9c2338508a9d5de6d4515\nbert-score==0.3.13 \\\n    --hash=sha256:8ffe5838eac8cdd988b8b1a896af7f49071188c8c011a1ed160d71a9899a2ba4 \\\n    --hash=sha256:bbbb4c7fcdaa46d7681aff49f37f96faa09ed74e1b150e659bdc6b58a66989b9\nbitsandbytes==0.48.1 \\\n    --hash=sha256:3e72cf07ba6d2169e69a61282a6f072fc675efee86049e56a33de099a0363ef2 \\\n    --hash=sha256:b7f440aee5ec8cb1d028b0d3b2d71e97c302766dc605232293f4a0f7e48b5c75 \\\n    --hash=sha256:d7d3f9b00b132bb25f09320ee07ccbfae8c1e0ea11cae48fbf7e1eff9943c7b4\ncertifi==2026.7.22 \\\n    --hash=sha256:62f22742b58a1a33014a2b6b706588a8d7e2a88ae7bd1a6ebe8c992928483775 \\\n    --hash=sha256:741e2c3b351ddf169a738da9f2c048608ff7f2c5cc02f1ebc6b118bb090d5d55\ncharset-normalizer==3.5.1 \\\n    --hash=sha256:00668ebb0609751758682eb0b5857e7c35b9f00e84dfdef062e103244ec94d45 \\\n    --hash=sha256:012a22b88a77ca2e59b98ac5889b0deb604147666032f45e6d6e217634d2550d \\\n    --hash=sha256:01e93745f7f219b703b60ba7afead36cfc4242782be5af484673fc500df12da5 \\\n    --hash=sha256:04368edf83514385ffc3e1cfd4546e595f4f1272dd23ba437a93a9cc3741d47b \\\n    --hash=sha256:0722590aabf9dc6a6c0343d523c05458fa2b5047dbe6302fd526bb570600753f \\\n    --hash=sha256:07ffd07412fc5d5e84cd8952acf9ff7e4ed7a708e69d1bada19d8ba91711353f \\\n    --hash=sha256:09a7bba9f739468c8e78c36a75c33768e53cb1959fc638f510454c14683f00d5 \\\n    --hash=sha256:0b2b1b3fa5670c127b246df1d0c059defd41f689a868a3b9d79df9b1cac42d22 \\\n    --hash=sha256:0c6dfb5ca6723eeed15aa8e564a014d69fcb8812f94eef11fe3631e0508199f5 \\\n    --hash=sha256:0d929fc574b4d6fd9e7c0f5c2ede8716a41911923aa7fa5fce38e0818aa4a1ac \\\n    --hash=sha256:13e3afe97712e8887cd516e960c63f0b93122971e5b5e4b2622fe7701771e838 \\\n    --hash=sha256:15f024313246a4ed976c60f440bb8d257815513a681d212ff74fd46f7d715a90 \\\n    --hash=sha256:195ce897c6153c0700078142cf8efe3e6454ca4cf4357499e4078dfd83396626 \\\n    --hash=sha256:19a3dd5aa73cef1c99687c4fc57db016a9c17104ae1185da88ba566a5d3bebe4 \\\n    --hash=sha256:1d1c7a53a6c2103925cdd6d7229f8c567379f211c869793df679f2e9f738c369 \\\n    --hash=sha256:1f5883d77fd409a261abb5dc8ccbe335720d798b1de4abb3b1d47ccbbc76b53b \\\n    --hash=sha256:21b82d8082f6f5e7f456ef0bd16323d08de1266efbfeb476e64b2a91d1471a4e \\\n    --hash=sha256:252d099029bcbea642f2a06c4ed5046bdf8b5a8150b64afa5e027e88b106e5ee \\\n    --hash=sha256:256dd4d85d9e4dc595e2bc983c980e73f62ddeb3165c58b4c3dfe78c5c8548c1 \\\n    --hash=sha256:26422d45fd13551cf564c58932f7d72b4f58b93b0fcf18c35ba6be12b46bb102 \\\n    --hash=sha256:2679de311c7946dde5d3b6f44941844133ff5c7cb86099c0061ab1e8901c20a8 \\\n    --hash=sha256:29880d17a8eb0b5cfdfd8944b468322928059aa35f1f5fa8ff22b149ec0b42f8 \\\n    --hash=sha256:2bced4061f000f7187254a02ad3433ae17eaf991747ceea2f478422590a5bba9 \\\n    --hash=sha256:2e9cf9253119d8e5d111f05d71626786fd3d6193817316eab1ca088cdb8593cf \\\n    --hash=sha256:2f06b7eae9dbe77fe1d644ca244dad508de8d302870a43f3c559b521270938a0 \\\n    --hash=sha256:2f293479cce755c75f1697e87c409b7ae4c555c7dfecb6e988ad13abba943031 \\\n    --hash=sha256:329fc3ccb63ad22d867d84c2adea759a64079a37ba4a343433b02c7a2816871e \\\n    --hash=sha256:343fb4f2821043bd87095f7b08a1a181febc8e36ac64212143bbfd0a0e1bc235 \\\n    --hash=sha256:3588e376b3ea2eea84976f67273d679f229e24c66dce7b82ae45aef04ff6e072 \\\n    --hash=sha256:35aea775dc2bd5f54cd84a1cd2696cc3207c479cb9cf0bd346f0d343e4300ddb \\\n    --hash=sha256:35fe081843b35aad20ffeccec3eeffbe637b15d14f3fb22cc1b59cd8ec17e93c \\\n    --hash=sha256:36047af20e17097c3bb9476c2b7655f2f7aa51322c0ba58c07695bedf755a950 \\\n    --hash=sha256:3617ac3cfd8b9888f145ad89dd6e692285834b0201c6074a5eeaad3fd4d668c2 \\\n    --hash=sha256:366ec70f5547c640d3ce1985722490f23faf4eb5216a7eeba78277490e78dacb \\\n    --hash=sha256:394fea06235c8543390050ed5f529187074b029fb027213f6c46ac11ab5d950e \\\n    --hash=sha256:3d27167433c0d5f18dc850f07d0b3816221984fecdc405d6c157a6f0b8f8e9e6 \\\n    --hash=sha256:3e5e1224c0a6a90e05843e07adfec669edebec17801c67072f51e59561d63c0b \\\n    --hash=sha256:41876ee62a3dddf48ff1121ad8f0798032aa03f2fd35f21f34a4cab14f18d8d2 \\\n    --hash=sha256:433c5a81eade63b47e522303bad236f59dba55ea6951746f5558355eeed8c75d \\\n    --hash=sha256:4582c27e8c889d64811987b5967fbd3ae0c823fe1fd933b543d55ac20bb475fa \\\n    --hash=sha256:485a0d363cafefcd2538a73c7c838daa2035f09b2c9f9b5e3133f80c6aeb84c2 \\\n    --hash=sha256:494b70049a4d69aec6e8137c13af4cf8db8c9f9820a1392ac293b0dd2987a818 \\\n    --hash=sha256:496846868fea80e479324862fa877f02411f2fd0f83b79ccee2607aa68b2a032 \\\n    --hash=sha256:4abdc5f9ad448c1ecbfae2974b820535d6bc6e7eef63babbab3d81cf46968c71 \\\n    --hash=sha256:4b599739b93b2cbeded49645ae3c8d1405c29ddfbceac1545c87a3f9580a9e96 \\\n    --hash=sha256:4bea7f8ebe90bbd7f0e4a2de42ca6924ba23e3e76418c408ff82f1d46fabd687 \\\n    --hash=sha256:4c4fb141a727957c93edfe5c32a26ceb6b5f6461d67146e2d39f51e16170bea8 \\\n    --hash=sha256:4c9548dc78002099910abaebc0a72ac58b7d30931869e0351c09b507dff4ece3 \\\n    --hash=sha256:4d26f14f041e83dd8edfd61f4cd4fa7285d31798b5bf1f28e70c367ba6c41d61 \\\n    --hash=sha256:4f298bdadb8f0b9e5672877f647d1be9373ef5320c9e2f049795e26cad28b6a9 \\\n    --hash=sha256:52ec005752a56ae79547a05c0139ca2501a0c866390b6115008456b9f0e7cde1 \\\n    --hash=sha256:55261ac0d2941c42f196dd576f543d87a8ee03cd6f5e30dfb4d807b2e3b9121a \\\n    --hash=sha256:56490c595a28b1bb27dfc583e816152a9767721ef58b2c03b13f954d2f707420 \\\n    --hash=sha256:58d3e12c88e0950bca850ae1f7c256055c097639c2edb9eb123af9807d8b15e4 \\\n    --hash=sha256:58d4aa13a59c969dbfdf9e6a9560e242cbfd9e8a8f50c2747714df1a423adf65 \\\n    --hash=sha256:59171c6e45bf07d0d5cab3b0bf81d945035530f6873398b3b531c31184d46663 \\\n    --hash=sha256:5b6d1386bf0096d26d3a863dc0a487a5b4eb9aa93cf5ba69683d29dde6b9d60f \\\n    --hash=sha256:5c0ea61a470e070686aa30892fed79e297d2c8d0ab46b8bcdf027d38c51da591 \\\n    --hash=sha256:5c84bec0ab5ae0c64bfe73a7d2adcb5ce73b467523fc27fd6a28ab2aa6cbe35a \\\n    --hash=sha256:5ca0555312ae2fe82715cada7fac375530c2f3349e1eaa1bcb33d0283ac79a18 \\\n    --hash=sha256:5d8531a6569d025f68e2321e7638fb7978f23db58e5f69f56913837aae03816e \\\n    --hash=sha256:5e2d0e146dcb57034f8b97dc58d2d512cb90aba253960ce449f695fec6a82c6f \\\n    --hash=sha256:5fc45d653ea8c9a20479167e11d4a0f8cb2fa3470737ab6f9c827532313187b7 \\\n    --hash=sha256:6117b84ea48435e5356dc737f5121485c30920ba43375fa7b434fd753df0eac3 \\\n    --hash=sha256:6199d5606e2bbf2b096cf64d03f8b6790c91081d5ac866b8e7bb6422738cc60c \\\n    --hash=sha256:62b55f6722735a6c472f88361cde6640608773d9443cebdbb51abf436a1fcdd3 \\\n    --hash=sha256:687c9ca3035544b113bea2055e180af96fb63c0c476e22a9180f51925186e7b7 \\\n    --hash=sha256:6b7430cf5728e68f6c462254009a6ef4086e1bea43cf2f57aa9c55fb4f50ff96 \\\n    --hash=sha256:6ba32c4d2abf1d2fe7cf27d280f4cca5664233b0f885549c7761719eb977f486 \\\n    --hash=sha256:6c9cdde8becb25a7fde49924511aa2644d6f8081cc8df8e9452724303348d8e3 \\\n    --hash=sha256:6df0ec430f9a831772c23ca5a224cba36517a58a84bb32c32bb59a9fa67c47f6 \\\n    --hash=sha256:6e2912d4babbc65196ac13c2f53468dc57fb8b9c25ef913e8c59ddf7c6dc0e1b \\\n    --hash=sha256:6e5e4d73d588ca5ed09df1b7dcd1b203d1df3c542e3f50d126c947d432b10731 \\\n    --hash=sha256:70055ff39b97c99e7ae40ea3e393fb62aa2e44dbd9b29f8d14f42fb0025c3959 \\\n    --hash=sha256:706bfd38730a5ac7a365793269a00f4e988178cec121391f4248d84ad8c972e9 \\\n    --hash=sha256:7235dc28fc6dd9d832ac7c7bce95367dedb85929f17368a0c2bee1e080b9acbf \\\n    --hash=sha256:774d157f112367ff4abd29019f38f023c24e00e56edc7829c20e358a5a913ad8 \\\n    --hash=sha256:77efcff2b23071c349402ac1066667a3d011f62398d81408c9b88ad991747c9e \\\n    --hash=sha256:789b8982559ae28dad2356519f841655756cdcd96616410590ae0b17454ee64f \\\n    --hash=sha256:7ac76cf9afd34929d76eb7fcb63be476a4853d8a96f0dcf2d0db68a0cbdf9885 \\\n    --hash=sha256:7c0c10730342b0c9b35dd1d619beb8214e520bd96a1f870f452680b238aab3e0 \\\n    --hash=sha256:823f82903d189af463d7df250ef1f7f696f3cee08cc8d91deb565e8d425f6506 \\\n    --hash=sha256:838648accb3a7fd9803fd45c87bce8509648eb0c11bc34e216141300977244f2 \\\n    --hash=sha256:854066be00447fa8de2ccbbe893e2ffc4b123ef16d897af794c1e18bd4a714b0 \\\n    --hash=sha256:85d5855daafc240cc045c026d7a15fd198a09b0fc8ff6f5ecbb5297b509cb11e \\\n    --hash=sha256:85de3134b5379856e323ba37c19c9256d39425f7b76a63af52b09fb4664c2e8f \\\n    --hash=sha256:87e4f41d375c0b9be2fb5251aee4b8a689169e134535aed81bf085c3b647451e \\\n    --hash=sha256:88ca277405c2d3b71c4e1c2ee0e7966e807bcba86a69d11e19ba199d18ae4491 \\\n    --hash=sha256:88e85ab89cb822c1e635f51d6d32e488f94e002e70e2f492bdb8b945543f345a \\\n    --hash=sha256:8ac8c94b6539074e0f40899301273ac8402b9b3e01c7b7ba269ff30340aaaf20 \\\n    --hash=sha256:8fe532b3c966d1fb794e0698e4589d0444017ae77fc0b31edea13c0e35bcc449 \\\n    --hash=sha256:9085f87b0e38a2b92b8923059b4e8789fe40d9279712d15dcc670048d77079af \\\n    --hash=sha256:90b7481fb62fbe172c558bc6fd1c4c98d82004a54a7551f20e11ac9bf0b8708c \\\n    --hash=sha256:92caef967d287a407085d61176fce4012b1dd62daed4eb6d5ceb26d3d2538712 \\\n    --hash=sha256:9362dd90aa7dab48c0054a21187791ccf05473f7dba5d92b8033ae62164675e7 \\\n    --hash=sha256:94d78ecec2605a8d0398b0f365d5f12a63248438516f5dac536a5eff7337df4a \\\n    --hash=sha256:94fbf1c0c6cc0d3d5e50f9a9313a8cdca90dd696d34b381cd1704f8c9e939f20 \\\n    --hash=sha256:950f23cb393f85543777b0433f082cddd25b51ab398eac7971146495679efe5f \\\n    --hash=sha256:96eefc178f8636b9c760c5829345307fd81cfae9ab1e80997dbddeb0f54ee9a3 \\\n    --hash=sha256:96fef3e886d6a9874b14f27fc193fbdc69d5d8035783d86aa4e1cea594e695f9 \\\n    --hash=sha256:977cdbd483a9cff38179bea4fd754289a6f2195c7abd414aba85410b3e66cc5e \\\n    --hash=sha256:978eab16f55b4ab2c2a745be9a0a840bf8f09a7f227d9c76eb30214d078865a5 \\\n    --hash=sha256:994e883d17c559cdfd38c84003c8b27d25424a1077272a17e7cd27bfe0bf57b2 \\\n    --hash=sha256:9ac4444d8d4fd4c4bd08bf451ed3167aa9e7ec6cdb41b648794f1d1103652e36 \\\n    --hash=sha256:9b5db6052055d34d41230fb78d7c439c23dc536a9896f6cb039e8dd92cfc1263 \\\n    --hash=sha256:9d9a0dc7cbe9bec24c3f767c9122c41fe5a1bc43f47cd099d00d393e09769de4 \\\n    --hash=sha256:9dbdd9205662134957cf0c324f639bdc5031c0ca056e2369e238db75187c0f11 \\\n    --hash=sha256:9eea3ab2597a5e65fe65296e2d6a84570845a6b55532d90333d740d48bbc850a \\\n    --hash=sha256:a2028475ba855475b8b4d3cfeb4994269c967aea8b9892dfba907f4263a863a3 \\\n    --hash=sha256:a3a370082ce34d0612f421e15fe011c53bb1feff21a26d06ad4fb244dab5a375 \\\n    --hash=sha256:a545775cfe815855ea32d7c27731d79da358ef2055b4a25830231b1622dd18aa \\\n    --hash=sha256:a5cbd90ecf0fc62e64726917ad083b73001f0563657a87ec3c0b504e277dc90d \\\n    --hash=sha256:a6d095662e73e74f0a49988e0593373e243e3a52e27bfeea0a859e88acf4a0f5 \\\n    --hash=sha256:a6dac12ff6b846103483683f60c5f8fee205121adc58ffd87e90a90a3af69e99 \\\n    --hash=sha256:a951ad59cad9145664a730d3036b40b844e74d2d3683da40111463cd3a83845d \\\n    --hash=sha256:aa1099b956fb795e686d073568f6dc002a0bb89765ea6d5b055dd7d9bf1b116c \\\n    --hash=sha256:aa2bb0b37202dca27175591f761108b5d34096ade1191ffe4808bdf6b1571488 \\\n    --hash=sha256:aae2ee51122d3ae968a3837d97dc24a0aeebb0dea23694422cd172bd30017cd6 \\\n    --hash=sha256:ab743e9bc90c1f73552ec33e10e3331315acd2c397b36065b591b0181de533cc \\\n    --hash=sha256:ac00177c4831ffa650f8609e4bdddd5fe09c03b1c0c47acece7e6ea20421598b \\\n    --hash=sha256:ac13b004224fb341e1e25a1ed5e19d32f57cdb2a403e01f003b46f051a550f6f \\\n    --hash=sha256:acaf604462bf330b0d07e7a07c1d6e4adac79e5fb13e9c5140590542cafacc00 \\\n    --hash=sha256:ae31a1a1db2ee6cc2942fccaf695c934bc7f3db9f2133a3fef1f367cf1a4ab10 \\\n    --hash=sha256:ae4a097991662cd4fff0ddc74e0fe7874f82e00042fa0ea00855645ed0c79598 \\\n    --hash=sha256:aea996a6aba25260827c9ea511d1addfde2da9eb686ac961838509086188b7e6 \\\n    --hash=sha256:b39b69b347e5e47a3b5b8cfc005c68c1ba347474e3960236c4944a8ecd174962 \\\n    --hash=sha256:b54e7e13267d49ffbfe68e25b3cbd774dab38fa37238f71265e91b36146eb21c \\\n    --hash=sha256:b9af956078716df40d985fb0dfeb2c2120c5ca92ba4ff4b388acfd01cdc14d08 \\\n    --hash=sha256:ba2f37ee79e6338845261a3c5b1784e5d1acdff2c0785b284f1b633033d136ab \\\n    --hash=sha256:ba501e667c17d8411f98e67a022d9604ef179aff0e459b7e292c796837c13573 \\\n    --hash=sha256:baf3775a2635e5a11fbd5e4e64ee69c7e86875d224a5c72aca4c141064589a90 \\\n    --hash=sha256:bb57753e36e4855b8ca375069482250a6246372331a3e4f3407eaebb007443f5 \\\n    --hash=sha256:bd6c173f04743d483881bffa1478d5a4624475b8cd1d2194956a75548e191c18 \\\n    --hash=sha256:be47f99644b208bff7766314013f9acf57b056b04191d570d68ad14022cf5b1d \\\n    --hash=sha256:c010f5581d9c612804cc59fcf7b524b707fbcb72828551237ab545bb5c7034af \\\n    --hash=sha256:c1dcc36dcb96abc02236e182d17e0f71430152a6c2c7447421da2d2dc144edea \\\n    --hash=sha256:c428c6c31eb5f4277d7f8eccaf767fbd548ddd5ce3c8b4f4cbbfab3d96b5904c \\\n    --hash=sha256:c658c50ac0c98cd755a2dd50b7977d3bca7df401dcc47fbdfa87db53ef7d4e8b \\\n    --hash=sha256:c71fb0d56c920c269cd3e2e3fe7c610e3f1fdb21a6ce60efa6430ff63676cea6 \\\n    --hash=sha256:c7b742bf31c88566b4bb6335a7f393bb322e580b6bb98df7bd0c25e6e3519ce8 \\\n    --hash=sha256:cc0329df4caaceb950d2f580b5ac716a377f7059624a0bafaeaf8a218c6ed774 \\\n    --hash=sha256:cc5d36d96478aa9c60654bd932525bf32964c62a7281eafdf16d85003a8d6004 \\\n    --hash=sha256:ce854f5f478050ade5a238731c4ca985a7d3b3cb53ff600a9b5c3b689b5f0a7a \\\n    --hash=sha256:ced3fdd71aaa83ce593746c2edb42b7a59cb4c19c8b5c407781c72e493aae55a \\\n    --hash=sha256:cee5dd7c6fb5dd52a0fe2a740f9bc6e3593f5f8b1788bde49de02086f30182b2 \\\n    --hash=sha256:cfa1c0cc3a8f9f53f1243a5a99ac36fd003880199383b37672e86ddda9cb07e2 \\\n    --hash=sha256:d1ee1e296209fdce05b81b663250eefa02213a2da7b41bf26f7829b8ba3545aa \\\n    --hash=sha256:d59b75732e9b6f27388e10c14b0259cc5f2e48c78627d185e6a177b58ad3cffe \\\n    --hash=sha256:d63600d620ad0064c3a748b950ac5ea38a80190e5498532efefa4b7b3f1da1f3 \\\n    --hash=sha256:dd732602a7009217f658d5863d12d79d373a4de0eebc111094bcdd3bb8e0a6cc \\\n    --hash=sha256:e06efa066f7dbadbc84ebc126a97c452a6451dfcf589d89d788484949e1cf795 \\\n    --hash=sha256:e199fb99720074809a7720f1c0b4d919eea8b87e88713e0f8f602f7bef543d9d \\\n    --hash=sha256:e4b018dc5a0eee4676e38fe84a47a427816c590b93b55d9025274ec4d6ffc2dc \\\n    --hash=sha256:e6621fb2a4988d6e53eedc455e5903e2679f3967b8acb3d639f1b63c14a2e893 \\\n    --hash=sha256:e71c909f353863b2b89c83de2ebed71ea6d0df8a6ef65a128193c5e650766bef \\\n    --hash=sha256:e90251c0c7bdd54a100a0dce3c07b7e637278c93af29dbf78ebb89a58c4bac7d \\\n    --hash=sha256:e9fbdce1e47394b09bc9f26ab117dfc8d6491977a11d86f592bb42c779db2fda \\\n    --hash=sha256:eb12fb2ba69ffa05f8695f61c69e591dc4b4a12ac3757ac8af8adb259bf56d17 \\\n    --hash=sha256:eda059b6bc8bc0812d626fd91a7ce01bf583df0a61296eff390fd94141a34e30 \\\n    --hash=sha256:f03ac127268b43ef4fe9e6ab6794a6794b49485a0cc0c1db79876d2f33f75bc7 \\\n    --hash=sha256:f298e218441525d3794428b4c8b8fb8662c6d3ea79925d4807ee6b9a96a3bca5 \\\n    --hash=sha256:f5542f9b941279d82d41eb0aa9f98eba36fe4df5c7086c651df7944935b37182 \\\n    --hash=sha256:f6f7deae3feb4edfa2efaf7c574fe88cbf055038a6abdb40188e4fff66d5699f \\\n    --hash=sha256:f9b1e28d0e8dbfa858abdba91d6b547beaf2df1a59bec6da6faae7b96a4991a9 \\\n    --hash=sha256:f9f8405c2c758532c74fed975dbee57be1f31a6e865c031870c79a6ed3212ada \\\n    --hash=sha256:fa48b1b63d639f9483e0633e092f5851e2348c352f1f9bb6c8182f87884ef876 \\\n    --hash=sha256:fb78f6e7fcd8ad785d28cd577168bc1aaee827b25bb8755638f694794ea98f0a \\\n    --hash=sha256:fbc597639158fd7c14d55e808718848319540f51b0e6746e3eefa59723a4a348 \\\n    --hash=sha256:fce8cbd4997efeb450bd298b54f755dcdff18d496f7a5ddbb4867c6d7c88fdc3 \\\n    --hash=sha256:fd0350afdc3aabd5576f60ea109228bd5538139713c7b094c5cd27c73a98bc6f \\\n    --hash=sha256:fd0a274c0e5f9a21565cd9d3dd749b61f96b7aa1e20a93aa1ba4029518f2e5c0 \\\n    --hash=sha256:fdb8a068947befafba9952162645dc2fecaeb400e64584829ed5e9b2fbe21a7f\nclick==8.4.2 \\\n    --hash=sha256:9a6cea6e60b17ebe0a44c5cc636d94f09bd66142c1cd7d8b4cd731c4917a15f6 \\\n    --hash=sha256:e6f9f66136c816745b9d65817da91d61d957fb16e02e4dcd0552553c5a197b76\ncolorama==0.4.6 \\\n    --hash=sha256:08695f5cb7ed6e0531a20572697297273c47b8cae5a63ffc6d6ed5c201be6e44 \\\n    --hash=sha256:4f1d9991f5acc0ca119f9d443620b77f9d6b33703e51011c16baf57afb285fc6\ncontourpy==1.3.3 \\\n    --hash=sha256:023b44101dfe49d7d53932be418477dba359649246075c996866106da069af69 \\\n    --hash=sha256:07ce5ed73ecdc4a03ffe3e1b3e3c1166db35ae7584be76f65dbbe28a7791b0cc \\\n    --hash=sha256:083e12155b210502d0bca491432bb04d56dc3432f95a979b429f2848c3dbe880 \\\n    --hash=sha256:0bf67e0e3f482cb69779dd3061b534eb35ac9b17f163d851e2a547d56dba0a3a \\\n    --hash=sha256:0c1fc238306b35f246d61a1d416a627348b5cf0648648a031e14bb8705fcdfe8 \\\n    --hash=sha256:13b68d6a62db8eafaebb8039218921399baf6e47bf85006fd8529f2a08ef33fc \\\n    --hash=sha256:15ff10bfada4bf92ec8b31c62bf7c1834c244019b4a33095a68000d7075df470 \\\n    --hash=sha256:177fb367556747a686509d6fef71d221a4b198a3905fe824430e5ea0fda54eb5 \\\n    --hash=sha256:1cadd8b8969f060ba45ed7c1b714fe69185812ab43bd6b86a9123fe8f99c3263 \\\n    --hash=sha256:1fd43c3be4c8e5fd6e4f2baeae35ae18176cf2e5cced681cca908addf1cdd53b \\\n    --hash=sha256:22e9b1bd7a9b1d652cd77388465dc358dafcd2e217d35552424aa4f996f524f5 \\\n    --hash=sha256:23416f38bfd74d5d28ab8429cc4d63fa67d5068bd711a85edb1c3fb0c3e2f381 \\\n    --hash=sha256:283edd842a01e3dcd435b1c5116798d661378d83d36d337b8dde1d16a5fc9ba3 \\\n    --hash=sha256:2a2a8b627d5cc6b7c41a4beff6c5ad5eb848c88255fda4a8745f7e901b32d8e4 \\\n    --hash=sha256:2b7e9480ffe2b0cd2e787e4df64270e3a0440d9db8dc823312e2c940c167df7e \\\n    --hash=sha256:322ab1c99b008dad206d406bb61d014cf0174df491ae9d9d0fac6a6fda4f977f \\\n    --hash=sha256:33c82d0138c0a062380332c861387650c82e4cf1747aaa6938b9b6516762e772 \\\n    --hash=sha256:348ac1f5d4f1d66d3322420f01d42e43122f43616e0f194fc1c9f5d830c5b286 \\\n    --hash=sha256:3519428f6be58431c56581f1694ba8e50626f2dd550af225f82fb5f5814d2a42 \\\n    --hash=sha256:3c30273eb2a55024ff31ba7d052dde990d7d8e5450f4bbb6e913558b3d6c2301 \\\n    --hash=sha256:3d1a3799d62d45c18bafd41c5fa05120b96a28079f2393af559b843d1a966a77 \\\n    --hash=sha256:451e71b5a7d597379ef572de31eeb909a87246974d960049a9848c3bc6c41bf7 \\\n    --hash=sha256:459c1f020cd59fcfe6650180678a9993932d80d44ccde1fa1868977438f0b411 \\\n    --hash=sha256:4d00e655fcef08aba35ec9610536bfe90267d7ab5ba944f7032549c55a146da1 \\\n    --hash=sha256:4debd64f124ca62069f313a9cb86656ff087786016d76927ae2cf37846b006c9 \\\n    --hash=sha256:4feffb6537d64b84877da813a5c30f1422ea5739566abf0bd18065ac040e120a \\\n    --hash=sha256:50ed930df7289ff2a8d7afeb9603f8289e5704755c7e5c3bbd929c90c817164b \\\n    --hash=sha256:51e79c1f7470158e838808d4a996fa9bac72c498e93d8ebe5119bc1e6becb0db \\\n    --hash=sha256:556dba8fb6f5d8742f2923fe9457dbdd51e1049c4a43fd3986a0b14a1d815fc6 \\\n    --hash=sha256:598c3aaece21c503615fd59c92a3598b428b2f01bfb4b8ca9c4edeecc2438620 \\\n    --hash=sha256:5ed3657edf08512fc3fe81b510e35c2012fbd3081d2e26160f27ca28affec989 \\\n    --hash=sha256:626d60935cf668e70a5ce6ff184fd713e9683fb458898e4249b63be9e28286ea \\\n    --hash=sha256:644a6853d15b2512d67881586bd03f462c7ab755db95f16f14d7e238f2852c67 \\\n    --hash=sha256:655456777ff65c2c548b7c454af9c6f33f16c8884f11083244b5819cc214f1b5 \\\n    --hash=sha256:66c8a43a4f7b8df8b71ee1840e4211a3c8d93b214b213f590e18a1beca458f7d \\\n    --hash=sha256:6afc576f7b33cf00996e5c1102dc2a8f7cc89e39c0b55df93a0b78c1bd992b36 \\\n    --hash=sha256:6c3d53c796f8647d6deb1abe867daeb66dcc8a97e8455efa729516b997b8ed99 \\\n    --hash=sha256:709a48ef9a690e1343202916450bc48b9e51c049b089c7f79a267b46cffcdaa1 \\\n    --hash=sha256:70f9aad7de812d6541d29d2bbf8feb22ff7e1c299523db288004e3157ff4674e \\\n    --hash=sha256:8153b8bfc11e1e4d75bcb0bff1db232f9e10b274e0929de9d608027e0d34ff8b \\\n    --hash=sha256:87acf5963fc2b34825e5b6b048f40e3635dd547f590b04d2ab317c2619ef7ae8 \\\n    --hash=sha256:88df9880d507169449d434c293467418b9f6cbe82edd19284aa0409e7fdb933d \\\n    --hash=sha256:929ddf8c4c7f348e4c0a5a3a714b5c8542ffaa8c22954862a46ca1813b667ee7 \\\n    --hash=sha256:92d9abc807cf7d0e047b95ca5d957cf4792fcd04e920ca70d48add15c1a90ea7 \\\n    --hash=sha256:95b181891b4c71de4bb404c6621e7e2390745f887f2a026b2d99e92c17892339 \\\n    --hash=sha256:9e999574eddae35f1312c2b4b717b7885d4edd6cb46700e04f7f02db454e67c1 \\\n    --hash=sha256:a15459b0f4615b00bbd1e91f1b9e19b7e63aea7483d03d804186f278c0af2659 \\\n    --hash=sha256:a22738912262aa3e254e4f3cb079a95a67132fc5a063890e224393596902f5a4 \\\n    --hash=sha256:ab2fd90904c503739a75b7c8c5c01160130ba67944a7b77bbf36ef8054576e7f \\\n    --hash=sha256:ab3074b48c4e2cf1a960e6bbeb7f04566bf36b1861d5c9d4d8ac04b82e38ba20 \\\n    --hash=sha256:afe5a512f31ee6bd7d0dda52ec9864c984ca3d66664444f2d72e0dc4eb832e36 \\\n    --hash=sha256:b08a32ea2f8e42cf1d4be3169a98dd4be32bafe4f22b6c4cb4ba810fa9e5d2cb \\\n    --hash=sha256:b20c7c9a3bf701366556e1b1984ed2d0cedf999903c51311417cf5f591d8c78d \\\n    --hash=sha256:b2e8faa0ed68cb29af51edd8e24798bb661eac3bd9f65420c1887b6ca89987c8 \\\n    --hash=sha256:b7301b89040075c30e5768810bc96a8e8d78085b47d8be6e4c3f5a0b4ed478a0 \\\n    --hash=sha256:b7448cb5a725bb1e35ce88771b86fba35ef418952474492cf7c764059933ff8b \\\n    --hash=sha256:ca0fdcd73925568ca027e0b17ab07aad764be4706d0a925b89227e447d9737b7 \\\n    --hash=sha256:ca658cd1a680a5c9ea96dc61cdbae1e85c8f25849843aa799dfd3cb370ad4fbe \\\n    --hash=sha256:cbedb772ed74ff5be440fa8eee9bd49f64f6e3fc09436d9c7d8f1c287b121d77 \\\n    --hash=sha256:cd5dfcaeb10f7b7f9dc8941717c6c2ade08f587be2226222c12b25f0483ed497 \\\n    --hash=sha256:cf9022ef053f2694e31d630feaacb21ea24224be1c3ad0520b13d844274614fd \\\n    --hash=sha256:d002b6f00d73d69333dac9d0b8d5e84d9724ff9ef044fd63c5986e62b7c9e1b1 \\\n    --hash=sha256:d06bb1f751ba5d417047db62bca3c8fde202b8c11fb50742ab3ab962c81e8216 \\\n    --hash=sha256:d304906ecc71672e9c89e87c4675dc5c2645e1f4269a5063b99b0bb29f232d13 \\\n    --hash=sha256:e4e6b05a45525357e382909a4c1600444e2a45b4795163d3b22669285591c1ae \\\n    --hash=sha256:e74a9a0f5e3fff48fb5a7f2fd2b9b70a3fe014a67522f79b7cca4c0c7e43c9ae \\\n    --hash=sha256:ea37e7b45949df430fe649e5de8351c423430046a2af20b1c1961cae3afcda77 \\\n    --hash=sha256:f64836de09927cba6f79dcd00fdd7d5329f3fccc633468507079c829ca4db4e3 \\\n    --hash=sha256:fd6ec6be509c787f1caf6b247f0b1ca598bef13f4ddeaa126b7658215529ba0f \\\n    --hash=sha256:fd907ae12cd483cd83e414b12941c632a969171bf90fc937d0c9f268a31cafff \\\n    --hash=sha256:fd914713266421b7536de2bfa8181aa8c699432b6763a0ea64195ebe28bff6a9 \\\n    --hash=sha256:fde6c716d51c04b1c25d0b90364d0be954624a0ee9d60e23e850e8d48353d07a\ncut-cross-entropy==25.1.1 \\\n    --hash=sha256:5fe5924509248b1aea5c890f8887c6a7759f7c8b1ebc0490e42c247c4f7c1e34 \\\n    --hash=sha256:e46f26d348f6a67927d17e65c5a212e795be13dcad5b10a77a200d6b8102d9d1\ncycler==0.12.1 \\\n    --hash=sha256:85cef7cff222d8644161529808465972e51340599459b8ac3ccbac5a854e0d30 \\\n    --hash=sha256:88bb128f02ba341da8ef447245a9e138fae777f6a23943da4540077d3601eb1c\ndatasets==4.3.0 \\\n    --hash=sha256:0ea157e72138b3ca6c7d2415f19a164ecf7d4c4fa72da2a570da286882e96903 \\\n    --hash=sha256:bc9118ed9afd92346c5be7ed3aaa00177eb907c25467f9d072a0d22777efbd2b\ndetoxify==0.5.2 \\\n    --hash=sha256:c119d0b47545bb076190ff583faad9aa3bb0a90ac01c7c4144758606d21da5b1 \\\n    --hash=sha256:e6135a2d85ad17bfb19d8cf53756a1b499e6d9c03cb1976ddefdc013bf29d32a\ndiffusers==0.39.0 \\\n    --hash=sha256:14bb1d98c85a0e463d734c99aaa73b480a7bc9bad22af30fbf730ef8f09c1d67 \\\n    --hash=sha256:912aca51b5787365110806e984d5555735bf8a461073bb8459029d0bca7870ef\ndill==0.4.0 \\\n    --hash=sha256:0633f1d2df477324f53a895b02c901fb961bdbf65a17122586ea7019292cbcf0 \\\n    --hash=sha256:44f54bf6412c2c8464c14e8243eb163690a9800dbe2c367330883b19c7561049\ndocstring-parser==0.18.0 \\\n    --hash=sha256:292510982205c12b1248696f44959db3cdd1740237a968ea1e2e7a900eeb2015 \\\n    --hash=sha256:b3fcbed555c47d8479be0796ef7e19c2670d428d72e96da63f3a40122860374b\net-xmlfile==2.0.0 \\\n    --hash=sha256:7a91720bc756843502c3b7504c77b8fe44217c85c537d85037f0f536151b2caa \\\n    --hash=sha256:dab3f4764309081ce75662649be815c4c9081e88f0837825f90fd28317d4da54\nfilelock==3.32.4 \\\n    --hash=sha256:22e58ca3b1ae3b98993b762d7338367ae64fe50252bf78d59da3bfebcdf1cedd \\\n    --hash=sha256:2bde2e4cf732e0153406d8a7bc80620ecf5e621fe0d25e41143c4e3b4733ff30\nfonttools==4.63.0 \\\n    --hash=sha256:032038247a96c1690f9f31e377c389383c902531b085aa4e4dabd6f57f870e69 \\\n    --hash=sha256:063e08bd17bd5a90127a14123de0d6a952dbc847695fd98b63c043d58057f90c \\\n    --hash=sha256:0c18358a155d75034911c5ee397a5b44cd19dd325dbb8b35fb60bf421d6a72ac \\\n    --hash=sha256:0eac00b9118c3c2f87d272e45341871c5b3066baa3c86897fa634a7c3fb59096 \\\n    --hash=sha256:1e874792a8212b44583ea02189d9e693906b2f78b261f372f95d6c563210ac1d \\\n    --hash=sha256:22135da48a348785c5e2d5d2d9d6bec5ed44adacbaeb9db12d9493bf6c6bfa68 \\\n    --hash=sha256:22693918177bd9ceabec4736d338045f357769416fc6b0b2508eefef75b08616 \\\n    --hash=sha256:27fdc65af8da6f88b9c6121c47a464cbe359fcfff7ff6fc2d37a1f395d755b78 \\\n    --hash=sha256:2b8ae05d9eacf6081414d759c0a352769ac28ce31280d6bb8e77b03f9e3c449f \\\n    --hash=sha256:2c14b4fd138c4bafcca294765c547914e1aa431ae1ca94ab99d8db08c958bd3b \\\n    --hash=sha256:308f957cdeaf8abe4e5f2f124902ef405448af92c90f80e302a3b771c2e6116b \\\n    --hash=sha256:37dd23e621e3b0aef1baa70a303b80aaf38449632cfc8fd2a55fb285bbccfc02 \\\n    --hash=sha256:445af2eab030a16b9171ea8bdda7ebf7d96bda2df88ee182a464252f6e05e20d \\\n    --hash=sha256:51394295f1a51de8b5f30bdb1e1b9a4231536c7064ef5c6e211eec19fa36036f \\\n    --hash=sha256:58dc6bb86a78d782f00f9190ca02c119cf5bbe2807536e361e18d42019f877d8 \\\n    --hash=sha256:59ac449f8cca9b4ffa08d2e7bbadad87ce710d69d1eda5c3c1ce579baa987272 \\\n    --hash=sha256:6b2248c5decb223562f7902ff6325077a073f608ee8e33e88ad88db734eb9f49 \\\n    --hash=sha256:6d4741eb179121cab9eea4cb2393d24492373a260d7945006358c08cfbf45419 \\\n    --hash=sha256:6db5140a60a5d731d21ec076745b40a310607731b0a565b50776393188649001 \\\n    --hash=sha256:6e528da43bc3791085f8cb6141b1d13e459226790240340fcbb4625649238b03 \\\n    --hash=sha256:796f27556dbe094c4824f75ca85267e4df776c79036c8441469a4df37038c196 \\\n    --hash=sha256:79cdc9f567aec74a72918fd060283911406750cbc9fd28c1316023deb6ce31a9 \\\n    --hash=sha256:7d76edbff9014094dbf03bd2d074709dfa6ec7aba13d838c937a2b33d2d6a86e \\\n    --hash=sha256:7d782fac32985914c351556f68ac0855391572bcd87de50e05970d3cd4c96fc5 \\\n    --hash=sha256:7dd683fef0663e9f0f45cf541d788d24caa3ec9db50796b588e1757d8b3bc007 \\\n    --hash=sha256:85be818f5506e8a7753153def2c9550178f0ecae6a47b5e0e8dbb23f7cc90380 \\\n    --hash=sha256:948428a275741f0b64b113c955425a953314f4b9ab9997f73a72c83e68e569c8 \\\n    --hash=sha256:9ced0bd02ac751dd6319b0da88aaef24414e3b0dbc32bb4f24944821a3741a27 \\\n    --hash=sha256:9e12f105d2b6342c559c298afb674006bb2893afc7102dcf8a1b55b0486b4e40 \\\n    --hash=sha256:a8b33a82979e0a6a34ff435cc81317be1f95ec1ebb7a3a2d1c8a6a54f02ae44e \\\n    --hash=sha256:a9faff9e0c1f76f9fd55899d2ce785832efebab37eb8ae13995853aef178bef0 \\\n    --hash=sha256:af2fd1664d00a397d75f806985ddb36282091c2131a73a6485c23b4a34722263 \\\n    --hash=sha256:afefc1ed0a59785a7fb06ea7e1678e849c193e1e387db783579bc7b3056fcfcb \\\n    --hash=sha256:b1cd75a03ad8cb5bc40c90bfde68c0c47de423aa19e5c0f362b43520645eea94 \\\n    --hash=sha256:ba04cb5891d4c0c21b6da95eda8d7b090021508a294fff33464fc7d241e0856b \\\n    --hash=sha256:bf00f21eb5fb721dbaf73d1e9da6d02a1af7768f2ebcf9798be98beab8ba90f6 \\\n    --hash=sha256:c0425b277a59cff3d80ca42162a8de360f318438a2ac83570842a678d826d579 \\\n    --hash=sha256:c1aaa4b9c75798400ac043ce04d74e7830376c85095a5a6ed7cba2f17a266bf4 \\\n    --hash=sha256:c2a2a42198b696a6f48fad91709afb55176e66a5e566131219dba372fb7f8c59 \\\n    --hash=sha256:caeb583deeb5168e694b65cda8b4ee62abedfa66cf88488734466f2366b9c4e0 \\\n    --hash=sha256:cb014d58140a38135f16064c74c652ed57aa0b75cbf8bb59cac821f7edb5334e \\\n    --hash=sha256:ccf41f2efdf56994d22d73bef4ced1052161958169428d06ba9724ea9e9a64be \\\n    --hash=sha256:cd7e9857e5e63738b9d9fd707bc1f59c8b09e5177726d23664db393c59bb08bd \\\n    --hash=sha256:d76ac49f929aecaf82d83250b8347e099d7aecba0f4726c1d9b6df3b8bb5fe18 \\\n    --hash=sha256:d7e5c9973aa04c95650c96e5f5ad865fbf42d62079163ecfab1e01cbc2504c22 \\\n    --hash=sha256:dcf076a4474fe0d7367e5bbf5b052c7284fa1feca729c04176ce513521afd8a0 \\\n    --hash=sha256:e3297a6a4059b4acc3a1e9a8b04741f240a80044eef08ebd32e8b5bcdddce75b \\\n    --hash=sha256:ee08ebfa58f6e1aeff5697ab9582105bb620008c1caafb681e4c557e7483027b \\\n    --hash=sha256:ef3048ef05dbb552b89817713d9cac912e00d0fde4a3105c00d29e52e10c89af \\\n    --hash=sha256:fd1e3094f42d806d3d7c79162fc59e5910fcbe3a7360c385b8da969bc4493745\nfrozenlist==1.8.0 \\\n    --hash=sha256:0325024fe97f94c41c08872db482cf8ac4800d80e79222c6b0b7b162d5b13686 \\\n    --hash=sha256:032efa2674356903cd0261c4317a561a6850f3ac864a63fc1583147fb05a79b0 \\\n    --hash=sha256:03ae967b4e297f58f8c774c7eabcce57fe3c2434817d4385c50661845a058121 \\\n    --hash=sha256:06be8f67f39c8b1dc671f5d83aaefd3358ae5cdcf8314552c57e7ed3e6475bdd \\\n    --hash=sha256:073f8bf8becba60aa931eb3bc420b217bb7d5b8f4750e6f8b3be7f3da85d38b7 \\\n    --hash=sha256:07cdca25a91a4386d2e76ad992916a85038a9b97561bf7a3fd12d5d9ce31870c \\\n    --hash=sha256:09474e9831bc2b2199fad6da3c14c7b0fbdd377cce9d3d77131be28906cb7d84 \\\n    --hash=sha256:0c18a16eab41e82c295618a77502e17b195883241c563b00f0aa5106fc4eaa0d \\\n    --hash=sha256:0f96534f8bfebc1a394209427d0f8a63d343c9779cda6fc25e8e121b5fd8555b \\\n    --hash=sha256:102e6314ca4da683dca92e3b1355490fed5f313b768500084fbe6371fddfdb79 \\\n    --hash=sha256:11847b53d722050808926e785df837353bd4d75f1d494377e59b23594d834967 \\\n    --hash=sha256:119fb2a1bd47307e899c2fac7f28e85b9a543864df47aa7ec9d3c1b4545f096f \\\n    --hash=sha256:13d23a45c4cebade99340c4165bd90eeb4a56c6d8a9d8aa49568cac19a6d0dc4 \\\n    --hash=sha256:154e55ec0655291b5dd1b8731c637ecdb50975a2ae70c606d100750a540082f7 \\\n    --hash=sha256:168c0969a329b416119507ba30b9ea13688fafffac1b7822802537569a1cb0ef \\\n    --hash=sha256:17c883ab0ab67200b5f964d2b9ed6b00971917d5d8a92df149dc2c9779208ee9 \\\n    --hash=sha256:1a7607e17ad33361677adcd1443edf6f5da0ce5e5377b798fba20fae194825f3 \\\n    --hash=sha256:1a7fa382a4a223773ed64242dbe1c9c326ec09457e6b8428efb4118c685c3dfd \\\n    --hash=sha256:1aa77cb5697069af47472e39612976ed05343ff2e84a3dcf15437b232cbfd087 \\\n    --hash=sha256:1b9290cf81e95e93fdf90548ce9d3c1211cf574b8e3f4b3b7cb0537cf2227068 \\\n    --hash=sha256:20e63c9493d33ee48536600d1a5c95eefc870cd71e7ab037763d1fbb89cc51e7 \\\n    --hash=sha256:21900c48ae04d13d416f0e1e0c4d81f7931f73a9dfa0b7a8746fb2fe7dd970ed \\\n    --hash=sha256:229bf37d2e4acdaf808fd3f06e854a4a7a3661e871b10dc1f8f1896a3b05f18b \\\n    --hash=sha256:2552f44204b744fba866e573be4c1f9048d6a324dfe14475103fd51613eb1d1f \\\n    --hash=sha256:27c6e8077956cf73eadd514be8fb04d77fc946a7fe9f7fe167648b0b9085cc25 \\\n    --hash=sha256:28bd570e8e189d7f7b001966435f9dac6718324b5be2990ac496cf1ea9ddb7fe \\\n    --hash=sha256:294e487f9ec720bd8ffcebc99d575f7eff3568a08a253d1ee1a0378754b74143 \\\n    --hash=sha256:29548f9b5b5e3460ce7378144c3010363d8035cea44bc0bf02d57f5a685e084e \\\n    --hash=sha256:2c5dcbbc55383e5883246d11fd179782a9d07a986c40f49abe89ddf865913930 \\\n    --hash=sha256:2dc43a022e555de94c3b68a4ef0b11c4f747d12c024a520c7101709a2144fb37 \\\n    --hash=sha256:2f05983daecab868a31e1da44462873306d3cbfd76d1f0b5b69c473d21dbb128 \\\n    --hash=sha256:33139dc858c580ea50e7e60a1b0ea003efa1fd42e6ec7fdbad78fff65fad2fd2 \\\n    --hash=sha256:332db6b2563333c5671fecacd085141b5800cb866be16d5e3eb15a2086476675 \\\n    --hash=sha256:33f48f51a446114bc5d251fb2954ab0164d5be02ad3382abcbfe07e2531d650f \\\n    --hash=sha256:34187385b08f866104f0c0617404c8eb08165ab1272e884abc89c112e9c00746 \\\n    --hash=sha256:342c97bf697ac5480c0a7ec73cd700ecfa5a8a40ac923bd035484616efecc2df \\\n    --hash=sha256:3462dd9475af2025c31cc61be6652dfa25cbfb56cbbf52f4ccfe029f38decaf8 \\\n    --hash=sha256:39ecbc32f1390387d2aa4f5a995e465e9e2f79ba3adcac92d68e3e0afae6657c \\\n    --hash=sha256:3e0761f4d1a44f1d1a47996511752cf3dcec5bbdd9cc2b4fe595caf97754b7a0 \\\n    --hash=sha256:3ede829ed8d842f6cd48fc7081d7a41001a56f1f38603f9d49bf3020d59a31ad \\\n    --hash=sha256:3ef2d026f16a2b1866e1d86fc4e1291e1ed8a387b2c333809419a2f8b3a77b82 \\\n    --hash=sha256:405e8fe955c2280ce66428b3ca55e12b3c4e9c336fb2103a4937e891c69a4a29 \\\n    --hash=sha256:42145cd2748ca39f32801dad54aeea10039da6f86e303659db90db1c4b614c8c \\\n    --hash=sha256:4314debad13beb564b708b4a496020e5306c7333fa9a3ab90374169a20ffab30 \\\n    --hash=sha256:433403ae80709741ce34038da08511d4a77062aa924baf411ef73d1146e74faf \\\n    --hash=sha256:44389d135b3ff43ba8cc89ff7f51f5a0bb6b63d829c8300f79a2fe4fe61bcc62 \\\n    --hash=sha256:48e6d3f4ec5c7273dfe83ff27c91083c6c9065af655dc2684d2c200c94308bb5 \\\n    --hash=sha256:494a5952b1c597ba44e0e78113a7266e656b9794eec897b19ead706bd7074383 \\\n    --hash=sha256:4970ece02dbc8c3a92fcc5228e36a3e933a01a999f7094ff7c23fbd2beeaa67c \\\n    --hash=sha256:4e0c11f2cc6717e0a741f84a527c52616140741cd812a50422f83dc31749fb52 \\\n    --hash=sha256:50066c3997d0091c411a66e710f4e11752251e6d2d73d70d8d5d4c76442a199d \\\n    --hash=sha256:517279f58009d0b1f2e7c1b130b377a349405da3f7621ed6bfae50b10adf20c1 \\\n    --hash=sha256:54b2077180eb7f83dd52c40b2750d0a9f175e06a42e3213ce047219de902717a \\\n    --hash=sha256:5500ef82073f599ac84d888e3a8c1f77ac831183244bfd7f11eaa0289fb30714 \\\n    --hash=sha256:581ef5194c48035a7de2aefc72ac6539823bb71508189e5de01d60c9dcd5fa65 \\\n    --hash=sha256:59a6a5876ca59d1b63af8cd5e7ffffb024c3dc1e9cf9301b21a2e76286505c95 \\\n    --hash=sha256:5a3a935c3a4e89c733303a2d5a7c257ea44af3a56c8202df486b7f5de40f37e1 \\\n    --hash=sha256:5c1c8e78426e59b3f8005e9b19f6ff46e5845895adbde20ece9218319eca6506 \\\n    --hash=sha256:5d63a068f978fc69421fb0e6eb91a9603187527c86b7cd3f534a5b77a592b888 \\\n    --hash=sha256:667c3777ca571e5dbeb76f331562ff98b957431df140b54c85fd4d52eea8d8f6 \\\n    --hash=sha256:6da155091429aeba16851ecb10a9104a108bcd32f6c1642867eadaee401c1c41 \\\n    --hash=sha256:6dc4126390929823e2d2d9dc79ab4046ed74680360fc5f38b585c12c66cdf459 \\\n    --hash=sha256:7398c222d1d405e796970320036b1b563892b65809d9e5261487bb2c7f7b5c6a \\\n    --hash=sha256:74c51543498289c0c43656701be6b077f4b265868fa7f8a8859c197006efb608 \\\n    --hash=sha256:776f352e8329135506a1d6bf16ac3f87bc25b28e765949282dcc627af36123aa \\\n    --hash=sha256:778a11b15673f6f1df23d9586f83c4846c471a8af693a22e066508b77d201ec8 \\\n    --hash=sha256:78f7b9e5d6f2fdb88cdde9440dc147259b62b9d3b019924def9f6478be254ac1 \\\n    --hash=sha256:799345ab092bee59f01a915620b5d014698547afd011e691a208637312db9186 \\\n    --hash=sha256:7bf6cdf8e07c8151fba6fe85735441240ec7f619f935a5205953d58009aef8c6 \\\n    --hash=sha256:8009897cdef112072f93a0efdce29cd819e717fd2f649ee3016efd3cd885a7ed \\\n    --hash=sha256:80f85f0a7cc86e7a54c46d99c9e1318ff01f4687c172ede30fd52d19d1da1c8e \\\n    --hash=sha256:8585e3bb2cdea02fc88ffa245069c36555557ad3609e83be0ec71f54fd4abb52 \\\n    --hash=sha256:878be833caa6a3821caf85eb39c5ba92d28e85df26d57afb06b35b2efd937231 \\\n    --hash=sha256:8a76ea0f0b9dfa06f254ee06053d93a600865b3274358ca48a352ce4f0798450 \\\n    --hash=sha256:8b7b94a067d1c504ee0b16def57ad5738701e4ba10cec90529f13fa03c833496 \\\n    --hash=sha256:8d92f1a84bb12d9e56f818b3a746f3efba93c1b63c8387a73dde655e1e42282a \\\n    --hash=sha256:908bd3f6439f2fef9e85031b59fd4f1297af54415fb60e4254a95f75b3cab3f3 \\\n    --hash=sha256:92db2bf818d5cc8d9c1f1fc56b897662e24ea5adb36ad1f1d82875bd64e03c24 \\\n    --hash=sha256:940d4a017dbfed9daf46a3b086e1d2167e7012ee297fef9e1c545c4d022f5178 \\\n    --hash=sha256:957e7c38f250991e48a9a73e6423db1bb9dd14e722a10f6b8bb8e16a0f55f695 \\\n    --hash=sha256:96153e77a591c8adc2ee805756c61f59fef4cf4073a9275ee86fe8cba41241f7 \\\n    --hash=sha256:96f423a119f4777a4a056b66ce11527366a8bb92f54e541ade21f2374433f6d4 \\\n    --hash=sha256:97260ff46b207a82a7567b581ab4190bd4dfa09f4db8a8b49d1a958f6aa4940e \\\n    --hash=sha256:974b28cf63cc99dfb2188d8d222bc6843656188164848c4f679e63dae4b0708e \\\n    --hash=sha256:9ff15928d62a0b80bb875655c39bf517938c7d589554cbd2669be42d97c2cb61 \\\n    --hash=sha256:a6483e309ca809f1efd154b4d37dc6d9f61037d6c6a81c2dc7a15cb22c8c5dca \\\n    --hash=sha256:a88f062f072d1589b7b46e951698950e7da00442fc1cacbe17e19e025dc327ad \\\n    --hash=sha256:ac913f8403b36a2c8610bbfd25b8013488533e71e62b4b4adce9c86c8cea905b \\\n    --hash=sha256:adbeebaebae3526afc3c96fad434367cafbfd1b25d72369a9e5858453b1bb71a \\\n    --hash=sha256:b2a095d45c5d46e5e79ba1e5b9cb787f541a8dee0433836cea4b96a2c439dcd8 \\\n    --hash=sha256:b3210649ee28062ea6099cfda39e147fa1bc039583c8ee4481cb7811e2448c51 \\\n    --hash=sha256:b37f6d31b3dcea7deb5e9696e529a6aa4a898adc33db82da12e4c60a7c4d2011 \\\n    --hash=sha256:b4dec9482a65c54a5044486847b8a66bf10c9cb4926d42927ec4e8fd5db7fed8 \\\n    --hash=sha256:b4f3b365f31c6cd4af24545ca0a244a53688cad8834e32f56831c4923b50a103 \\\n    --hash=sha256:b6db2185db9be0a04fecf2f241c70b63b1a242e2805be291855078f2b404dd6b \\\n    --hash=sha256:b9be22a69a014bc47e78072d0ecae716f5eb56c15238acca0f43d6eb8e4a5bda \\\n    --hash=sha256:bac9c42ba2ac65ddc115d930c78d24ab8d4f465fd3fc473cdedfccadb9429806 \\\n    --hash=sha256:bf0a7e10b077bf5fb9380ad3ae8ce20ef919a6ad93b4552896419ac7e1d8e042 \\\n    --hash=sha256:c23c3ff005322a6e16f71bf8692fcf4d5a304aaafe1e262c98c6d4adc7be863e \\\n    --hash=sha256:c4c800524c9cd9bac5166cd6f55285957fcfc907db323e193f2afcd4d9abd69b \\\n    --hash=sha256:c7366fe1418a6133d5aa824ee53d406550110984de7637d65a178010f759c6ef \\\n    --hash=sha256:c8d1634419f39ea6f5c427ea2f90ca85126b54b50837f31497f3bf38266e853d \\\n    --hash=sha256:c9a63152fe95756b85f31186bddf42e4c02c6321207fd6601a1c89ebac4fe567 \\\n    --hash=sha256:cb89a7f2de3602cfed448095bab3f178399646ab7c61454315089787df07733a \\\n    --hash=sha256:cba69cb73723c3f329622e34bdbf5ce1f80c21c290ff04256cff1cd3c2036ed2 \\\n    --hash=sha256:cee686f1f4cadeb2136007ddedd0aaf928ab95216e7691c63e50a8ec066336d0 \\\n    --hash=sha256:cf253e0e1c3ceb4aaff6df637ce033ff6535fb8c70a764a8f46aafd3d6ab798e \\\n    --hash=sha256:d1eaff1d00c7751b7c6662e9c5ba6eb2c17a2306ba5e2a37f24ddf3cc953402b \\\n    --hash=sha256:d3bb933317c52d7ea5004a1c442eef86f426886fba134ef8cf4226ea6ee1821d \\\n    --hash=sha256:d4d3214a0f8394edfa3e303136d0575eece0745ff2b47bd2cb2e66dd92d4351a \\\n    --hash=sha256:d6a5df73acd3399d893dafc71663ad22534b5aa4f94e8a2fabfe856c3c1b6a52 \\\n    --hash=sha256:d8b7138e5cd0647e4523d6685b0eac5d4be9a184ae9634492f25c6eb38c12a47 \\\n    --hash=sha256:db1e72ede2d0d7ccb213f218df6a078a9c09a7de257c2fe8fcef16d5925230b1 \\\n    --hash=sha256:e25ac20a2ef37e91c1b39938b591457666a0fa835c7783c3a8f33ea42870db94 \\\n    --hash=sha256:e2de870d16a7a53901e41b64ffdf26f2fbb8917b3e6ebf398098d72c5b20bd7f \\\n    --hash=sha256:e4a3408834f65da56c83528fb52ce7911484f0d1eaf7b761fc66001db1646eff \\\n    --hash=sha256:eaa352d7047a31d87dafcacbabe89df0aa506abb5b1b85a2fb91bc3faa02d822 \\\n    --hash=sha256:eab8145831a0d56ec9c4139b6c3e594c7a83c2c8be25d5bcf2d86136a532287a \\\n    --hash=sha256:ec3cc8c5d4084591b4237c0a272cc4f50a5b03396a47d9caaf76f5d7b38a4f11 \\\n    --hash=sha256:edee74874ce20a373d62dc28b0b18b93f645633c2943fd90ee9d898550770581 \\\n    --hash=sha256:eefdba20de0d938cec6a89bd4d70f346a03108a19b9df4248d3cf0d88f1b0f51 \\\n    --hash=sha256:ef2b7b394f208233e471abc541cc6991f907ffd47dc72584acee3147899d6565 \\\n    --hash=sha256:f21f00a91358803399890ab167098c131ec2ddd5f8f5fd5fe9c9f2c6fcd91e40 \\\n    --hash=sha256:f4be2e3d8bc8aabd566f8d5b8ba7ecc09249d74ba3c9ed52e54dc23a293f0b92 \\\n    --hash=sha256:f57fb59d9f385710aa7060e89410aeb5058b99e62f4d16b08b91986b9a2140c2 \\\n    --hash=sha256:f6292f1de555ffcc675941d65fffffb0a5bcd992905015f85d0592201793e0e5 \\\n    --hash=sha256:f833670942247a14eafbb675458b4e61c82e002a148f49e68257b79296e865c4 \\\n    --hash=sha256:fa47e444b8ba08fffd1c18e8cdb9a75db1b6a27f17507522834ad13ed5922b93 \\\n    --hash=sha256:fb30f9626572a76dfe4293c7194a09fb1fe93ba94c7d4f720dfae3b646b45027 \\\n    --hash=sha256:fe3c58d2f5db5fbd18c2987cba06d51b0529f52bc3a6cdc33d3f4eab725104bd\nfsspec==2025.9.0 \\\n    --hash=sha256:19fd429483d25d28b65ec68f9f4adc16c17ea2c7c7bf54ec61360d478fb19c19 \\\n    --hash=sha256:530dc2a2af60a414a832059574df4a6e10cce927f6f4a78209390fe38955cfb7\ngrpcio==1.83.0 \\\n    --hash=sha256:009667eaf3dcd5224c713589cdc98e7ca4ed0ff0b61132c6b276e930eb83a2df \\\n    --hash=sha256:10b3fa0475eb572c9a81a6fe37fa16a9c500c0c91cfc148cac15692b7e3c2867 \\\n    --hash=sha256:1aa567f8c3f19850ffd5d2858c9a8ea7c80f0db6c01186b71eb31e923ec984f5 \\\n    --hash=sha256:1c699bbb20f143c8f2bff219de578aa2dc1f919399d67dc702b038b986ee62df \\\n    --hash=sha256:28f6c35ac8fcf10e4594f138e468f194360089dde40d126a7033e863fc479930 \\\n    --hash=sha256:2b5e75c34842cd9c1b95285ca395c6a569664b81e3ffa6b714125922942abaaf \\\n    --hash=sha256:2bb48cb5e6dd005ca12b89ce4b6ac0b48ff3112c747542ee7986ef611a8ca6d9 \\\n    --hash=sha256:32e11c37f5285b0c6fa3042c05fe06903696689749833fc64e67dec71b9bbe33 \\\n    --hash=sha256:33898e6a28e4ae598f1577cb1c4fec2a15c033d0ec52b9b45a09610dd045b9da \\\n    --hash=sha256:35a5b1c192496b6c25956eebfa963468935612206fd2543ac3ce981e6a5e0f03 \\\n    --hash=sha256:3f351629f6ae16ecc0ec3553e586a6763ffd9f6114044286d0cbec3e09241bfa \\\n    --hash=sha256:4772402f43517b4824980be4b3b2274a81eec0004a70009473c31b340d43e223 \\\n    --hash=sha256:4e3eedfc92b6b9f2960115e7e620cf0cbf80bb7849a51ce3820dc54dfd88b6b9 \\\n    --hash=sha256:4fcaa7c45c45b4a89e2867d1f1785d9481a788399d915e341ed2eb49aeef9dd4 \\\n    --hash=sha256:5882c1a721b50ce0123ee5e839e1ab059ad72a7ade76cdf2d5bd833b56791acf \\\n    --hash=sha256:5f20a988480b0f28207f057f7f7ae1313393c3cef0adcfeae8248f9947eaf881 \\\n    --hash=sha256:61007cd08640abc5c54547ee32505474c482cd733a53cb87551ea81faa6350af \\\n    --hash=sha256:62003babc444a606dcd1f009cd16391ce23669ae4ad6ec267a873da7937a69f5 \\\n    --hash=sha256:6662f3b1e07cc7493d437351860dc867bddc6a93c83ecf33bbfdaf0c217ab2d0 \\\n    --hash=sha256:6755ed67cc3e454d51ae9f6e1915b80d3942fa4de956ef48dacd45ab7f40b727 \\\n    --hash=sha256:6b6c666a1d5613ff360c9e90f44665e3a88b25a815209ddbc0917eec281931cb \\\n    --hash=sha256:6be5c807b717be3dd649446f021301fd7907e376318675d2147823071034112a \\\n    --hash=sha256:6e01ecd9d8ef280abe1365138a4dc318f9a5287f4cb1b41d07816f796653f735 \\\n    --hash=sha256:6fb8a1dd0c6f0f931e69e9d0dc6d1c406ed2a44fa963414eafba07b7fb685d16 \\\n    --hash=sha256:7416952ca770477990257206276999056f8316d79196f2f25942393e58a20b49 \\\n    --hash=sha256:74fe6f9e8a35c7dbf32255ee154d15e3e5338a81ed39173d079d594d2e544cd1 \\\n    --hash=sha256:7674587248fbbb2ac6e4eecf83a8a0f3d91a928f941de571acfd3a2f007fbc24 \\\n    --hash=sha256:7936f2a56cf04f6514705c0fedf400971de01b6aa1719327e4718f410a765e2b \\\n    --hash=sha256:7bd82671b39065ba18cd536e9cd45b27ff649053f81ddd2c6a966d595067080f \\\n    --hash=sha256:8f6c395e493d20c39b29392ca200e9aaeb78d0bc2f04db0c0a7da7ddc939aa57 \\\n    --hash=sha256:8fe04f1050a59f875601eb55d42b4f66946fe89817f967e34db1462ccd07dadf \\\n    --hash=sha256:8ff0b8767ddd62704e0d9571c1890af08d84a3a689ebba1807e62519d0b3277f \\\n    --hash=sha256:a21cb4eeeba124443f399be2e8b624943cde864dcbe588cb42e5c483a52a906c \\\n    --hash=sha256:aa074041231f03959cb097dd5517b0677b8ea49215bae01d5710a7b69dd59969 \\\n    --hash=sha256:aeb339838db07600481ef869507279b75326c75eac6d10f7afa62a0da1d2bcdd \\\n    --hash=sha256:b0a0be840e51b6b7ee9df9269770faf77bdf4b771053c257c21d12bad607714c \\\n    --hash=sha256:bb669918fd88936b15599caff4160a77ab74bdeb25f2231f6e45b61282d6107b \\\n    --hash=sha256:bc60215b5cb9fc8ca72942c498b551ac2305bd08f6ef8d4e3f0d21b64fbecd61 \\\n    --hash=sha256:c19b454d3d3f28db81f2c7c4dbaee96e7f6fd149721733ffe79d6bc530f17404 \\\n    --hash=sha256:c6444666317338e903093c7c756e6cc88eee59f798cb8dd41e87725bf54e1617 \\\n    --hash=sha256:c834e86d8fd2f03d7e4db49a027f7c5b89c5b88eed305543a5295bd6fee61e40 \\\n    --hash=sha256:cb056f6e171c42639a50460b2929c82241fda51f71cf3dcdd68090fe45095a45 \\\n    --hash=sha256:cb2906c61db4f9c64cc360054b5df70eeb81846228e9e56a4944bd415a63dadc \\\n    --hash=sha256:d05ff664100d429335b93c91b8b34ddf9e94a112205e7fa06dede309e44a4e4c \\\n    --hash=sha256:ee94a4016fdf8699fb1fd8a38652475ff677f1c72074cee44deeeb9a7e95e745 \\\n    --hash=sha256:f1c3e5689d4b90987b1d72022bcfe866a9a3dc66197484cf856d96b6150e7f45 \\\n    --hash=sha256:f47d62808b4c0a97b78bff88a6d4ca283a2a492b9a04a87d814af95ca3b9c19c \\\n    --hash=sha256:f4cee5fc86e84a0cf7ad1574b454c3320e087c07f55b7df5dc0ac6a873fb90c0 \\\n    --hash=sha256:f5e822a7e7d03282f6ad225e710493c48b9057a353358344a5f7c42b2b37618d \\\n    --hash=sha256:f5f410d7c2903eabb34789dfd6342eef04af1ad459943936b7e09a9f5bd417b9 \\\n    --hash=sha256:fba099b716e73512d61b97f71ea3c31a72abb36904036e316bf4dd148ca8dcc8\nh11==0.16.0 \\\n    --hash=sha256:4e35b956cf45792e4caa5885e69fba00bdbc6ffafbfa020300e549b208ee5ff1 \\\n    --hash=sha256:63cf8bbe7522de3bf65932fda1d9c2772064ffb3dae62d55932da54b31cb6c86\nh2==4.4.1 \\\n    --hash=sha256:0e25f1462b23c9cb82d9eb02e28bc706dac2a68cb457c6a0d74d63c8a2a5d0e6 \\\n    --hash=sha256:4e866ffb1a869ae14dd9b5e6beb5c24a13da0495ad72b65925ded182521c1516\nhf-transfer==0.1.9 \\\n    --hash=sha256:035572865dab29d17e783fbf1e84cf1cb24f3fcf8f1b17db1cfc7fdf139f02bf \\\n    --hash=sha256:0d991376f0eac70a60f0cbc95602aa708a6f7c8617f28b4945c1431d67b8e3c8 \\\n    --hash=sha256:16f208fc678911c37e11aa7b586bc66a37d02e636208f18b6bc53d29b5df40ad \\\n    --hash=sha256:1a6bd16c667ebe89a069ca163060127a794fa3a3525292c900b8c8cc47985b0d \\\n    --hash=sha256:2c7fc1b85f4d0f76e452765d7648c9f4bfd0aedb9ced2ae1ebfece2d8cfaf8e2 \\\n    --hash=sha256:3a736dfbb2c84f5a2c975478ad200c0c8bfcb58a25a35db402678fb87ce17fa4 \\\n    --hash=sha256:3ebc4ab9023414880c8b1d3c38174d1c9989eb5022d37e814fa91a3060123eb0 \\\n    --hash=sha256:435cc3cdc8524ce57b074032b8fd76eed70a4224d2091232fa6a8cef8fd6803e \\\n    --hash=sha256:504b8427fd785dd8546d53b9fafe6e436bd7a3adf76b9dce556507650a7b4567 \\\n    --hash=sha256:57fd9880da1ee0f47250f735f791fab788f0aa1ee36afc49f761349869c8b4d9 \\\n    --hash=sha256:5828057e313de59300dd1abb489444bc452efe3f479d3c55b31a8f680936ba42 \\\n    --hash=sha256:5d561f0520f493c66b016d99ceabe69c23289aa90be38dd802d2aef279f15751 \\\n    --hash=sha256:6e94e8822da79573c9b6ae4d6b2f847c59a7a06c5327d7db20751b68538dc4f6 \\\n    --hash=sha256:8669dbcc7a3e2e8d61d42cd24da9c50d57770bd74b445c65123291ca842a7e7a \\\n    --hash=sha256:8674026f21ed369aa2a0a4b46000aca850fc44cd2b54af33a172ce5325b4fc82 \\\n    --hash=sha256:89a23f58b7b7effbc047b8ca286f131b17728c99a9f972723323003ffd1bb916 \\\n    --hash=sha256:8fd0167c4407a3bc4cdd0307e65ada2294ec04f1813d8a69a5243e379b22e9d8 \\\n    --hash=sha256:a5b366d34cd449fe9b20ef25941e6eef0460a2f74e7389f02e673e1f88ebd538 \\\n    --hash=sha256:cdca9bfb89e6f8f281890cc61a8aff2d3cecaff7e1a4d275574d96ca70098557 \\\n    --hash=sha256:d2fde99d502093ade3ab1b53f80da18480e9902aa960dab7f74fb1b9e5bc5746 \\\n    --hash=sha256:dc7fff1345980d6c0ebb92c811d24afa4b98b3e07ed070c8e38cc91fd80478c5 \\\n    --hash=sha256:e66acf91df4a8b72f60223059df3003062a5ae111757187ed1a06750a30e911b \\\n    --hash=sha256:e6ac4eddcd99575ed3735ed911ddf9d1697e2bd13aa3f0ad7e3904dd4863842e \\\n    --hash=sha256:ee8b10afedcb75f71091bcc197c526a6ebf5c58bbbadb34fdeee6160f55f619f \\\n    --hash=sha256:fc6bd19e1cc177c66bdef15ef8636ad3bde79d5a4f608c158021153b4573509d\nhf-xet==1.6.0 \\\n    --hash=sha256:0e6e21fa3cdfcdcd76748564bf593870a5e013f47d97cf10aed63aa222cff5b7 \\\n    --hash=sha256:23379c2f9ec8696d952b16414a2bae72cad86a52df869b050698ba60f538c675 \\\n    --hash=sha256:2e58454a340b3556dfa4972d5451aff4fba8dd42a236600ba1a1d2b1514f0fef \\\n    --hash=sha256:35cec30d75c6f9eb9c16a77cef68e85a103b72e24d4b473714ec9ff06428bab9 \\\n    --hash=sha256:3dc3e35441ba395006af5aaacc40ef2e603c51ef46c3530b9156185f00935ea3 \\\n    --hash=sha256:4fc74352a17015bd0ee90038bc9efe38db894cde45f268b6712b04fce8cd0acb \\\n    --hash=sha256:5153e6bb103ad49d6ea9f1b2e230db5a2ea32551ad09a706d2f61d7c7c80d80e \\\n    --hash=sha256:5789835d7c6bc9436962853192082374297fb72d7eff7e7762ec25ceb7e25338 \\\n    --hash=sha256:633dc0cd71d32da58ab8c03ad38e2fac452c15c2b0a2866ebf6ededfe0a5061d \\\n    --hash=sha256:70cbb9c896901600128cb9b6f06e132954fbede1db30f31f7c6c63f84cb7c31d \\\n    --hash=sha256:75765820ce4700db3750c94acc8fe27c5fae4c9ec000a0dbac3ca082acf97765 \\\n    --hash=sha256:8fb4f71cba6129110c3374a33f919001ff130488fc23553698e34cc1c2a1198c \\\n    --hash=sha256:948f15d3a9545cfe5932f6bd8b440f6ae630aee108f14b7bd6c561f7c2dcc522 \\\n    --hash=sha256:d62671bb130879cef0ee4c9ebe47a14af6c66ec53e6d84dc15936e5ffdfac82f \\\n    --hash=sha256:f0906082d9932ae0c0057fa194041c22b4e2cdb46b2592ef3b91f020d62a081a \\\n    --hash=sha256:f2f7278c05c22fd60cb436cda1269649b3e81db65ecdc8496e5e164aa4143e7b \\\n    --hash=sha256:fb4fadde1b2b70bf4c0c14a6dccbe7194b1c28947fefd5bbe3fed9d940676c3b\nhpack==4.2.0 \\\n    --hash=sha256:0895cfa3b5531fc65fe439c05eb65144f123bf7a394fcaa56aa423548d8e45c0 \\\n    --hash=sha256:858ac0b02280fa582b5080d68db0899c62a80375e0e5413a74970c5e518b6986\nhttpcore==1.0.9 \\\n    --hash=sha256:2d400746a40668fc9dec9810239072b40b4484b640a8c38fd654a024c7a1bf55 \\\n    --hash=sha256:6e34463af53fd2ab5d807f399a9b45ea31c3dfa2276f15a2c3f00afff6e176e8\nhttpx==0.28.1 \\\n    --hash=sha256:75e98c5f16b0f35b567856f597f06ff2270a374470a5c2392242528e3e3e42fc \\\n    --hash=sha256:d909fcccc110f8c7faf814ca82a9a4d816bc5a6dbfea25d6591d6985b8ba59ad\nhuggingface-hub==1.3.2 \\\n    --hash=sha256:15d7902e154f04174a0816d1e9594adcf15cdad57596920a5dc70fadb5d896c7 \\\n    --hash=sha256:b552b9562a5532102a041fa31a6966bb9de95138fc7aa578bb3703198c25d1b6\nhyperframe==6.1.0 \\\n    --hash=sha256:b03380493a519fce58ea5af42e4a42317bf9bd425596f7a0835ffce80f1a42e5 \\\n    --hash=sha256:f630908a00854a7adeabd6382b43923a4c4cd4b821fcb527e6ab9e15382a3b08\nidna==3.19 \\\n    --hash=sha256:5e0811a4383b21dc5838069f801c4fb62113b7447663d2530d2bd6e77b49bf15 \\\n    --hash=sha256:815e7be7a7806d54abb586dc943addc79e8b2ee16915059658cbeff4b1b43bf4\nimportlib-metadata==9.0.0 \\\n    --hash=sha256:2d21d1cc5a017bd0559e36150c21c830ab1dc304dedd1b7ea85d20f45ef3edd7 \\\n    --hash=sha256:a4f57ab599e6a2e3016d7595cfd72eb4661a5106e787a95bcc90c7105b831efc\ninteregular==0.3.3 \\\n    --hash=sha256:b0c07007d48c89d6d19f7204972d369b2a77222722e126b6aa63aa721dc3b19c \\\n    --hash=sha256:d9b697b21b34884711399ba0f0376914b81899ce670032486d0d048344a76600\njinja2==3.1.6 \\\n    --hash=sha256:0137fb05990d35f1275a587e9aee6d56da821fc83491a0fb838183be43f66d6d \\\n    --hash=sha256:85ece4451f492d0c13c5dd7c13a64681a86afae63a5f347908daf103ce6d2f67\njoblib==1.5.3 \\\n    --hash=sha256:5fc3c5039fc5ca8c0276333a188bbd59d6b7ab37fe6632daa76bc7f9ec18e713 \\\n    --hash=sha256:8561a3269e6801106863fd0d6d84bb737be9e7631e33aaed3fb9ce5953688da3\nkiwisolver==1.5.0 \\\n    --hash=sha256:012b1eb16e28718fa782b5e61dc6f2da1f0792ca73bd05d54de6cb9561665fc9 \\\n    --hash=sha256:01808c6d15f4c3e8559595d6d1fe6411c68e4a3822b4b9972b44473b24f4e679 \\\n    --hash=sha256:0255a027391d52944eae1dbb5d4cc5903f57092f3674e8e544cdd2622826b3f0 \\\n    --hash=sha256:0b85aad90cea8ac6797a53b5d5f2e967334fa4d1149f031c4537569972596cb8 \\\n    --hash=sha256:0bf3acf1419fa93064a4c2189ac0b58e3be7872bf6ee6177b0d4c63dc4cea276 \\\n    --hash=sha256:0c50b89ffd3e1a911c69a1dd3de7173c0cd10b130f56222e57898683841e4f96 \\\n    --hash=sha256:0cbe94b69b819209a62cb27bdfa5dc2a8977d8de2f89dfd97ba4f53ed3af754e \\\n    --hash=sha256:0df54df7e686afa55e6f21fb86195224a6d9beb71d637e8d7920c95cf0f89aac \\\n    --hash=sha256:0e3aafb33aed7479377e5e9a82e9d4bf87063741fc99fc7ae48b0f16e32bdd6f \\\n    --hash=sha256:12e91c215a96e39f57989c8912ae761286ac5a9584d04030ceb3368a357f017a \\\n    --hash=sha256:1465387ac63576c3e125e5337a6892b9e99e0627d52317f3ca79e6930d889d15 \\\n    --hash=sha256:16b85d37c2cbb3253226d26e64663f755d88a03439a9c47df6246b35defbdfb7 \\\n    --hash=sha256:1b0feb50971481a2cc44d94e88bdb02cdd497618252ae226b8eb1201b957e368 \\\n    --hash=sha256:1d49a49ac4cbfb7c1375301cd1ec90169dfeae55ff84710d782260ce77a75a02 \\\n    --hash=sha256:1d9daea4ea6b9be74fe2f01f7fbade8d6ffab263e781274cffca0dba9be9eec9 \\\n    --hash=sha256:1dd9b0b119a350976a6d781e7278ec7aca0b201e1a9e2d23d9804afecb6ca681 \\\n    --hash=sha256:1f1489f769582498610e015a8ef2d36f28f505ab3096d0e16b4858a9ec214f57 \\\n    --hash=sha256:2517e24d7315eb51c10664cdb865195df38ab74456c677df67bb47f12d088a27 \\\n    --hash=sha256:295d9ffe712caa9f8a3081de8d32fc60191b4b51c76f02f951fd8407253528f4 \\\n    --hash=sha256:2a075bd7bd19c70cf67c8badfa36cf7c5d8de3c9ddb8420c51e10d9c50e94920 \\\n    --hash=sha256:32cc0a5365239a6ea0c6ed461e8838d053b57e397443c0ca894dcc8e388d4374 \\\n    --hash=sha256:332b4f0145c30b5f5ad9374881133e5aa64320428a57c2c2b61e9d891a51c2f3 \\\n    --hash=sha256:377815a8616074cabbf3f53354e1d040c35815a134e01d7614b7692e4bf8acfa \\\n    --hash=sha256:38f4a703656f493b0ad185211ccfca7f0386120f022066b018eb5296d8613e23 \\\n    --hash=sha256:3ac2360e93cb41be81121755c6462cff3beaa9967188c866e5fce5cf13170859 \\\n    --hash=sha256:3c4923e404d6bcd91b6779c009542e5647fef32e4a5d75e115e3bbac6f2335eb \\\n    --hash=sha256:3cdcb35dc9d807259c981a85531048ede628eabcffb3239adf3d17463518992d \\\n    --hash=sha256:41024ed50e44ab1a60d3fe0a9d15a4ccc9f5f2b1d814ff283c8d01134d5b81bc \\\n    --hash=sha256:413b820229730d358efd838ecbab79902fe97094565fdc80ddb6b0a18c18a581 \\\n    --hash=sha256:4432b835675f0ea7414aab3d37d119f7226d24869b7a829caeab49ebda407b0c \\\n    --hash=sha256:4db576bb8c3ef9365f8b40fe0f671644de6736ae2c27a2c62d7d8a1b4329f099 \\\n    --hash=sha256:4e7f886f47ab881692f278ae901039a234e4025a68e6dfab514263a0b1c4ae05 \\\n    --hash=sha256:4e9750bc21b886308024f8a54ccb9a2cc38ac9fa813bf4348434e3d54f337ff9 \\\n    --hash=sha256:5060731cc3ed12ca3a8b57acd4aeca5bbc2f49216dd0bec1650a1acd89486bcd \\\n    --hash=sha256:50847dca5d197fcbd389c805aa1a1cf32f25d2e7273dc47ab181a517666b68cc \\\n    --hash=sha256:5092eb5b1172947f57d6ea7d89b2f29650414e4293c47707eb499ec07a0ac796 \\\n    --hash=sha256:5124d1ea754509b09e53738ec185584cc609aae4a3b510aaf4ed6aa047ef9303 \\\n    --hash=sha256:51e8c4084897de9f05898c2c2a39af6318044ae969d46ff7a34ed3f96274adca \\\n    --hash=sha256:530a3fd64c87cffa844d4b6b9768774763d9caa299e9b75d8eca6a4423b31314 \\\n    --hash=sha256:56fa888f10d0f367155e76ce849fa1166fc9730d13bd2d65a2aa13b6f5424489 \\\n    --hash=sha256:58f812017cd2985c21fbffb4864d59174d4903dd66fa23815e74bbc7a0e2dd57 \\\n    --hash=sha256:59cd8683f575d96df5bb48f6add94afc055012c29e28124fcae2b63661b9efb1 \\\n    --hash=sha256:5ae8e62c147495b01a0f4765c878e9bfdf843412446a247e28df59936e99e797 \\\n    --hash=sha256:5b233ea3e165e43e35dba1d2b8ecc21cf070b45b65ae17dd2747d2713d942021 \\\n    --hash=sha256:6176c1811d9d5a04fa391c490cc44f451e240697a16977f11c6f722efb9041db \\\n    --hash=sha256:62f59da443c4f4849f73a51a193b1d9d258dcad0c41bc4d1b8fb2bcc04bfeb22 \\\n    --hash=sha256:6783e069732715ad0c3ce96dbf21dbc2235ab0593f2baf6338101f70371f4028 \\\n    --hash=sha256:6ab8ba9152203feec73758dad83af9a0bbe05001eb4639e547207c40cfb52083 \\\n    --hash=sha256:70d593af6a6ca332d1df73d519fddb5148edb15cd90d5f0155e3746a6d4fcc65 \\\n    --hash=sha256:72ec46b7eba5b395e0a7b63025490d3214c11013f4aacb4f5e8d6c3041829588 \\\n    --hash=sha256:7a32f72973f0f950c1920475d5c5ea3d971b81b6f0ec53b8d0a956cc965f22e0 \\\n    --hash=sha256:7a4aa69609f40fce3cbc3f87b2061f042eee32f94b8f11db707b66a26461591a \\\n    --hash=sha256:7c60d3c9b06fb23bd9c6139281ccbdc384297579ae037f08ae90c69f6845c0b1 \\\n    --hash=sha256:800ee55980c18545af444d93fdd60c56b580db5cc54867d8cbf8a1dc0829938c \\\n    --hash=sha256:80aa065ffd378ff784822a6d7c3212f2d5f5e9c3589614b5c228b311fd3063ac \\\n    --hash=sha256:86e0287879f75621ae85197b0877ed2f8b7aa57b511c7331dce2eb6f4de7d476 \\\n    --hash=sha256:893ff3a711d1b515ba9da14ee090519bad4610ed1962fbe298a434e8c5f8db53 \\\n    --hash=sha256:89fc958c702ee9a745e4700378f5d23fddbc46ff89e8fdbf5395c24d5c1452a3 \\\n    --hash=sha256:8c63c91f95173f9c2a67c7c526b2cea976828a0e7fced9cdcead2802dc10f8a4 \\\n    --hash=sha256:8df31fe574b8b3993cc61764f40941111b25c2d9fea13d3ce24a49907cd2d615 \\\n    --hash=sha256:8f9baf6f0a6e7571c45c8863010b45e837c3ee1c2c77fcd6ef423be91b21fedb \\\n    --hash=sha256:9027d773c4ff81487181a925945743413f6069634d0b122d0b37684ccf4f1e18 \\\n    --hash=sha256:9190426b7aa26c5229501fa297b8d0653cfd3f5a36f7990c264e157cbf886b3b \\\n    --hash=sha256:940dda65d5e764406b9fb92761cbf462e4e63f712ab60ed98f70552e496f3bf1 \\\n    --hash=sha256:94eff26096eb5395136634622515b234ecb6c9979824c1f5004c6e3c3c85ccd2 \\\n    --hash=sha256:9eed0f7edbb274413b6ee781cca50541c8c0facd3d6fd289779e494340a2b85c \\\n    --hash=sha256:ad4ae4ffd1ee9cd11357b4c66b612da9888f4f4daf2f36995eda64bd45370cac \\\n    --hash=sha256:b0f172dc8ffaccb8522d7c5d899de00133f2f1ca7b0a49b7da98e901de87bf2d \\\n    --hash=sha256:b2af221f268f5af85e776a73d62b0845fc8baf8ef0abfae79d29c77d0e776aaf \\\n    --hash=sha256:b7d335370ae48a780c6e6a6bbfa97342f563744c39c35562f3f367665f5c1de2 \\\n    --hash=sha256:b83af57bdddef03c01a9138034c6ff03181a3028d9a1003b301eb1a55e161a3f \\\n    --hash=sha256:bb5136fb5352d3f422df33f0c879a1b0c204004324150cc3b5e3c4f310c9049f \\\n    --hash=sha256:bc4d8e252f532ab46a1de9349e2d27b91fce46736a9eedaa37beaca66f574ed4 \\\n    --hash=sha256:bdd3e53429ff02aa319ba59dfe4ceeec345bf46cf180ec2cf6fd5b942e7975e9 \\\n    --hash=sha256:be12f931839a3bdfe28b584db0e640a65a8bcbc24560ae3fdb025a449b3d754e \\\n    --hash=sha256:be4a51a55833dc29ab5d7503e7bcb3b3af3402d266018137127450005cdfe737 \\\n    --hash=sha256:beb7f344487cdcb9e1efe4b7a29681b74d34c08f0043a327a74da852a6749e7b \\\n    --hash=sha256:bf4679a3d71012a7c2bf360e5cd878fbd5e4fcac0896b56393dec239d81529ed \\\n    --hash=sha256:c0e1403fd7c26d77c1f03e096dc58a5c726503fa0db0456678b8668f76f521e3 \\\n    --hash=sha256:c31c13da98624f957b0fb1b5bae5383b2333c2c3f6793d9825dd5ce79b525cb7 \\\n    --hash=sha256:c438f6ca858697c9ab67eb28246c92508af972e114cac34e57a6d4ba17a3ac08 \\\n    --hash=sha256:c8277104ded0a51e699c8c3aff63ce2c56d4ed5519a5f73e0fd7057f959a2b9e \\\n    --hash=sha256:c95cab08d1965db3d84a121f1c7ce7479bdd4072c9b3dafd8fecce48a2e6b902 \\\n    --hash=sha256:cc0b66c1eec9021353a4b4483afb12dfd50e3669ffbb9152d6842eb34c7e29fd \\\n    --hash=sha256:cdee07c4d7f6d72008d3f73b9bf027f4e11550224c7c50d8df1ae4a37c1402a6 \\\n    --hash=sha256:ce9bf03dad3b46408c08649c6fbd6ca28a9fce0eb32fdfffa6775a13103b5310 \\\n    --hash=sha256:cff8e5383db4989311f99e814feeb90c4723eb4edca425b9d5d9c3fefcdd9537 \\\n    --hash=sha256:d168fda2dbff7b9b5f38e693182d792a938c31db4dac3a80a4888de603c99554 \\\n    --hash=sha256:d1ffeb80b5676463d7a7d56acbe8e37a20ce725570e09549fe738e02ca6b7e1e \\\n    --hash=sha256:d36ca54cb4c6c4686f7cbb7b817f66f5911c12ddb519450bbe86707155028f87 \\\n    --hash=sha256:d4193f3d9dc3f6f79aaed0e5637f45d98850ebf01f7ca20e69457f3e8946b66a \\\n    --hash=sha256:d5cd5189fc2b6a538b75ae45433140c4823463918f7b1617c31e68b085c0022c \\\n    --hash=sha256:d618fd27420381a4f6044faa71f46d8bfd911bd077c555f7138ed88729bfbe79 \\\n    --hash=sha256:d76e2d8c75051d58177e762164d2e9ab92886534e3a12e795f103524f221dd8e \\\n    --hash=sha256:daae526907e262de627d8f70058a0f64acc9e2641c164c99c8f594b34a799a16 \\\n    --hash=sha256:db485b3847d182b908b483b2ed133c66d88d49cacf98fd278fadafe11b4478d1 \\\n    --hash=sha256:dd952e03bfbb096cfe2dd35cd9e00f269969b67536cb4370994afc20ff2d0875 \\\n    --hash=sha256:dda366d548e89a90d88a86c692377d18d8bd64b39c1fb2b92cb31370e2896bbd \\\n    --hash=sha256:e315e5ec90d88e140f57696ff85b484ff68bb311e36f2c414aa4286293e6dee0 \\\n    --hash=sha256:e4415a8db000bf49a6dd1c478bf70062eaacff0f462b92b0ba68791a905861f9 \\\n    --hash=sha256:e7a116ae737f0000343218c4edf5bd45893bfeaff0993c0b215d7124c9f77646 \\\n    --hash=sha256:e7c4c09a490dc4d4a7f8cbee56c606a320f9dc28cf92a7157a39d1ce7676a657 \\\n    --hash=sha256:ebae99ed6764f2b5771c522477b311be313e8841d2e0376db2b10922daebbba4 \\\n    --hash=sha256:ec4c85dc4b687c7f7f15f553ff26a98bfe8c58f5f7f0ac8905f0ba4c7be60232 \\\n    --hash=sha256:ed3a984b31da7481b103f68776f7128a89ef26ed40f4dc41a2223cda7fb24819 \\\n    --hash=sha256:f18c2d9782259a6dc132fdc7a63c168cbc74b35284b6d75c673958982a378384 \\\n    --hash=sha256:f1f9f4121ec58628c96baa3de1a55a4e3a333c5102c8e94b64e23bf7b2083309 \\\n    --hash=sha256:f42c23db5d1521218a3276bb08666dcb662896a0be7347cba864eca45ff64ede \\\n    --hash=sha256:f443b4825c50a51ee68585522ab4a1d1257fac65896f282b4c6763337ac9f5d2 \\\n    --hash=sha256:f6764a4ccab3078db14a632420930f6186058750df066b8ea2a7106df91d3203 \\\n    --hash=sha256:f7c7553b13f69c1b29a5bde08ddc6d9d0c8bfb84f9ed01c30db25944aeb852a7 \\\n    --hash=sha256:fa6248cd194edff41d7ea9425ced8ca3a6f838bfb295f6f1d6e6bb694a8518df \\\n    --hash=sha256:fa8eb9ecdb7efb0b226acec134e0d709e87a909fa4971a54c0c4f6e88635484c \\\n    --hash=sha256:fc20894c3d21194d8041a28b65622d5b86db786da6e3cfe73f0c762951a61167 \\\n    --hash=sha256:fc4d3f1fb9ca0ae9f97b095963bc6326f1dbfd3779d6679a1e016b9baaa153d3 \\\n    --hash=sha256:fd40bb9cd0891c4c3cb1ddf83f8bbfa15731a248fdc8162669405451e2724b09 \\\n    --hash=sha256:ff710414307fefa903e0d9bdf300972f892c23477829f49504e59834f4195398\nlingua-language-detector==2.1.1 \\\n    --hash=sha256:07a2447576fabbf7f381b82be6de2336f55f54d7836b0ac6eed721f3b79cabe4 \\\n    --hash=sha256:0a18d3bf0039ef8746f8df391cff885b47e2a3762bb30883eceac3d449fd1fc8 \\\n    --hash=sha256:1703cd369d74bde4fd6df8f21988c66231d8c85589e7ce535c3e251e0d4ee4c5 \\\n    --hash=sha256:17110a40f9346a4c24291b170d0deb815bd615427c4857342a7f513813717148 \\\n    --hash=sha256:178b65db951cdfbd17d05a2eb629e177e5495e57e2b41b6789e82db4df126ff7 \\\n    --hash=sha256:197b2a394015298b80a12f61294094800560761227dcb27003f957c378059b20 \\\n    --hash=sha256:2477f75cc871d20bafdcb56f9d51be25d57c37d9ea5d7301cc592761dc68c963 \\\n    --hash=sha256:284077366b7ca3b2c4ecc492f3e40570f6afeaee2bf44153d347fa100137561b \\\n    --hash=sha256:2a468c3fc9eaa6db733a347fee768fe171e76fac2c4bc49951e26bc79aec6a2a \\\n    --hash=sha256:354040a3c2ac748623966373cc64de34e8f14b0043c09aa334fdccf3456bf5d0 \\\n    --hash=sha256:3de2070ab293457a4f0fc1bd87f34de2e98c8348205b40b5667e043485950a64 \\\n    --hash=sha256:430e9e517427070f20d1b9eaff88633614d332cdcd119a519cefcd8c9f3d67e9 \\\n    --hash=sha256:53cc131d9c7be64a88b1e20633d66c62a2779776e11a26e32bac80fb19b43f33 \\\n    --hash=sha256:5a6370392683607a34e941ea05088fde197d5dc37b8abb088fada7a51749ca44 \\\n    --hash=sha256:5f0917b9210b59acb0d2c0979a78f3e54dcf9967d54885a4a8d34264d5a07720 \\\n    --hash=sha256:6190e7632d08467dd3d148134743c4c40ccb4c84d6f3313508ffd73a8210c614 \\\n    --hash=sha256:69a408fc0bb372a46afec6d0744077b9932c64df7c02ad517bf37c5c7e3734bb \\\n    --hash=sha256:6a398e4871fe8e32ff5711eaabb09ebdf4f80420d73e5d646a6cae0468c7c47e \\\n    --hash=sha256:6e1c950dceafd5ee12b6267d9f71987c572a1ee9f18b6465c722fcde9b0d5149 \\\n    --hash=sha256:7e82bcb924d09e552a52bc79265d5e49a5863b6b7297524adc4fb7c4564ebb5d \\\n    --hash=sha256:80acd7652f95ae569e6a03ffcb0cb9522ea3fde2328b9c6fac15b1d24ba382d9 \\\n    --hash=sha256:8230d7a08d0477d136f22ec53cf8eaffaa90b273a7ab616d972298360c2a4090 \\\n    --hash=sha256:98f6128ea7b6122b23dab9168cd447fe85a3cc90d0272b7ea67034453715d306 \\\n    --hash=sha256:9c195a39f3b9ebeec9af72acf03a0b13132bff09147cd50d99001c060a998eaa \\\n    --hash=sha256:a54d976a1daa8ecb5fd3f36e1ed5d3f9a363beed6edca22e88e49a8af9c4757a \\\n    --hash=sha256:a857dd48a801f6a492a7832dc812c1a256f83da93725897abff7534321b0b7b1 \\\n    --hash=sha256:b359f2571ff7ca6b4b998a3fc10fa87f136dbd4e4809af5283bb6b7093e88a07 \\\n    --hash=sha256:bbf22f8b1715f577f8cda4758d61ff3c1f5238d48b8cbe035a3c2064edf7b0a5 \\\n    --hash=sha256:c121e6340bb4cb051e1469adc4575790a94a47f03999155c028fd5cfb4a7516b \\\n    --hash=sha256:cd9f734f67da00d37d93a1354f7c44d86c83b297cc078d7f31f1fd8a7deddef9 \\\n    --hash=sha256:d0429b482e8b3def24ad4e5565eb187a389c35cf0506ac7ed7ef7360e7369ec7 \\\n    --hash=sha256:d3790a8761c37d2c4c2aa287c1a6a1f8d3d5d9b0d74c276ffd37243385cc33f0 \\\n    --hash=sha256:d55300513d9e2e0e034f6fd6b8cf111ba9faf49ba20cc060d67b403c3d356148 \\\n    --hash=sha256:db18582802ae49d03fe4330d13bbb9285532f157f3e7be291f83ed2ea90ea190 \\\n    --hash=sha256:deec7c9d5a72d7434144952cd5c3b85923d7a858264baa9c21018cb0e2e929fb \\\n    --hash=sha256:e94ad32a46f04b670a939623b69d7d7008221d15e20b3d975d284304d9d6c788 \\\n    --hash=sha256:f9e116206431eb283a4bc9107407a58b7d093870ae9d50ab43e39796db029fc5 \\\n    --hash=sha256:fa3c9cfe7f7d9dc857ba22f14c2dfc7834e2dac131afaa737f3a59adae3ea553 \\\n    --hash=sha256:fbd56b2f830f77819f8f97ecc11bff421133ced2da7c2cf9e05ebfe0e9f625f5 \\\n    --hash=sha256:fe21c948a387fd9aa0b994853eb47cfedcc738a91530193de839ef0977ecc0de\nlm-format-enforcer==0.10.12 \\\n    --hash=sha256:130bd7ce8a6b224f25b6314ba9ae78ee4b48594db1767c74391c9182e2902a6c \\\n    --hash=sha256:267c2b421c77f7cd51ac2e0e3af8db278a373704d834b49ff55f18a2c05e9800\nlxml==6.1.2 \\\n    --hash=sha256:0349321a0537d4fdbebb2af06dd1b64676132c72e2ae250de8cdb58f8c43019c \\\n    --hash=sha256:04cf9e3f4ee9cab9d9ba05401bef8668840fa9620fcd4d8e85a2d2fd0b0fa960 \\\n    --hash=sha256:054175250531a5fb102d485743ff16412279c93add12385b3b1c3d7b16d8deaa \\\n    --hash=sha256:058c79e172926ef524fb3c7c6beea4b55e15886ac99cb0c139ecaac6b375f1e2 \\\n    --hash=sha256:0666943ee1576fa890a6dc6316ef42e8241b5dd56f67bc5475acb2ac298c6ca9 \\\n    --hash=sha256:074a88f70a7360a4a0c5be5d898062cd26f898c25b459efb1bdd43ae700c5a1a \\\n    --hash=sha256:08cd52e6487435c75f2da0a5b276beef7fed161681b93ab766e66b954f0c349a \\\n    --hash=sha256:08f0c9ed7cded07c5e798b17c9c25bbba5d0650c8ff0a7f65f84c634966f0f10 \\\n    --hash=sha256:093fbf547d0f3ca02705381f795a050fbb58988be4aac7f79f99f280c4082313 \\\n    --hash=sha256:0aa07065497f191ad26c4b587ce5dbb5a7105285a3789aafd0661750e8bac537 \\\n    --hash=sha256:1055241852f2b02068af4a625a5d32c087db193c12251928af2562ecd2239f18 \\\n    --hash=sha256:1133bd969f2bfcc6b0c0cf7cdf5f2631e62b23fa2471ee8bd44f6ab73554ee9a \\\n    --hash=sha256:11f529062255209a421ae4de5b1bb36b2f0a2e1a700745e675a4bf4084d13c00 \\\n    --hash=sha256:12acd337d2821cb8b9247dfe4b7aa2f2769a3df5ae8511b7e550df42b8f4d3c3 \\\n    --hash=sha256:12ecfea07d767f6accbf30b014e1c477b5eabb13eb4e8c748215efb52c0e314a \\\n    --hash=sha256:14879fa5eb2b793c040bbfcb62011aa3015c65d6c9875e063ea98ce2029d51fb \\\n    --hash=sha256:18467b0e9f7f0bc477df69e99829a59ae17fb37d34e5f68399371c7c67be9002 \\\n    --hash=sha256:1a2331da06dd55a8184985306eb2afd72d708283ce7e85d67bba77317b785060 \\\n    --hash=sha256:1c0173595dc1c25768f42681a1517dcfc74bb18a34695f127931cbd05f4dead6 \\\n    --hash=sha256:1c4c6dc1b2485aaa4adfb6ed754f90dddcb2b96a66bbebc9e1ac242b5ce5e818 \\\n    --hash=sha256:1d55a614d2f0457b1f7511c1b7bec0db0dcdd4af4d09d226829eb054c647527c \\\n    --hash=sha256:1e3c67b817867c484794d7fe0d73045d7d0c67460c78a0a1249a9e92266e6a0e \\\n    --hash=sha256:1edca8f4a92b94e873093df959f141d388f2141fcad0c47598442fb4730ef57a \\\n    --hash=sha256:1fcfe8481302e6dec07909914b8f3f9e1739ae1615209d4b9e7544325fb699c4 \\\n    --hash=sha256:20134744db7abcbd5232214e767814ef64e5ab57a5b7df93a2bd68b74ef0a6c0 \\\n    --hash=sha256:215bb3cc4be015ccac3c7d4f25eb7b941f857fe5b02c0e3504cca61f7fb12455 \\\n    --hash=sha256:2170d0a280c877b6e2dc6738217db947be35dd8cf09ca458b355aa1bab2a9e70 \\\n    --hash=sha256:2374235206ec83d4827ad219c93c0f7366b93626eab85392c0ee7c8026649376 \\\n    --hash=sha256:243ecef7cb7415766dd742336cd5b8361a84c6f297e2773c865b783724cbbe74 \\\n    --hash=sha256:261d98065326676d7253882db0198d0aa06748d7ee0443367acf10b148273f99 \\\n    --hash=sha256:26ff164c6629e5c4d11c9e55d5ea3d6eed0be2a420eee1f55cbce6e2c23e231a \\\n    --hash=sha256:2afd1688e372d8eafaa6f56c589399e0a87d086a0c110f6346b0b50f42e67e25 \\\n    --hash=sha256:2dcc69e307e0916c7a0b552212010938d02a664d29b6bda75ab2bc5fa487c861 \\\n    --hash=sha256:2e37fe49fe2d5aa40a2cb1cc8176673ad7de0d124e6f4a509d9318f5979c7871 \\\n    --hash=sha256:2f3194777c0d05945ac91d8594be25d2679d1d826e01e1fc90bae568ff3a547b \\\n    --hash=sha256:351318f5c0eb7fcab5b4fdb507c6f88fb2c4b5e67784c7e5911448c91fffb5d4 \\\n    --hash=sha256:351855814dec4ad55ca5f24d0f4b1cdaca7927fe48023a2965351845f3b60cff \\\n    --hash=sha256:3a698fad6f122a9b3e2dc2fb598c1de7329c74a67c7a334c9109a440de2508e5 \\\n    --hash=sha256:3be94d2464f19e42d8c39a299f356b12f2fd095c28793671eabfcd9db9c76987 \\\n    --hash=sha256:3e3b666f57a5d81562f38c766c762416b0f6eb58a00590546911514b48412abd \\\n    --hash=sha256:40366c23a938008a3bedfcfd80709b3a857c188b4d710b083e978ef5d2c1c715 \\\n    --hash=sha256:4303f904fb6c41b58dc70743b1d8a470aba6c9897427c48324cff1a95673ddb4 \\\n    --hash=sha256:442766b326d9892585a64e8c6c4b5ab81d0e6c0538c9f0fc11a84dc101a5d97f \\\n    --hash=sha256:446f1f92c137e0cbb97eb7e932e15315c11a7c86974f43f15e68c9707ac6a9f6 \\\n    --hash=sha256:4618b20f43dc98b49569b1dc822176140ea0f2598d672a6989187ba49bcbfec1 \\\n    --hash=sha256:4622c5616683faf63791b349e6c8dad7717412dc5f29f4febe7575f110609a86 \\\n    --hash=sha256:47c92dc5167de16e27ace8332454f12ba172dcab04f7a78a9eae14e2e41b6a41 \\\n    --hash=sha256:47e367dfe341521426692819803e260d0673899c0ff611f14af978d725e2c999 \\\n    --hash=sha256:48e912f37c99a297175ba955f55a47c0e1c834b506ef162e52a6e4fe276e6e45 \\\n    --hash=sha256:4a16457e330b7099aa5a8e8bfa5d53a33a1672a819fa656157e9e6dc433ac7a4 \\\n    --hash=sha256:4aced3284e0353c798b060fe2c175eb81410e99b9a7e2ae6951be5333732b111 \\\n    --hash=sha256:4b0fa7109b1d0bc1747d8241a0853e135eefb1c978685241b544c46937383efd \\\n    --hash=sha256:4bf14db2f0214003ec7f46c4300e2065668fc93e20448c1c95bac2e952072168 \\\n    --hash=sha256:4e220a9c297e5d36895d489a08c9a3f1f6193b6414e702c5fb751e4a3767f8d0 \\\n    --hash=sha256:4f4d2c36fd5997d30ff19c29fb93293401d0daaf87512297d47610e6883964b5 \\\n    --hash=sha256:5078ff51e6316c0f75ea8127c2cd24374747fb351f62fb93d1761f8ae5a04a40 \\\n    --hash=sha256:50ee0c360862f4152db835b456e38614f94b674bca2a47bc8de7171ee6ccbbb8 \\\n    --hash=sha256:522387e05cd015a81d1dc621fb167fb42b8f629ccd2e8b39de583828f165aae6 \\\n    --hash=sha256:5295205fd57510c19a0e46385b516119f3a781d45c2672159bce02949238981a \\\n    --hash=sha256:52f6d4dff133c9778a24e9a2cfc1608930b15869866171aacc5131b5a418a003 \\\n    --hash=sha256:57188e441ab24f906bd5a5c14eb55363ab51aa6c0de549f3dd320043721cc118 \\\n    --hash=sha256:575fef7f30048b744dffb3e4ff64a18cac7dba3fd26efdea5730ade9d1bdeb33 \\\n    --hash=sha256:5848f3de6a8de8a93cff9f068134393ff5fa69ac2a04399f7d49cd67c61c348c \\\n    --hash=sha256:5a096d6a5f96b776a5b020cb45c17c545effd2a3b6639e6fa97bc95537600923 \\\n    --hash=sha256:5c2bae42b3a09f977330a08f4a8fe72aec58c4bdb89069d3fe7272a71d885881 \\\n    --hash=sha256:5d78ba560f3dd404d87b1fcc89b2b382d638ea2998431a3b2e5cda0f3ba2da91 \\\n    --hash=sha256:604f4778632588d7c000e7e19430639dc12fca58b5b6e99edffba7631725ef0e \\\n    --hash=sha256:614d4c5a34556e369b86cfcc8d0cf71cd0759a3444a464a07a9427ab0f5e3a99 \\\n    --hash=sha256:6330cf0ce83f6273ad8ad99bdd25d6ebb3863912f9ac717f96bc8942706e0e26 \\\n    --hash=sha256:633ac039cb32366dd5935868e041e385875c017b8cd54ea56aeee3fe29ca5935 \\\n    --hash=sha256:6454d184d556eaf4cb3d6f69e405d21602d6fdcf08b8d57796824275986c6595 \\\n    --hash=sha256:648861c19b775b89ebefa14586f85090b10163367476d77f242c4131c835ce73 \\\n    --hash=sha256:65c32ddc5d0750129c7b119fb57d48192b76d334c21e6b690d19dfb06b34af79 \\\n    --hash=sha256:662432a6103e671d971e06e75ed146d9ff67f39d2c98c2f26613b6057f54eafc \\\n    --hash=sha256:678e35f1cbca98f55107511ee21a60568535c950f3c2371819bd64504c980d20 \\\n    --hash=sha256:69df1856cb6c065e5bfd23adcc7408bfa6dcf32b0018373a99b0769bd86e2256 \\\n    --hash=sha256:6c9cc4b6532abe154dbdebb42aaba8d52c852919591e45067f5b7d46a0405e88 \\\n    --hash=sha256:6cb0c87421946030b92b558be416852780a912454e3dcba0998e4497c9c588d5 \\\n    --hash=sha256:733dfb492ec3dfef8350a5cc896e90d202c5171e791e1609e77563751d69a15d \\\n    --hash=sha256:75530642d8471327e691ab9b0513a5f9c77f38871014ceda40f51bb51765c0a1 \\\n    --hash=sha256:7766e525282dd38fd89567311323e441996eb958e8e816d16b38f782e3aecd2a \\\n    --hash=sha256:785761d5123f222cd97f2263a510107226fe32ce7aa7824a90616a41c574ace1 \\\n    --hash=sha256:79b428c3242e63bdacf3b526a34e0b8b26583846fc597da84b8f0c3d5ea446b2 \\\n    --hash=sha256:7c444c3a6e8e75334879980eed96568f0e12064c8b1913424eac1805e976736b \\\n    --hash=sha256:7c482e87cc86bed78a50462560675bc2c348ef72c47596f9b933346d5a8e920e \\\n    --hash=sha256:7c534ed898413f439b048130011e99a4245ee13d62d431f6b4f7f2484d02a93a \\\n    --hash=sha256:7c687fd8e558c7d169f6f1987b696f37824d3a097f291bffd0ab4a2ea2307dfb \\\n    --hash=sha256:7d506bdba580ecb1a6ad2e2b5c49445e66d3e1f95894885739094393a1aad237 \\\n    --hash=sha256:7e81fc065ede5d58dd0bf0912025aee1bd04c52c2affd61fdb93226a97ce2fc6 \\\n    --hash=sha256:7f35ba7667004ecdafebbe08da7c9fa06ee6195275bb7ef7a29ee1901e69519c \\\n    --hash=sha256:7feb72424f19a893ae4f3373c7aae821b1aacb6076b708915c651f0683a97c49 \\\n    --hash=sha256:822d9397033edbe530a13bb1e0091c0e817536b6aba87a9b4ad626ed779ca0bd \\\n    --hash=sha256:827438bf6c8292d22a409bb7990d7cffce410f33e7664e46ca74d2ecc26975ef \\\n    --hash=sha256:83e7510a6dda8df41d1b68b783de2953b3feb55a11dcebf693201ebaa5cc0c4a \\\n    --hash=sha256:841630176c15fa5d3c5cd6f755435d3c5540a82e1dd2a7de1799401f92ee6d24 \\\n    --hash=sha256:84a2a46b93b789d8acb44cfcb3d967ce9dbe29884ddb93fbb1a33f0e0c8fcd86 \\\n    --hash=sha256:8512b3775d68994dd1d6d533161e0a214f2ad9c634659d34a99c98e86c6c3d68 \\\n    --hash=sha256:85690cfc8ed54c4292e36a08bcf984dde7957e653fd6d94f59184244bcc35843 \\\n    --hash=sha256:86d93dc3882c283e9aa2124d7d2b50c85579485216a2b3b7f91ba479e31a128f \\\n    --hash=sha256:87534cec6ea325435e4adf2326b0cf3110eee9a47abf73652eb155db639c08c6 \\\n    --hash=sha256:878e7c8ada8f92c52f13f35a2ab98ef0adf7fd0211d164fc2af589e4c3cfed63 \\\n    --hash=sha256:87e9673cd8a3445024fe38e7f91b55fa3428437eec9b7a7ff7d81979520c0d2d \\\n    --hash=sha256:8807998c1023d1e9d60e02500f90e85a0752dbc0b670989806bba87b82dd5b42 \\\n    --hash=sha256:8b68f2548259bb04e0b3d5df0c397abe8b0080f5e1ffe4019fb7a8bf01a9339e \\\n    --hash=sha256:8e613018a5ac66de7abaf1acaae0d7af37a5e1b9bf1ae190a1198b0fdb988ad8 \\\n    --hash=sha256:8ec111ff8067325f85c08aa9c2b26179ec0537bb89c003fde31127139f85f82d \\\n    --hash=sha256:8ffb17ec0a8bae18b6628ae40b0896eb264dd285e39a0faa864965c00933b64c \\\n    --hash=sha256:9031f5f01452681abf39fdd65f84a70cb01a7572a1bbf570042e826b1232d07b \\\n    --hash=sha256:9088da25ecd609965f838d89fda0465a905b48f4dd90331db9845518f2177372 \\\n    --hash=sha256:9221442682c27417f10fe11184ea4cce174b25ab52465570b1f3ee3f85f320fa \\\n    --hash=sha256:927f3e1d04dc0906265fc0416c13500363e42cd683bbb8d46911c79b73d26800 \\\n    --hash=sha256:92c2b366028ac01e90399e6d17734ce6e4f4aeddd8ba75fbaf80ea11d6c6d645 \\\n    --hash=sha256:94162456ed0a64fb1c06915df5bd06af4675ae3966d6048fcb73b0906e0e0222 \\\n    --hash=sha256:9429d2371d406344ed1da5b5686d9412e74137c07b0171278368ff704f470ed5 \\\n    --hash=sha256:9477e14217c212e6023c994a71a1a349db19b0e10fd5bf189666b281ae63b1fd \\\n    --hash=sha256:962c12b51d0b164f12569af225dea57568477e24a845b96eaccbef6c07e4cc03 \\\n    --hash=sha256:9b52ea73a37fc64aa3357ff8607801d46dd170506d3cf8253a91a1d91639d4f9 \\\n    --hash=sha256:9bdc2db9e04538f917bba0242920764dd740649d8df58700d6d687ead4429429 \\\n    --hash=sha256:a02164a8cd3e2dc028918e51af844c934c7a24a0b8f4064368360aa14ad1aac4 \\\n    --hash=sha256:a2b7fe53abced1fe8bd984a9ab3c8c98bc093ec4f9f543089a8817a493818208 \\\n    --hash=sha256:a5005c0c9e4d749a76a2ff8bd5918a8bb248df8e08e73a55654b9f79c9cd1e2b \\\n    --hash=sha256:a7fd1dd6faa3df9dcd8f1765237362cd885ca62cdf77a7c5f5ea383ae5b6048b \\\n    --hash=sha256:a8326e24ae6c3a6bfb03fa8b4793f9a5d804c125228aa067f652b0428e31b87c \\\n    --hash=sha256:aa224ecc613d411690aa650dbf01daafbe385cd6c67145e80bc5fc01b3a71469 \\\n    --hash=sha256:adbecbfe44a497c742792457b1c27300617967c18c3934d2416023eba8d8c553 \\\n    --hash=sha256:ae520f189895c5dd7eeb2b7a372d464da6f4a1ba1d0ecb741b1d4fe4c1f699ac \\\n    --hash=sha256:aea814342f6afd20d832937ff8b333cd6506428a39c0c4c70c2380aab1887bfb \\\n    --hash=sha256:aebcc6b184c935e1f7091c09124cfe5107b7c2253894ba23ad646828c17e4c3b \\\n    --hash=sha256:af6585a466cee2c5a524f7fffc591844bd604a29fdd9cade964f548512b5ef7e \\\n    --hash=sha256:b1c0d2dde8a50520efc51644587f0fc4810e3af7d3e029d7af0be93bf39e2b5c \\\n    --hash=sha256:b20440e578d269c5e8a722ab602ddd0f0cedb8b080006b3f936da9991a593d3b \\\n    --hash=sha256:b28842b30c4bc2e6afe137d98a5d2071a62589471e76d053bea55b0e53298af9 \\\n    --hash=sha256:b3ca02ef3b5920b88119c82eb6badfb2d082b1f681d528a856dcce17c8706da8 \\\n    --hash=sha256:b3db5497af55f7a557c95265dd3b91c75dc56364a7b59f258c45fa5576dce058 \\\n    --hash=sha256:b631174cd2e4d9f8a94ef17f911c6ded10ede93b5e7860dee7bbf85961d321e9 \\\n    --hash=sha256:b7233a987a101bdf79059014130262a01339094a0a709f175162542f33b55d4e \\\n    --hash=sha256:b97153ca609b434b712ddfb92cd6af101a7045a7724c542258bd4727a344472f \\\n    --hash=sha256:ba0dfead73be5be9ad0b7fbf9f31ff29c1b1eae858816dfc8d85099d6e4af0d6 \\\n    --hash=sha256:ba58574d710b82ead7cbedea01cac3e110bc3ef82d4731519b74a2c11f7cf5e9 \\\n    --hash=sha256:be365ce8d2d411cf2fb573747684b4fd470fa6224e0094d9d5a21155acc369d3 \\\n    --hash=sha256:be6f87cd224254a8f81324e34cc655508b83f1d70458a1a39857ad2aa9925852 \\\n    --hash=sha256:bfcbee8ffff4188f4c6d97eceeff36d8eb983cf838933cbc12ce5f5dd51476c6 \\\n    --hash=sha256:c0edde95e4b4278dcc0175eda06dc8aa2631ad9f83ae5dbdbc4f0925e200b0b0 \\\n    --hash=sha256:c20fa05d128c463209ef5323ebf33ee1cac6d87cdc3933fd789fd3c101017c8e \\\n    --hash=sha256:c470d192e27f97842a068cf12a1c1296b20ca716c56a9249715c6654bc192d19 \\\n    --hash=sha256:c67f3c1278f942e97d8665c2a690324aaea5137de16f056583a21f0ac706177f \\\n    --hash=sha256:cb0cf498efa3204621b3c5576f0accd80ad2ee85575f1cae5d2f98de32c8d9cc \\\n    --hash=sha256:cdd35422de747237f451e821766e2b6be3dd2c31955c1ecd7f17984c5b9bb62d \\\n    --hash=sha256:cde6b8db7d2e5135129eb5e74b7b44dd2053aa767cd5023541fccedddc262453 \\\n    --hash=sha256:ceafa5e0536c62a5cd9f65327fa0b57d6f0b0e3435daf2c98a78d0dde7ecbae1 \\\n    --hash=sha256:cfeac14425fc7a6fca7864b774d4ee63547926158f4a18c67d77b2c9a948acf1 \\\n    --hash=sha256:d0bfd719c254bbe60ea022cff0e6ffb799a6fa7d4d72852cebe0257957b32d68 \\\n    --hash=sha256:d117f39b28ab8a330a74abdbe61c2255b51973b238db25fd6c2448de1eb2a02d \\\n    --hash=sha256:d3e97ac4353cca3fbbfa829bc0c6a913771573d1c6d46932d4335c46f2b7796a \\\n    --hash=sha256:d50a44113fe6800dcc8a859332b823a4735b1e6ae1b0063882e4cca569ec3e29 \\\n    --hash=sha256:d858e718b94033ab4b67e4a58fe3114c65bae01ae2314a62fb39ae8897ed4324 \\\n    --hash=sha256:d86130d70a2557cdf825dffc56255f1f16b83a7bbeab677b4cd040c4c53d8c52 \\\n    --hash=sha256:da6a4f55f0e3308c07354b1ee239c5550afc212f81629a6067db505ace3b667a \\\n    --hash=sha256:dd7ea3fa47154b9fff90591b961e41b3718bd7fcd5bc2d9bb47e9845c8ace088 \\\n    --hash=sha256:e062f5ac1255dfa6c98e3e3863ec18bc79d0947d22d08921a3ca60cee40559fd \\\n    --hash=sha256:e17e2c30e27f56da5551e7a425888b45f013e940b99ab07d125a1c33f77a4605 \\\n    --hash=sha256:e7269cc410f3cdf84a66914fc0ef54b1618115c87fb4f9a59a05c5dfc23bece1 \\\n    --hash=sha256:e8b9a92652e75e7731309ea51db5dee892eef414ce70a6ec3441e5d36bf5189f \\\n    --hash=sha256:e8dc3d29f2ed2bbf24c205a86326d6681230ace55abfb3f9d5230f42078ad63d \\\n    --hash=sha256:e92e4419cad18d60b14bf18b82152fbae67f4b1128be7d73b172df275554f5d9 \\\n    --hash=sha256:ec8d09f460fdeb65f9ead9b75941e312def4bcbb23e1f951b7def061eb99501d \\\n    --hash=sha256:ee23f6599682bd4d48bb757c0633e78774eedfb65a7e52851f9ad182eeeb625e \\\n    --hash=sha256:ee7410c98222070fd717ad881ee2a80cc11826b7001b9a5a807155d8918bfc7a \\\n    --hash=sha256:ef0b8ba6e13597f681b2b4924ca9c4e8c88420bf0e21d9a9006c757f2fc39d1f \\\n    --hash=sha256:eff128ffdc093cc6317955934ad9751105d37ed8dbca3ff4ccd751af6be37185 \\\n    --hash=sha256:f16a407766bac51c65d605b06d900821751a79aa20e12185f273f14a17180e7b \\\n    --hash=sha256:f86e23ed610727a7f025ebbff788f22a7956d3f1b24a25bb1d9286fc7b7642b0 \\\n    --hash=sha256:f8b89b3be75a37509602b03f9cfa1a28298d4eed4625748148307aeb907901b7 \\\n    --hash=sha256:f93bc5e25992f5545709000d840c6cafdbd022781a7a0ed79d58a5633733a4e8 \\\n    --hash=sha256:fa813b0247d0543a563b993ac3dba6168eef59e3a61448432cf5453300c2412b \\\n    --hash=sha256:feda2ef68c339987dfb370af3a4b785dbc40f925723fe2365e68e43c2640f85a\nmarkdown-it-py==4.2.0 \\\n    --hash=sha256:04a21681d6fbb623de53f6f364d352309d4094dd4194040a10fd51833e418d49 \\\n    --hash=sha256:9f7ebbcd14fe59494226453aed97c1070d83f8d24b6fc3a3bcf9a38092641c4a\nmarkupsafe==3.0.3 \\\n    --hash=sha256:0303439a41979d9e74d18ff5e2dd8c43ed6c6001fd40e5bf2e43f7bd9bbc523f \\\n    --hash=sha256:068f375c472b3e7acbe2d5318dea141359e6900156b5b2ba06a30b169086b91a \\\n    --hash=sha256:0bf2a864d67e76e5c9a34dc26ec616a66b9888e25e7b9460e1c76d3293bd9dbf \\\n    --hash=sha256:0db14f5dafddbb6d9208827849fad01f1a2609380add406671a26386cdf15a19 \\\n    --hash=sha256:0eb9ff8191e8498cca014656ae6b8d61f39da5f95b488805da4bb029cccbfbaf \\\n    --hash=sha256:0f4b68347f8c5eab4a13419215bdfd7f8c9b19f2b25520968adfad23eb0ce60c \\\n    --hash=sha256:1085e7fbddd3be5f89cc898938f42c0b3c711fdcb37d75221de2666af647c175 \\\n    --hash=sha256:116bb52f642a37c115f517494ea5feb03889e04df47eeff5b130b1808ce7c219 \\\n    --hash=sha256:12c63dfb4a98206f045aa9563db46507995f7ef6d83b2f68eda65c307c6829eb \\\n    --hash=sha256:133a43e73a802c5562be9bbcd03d090aa5a1fe899db609c29e8c8d815c5f6de6 \\\n    --hash=sha256:1353ef0c1b138e1907ae78e2f6c63ff67501122006b0f9abad68fda5f4ffc6ab \\\n    --hash=sha256:15d939a21d546304880945ca1ecb8a039db6b4dc49b2c5a400387cdae6a62e26 \\\n    --hash=sha256:177b5253b2834fe3678cb4a5f0059808258584c559193998be2601324fdeafb1 \\\n    --hash=sha256:1872df69a4de6aead3491198eaf13810b565bdbeec3ae2dc8780f14458ec73ce \\\n    --hash=sha256:1b4b79e8ebf6b55351f0d91fe80f893b4743f104bff22e90697db1590e47a218 \\\n    --hash=sha256:1b52b4fb9df4eb9ae465f8d0c228a00624de2334f216f178a995ccdcf82c4634 \\\n    --hash=sha256:1ba88449deb3de88bd40044603fafffb7bc2b055d626a330323a9ed736661695 \\\n    --hash=sha256:1cc7ea17a6824959616c525620e387f6dd30fec8cb44f649e31712db02123dad \\\n    --hash=sha256:218551f6df4868a8d527e3062d0fb968682fe92054e89978594c28e642c43a73 \\\n    --hash=sha256:26a5784ded40c9e318cfc2bdb30fe164bdb8665ded9cd64d500a34fb42067b1c \\\n    --hash=sha256:2713baf880df847f2bece4230d4d094280f4e67b1e813eec43b4c0e144a34ffe \\\n    --hash=sha256:2a15a08b17dd94c53a1da0438822d70ebcd13f8c3a95abe3a9ef9f11a94830aa \\\n    --hash=sha256:2f981d352f04553a7171b8e44369f2af4055f888dfb147d55e42d29e29e74559 \\\n    --hash=sha256:32001d6a8fc98c8cb5c947787c5d08b0a50663d139f1305bac5885d98d9b40fa \\\n    --hash=sha256:3524b778fe5cfb3452a09d31e7b5adefeea8c5be1d43c4f810ba09f2ceb29d37 \\\n    --hash=sha256:3537e01efc9d4dccdf77221fb1cb3b8e1a38d5428920e0657ce299b20324d758 \\\n    --hash=sha256:35add3b638a5d900e807944a078b51922212fb3dedb01633a8defc4b01a3c85f \\\n    --hash=sha256:38664109c14ffc9e7437e86b4dceb442b0096dfe3541d7864d9cbe1da4cf36c8 \\\n    --hash=sha256:3a7e8ae81ae39e62a41ec302f972ba6ae23a5c5396c8e60113e9066ef893da0d \\\n    --hash=sha256:3b562dd9e9ea93f13d53989d23a7e775fdfd1066c33494ff43f5418bc8c58a5c \\\n    --hash=sha256:457a69a9577064c05a97c41f4e65148652db078a3a509039e64d3467b9e7ef97 \\\n    --hash=sha256:4bd4cd07944443f5a265608cc6aab442e4f74dff8088b0dfc8238647b8f6ae9a \\\n    --hash=sha256:4e885a3d1efa2eadc93c894a21770e4bc67899e3543680313b09f139e149ab19 \\\n    --hash=sha256:4faffd047e07c38848ce017e8725090413cd80cbc23d86e55c587bf979e579c9 \\\n    --hash=sha256:509fa21c6deb7a7a273d629cf5ec029bc209d1a51178615ddf718f5918992ab9 \\\n    --hash=sha256:5678211cb9333a6468fb8d8be0305520aa073f50d17f089b5b4b477ea6e67fdc \\\n    --hash=sha256:591ae9f2a647529ca990bc681daebdd52c8791ff06c2bfa05b65163e28102ef2 \\\n    --hash=sha256:5a7d5dc5140555cf21a6fefbdbf8723f06fcd2f63ef108f2854de715e4422cb4 \\\n    --hash=sha256:69c0b73548bc525c8cb9a251cddf1931d1db4d2258e9599c28c07ef3580ef354 \\\n    --hash=sha256:6b5420a1d9450023228968e7e6a9ce57f65d148ab56d2313fcd589eee96a7a50 \\\n    --hash=sha256:722695808f4b6457b320fdc131280796bdceb04ab50fe1795cd540799ebe1698 \\\n    --hash=sha256:729586769a26dbceff69f7a7dbbf59ab6572b99d94576a5592625d5b411576b9 \\\n    --hash=sha256:77f0643abe7495da77fb436f50f8dab76dbc6e5fd25d39589a0f1fe6548bfa2b \\\n    --hash=sha256:795e7751525cae078558e679d646ae45574b47ed6e7771863fcc079a6171a0fc \\\n    --hash=sha256:7be7b61bb172e1ed687f1754f8e7484f1c8019780f6f6b0786e76bb01c2ae115 \\\n    --hash=sha256:7c3fb7d25180895632e5d3148dbdc29ea38ccb7fd210aa27acbd1201a1902c6e \\\n    --hash=sha256:7e68f88e5b8799aa49c85cd116c932a1ac15caaa3f5db09087854d218359e485 \\\n    --hash=sha256:83891d0e9fb81a825d9a6d61e3f07550ca70a076484292a70fde82c4b807286f \\\n    --hash=sha256:8485f406a96febb5140bfeca44a73e3ce5116b2501ac54fe953e488fb1d03b12 \\\n    --hash=sha256:8709b08f4a89aa7586de0aadc8da56180242ee0ada3999749b183aa23df95025 \\\n    --hash=sha256:8f71bc33915be5186016f675cd83a1e08523649b0e33efdb898db577ef5bb009 \\\n    --hash=sha256:915c04ba3851909ce68ccc2b8e2cd691618c4dc4c4232fb7982bca3f41fd8c3d \\\n    --hash=sha256:949b8d66bc381ee8b007cd945914c721d9aba8e27f71959d750a46f7c282b20b \\\n    --hash=sha256:94c6f0bb423f739146aec64595853541634bde58b2135f27f61c1ffd1cd4d16a \\\n    --hash=sha256:9a1abfdc021a164803f4d485104931fb8f8c1efd55bc6b748d2f5774e78b62c5 \\\n    --hash=sha256:9b79b7a16f7fedff2495d684f2b59b0457c3b493778c9eed31111be64d58279f \\\n    --hash=sha256:a320721ab5a1aba0a233739394eb907f8c8da5c98c9181d1161e77a0c8e36f2d \\\n    --hash=sha256:a4afe79fb3de0b7097d81da19090f4df4f8d3a2b3adaa8764138aac2e44f3af1 \\\n    --hash=sha256:ad2cf8aa28b8c020ab2fc8287b0f823d0a7d8630784c31e9ee5edea20f406287 \\\n    --hash=sha256:b8512a91625c9b3da6f127803b166b629725e68af71f8184ae7e7d54686a56d6 \\\n    --hash=sha256:bc51efed119bc9cfdf792cdeaa4d67e8f6fcccab66ed4bfdd6bde3e59bfcbb2f \\\n    --hash=sha256:bdc919ead48f234740ad807933cdf545180bfbe9342c2bb451556db2ed958581 \\\n    --hash=sha256:bdd37121970bfd8be76c5fb069c7751683bdf373db1ed6c010162b2a130248ed \\\n    --hash=sha256:be8813b57049a7dc738189df53d69395eba14fb99345e0a5994914a3864c8a4b \\\n    --hash=sha256:c0c0b3ade1c0b13b936d7970b1d37a57acde9199dc2aecc4c336773e1d86049c \\\n    --hash=sha256:c47a551199eb8eb2121d4f0f15ae0f923d31350ab9280078d1e5f12b249e0026 \\\n    --hash=sha256:c4ffb7ebf07cfe8931028e3e4c85f0357459a3f9f9490886198848f4fa002ec8 \\\n    --hash=sha256:ccfcd093f13f0f0b7fdd0f198b90053bf7b2f02a3927a30e63f3ccc9df56b676 \\\n    --hash=sha256:d2ee202e79d8ed691ceebae8e0486bd9a2cd4794cec4824e1c99b6f5009502f6 \\\n    --hash=sha256:d53197da72cc091b024dd97249dfc7794d6a56530370992a5e1a08983ad9230e \\\n    --hash=sha256:d6dd0be5b5b189d31db7cda48b91d7e0a9795f31430b7f271219ab30f1d3ac9d \\\n    --hash=sha256:d88b440e37a16e651bda4c7c2b930eb586fd15ca7406cb39e211fcff3bf3017d \\\n    --hash=sha256:de8a88e63464af587c950061a5e6a67d3632e36df62b986892331d4620a35c01 \\\n    --hash=sha256:df2449253ef108a379b8b5d6b43f4b1a8e81a061d6537becd5582fba5f9196d7 \\\n    --hash=sha256:e1c1493fb6e50ab01d20a22826e57520f1284df32f2d8601fdd90b6304601419 \\\n    --hash=sha256:e1cf1972137e83c5d4c136c43ced9ac51d0e124706ee1c8aa8532c1287fa8795 \\\n    --hash=sha256:e2103a929dfa2fcaf9bb4e7c091983a49c9ac3b19c9061b6d5427dd7d14d81a1 \\\n    --hash=sha256:e56b7d45a839a697b5eb268c82a71bd8c7f6c94d6fd50c3d577fa39a9f1409f5 \\\n    --hash=sha256:e8afc3f2ccfa24215f8cb28dcf43f0113ac3c37c2f0f0806d8c70e4228c5cf4d \\\n    --hash=sha256:e8fc20152abba6b83724d7ff268c249fa196d8259ff481f3b1476383f8f24e42 \\\n    --hash=sha256:eaa9599de571d72e2daf60164784109f19978b327a3910d3e9de8c97b5b70cfe \\\n    --hash=sha256:ec15a59cf5af7be74194f7ab02d0f59a62bdcf1a537677ce67a2537c9b87fcda \\\n    --hash=sha256:f190daf01f13c72eac4efd5c430a8de82489d9cff23c364c3ea822545032993e \\\n    --hash=sha256:f34c41761022dd093b4b6896d4810782ffbabe30f2d443ff5f083e0cbbb8c737 \\\n    --hash=sha256:f3e98bb3798ead92273dc0e5fd0f31ade220f59a266ffd8a4f6065e0a3ce0523 \\\n    --hash=sha256:f42d0984e947b8adf7dd6dde396e720934d12c506ce84eea8476409563607591 \\\n    --hash=sha256:f71a396b3bf33ecaa1626c255855702aca4d3d9fea5e051b41ac59a9c1c41edc \\\n    --hash=sha256:f9e130248f4462aaa8e2552d547f36ddadbeaa573879158d721bbd33dfe4743a \\\n    --hash=sha256:fed51ac40f757d41b7c48425901843666a6677e3e8eb0abcff09e4ba6e664f50\nmatplotlib==3.10.7 \\\n    --hash=sha256:07124afcf7a6504eafcb8ce94091c5898bbdd351519a1beb5c45f7a38c67e77f \\\n    --hash=sha256:09d7945a70ea43bf9248f4b6582734c2fe726723204a76eca233f24cffc7ef67 \\\n    --hash=sha256:0d8c32b7ea6fb80b1aeff5a2ceb3fb9778e2759e899d9beff75584714afcc5ee \\\n    --hash=sha256:11ae579ac83cdf3fb72573bb89f70e0534de05266728740d478f0f818983c695 \\\n    --hash=sha256:15112bcbaef211bd663fa935ec33313b948e214454d949b723998a43357b17b0 \\\n    --hash=sha256:1d9d3713a237970569156cfb4de7533b7c4eacdd61789726f444f96a0d28f57f \\\n    --hash=sha256:1e4bbad66c177a8fdfa53972e5ef8be72a5f27e6a607cec0d8579abd0f3102b1 \\\n    --hash=sha256:2222c7ba2cbde7fe63032769f6eb7e83ab3227f47d997a8453377709b7fe3a5a \\\n    --hash=sha256:22df30ffaa89f6643206cf13877191c63a50e8f800b038bc39bee9d2d4957632 \\\n    --hash=sha256:31963603041634ce1a96053047b40961f7a29eb8f9a62e80cc2c0427aa1d22a2 \\\n    --hash=sha256:37a1fea41153dd6ee061d21ab69c9cf2cf543160b1b85d89cd3d2e2a7902ca4c \\\n    --hash=sha256:3886e47f64611046bc1db523a09dd0a0a6bed6081e6f90e13806dd1d1d1b5e91 \\\n    --hash=sha256:4645fc5d9d20ffa3a39361fcdbcec731382763b623b72627806bf251b6388866 \\\n    --hash=sha256:4a11c2e9e72e7de09b7b72e62f3df23317c888299c875e2b778abf1eda8c0a42 \\\n    --hash=sha256:4a74f79fafb2e177f240579bc83f0b60f82cc47d2f1d260f422a0627207008ca \\\n    --hash=sha256:4c14b6acd16cddc3569a2d515cfdd81c7a68ac5639b76548cfc1a9e48b20eb65 \\\n    --hash=sha256:53b492410a6cd66c7a471de6c924f6ede976e963c0f3097a3b7abfadddc67d0a \\\n    --hash=sha256:53cc80662dd197ece414dd5b66e07370201515a3eaf52e7c518c68c16814773b \\\n    --hash=sha256:5c09cf8f2793f81368f49f118b6f9f937456362bee282eac575cca7f84cda537 \\\n    --hash=sha256:5e38c2d581d62ee729a6e144c47a71b3f42fb4187508dbbf4fe71d5612c3433b \\\n    --hash=sha256:5f3f6d315dcc176ba7ca6e74c7768fb7e4cf566c49cb143f6bc257b62e634ed8 \\\n    --hash=sha256:6516ce375109c60ceec579e699524e9d504cd7578506f01150f7a6bc174a775e \\\n    --hash=sha256:667ecd5d8d37813a845053d8f5bf110b534c3c9f30e69ebd25d4701385935a6d \\\n    --hash=sha256:6f1851eab59ca082c95df5a500106bad73672645625e04538b3ad0f69471ffcc \\\n    --hash=sha256:702590829c30aada1e8cef0568ddbffa77ca747b4d6e36c6d173f66e301f89cc \\\n    --hash=sha256:7146d64f561498764561e9cd0ed64fcf582e570fc519e6f521e2d0cfd43365e1 \\\n    --hash=sha256:744991e0cc863dd669c8dc9136ca4e6e0082be2070b9d793cbd64bec872a6815 \\\n    --hash=sha256:786656bb13c237bbcebcd402f65f44dd61ead60ee3deb045af429d889c8dbc67 \\\n    --hash=sha256:7a0edb7209e21840e8361e91ea84ea676658aa93edd5f8762793dec77a4a6748 \\\n    --hash=sha256:7ac81eee3b7c266dd92cee1cd658407b16c57eed08c7421fa354ed68234de380 \\\n    --hash=sha256:90ad854c0a435da3104c01e2c6f0028d7e719b690998a2333d7218db80950722 \\\n    --hash=sha256:9257be2f2a03415f9105c486d304a321168e61ad450f6153d77c69504ad764bb \\\n    --hash=sha256:932c55d1fa7af4423422cb6a492a31cbcbdbe68fd1a9a3f545aa5e7a143b5355 \\\n    --hash=sha256:a06ba7e2a2ef9131c79c49e63dad355d2d878413a0376c1727c8b9335ff731c7 \\\n    --hash=sha256:aebed7b50aa6ac698c90f60f854b47e48cd2252b30510e7a1feddaf5a3f72cbf \\\n    --hash=sha256:b172db79759f5f9bc13ef1c3ef8b9ee7b37b0247f987fbbbdaa15e4f87fd46a9 \\\n    --hash=sha256:b3c4ea4948d93c9c29dc01c0c23eef66f2101bf75158c291b88de6525c55c3d1 \\\n    --hash=sha256:b498e9e4022f93de2d5a37615200ca01297ceebbb56fe4c833f46862a490f9e3 \\\n    --hash=sha256:b4d41379b05528091f00e1728004f9a8d7191260f3862178b88e8fd770206318 \\\n    --hash=sha256:b69676845a0a66f9da30e87f48be36734d6748024b525ec4710be40194282c84 \\\n    --hash=sha256:c17398b709a6cce3d9fdb1595c33e356d91c098cd9486cb2cc21ea2ea418e715 \\\n    --hash=sha256:c380371d3c23e0eadf8ebff114445b9f970aff2010198d498d4ab4c3b41eea4f \\\n    --hash=sha256:cb783436e47fcf82064baca52ce748af71725d0352e1d31564cbe9c95df92b9c \\\n    --hash=sha256:cc1c51b846aca49a5a8b44fbba6a92d583a35c64590ad9e1e950dc88940a4297 \\\n    --hash=sha256:d0b181e9fa8daf1d9f2d4c547527b167cb8838fc587deabca7b5c01f97199e84 \\\n    --hash=sha256:d2a959c640cdeecdd2ec3136e8ea0441da59bcaf58d67e9c590740addba2cb68 \\\n    --hash=sha256:d5f256d49fea31f40f166a5e3131235a5d2f4b7f44520b1cf0baf1ce568ccff0 \\\n    --hash=sha256:d883460c43e8c6b173fef244a2341f7f7c0e9725c7fe68306e8e44ed9c8fb100 \\\n    --hash=sha256:d8eb7194b084b12feb19142262165832fc6ee879b945491d1c3d4660748020c4 \\\n    --hash=sha256:d9749313deb729f08207718d29c86246beb2ea3fdba753595b55901dee5d2fd6 \\\n    --hash=sha256:de66744b2bb88d5cd27e80dfc2ec9f0517d0a46d204ff98fe9e5f2864eb67657 \\\n    --hash=sha256:e91f61a064c92c307c5a9dc8c05dc9f8a68f0a3be199d9a002a0622e13f874a1 \\\n    --hash=sha256:f19410b486fdd139885ace124e57f938c1e6a3210ea13dd29cab58f5d4bc12c7 \\\n    --hash=sha256:f79d5de970fc90cd5591f60053aecfce1fcd736e0303d9f0bf86be649fa68fb8 \\\n    --hash=sha256:fba2974df0bf8ce3c995fa84b79cde38326e0f7b5409e7a3a481c1141340bcf7\nmdurl==0.1.2 \\\n    --hash=sha256:84008a41e51615a49fc9966191ff91509e3c40b939176e643fd50a5c2196b8f8 \\\n    --hash=sha256:bb413d29f5eea38f31dd4754dd7377d4465116fb207585f97bf925588687c1ba\nmpmath==1.3.0 \\\n    --hash=sha256:7a28eb2a9774d00c7bc92411c19a89209d5da7c4c9a9e227be8330a23a25b91f \\\n    --hash=sha256:a0b2b9fe80bbcd81a6647ff13108738cfb482d481d826cc0e02f5b35e5c88d2c\nmsgspec==0.21.1 \\\n    --hash=sha256:0d03867786e5d7ba25d666df4b11320c27170f4aeafcb8e3a8b0a50a4fb742ca \\\n    --hash=sha256:0d1009f6715f5bff3b54d4ff5c7428ad96197e0534e1645b8e9b955890c84664 \\\n    --hash=sha256:0d2cc73df6058d811a126ac3a8ad63a4dfa210c82f9cf5a004802eaf4712de90 \\\n    --hash=sha256:15f523d51c00ebad412213bfe9f06f0a50ec2b93e0c19e824a2d267cabb48ea2 \\\n    --hash=sha256:1bf17cbd7b28a5dffc7e764c654eed8ccde5e0f1de7970628608304640d4ce4e \\\n    --hash=sha256:21995e74b5c598c2e004110ad66ec7f1b8c20bf2bcf3b2de8fd9a3094422d3ff \\\n    --hash=sha256:2313508e394b0d208f8f56892ca9b2799e2561329de9763b19619595a6c0f72c \\\n    --hash=sha256:344c7cd0eaed1fb81d7959f99100ef71ec9b536881a376f11b9a6c4803365697 \\\n    --hash=sha256:38fe93e86b61328fe544cb7fd871fad5a27c8734bfda90f65e5dbe288ae50f61 \\\n    --hash=sha256:3cb779ea0c35bc807ff941d415875c1f69ca0be91a2e907ab99a171811d86a9a \\\n    --hash=sha256:3d6b9dc50948eaf65df54d2fd0ff66e6d8c32f116037209ee861810eb9b676cb \\\n    --hash=sha256:42bb1241e0750c1a4346f2aa84db26c5ffd99a4eb3a954927d9f149ff2f42898 \\\n    --hash=sha256:4692b7c1609155708c4418f88e92f63c13fdf08aa095c84bae82bad75b53389b \\\n    --hash=sha256:48943e278b3854c2f89f955ddc6f9f430d3f0784b16e47d10604ee0463cd21f5 \\\n    --hash=sha256:49880fd20fdbcfe1b793f07dd83f12572bab679c9800352c8b2240289aa46a06 \\\n    --hash=sha256:4e47390360583ba3d5c6cb44cf0a9f61b0a06a899d3c2c00627cedebb2e2884b \\\n    --hash=sha256:5102c7e9b3acff82178449b85006d96310e690291bb1ea0142f1b24bcb8aabcb \\\n    --hash=sha256:52c5e21930942302394429c5a582ce7e6b62c7f983b3760834c2ce107e0dd6df \\\n    --hash=sha256:5666b1b560b97b6ec2eb3fca8a502298ebac56e13bbca1f88523538ce83d01ea \\\n    --hash=sha256:5d2d4116ebe3035a78d9ec76e99a9d64e5fa6d44fe61a9c5de7fd1acf54bcc69 \\\n    --hash=sha256:5f8e9dfcd98419cf7568808470c4317a3fb30bef0e3715b568730a2b272a20d7 \\\n    --hash=sha256:6129f0cca52992e898fd5344187f7c8127b63d810b2fd73e36fca73b4c6475ee \\\n    --hash=sha256:628aaa35c74950a8c59da330d7e98917e1c7188f983745782027748ee4ca573e \\\n    --hash=sha256:68604db36b3b4dd9bf160e436e12798a4738848144cea1aca1cb984011eb160f \\\n    --hash=sha256:6badc03b9725352219cca017bfe71c61f2fbd0fb5982b410ac17c97c213deb30 \\\n    --hash=sha256:72d9cd03241b8b2edb2e12dcc66c500fa480d8cbd71a8bac105809d468882064 \\\n    --hash=sha256:740fbf1c9d59992ca3537d6fbe9ebbf9eaf726a65fbf31448e0ecbc710697a63 \\\n    --hash=sha256:764173717a01743f007e9f74520ed281f24672c604514f7d76c1c3a10e8edb66 \\\n    --hash=sha256:846758412e9518252b2ac9bffd6f0e54d9ff614f5f9488df7749f81ff5c80920 \\\n    --hash=sha256:8bc666331c35fcce05a7cd2d6221adbe0f6058f8e750711413d22793c080ac6a \\\n    --hash=sha256:92d89dfad13bd1ea640dc3e37e724ed380da1030b272bdf5ecafb983c3ad7c75 \\\n    --hash=sha256:a9aa659ebb0101b1cbc31461212b87e341d961f0ab0772aaf068a99e001ec4aa \\\n    --hash=sha256:abbb39d65681fa24ed394e01af3d59d869068324f900c61d06062b7fb9980f2f \\\n    --hash=sha256:ae0162e22849a5e91eaad907766525107523b0daea3df267a9fcb5ba4e0936ae \\\n    --hash=sha256:b504b6e7f7a22a24b27232b73034421692147865162daaec9f3bf62439007c87 \\\n    --hash=sha256:c6faffe5bb644ec884052679af4dfd776d4b5ca90e4a7ec7e7e319e4e6b93a6e \\\n    --hash=sha256:d3124010b3815451494c85ff345e693cb9fe5889cfcbbef39ed8622e0e72319c \\\n    --hash=sha256:d4248cf0b6129b7d230eacd493c17cc2d4f3989f3bb7f633a928a85b7dcfa251 \\\n    --hash=sha256:d4ab834a054c6f0cbeef6df9e7e1b33d5f1bc7b86dea1d2fd7cad003873e783d \\\n    --hash=sha256:d8b8578e4c83b14ceea4cef0d0b747e31d9330fe4b03b2b2ad4063866a178f93 \\\n    --hash=sha256:dd677e3001fdfed9186de72eab434da2976303cd5eb9550921d3d0c3e3e168ce \\\n    --hash=sha256:ed2ab278200e743a1d2610a4e0c8fc74f6cecb8548544cdec43f927bd9265238 \\\n    --hash=sha256:ee9e3f11fa94603f7d673bf795cfa31b549c4a2c723bc39b45beb1e7f5a3fb99 \\\n    --hash=sha256:ef3ec2296248d1f8b9231acb051b6d471dfde8f21819e86c9adaaa9f42918521 \\\n    --hash=sha256:f041a2279f31e3a53319005e4d60ba77c085cfcbe394cdc7ce803c2d01fe9449 \\\n    --hash=sha256:f60800e6299b798142dc40b0644da77ceac5ea0568be58228417eae14135c847 \\\n    --hash=sha256:f667b90b37fad734a91671abd68e0d7f4d066862771b87e91c53996dcb7a9027 \\\n    --hash=sha256:f7b27d1a8ead2b6f5b0c4f2d07b8be1ccfcc041c8a0e704781edebe3ae13c484 \\\n    --hash=sha256:fab48eb45fdbfbdb2c0edfec00ffc53b6b6085beefc6b50b61e01659f9f8757f\nmultidict==6.7.1 \\\n    --hash=sha256:026d264228bcd637d4e060844e39cdc60f86c479e463d49075dedc21b18fbbe0 \\\n    --hash=sha256:03ede2a6ffbe8ef936b92cb4529f27f42be7f56afcdab5ab739cd5f27fb1cbf9 \\\n    --hash=sha256:0458c978acd8e6ea53c81eefaddbbee9c6c5e591f41b3f5e8e194780fe026581 \\\n    --hash=sha256:067343c68cd6612d375710f895337b3a98a033c94f14b9a99eff902f205424e2 \\\n    --hash=sha256:08ccb2a6dc72009093ebe7f3f073e5ec5964cba9a706fa94b1a1484039b87941 \\\n    --hash=sha256:0b38ebffd9be37c1170d33bc0f36f4f262e0a09bc1aac1c34c7aa51a7293f0b3 \\\n    --hash=sha256:0b4c48648d7649c9335cf1927a8b87fa692de3dcb15faa676c6a6f1f1aabda43 \\\n    --hash=sha256:0d17522c37d03e85c8098ec8431636309b2682cf12e58f4dbc76121fb50e4962 \\\n    --hash=sha256:0e161ddf326db5577c3a4cc2d8648f81456e8a20d40415541587a71620d7a7d1 \\\n    --hash=sha256:0e697826df7eb63418ee190fd06ce9f1803593bb4b9517d08c60d9b9a7f69d8f \\\n    --hash=sha256:10ae39c9cfe6adedcdb764f5e8411d4a92b055e35573a2eaa88d3323289ef93c \\\n    --hash=sha256:121a34e5bfa410cdf2c8c49716de160de3b1dbcd86b49656f5681e4543bcd1a8 \\\n    --hash=sha256:128441d052254f42989ef98b7b6a6ecb1e6f708aa962c7984235316db59f50fa \\\n    --hash=sha256:12fad252f8b267cc75b66e8fc51b3079604e8d43a75428ffe193cd9e2195dfd6 \\\n    --hash=sha256:14525a5f61d7d0c94b368a42cff4c9a4e7ba2d52e2672a7b23d84dc86fb02b0c \\\n    --hash=sha256:17207077e29342fdc2c9a82e4b306f1127bf1ea91f8b71e02d4798a70bb99991 \\\n    --hash=sha256:17307b22c217b4cf05033dabefe68255a534d637c6c9b0cc8382718f87be4262 \\\n    --hash=sha256:1b99af4d9eec0b49927b4402bcbb58dea89d3e0db8806a4086117019939ad3dd \\\n    --hash=sha256:1d540e51b7e8e170174555edecddbd5538105443754539193e3e1061864d444d \\\n    --hash=sha256:1e3a8bb24342a8201d178c3b4984c26ba81a577c80d4d525727427460a50c22d \\\n    --hash=sha256:1fa6609d0364f4f6f58351b4659a1f3e0e898ba2a8c5cac04cb2c7bc556b0bc5 \\\n    --hash=sha256:21f830fe223215dffd51f538e78c172ed7c7f60c9b96a2bf05c4848ad49921c3 \\\n    --hash=sha256:233b398c29d3f1b9676b4b6f75c518a06fcb2ea0b925119fb2c1bc35c05e1601 \\\n    --hash=sha256:24c0cf81544ca5e17cfcb6e482e7a82cd475925242b308b890c9452a074d4505 \\\n    --hash=sha256:25167cc263257660290fba06b9318d2026e3c910be240a146e1f66dd114af2b0 \\\n    --hash=sha256:253282d70d67885a15c8a7716f3a73edf2d635793ceda8173b9ecc21f2fb8292 \\\n    --hash=sha256:273d23f4b40f3dce4d6c8a821c741a86dec62cded82e1175ba3d99be128147ed \\\n    --hash=sha256:283ddac99f7ac25a4acadbf004cb5ae34480bbeb063520f70ce397b281859362 \\\n    --hash=sha256:28ca5ce2fd9716631133d0e9a9b9a745ad7f60bac2bccafb56aa380fc0b6c511 \\\n    --hash=sha256:2b41f5fed0ed563624f1c17630cb9941cf2309d4df00e494b551b5f3e3d67a23 \\\n    --hash=sha256:2bbd113e0d4af5db41d5ebfe9ccaff89de2120578164f86a5d17d5a576d1e5b2 \\\n    --hash=sha256:2e1425e2f99ec5bd36c15a01b690a1a2456209c5deed58f95469ffb46039ccbb \\\n    --hash=sha256:2e2d2ed645ea29f31c4c7ea1552fcfd7cb7ba656e1eafd4134a6620c9f5fdd9e \\\n    --hash=sha256:3758692429e4e32f1ba0df23219cd0b4fc0a52f476726fff9337d1a57676a582 \\\n    --hash=sha256:38fb49540705369bab8484db0689d86c0a33a0a9f2c1b197f506b71b4b6c19b0 \\\n    --hash=sha256:3943debf0fbb57bdde5901695c11094a9a36723e5c03875f87718ee15ca2f4d2 \\\n    --hash=sha256:398c1478926eca669f2fd6a5856b6de9c0acf23a2cb59a14c0ba5844fa38077e \\\n    --hash=sha256:3ab8b9d8b75aef9df299595d5388b14530839f6422333357af1339443cff777d \\\n    --hash=sha256:3bd231490fa7217cc832528e1cd8752a96f0125ddd2b5749390f7c3ec8721b65 \\\n    --hash=sha256:3d51ff4785d58d3f6c91bdbffcb5e1f7ddfda557727043aa20d20ec4f65e324a \\\n    --hash=sha256:3fccb473e87eaa1382689053e4a4618e7ba7b9b9b8d6adf2027ee474597128cd \\\n    --hash=sha256:401c5a650f3add2472d1d288c26deebc540f99e2fb83e9525007a74cd2116f1d \\\n    --hash=sha256:41f2952231456154ee479651491e94118229844dd7226541788be783be2b5108 \\\n    --hash=sha256:432feb25a1cb67fe82a9680b4d65fb542e4635cb3166cd9c01560651ad60f177 \\\n    --hash=sha256:439cbebd499f92e9aa6793016a8acaa161dfa749ae86d20960189f5398a19144 \\\n    --hash=sha256:4885cb0e817aef5d00a2e8451d4665c1808378dc27c2705f1bf4ef8505c0d2e5 \\\n    --hash=sha256:497394b3239fc6f0e13a78a3e1b61296e72bf1c5f94b4c4eb80b265c37a131cd \\\n    --hash=sha256:497bde6223c212ba11d462853cfa4f0ae6ef97465033e7dc9940cdb3ab5b48e5 \\\n    --hash=sha256:4cfb48c6ea66c83bcaaf7e4dfa7ec1b6bbcf751b7db85a328902796dfde4c060 \\\n    --hash=sha256:538cec1e18c067d0e6103aa9a74f9e832904c957adc260e61cd9d8cf0c3b3d37 \\\n    --hash=sha256:55d97cc6dae627efa6a6e548885712d4864b81110ac76fa4e534c03819fa4a56 \\\n    --hash=sha256:563fe25c678aaba333d5399408f5ec3c383ca5b663e7f774dd179a520b8144df \\\n    --hash=sha256:57b46b24b5d5ebcc978da4ec23a819a9402b4228b8a90d9c656422b4bdd8a963 \\\n    --hash=sha256:5884a04f4ff56c6120f6ccf703bdeb8b5079d808ba604d4d53aec0d55dc33568 \\\n    --hash=sha256:59bc83d3f66b41dac1e7460aac1d196edc70c9ba3094965c467715a70ecb46db \\\n    --hash=sha256:5a37ca18e360377cfda1d62f5f382ff41f2b8c4ccb329ed974cc2e1643440118 \\\n    --hash=sha256:5c4b9bfc148f5a91be9244d6264c53035c8a0dcd2f51f1c3c6e30e30ebaa1c84 \\\n    --hash=sha256:5e01429a929600e7dab7b166062d9bb54a5eed752384c7384c968c2afab8f50f \\\n    --hash=sha256:5fa6a95dfee63893d80a34758cd0e0c118a30b8dcb46372bf75106c591b77889 \\\n    --hash=sha256:619e5a1ac57986dbfec9f0b301d865dddf763696435e2962f6d9cf2fdff2bb71 \\\n    --hash=sha256:65573858d27cdeaca41893185677dc82395159aa28875a8867af66532d413a8f \\\n    --hash=sha256:6704fa2b7453b2fb121740555fa1ee20cd98c4d011120caf4d2b8d4e7c76eec0 \\\n    --hash=sha256:6aac4f16b472d5b7dc6f66a0d49dd57b0e0902090be16594dc9ebfd3d17c47e7 \\\n    --hash=sha256:6b10359683bd8806a200fd2909e7c8ca3a7b24ec1d8132e483d58e791d881048 \\\n    --hash=sha256:6b83cabdc375ffaaa15edd97eb7c0c672ad788e2687004990074d7d6c9b140c8 \\\n    --hash=sha256:6d3bc717b6fe763b8be3f2bee2701d3c8eb1b2a8ae9f60910f1b2860c82b6c49 \\\n    --hash=sha256:6f77ce314a29263e67adadc7e7c1bc699fcb3a305059ab973d038f87caa42ed0 \\\n    --hash=sha256:749aa54f578f2e5f439538706a475aa844bfa8ef75854b1401e6e528e4937cf9 \\\n    --hash=sha256:7a7e590ff876a3eaf1c02a4dfe0724b6e69a9e9de6d8f556816f29c496046e59 \\\n    --hash=sha256:7dfb78d966b2c906ae1d28ccf6e6712a3cd04407ee5088cd276fe8cb42186190 \\\n    --hash=sha256:7eee46ccb30ff48a1e35bb818cc90846c6be2b68240e42a78599166722cea709 \\\n    --hash=sha256:7ff981b266af91d7b4b3793ca3382e53229088d193a85dfad6f5f4c27fc73e5d \\\n    --hash=sha256:841189848ba629c3552035a6a7f5bf3b02eb304e9fea7492ca220a8eda6b0e5c \\\n    --hash=sha256:844c5bca0b5444adb44a623fb0a1310c2f4cd41f402126bb269cd44c9b3f3e1e \\\n    --hash=sha256:84e61e3af5463c19b67ced91f6c634effb89ef8bfc5ca0267f954451ed4bb6a2 \\\n    --hash=sha256:8affcf1c98b82bc901702eb73b6947a1bfa170823c153fe8a47b5f5f02e48e40 \\\n    --hash=sha256:8be1802715a8e892c784c0197c2ace276ea52702a0ede98b6310c8f255a5afb3 \\\n    --hash=sha256:8f333ec9c5eb1b7105e3b84b53141e66ca05a19a605368c55450b6ba208cb9ee \\\n    --hash=sha256:9004d8386d133b7e6135679424c91b0b854d2d164af6ea3f289f8f2761064609 \\\n    --hash=sha256:90efbcf47dbe33dcf643a1e400d67d59abeac5db07dc3f27d6bdeae497a2198c \\\n    --hash=sha256:935434b9853c7c112eee7ac891bc4cb86455aa631269ae35442cb316790c1445 \\\n    --hash=sha256:93b1818e4a6e0930454f0f2af7dfce69307ca03cdcfb3739bf4d91241967b6c1 \\\n    --hash=sha256:95922cee9a778659e91db6497596435777bd25ed116701a4c034f8e46544955a \\\n    --hash=sha256:960c83bf01a95b12b08fd54324a4eb1d5b52c88932b5cba5d6e712bb3ed12eb5 \\\n    --hash=sha256:97231140a50f5d447d3164f994b86a0bed7cd016e2682f8650d6a9158e14fd31 \\\n    --hash=sha256:974e72a2474600827abaeda71af0c53d9ebbc3c2eb7da37b37d7829ae31232d8 \\\n    --hash=sha256:97891f3b1b3ffbded884e2916cacf3c6fc87b66bb0dde46f7357404750559f33 \\\n    --hash=sha256:98655c737850c064a65e006a3df7c997cd3b220be4ec8fe26215760b9697d4d7 \\\n    --hash=sha256:98bc624954ec4d2c7cb074b8eefc2b5d0ce7d482e410df446414355d158fe4ca \\\n    --hash=sha256:98c5787b0a0d9a41d9311eae44c3b76e6753def8d8870ab501320efe75a6a5f8 \\\n    --hash=sha256:9b0d9b91d1aa44db9c1f1ecd0d9d2ae610b2f4f856448664e01a3b35899f3f92 \\\n    --hash=sha256:9c90fed18bffc0189ba814749fdcc102b536e83a9f738a9003e569acd540a733 \\\n    --hash=sha256:9d624335fd4fa1c08a53f8b4be7676ebde19cd092b3895c421045ca87895b429 \\\n    --hash=sha256:9f9af11306994335398293f9958071019e3ab95e9a707dc1383a35613f6abcb9 \\\n    --hash=sha256:a0543217a6a017692aa6ae5cc39adb75e587af0f3a82288b1492eb73dd6cc2a4 \\\n    --hash=sha256:a088b62bd733e2ad12c50dad01b7d0166c30287c166e137433d3b410add807a6 \\\n    --hash=sha256:a407f13c188f804c759fc6a9f88286a565c242a76b27626594c133b82883b5c2 \\\n    --hash=sha256:a90f75c956e32891a4eda3639ce6dd86e87105271f43d43442a3aedf3cddf172 \\\n    --hash=sha256:a9fc4caa29e2e6ae408d1c450ac8bf19892c5fca83ee634ecd88a53332c59981 \\\n    --hash=sha256:aa23b001d968faef416ff70dc0f1ab045517b9b42a90edd3e9bcdb06479e31d5 \\\n    --hash=sha256:ac1c665bad8b5d762f5f85ebe4d94130c26965f11de70c708c75671297c776de \\\n    --hash=sha256:af959b9beeb66c822380f222f0e0a1889331597e81f1ded7f374f3ecb0fd6c52 \\\n    --hash=sha256:b0fa96985700739c4c7853a43c0b3e169360d6855780021bfc6d0f1ce7c123e7 \\\n    --hash=sha256:b26684587228afed0d50cf804cc71062cc9c1cdf55051c4c6345d372947b268c \\\n    --hash=sha256:b4938326284c4f1224178a560987b6cf8b4d38458b113d9b8c1db1a836e640a2 \\\n    --hash=sha256:b8c990b037d2fff2f4e33d3f21b9b531c5745b33a49a7d6dbe7a177266af44f6 \\\n    --hash=sha256:ba0a9fb644d0c1a2194cf7ffb043bd852cea63a57f66fbd33959f7dae18517bf \\\n    --hash=sha256:bb08271280173720e9fea9ede98e5231defcbad90f1624bea26f32ec8a956e2f \\\n    --hash=sha256:bdbf9f3b332abd0cdb306e7c2113818ab1e922dc84b8f8fd06ec89ed2a19ab8b \\\n    --hash=sha256:bfde23ef6ed9db7eaee6c37dcec08524cb43903c60b285b172b6c094711b3961 \\\n    --hash=sha256:c0abd12629b0af3cf590982c0b413b1e7395cd4ec026f30986818ab95bfaa94a \\\n    --hash=sha256:c102791b1c4f3ab36ce4101154549105a53dc828f016356b3e3bcae2e3a039d3 \\\n    --hash=sha256:c3a32d23520ee37bf327d1e1a656fec76a2edd5c038bf43eddfa0572ec49c60b \\\n    --hash=sha256:c524c6fb8fc342793708ab111c4dbc90ff9abd568de220432500e47e990c0358 \\\n    --hash=sha256:c5f0c21549ab432b57dcc82130f388d84ad8179824cc3f223d5e7cfbfd4143f6 \\\n    --hash=sha256:c6b3228e1d80af737b72925ce5fb4daf5a335e49cd7ab77ed7b9fdfbf58c526e \\\n    --hash=sha256:c76c4bec1538375dad9d452d246ca5368ad6e1c9039dadcf007ae59c70619ea1 \\\n    --hash=sha256:c9035dde0f916702850ef66460bc4239d89d08df4d02023a5926e7446724212c \\\n    --hash=sha256:c93c3db7ea657dd4637d57e74ab73de31bccefe144d3d4ce370052035bc85fb5 \\\n    --hash=sha256:cb2a55f408c3043e42b40cc8eecd575afa27b7e0b956dfb190de0f8499a57a53 \\\n    --hash=sha256:cdea2e7b2456cfb6694fb113066fd0ec7ea4d67e3a35e1f4cbeea0b448bf5872 \\\n    --hash=sha256:ce1bbd7d780bb5a0da032e095c951f7014d6b0a205f8318308140f1a6aba159e \\\n    --hash=sha256:cf37cbe5ced48d417ba045aca1b21bafca67489452debcde94778a576666a1df \\\n    --hash=sha256:d4f49cb5661344764e4c7c7973e92a47a59b8fc19b6523649ec9dc4960e58a03 \\\n    --hash=sha256:d54ecf9f301853f2c5e802da559604b3e95bb7a3b01a9c295c6ee591b9882de8 \\\n    --hash=sha256:d62b7f64ffde3b99d06b707a280db04fb3855b55f5a06df387236051d0668f4a \\\n    --hash=sha256:d82dd730a95e6643802f4454b8fdecdf08667881a9c5670db85bc5a56693f122 \\\n    --hash=sha256:da62917e6076f512daccfbbde27f46fed1c98fee202f0559adec8ee0de67f71a \\\n    --hash=sha256:dd96c01a9dcd4889dcfcf9eb5544ca0c77603f239e3ffab0524ec17aea9a93ee \\\n    --hash=sha256:df9f19c28adcb40b6aae30bbaa1478c389efd50c28d541d76760199fc1037c32 \\\n    --hash=sha256:e1c5988359516095535c4301af38d8a8838534158f649c05dd1050222321bcb3 \\\n    --hash=sha256:e628ef0e6859ffd8273c69412a2465c4be4a9517d07261b33334b5ec6f3c7489 \\\n    --hash=sha256:e82d14e3c948952a1a85503817e038cba5905a3352de76b9a465075d072fba23 \\\n    --hash=sha256:e954b24433c768ce78ab7929e84ccf3422e46deb45a4dc9f93438f8217fa2d34 \\\n    --hash=sha256:eb0ce7b2a32d09892b3dd6cc44877a0d02a33241fafca5f25c8b6b62374f8b75 \\\n    --hash=sha256:eb304767bca2bb92fb9c5bd33cedc95baee5bb5f6c88e63706533a1c06ad08c8 \\\n    --hash=sha256:eb351f72c26dc9abe338ca7294661aa22969ad8ffe7ef7d5541d19f368dc854a \\\n    --hash=sha256:ec6652a1bee61c53a3e5776b6049172c53b6aaba34f18c9ad04f82712bac623d \\\n    --hash=sha256:f2a0a924d4c2e9afcd7ec64f9de35fcd96915149b2216e1cb2c10a56df483855 \\\n    --hash=sha256:f33dc2a3abe9249ea5d8360f969ec7f4142e7ac45ee7014d8f8d5acddf178b7b \\\n    --hash=sha256:f537b55778cd3cbee430abe3131255d3a78202e0f9ea7ffc6ada893a4bcaeea4 \\\n    --hash=sha256:f5dd81c45b05518b9aa4da4aa74e1c93d715efa234fd3e8a179df611cc85e5f4 \\\n    --hash=sha256:f99fe611c312b3c1c0ace793f92464d8cd263cc3b26b5721950d977b006b6c4d \\\n    --hash=sha256:fa263a02f4f2dd2d11a7b1bb4362aa7cb1049f84a9235d31adf63f30143469a0 \\\n    --hash=sha256:fc5907494fccf3e7d3f94f95c91d6336b092b5fc83811720fae5e2765890dfba \\\n    --hash=sha256:fcee94dfbd638784645b066074b338bc9cc155d4b4bffa4adce1615c5a426c19\nmultiprocess==0.70.16 \\\n    --hash=sha256:0dfd078c306e08d46d7a8d06fb120313d87aa43af60d66da43ffff40b44d2f41 \\\n    --hash=sha256:161af703d4652a0e1410be6abccecde4a7ddffd19341be0a7011b94aeb171ac1 \\\n    --hash=sha256:37b55f71c07e2d741374998c043b9520b626a8dddc8b3129222ca4f1a06ef67a \\\n    --hash=sha256:476887be10e2f59ff183c006af746cb6f1fd0eadcfd4ef49e605cbe2659920ee \\\n    --hash=sha256:a0bafd3ae1b732eac64be2e72038231c1ba97724b60b09400d68f229fcc2fbf3 \\\n    --hash=sha256:a71d82033454891091a226dfc319d0cfa8019a4e888ef9ca910372a446de4435 \\\n    --hash=sha256:af4cabb0dac72abfb1e794fa7855c325fd2b55a10a44628a3c1ad3311c04127a \\\n    --hash=sha256:ba8c31889abf4511c7308a8c52bb4a30b9d590e7f58523302ba00237702ca054 \\\n    --hash=sha256:c4a9944c67bd49f823687463660a2d6daae94c289adff97e0f9d696ba6371d02 \\\n    --hash=sha256:d951bed82c8f73929ac82c61f01a7b5ce8f3e5ef40f5b52553b4f547ce2b08ec \\\n    --hash=sha256:e7b9d0f307cd9bd50851afaac0dba2cb6c44449efff697df7c7645f7d3f2be3a \\\n    --hash=sha256:fc0544c531920dde3b00c29863377f87e1632601092ea2daca74e4beb40faa2e\nnetworkx==3.6.1 \\\n    --hash=sha256:26b7c357accc0c8cde558ad486283728b65b6a95d85ee1cd66bafab4c8168509 \\\n    --hash=sha256:d47fbf302e7d9cbbb9e2555a0d267983d2aa476bac30e90dfbe5669bd57f3762\nnltk==3.9.1 \\\n    --hash=sha256:4fa26829c5b00715afe3061398a8989dc643b92ce7dd93fb4585a70930d168a1 \\\n    --hash=sha256:87d127bd3de4bd89a4f81265e5fa59cb1b199b27440175370f7417d2bc7ae868\nnumpy==2.4.6 \\\n    --hash=sha256:001fbb8e08d942dd57599e781f2472269ee7f2755fae407b4f67b2f0b17da3f1 \\\n    --hash=sha256:0280e0356c0829a18d9de1cb7eee50ec22ca639878d7240307ca0943d73cd2c4 \\\n    --hash=sha256:043191bfa8eab18c776647b62723ac9dddece59743b13f49b2016094129c2b3f \\\n    --hash=sha256:06ca2f61ec4385a07a6977c55ba998a4466c123642b4a32694d3128fce18c079 \\\n    --hash=sha256:0a041d3d761dc3c35cc56ce0351506a02bcbc25f7b169f652435141a17db9096 \\\n    --hash=sha256:0ab0a9c4ffb1a6d95ef519fe4247dba8eb6b18ad93999f76b7f657039acabd47 \\\n    --hash=sha256:0c9136e14ed34a9e343a31c533d78a9813a69a3148332bce5e9821cb2f996e66 \\\n    --hash=sha256:110f8b71aacb688ec69062bb7f6938a0f8acb01b7c1c4beb453c65b6d234584d \\\n    --hash=sha256:112b06a867b235ef466ed3508ddf0238050df9c727cafb5301ac385b899189a1 \\\n    --hash=sha256:17f9ade344e7d9b464a084d69bcf18fc691cb1db67c62ed80820bf4926d78f0e \\\n    --hash=sha256:1e254a00cdf42b1e4d5b3d68d33af63268d41340d8885df2ab6470f2e1500147 \\\n    --hash=sha256:1e978ec1e8bd0e0e4de6bb75de9d30cbb74db6b6a2bb727618613703ca0167dd \\\n    --hash=sha256:25c692919ac5a01f170a3bfcd62d745b24fd095c353d50812637d6fcab442e75 \\\n    --hash=sha256:260a5d70215b61ab4fadf5c7baacd64821842975eea312125ed3c39a6391b063 \\\n    --hash=sha256:2803abfebfc990042cd494d8ce2d5f82e9d847af6d35ec486923aa19dbad5e73 \\\n    --hash=sha256:29a287e0cf63ff528da061de6b9f64a4618da591ca1046aafc54062e40ca7eab \\\n    --hash=sha256:29cb7f67d10b479ff07c17d33e39f78c07f71c40ef30d63c153d340e96cd3fb4 \\\n    --hash=sha256:3213d622a0283a39a93d188f3cf72b26862df52fbb4ca3697f51705016523d41 \\\n    --hash=sha256:33111801a01c12a8a1e3721f0a9232f8cfc8ae2c6b7098167e6f623c6073f402 \\\n    --hash=sha256:357cc07a6d7b0b182ff02249616a03742827ebb1277546b5c7cd7f7620a45698 \\\n    --hash=sha256:38efbc8de75c7a0fc1ac190162d892787f3f47b57cc291231aafee36b80982b7 \\\n    --hash=sha256:4081eb135ac24158bd51cdfbef16f1c64df7063b1143f24731387137c092bec8 \\\n    --hash=sha256:40fdc1ae7125e518ea98e53e69a4ebc27e1fd50510c47b7ea130cf21e5e1d42b \\\n    --hash=sha256:4cfe66903cc32a9921a6733d96b19bb6abf310397581bbad89c228f5abaf0ee8 \\\n    --hash=sha256:511dbaf848decaaaf4b4ca48032619fb3138710c4bf7da7617765edad1ef96b0 \\\n    --hash=sha256:55cced7c52e981362f708ad635198e97a752dfba412cc03c23bbf3bd8d5cd662 \\\n    --hash=sha256:56b39e5e0622a09a25bf5baf62f4bcf0cb8a41ae6e2819cf49bbc5a74c083f91 \\\n    --hash=sha256:5dbbdb29840ca3d91ee0fece42fc29278886d908280bfec0a5846c6f901a3eb0 \\\n    --hash=sha256:5f9fb9157b4ce2971008323afe46053787b526ef624fea915b261468a8421a0f \\\n    --hash=sha256:6180d8b35af935aed8ece3a85e0a43f87393ae0ac87c8d2c8bd2c993f7270ef3 \\\n    --hash=sha256:68a5124b13fa6cc2086764a20005d30bc0548146f7f5322f02fce212ca14317f \\\n    --hash=sha256:68bb27509ac1b9a3443094260f6326150663b06abe40b73a2f81160623da5b67 \\\n    --hash=sha256:6f41ae150c4e32db4f3310cdaf64b1593a03dbabe29eec77fc9b50fe64061df6 \\\n    --hash=sha256:7265a2f3d436e54ef9f2b52b5c937e6be778781bd97a590319d7348f1c1ca997 \\\n    --hash=sha256:72fbe16c6fac95aedf5937fa873445cec2110be35d8a4e9433d7501fd98dae6b \\\n    --hash=sha256:7d92c3819208a60205a12a245c91ad70cb0a85336659b19b834205573ac8456e \\\n    --hash=sha256:8155154c7c691289fe18f510b5d4657c68c67989f293f0535a91360392ff6538 \\\n    --hash=sha256:81a1cca95ed5bb92aa8b10dd2cdc9a0d3853a50fad926c28b5d7e8ea54389627 \\\n    --hash=sha256:89cd468399cfd2504718f0ba50e410dca55a170b61a02ad92bb18c8a65186e93 \\\n    --hash=sha256:8ad03c0965fb3c692200e74d458ca28c1dbb4ce96f9a479a8aa041ad5fabca02 \\\n    --hash=sha256:90f9849678c75fe7afa2d348ac842c168b0a4d3d61919687216dfc547976d853 \\\n    --hash=sha256:948424b06129ce883307e8cff868c31396d8dc7630a59c61d70d98dbe70f222c \\\n    --hash=sha256:9cd5ffd25db4e7ba6a375693b3fc0fc1791ec636c17db3720da19bde7180ec43 \\\n    --hash=sha256:a0df0043bdb289bde1f62da130d20df23d58b45429f752bc7a8fc5325a225ecd \\\n    --hash=sha256:a2c306dea656c12c68f51f4cea133cbe78ca7435eb28c735eac1d3ebe73be6e8 \\\n    --hash=sha256:a7830bab239b79cda9c08c2da014761cafb48da6150e1da17ac06283f43b6089 \\\n    --hash=sha256:a7c711e21628b52034bb5ab8d1bce291f752fcc5e92accc615778acee1ff4778 \\\n    --hash=sha256:aaf159caa35993cb1f56fb9b8e4610d35758e7ca005412eb1daa856a78c9c4b1 \\\n    --hash=sha256:ae506e6902902557576a26ff33eda8695e7ecb3cb36c3b573a0765dee114ebdb \\\n    --hash=sha256:b507f5c4c1d508876d1819b6bf9a49d365b96320b5d4993426b33a23ca4b8261 \\\n    --hash=sha256:bf162abab1c1a736333192707cef898e735a5ca00f38f27eeedf44b39d9e85eb \\\n    --hash=sha256:c1a2af6c6ef86344a6b0db6b97834208bf598db514f2b155042439b62605601a \\\n    --hash=sha256:c2d37ab77531417474168eb79d6d80b14f821a966818505d03013d0833edb7a8 \\\n    --hash=sha256:c4fc99836233ea196540b17ab0983aff60ed07941751930f5f4d05bc3b3b7359 \\\n    --hash=sha256:d581b735e177fdcdce6fed8e7e8880a3fb6ee4e3653a3ac6af01c6f4c03effc5 \\\n    --hash=sha256:d6da64deb6b8ed903e7560180a92f2d804ee1ba5eeb849ac2748b8c1aba1f6d7 \\\n    --hash=sha256:d8e8286dd7cea7895157318d1b91cdacac64c479f3cbc8dce548331728484751 \\\n    --hash=sha256:ddea102b48f9e339f3948bf22040944184627a30fdf7f858667673b9c5f033c8 \\\n    --hash=sha256:dfa20cc6ca228e6b155b11da03825975ce66aea520985dbbddf0f2a5a495c605 \\\n    --hash=sha256:e3e5193ef5a3dc73bceee50f7fdc2c90dbb76c42df8d8fae3d1067a583df579e \\\n    --hash=sha256:e3eeb0aabd6bd5ce64faae67e9935203a6991b4bc2a485a767fbafb2c5125f45 \\\n    --hash=sha256:e5805d5a22fd19c8ccff10a9561f9df94436b0545619ea579db2d3c35294bce2 \\\n    --hash=sha256:e85b752a1e912b70eaad4fafbd4d1238007ab221de2009b9a2f5ae7461239895 \\\n    --hash=sha256:eaf7fa2de5c0be8ae6ff8e9bea2ccd725e980541244521d8d4b5f3354a27babe \\\n    --hash=sha256:ebfb099f8dcf083deef3ac1ca4c1503f387cf76296fcb3816b66f5ecb5f54fdb \\\n    --hash=sha256:ece3d2cfe132e7d51f44a832b303895e6f2d499c5e74dfbdb06ee246147a304a \\\n    --hash=sha256:ed9749eef4cbd126da3dc1d6bcb3a57f5eb7ac6a6484146bdbf743f552dfc577 \\\n    --hash=sha256:ede83e07a75dd06bc501566c1eca2afc0d61677c1472ac9ad93fdee6e638a48d \\\n    --hash=sha256:ef4aea96ce4d3b074422cb4f2f64e216bf9e213004bb58ecfdf50ea02ea8eb9a \\\n    --hash=sha256:f3a3570c4a2a16746ac2c31a7c7c7b0c186b95ce902e33db6f28094ed7387dda \\\n    --hash=sha256:f407cb6b8e9d6d8c626bc73c945db1706035af8fd632295547bf1c9e46d092d6 \\\n    --hash=sha256:f74a575920ab21fe304421a3fc28793d82e299cae9eccb37084e9fc7f3617c20\nopenpyxl==3.1.5 \\\n    --hash=sha256:5282c12b107bffeef825f4617dc029afaf41d0ea60823bbb665ef3079dc79de2 \\\n    --hash=sha256:cf0e3cf56142039133628b5acffe8ef0c12bc902d2aadd3e0fe5878dc08d1050\npackaging==26.3 \\\n    --hash=sha256:94edc256424af38762eb31306eed28beb9f0efc50a8837492c9d6fd6004aed79 \\\n    --hash=sha256:d7193f7c8e4e93f444fde0262bf90af30e16fa0ad0ad44cb553c87339b23cd1c\npandas==2.3.3 \\\n    --hash=sha256:0242fe9a49aa8b4d78a4fa03acb397a58833ef6199e9aa40a95f027bb3a1b6e7 \\\n    --hash=sha256:1611aedd912e1ff81ff41c745822980c49ce4a7907537be8692c8dbc31924593 \\\n    --hash=sha256:1b07204a219b3b7350abaae088f451860223a52cfb8a6c53358e7948735158e5 \\\n    --hash=sha256:1d37b5848ba49824e5c30bedb9c830ab9b7751fd049bc7914533e01c65f79791 \\\n    --hash=sha256:23ebd657a4d38268c7dfbdf089fbc31ea709d82e4923c5ffd4fbd5747133ce73 \\\n    --hash=sha256:2462b1a365b6109d275250baaae7b760fd25c726aaca0054649286bcfbb3e8ec \\\n    --hash=sha256:28083c648d9a99a5dd035ec125d42439c6c1c525098c58af0fc38dd1a7a1b3d4 \\\n    --hash=sha256:2e3ebdb170b5ef78f19bfb71b0dc5dc58775032361fa188e814959b74d726dd5 \\\n    --hash=sha256:318d77e0e42a628c04dc56bcef4b40de67918f7041c2b061af1da41dcff670ac \\\n    --hash=sha256:371a4ab48e950033bcf52b6527eccb564f52dc826c02afd9a1bc0ab731bba084 \\\n    --hash=sha256:376c6446ae31770764215a6c937f72d917f214b43560603cd60da6408f183b6c \\\n    --hash=sha256:3869faf4bd07b3b66a9f462417d0ca3a9df29a9f6abd5d0d0dbab15dac7abe87 \\\n    --hash=sha256:3fd2f887589c7aa868e02632612ba39acb0b8948faf5cc58f0850e165bd46f35 \\\n    --hash=sha256:4793891684806ae50d1288c9bae9330293ab4e083ccd1c5e383c34549c6e4250 \\\n    --hash=sha256:4e0a175408804d566144e170d0476b15d78458795bb18f1304fb94160cabf40c \\\n    --hash=sha256:503cf027cf9940d2ceaa1a93cfb5f8c8c7e6e90720a2850378f0b3f3b1e06826 \\\n    --hash=sha256:5554c929ccc317d41a5e3d1234f3be588248e61f08a74dd17c9eabb535777dc9 \\\n    --hash=sha256:56851a737e3470de7fa88e6131f41281ed440d29a9268dcbf0002da5ac366713 \\\n    --hash=sha256:5caf26f64126b6c7aec964f74266f435afef1c1b13da3b0636c7518a1fa3e2b1 \\\n    --hash=sha256:602b8615ebcc4a0c1751e71840428ddebeb142ec02c786e8ad6b1ce3c8dec523 \\\n    --hash=sha256:6253c72c6a1d990a410bc7de641d34053364ef8bcd3126f7e7450125887dffe3 \\\n    --hash=sha256:6435cb949cb34ec11cc9860246ccb2fdc9ecd742c12d3304989017d53f039a78 \\\n    --hash=sha256:6d21f6d74eb1725c2efaa71a2bfc661a0689579b58e9c0ca58a739ff0b002b53 \\\n    --hash=sha256:6d2cefc361461662ac48810cb14365a365ce864afe85ef1f447ff5a1e99ea81c \\\n    --hash=sha256:74ecdf1d301e812db96a465a525952f4dde225fdb6d8e5a521d47e1f42041e21 \\\n    --hash=sha256:75ea25f9529fdec2d2e93a42c523962261e567d250b0013b16210e1d40d7c2e5 \\\n    --hash=sha256:854d00d556406bffe66a4c0802f334c9ad5a96b4f1f868adf036a21b11ef13ff \\\n    --hash=sha256:8fe25fc7b623b0ef6b5009149627e34d2a4657e880948ec3c840e9402e5c1b45 \\\n    --hash=sha256:900f47d8f20860de523a1ac881c4c36d65efcb2eb850e6948140fa781736e110 \\\n    --hash=sha256:93c2d9ab0fc11822b5eece72ec9587e172f63cff87c00b062f6e37448ced4493 \\\n    --hash=sha256:a16dcec078a01eeef8ee61bf64074b4e524a2a3f4b3be9326420cabe59c4778b \\\n    --hash=sha256:a21d830e78df0a515db2b3d2f5570610f5e6bd2e27749770e8bb7b524b89b450 \\\n    --hash=sha256:a45c765238e2ed7d7c608fc5bc4a6f88b642f2f01e70c0c23d2224dd21829d86 \\\n    --hash=sha256:a637c5cdfa04b6d6e2ecedcb81fc52ffb0fd78ce2ebccc9ea964df9f658de8c8 \\\n    --hash=sha256:a68e15f780eddf2b07d242e17a04aa187a7ee12b40b930bfdd78070556550e98 \\\n    --hash=sha256:b3d11d2fda7eb164ef27ffc14b4fcab16a80e1ce67e9f57e19ec0afaf715ba89 \\\n    --hash=sha256:b468d3dad6ff947df92dcb32ede5b7bd41a9b3cceef0a30ed925f6d01fb8fa66 \\\n    --hash=sha256:b98560e98cb334799c0b07ca7967ac361a47326e9b4e5a7dfb5ab2b1c9d35a1b \\\n    --hash=sha256:bdcd9d1167f4885211e401b3036c0c8d9e274eee67ea8d0758a256d60704cfe8 \\\n    --hash=sha256:bf1f8a81d04ca90e32a0aceb819d34dbd378a98bf923b6398b9a3ec0bf44de29 \\\n    --hash=sha256:c46467899aaa4da076d5abc11084634e2d197e9460643dd455ac3db5856b24d6 \\\n    --hash=sha256:c4fc4c21971a1a9f4bdb4c73978c7f7256caa3e62b323f70d6cb80db583350bc \\\n    --hash=sha256:c503ba5216814e295f40711470446bc3fd00f0faea8a086cbc688808e26f92a2 \\\n    --hash=sha256:d051c0e065b94b7a3cea50eb1ec32e912cd96dba41647eb24104b6c6c14c5788 \\\n    --hash=sha256:d3e28b3e83862ccf4d85ff19cf8c20b2ae7e503881711ff2d534dc8f761131aa \\\n    --hash=sha256:db4301b2d1f926ae677a751eb2bd0e8c5f5319c9cb3f88b0becbbb0b07b34151 \\\n    --hash=sha256:dd7478f1463441ae4ca7308a70e90b33470fa593429f9d4c578dd00d1fa78838 \\\n    --hash=sha256:e05e1af93b977f7eafa636d043f9f94c7ee3ac81af99c13508215942e64c993b \\\n    --hash=sha256:e19d192383eab2f4ceb30b412b22ea30690c9e618f78870357ae1d682912015a \\\n    --hash=sha256:e32e7cc9af0f1cc15548288a51a3b681cc2a219faa838e995f7dc53dbab1062d \\\n    --hash=sha256:ecaf1e12bdc03c86ad4a7ea848d66c685cb6851d807a26aa245ca3d2017a1908 \\\n    --hash=sha256:ee15f284898e7b246df8087fc82b87b01686f98ee67d85a17b7ab44143a3a9a0 \\\n    --hash=sha256:ee67acbbf05014ea6c763beb097e03cd629961c8a632075eeb34247120abcb4b \\\n    --hash=sha256:f086f6fe114e19d92014a1966f43a3e62285109afe874f067f5abbdcbb10e59c \\\n    --hash=sha256:f8bfc0e12dc78f777f323f55c58649591b2cd0c43534e8355c51d3fede5f4dee\npeft==0.20.0 \\\n    --hash=sha256:0fbba16ffebfad3de96e06f2da6860fd860292324b85b6141909fa1e26ea9233 \\\n    --hash=sha256:4769c8093a4ca145fd6fb3fd4dd50449675f5fe46434ad1e98b285a132d4b1d0\npillow==12.3.0 \\\n    --hash=sha256:00808c5e14ef63ac5161091d242999076604ff74b883423a11e5d7bbb38bf756 \\\n    --hash=sha256:04f01d28a6aaff387bf842a13be313df23ba0597a44f1a976c9feb3c6ff4711a \\\n    --hash=sha256:06ff022112bc9cbf83b60f8e028d94ad87b60621706487e65f673de61610ab59 \\\n    --hash=sha256:0740a512dc522224c77d9aa5a8d70d8b7d73fb91f2c21125d8d025d3b8990e45 \\\n    --hash=sha256:0847a763afefb695bc912d7c131e7e0632d4edc1d8698f58ddabec8e46b8b6d3 \\\n    --hash=sha256:0dd2064cbc55aaec028ef5fbb60fa47bb6c3e7918e07ff17935284b227a9d2df \\\n    --hash=sha256:0feb2e9d6ad6c9e3c06effe9d00f3f1e618a6643273576b016f591e9315a7139 \\\n    --hash=sha256:10e41f0fbf1eec8cfd234b8fe17a4caac7c9d0db4c204d3c173a8f9f6ef3232b \\\n    --hash=sha256:1182d52bc2d5e5d7d0949503aa7e36d12f42205dc287e4883f407b1988820d39 \\\n    --hash=sha256:164b31cd1a0490ab6efae01aa5df49da7061be0af1b30e035b6e9a1bfe34ee6e \\\n    --hash=sha256:1657923d2d45afb66526e5b933e5b3052e6bdea196c90d3abb2424e18c77dae8 \\\n    --hash=sha256:186941b6aef820ad110fb01fb06eb925374dc3a21b17e37ec9a53b250c6fe2d1 \\\n    --hash=sha256:1cca606cd25738df4ed873d5ad46bbdb3d83b5cbca291f6b4ff13a4df6b0bbe8 \\\n    --hash=sha256:21900ce7ba264168cd50defae43cd75d25c833ad4ad6e73ffc5596d12e25ac89 \\\n    --hash=sha256:236ff70b9312fb68943c703aa842ca6a758abfa45ac187a5e7c1452e96ef72b5 \\\n    --hash=sha256:23aceaa007d6172b02c277f0cd359c79492bbb14f7072b4ede9fbcaf20648130 \\\n    --hash=sha256:23d27a3e0307ec2244cc51e7287b919aa68d097504ebe19df4e76a98a3eea5bd \\\n    --hash=sha256:24870b09b224f7ae3c39ed07d10e819d06f8720bc551847b1d623832b5b0e28d \\\n    --hash=sha256:251bf95b67017e27b13d82f5b326234ca62d70f9cf4c2b9032de2358a3b12c7b \\\n    --hash=sha256:25b9b82bb22e6e2b3cd07b39c68b7b862001226cb3dff7130d1cb914121b39ed \\\n    --hash=sha256:28ce87c5ab450a9dd970b52e5aca5fe63ed432d18a2eaddd1979a00a1ba24ace \\\n    --hash=sha256:300557495eb45ebb8aec96c2da9c4be642fbf7cd937278b4013ba894ea8eb0eb \\\n    --hash=sha256:30f2aa603c41533cc25c05acd0da21636e84a315768feb631c937177db558931 \\\n    --hash=sha256:331b624368d4f1d069149002f25f44bc61c8919ce8ddb3c45bdad8f6e2d89510 \\\n    --hash=sha256:37d6d0a00072fd2948eb22bce7e1475f34569d90c87c59f7a2ec59541b77f7a6 \\\n    --hash=sha256:37dc8f7bbb66efe481bb60defacef820c950c24713fb44962ed6aa2a50966de1 \\\n    --hash=sha256:3b8182a766685eaa002637e28b4ec8d6b18819a0c71f579bf0dbaa5830297cce \\\n    --hash=sha256:3edce1d53195db527e0191f84b71d02022de0540bf43a16ed734ed7537b07385 \\\n    --hash=sha256:446c34dcc4324b084a53b705127dc15717b22c5e140ae0a3c38349d4efec071e \\\n    --hash=sha256:4998562bf62a445225f22e07c896bb04b35b1b1f2eb6d760584c9c51d7a5f78c \\\n    --hash=sha256:4b0a7fe987b14c31ebda6083f74f22b561fd3739bc0ac51e019622e3d72668c7 \\\n    --hash=sha256:4e8c2a84d977f50b9daed6eeaf3baef67d00d5d74d932288f02cb94518ee3ace \\\n    --hash=sha256:4f883547d4b7f0495ebe7056b0cc2aea76094e7a4abc8e933540f3271df27d9c \\\n    --hash=sha256:514435a37670e3e5e08f3945b68718b6ed329bb84367777e16f9f4dfe1e61a0f \\\n    --hash=sha256:53aa02d20d10c3d814d536aa4e5ac9b84ca0ff5a88377963b085ad6822f93e64 \\\n    --hash=sha256:5594fc43d548a7ed94949d139aa1341b270f1863f11cfd37f5a6c8b778a6b67f \\\n    --hash=sha256:571b9fcb07b97ef3a492028fb3d2dc0993ca23a06138b0315286566d29ef718a \\\n    --hash=sha256:57b3d78c95ba9059768b10e28b813002261d3f3dfc55cc48b0c988f625175827 \\\n    --hash=sha256:5afb51d599ea772b8365ae807ae557f18bccfe46ab261fd1c2a9ed700fc6eb17 \\\n    --hash=sha256:6b02afb9b97f65fbca5f31db6a2a3ba21aa93030225f150fa3f249717e938fb4 \\\n    --hash=sha256:6c0016e7b354317c4e9e525b937ac8596c38d2d232b419529b9cd7a1cd46e39a \\\n    --hash=sha256:71d6097b330eea8fd15097780c8e89cb1a8ce7838669f48c5bacd6f663dd4701 \\\n    --hash=sha256:756c768d0c9c2955feb7a56c37ea24aea2e369f8d36a88da270b6a9f19e62b5e \\\n    --hash=sha256:78cb2c6865a35ab8ff8b75fd122f6033b92a62c82801110e48ddd6c936a45d91 \\\n    --hash=sha256:7a743ff716f746fc19a9557f60dab1600d4613255f8a7aeb3cdde4db7eb15a66 \\\n    --hash=sha256:85f998ea1848bc6757289e739cfbdda3a04adfd58b02fc018ce54d754a5ce468 \\\n    --hash=sha256:8728f216dcdb6e6d555cf971cb34076139ad74b31fc2c14da4fafc741c5f6217 \\\n    --hash=sha256:877c3f311ff35410f690861c4409e7ccbf0cd2f878e50628a28e5a0bb689e658 \\\n    --hash=sha256:8cd2f7bdda092d99c9fc2fb7391354f306d01443d22785d0cbfafa2e2c8bb418 \\\n    --hash=sha256:8e95e1385e4998ae9694eeaa4730ba5457ff61185b3a55e2e7bea0880aef452a \\\n    --hash=sha256:962864dc93511324d51ddbb5b9f8731bf71675b93ca612a07441896f4688fb8c \\\n    --hash=sha256:9cf95fe4d0f84c82d282745d9bb08ad9f926efa00be4697e767b814ce40d4330 \\\n    --hash=sha256:9e881fca225083806662a5c43d627d215f258ff43c890f831966c7d7ba9c7402 \\\n    --hash=sha256:a2b55dd6b2a4c4b7d87ffa56bdb33fdc5fdb9a462173861a7bc097f17d91cb09 \\\n    --hash=sha256:a45650e8ce7fafffd731db8550230db6b0d306d181a90b67d3e6bca2f1990930 \\\n    --hash=sha256:a876864214e136f0eb367788dbd7df045f4806801518e2cfe9e13229cfe06d8f \\\n    --hash=sha256:ae26d61dfa7a47befdc7572b521024e8745f3d809bd95ca9505a7bba9ef849ec \\\n    --hash=sha256:af8d94b0db561cf68b88a267c5c44b49e134f525d0dc2cb7ed413a66bc23559a \\\n    --hash=sha256:b343699e8308bdc51978310e1c959c584e7869cc8c40780058c87da7781a1e94 \\\n    --hash=sha256:b3c777e849237620b022f7f297dd67705f9f5cf1685f09f02e46f93e92725468 \\\n    --hash=sha256:b629de27fda84b42cde7edef0d85f13b958b47f6e9bbcbba9b673c562a89bd8b \\\n    --hash=sha256:ba09209fbe443b4acccebe845d8a138b89a8f4fbaeedd44953490b5315d5e965 \\\n    --hash=sha256:ba54cfebe86920a559a7c4d6b9050791c20513650a1952ebe3368c7dc70306f8 \\\n    --hash=sha256:bcb46e2f9feff8d06323983bd83ed00c201fdcab3d74973e7072a889b3979fcd \\\n    --hash=sha256:bcc33feacfaefce60c12fd500a277533bdc02b10a19f7f6d348763d8140bbba7 \\\n    --hash=sha256:bf16ba1b4d0b6b7c8e534936632270cf70eb00dbe09005bc345b2677b726855c \\\n    --hash=sha256:cf1845d02ad822a369a49f2bb9345b1614744267682e7a03527dc3bf6eea1777 \\\n    --hash=sha256:d69141514cc30b774ceea5e3ed3a6635c8d8a96edf664689b890f4089111fb35 \\\n    --hash=sha256:d9c7f76c0673154f044e9d78c8655fb4213f6ca31a836df48b40fe5d187717b9 \\\n    --hash=sha256:dbce0b29841537a2fa4a214c2bbf14de3587c9680caa9b4e217568472490b28f \\\n    --hash=sha256:dc624f6bc473dacdf7ef7eb8678d0d08edf15cd94fad6ae5c7d6cc67a4e4902f \\\n    --hash=sha256:e158cb00350dc278f3b91551101aa7d12415a66ebf2c91d8d5ac14e56ddd3ad0 \\\n    --hash=sha256:e491916b378fba47242221bb9ead245211b70d504f495d105d17b14a24b4907c \\\n    --hash=sha256:e795b7eb908249c4e43c7c99fac7c2c75dab0c43566e37db472a355f63693d71 \\\n    --hash=sha256:e7e480451b9fa137494bccd3a7d69adbe8ac65a87d97be61e11f1b1050a5bac3 \\\n    --hash=sha256:e91206ee562682b51b98ef4b26a6ef48fd84e15fd4c4bc5ec768eb641d206838 \\\n    --hash=sha256:e9871b1ffbfa9656b60aeee92ed5136a5742696006fa322b29ea3d8da0ecc9cf \\\n    --hash=sha256:e9aeb04d6aef139de265b29683e119b638208f88cf73cdd1658aa07221165321 \\\n    --hash=sha256:ebaea975e03d3141d9d3a507df75c9b3ec90fa9d2ffd07567b3a978d9d790b26 \\\n    --hash=sha256:f0606c8bf2cdefea14a43530f7657cbbb7ecf1c4222512492ef4a4434a9501ec \\\n    --hash=sha256:f13c32a3abd6079a66d9526e18dad9b6d280384d49d7c54040cd57b6424041d9 \\\n    --hash=sha256:f7401aebd7f581d7f83a439d87d474999317ee099218e5ad25d125290990ba65 \\\n    --hash=sha256:fa4ecea169a355be7a3ade2c783e2ed12f0e40d2c5621cda8b3297faf7fbb9f5 \\\n    --hash=sha256:fbd139c8447d25dd750ab79ee274cc5e1fe80fc56340ab10b18a195e1b6eca3e \\\n    --hash=sha256:fdafc9cce40277e0f7a0feabce0ee50dd2fa1800f3b38015e51296b5e814048d \\\n    --hash=sha256:fe3cca2e4e8a592be0f269a1ca4835c25199d9f3ce815c8491048f785b0a0198 \\\n    --hash=sha256:ffd0c5368496f41b0944be820fcb7a838aa6e623d250b01acf2643939c3f99d7\nportalocker==3.2.0 \\\n    --hash=sha256:1f3002956a54a8c3730586c5c77bf18fae4149e07eaf1c29fc3faf4d5a3f89ac \\\n    --hash=sha256:3cdc5f565312224bc570c49337bd21428bba0ef363bbcf58b9ef4a9f11779968\npropcache==0.5.2 \\\n    --hash=sha256:01c4fc7480cd0598bb4b57022df55b9ca296da7fc5a8760bd8451a7e63a7d427 \\\n    --hash=sha256:04dc2390d9edbbaef7461f33322555976ffddf0b650a038649d026358714e6c5 \\\n    --hash=sha256:06187263ddad280d05b4d8a8b3bb7d164cbebd469236544a42e6d9b28ac6a4fa \\\n    --hash=sha256:0958834041a0166d343b8d2cedcd8bcbaeb4fdbe0cf08320c5379f143c3be6e7 \\\n    --hash=sha256:099aaf4b4d1a02265b92a977edf00b5c4f63b3b17ac6de39b0d637c9cac0188a \\\n    --hash=sha256:0d2c9bf8528f135dbb805ce027567e09164f7efa51a2be07458a2c0420f292d0 \\\n    --hash=sha256:0fd59b5af35f74da48d905dcbad55449ba13be91823cb05a9bd590bbf5b61660 \\\n    --hash=sha256:10734b5484ea113152ee25a91dccedf81631791805d2c9ccb054958e51842c94 \\\n    --hash=sha256:13fef48778b5a2a756523fdb781326b028ca75e32858b04f2cdd19f394564917 \\\n    --hash=sha256:178b4a2cdaac1818e2bf1c5a99b94383fa73ea5382e032a48dec07dc5668dc42 \\\n    --hash=sha256:196913dea116aeb5a2ba95af4ddcb7ea85559ae07d8eee8751688310d09168c3 \\\n    --hash=sha256:1b31822f4474c4036bae62de9402710051d431a606d6a0f907fec79935a071aa \\\n    --hash=sha256:1ca071adabaab6e9219924bbe00af821f1ee7de113a9eca1cdc292de3d120f4d \\\n    --hash=sha256:1d1ad32d9d4355e2be65574fd0bfd3677e7066b009cd5b9b2dee8aa6a6393b33 \\\n    --hash=sha256:1dbcf7675229b35d31abb6547d8ebc8c27a830ac3f9a794edff6254873ec7c0a \\\n    --hash=sha256:2293949b855ce597f2826452d17c2d545fb5622379c4ea6fdf525e9b8e8a2511 \\\n    --hash=sha256:26a4dca084132874e639895c3135dfad5eb20bae209f62d1aeb31b03e601c3c0 \\\n    --hash=sha256:2800a4a8ead6b28cccd1ec54b59346f0def7922ee1c7598e8499c733cfbb7c84 \\\n    --hash=sha256:29cbaac5ea0212663e6845e04b5e188d5a6ae6dd919810ac835bf1d3b42c3f4c \\\n    --hash=sha256:29f9309a2e42b0d273be006fdb4be2d6c39a47f6f57d8fb1cf9f81481df81b66 \\\n    --hash=sha256:2d7aa89ebca5acc98cba9d1472d976e394782f587bad6661003602a619fd1821 \\\n    --hash=sha256:2f22cbbac9e26a8e864c0985ff1268d5d939d53d9d9411a9824279097e03a2cb \\\n    --hash=sha256:2f8ea531c794b9d6274acd4e8d2c2ebcac590a4361d27482edd3010b79f1325e \\\n    --hash=sha256:3115559b8effafd63b142ea5ed53d63a16ea6469cbc63dce4ee194b42db5d853 \\\n    --hash=sha256:32775082acd2d807ee3db715c7770d38767b817870acfa08c29e057f3c4d5b56 \\\n    --hash=sha256:3430bb2bfe1331885c427745a751e774ee679fd4344f80b97bf879815fe8fa55 \\\n    --hash=sha256:3b199b9b2b3d6a7edf3183ba8a9a137a22b97f7df525feb5ae1eccf026d2a9c6 \\\n    --hash=sha256:40314bca9ac559716fe374094fc81c11dcc34b64fd6c585360f5775690505704 \\\n    --hash=sha256:44e488ef40dbb452700b2b1f8188934121f6648f52c295055662d2191959ff82 \\\n    --hash=sha256:452b5065457eb9991ec5eb38ff41d6cd4c991c9ac7c531c4d5849ae473a9a13f \\\n    --hash=sha256:45f11346f884bc47444f6e6647131055844134c3175b629f84952e2b5cd62b64 \\\n    --hash=sha256:46088abff4cba581dea21ae0467a480526cb25aa5f3c269e909f800328bc3999 \\\n    --hash=sha256:4621064bbf28fa77ff64dd5d94367c04684c67d3a5bf1dff25f0cd0d98a38f3b \\\n    --hash=sha256:4bc8ff1feffc6a61c7002ffe84634c41b822e104990ae009f44a0834430070bb \\\n    --hash=sha256:4db0ba63d693afd40d249bd93f842b5f144f8fcbb83de05660373bcf30517b1d \\\n    --hash=sha256:51f96d685ab16e88cab128cd37a52c5da540809c8b879fa047731bfcb4ad35a4 \\\n    --hash=sha256:54adaa85a22078d1e306304a40984dc5be99d599bf3dc0a24dc98f7daeab89ab \\\n    --hash=sha256:552ffadf6ad409844bc5919c42a0a83d88314cedddaea0e41e80a8b8fffe881f \\\n    --hash=sha256:5538d2c13d93e4698af7e092b57bc7298fd35d1d58e656ae18f23ee0d0378e03 \\\n    --hash=sha256:5570dbcc97571c15f68068e529c92715a12f8d54030e272d264b377e22bd17a5 \\\n    --hash=sha256:5671d09a36b06d0fd4a3da0fccbcae360e9b1570924171a15e9e0997f0249fba \\\n    --hash=sha256:583c19759d9eec1e5b69e2fbef36a7d9c326041be9746cb822d335c8cedc2979 \\\n    --hash=sha256:5aaa2b923c1944ac8febd6609cb373540a5563e7cbcb0fd770f75dace2eb817b \\\n    --hash=sha256:5dbc581d2814337da56222fab8dc5f161cd798a434e49bac27930aaef798e144 \\\n    --hash=sha256:5fcb98e7598b1ee0addab320d90f65b530297a867dbfe9de52ea838077e16e3d \\\n    --hash=sha256:6041d31504dc1779d700e1edcfb08eea334b357620b06681a4eabb57a74e574e \\\n    --hash=sha256:66ea454f095ddf5b6b14f56c064c0941c4788be11e18d2464cf643bf7203ff67 \\\n    --hash=sha256:68ce1c44c7a813a7f71ea04315a8c7b330b63db99d059a797a4651bb6f69f117 \\\n    --hash=sha256:6a997d0489e9668a384fcfd5061b857aa5361de73191cac204d04b889cfbbafa \\\n    --hash=sha256:6bf3be92233808fcd338eba0fb4d0b59ec5772af4f4ecfcec450d1bfc0f8b5eb \\\n    --hash=sha256:6de8bd93ddde9b992cf2b2e0d796d501a19026b5b9fd87356d7d0779531a8d96 \\\n    --hash=sha256:6e7b8719005dd1175be4ab1cd25e9b98659a5e0347331506ec6760d2773a7fb5 \\\n    --hash=sha256:6f328175a2cde1f0ff2c4ed8ce968b9dcfb55f3a7153f39e2957ed994da13476 \\\n    --hash=sha256:72d61e16dd78228b58c5d47be830ff3da7e5f139abdf0aef9d86cde1c5cf2191 \\\n    --hash=sha256:74b70780220e2dd89175ca24b81b68b67c83db499ae611e7f2313cb329801c78 \\\n    --hash=sha256:79aa3ff0a9b566633b642fa9caf7e21ed1c13d6feca718187873f199e1514078 \\\n    --hash=sha256:7afa37062e6650640e932e4cc9297d81f9f42d9944029cc386b8247dea4da837 \\\n    --hash=sha256:80168e2ebe4d3ec6599d10ad8f520304ae1cad9b6c5a95372aef1b66b7bfb53a \\\n    --hash=sha256:806719138ecd720339a12410fb9614ac9b2b2d3a5fdf8235d56981c36f4039ba \\\n    --hash=sha256:8114f28879e0904748e831c3a7774261bd9e75f49be089f389a76f959dcd13fe \\\n    --hash=sha256:81e3a30b0bb60caa22033dd0f8a3618d1d67356212514f62c57db75cb0ef410c \\\n    --hash=sha256:823581fd5cb08b12a48bfa11fe962a7916766b6170c17b028fbdf762b85eb9bf \\\n    --hash=sha256:85341b12b9d55bad0bded24cac341bb34289469e03a11f3f583ea1cc1db0326c \\\n    --hash=sha256:857187f381f88c8e2fa2fe56ab94879d011b883d5a2ee5a1b60a8cd2a06846d9 \\\n    --hash=sha256:8a90efd5777e996e42d568db9ac740b944d691e565cbfd31b2f7832f9184b2b8 \\\n    --hash=sha256:8b73ab70f1a3351fbc71f663b3e645af6dd0329100c353081cf69c37433fc6fe \\\n    --hash=sha256:8c7972d8f193740d9175f0998ab38717e6cd322d5935c5b0fef8c0d323fd9031 \\\n    --hash=sha256:8e778ebd44ef4f66ed60a0416b06b489687db264a9c0b3620362f26489492913 \\\n    --hash=sha256:9282fb1a3bccd038da9f768b927b24a0c753e466c086b7c4f3c6982851eefb2d \\\n    --hash=sha256:949c91d1a990cf3b2e8188dfcfb25005e0b834a06c63fa4ef9f360878ce21ecf \\\n    --hash=sha256:95f1e3f4760d404b13c9976c0229b2b49a3c8e2c62a9ce92efdd2b11ada75e3f \\\n    --hash=sha256:97797ebb098e670a2f92dd66f32897e30d7615b14e7f59711de23e30a9072539 \\\n    --hash=sha256:a0e399a2eccb91ed18721f86aa85757727400b6865c89e88934781deb9c8498b \\\n    --hash=sha256:a473b3440261e0c60706e732b2ed2f517857344fc21bf48fdfe211e2d98eb285 \\\n    --hash=sha256:a4840ab0ae0216d952f4b53dc6d0b992bfc2bedbfe360bdd9b548bc184c08959 \\\n    --hash=sha256:a592f5f3da71c8691c788c13cb6734b6d17663d2e1cb8caddf0673d01ef8847d \\\n    --hash=sha256:a6ae2198be502c10f09b2516e7b5d019816924bc3183a43ce792a7bd6625e6f4 \\\n    --hash=sha256:a6ddc6ac9e25de626c1f129c1b467d7ecd33ce2237d3fd0c4e429feef0a7ee1f \\\n    --hash=sha256:acd2c8edba48e31e58a363b8cf4e5c7db3b04b3f9e371f601df30d9b0d244836 \\\n    --hash=sha256:b05d643f944a8c3c4bd86d65ffd87bf3264b617f87791940302bc474d2ff5274 \\\n    --hash=sha256:b96db7141a592cbc968daf1feea83a118e6ab378af4abbc72b248c895414c22d \\\n    --hash=sha256:ba338430e87ceb9c8f0cf754de38a9860560261e56c00376debd628698a7364f \\\n    --hash=sha256:ba57fffe4ac99c5d30076161b5866336d97600769bad35cc68f7774b15298a4e \\\n    --hash=sha256:be1ddfcbb376e3de5d2e2db1d58d6d67463e6b4f9f040c000de8e300295465fe \\\n    --hash=sha256:c0cb9ed24c8964e172768d455a38254c2dd8a552905729ce006cad3d3dda59b1 \\\n    --hash=sha256:c60462af8e6dc30c35407c7237ea908d777b22862bbee27bc4699c0d8bcdc45a \\\n    --hash=sha256:c66afea89b1e43725731d2004732a046fe6fe955d51f952c3e95a7314a284a39 \\\n    --hash=sha256:c6844ba6364fb12f403928a82cfd295ab103a2b315c77c747b2dbe4a41894ea7 \\\n    --hash=sha256:c80f4ba3e8f00189165999a742ee526ebeccedf6c3f7beb0c7df821e9772435a \\\n    --hash=sha256:cafca7e56c12bb02ae16d283742bef25a61122e9dab2b5b3f2ccbe589ce32164 \\\n    --hash=sha256:cc1177027eda740fdb152706bd215a3f124e3eea15afc39f2cb9fe351b50619e \\\n    --hash=sha256:cc49723e2f60d6b32a0f0b08a3fd6d13203c07f1cd9566cfce0f12a917c967a2 \\\n    --hash=sha256:cc6fc3cc62e8501d3ed62894425040d2728ecddb1ed072737a5c70bd537aa9f0 \\\n    --hash=sha256:cd416c1de191973c52ff1a12a57446bfc7642797b282d7caf2162d7d1b8aa9a0 \\\n    --hash=sha256:cd645f03898405cabe694fb8bc35241e3a9c332ec85627584fe3de201452b335 \\\n    --hash=sha256:cef6cea3922890dd6c9654971001fa797b526c16ab5e1e46c05fd6f877be7568 \\\n    --hash=sha256:cfa21e036ce1e1db2be04ba3b85d2df1bb1702fa01932d984c5464c665228ff4 \\\n    --hash=sha256:d0326e2e5e1f3163fa306c834e48e8d490e5fae607a097a40c0648109b47ba80 \\\n    --hash=sha256:d310c013aad2c72f1c3f2f8dd3279d460a858c551f97aeb8c63e4693cca7b4d2 \\\n    --hash=sha256:d447bb0b3054be5818458fbb171208b1d9ff11eba14e18ca18b90cbb45767370 \\\n    --hash=sha256:d4dc37dec6c6cdad0b57881a5658fd14fbf53e333b1a86cf86559f190e1d9ec4 \\\n    --hash=sha256:d5a81be28596d6559f6131ef33e10200de6e17643b3c74ce03f9eb103be6ae8b \\\n    --hash=sha256:d9ee8826a7d47863a08ac44e1a5f611a462eefc3a194b492da242128bec75b42 \\\n    --hash=sha256:db2b80ea58eab4f86b2beec3cc8b39e8ff9276ac20e96b7cce43c8ae84cd6b5a \\\n    --hash=sha256:decfca4c79dd53ebab484b00cc4b6717d8c369f86e74aa4ca395a64ac651495e \\\n    --hash=sha256:dfed59d0a5aeb01e242e66ff0300bc4a265a7c05f612d30016f0b60b1017d757 \\\n    --hash=sha256:e00820e192c8dbebcafb383ebbf99030895f09905e7a0eb2e0340a0bcc2bc825 \\\n    --hash=sha256:e4294d04a94dcab1b3bccd8b66d962dcad411a1d19414b2a41d1445f1de32ad0 \\\n    --hash=sha256:e59bc9e66329185b93dab73f210f1a37f81cb40f321501db8017c9aea15dba27 \\\n    --hash=sha256:e5cbfac9f61484f7e9f3597775500cd3ebe8274e9b050c38f9525c77c97520bf \\\n    --hash=sha256:f064f8d2b59177878b7615df1735cd8fe3462ed6be8c7b217d17a276489c2b7f \\\n    --hash=sha256:f156a3529f38063b6dbaf356e15602a7f95f8055b1295a438433a6386f10463d \\\n    --hash=sha256:f19bb891234d72535764d703bfed1153cc34f4214d5bd7150aee1eec9e8f4366 \\\n    --hash=sha256:f7467da8a9822bf1a55336f877340c5bcbd3c482afc43a99771169f74a26dedc \\\n    --hash=sha256:f78abfa8dfc32376fd1aacf597b2f2fbbe0ea751419aee718af5d4f82537ef8c \\\n    --hash=sha256:f7eabc04151c78a9f4d5bbb5f1faf571e4defeb4b585e0fe95b60ff2dbe4d3d7 \\\n    --hash=sha256:f814362777a9f841adddb200ecdf8f5cb1e5a3c4b7a86378edbd6ccb26edd702 \\\n    --hash=sha256:fc299c129490f55f254cd90be0deca4764e36e9a7c08b4aa588479a3bbed3098 \\\n    --hash=sha256:fc76378c62a0f04d0cd82fbb1a2cd2d7e28fcb40d5873f28a6c44e388aaa2751 \\\n    --hash=sha256:fc88b26f08d634f7bc819a7852e5214f5802641ab8d9fd5326892292eee1993e \\\n    --hash=sha256:fe67a3d11cd9b4efabfa45c3d00ffba2b26811442a73a581a94b67c2b5faccf6\nprotobuf==7.36.0 \\\n    --hash=sha256:1781cc1de61249b750848029bca452c0a8b7e990080316b9bbc2518b2117b488 \\\n    --hash=sha256:3297e60abdff301e5f74393d87f6cc59dacab5f024a89548a6e8de1d26576b16 \\\n    --hash=sha256:53374d53fc29a67f7dbbf0ade47d7526a0f0137bf0f9c90e48d8a60790ef748c \\\n    --hash=sha256:70f5ec8eb0da81a44360c0dc0beac99a0d78071d21956a7076bae8bd2051841b \\\n    --hash=sha256:7326fd717bdc419162a735938d89d4032332bcc3408804012b24ff3a37086071 \\\n    --hash=sha256:9103532dffd80c6fab7e50c65a31007680a06eb57537d437bb1b35812c138a37 \\\n    --hash=sha256:bf94a5917c71058262de683669bc0a797a7669d3de71f0b36d058e3194f47b44 \\\n    --hash=sha256:e8e09cb0d794c6687926fa558a8a6e72aa10edb997d5ca61da0765f12a3e00ea\npsutil==7.2.2 \\\n    --hash=sha256:0746f5f8d406af344fd547f1c8daa5f5c33dbc293bb8d6a16d80b4bb88f59372 \\\n    --hash=sha256:076a2d2f923fd4821644f5ba89f059523da90dc9014e85f8e45a5774ca5bc6f9 \\\n    --hash=sha256:11fe5a4f613759764e79c65cf11ebdf26e33d6dd34336f8a337aa2996d71c841 \\\n    --hash=sha256:1a571f2330c966c62aeda00dd24620425d4b0cc86881c89861fbc04549e5dc63 \\\n    --hash=sha256:1a7b04c10f32cc88ab39cbf606e117fd74721c831c98a27dc04578deb0c16979 \\\n    --hash=sha256:1fa4ecf83bcdf6e6c8f4449aff98eefb5d0604bf88cb883d7da3d8d2d909546a \\\n    --hash=sha256:2edccc433cbfa046b980b0df0171cd25bcaeb3a68fe9022db0979e7aa74a826b \\\n    --hash=sha256:7b6d09433a10592ce39b13d7be5a54fbac1d1228ed29abc880fb23df7cb694c9 \\\n    --hash=sha256:8c233660f575a5a89e6d4cb65d9f938126312bca76d8fe087b947b3a1aaac9ee \\\n    --hash=sha256:917e891983ca3c1887b4ef36447b1e0873e70c933afc831c6b6da078ba474312 \\\n    --hash=sha256:ab486563df44c17f5173621c7b198955bd6b613fb87c71c161f827d3fb149a9b \\\n    --hash=sha256:ae0aefdd8796a7737eccea863f80f81e468a1e4cf14d926bd9b6f5f2d5f90ca9 \\\n    --hash=sha256:b0726cecd84f9474419d67252add4ac0cd9811b04d61123054b9fb6f57df6e9e \\\n    --hash=sha256:b58fabe35e80b264a4e3bb23e6b96f9e45a3df7fb7eed419ac0e5947c61e47cc \\\n    --hash=sha256:c7663d4e37f13e884d13994247449e9f8f574bc4655d509c3b95e9ec9e2b9dc1 \\\n    --hash=sha256:e452c464a02e7dc7822a05d25db4cde564444a67e58539a00f929c51eddda0cf \\\n    --hash=sha256:e78c8603dcd9a04c7364f1a3e670cea95d51ee865e4efb3556a3a63adef958ea \\\n    --hash=sha256:eb7e81434c8d223ec4a219b5fc1c47d0417b12be7ea866e24fb5ad6e84b3d988 \\\n    --hash=sha256:ed0cace939114f62738d808fdcecd4c869222507e266e574799e9c0faa17d486 \\\n    --hash=sha256:eed63d3b4d62449571547b60578c5b2c4bcccc5387148db46e0c2313dad0ee00 \\\n    --hash=sha256:fd04ef36b4a6d599bbdb225dd1d3f51e00105f6d48a28f006da7f9822f2606d8\npyarrow==21.0.0 \\\n    --hash=sha256:067c66ca29aaedae08218569a114e413b26e742171f526e828e1064fcdec13f4 \\\n    --hash=sha256:072116f65604b822a7f22945a7a6e581cfa28e3454fdcc6939d4ff6090126623 \\\n    --hash=sha256:0c4e75d13eb76295a49e0ea056eb18dbd87d81450bfeb8afa19a7e5a75ae2ad7 \\\n    --hash=sha256:186aa00bca62139f75b7de8420f745f2af12941595bbbfa7ed3870ff63e25636 \\\n    --hash=sha256:1e005378c4a2c6db3ada3ad4c217b381f6c886f0a80d6a316fe586b90f77efd7 \\\n    --hash=sha256:203003786c9fd253ebcafa44b03c06983c9c8d06c3145e37f1b76a1f317aeae1 \\\n    --hash=sha256:222c39e2c70113543982c6b34f3077962b44fca38c0bd9e68bb6781534425c10 \\\n    --hash=sha256:26bfd95f6bff443ceae63c65dc7e048670b7e98bc892210acba7e4995d3d4b51 \\\n    --hash=sha256:3a302f0e0963db37e0a24a70c56cf91a4faa0bca51c23812279ca2e23481fccd \\\n    --hash=sha256:3a81486adc665c7eb1a2bde0224cfca6ceaba344a82a971ef059678417880eb8 \\\n    --hash=sha256:3b4d97e297741796fead24867a8dabf86c87e4584ccc03167e4a811f50fdf74d \\\n    --hash=sha256:40ebfcb54a4f11bcde86bc586cbd0272bac0d516cfa539c799c2453768477569 \\\n    --hash=sha256:479ee41399fcddc46159a551705b89c05f11e8b8cb8e968f7fec64f62d91985e \\\n    --hash=sha256:5051f2dccf0e283ff56335760cbc8622cf52264d67e359d5569541ac11b6d5bc \\\n    --hash=sha256:555ca6935b2cbca2c0e932bedd853e9bc523098c39636de9ad4693b5b1df86d6 \\\n    --hash=sha256:585e7224f21124dd57836b1530ac8f2df2afc43c861d7bf3d58a4870c42ae36c \\\n    --hash=sha256:58c30a1729f82d201627c173d91bd431db88ea74dcaa3885855bc6203e433b82 \\\n    --hash=sha256:6299449adf89df38537837487a4f8d3bd91ec94354fdd2a7d30bc11c48ef6e79 \\\n    --hash=sha256:65f8e85f79031449ec8706b74504a316805217b35b6099155dd7e227eef0d4b6 \\\n    --hash=sha256:689f448066781856237eca8d1975b98cace19b8dd2ab6145bf49475478bcaa10 \\\n    --hash=sha256:69cbbdf0631396e9925e048cfa5bce4e8c3d3b41562bbd70c685a8eb53a91e61 \\\n    --hash=sha256:731c7022587006b755d0bdb27626a1a3bb004bb56b11fb30d98b6c1b4718579d \\\n    --hash=sha256:7be45519b830f7c24b21d630a31d48bcebfd5d4d7f9d3bdb49da9cdf6d764edb \\\n    --hash=sha256:898afce396b80fdda05e3086b4256f8677c671f7b1d27a6976fa011d3fd0a86e \\\n    --hash=sha256:8d58d8497814274d3d20214fbb24abcad2f7e351474357d552a8d53bce70c70e \\\n    --hash=sha256:9b0b14b49ac10654332a805aedfc0147fb3469cbf8ea951b3d040dab12372594 \\\n    --hash=sha256:9d9f8bcb4c3be7738add259738abdeddc363de1b80e3310e04067aa1ca596634 \\\n    --hash=sha256:a7a102574faa3f421141a64c10216e078df467ab9576684d5cd696952546e2da \\\n    --hash=sha256:a7f6524e3747e35f80744537c78e7302cd41deee8baa668d56d55f77d9c464b3 \\\n    --hash=sha256:b6b27cf01e243871390474a211a7922bfbe3bda21e39bc9160daf0da3fe48876 \\\n    --hash=sha256:b7ae0bbdc8c6674259b25bef5d2a1d6af5d39d7200c819cf99e07f7dfef1c51e \\\n    --hash=sha256:bd04ec08f7f8bd113c55868bd3fc442a9db67c27af098c5f814a3091e71cc61a \\\n    --hash=sha256:c077f48aab61738c237802836fc3844f85409a46015635198761b0d6a688f87b \\\n    --hash=sha256:cdc4c17afda4dab2a9c0b79148a43a7f4e1094916b3e18d8975bfd6d6d52241f \\\n    --hash=sha256:cf56ec8b0a5c8c9d7021d6fd754e688104f9ebebf1bf4449613c9531f5346a18 \\\n    --hash=sha256:d2fe8e7f3ce329a71b7ddd7498b3cfac0eeb200c2789bd840234f0dc271a8efe \\\n    --hash=sha256:dc56bc708f2d8ac71bd1dcb927e458c93cec10b98eb4120206a4091db7b67b99 \\\n    --hash=sha256:e563271e2c5ff4d4a4cbeb2c83d5cf0d4938b891518e676025f7268c6fe5fe26 \\\n    --hash=sha256:e72a8ec6b868e258a2cd2672d91f2860ad532d590ce94cdf7d5e7ec674ccf03d \\\n    --hash=sha256:e99310a4ebd4479bcd1964dff9e14af33746300cb014aa4a3781738ac63baf4a \\\n    --hash=sha256:f522e5709379d72fb3da7785aa489ff0bb87448a9dc5a75f45763a795a089ebd \\\n    --hash=sha256:fc0d2f88b81dcf3ccf9a6ae17f89183762c8a94a5bdcfa09e05cfe413acf0503 \\\n    --hash=sha256:fee33b0ca46f4c85443d6c450357101e47d53e6c3f008d658c27a2d020d44c79\npydantic==2.13.4 \\\n    --hash=sha256:45a282cde31d808236fd7ea9d919b128653c8b38b393d1c4ab335c62924d9aba \\\n    --hash=sha256:c40756b57adaa8b1efeeced5c196f3f3b7c435f90e84ea7f443901bec8099ef6\npydantic-core==2.46.4 \\\n    --hash=sha256:00c603d540afdd6b80eb39f078f33ebd46211f02f33e34a32d9f053bba711de0 \\\n    --hash=sha256:0186750b482eefa11d7f435892b09c5c606193ef3375bcf94aa00ae6bfb66262 \\\n    --hash=sha256:041bde0a48fd37cf71cab1c9d56d3e8625a3793fef1f7dd232b3ff37e978ecda \\\n    --hash=sha256:0c563b08bca408dc7f65f700633d8442fffb2421fc47b8101377e9fd65051ff0 \\\n    --hash=sha256:0cbe8b01f948de4286c74cdd6c667aceb38f5c1e26f0693b3983d9d74887c65e \\\n    --hash=sha256:0ce40cd7b21210e99342afafbd4d0f76d784eb5b1d60f3bdc566be4983c6c73b \\\n    --hash=sha256:0e96592440881c74a213e5ad528e2b24d3d4f940de2766bed9010ab1d9e51594 \\\n    --hash=sha256:10e17cbb10a330363733efc4d7c4d0dd827ac0909b8f6a6542298fed1ea62f29 \\\n    --hash=sha256:133878133d271ade3d41d1bfb2a45ec38dbdbda40bc065921c6b04e4630127e2 \\\n    --hash=sha256:14d4edf427bdcf950a8a02d7cb44a08614388dd6e1bdcbf4f67504fa7887da9c \\\n    --hash=sha256:14f4c5d6db102bd796a627bbb3a17b4cf4574b9ae861d8b7c9a9661c6dd3362d \\\n    --hash=sha256:17299feefe090f2caa5b8e37222bb5f663e4935a8bfa6931d4102e5df1a9f398 \\\n    --hash=sha256:184c081504d17f1c1066e430e117142b2c77d9448a97f7b65c6ac9fd9aee238d \\\n    --hash=sha256:18e5ceec2ab67e6d5f1a9085e5a24c9c4e2ac4545730bfe668680bca05e555f3 \\\n    --hash=sha256:19e51f073cd3df251856a8a4189fbdf1de4012c3ebacfb1884f94f1eb406079f \\\n    --hash=sha256:1a7dd0b3ee80d90150e3495a3a13ac34dbcbfd4f012996a6a1d8900e91b5c0fb \\\n    --hash=sha256:1d8ba486450b14f3b1d63bc521d410ec7565e52f887b9fb671791886436a42f7 \\\n    --hash=sha256:2108ba5c1c1eca18030634489dc544844144ee36357f2f9f780b93e7ddbb44b5 \\\n    --hash=sha256:228ee9bae8bef5b1e97ec58302f80357c37199e0d0a99174e138d28e6957b9d9 \\\n    --hash=sha256:23ace664830ee0bfe014a0c7bc248b1f7f25ed7ad103852c317624a1083af462 \\\n    --hash=sha256:2412e734dcb48da14d4e4006b82b46b74f2518b8a26ee7e58c6844a6cd6d03c4 \\\n    --hash=sha256:29c61fc04a3d840155ff08e475a04809278972fe6aef51e2720554e96367e34b \\\n    --hash=sha256:2f84c03c8607173d16b5a854ec68a2f9079ae03237a54fb506d13af47e1d018d \\\n    --hash=sha256:3009f12e4e90b7f88b4f9adb1b0c4a3d58fe7820f3238c190047209d148026df \\\n    --hash=sha256:3245406455a5d98187ec35530fd772b1d799b26667980872c8d4614991e2c4a2 \\\n    --hash=sha256:3447661d99f75a3683a4cf5c87da72f2161964611864dbbeac7fbb118bb4bfc0 \\\n    --hash=sha256:372429a130e469c9cd698925ce5fc50940b7a1336b0d82038e63d5bbc4edc519 \\\n    --hash=sha256:395aebd9183f9d112f569aeb5b2214d1a10a33bec8456447f7fbdfa51d38d4cd \\\n    --hash=sha256:3a233125ac121aa3ffba9a2b59edfc4a985a76092dc8279586ab4b71390875e7 \\\n    --hash=sha256:3be77f45df024d789a672ae34f8b06fb346c4f9f46ea714956660ea4862e89ac \\\n    --hash=sha256:3bf92c5d0e00fefaab325a4d27828fe6b6e2a21848686b5b60d2d9eeb09d76c6 \\\n    --hash=sha256:3ecbc122d18468d06ca279dc26a8c2e2d5acb10943bb35e36ae92096dc3b5565 \\\n    --hash=sha256:3fb702cd90b0446a3a1c5e470bfa0dd23c0233b676a9099ddcc964fa6ca13898 \\\n    --hash=sha256:428e04521a40150c85216fc8b85e8d39fece235a9cf5e383761238c7fa9b96fb \\\n    --hash=sha256:432c179df7874eeb73307aad2df0755e1ae0efa61ff0ea89b93e194411ae3928 \\\n    --hash=sha256:4a05d69cba51d852c5c3e92758653245a50c0b646ced0cf05bd793ed592839d6 \\\n    --hash=sha256:4c63ebc82684aa89d9a3bcbd13d515b3be44250dc68dd3bd81526c1cb31286c3 \\\n    --hash=sha256:4fc73cb559bdb54b1134a706a2802a4cddd27a0633f5abb7e53056268751ac6a \\\n    --hash=sha256:4fcbe087dbc2068af7eda3aa87634eba216dbda64d1ae73c8684b621d33f6596 \\\n    --hash=sha256:56cb4851bcaf3d117eddcef4fe66afd750a50274b0da8e22be256d10e5611987 \\\n    --hash=sha256:5855698a4856556d86e8e6cd8434bc3ac0314ee8e12089ae0e143f64c6256e4e \\\n    --hash=sha256:5a4330cdbc57162e4b3aa303f588ba752257694c9c9be3e7ebb11b4aca659b5d \\\n    --hash=sha256:5b712b53160b79a5850310b912a5ef8e57e56947c8ad690c227f5c9d7e561712 \\\n    --hash=sha256:5d5902252db0d3cedf8d4a1bc68f70eeb430f7e4c7104c8c476753519b423008 \\\n    --hash=sha256:617d7e2ca7dcb8c5cf6bcb8c59b8832c94b36196bbf1cbd1bfb56ed341905edd \\\n    --hash=sha256:62f875393d7f270851f20523dd2e29f082bcc82292d66db2b64ea71f64b6e1c1 \\\n    --hash=sha256:633147d34cf4550417f12e2b1a0383973bdf5cdfde212cb09e9a581cf10820be \\\n    --hash=sha256:66ce7632c22d837c95301830e111ad0128a32b8207533b60896a96c4915192ea \\\n    --hash=sha256:6b3ace8194b0e5204818c92802dcdca7fc6d88aabbb799d7c795540d9cd6d292 \\\n    --hash=sha256:6f2eeda33a839975441c86a4119e1383c50b47faf0cbb5176985565c6bb02c33 \\\n    --hash=sha256:7027560ee92211647d0d34e3f7cd6f50da56399d26a9c8ad0da286d3869a53f3 \\\n    --hash=sha256:7283d57845ecf5a163403eb0702dfc220cc4fbdd18919cb5ccea4f95ee1cdab4 \\\n    --hash=sha256:7a5f930472650a82629163023e630d160863fce524c616f4e5186e5de9d9a49b \\\n    --hash=sha256:7bfb192b3f4b9e8a89b6277b6ce787564f62cfd272055f6e685726b111dc7826 \\\n    --hash=sha256:811ff8e9c313ab425368bcbb36e5c4ebd7108c2bbf4e4089cfbb0b01eff63fac \\\n    --hash=sha256:8233f2947cf85404441fd7e0085f53b10c93e0ee78611099b5c7237e36aacbf7 \\\n    --hash=sha256:82cf5301172168103724d49a1444d3378cb20cdee30b116a1bd6031236298a5d \\\n    --hash=sha256:8358a950c8909158e3df31538a7e4edc2d7265a7c54b47f0864d9e5bae9dcebf \\\n    --hash=sha256:85bb3611ff1802f3ee7fdd7dbff26b56f343fb432d57a4728fdd49b6ef35e2f4 \\\n    --hash=sha256:86e1a4418c6cd97d60c95c71164158eaf7324fae7b0923264016baa993eba6fc \\\n    --hash=sha256:8b9bab013d1c7a79d3501ff86d0bc9c31bf587db4551677b96bec07df78c6b15 \\\n    --hash=sha256:8c5dac79fa1614d1e06ca695109c6105923bd9c7d1d6c918d4e637b7e6b32fd3 \\\n    --hash=sha256:8d0820e8192167f80d88d64038e609c31452eeca865b4e1d9950a27a4609b00b \\\n    --hash=sha256:8daafc69c93ee8a0204506a3b6b30f586ef54028f52aeeeb5c4cfc5184fd5914 \\\n    --hash=sha256:9037063db01f09b09e237c282b6792bd4da634b5402c4e7f0c61effed7701a04 \\\n    --hash=sha256:905a0ed8ea6f2d61c1738835f99b699348d7857379083e5fc497fa0c967a407c \\\n    --hash=sha256:90884113d8b48f760e9587002789ddd741e76ab9f89518cd1e43b1f1a52ec44b \\\n    --hash=sha256:91a06d2e259ecfbd8c901d70c3c507900458498142b3026a296b7de4d1322cc9 \\\n    --hash=sha256:926c9541b14b12b1681dca8a0b75feb510b06c6341b70a8e500c2fdcff837cce \\\n    --hash=sha256:9401557acd873c3a7f3eb9383edef8ac4968f9510e340f4808d427e75667e7b4 \\\n    --hash=sha256:9551187363ffc0de2a00b2e47c25aeaeb1020b69b668762966df15fc5659dd5a \\\n    --hash=sha256:962ccbab7b642487b1d8b7df90ef677e03134cf1fd8880bf698649b22a69371f \\\n    --hash=sha256:97e7cf2be5c77b7d1a9713a05605d49460d02c6078d38d8bef3cbe323c548424 \\\n    --hash=sha256:9aa768456404a8bf48a4406685ac2bec8e72b62c69313734fa3b73cf33b3a894 \\\n    --hash=sha256:9bc519fbf2b7578398853d815009ae5e4d4603d12f4e3f91da8c06852d3da3e9 \\\n    --hash=sha256:9d56801be94b86a9da183e5f3766e6310752b99ff647e38b09a9500d88e46e76 \\\n    --hash=sha256:9f444c499b3eefd3a92e348059471ea0c3a6e303d9c1cec09fa748fd9f895201 \\\n    --hash=sha256:9fa8ae11da9e2b3126c6426f147e0fba88d96d65921799bb30c6abd1cb2c97fb \\\n    --hash=sha256:a0f62d0a58f4e7da165457e995725421e0064f2255d8eccebc49f41bbc23b109 \\\n    --hash=sha256:a396dcc17e5a0b164dbe026896245a4fa9ff402edca1dff0be3d53a517f74de4 \\\n    --hash=sha256:aaa2a54443eff1950ba5ddc6b6ccda0d9c84a364276a62f969bdf2a390650848 \\\n    --hash=sha256:ad785e92e6dc634c21555edc8bd6b64957ab844541bcb96a1366c202951ae526 \\\n    --hash=sha256:af8244b2bef6aaad6d92cda81372de7f8c8d36c9f0c3ea36e827c60e7d9467a0 \\\n    --hash=sha256:b078afbc25f3a1436c7a1d2cd3e322497ee99615ba97c563566fdf46aff1ee01 \\\n    --hash=sha256:b2f69dec1725e79a012d920df1707de5caf7ed5e08f3be4435e25803efc47458 \\\n    --hash=sha256:b8458003118a712e66286df6a707db01c52c0f52f7db8e4a38f0da1d3b94fc4e \\\n    --hash=sha256:bb63e0198ca18aad131c089b9204c23079c3afa95487e561f4c522d519e55aba \\\n    --hash=sha256:bfec22eab3c8cc2ceec0248aec886624116dc079afa027ecc8ad4a7e62010f8a \\\n    --hash=sha256:c1747f85cee84c26985853c6f3d9bd3e75da5212912443fa111c113b9c246f39 \\\n    --hash=sha256:c1b3f518abeca3aa13c712fd202306e145abf59a18b094a6bafb2d2bbf59192c \\\n    --hash=sha256:c50f2528cf200c5eed56faf3f4e22fcd5f38c157a8b78576e6ba3168ec35f000 \\\n    --hash=sha256:c68fcd102d71ea85c5b2dfac3f4f8476eff42a9e078fd5faefff6d145063536b \\\n    --hash=sha256:c7a7bd4e39e8e4c12c39cd480356842b6a8a06e41b23a55a5e3e191718838ddf \\\n    --hash=sha256:c94f0688e7b8d0a67abf40e57a7eaaecd17cc9586706a31b76c031f63df052b4 \\\n    --hash=sha256:cbaf13819775b7f769bf4a1f066cb6df7a28d4480081a589828ef190226881cd \\\n    --hash=sha256:cd2213145bcc2ba85884d0ac63d222fece9209678f77b9b4d76f054c561adb28 \\\n    --hash=sha256:ce5c1d2a8b27468f433ca974829c44060b8097eedc39933e3c206a90ee49c4a9 \\\n    --hash=sha256:d396ec2b979760aaf3218e76c24e65bd0aca24983298653b3a9d7a45f9e47b30 \\\n    --hash=sha256:d51026d73fcfd93610abc7b27789c26b313920fcfb20e27462d74a7f8b06e983 \\\n    --hash=sha256:d80ee3d731373b24cebbc10d689ca4ee1875caf0d5703a245db18efd4dd37fc1 \\\n    --hash=sha256:d995260fdf4e1db774581b4900e0f832abe3c7c84996726bbc161b19c8f29e76 \\\n    --hash=sha256:da4b951fe36dc7c3a1ccb4e3cd1747c3542b8c9ceede8fc86cae054e764485f5 \\\n    --hash=sha256:daa27d92c36f24388fe3ad306b174781c747627f134452e4f128ea00ce1fe8c4 \\\n    --hash=sha256:db06ffe51636ffe9ca531fe9023dd64bdd794be8754cb5df57c5498ae5b518a7 \\\n    --hash=sha256:e0d65b8c354be7fb5f720c3caa8bc940bc2d20ce749c8e06135f07f8ed95dd7c \\\n    --hash=sha256:e68b7a074f65a2fd746c52a7ce6142ab7006074ac269ace0c25cd8ba171f8066 \\\n    --hash=sha256:e739fee756ba1010f8bcccb534252e85a35fe45ae92c295a06059ce58b74ccd3 \\\n    --hash=sha256:e846ae7835bf0703ae43f534ab79a867146dadd59dc9ca5c8b53d5c8f7c9ef02 \\\n    --hash=sha256:e9c26f834c65f5752f3f06cb08cb86a913ceb7274d0db6e267808a708b46bc89 \\\n    --hash=sha256:ea793e075b70290d89d8142074262885d3f7da19634845135751bd6344f73b50 \\\n    --hash=sha256:f027324c56cd5406ca49c124b0db10e56c69064fec039acc571c29020cc87c76 \\\n    --hash=sha256:f13a646d65d09fbf1bc6b3a9635d30095c8e7e5cc419ff35ecc563c5fd04cd49 \\\n    --hash=sha256:f47286a97f0bc9b8859519809077b91b2cefe4ae47fcbf5e466a009c1c5d742b \\\n    --hash=sha256:f747929cf940cddb5b3668a390056ddd5ba2e5010615ea2dcf4f9c4f3ab8791d \\\n    --hash=sha256:f99626688942fb746e545232e7726926f3be91b5975f8b55327665fafda991c7 \\\n    --hash=sha256:f9fa868638bf362d3d138ea55829cefb3d5f4b0d7f142234382a15e2485dbec4 \\\n    --hash=sha256:fbdb89b3e1c94a30cc5edfce477c6e6a5dc4d8f84665b455c27582f211a1c72c \\\n    --hash=sha256:fc010ab034c8c7452522748bf937df58020d256ccae0874463d1f4d01758af8e \\\n    --hash=sha256:fc3e9034a63de20e15e8ade85358bc6efc614008cab72898b4b4952bea0509ff \\\n    --hash=sha256:fd8b3d9fd264be37976686c7f65cd52a83f5e84f4bfd2adf9c1d469676bbb6ae\npygments==2.21.0 \\\n    --hash=sha256:2363c69b61c4a97c838da3b130dcd6468f4848992b21a82f2a63ec34377137d9 \\\n    --hash=sha256:610ca751c9bc2492b38eb9a38a7fbc93edbbb2d7182edaf34e66ae493dee5c8c\npymupdf==1.26.5 \\\n    --hash=sha256:2bfb58f07ad631e5f71ad0bd6f1ff52700f7ba7ebb4973130e81e75b721beae1 \\\n    --hash=sha256:39a6fb58182b27b51ea8150a0cd2e4ee7e0cf71e9d6723978f28699b42ee61ae \\\n    --hash=sha256:7dfea81fdd73437a6a6ce83e1fcf556faee9327a6540571e58bf04fa362bb0cd \\\n    --hash=sha256:8ef335e07f648492df240f2247854d0e7c0467afb9c4dc2376ec30978ec158c3 \\\n    --hash=sha256:a2a42f5911d153a47bf5c3e162a0bfe8745eb9bec3e59fbaf87617b4003d8270 \\\n    --hash=sha256:caad0ffeb63dcc4a29ca40f3c68d7b78d32a932e834b0056b529cc0bdbaaffc9 \\\n    --hash=sha256:d58599479bc471d3ae56c3d68d9160d0b7de8a3bd40221ddc3a4eaae2d281b86 \\\n    --hash=sha256:e24e7a7d696bd398543cc5c147869edb2026d5d5a21b7f8e35db2f20170b389e\npyparsing==3.3.2 \\\n    --hash=sha256:850ba148bd908d7e2411587e247a1e4f0327839c40e2e5e6d05a007ecc69911d \\\n    --hash=sha256:c777f4d763f140633dcb6d8a3eda953bf7a214dc4eff598413c070bcdc117cbc\npython-dateutil==2.9.0.post0 \\\n    --hash=sha256:37dd54208da7e1cd875388217d5e00ebd4179249f90fb72437e91a35459a0ad3 \\\n    --hash=sha256:a8b2bc7bffae282281c8140a97d3aa9c14da0b136dfe83f850eea9a5f7470427\npytz==2026.3.post1 \\\n    --hash=sha256:2211d3fcf9a797d3405cac96ac7f61d80e6a644f72a3309607282fe8a2010c5d \\\n    --hash=sha256:dd95840dd199baea12d9cc096a1d452caa6596a1c1e4b5f3dbd1541855d5e815\npyyaml==6.0.3 \\\n    --hash=sha256:00c4bdeba853cc34e7dd471f16b4114f4162dc03e6b7afcc2128711f0eca823c \\\n    --hash=sha256:0150219816b6a1fa26fb4699fb7daa9caf09eb1999f3b70fb6e786805e80375a \\\n    --hash=sha256:02893d100e99e03eda1c8fd5c441d8c60103fd175728e23e431db1b589cf5ab3 \\\n    --hash=sha256:02ea2dfa234451bbb8772601d7b8e426c2bfa197136796224e50e35a78777956 \\\n    --hash=sha256:0f29edc409a6392443abf94b9cf89ce99889a1dd5376d94316ae5145dfedd5d6 \\\n    --hash=sha256:10892704fc220243f5305762e276552a0395f7beb4dbf9b14ec8fd43b57f126c \\\n    --hash=sha256:16249ee61e95f858e83976573de0f5b2893b3677ba71c9dd36b9cf8be9ac6d65 \\\n    --hash=sha256:1d37d57ad971609cf3c53ba6a7e365e40660e3be0e5175fa9f2365a379d6095a \\\n    --hash=sha256:1ebe39cb5fc479422b83de611d14e2c0d3bb2a18bbcb01f229ab3cfbd8fee7a0 \\\n    --hash=sha256:214ed4befebe12df36bcc8bc2b64b396ca31be9304b8f59e25c11cf94a4c033b \\\n    --hash=sha256:2283a07e2c21a2aa78d9c4442724ec1eb15f5e42a723b99cb3d822d48f5f7ad1 \\\n    --hash=sha256:22ba7cfcad58ef3ecddc7ed1db3409af68d023b7f940da23c6c2a1890976eda6 \\\n    --hash=sha256:27c0abcb4a5dac13684a37f76e701e054692a9b2d3064b70f5e4eb54810553d7 \\\n    --hash=sha256:28c8d926f98f432f88adc23edf2e6d4921ac26fb084b028c733d01868d19007e \\\n    --hash=sha256:2e71d11abed7344e42a8849600193d15b6def118602c4c176f748e4583246007 \\\n    --hash=sha256:34d5fcd24b8445fadc33f9cf348c1047101756fd760b4dacb5c3e99755703310 \\\n    --hash=sha256:37503bfbfc9d2c40b344d06b2199cf0e96e97957ab1c1b546fd4f87e53e5d3e4 \\\n    --hash=sha256:3c5677e12444c15717b902a5798264fa7909e41153cdf9ef7ad571b704a63dd9 \\\n    --hash=sha256:3ff07ec89bae51176c0549bc4c63aa6202991da2d9a6129d7aef7f1407d3f295 \\\n    --hash=sha256:41715c910c881bc081f1e8872880d3c650acf13dfa8214bad49ed4cede7c34ea \\\n    --hash=sha256:418cf3f2111bc80e0933b2cd8cd04f286338bb88bdc7bc8e6dd775ebde60b5e0 \\\n    --hash=sha256:44edc647873928551a01e7a563d7452ccdebee747728c1080d881d68af7b997e \\\n    --hash=sha256:4a2e8cebe2ff6ab7d1050ecd59c25d4c8bd7e6f400f5f82b96557ac0abafd0ac \\\n    --hash=sha256:4ad1906908f2f5ae4e5a8ddfce73c320c2a1429ec52eafd27138b7f1cbe341c9 \\\n    --hash=sha256:501a031947e3a9025ed4405a168e6ef5ae3126c59f90ce0cd6f2bfc477be31b7 \\\n    --hash=sha256:5190d403f121660ce8d1d2c1bb2ef1bd05b5f68533fc5c2ea899bd15f4399b35 \\\n    --hash=sha256:5498cd1645aa724a7c71c8f378eb29ebe23da2fc0d7a08071d89469bf1d2defb \\\n    --hash=sha256:5cf4e27da7e3fbed4d6c3d8e797387aaad68102272f8f9752883bc32d61cb87b \\\n    --hash=sha256:5e0b74767e5f8c593e8c9b5912019159ed0533c70051e9cce3e8b6aa699fcd69 \\\n    --hash=sha256:5ed875a24292240029e4483f9d4a4b8a1ae08843b9c54f43fcc11e404532a8a5 \\\n    --hash=sha256:5fcd34e47f6e0b794d17de1b4ff496c00986e1c83f7ab2fb8fcfe9616ff7477b \\\n    --hash=sha256:5fdec68f91a0c6739b380c83b951e2c72ac0197ace422360e6d5a959d8d97b2c \\\n    --hash=sha256:6344df0d5755a2c9a276d4473ae6b90647e216ab4757f8426893b5dd2ac3f369 \\\n    --hash=sha256:64386e5e707d03a7e172c0701abfb7e10f0fb753ee1d773128192742712a98fd \\\n    --hash=sha256:652cb6edd41e718550aad172851962662ff2681490a8a711af6a4d288dd96824 \\\n    --hash=sha256:66291b10affd76d76f54fad28e22e51719ef9ba22b29e1d7d03d6777a9174198 \\\n    --hash=sha256:66e1674c3ef6f541c35191caae2d429b967b99e02040f5ba928632d9a7f0f065 \\\n    --hash=sha256:6adc77889b628398debc7b65c073bcb99c4a0237b248cacaf3fe8a557563ef6c \\\n    --hash=sha256:79005a0d97d5ddabfeeea4cf676af11e647e41d81c9a7722a193022accdb6b7c \\\n    --hash=sha256:7c6610def4f163542a622a73fb39f534f8c101d690126992300bf3207eab9764 \\\n    --hash=sha256:7f047e29dcae44602496db43be01ad42fc6f1cc0d8cd6c83d342306c32270196 \\\n    --hash=sha256:8098f252adfa6c80ab48096053f512f2321f0b998f98150cea9bd23d83e1467b \\\n    --hash=sha256:850774a7879607d3a6f50d36d04f00ee69e7fc816450e5f7e58d7f17f1ae5c00 \\\n    --hash=sha256:8d1fab6bb153a416f9aeb4b8763bc0f22a5586065f86f7664fc23339fc1c1fac \\\n    --hash=sha256:8da9669d359f02c0b91ccc01cac4a67f16afec0dac22c2ad09f46bee0697eba8 \\\n    --hash=sha256:8dc52c23056b9ddd46818a57b78404882310fb473d63f17b07d5c40421e47f8e \\\n    --hash=sha256:9149cad251584d5fb4981be1ecde53a1ca46c891a79788c0df828d2f166bda28 \\\n    --hash=sha256:93dda82c9c22deb0a405ea4dc5f2d0cda384168e466364dec6255b293923b2f3 \\\n    --hash=sha256:96b533f0e99f6579b3d4d4995707cf36df9100d67e0c8303a0c55b27b5f99bc5 \\\n    --hash=sha256:9c57bb8c96f6d1808c030b1687b9b5fb476abaa47f0db9c0101f5e9f394e97f4 \\\n    --hash=sha256:9c7708761fccb9397fe64bbc0395abcae8c4bf7b0eac081e12b809bf47700d0b \\\n    --hash=sha256:9f3bfb4965eb874431221a3ff3fdcddc7e74e3b07799e0e84ca4a0f867d449bf \\\n    --hash=sha256:a33284e20b78bd4a18c8c2282d549d10bc8408a2a7ff57653c0cf0b9be0afce5 \\\n    --hash=sha256:a80cb027f6b349846a3bf6d73b5e95e782175e52f22108cfa17876aaeff93702 \\\n    --hash=sha256:b30236e45cf30d2b8e7b3e85881719e98507abed1011bf463a8fa23e9c3e98a8 \\\n    --hash=sha256:b3bc83488de33889877a0f2543ade9f70c67d66d9ebb4ac959502e12de895788 \\\n    --hash=sha256:b865addae83924361678b652338317d1bd7e79b1f4596f96b96c77a5a34b34da \\\n    --hash=sha256:b8bb0864c5a28024fac8a632c443c87c5aa6f215c0b126c449ae1a150412f31d \\\n    --hash=sha256:ba1cc08a7ccde2d2ec775841541641e4548226580ab850948cbfda66a1befcdc \\\n    --hash=sha256:bdb2c67c6c1390b63c6ff89f210c8fd09d9a1217a465701eac7316313c915e4c \\\n    --hash=sha256:c1ff362665ae507275af2853520967820d9124984e0f7466736aea23d8611fba \\\n    --hash=sha256:c2514fceb77bc5e7a2f7adfaa1feb2fb311607c9cb518dbc378688ec73d8292f \\\n    --hash=sha256:c3355370a2c156cffb25e876646f149d5d68f5e0a3ce86a5084dd0b64a994917 \\\n    --hash=sha256:c458b6d084f9b935061bc36216e8a69a7e293a2f1e68bf956dcd9e6cbcd143f5 \\\n    --hash=sha256:d0eae10f8159e8fdad514efdc92d74fd8d682c933a6dd088030f3834bc8e6b26 \\\n    --hash=sha256:d76623373421df22fb4cf8817020cbb7ef15c725b9d5e45f17e189bfc384190f \\\n    --hash=sha256:ebc55a14a21cb14062aa4162f906cd962b28e2e9ea38f9b4391244cd8de4ae0b \\\n    --hash=sha256:eda16858a3cab07b80edaf74336ece1f986ba330fdb8ee0d6c0d68fe82bc96be \\\n    --hash=sha256:ee2922902c45ae8ccada2c5b501ab86c36525b883eff4255313a253a3160861c \\\n    --hash=sha256:efd7b85f94a6f21e4932043973a7ba2613b059c4a000551892ac9f1d11f5baf3 \\\n    --hash=sha256:f7057c9a337546edc7973c0d3ba84ddcdf0daa14533c2065749c9075001090e6 \\\n    --hash=sha256:fa160448684b4e94d80416c0fa4aac48967a969efe22931448d853ada8baf926 \\\n    --hash=sha256:fc09d0aa354569bc501d4e787133afc08552722d3ab34836a80547331bb5d4a0\nqdrant-client==1.15.1 \\\n    --hash=sha256:2b975099b378382f6ca1cfb43f0d59e541be6e16a5892f282a4b8de7eff5cb63 \\\n    --hash=sha256:631f1f3caebfad0fd0c1fba98f41be81d9962b7bf3ca653bed3b727c0e0cbe0e\nrank-bm25==0.2.2 \\\n    --hash=sha256:096ccef76f8188563419aaf384a02f0ea459503fdf77901378d4fd9d87e5e51d \\\n    --hash=sha256:7bd4a95571adadfc271746fa146a4bcfd89c0cf731e49c3d1ad863290adbe8ae\nregex==2026.7.19 \\\n    --hash=sha256:062f8cb7a9739c4835d22bd96f370c59aba89f257adcfa53be3cc209e08d3ae0 \\\n    --hash=sha256:064f1760a5a4ade65c5419be23e782f29147528e8a66e0c42dd4cedb8d4e9fc6 \\\n    --hash=sha256:09523a592938aa9f587fb74467c63ff0cf88fc3df14c82ab0f0517dcf76aaa62 \\\n    --hash=sha256:09d3007fc76249a83cdd33de160d50e6cb77f54e09d8fa9e7148e10607ce24af \\\n    --hash=sha256:09f3e5287f94f17b709dc9a9e70865855feee835c861613be144218ce4ca82cc \\\n    --hash=sha256:0c41c63992bf1874cebb6e7f56fd7d3c007924659a604ae3d90e427d40d4fd13 \\\n    --hash=sha256:0e9554c8785eac5cffe6300f69a91f58ba72bc88a5f8d661235ad7c6aa5b8ccd \\\n    --hash=sha256:1123ef4211d763ee771d47916a1596e2f4915794f7aabdc1adcb20e4249a6951 \\\n    --hash=sha256:15b364b9b98d6d2fe1a85034c23a3180ff913f46caddc3895f6fd65186255ccc \\\n    --hash=sha256:1649eb39fcc9ea80c4d2f110fde2b8ab2aef3877b98f02ab9b14e961f418c511 \\\n    --hash=sha256:17ed5692f6acc4183e98331101a5f9e4f64d72fe58b753da4d444a2c77d05b12 \\\n    --hash=sha256:199535629f25caf89698039af3d1ad5fcae7f933e2112c73f1cdf49165c99518 \\\n    --hash=sha256:1c398716054621aa300b3d411f467dda903806c5da0df6945ab73982b8d115db \\\n    --hash=sha256:1d3372064506b94dd2c67c845f2db8062e9e9ba84d04e33cb96d7d33c11fe1ae \\\n    --hash=sha256:1d58561843f0ff7dc78b4c28b5e2dc388f3eff94ebc8a232a3adba961fc00009 \\\n    --hash=sha256:1d793a7988e04fcb1e2e135567443d82173225d657419ec09414a9b5a145b986 \\\n    --hash=sha256:1ebac3474b8589fce2f9b225b650afd61448f7c73a5d0255a10cc6366471aed1 \\\n    --hash=sha256:20568e182eb82d39a6bf7cff3fd58566f14c75c6f74b2c8c96537eecf9010e3a \\\n    --hash=sha256:22a992de9a0d91bda927bf02b94351d737a0302905432c88a53de7c4b9ce62e2 \\\n    --hash=sha256:2955907b7157a6660f27079edf7e0229e9c9c5325c77a2ef6a890cba91efa6f0 \\\n    --hash=sha256:2c4e61e2e1be56f63ec3cc618aa9e0de81ef6f43d177205451840022e24f5b78 \\\n    --hash=sha256:2cc3460cedf7579948486eab03bc9ad7089df4d7281c0f47f4afe03e8d13f02d \\\n    --hash=sha256:2ce9e679f776649746729b6c86382da519ef649c8e34cc41df0d2e5e0f6c36d4 \\\n    --hash=sha256:2ef7eeb108c47ce7bcc9513e51bcb1bf57e8f483d52fce68a8642e3527141ae0 \\\n    --hash=sha256:3080a7fd38ef049bd489e01c970c97dd84ff446a885b0f1f6b26d9b1ad13ce11 \\\n    --hash=sha256:343a4504e3fb688c47cad451221ca5d4814f42b1e16c0065bde9cbf7f473bd52 \\\n    --hash=sha256:36aacfb15faaff3ced55afbf35ec72f50d4aee22082c4f7fe0573a33e2fca92e \\\n    --hash=sha256:3d3143f159261b1ce5b24c261c590e5913370c3200c5e9ebbb92b5aa5e111902 \\\n    --hash=sha256:40b34dd88658e4fedd2fddbf0275ac970d00614b731357f425722a3ed1983d11 \\\n    --hash=sha256:4458124d71339f505bf1fb94f69fd1bb8fa9d2481eebfef27c10ef4f2b9e12f6 \\\n    --hash=sha256:4896db1f4ce0576765b8272aa922df324e0f5b9bb2c3d03044ff32a7234a9aba \\\n    --hash=sha256:4a0530bb1b8c1c985e7e2122e2b4d3aedd8a3c21c6bfddae6767c4405668b56e \\\n    --hash=sha256:4aa5435cdb3eb6f55fe98a171b05e3fbcd95fadaa4aa32acf62afd9b0cfdbcac \\\n    --hash=sha256:4c3501bfa814ab07b5580741f9bf78dfdfe146a04057f82df9e2402d2a975939 \\\n    --hash=sha256:4e5413bd5f13d3a4e3539ca98f70f75e7fca92518dd7f117f030ebedd10b60cb \\\n    --hash=sha256:4e6883a021db30511d9fb8cfb0f222ce1f2c369f7d4d8b0448f449a93ba0bdfc \\\n    --hash=sha256:52579c60a6078be70a0e49c81d6e56d677f34cd439af281a0083b8c7bc75c095 \\\n    --hash=sha256:555497390743af1a65045fa4527782d10ff5b88970359412baa4a1e628fe393b \\\n    --hash=sha256:56ad4d9f77df871a99e25c37091052a02528ec0eb059de928ee33956b854b45b \\\n    --hash=sha256:571fde9741eb0ccde23dd4e0c1d50fbae910e901fa7e629faf39b2dda740d220 \\\n    --hash=sha256:572fc57b0009c735ee56c175ea021b637a15551a312f56734277f923d6fd0f6c \\\n    --hash=sha256:59787bd5f8c70aa339084e961d2996b53fbdeab4d5393bba5c1fe1fc32e02bae \\\n    --hash=sha256:5a2721c8720e2cb3c209925dfb9200199b4b07361c9e01d321719404b21458b3 \\\n    --hash=sha256:5cc26a66e212fa5d6c6170c3a40d99d888db3020c6fdab1523250d4341382e44 \\\n    --hash=sha256:5ebee1ee89c39c953baac6924fcde08c5bb427c4057510862f9d7c7bdb3d8665 \\\n    --hash=sha256:60be8693a1dadc210bbcbc0db3e26da5f7d01d1d5a3da594e99b4fa42df404f5 \\\n    --hash=sha256:618a0aed532be87294c4477b0481f3aa0f1520f4014a4374dd4cf789b4cd2c97 \\\n    --hash=sha256:61bb1bd45520aacd56dd80943bd34991fb5350afdd1f36f2282230fd5154a218 \\\n    --hash=sha256:6383cd2ed53a646c659ba1fe65727db76437fdaa069e697a0b44a51d5843d864 \\\n    --hash=sha256:64729333167c2dcaaa56a331d40ee097bd9c5617ffd51dabb09eaddafb1b532e \\\n    --hash=sha256:64b6ca7391a1395c2638dd5c7456d67bea44fc6c5e8e92c5dc8aa6a8f23292b4 \\\n    --hash=sha256:65dcd28d3eba2ab7c2fd906485cc301392b47cc2234790d27d4e4814e02cdfda \\\n    --hash=sha256:65fa6cb38ed5e9c3637e68e544f598b39c3b86b808ed0627a67b68320384b459 \\\n    --hash=sha256:66bd62c59a5427746e8c44becae1d9b99d22fb13f30f492083dfb9ad7c45cc18 \\\n    --hash=sha256:6e44c0e7c5664be20aee92085153150c0a7967310a73a43c0f832b7cd35d0dd3 \\\n    --hash=sha256:6f8c6e7a1cfa3dc9d0ee2de0e65e834537fa29992cc3976ffec914afc35c5dd5 \\\n    --hash=sha256:7322ec6cc9fba9d49ab888bb82d67ac5625627aa168f0165139b17018df3fb8a \\\n    --hash=sha256:73b133a9e6fb512858e7f065e96f1180aa46646bc74a83aea62f1d314f3dd035 \\\n    --hash=sha256:73f272fba87b8ccfe70a137d02a54af386f6d27aa509fbffdd978f5947aae1aa \\\n    --hash=sha256:7e77b324909c1617cbb4c668677e2c6ae13f44d7c1de0d4f15f2e3c10f3315b5 \\\n    --hash=sha256:80115dd39481fd3a4b4080220799dbcacb921a844de4b827264ececacbe17c78 \\\n    --hash=sha256:87ccab0db8d5f4fbb0272642113c1adb2ffc698c16d3a0944580222331fa7a20 \\\n    --hash=sha256:89dfee3319f5ae3f75ebd5c2445a809bb320252ba5529ffdafea4ef25d79cf1a \\\n    --hash=sha256:8ac59a0900474a52b7c04af8196affc22bd9842acb0950df12f7b813e983609a \\\n    --hash=sha256:8cae6fd77a5b72dae505084b1a2ee0360139faf72fedbab667cd7cc65aae7a6a \\\n    --hash=sha256:8d3469c91dd92ee41b7c95280edbd975ef1ba9195086686623a1c6e8935ce965 \\\n    --hash=sha256:90c633e7e8d6bf4e992b8b36ce69e018f834b641dd6de8cea6d78c06ffa119c5 \\\n    --hash=sha256:93db40c8de0815baab96a06e08a984bac71f989d13bab789e382158c5d426797 \\\n    --hash=sha256:9724e6cb5e478cd7d8cabf027826178739cb18cf0e117d0e32814d479fa02276 \\\n    --hash=sha256:98c6ac18480fcdb33f35439183f1d2e79760ab41930309c6d951cb1f8e46694c \\\n    --hash=sha256:9a15e785f244f3e07847b984ce8773fc3da10a9f3c131cc49a4c5b4d672b4547 \\\n    --hash=sha256:9b60d7814174f059e5de4ab98271cc5ba9259cfea55273a81544dceea32dc8d9 \\\n    --hash=sha256:9be2a6647740dd3cca6acb24e87f03d7632cd280dbce9bbe40c26353a215a45d \\\n    --hash=sha256:9c7472192ebfad53a6be7c4a8bfb2d64b81c0e93a1fc8c57e1dd0b638297b5d1 \\\n    --hash=sha256:9dce8ec9695f531a1b8a6f314fd4b393adcccf2ea861db480cdf97a301d01a68 \\\n    --hash=sha256:9e50d748a32da622f256e8d505867f5d3c43a837c6a9f0efb149655fadd1042a \\\n    --hash=sha256:a81758ed242b861b72e778ba34d41366441a2e10b16b472784c88da2dea7e2dd \\\n    --hash=sha256:ac777001cdfc28b72477d93c8564bb7583081ea8fb45cdca3d568e0a4f87183c \\\n    --hash=sha256:b2b506b1788df5fecd270a10d5e70a95fe77b87ea2b370a318043f6f5f817ee6 \\\n    --hash=sha256:b2ea4a3e8357be8849e833beeae757ac3c7a6b3fc055c03c808a53c91ad30d82 \\\n    --hash=sha256:bf1516fe58fc104f39b2d1dbe2d5e27d0cd45c4be2e42ba6ee0cc763701ec3c7 \\\n    --hash=sha256:c0d702548d89d572b2929879bc883bb7a4c4709efafe4512cadee56c55c9bd15 \\\n    --hash=sha256:c10b82c2634df08dfb13b1f04e38fe310d086ee092f4f69c0c8da234251e556e \\\n    --hash=sha256:c42572142ed0b9d5d261ba727157c426510da78e20828b66bbb855098b8a4e38 \\\n    --hash=sha256:c4585c3e64b4f9e583b4d2683f18f5d5d872b3d71dcf24594b74ecc23602fa96 \\\n    --hash=sha256:c639ea314df70a7b2811e8020448c75af8c9445f5a60f8a4ced81c306a9380c2 \\\n    --hash=sha256:c670fe7be5b6020b76bc6e8d2196074657e1327595bca93a389e1a76ab130ad8 \\\n    --hash=sha256:cc1b2440423a851fad781309dd87843868f4f66a6bcd1ddb9225cf4ec2c84732 \\\n    --hash=sha256:cd3584591ea4429026cdb931b054342c2bcf189b44ff367f8d5c15bc092a2966 \\\n    --hash=sha256:d15df07081d91b76ff20d43f94592ee110330152d617b730fdbe5ef9fb680053 \\\n    --hash=sha256:d19662dbedbe783d323196312d38f5ba53cf56296378252171985da6899887d3 \\\n    --hash=sha256:d24ecb4f5e009ea0bd275ee37ad9953b32005e2e5e60f8bbae16da0dbbf0d3a0 \\\n    --hash=sha256:d446c6ac40bb6e05025ccee55b84d80fe9bf8e93010ffc4bb9484f13d498835f \\\n    --hash=sha256:d51ffd3427640fa2da6ade574ceba932f210ad095f65fcc450a2b0a0d454868e \\\n    --hash=sha256:d6ce43a0269d68cee79a7d1ade7def53c20f8f2a047b92d7b5d5bcc73ae88327 \\\n    --hash=sha256:d721e53758b2cca74990185eb0671dd466d7a388a1a45d0c6f4c13cef41a68ac \\\n    --hash=sha256:d7da47a0f248977f08e2cb659ff3c17ddc13a4d39b3a7baa0a81bf5b415430f6 \\\n    --hash=sha256:db47b561c9afd884baa1f96f797c9ca369872c4b65912bc691cfa99e68340af2 \\\n    --hash=sha256:dbe6493fbd27321b1d1f2dd4f5c7e5bd4d8b1d7cab7f32fd67db3d0b2ed8248a \\\n    --hash=sha256:dbece16025afda5e3031af0c4059207e61dcf73ef13af844964f57f387d1c435 \\\n    --hash=sha256:ddd67571c10869f65a5d7dde536d1e066e306cc90de57d7de4d5f34802428bb5 \\\n    --hash=sha256:de9208bb427130c82a5dbfd104f92c8876fc9559278c880b3002755bbbe9c83d \\\n    --hash=sha256:e30d40268a28d54ce0437031750497004c22602b8e3ab891f759b795a003b312 \\\n    --hash=sha256:e8b0abe7d870f53ca5143895fef7d1041a0c831a140d3dc2c760dd7ba25d4a8b \\\n    --hash=sha256:f035d9dc1d25eff9d361456572231c7d27b5ccd473ca7dc0adfce732bd006d40 \\\n    --hash=sha256:f04b9f56b0e0614c0126be12c2c2d9f8850c1e57af302bd0a63bed379d4af974 \\\n    --hash=sha256:f0fa4fa9c3632d708742baf2282f2055c11d888a790362670a403cbf48a2c404 \\\n    --hash=sha256:f2e7f8e2ab6c2922be02c7ec45185aa5bd771e2e57b95455ee343a44d8130dff \\\n    --hash=sha256:f8f6fa298bb4f7f58a33334406218ba74716e68feddf5e4e54cd5d8082705abf \\\n    --hash=sha256:fbf300e2070bb35038660b3be1be4b91b0024edb41517e6996320b49b92b4175 \\\n    --hash=sha256:fce7760bf283405b2c7999cab3da4e72f7deca6396013115e3f7a955db9760da \\\n    --hash=sha256:fcee38cd8e5089d6d4f048ba1233b3ad76e5954f545382180889112ff5cb712d \\\n    --hash=sha256:fe31f28c94402043161876a258a9c6f757cb485905c7614ce8d6cd40e6b7bdc1 \\\n    --hash=sha256:ffd8893ccc1c2fce6e0d6ca402d716fe1b29db70c7132609a05955e31b2aa8f2\nrequests==2.34.2 \\\n    --hash=sha256:2a0d60c172f83ac6ab31e4554906c0f3b3588d37b5cb939b1c061f4907e278e0 \\\n    --hash=sha256:f288924cae4e29463698d6d60bc6a4da69c89185ad1e0bcc4104f584e960b9ed\nrich==15.0.0 \\\n    --hash=sha256:33bd4ef74232fb73fe9279a257718407f169c09b78a87ad3d296f548e27de0bb \\\n    --hash=sha256:edd07a4824c6b40189fb7ac9bc4c52536e9780fbbfbddf6f1e2502c31b068c36\nrouge-score==0.1.2 \\\n    --hash=sha256:c7d4da2683e68c9abf0135ef915d63a46643666f848e558a1b9f7ead17ff0f04\nsacrebleu==2.5.1 \\\n    --hash=sha256:1a088cc1c74ffaff0759c3191a85db09eecfa7a52e09be244e319d8d64e2fb11 \\\n    --hash=sha256:7c9f7ee75bec3a5bf19dd87112dfd654952130e403ad30c48298fb7da3212d5d\nsafetensors==0.8.0 \\\n    --hash=sha256:040070828e36dc8e122178bbbd5830ff9e97920affb84cbe0f46442497bed358 \\\n    --hash=sha256:096ec1a98435df7beb08853bb5aa9081a84f23d0adc67ed1a0a10550f608373f \\\n    --hash=sha256:2ddf52eac562eda224f99acfa7889d02968c1fd59a5b011ae7d8137c37e9c02d \\\n    --hash=sha256:3ae091f16662658bdc019a4ff6cb4c085bb7d725eb5978b183ffd265863b6d2d \\\n    --hash=sha256:4124502b78f03534117c848f87a39b8f31e577b15eff423bf8bfb95f2a8c30d0 \\\n    --hash=sha256:4a95ae2b05d7726d751da4ebf626a2ca782b706e101bd894c95bc2450b1cffcc \\\n    --hash=sha256:7a46e5ff292c356d6991e60942ba7f79817682d3a2cef0702136448cb9c4d235 \\\n    --hash=sha256:7bc0a787ba8a35be368ee3574edfa2b1ad389eebd0a72e482ae275490e3f6c98 \\\n    --hash=sha256:87eec7ffed2b809f05a398a8becb7d013f19f7837cd15d9748580d6cf30dbaf4 \\\n    --hash=sha256:8e080062fcde23be189565e1c3305d16751a218ecf9412c8601e64204eb6f846 \\\n    --hash=sha256:8e9f537aa183a38ace122d27303dcd986b26bd2a7591f9181d7f0c396f4677ca \\\n    --hash=sha256:c554f85858e05226d3c2828e32395e677434685d6d94594a41643361c5e837f0 \\\n    --hash=sha256:c80201d22cbf405b80647a60ada77bba06c8fba2da2743ba1e89cdcc39a81f25 \\\n    --hash=sha256:f7838e5135a406ad3e02efdcb8cf2e5397d368b0154537c4fec682dbc544d452 \\\n    --hash=sha256:fabaf3e0f18a6618d9b36560682562157f77c2b71fcffc7b432be2baed9d753d \\\n    --hash=sha256:fcdd41ec4628fee5799f807c73c353629130fbd942aa23d83c623dd6c9d52d78 \\\n    --hash=sha256:fd6f3f93c9a0a7cc2788ee63fb763353d4bd2e89b0751bc78fcf7dda00bea774\nscikit-learn==1.7.2 \\\n    --hash=sha256:0486c8f827c2e7b64837c731c8feff72c0bd2b998067a8a9cbc10643c31f0fe1 \\\n    --hash=sha256:0b7dacaa05e5d76759fb071558a8b5130f4845166d88654a0f9bdf3eb57851b7 \\\n    --hash=sha256:191e5550980d45449126e23ed1d5e9e24b2c68329ee1f691a3987476e115e09c \\\n    --hash=sha256:20e9e49ecd130598f1ca38a1d85090e1a600147b9c02fa6f15d69cb53d968fda \\\n    --hash=sha256:2a41e2a0ef45063e654152ec9d8bcfc39f7afce35b08902bfe290c2498a67a6a \\\n    --hash=sha256:36749fb62b3d961b1ce4fedf08fa57a1986cd409eff2d783bca5d4b9b5fce51c \\\n    --hash=sha256:4a847fea807e278f821a0406ca01e387f97653e284ecbd9750e3ee7c90347f18 \\\n    --hash=sha256:502c18e39849c0ea1a5d681af1dbcf15f6cce601aebb657aabbfe84133c1907f \\\n    --hash=sha256:57dc4deb1d3762c75d685507fbd0bc17160144b2f2ba4ccea5dc285ab0d0e973 \\\n    --hash=sha256:6088aa475f0785e01bcf8529f55280a3d7d298679f50c0bb70a2364a82d0b290 \\\n    --hash=sha256:63a9afd6f7b229aad94618c01c252ce9e6fa97918c5ca19c9a17a087d819440c \\\n    --hash=sha256:6b33579c10a3081d076ab403df4a4190da4f4432d443521674637677dc91e61f \\\n    --hash=sha256:7a4c328a71785382fe3fe676a9ecf2c86189249beff90bf85e22bdb7efaf9ae0 \\\n    --hash=sha256:7a58814265dfc52b3295b1900cfb5701589d30a8bb026c7540f1e9d3499d5ec8 \\\n    --hash=sha256:89877e19a80c7b11a2891a27c21c4894fb18e2c2e077815bcade10d34287b20d \\\n    --hash=sha256:8d91a97fa2b706943822398ab943cde71858a50245e31bc71dba62aab1d60a96 \\\n    --hash=sha256:8da8bf89d4d79aaec192d2bda62f9b56ae4e5b4ef93b6a56b5de4977e375c1f1 \\\n    --hash=sha256:9656e4a53e54578ad10a434dc1f993330568cfee176dff07112b8785fb413106 \\\n    --hash=sha256:96dc05a854add0e50d3f47a1ef21a10a595016da5b007c7d9cd9d0bffd1fcc61 \\\n    --hash=sha256:98335fb98509b73385b3ab2bd0639b1f610541d3988ee675c670371d6a87aa7c \\\n    --hash=sha256:9acb6c5e867447b4e1390930e3944a005e2cb115922e693c08a323421a6966e8 \\\n    --hash=sha256:9b7ed8d58725030568523e937c43e56bc01cadb478fc43c042a9aca1dacb3ba1 \\\n    --hash=sha256:abebbd61ad9e1deed54cca45caea8ad5f79e1b93173dece40bb8e0c658dbe6fe \\\n    --hash=sha256:acbc0f5fd2edd3432a22c69bed78e837c70cf896cd7993d71d51ba6708507476 \\\n    --hash=sha256:b4d6e9deed1a47aca9fe2f267ab8e8fe82ee20b4526b2c0cd9e135cea10feb44 \\\n    --hash=sha256:bb24510ed3f9f61476181e4db51ce801e2ba37541def12dc9333b946fc7a9cf8 \\\n    --hash=sha256:c7509693451651cd7361d30ce4e86a1347493554f172b1c72a39300fa2aea79e \\\n    --hash=sha256:ca250e6836d10e6f402436d6463d6c0e4d8e0234cfb6a9a47835bd392b852ce5 \\\n    --hash=sha256:e5bf3d930aee75a65478df91ac1225ff89cd28e9ac7bd1196853a9229b6adb0b \\\n    --hash=sha256:f95dc55b7902b91331fa4e5845dd5bde0580c9cd9612b1b2791b7e80c3d32615 \\\n    --hash=sha256:fa8f63940e29c82d1e67a45d5297bdebbcb585f5a5a50c4914cc2e852ab77f33\nscipy==1.16.2 \\\n    --hash=sha256:024dd4a118cccec09ca3209b7e8e614931a6ffb804b2a601839499cb88bdf925 \\\n    --hash=sha256:033570f1dcefd79547a88e18bccacff025c8c647a330381064f561d43b821232 \\\n    --hash=sha256:03dfc75e52f72cf23ec2ced468645321407faad8f0fe7b1f5b49264adbc29cb1 \\\n    --hash=sha256:0a17541827a9b78b777d33b623a6dcfe2ef4a25806204d08ead0768f4e529a88 \\\n    --hash=sha256:0ce54e07bbb394b417457409a64fd015be623f36e330ac49306433ffe04bc97e \\\n    --hash=sha256:116296e89fba96f76353a8579820c2512f6e55835d3fad7780fece04367de351 \\\n    --hash=sha256:17d9bb346194e8967296621208fcdfd39b55498ef7d2f376884d5ac47cec1a70 \\\n    --hash=sha256:26284797e38b8a75e14ea6631d29bda11e76ceaa6ddb6fdebbfe4c4d90faf2f9 \\\n    --hash=sha256:2a8ffaa4ac0df81a0b94577b18ee079f13fecdb924df3328fc44a7dc5ac46851 \\\n    --hash=sha256:2cc73a33305b4b24556957d5857d6253ce1e2dcd67fa0ff46d87d1670b3e1e1d \\\n    --hash=sha256:2f5350da923ccfd0b00e07c3e5cfb316c1c0d6c1d864c07a72d092e9f20db104 \\\n    --hash=sha256:4e409eac067dcee96a57fbcf424c13f428037827ec7ee3cb671ff525ca4fc34d \\\n    --hash=sha256:5221c0b2a4b58aa7c4ed0387d360fd90ee9086d383bb34d9f2789fafddc8a936 \\\n    --hash=sha256:53d8d2ee29b925344c13bda64ab51785f016b1b9617849dac10897f0701b20c1 \\\n    --hash=sha256:567e77755019bb7461513c87f02bb73fb65b11f049aaaa8ca17cfaa5a5c45d77 \\\n    --hash=sha256:5c39026d12edc826a1ef2ad35ad1e6d7f087f934bb868fc43fa3049c8b8508f9 \\\n    --hash=sha256:5c66511f29aa8d233388e7416a3f20d5cae7a2744d5cee2ecd38c081f4e861b3 \\\n    --hash=sha256:5e9feab931bd2aea4a23388c962df6468af3d808ddf2d40f94a81c5dc38f32ef \\\n    --hash=sha256:63870a84cd15c44e65220eaed2dac0e8f8b26bbb991456a033c1d9abfe8a94f8 \\\n    --hash=sha256:6406d2ac6d40b861cccf57f49592f9779071655e9f75cd4f977fa0bdd09cb2e4 \\\n    --hash=sha256:654324826654d4d9133e10675325708fb954bc84dae6e9ad0a52e75c6b1a01d7 \\\n    --hash=sha256:6ab88ea43a57da1af33292ebd04b417e8e2eaf9d5aa05700be8d6e1b6501cd92 \\\n    --hash=sha256:70327d6aa572a17c2941cdfb20673f82e536e91850a2e4cb0c5b858b690e1548 \\\n    --hash=sha256:7280d926f11ca945c3ef92ba960fa924e1465f8d07ce3a9923080363390624c4 \\\n    --hash=sha256:7a5dc7ee9c33019973a470556081b0fd3c9f4c44019191039f9769183141a4d9 \\\n    --hash=sha256:7f3a337d9ae06a1e8d655ee9d8ecb835ea5ddcdcbd8d23012afa055ab014f374 \\\n    --hash=sha256:7fe65b36036357003b3ef9d37547abeefaa353b237e989c21027b8ed62b12d4f \\\n    --hash=sha256:84f7bf944b43e20b8a894f5fe593976926744f6c185bacfcbdfbb62736b5cc70 \\\n    --hash=sha256:87eb178db04ece7c698220d523c170125dbffebb7af0345e66c3554f6f60c173 \\\n    --hash=sha256:89d6c100fa5c48472047632e06f0876b3c4931aac1f4291afc81a3644316bb0d \\\n    --hash=sha256:8afae1756f6a1fe04636407ef7dbece33d826a5d462b74f3d0eb82deabefd831 \\\n    --hash=sha256:912f46667d2d3834bc3d57361f854226475f695eb08c08a904aadb1c936b6a88 \\\n    --hash=sha256:91e9e8a37befa5a69e9cacbe0bcb79ae5afb4a0b130fd6db6ee6cc0d491695fa \\\n    --hash=sha256:9702c4c023227785c779cba2e1d6f7635dbb5b2e0936cdd3a4ecb98d78fd41eb \\\n    --hash=sha256:98e22834650be81d42982360382b43b17f7ba95e0e6993e2a4f5b9ad9283a94d \\\n    --hash=sha256:9e05e33657efb4c6a9d23bd8300101536abd99c85cca82da0bffff8d8764d08a \\\n    --hash=sha256:9ea2a3fed83065d77367775d689401a703d0f697420719ee10c0780bcab594d8 \\\n    --hash=sha256:9fb1eb735fe3d6ed1f89918224e3385fbf6f9e23757cacc35f9c78d3b712dd6e \\\n    --hash=sha256:af029b153d243a80afb6eabe40b0a07f8e35c9adc269c019f364ad747f826a6b \\\n    --hash=sha256:af80196eaa84f033e48444d2e0786ec47d328ba00c71e4299b602235ffef9acb \\\n    --hash=sha256:b0348d8ddb55be2a844c518cd8cc8deeeb8aeba707cf834db5758fc89b476a2c \\\n    --hash=sha256:bab3605795d269067d8ce78a910220262711b753de8913d3deeaedb5dded3bb6 \\\n    --hash=sha256:c2275ff105e508942f99d4e3bc56b6ef5e4b3c0af970386ca56b777608ce95b7 \\\n    --hash=sha256:c95e96c7305c96ede73a7389f46ccd6c659c4da5ef1b2789466baeaed3622b6e \\\n    --hash=sha256:ca748936cd579d3f01928b30a17dc474550b01272d8046e3e1ee593f23620371 \\\n    --hash=sha256:d1cdf0ac28948d225decdefcc45ad7dd91716c29ab56ef32f8e0d50657dffcc7 \\\n    --hash=sha256:d2a4472c231328d4de38d5f1f68fdd6d28a615138f842580a8a321b5845cf779 \\\n    --hash=sha256:d7d4c6ba016ffc0f9568d012f5f1eb77ddd99412aea121e6fa8b4c3b7cbad91f \\\n    --hash=sha256:e52729ffd45b68777c5319560014d6fd251294200625d9d70fd8626516fc49f5 \\\n    --hash=sha256:e574be127bb760f0dad24ff6e217c80213d153058372362ccb9555a10fc5e8d2 \\\n    --hash=sha256:ea3421209bf00c8a5ef2227de496601087d8f638a2363ee09af059bd70976dc1 \\\n    --hash=sha256:ec6e74c4e884104ae006d34110677bfe0098203a3fec2f3faf349f4cb05165e3 \\\n    --hash=sha256:efe6305aeaa0e96b0ccca5ff647a43737d9a092064a3894e46c414db84bc54ac \\\n    --hash=sha256:f3bf75a6dcecab62afde4d1f973f1692be013110cad5338007927db8da73249c \\\n    --hash=sha256:f5a85d7b2b708025af08f060a496dd261055b617d776fc05a1a1cc69e09fe9ff \\\n    --hash=sha256:f5db5ba6188d698ba7abab982ad6973265b74bb40a1efe1821b58c87f73892b9 \\\n    --hash=sha256:f66bd07ba6f84cd4a380b41d1bf3c59ea488b590a2ff96744845163309ee8e2f \\\n    --hash=sha256:fa01f0f6a3050fa6a9771a95d5faccc8e2f5a92b4a2e5440a0fa7264a2398472 \\\n    --hash=sha256:fac4f8ce2ddb40e2e3d0f7ec36d2a1e7f92559a2471e59aec37bd8d9de01fec0 \\\n    --hash=sha256:fda714cf45ba43c9d3bae8f2585c777f64e3f89a2e073b668b32ede412d8f52c \\\n    --hash=sha256:ff4dc42bd321991fbf611c23fc35912d690f731c9914bf3af8f417e64aca0f21\nseaborn==0.13.2 \\\n    --hash=sha256:636f8336facf092165e27924f223d3c62ca560b1f2bb5dff7ab7fad265361987 \\\n    --hash=sha256:93e60a40988f4d65e9f4885df477e2fdaff6b73a9ded434c1ab356dd57eefff7\nsentence-transformers==5.2.2 \\\n    --hash=sha256:280ac54bffb84c110726b4d8848ba7b7c60813b9034547f8aea6e9a345cd1c23 \\\n    --hash=sha256:7033ee0a24bc04c664fd490abf2ef194d387b3a58a97adcc528783ff505159fa\nsentencepiece==0.2.2 \\\n    --hash=sha256:046b15ea22d8042e2e173561d464ec3b64a9c2081324df70ebce7bf7ebb3e497 \\\n    --hash=sha256:0e2aae42960392d6dcb9a72d8e1e65a97294c965071b43c7b3429a42f350250e \\\n    --hash=sha256:1120e0791540615e650b2e9bea835bf38a7362455d8ab62dee7968219c2d79a0 \\\n    --hash=sha256:1402d8ee36f0d851cea8eee4dbb85fea14643b7503cf4d00d102eec0fe3ca719 \\\n    --hash=sha256:1416b92f2f010333786fe6306ed2631121d5ea492219b0841e967b6765e64107 \\\n    --hash=sha256:16c84ddef8d3084a8af37208acd365b08092ca089080f1a71fbfdd911adda9b3 \\\n    --hash=sha256:1edb10e520e4bddf74d85b0f5ae74cc2d60c2b448885080bfb618bc2b3a49f6b \\\n    --hash=sha256:201a8e0f55501a76e08dbf2c54bc45f4642b379271e89c667d517bfbc2191f2a \\\n    --hash=sha256:252908153eeec06c3ca3a32077e64a49d572e3d89881475b4e0f02d99d9fcc7c \\\n    --hash=sha256:38111ed1f79268f399c505028023d5eaaf0ab4e5eafceb709468b0d3323e7838 \\\n    --hash=sha256:3ab3f1ae98970b5590e2209341522718900ba19bcc2c207ffaa6bd417ad960c5 \\\n    --hash=sha256:3d2b5e824b5622038dc7b490897efe05ebbbb9e7350fc142f3ecc8789ef9bdf6 \\\n    --hash=sha256:3ec27c152a1f1b24bc9168b55a5880f3c16e2334e697da6f55a1046a22405a3d \\\n    --hash=sha256:3f5851441ab1ef8634963a5100b733a8bbeefe623e0c5c005b1f1f3880e574cf \\\n    --hash=sha256:3fd9ce2ab4460c713cfdeb4aca693ca6732a11538e05fb332d5af42e3d7fde25 \\\n    --hash=sha256:44284adc6fbe9d5bdd480541431a3d93f674fa44736714d3ad4bcee8283ace7d \\\n    --hash=sha256:443ac618c7a2a1377cf5c82581fbb849591d14e656d5e5a3e4682d4e36a34e4e \\\n    --hash=sha256:46ba07b543add034de0ff47ac5f907e9a06682f91d85121a972764628933be6b \\\n    --hash=sha256:4f0603267cd15b92b68c2c0e852a441507614b70dc7773659baa6b8c214a91fd \\\n    --hash=sha256:524e2a85c028a0d2f9935191fa751e5ef9d9bcc39616f70ab14b28d0369c9936 \\\n    --hash=sha256:54a83df9260a89c1734256e620fe1f1a6bfedd7547139d4dc1384efac11a3a85 \\\n    --hash=sha256:59d6588712101ccfcae9b03692be3aaae1514c2078666d7b05f15ba3a702e41b \\\n    --hash=sha256:63250cfab8b80a1ef82a614eb2b3cadfec2c405f870cedc139d08e2f063eb708 \\\n    --hash=sha256:64b656f025355cf8c51abe9fbe3848540756c6d7ca5e6791b1afa664bc24c7cb \\\n    --hash=sha256:65d84ec36888de4a848eee5f910e67fbc79b064685ef1e10a502e14520ead9c9 \\\n    --hash=sha256:69e9dc8078e128286ed3b975e37c837ba96e215a50c3ef9f3f8b7ab9e5a832a0 \\\n    --hash=sha256:6dd76f3e5c8b2eb8a3a3efee787bbf5b9a66e52a048fe09cab85eca33fec6790 \\\n    --hash=sha256:70d4ca6f4d06df7f0ccab6fe4f49c8a712c8c8b6847b4f0af9a0e1dbb0e0337e \\\n    --hash=sha256:72b7825b331b1b7e7c45be2e674b3e3c65af608fa376bad2d851b20aaf0cdc78 \\\n    --hash=sha256:741b4b367140e9b5c36b5a14c72179f2c946d991ea9a7c031a2a1ee6ad097b99 \\\n    --hash=sha256:74f0ee601047c0c12a783088b51be4e6214a62ecd9e02278c477433cd16e0ed9 \\\n    --hash=sha256:76ff5814db72e7462dece042d7593cdf102b8ec82c2b1cc201a2add34ee3050d \\\n    --hash=sha256:77c3ce990b23441e5ecfa5bce181fd6f408b564aeb6d7e1d1e7de9c5612501c8 \\\n    --hash=sha256:79bac5a251f23a7341e28fda9ce0d5319edf45328239ce037c0682936f137906 \\\n    --hash=sha256:7c6e7bf684dc12145bfa685d3060beaea55139134ba848289bee514ed42e7383 \\\n    --hash=sha256:7fc14c1585139fa6b68775e616a6b90cf622ebf219f9558c0aeaf5d253ee6c9b \\\n    --hash=sha256:89625fb43765cccaa1443b9adb61f283e5fe4cb1536728205d06bada730caa53 \\\n    --hash=sha256:8b2db2056c97224e122054fd794543cde5d24b7cae28424f6e3eb79bbe08e42b \\\n    --hash=sha256:8d44b20234905ff022b7d535f79d1f823ad7670c9851cc4f03cdc34787cdb3ab \\\n    --hash=sha256:8eed98514bffe5ecac37f493f91869c351fbb05629328bfdbc08502c6c094dc0 \\\n    --hash=sha256:8f1f61592e7cabd45d49ce8cc0ef42ca655c091e037153754fb3fa59725b5914 \\\n    --hash=sha256:b23fe17779834d3c27aaf2edac9486d04cca1a7deb8f5facda35150ac6263a91 \\\n    --hash=sha256:bc7b0b1da20f856bfac5f84b2673fe534b167e41980b27442ca8f78c2b7eb77e \\\n    --hash=sha256:c62bd361cec1f5b556eb8210264ecfff37486cd990c3386cc00310f26c54090a \\\n    --hash=sha256:c76c9b3324efd79029eeb0fd2ced1964bdbeca7d45e030b46fa3ef3cf74f8032 \\\n    --hash=sha256:c798f0b327bac10dc95cdac77b9a197ab2bd7dd1e60ebd7586a12d918d4be711 \\\n    --hash=sha256:c8a168b040bc61681293f79a949b5d911c8e25086f4260285b8d97ab5f1195da \\\n    --hash=sha256:caad9566e2ef0e5640d36032c69b0edc7ac6028277b93d93815898804fac450c \\\n    --hash=sha256:cbce24284f51f71d10a42b7b9c964dcb9048b28f1c8e5db40bcbcb6f428cba6a \\\n    --hash=sha256:cd810878180a52950e5a61f25ada5248a453bbdbafe474f89514135fbc1f633d \\\n    --hash=sha256:d254c98ca6387655400b3959c33c83efd807f5edeb608e3aca45800ceaa77151 \\\n    --hash=sha256:d795c4ac689a57f9d4ba2288126ec7901d389ad5827d2f8b8533c883974fe563 \\\n    --hash=sha256:df88b0c34f2fa909d322f7b06b1398e1e81af4b2f42a7b8e3556f928b25d1811 \\\n    --hash=sha256:eb8da9d9a9b418422c21a07fd19b9d9228692b7a7468a45eec6b11642d3c808b \\\n    --hash=sha256:f7c06c751c19d923435a54bff4f7e66e728fad160e8da28254f133abc9725820 \\\n    --hash=sha256:fa9f5ef0e2a82233dd0b8b32ea3f5710e0c44afbc07ed3620219f32601e56090 \\\n    --hash=sha256:fd523c4992041faa5c2b3cde62253d11a96c30d73a34afe48a486e8e2254cd1c\nsetuptools==81.0.0 \\\n    --hash=sha256:487b53915f52501f0a79ccfd0c02c165ffe06631443a886740b91af4b7a5845a \\\n    --hash=sha256:fdd925d5c5d9f62e4b74b30d6dd7828ce236fd6ed998a08d81de62ce5a6310d6\nshellingham==1.5.4 \\\n    --hash=sha256:7ecfff8f2fd72616f7481040475a65b2bf8af90a56c89140852d1120324e8686 \\\n    --hash=sha256:8dbca0739d487e5bd35ab3ca4b36e11c4078f3a234bfce294b0a0291363404de\nsix==1.17.0 \\\n    --hash=sha256:4721f391ed90541fddacab5acf947aa0d3dc7d27b2e1e8eda2be8970586c3274 \\\n    --hash=sha256:ff70335d468e7eb6ec65b95b99d3a2836546063f63acc5171de367e834932a81\nsoupsieve==2.9.2 \\\n    --hash=sha256:4a55d8cf158a9c2e587fa4922f1bbb91d68ac829e2d6f25403a85747c71daf74 \\\n    --hash=sha256:8089a26fd974ca7a1f30276d3d8492ab266ab15af581642dfe8aa162e0c1c823\nsympy==1.14.0 \\\n    --hash=sha256:d3d3fe8df1e5a0b42f0e7bdf50541697dbe7d23746e894990c030e2b05e72517 \\\n    --hash=sha256:e091cc3e99d2141a0ba2847328f5479b05d94a6635cb96148ccb3f34671bd8f5\ntabulate==0.10.0 \\\n    --hash=sha256:e2cfde8f79420f6deeffdeda9aaec3b6bc5abce947655d17ac662b126e48a60d \\\n    --hash=sha256:f0b0622e567335c8fabaaa659f1b33bcb6ddfe2e496071b743aa113f8774f2d3\nthreadpoolctl==3.6.0 \\\n    --hash=sha256:43a0b8fd5a2928500110039e43a5eed8480b918967083ea48dc3ab9f13c4a7fb \\\n    --hash=sha256:8ab8b4aa3491d812b623328249fab5302a68d2d71745c8a4c719a2fcaba9f44e\ntokenizers==0.22.2 \\\n    --hash=sha256:143b999bdc46d10febb15cbffb4207ddd1f410e2c755857b5a0797961bbdc113 \\\n    --hash=sha256:1a62ba2c5faa2dd175aaeed7b15abf18d20266189fb3406c5d0550dd34dd5f37 \\\n    --hash=sha256:1c774b1276f71e1ef716e5486f21e76333464f47bece56bbd554485982a9e03e \\\n    --hash=sha256:1e418a55456beedca4621dbab65a318981467a2b188e982a23e117f115ce5001 \\\n    --hash=sha256:1e50f8554d504f617d9e9d6e4c2c2884a12b388a97c5c77f0bc6cf4cd032feee \\\n    --hash=sha256:2249487018adec45d6e3554c71d46eb39fa8ea67156c640f7513eb26f318cec7 \\\n    --hash=sha256:25b85325d0815e86e0bac263506dd114578953b7b53d7de09a6485e4a160a7dd \\\n    --hash=sha256:29c30b83d8dcd061078b05ae0cb94d3c710555fbb44861139f9f83dcca3dc3e4 \\\n    --hash=sha256:319f659ee992222f04e58f84cbf407cfa66a65fe3a8de44e8ad2bc53e7d99012 \\\n    --hash=sha256:369cc9fc8cc10cb24143873a0d95438bb8ee257bb80c71989e3ee290e8d72c67 \\\n    --hash=sha256:37ae80a28c1d3265bb1f22464c856bd23c02a05bb211e56d0c5301a435be6c1a \\\n    --hash=sha256:38337540fbbddff8e999d59970f3c6f35a82de10053206a7562f1ea02d046fa5 \\\n    --hash=sha256:473b83b915e547aa366d1eee11806deaf419e17be16310ac0a14077f1e28f917 \\\n    --hash=sha256:544dd704ae7238755d790de45ba8da072e9af3eea688f698b137915ae959281c \\\n    --hash=sha256:64d94e84f6660764e64e7e0b22baa72f6cd942279fdbb21d46abd70d179f0195 \\\n    --hash=sha256:753d47ebd4542742ef9261d9da92cd545b2cacbb48349a1225466745bb866ec4 \\\n    --hash=sha256:791135ee325f2336f498590eb2f11dc5c295232f288e75c99a36c5dbce63088a \\\n    --hash=sha256:9ce725d22864a1e965217204946f830c37876eee3b2ba6fc6255e8e903d5fcbc \\\n    --hash=sha256:a6bf3f88c554a2b653af81f3204491c818ae2ac6fbc09e76ef4773351292bc92 \\\n    --hash=sha256:bfb88f22a209ff7b40a576d5324bf8286b519d7358663db21d6246fb17eea2d5 \\\n    --hash=sha256:c9ea31edff2968b44a88f97d784c2f16dc0729b8b143ed004699ebca91f05c48 \\\n    --hash=sha256:df6c4265b289083bf710dff49bc51ef252f9d5be33a45ee2bed151114a56207b \\\n    --hash=sha256:e10bf9113d209be7cd046d40fbabbaf3278ff6d18eb4da4c500443185dc1896c \\\n    --hash=sha256:f01a9c019878532f98927d2bacb79bbb404b43d3437455522a00a30718cdedb5\ntqdm==4.67.1 \\\n    --hash=sha256:26445eca388f82e72884e0d580d5464cd801a3ea01e63e5601bdff9ba6a48de2 \\\n    --hash=sha256:f8aef9c52c08c13a65f30ea34f4e5aac3fd1a34959879d7e59e63027286627f2\ntransformers==5.2.0 \\\n    --hash=sha256:0088b8b46ccc9eff1a1dca72b5d618a5ee3b1befc3e418c9512b35dea9f9a650 \\\n    --hash=sha256:9ecaf243dc45bee11a7d93f8caf03746accc0cb069181bbf4ad8566c53e854b4\ntrl==0.24.0 \\\n    --hash=sha256:a9145b7d4a4a33778de117bda48530f0cf5b2ac25acc07db80ad04836f490dfc \\\n    --hash=sha256:eee495223725d3da0596be2607581969db89ba0f7c00b075802addc31e61eac9\ntypeguard==4.6.0 \\\n    --hash=sha256:79878165bb86f2cf5d41d159a0ff1792a796cf496882d2fe1b1c6c7049b9cdd7 \\\n    --hash=sha256:e7414f09111317de3e335de92cd397c5c0ca00b1cc1676de12e1d444a79b3f21\ntyper==0.27.1 \\\n    --hash=sha256:53150287edd11baeb4e4722c8e394fcdf8181c0ae89485cba8d25c778d5edd56 \\\n    --hash=sha256:a79bef8469a79c45498e7b814ecf8d603cc7644e9acbd9e19cac0334240b18df\ntyper-slim==0.24.0 \\\n    --hash=sha256:d5d7ee1ee2834d5020c7c616ed5e0d0f29b9a4b1dd283bdebae198ec09778d0e \\\n    --hash=sha256:f0ed36127183f52ae6ced2ecb2521789995992c521a46083bfcdbb652d22ad34\ntyping-extensions==4.16.0 \\\n    --hash=sha256:481caa481374e813c1b176ada14e97f1f67a4539ce9cfeb3f350d78d6370c2e8 \\\n    --hash=sha256:dc983d19a509c94dba722ee6abd33940f7c05a89e243c47e907eb4db6f1a43e5\ntyping-inspection==0.4.4 \\\n    --hash=sha256:547274fa6b0a561ccf549cc9524b999a578e737d015d8709d021f9d0d13bea47 \\\n    --hash=sha256:65b8397ba37ccbce054456aaccddfc91e6e3083c92824df348d96ca832f3f147\ntyro==1.0.16 \\\n    --hash=sha256:2f9087196113699647190459a12b6b4f2784dd0703c50ab1f7bf0f85a3a794ea \\\n    --hash=sha256:edf23c9c3eecba5a2b5af152df8f34a723b689a001ff5b7311f93eff1893a636\ntzdata==2026.3 \\\n    --hash=sha256:4a1518b8993086a7982523e071643f3c0e5f213e75b21318e78bcabfff9d1415 \\\n    --hash=sha256:dc096730c87af6cab1b171c9d532be840741ff5d459015e7f6947bd7d7e54931\nunsloth==2026.3.4 \\\n    --hash=sha256:7b260d4a66fe492e63509ccfbd279c1f8824e26426433e38a603f119bc0c9054 \\\n    --hash=sha256:f5d40405e61d23018f61c1fd44c19b28b0a4a3c95c84c594c4d3313735db2c4d\nunsloth-zoo==2026.3.4 \\\n    --hash=sha256:73dd4f7fedc3e8c3187db303458e16b67365c9c618623359eb79eff8dd1e5f69 \\\n    --hash=sha256:db8c3c515e5c2c6e575ab53bdf17e683e264df6ce5f9048f76a5cbcccb663f51\nurllib3==2.7.0 \\\n    --hash=sha256:231e0ec3b63ceb14667c67be60f2f2c40a518cb38b03af60abc813da26505f4c \\\n    --hash=sha256:9fb4c81ebbb1ce9531cce37674bbc6f1360472bc18ca9a553ede278ef7276897\nwheel==0.48.0 \\\n    --hash=sha256:3217dcc807155e45db462d7ef2431f5ddda0d7273b700d05a67b271ceb1287ab \\\n    --hash=sha256:94800765601e9171bf5d58d066e640662842bcedcbab982b2c90787a2c987322\nxxhash==4.0.1 \\\n    --hash=sha256:0163b5d259de23ae9e07b7eabf435ce4704f6f205589a2b154e6af4be985ce1b \\\n    --hash=sha256:03600a8987849b2bef7be795a60a6052b635c63fa98b718b08ca5ee823691cfc \\\n    --hash=sha256:04f9a24de11a6647666d5302fd73d6a5224ce50ddc965fb0bb44cee736e6bd7c \\\n    --hash=sha256:06713a5aaf1d0905c5579416c020c02e42b3ceb931e86c7d3b7fb85403dee3f3 \\\n    --hash=sha256:06d7fbd609503c3be5e65cdb6bb2f040d6a98574404e2e1d5c60815c97fff4aa \\\n    --hash=sha256:0718ad66f4ded2411f8e62bdba549ee71e313a2d26ef5060ca3fdbf29897dd3c \\\n    --hash=sha256:08ed8da18cd4fd0a6a5d6a444852d8fbd0e565388a74a4937085451b5f1a312a \\\n    --hash=sha256:09f9feb118966cc6650e1806205d577eae7ca394aa6acf349a0b62a94bbeb329 \\\n    --hash=sha256:0ab851b45c70d4992be7cdeeee16f97a0b677408c758c4b1efb1cfe8030bfd37 \\\n    --hash=sha256:0b1082fd0f089ce9098ed77aad8b777b5d156f8ac601c69cab73811822b8ef07 \\\n    --hash=sha256:0b20a06454b34f1531fc677c54efe2ecdec691ef9224f7fa919bf2c1363f7ff1 \\\n    --hash=sha256:0b42a5a26607e4b2409fea174773a66f2dff9dfdbf2c1a851bb7b804e2c97535 \\\n    --hash=sha256:101aa300de6ceef3d9c77569706330d8921fc45dd82bceed2084f1e9f2557a24 \\\n    --hash=sha256:1216f7ba5683f17a89eb7dcb4bc50a0b743dfe1902278d7b3d0786f538118433 \\\n    --hash=sha256:1642907941ee4b75aacc3db688af52ea02ca2305ab22af7ee686ed726b332684 \\\n    --hash=sha256:168dd6b51725a222abc722832e56624d15a63fc2e8249021509c93f1063913f6 \\\n    --hash=sha256:1749f0688020209fe0d357ce1e1cd9ec9c6161ed0405ea949d24581c4c43fa91 \\\n    --hash=sha256:1b3cccf75eeb5b01639b2feadb042a8e07889293b7ca72fa2985e7dcb64763cf \\\n    --hash=sha256:1b50223d92df94d54e1a31469335a2c74b16692e6c1cb726f1e6949514458706 \\\n    --hash=sha256:1bc591533fc975614f7e13594daee76af96b8e1fbcf8de76c8773858fa9e7cea \\\n    --hash=sha256:1c2200b98a805351cb3142ae4e1fdcc9e91b5e20f5d30d4862b0b96f92558f4e \\\n    --hash=sha256:1c7c642a0f79c3e3cf2965475507574d3d1a50ec71060039d60cb87358667cb2 \\\n    --hash=sha256:1ee523f51718e41753f04f7102bb4dc55a18d2ea5cbaceef8ec7ca08571bd428 \\\n    --hash=sha256:1f3346c5c287ac3c7f38b20380f55e8768230e7252af59fabcf3b87ab21e4256 \\\n    --hash=sha256:2194bf96d5f3d4e0cb65deba370ec83dda3edfba42155f9384190ed5e51ea5e2 \\\n    --hash=sha256:237b8f63a2a0fcfb1ffc06e21dad23add44e6d354b2b014364a1d41e419a4dee \\\n    --hash=sha256:23a4376b4a3183cb50d4d2a3179f887a7773cc695eb2c908e551bec3221b8c60 \\\n    --hash=sha256:247ece770647c0aef080561fa996f9774b4dadce2d0c42eeb98229db7dcf820d \\\n    --hash=sha256:2696bbac613f6880fed60316c298bf3091d4f8eee3ae2e9466f70bb76204fb0c \\\n    --hash=sha256:26fe6238c2d5b11ed5063b9bf4eb290624b004fd074688da6bb079bd564f10d7 \\\n    --hash=sha256:2d52dc7c33c1b83082b707f6b7814dc76d2faaa2ea62bd9c5fab4b36f83c087f \\\n    --hash=sha256:2df3ca8757dc381e75e90a4d7995a6324f58a923c7145220a7b2c0231f66fddc \\\n    --hash=sha256:303121aab4b7f898058582d7962ea79d9e26e2379d7b6d8743f70f2671674481 \\\n    --hash=sha256:3088dadbffa33c29e0518578430a7dff2e901a212e487aefa5faaa0dc06dad34 \\\n    --hash=sha256:31d86f9e81f3e84e00131ac7c54caf5119ae4ddd82c09c31cff597c813ce1ee2 \\\n    --hash=sha256:3358097d333d40657569ec1121e21043dd7d0efa10aead1b50e8b4fa83077d7b \\\n    --hash=sha256:33e270d302c95ec426dfa0f5a4e16bff2ab8d7b8a46faa4746affb05e684ac77 \\\n    --hash=sha256:33fd538191f47071deef6b1f676535e2aa770f1fd150ae4cc75a34c9e930be3d \\\n    --hash=sha256:348c8f288dc961d6bbd1985c8152a3ed7a85c95df00e82320f0c5215d922a399 \\\n    --hash=sha256:349775ac30372b344d2338b2a168c0a1312a644194da25b8bec476d55761a128 \\\n    --hash=sha256:34ed93e20bfd98d722b902121643791eeb4b1641871e2dc63d0d4c2d93f187df \\\n    --hash=sha256:37f667dee0f867c42894b34e2a6fe26bf195c0ea4683d9d2b713db023f242c3a \\\n    --hash=sha256:3891efe3d7a531ce6da0a4a50a99dd41c75b8fd4ca19d73c86431b4db5c305f0 \\\n    --hash=sha256:38c3d22129a6958846a3098d68bc8e661704461c0be4793ae28836e4690c8478 \\\n    --hash=sha256:3c2445edafc300cc40feb6a25a8356a971c30cd0bf47b5349c2ad74c508343b1 \\\n    --hash=sha256:3f68fe400ceec235f3e4a4b02a28c2fd2d283584a193223c921dd4c48f1d0754 \\\n    --hash=sha256:3fb1d30d4b6d6e2c4a08e5ac6fffdb2b572d2cfcca15a5509cf4e7a1350f955c \\\n    --hash=sha256:41e579025a6e13a99e6d71e39c9cfc621a0dcdbbf19106325e145fa858f2d794 \\\n    --hash=sha256:421b94f3ba7067958d02e38960d987756347aa150df06df11aa68ae1af78c619 \\\n    --hash=sha256:427b62d62d4f967fbb10b82a3813e4875c2a6e7e7634739f17265b650c7f65a6 \\\n    --hash=sha256:436e11b4dd966afe5f7f665e4cc4c5485ffe3ceb42f25a22e1701d236abf1853 \\\n    --hash=sha256:43bcf2a871f28f16135545415cab3ec43904d4c80425a64598a9e6cebfb2b5ba \\\n    --hash=sha256:43e5f9169e73d0f0db33b5f6b8554bcce69ac278c966daf83d5eb4eb2f13829f \\\n    --hash=sha256:440c401e146ce64bdb3beb8ff0c84677b6f21307c28a34779071cecee5d4d70c \\\n    --hash=sha256:44ab12e8cd17d4f001769f00ad465208b4bcb897ed29e65f058f74466b57a98f \\\n    --hash=sha256:4528cf80ebbbf57d40edfb31521ae265daa6dd636d615b1cf0ac86209579e59d \\\n    --hash=sha256:45e88111ebe331de478ef8d4293efbe88f3cf8b863386c9a2357136b838e1af0 \\\n    --hash=sha256:4741d42d59e4e5fa1a86c17ab9c27dc8ea459c700d91b6742fdb9138d9a516cb \\\n    --hash=sha256:4751f1d7eecae6b2d2a773630f1a7248f125c9a92a456694d03c15bceffc9d68 \\\n    --hash=sha256:488ca5c5e28ef56ec4bbb12f835b3f1cbecc5f3510062e70117bc6594851932a \\\n    --hash=sha256:4972332c079d6aad69c4620a68d015a4ecb33141583f70d642cf9edf6a713763 \\\n    --hash=sha256:4a252fb862b0ae2590587e625f47a0e03da05cf0205e8830b67b6596c06038b1 \\\n    --hash=sha256:4a76345f5aceb4ec404918edf9c7f2b5507db864dc0d7455982009ac0890b57b \\\n    --hash=sha256:4af350bc3f329970c0e3a59af84a8a30998bf8a9167eb50cd48e59baaa1d7bec \\\n    --hash=sha256:4bbf3ff651e0f1a19beb5d0f48e0874a9bad2482a588c9d214c96ef1fff1cd9c \\\n    --hash=sha256:4e5141543c7f7fe3087500bbb4ac2845cb528a980aa91f8f1e661e2292ff4a5d \\\n    --hash=sha256:4f5e5c6df4b703afcbe9352d238a51efd97c3b91fdc3a2052e40fdacb1e7505f \\\n    --hash=sha256:515a822c73abbf6a0b7c70976d9662be342835c9d78b8dc7c023411f39c35dbc \\\n    --hash=sha256:554f87034635bcec47c5d72447bf3db7e02da1bf493a0ada010db28a76f891c6 \\\n    --hash=sha256:567cbc630302a46a8ecfd943b309ccf5372bb3718f1f3762d452df30f033bcf0 \\\n    --hash=sha256:57d7fa8f23908d173001c21a9e82bfc6ad997d1b6c270fb121812b7ed158891c \\\n    --hash=sha256:5adf927dca8c47fde7e683fe69efdd81bc865c4db1fb6bb00b391e2b6185207b \\\n    --hash=sha256:5b7875ac1a2edcb691f27642b8b94b904baa6bcecb7d79c72df2228ba8cb5c51 \\\n    --hash=sha256:5b7979f71d06ae45a769de0699900a246d8cb632db1e8bfdc79ec019063a503c \\\n    --hash=sha256:5c2d525a3afabcd8e3549d85fc7e111fde6bc302d06a1893fe73adb79823415e \\\n    --hash=sha256:5dc434c946012e6d8a72b10f970ea30755b718251dd7591dbfdabafd3bcb21bc \\\n    --hash=sha256:5f1ea31d61bcd2cd2f3ec4ca80a64187bbd7948f490b63cf0dcbc6e717b4c1e9 \\\n    --hash=sha256:62198213fc3e0c56e567894b318ba45834e007d065f84ba6dc9165d21546fc56 \\\n    --hash=sha256:63aa52659bc32bb9bd7cb5caf523b4d14429a477762cfac886132d687c1f80fc \\\n    --hash=sha256:649f2682c090cca1ac4037866381f3652eaacbd56e5178030f4ce1325b8f945b \\\n    --hash=sha256:67e57b834e07ed973cee7b6da1548ff28a56458d77696fd2a5f397f340694848 \\\n    --hash=sha256:684160b3c0a9b62c6f0de90f44e11dc5d8643dcfa18a5856b45fb1c47478bb71 \\\n    --hash=sha256:6a8c5ce76b94ba49f3be8a8f2611abc6564210702c72ac9e237ca2bebfd17794 \\\n    --hash=sha256:6a9f98af872355e0c02439e48583958eee00e60b928bb20476460d9d40cb7b4e \\\n    --hash=sha256:6c45258a37fc22721395c09927cb982d3e7a83607cab15be7e2416501bd3a330 \\\n    --hash=sha256:6cbf4e21ef0890804b5bb9ad25c48f9c127758d7f6c66bef374efcacc63c738a \\\n    --hash=sha256:6cf633df84d80a1668fcf61e330791dae46825e395549e7d34f376411e75088a \\\n    --hash=sha256:6efb8f21cc136c79b3e5bb747c8682d37916fb202cdbbc32182de5c4e47f821f \\\n    --hash=sha256:70129ebb8f20e1ac1da58b78ed381624bd689a43a9a7366560bd8fabea145105 \\\n    --hash=sha256:704381264b36a18b9c62ecbabe2e71d0fc58c77c129c15355c989b10bf05b6b0 \\\n    --hash=sha256:7236be540d6be9ce448d98b940dd26ddf70ca41012e8a14a53fd9354cefe4e8d \\\n    --hash=sha256:72f34834518157a75e7090f328ee7a16c70c804cfc7c694fa069cc888e9fc03e \\\n    --hash=sha256:74379a577a9f3b6afbdedf1b90e5c7764467051977f18a326d7d607336d743bd \\\n    --hash=sha256:74a164e8b63f1e9cf35c9a7809d082b033d1a00e7375d5d814415436e7867e57 \\\n    --hash=sha256:760de77279e9cf9c81d012ce0705cba13afccee9b09c480f17d778c8c5cefae8 \\\n    --hash=sha256:764b32d52d15b8b95ac8160e540772fa1adeb611fe40bffaeb42e7bf98279e44 \\\n    --hash=sha256:79a3203aadf39637869dfea1185227d8452844d78b837e54fb1117b4d34ba5c3 \\\n    --hash=sha256:7c343ee174d417a44d0c3355602c0cbbfa52a04d1bbbf1723378c7d2c8f60626 \\\n    --hash=sha256:7e27dbed5c4ba033919e4b4ed8dc14e029e91d14a93cd9f920d25277c7df6781 \\\n    --hash=sha256:81507a68ba84c55241fb61cce1469f473a5da4205fc8ef6f698e5948eea8dd88 \\\n    --hash=sha256:81664268dba92e037b740ecf37fa02f1cab4a391f93f28e35792b3341c60648f \\\n    --hash=sha256:839f58c5bd9989875be0fd28446dbf32cace2c2cd8bf2f6762acdc38a95cd1aa \\\n    --hash=sha256:83b8c2013edb5dc1f9e7268b6496130705bc48d79c86bb8817b3d210b81a5513 \\\n    --hash=sha256:84df5f8da574caadbc0cb1b8866ecc2368cc941f0cd05f677756c802f370dafa \\\n    --hash=sha256:8580aab306888224074c7edeec734de0c3c5ccde65b2da4e6c9a5e28f7c0a1bd \\\n    --hash=sha256:85bdd40cb505a11e0ca04191711266c5fd696ed786ae83849955e457774edc96 \\\n    --hash=sha256:85e402dab0f9acd3604539747c6fcc57dc188a18af6ab07eb8189351cd32466c \\\n    --hash=sha256:863f3d3b44110f7243e86cf994aa5c5d88f2348b6e84ab4402fadadfbf9f7da7 \\\n    --hash=sha256:86b2b12bec60c678ed8f5cca0258ad93a8928ebddb6ca7732f0875afe1451d1a \\\n    --hash=sha256:87aa309a93bd5ec13f14309a305ff4e9bf74c5363fc46c264c0a22edfd5b0670 \\\n    --hash=sha256:87cbdec1a7dd930079671a60b249f3ca4e773e6fbd0676e21e36fdc9dd0f3b00 \\\n    --hash=sha256:87da13df72c5612771cd905a8b121e0bfea62d7659b1c92198736eb722220e83 \\\n    --hash=sha256:88d87719fe6bddf117238b341c5db851f8e96ba68ad9832b450e4a43dc60b37f \\\n    --hash=sha256:8b4477edc03091f51f5309406d230851c23cf4822029e3bf40b8df53093fff1c \\\n    --hash=sha256:8b99ebaf9e816ac5069423b1367ee7e8078fbcebcf62545506bb0608d2f4f468 \\\n    --hash=sha256:8ba782ca3bf1e81492611152b9a0d5264971339e95e34d69de0ac2c926be496d \\\n    --hash=sha256:8bcba9456242ebf180a04d9443812fd85ffe6bd12bda464dd116fcece8886ff3 \\\n    --hash=sha256:8c9fe122444e129881afd1d4d1c7ac0d3ce2d91b68c2b40173b6025ff1c31f9a \\\n    --hash=sha256:8ec4777d92fd61a5c8fdeddab894fd65bea301a8092fb5419ec6472aa4d458d7 \\\n    --hash=sha256:90cb2a1c9cc503a054a19612b48ff6e8e47805f618bdb3224a07568aad03a37e \\\n    --hash=sha256:9283d9dd6b44acad35118e2976fc763a065509e4118debdb61916ec322ed17b9 \\\n    --hash=sha256:94ac8a6b8c47951173f0b67bf862bcb971bf24e493b9fbbdb0e010cbbc7d9f54 \\\n    --hash=sha256:96d8de55029d42251945531f6aa7590c32b48163c66a43bf29d8657d7446a377 \\\n    --hash=sha256:96dedccfb09a73a25751053a183159b88f4ee75f388df8166040c152ac0531c6 \\\n    --hash=sha256:9761ff4a0ffa583fe850731ad24fe82c88cccb7a2294727db0955f3279a4cb3f \\\n    --hash=sha256:97b455de3e8b1b0b1e4594cb61a468992563f03ca264062fbb0a66b393c01d90 \\\n    --hash=sha256:97b94fb29abf21f5f0bde15f7dbdd3a4aa2dc59f37026adc7b4bee8563b84375 \\\n    --hash=sha256:99054b838b74d8d3995ea0d410976ae967c46207ae22d6ddfc535e809197dab9 \\\n    --hash=sha256:99166cc98637e8bf550cda2aab07f4f1d5f899c45fbd721801aeabcc9d404824 \\\n    --hash=sha256:9a51b061d54cda8b83e62c44458bfbf0dabbef9b975dd9649952ba5076b9f349 \\\n    --hash=sha256:9b1dddc257279417d93c9e59420d49ef90aece90d7a01996db3aade74b0281b1 \\\n    --hash=sha256:9c3c4b9aa9a27196b921197f7daf9e6c1412739df06a99cfa6e923879362eff6 \\\n    --hash=sha256:a14578102a6081465aec9cf73c76c3cd3f79f0709bdb3b8ae7ab0b54c9d8b089 \\\n    --hash=sha256:a16a3fa6936e36bb1414d16a6bd012c9033e5161b68b426805b61d895392437d \\\n    --hash=sha256:a33de7633c948ab2dc144af370a66e7e7af29b425dcd0f7e4f59689fb9391b53 \\\n    --hash=sha256:a43418e1a90b4809a9caf64aeb8b0696e3e1f300a323acc1e6ee2f93ae319fcf \\\n    --hash=sha256:a4553d36cc0b7fce1f35ba8a94dfd775aa3ed12f5eab2dc3b46ac75a0706b0bb \\\n    --hash=sha256:a5b21b42a01a343096a1c018d35e9b7aec9c7065dda53ae8da071e37478b2cea \\\n    --hash=sha256:a65785e653573fcd1e33062760ab4c3c3440e8e910765018e4b6ed4ad07b54a0 \\\n    --hash=sha256:a6671a8f6ea4f2101ce11fab5023a2e59391cff249fc3928cecb69d971525fd5 \\\n    --hash=sha256:a69e8946e4902ea11fc1c557740cdbfe7d75c78fcc5e4324ff89a696a634357d \\\n    --hash=sha256:a6e3653df1a70b8ac4191216324242e4be2bca18c9a7c10934e1bd56dc7ca15e \\\n    --hash=sha256:a865d2d470220e659220fdb59d5b6c4422802d8d6098e1324bc4d12444798914 \\\n    --hash=sha256:a949b072ea59c6eca0811ccd9e95133cc50d2afda8d464b5b077c78f78efa269 \\\n    --hash=sha256:aa6ccc7f31018484d652cf52db020003433f3c9fa83189c028bd807d2adde503 \\\n    --hash=sha256:ac0f291ab6485bd71f33941f9b92771318332a05d505460b41e893a549caadc0 \\\n    --hash=sha256:acb31ecdd1a97fab5cd39a84ee9f515e727d319f796fec48703b8339b9998360 \\\n    --hash=sha256:acf52474b2494ef66dc7e0fb6d5e2b50c18313039ad4d275fbf9f9907c804bc5 \\\n    --hash=sha256:ad889d58361a26ba75f5d6a1a0da08ed4950ec4ac8a6da86e1c5ce1b95ccb43f \\\n    --hash=sha256:adbd48b30e3f82c89fb2b3e6a87cdd28d113b190a5ed0ee2dee286323ee9a621 \\\n    --hash=sha256:af05a3f650220a6c59fa0ad2410249f2d2470a05225807c378fb67458693f8df \\\n    --hash=sha256:b3662719007e059abde7eddacf8517142ba076ddc7b30c807260e57d28c3c191 \\\n    --hash=sha256:b3bece52127ac20044311ee73567f9f0893b5de64f9028aecc90cc740cfd525a \\\n    --hash=sha256:b4c8842fb19d78b5e8c2a52baf4c8357658cc56c62bc822b86ce0f942f28e286 \\\n    --hash=sha256:b659fad79c99b0238c7ad7e9d7dbf4eebfea9097c2dba65fa0a4d18a25b29a2f \\\n    --hash=sha256:b6c1f9c59bbe593f88a0aad30be4150f15bd57bd64efb95feeabcb8e563f1ecd \\\n    --hash=sha256:bdd16718b63aa3ebd68aabb79021a40e47c81374852d41a306b9453141bbcbee \\\n    --hash=sha256:bf430c587f447a554c53768ad76b9846fe7c5632180ef6f69c4fce8b0552fbd0 \\\n    --hash=sha256:bfed61996d618eb90d6eaae0178002e3466a28b06bfc557a7a3a7266378d8c5a \\\n    --hash=sha256:c09ada495567c9c9a8156c5ebcfb93be7fece0755062d738c972dcbecd0d84b5 \\\n    --hash=sha256:c0e6ccc2b19ec8a726b2e26062ac71ea63e15500d6bf85910e42481844fdffc1 \\\n    --hash=sha256:c101180495cb4ba3617b279a944345c53a5e73b0c150053d1fa8d8af32de9579 \\\n    --hash=sha256:c10b9206753b64aa791b35b201485477525b26fdec5bf86e8364c388a03e2592 \\\n    --hash=sha256:c3074db513c81f764053e3da079312ecf85a50d8350c71f4cc0105d9662a9e6c \\\n    --hash=sha256:c30dd1af66a820820398b26e0d74e7a9aa43cae705924f23ed828cd8e5c26c3d \\\n    --hash=sha256:c57963970d359a72262f7fe6be88f945e2334d4bc41462b7f08c37b0abf35ca6 \\\n    --hash=sha256:c6301d92545c591ad31c3e050aa40a5f8a4c16413f1f9e6f9322c6f0f9d2b736 \\\n    --hash=sha256:c6370189e8e66b7e608f533b939a9de092ddca6cce084ca0d3d414d2ed5b5d59 \\\n    --hash=sha256:c6fc415b5568bd9accc7187f1729a99707330c0a67a8b9f93c1149ed573ed75d \\\n    --hash=sha256:c7484fea54964edd417cc3a104d5180562514aa7c4e2a2bc26d776ef0c4cb4a1 \\\n    --hash=sha256:cba763d84b06bda2c38d5185dee76f1b9dfdc0789e96e476d9e10005526d0788 \\\n    --hash=sha256:cd878d32f5c6cbce9783f8d6897561fb772211edba9dde49d85672b88ed45276 \\\n    --hash=sha256:ce6d5cc94a50291d080259a126cbf1e9ba4ac861e6429d2f3cdbb1474f51945d \\\n    --hash=sha256:d0d24a4f3fb63852cd09af46ae4b7a4d00cc8b8615a046dca543786e728d1056 \\\n    --hash=sha256:d1e0d1ea6e44f51808a9e8469c8afdebcdf6fa23d1ea524a0303d57d23919712 \\\n    --hash=sha256:d54b8ae068af532c8cdf56abb9e09a60fbe7b10792444c9c27987bb6d3b450fa \\\n    --hash=sha256:d55bf4ef10eb09b8b6866790e083d26d087d84caa3cc0946ba87c3ca7ecaf7b7 \\\n    --hash=sha256:d9f3848ffaf010bdbabdbf4c25641fa258b6227ff27bc74a4d06edef521a4873 \\\n    --hash=sha256:da0264844a09b538c894e5eff25313d941deb4dedec2131b98418a71a3c9944e \\\n    --hash=sha256:da544672efd9ad76077928a3e6c5d894e52ce82d3bf14002db4a1bf17d1a36a2 \\\n    --hash=sha256:daade8936c4deaaf7b01561324ce438ba4f885d717e9adc62b4d67212ad7d7bd \\\n    --hash=sha256:dd649663ddeafbfd4734eb8abae921dd5baa1242f20bda54e8bc927369ccded4 \\\n    --hash=sha256:deca2a30d983d240b8375ec2ee0a4288e72042827fc61df2f7671f8467e4cb2f \\\n    --hash=sha256:e259bb7e1e2d8de6b35f430f5c7220b1c0ebf3962d1ba7ec7545980d5931edb8 \\\n    --hash=sha256:e3996ff9b6f99180357024336bf5749a8ad6476a9a2523e535c5212b995b12a2 \\\n    --hash=sha256:e3eba72f9bb84fe696516f4cbca68d3d74a376157e68bacddbb7f2516af61523 \\\n    --hash=sha256:e4296fcc790876a8b0f297edc83d3b088457b774d8f67b4636807f8a2ec69a79 \\\n    --hash=sha256:e53926e76131a74e79cc0b39fa712c227875f180afc68646bd1e1d8a17e60313 \\\n    --hash=sha256:e681a6fc7e4f715252b9b5acfb30536ec7dd1f75033a32dc617e6fa95af1a3fd \\\n    --hash=sha256:e71b34978e77868cbf2d18c5206a4603f9c644dd7181bec5643bd40141d3b8c5 \\\n    --hash=sha256:e8cda075b10bb3917b002c74a04f9e02b7d13b5bf732571404d51c52b11c7329 \\\n    --hash=sha256:e90b4bcf1d9eb1010fdaee7c9209fb667e74c0684f3ba17f9032bd7319da90c9 \\\n    --hash=sha256:e961093277ff9d42addb9dad5614dfb7800ccba07c245c39c8e9b4daa35d160c \\\n    --hash=sha256:e9701c073bd062fb6bf6be51b47186ad15f1e87feedf4ea07198e0333ec068dc \\\n    --hash=sha256:e998cb3685b92101ec5de0fb4d9485cf01e50bc418211955c55d98064664cf4c \\\n    --hash=sha256:ea5ecf800b45bdb34afe05a1d0dae1f8ea02a290e50636dccd399063f6b180f8 \\\n    --hash=sha256:ec1a470c6db94ac4589c203921e89ac1bc13e796a8b1784d8135e1893559cd3b \\\n    --hash=sha256:edccc2ec58435a580f96a48a3ccae8cd0a480824119165dd90108718ad81ae6e \\\n    --hash=sha256:f00330ac7e24769e2032203f2b01794d670916b0c1799fd261340f1af9499875 \\\n    --hash=sha256:f09ee747e2a5f876cc5ad56947734811828335e13b403dd8ea1e06d77a9dd48d \\\n    --hash=sha256:f18732adcc271741bd651c3e56fa519d8a237d2cccda01fe3afb226bf87f783b \\\n    --hash=sha256:f1b603d0686c99fa0879f104a74e7db58367634c6e50ba827bee9aa095e23205 \\\n    --hash=sha256:f33cf0baa91eccd2cb7b62bf00f10c2264ef578b71dd33a12962e71a36eb4d32 \\\n    --hash=sha256:f3e1a44af01b6692de0ec6caba5f0bf93ceb36896e02b7fc00952c6ea7ef39e1 \\\n    --hash=sha256:f484ed57bb3e4142f9d6439568658c38be5f94b702ba00a1ff32c69783b6c66d \\\n    --hash=sha256:f5d031f35962e5483a613214e61f09fe24ab523062c3646d592dc16c4a217451 \\\n    --hash=sha256:f6247f5e23ee94f2557ac9dab738a336f607c6ff476fcf66ca70c3aef5eee15a \\\n    --hash=sha256:f7db035447a0ac8959aa230c5d36545ecf9f547413eb1711c0ca6f0ba1418925 \\\n    --hash=sha256:f83295394d34e1287e5b30fcc496c13b92cf886a131f3dae5444e38da8757efb \\\n    --hash=sha256:fac4832b638000106207bc44e44b9616a6a416aaee56c62b01d61f3705e49f58 \\\n    --hash=sha256:fb59a0dd61fb2ad481c03fda399d78ce57dab6bb62c2c8fdb446a7ba4754b89a \\\n    --hash=sha256:fc737c05ca2d48e5dcdbbb249314df3fc6c2a0be6da8b0aa28e13d72afaad7cd \\\n    --hash=sha256:ff48915bf1871a1f19f74c11834c6329443d306cedc0c05fe7fe617810422a80 \\\n    --hash=sha256:ffa44b4c7c5d0ffa31356b4428659516c0e47647825c74079a296b3857b6d99d\nyarl==1.24.5 \\\n    --hash=sha256:0055afc45e864b92729ac7600e2d102c17bef060647e74bca75fa84d66b9ff36 \\\n    --hash=sha256:0465ec8cedc2349b97a6b595ace64084a50c6e839eca40aa0626f38b8350e331 \\\n    --hash=sha256:0ebfaffe1a16cb72141c8e09f18cc76856dbe58639f393a4f2b26e474b96b871 \\\n    --hash=sha256:16a2f5010280020e90f5330257e6944bc33e73593b136cc5a241e6c1dc292498 \\\n    --hash=sha256:17f57620f5475b3c69109376cc87e42a7af5db13c9398e4292772a706ff10780 \\\n    --hash=sha256:2120b96872df4a117cde97d270bac96aea7cc52205d305cf4611df694a487027 \\\n    --hash=sha256:240cbec09667c1fed4c6cd0060b9ec57332427d7441289a2ed8875dc9fb2b224 \\\n    --hash=sha256:24e861e9630e0daddcb9191fb187f60f034e17a4426f8101279f0c475cd74144 \\\n    --hash=sha256:2729fcfc4f6a596fb0c50f32090400aa9367774ac296a00387e65098c0befa76 \\\n    --hash=sha256:2c1fe720934a16ea8e7146175cba2126f87f54912c8c5435e7f7c7a51ef808d3 \\\n    --hash=sha256:2cabe6546e41dabe439999a23fcb5246e0c3b595b4315b96ef755252be90caeb \\\n    --hash=sha256:2dbe06fc16bc91502bca713704022182e5729861ae00277c3a23354b40929740 \\\n    --hash=sha256:3363fcc96e665878946ad7a106b9a13eac0541766a690ef287c0232ac768b6ec \\\n    --hash=sha256:377fe3732edbaf78ee74efdf2c9f49f6e99f20e7f9d2649fda3eb4badd77d76e \\\n    --hash=sha256:3ac6aff147deb9c09461b2d4bbdf6256831198f5d8a23f5d37138213090b6d8a \\\n    --hash=sha256:3f45789ce415a7ec0820dc4f82925f9b5f7732070be1dec1f5f23ec381435a24 \\\n    --hash=sha256:4103b77b8a8225e413107d2349b65eb3c1c52627b5cc5c3c4c1c6a798b218950 \\\n    --hash=sha256:4377407001ca3c057773f44d8ddd6358fa5f691407c1ba92210bd3cf8d9e4c95 \\\n    --hash=sha256:46c2f213e23a04b93a392942d782eb9e413e6ef6bf7c8c53884e599a5c174dcb \\\n    --hash=sha256:47e98aab9d8d82ff682e7b0b5dded33bf138a32b817fcf7fa3b27b2d7c412928 \\\n    --hash=sha256:4a36f9becdd4c5c52a20c3e9484128b070b1dcfc8944c006f3a528295a359a9c \\\n    --hash=sha256:4af7b7e1be0a69bee8210735fe6dcfc38879adfac6d62e789d53ba432d1ffa41 \\\n    --hash=sha256:4d97a951a81039050e45f04e96689b58b8243fa5e62aa14fe67cb6075300885e \\\n    --hash=sha256:4db9aecb141cb7a5447171b57aa1ed3a8fee06af40b992ffc31206c0b0121550 \\\n    --hash=sha256:53e549287ef628fecba270045c9701b0c564563a9b0577d24a4ec75b8ab8040f \\\n    --hash=sha256:56b149b22de33b23b0c6077ab9518c6dcb538ad462e1830e68d06591ccf6e38b \\\n    --hash=sha256:570fec8fbd22b032733625f03f10b7ff023bc399213db15e72a7acaef28c2f4e \\\n    --hash=sha256:5b8ee53be440a0cffc991a27be3057e0530122548dbe7c0892df08822fce5ede \\\n    --hash=sha256:5ba4f78df2bcc19f764a4b26a8a4f5049c110090ad5825993aacb052bf8003ad \\\n    --hash=sha256:5c55256dee8f4b27bfbf636c8363383c7c8db7890c7cba5217d7bd5f5f21dab6 \\\n    --hash=sha256:5c88e5815a49d289e599f3513aa7fde0bc2092ff188f99c940f007f90f53d104 \\\n    --hash=sha256:5fede79c6f73ff2c3ef822864cb1ada23196e62756df53bc6231d351a49516a2 \\\n    --hash=sha256:65be18ec59496c13908f02a2472751d9ef840b4f3fb5726f129306bf6a2a7bba \\\n    --hash=sha256:66410eb6345d467151934b49bfa70fb32f5b35a6140baa40ad97d6436abea2e9 \\\n    --hash=sha256:665b0a2c463cc9423dd647e0bfd9f4ccc9b50f768c55304d5e9f80b177c1de12 \\\n    --hash=sha256:6b8536851f9f65e7f00c7a1d49ba7f2be0ffe2c11555367fc9f50d9f842410a1 \\\n    --hash=sha256:6c95b17fe34ed802f17e205112e6e10db92275c34fee290aa9bdc55a9c724027 \\\n    --hash=sha256:6e73e7fe93f17a7b191f52ec9da9dd8c06a8fe735a1ecbd13b97d1c723bff385 \\\n    --hash=sha256:6efbccc3d7f75d5b03105172a8dc86d82ba4da86817952529dd93185f4a88be2 \\\n    --hash=sha256:709f1efed56c4a145793c046cd4939f9959bcd818979a787b77d8e09c57a0840 \\\n    --hash=sha256:79af890482fc94648e8cde4c68620378f7fef60932710fa17a66abc039244da2 \\\n    --hash=sha256:7bcbe0fcf850eae67b6b01749815a4f7161c560a844c769ad7b48fcd99f791c4 \\\n    --hash=sha256:7c0494a31a1ac5461a226e7947a9c9b78c44e1dc7185164fa7e9651557a5d9bc \\\n    --hash=sha256:7ce27823052e2013b597e0c738b13e7e36b8ccb9400df8959417b052ab0fd92c \\\n    --hash=sha256:7f72c74aa99359e27a2ee8d6613fefa28b5f76a983c083074dfc2aaa4ab46213 \\\n    --hash=sha256:7fa5e51397466ea7e98de493fa2ff1b8193cfef8a7b0f9b4842f92d342df0dba \\\n    --hash=sha256:82632daed195dcc8ea664e8556dc9bdbd671960fb3776bd92806ce05792c2448 \\\n    --hash=sha256:82f75e05912e84b7a0fe57075d9c59de3cb352b928330f2eb69b2e1f54c3e1f0 \\\n    --hash=sha256:841f0852f48fefea3b12c9dfec00704dfa3aef5215d0e3ce564bb3d7cd8d57c6 \\\n    --hash=sha256:874019bd513008b009f58657134e5d0c5e030b3559bd0553976837adf52fe966 \\\n    --hash=sha256:88f50c94e21a0a7f14042c015b0eba1881af78562e7bf007e0033e624da59750 \\\n    --hash=sha256:89a1bbb58e0e3f7a283653d854b1e95d65e5cfd4af224dac5f02629ec1a3e621 \\\n    --hash=sha256:8a6987eaad834cb32dd57d9d582225f0054a5d1af706ccfbbdba735af4927e13 \\\n    --hash=sha256:8ac73abdc7ab75610f95a8fd994c6457e87752b02a63987e188f937a1fc180f0 \\\n    --hash=sha256:8ccf9aca873b767977c73df497a85dbedee4ee086ae9ae49dc461333b9b79f58 \\\n    --hash=sha256:90333fd89b43c0d08ac85f3f1447593fc2c66de18c3d6378d7125ea118dc7a54 \\\n    --hash=sha256:92ab3e11448f2ff7bf53c5a26eff0edc086898ec8b21fb154b85839ce1d88075 \\\n    --hash=sha256:9335a099ad87287c37fe5d1a982ff392fa5efe5d14b40a730b1ec1d6a41382b4 \\\n    --hash=sha256:96d30286dd02679e32a39aa8f0b7498fc847fcda46cfc09df5513e82ce252440 \\\n    --hash=sha256:9baafc71b04f8f4bb0703b21d6fc9f0c30b346c636a532ff16ec8491a5ea4b1f \\\n    --hash=sha256:9d1216a7f6f77836617dba35687c5b78a4170afc3c3f18fc788f785ba26565c4 \\\n    --hash=sha256:9d399bdcfb4a0f659b9b3788bbc89babe63d9a6a65aacdf4d4e7065ff2e6316c \\\n    --hash=sha256:9e4e16c73d717c5cf27626c524d0a2e261ad20e46932b2670f64ad5dde23e26f \\\n    --hash=sha256:9f4d8cf085a4c6a40fb97ea0f46938a8df43c85d31f9d45e2a8867ea9293790d \\\n    --hash=sha256:a33700d13d9b7d84fd10947b09ff69fb9a792e519c8cb9764a3ca70baa6c23a7 \\\n    --hash=sha256:a3732e66413163e72508da9eff9ce9d2846fde51fae45d3605393d3e6cd303e9 \\\n    --hash=sha256:a4582acf7ef76482f6f511ebaf1946dae7f2e85ec4728b81a678c01df63bd723 \\\n    --hash=sha256:a61834fb15d81322d872eaafd333838ae7c9cea84067f232656f75965933d047 \\\n    --hash=sha256:a7cff474ab7cd149765bb784cf6d78b32e18e20473fb7bda860bce98ab58e9da \\\n    --hash=sha256:a8fe66b8f300da93798025a785a5b90b42f3810dc2b72283ff84a41aaaebc293 \\\n    --hash=sha256:a929d878fec099030c292803b31e5d5540a7b6a31e6a3cc76cb4685fc2a2f51b \\\n    --hash=sha256:ad5d8201d310b031e6cd839d9bac2d4e5a01533ce5d3d5b50b7de1ef3af1de61 \\\n    --hash=sha256:af3aefa655adb5869491fa907e652290386800ae99cc50095cba71e2c6aefdca \\\n    --hash=sha256:c0ebc836c47a6477e182169c6a476fc691d12b518894bf7dd2572f0d59f1c7ed \\\n    --hash=sha256:c687ed078e145f5fd53a14854beff320e1d2ab76df03e2009c98f39a0f68f39a \\\n    --hash=sha256:cbb833ccacdb5519eff9b8b71ee618cc2801c878e77e288775d77c3a2ced858a \\\n    --hash=sha256:cf139c02f5f23ef6532040a30ff662c00a318c952334f211046b8e60b7f17688 \\\n    --hash=sha256:d46b86567dd4e248c6c159fcbcdcce01e0a5c8a7cd2334a0fff759d0fa075b16 \\\n    --hash=sha256:d693396e5aea78db03decd60aec9ece16c9b40ba00a587f089615ff4e718a81d \\\n    --hash=sha256:d897129df1a22b12aeed2c2c98df0785a2e8e6e0bde87b389491d0025c187077 \\\n    --hash=sha256:daba5e594f06114e37db186efd2dd916609071e59daca901a0a2e71f02b142ce \\\n    --hash=sha256:dd625535328fd9882374356269227670189adfcc6a2d90284f323c05862eecbd \\\n    --hash=sha256:e006d3a974c4ee19512e5f058abedb6eef36a5e553c14812bdeba1758d812e6d \\\n    --hash=sha256:e1ae548a9d901adca07899a4147a7c826bbcc06239d3ce9a59f57886a28a4c88 \\\n    --hash=sha256:e2935f8c39e3b03e83519292d78f075189978f3f4adc15a78144c7c8e2a1cba5 \\\n    --hash=sha256:e42d75862735da90e7fc5a7b23db0c976f737113a54b3c9777a9b665e9cbff75 \\\n    --hash=sha256:e7d42c531243450ef0d4d9c172e7ed6ef052640f195629065041b5add4e058d1 \\\n    --hash=sha256:e81b83143bee16329c23db3c1b2d82b29892fcbcb849186d2f6e98a5abe9a57f \\\n    --hash=sha256:e8ffa78582120024f476a611d7befc123cee59e47e8309d470cf667d806e613b \\\n    --hash=sha256:ebb0ec7f17803063d5aeb982f3b1bd2b2f4e4fae6751226cbd6ba1fcfe9e63ff \\\n    --hash=sha256:f08c7513ecef5aad65687bfdf6bc601ae9fccd04a42904501f8f7141abad9eb9 \\\n    --hash=sha256:f0a658a6d3fafee5c6f63c58f3e785c8c43c93fbc02bf9f2b6663f8185e0971f \\\n    --hash=sha256:f0e466ed7511fe9d459a819edbc6c2585c0b6eabde9fa8a8947552468a7a6ef0 \\\n    --hash=sha256:f141474e85b7e54998ec5180530a7cda99ab29e282fa50e0756d89981a9b43c5 \\\n    --hash=sha256:f4239bbec5a3577ddb49e4b50aeb32d8e5792098262ae2f63723f916a29b1a25 \\\n    --hash=sha256:f540c013589084679a6c7fac07096b10159737918174f5dfc5e11bf5bca4dfe6 \\\n    --hash=sha256:f9f3e9c8a9ecffa57bef8fb4fa19e5fa4d2d8307cf6bac5b1fca5e5860f4ba00 \\\n    --hash=sha256:fa139875ff98ab97da323cfadfaff08900d1ad42f1b5087b0b812a55c5a06373 \\\n    --hash=sha256:fcd3b77e2f17bbe4ca56ec7bcb07992647d19d0b9c05d84886dcd6f9eb810afd \\\n    --hash=sha256:fd8c81f346b58f45818d09ea11db69a8d5fd34a224b79871f6d44f12cd7977b1 \\\n    --hash=sha256:fe7b7bb170daccbba19ad33012d2b15f1e7942296fd4d45fc1b79013da8cc0f2 \\\n    --hash=sha256:ff330d3c30db4eb6b01d79e29d2d0b407a7ecad39cfd9ec993ece57396a2ec0d \\\n    --hash=sha256:ff405d91509d88e8d44129cd87b18d70acd1f0c1aeabd7bc3c46792b1fe2acba \\\n    --hash=sha256:ffcd54362564dc1a30fb74d8b8a6e5a6b11ebd5e27266adc3b7427a21a6c9104\nzipp==4.1.0 \\\n    --hash=sha256:25ad4e16390cd314347dd8f1de67a2ac538ae658ed4ab9db16029c07c188e97f \\\n    --hash=sha256:4cb57381f544315db7688e976e922a2b18cdb513d21cc194eb42232ba2a3e602\n\n# The following packages were excluded from the output:\n# torch\n# torchvision\n# triton\n# xformers\n# torchao\n# cuda-bindings\n# cuda-pathfinder\n# cuda-toolkit\n# nvidia-cublas-cu12\n# nvidia-cuda-cupti-cu12\n# nvidia-cuda-nvrtc-cu12\n# nvidia-cuda-runtime-cu12\n# nvidia-cudnn-cu12\n# nvidia-cufft-cu12\n# nvidia-cufile-cu12\n# nvidia-curand-cu12\n# nvidia-cusolver-cu12\n# nvidia-cusparse-cu12\n# nvidia-cusparselt-cu12\n# nvidia-nccl-cu12\n# nvidia-nvshmem-cu12\n# nvidia-nvjitlink-cu12\n# nvidia-nvtx-cu12\n')
LOCKED_REQUIREMENTS = APPLICATION_CONSTRAINTS
def installed_package_versions():
    return {str(dist.metadata.get("Name") or "").lower().replace("_", "-"): str(dist.version) for dist in importlib.metadata.distributions() if dist.metadata.get("Name")}
def verify_locked_environment(expected=None):
    expected = dict(expected or LOCKED_REQUIREMENTS)
    installed = installed_package_versions()
    missing = sorted(name for name in expected if name not in installed)
    mismatched = {name: {"expected": version, "installed": installed.get(name)} for name, version in expected.items() if installed.get(name) != version}
    if missing or mismatched: raise RuntimeError(json.dumps({"locked_environment_mismatch": True, "missing": missing, "mismatched": mismatched}, sort_keys=True))
    return {"locked_requirements": expected, "installed_package_versions": installed, "verified": True}
LOCK_ENVIRONMENT = verify_locked_environment()
def _cuda_driver_version():
    try: return torch._C._cuda_getDriverVersion()
    except Exception: return None
ENVIRONMENT_FINGERPRINT = {"python_version": platform.python_version(), "python_implementation": platform.python_implementation(), "python_executable": sys.executable, "python_hash_seed": os.environ.get("PYTHONHASHSEED"), "hash_randomization": bool(sys.flags.hash_randomization), "locked_requirements": LOCK_ENVIRONMENT["locked_requirements"], "installed_package_versions": LOCK_ENVIRONMENT["installed_package_versions"], "lock_package_counts": LOCK_PACKAGE_COUNTS, "pytorch_version": getattr(torch, "__version__", None), "torch_cuda_version": getattr(torch.version, "cuda", None), "cuda_driver_version": _cuda_driver_version(), "cudnn_version": torch.backends.cudnn.version() if torch.cuda.is_available() else None, "cuda_available": bool(torch.cuda.is_available()), "managed_accelerator": MANAGED_ACCELERATOR, "qwen35_transformers_compatibility": QWEN35_TRANSFORMERS_COMPATIBILITY, "gpu_name": torch.cuda.get_device_name(0) if torch.cuda.is_available() else None, "gpu_count": int(torch.cuda.device_count()) if torch.cuda.is_available() else 0, "model_revisions": {"bertscore_model_id": BERTSCORE_MODEL_ID, "bertscore_model_revision": BERTSCORE_MODEL_REVISION, "nli_model_id": NLI_MODEL_ID, "nli_model_revision": NLI_MODEL_REVISION, "generator_model_id": (os.environ.get("MODEL_ID") or "unsloth/" + "Qwen3.5-4B"), "generator_model_revision": os.environ.get("MODEL_REVISION"), "embedding_model_id": "BAAI/bge-m3", "embedding_model_revision": os.environ.get("EMBEDDING_MODEL_REVISION"), "reranker_model_id": "BAAI/bge-reranker-v2-m3", "reranker_model_revision": os.environ.get("RERANKER_MODEL_REVISION")}}
ENVIRONMENT_FINGERPRINT["managed_accelerator_contract_hash"] = MANAGED_ACCELERATOR_CONTRACT_HASH
ENVIRONMENT_FINGERPRINT["detoxify_metric_contract"] = {"device": "cpu", "required": True, "language": "en"}
ENVIRONMENT_FINGERPRINT_HASH = stable_id("environment-fingerprint.v1", json.dumps(ENVIRONMENT_FINGERPRINT, ensure_ascii=False, sort_keys=True, default=str))



# %% [notebook cell 5]

# 03 - Configuration and frozen run identity.
ENVIRONMENT_FINGERPRINT["model_revisions"].update({"generator_model_revision": GENERATOR_MODEL_REVISION, "embedding_model_revision": EMBEDDING_MODEL_REVISION, "reranker_model_revision": RERANKER_MODEL_REVISION})
ENVIRONMENT_FINGERPRINT_HASH = stable_id("environment-fingerprint.v1", json.dumps(ENVIRONMENT_FINGERPRINT, ensure_ascii=False, sort_keys=True, default=str))
CONFIG = {
    "dataset_xlsx": PROJECT_ROOT / "input" / "dataset (1).xlsx", "corpus_root": PROJECT_ROOT / "input" / "lgbt_hate_speech_kg_sources", "sheet_name": "Final_Dataset",
    "evaluation_categories": ["Homophobic", "Non-Homophobic"], "max_experiment_rows": MAX_EXPERIMENT_ROWS, "split_name": os.environ.get("SPLIT_NAME", "test"), "smoke_test": False, "smoke_rows": 8,
    "minimum_parse_rate": 0.98, "require_citation_nli": True, "require_detoxify": True, "detoxify_device": "cpu", "evidence_char_budget": 9000, "embedding_model": "BAAI/bge-m3", "reranker_model": "BAAI/bge-reranker-v2-m3", "generator_model": os.environ.get("MODEL_ID", "unsloth/Qwen3.5-4B"), "load_in_4bit": True, "max_seq_length": 4096,
    "perspective_batch_size": 5, "extraction_batch_size": 24, "perspective_max_new_tokens": 768, "plan_max_new_tokens": 768, "answer_max_new_tokens": 768, "reasoning_max_new_tokens": 192, "thinking_enabled": True, "minimum_healthy_factual_documents": 40, "minimum_factual_document_coverage": 0.80, "dense_top_k": 12, "bm25_top_k": 12, "graph_top_k": 8, "rerank_top_k": 5,
    "max_chunk_chars": 1800, "chunk_overlap_chars": 220, "minimum_authority": 0.60, "minimum_dense_score": 0.25, "minimum_graph_score": 0.0, "minimum_rerank_probability": 0.55, "max_graph_hops": 2, "rrf_constant": 60.0,
    "precompute_evidence": True, "graph_ablation": False, "generation_variants": ["qwen_zero_shot", "qwen_few_shot", "kg_rag", "mp_kg_rag"], "shard_index": SHARD_INDEX, "shard_count": SHARD_COUNT, "cache_max_records": EFFECTIVE_CACHE_CAPACITY,
    "evaluation_languages": ["en", "hi", "ta"], "bertscore_model_id": BERTSCORE_MODEL_ID, "bertscore_model_revision": BERTSCORE_MODEL_REVISION, "lockfile_sha256": LOCKFILE_SHA256, "lock_package_counts": LOCK_PACKAGE_COUNTS, "environment_fingerprint_hash": ENVIRONMENT_FINGERPRINT_HASH,
    "nli_model_id": NLI_MODEL_ID, "nli_model_revision": NLI_MODEL_REVISION, "nli_model_label_mapping": MODEL_LABEL_MAPPING, "nli_dataset_label_mapping": DATASET_LABEL_MAPPING, "nli_calibration_examples": 600, "nli_calibration_min_support": 50, "nli_calibration_bootstrap": 2000, "nli_min_accuracy": 0.70, "nli_min_entailment_precision": 0.70, "nli_min_entailment_recall": 0.70, "nli_min_per_label_support": 50, "nli_calibration_provenance": {language: {**NLI_DATASET_PROVENANCE[language], "dataset_content_hash": None, "code_hash": CORE_SOURCE_SHA256, "eval_core_hash": EVAL_CORE_SOURCE_SHA256, "model_id": NLI_MODEL_ID, "model_revision": NLI_MODEL_REVISION, "model_label_mapping": MODEL_LABEL_MAPPING, "dataset_label_mapping": DATASET_LABEL_MAPPING} for language in ["en", "hi", "ta"]}, "annotation_max_ids": 200,
}
CONFIG["generator_model_revision"] = GENERATOR_MODEL_REVISION
CONFIG["embedding_model_revision"] = EMBEDDING_MODEL_REVISION
CONFIG["reranker_model_revision"] = RERANKER_MODEL_REVISION
CONFIG["managed_accelerator_contract"] = MANAGED_ACCELERATOR_CONTRACT
CONFIG["managed_accelerator_contract_hash"] = MANAGED_ACCELERATOR_CONTRACT_HASH
CONFIG["qwen35_transformers_compatibility"] = QWEN35_TRANSFORMERS_COMPATIBILITY
CONFIG["nli_model_label_mapping"] = MODEL_LABEL_MAPPING
CONFIG["nli_dataset_label_mapping"] = DATASET_LABEL_MAPPING
CONFIG["nli_min_per_label_support"] = 50
CONFIG["nli_min_accuracy_lower"] = 0.70
CONFIG["nli_min_entailment_precision_lower"] = 0.70
CONFIG["nli_min_entailment_recall_lower"] = 0.70
for _language in ["en", "hi", "ta"]:
    CONFIG["nli_calibration_provenance"][_language]["dataset_content_hash"] = None
    CONFIG["nli_calibration_provenance"][_language]["model_label_mapping"] = MODEL_LABEL_MAPPING
    CONFIG["nli_calibration_provenance"][_language]["dataset_label_mapping"] = DATASET_LABEL_MAPPING
IDENTITY = identity_config(CONFIG)
CONFIG_HASH = IDENTITY["config_hash"]
SPLIT_IDENTITY_HASH = IDENTITY["split_identity_hash"]
RUN_IDENTITY_HASH = IDENTITY["run_identity_hash"]
PROMPT_TEMPLATE_HASH = stable_id("mp_kg_rag_prompt_templates_v4_closed_catalog_fail_closed")
FEW_SHOT_PROMPT_REVISION = "few-shot.v1.contamination-safe.static"
FEW_SHOT_EXAMPLES = [
    {"post": "A claim says a group is dangerous because of identity.", "target": "identity-based hate", "response": "Identity does not determine a person's character. Respond to people as individuals and avoid generalizing about a protected group."},
    {"post": "Someone says respect means staying silent about discrimination.", "target": "respect and equality", "response": "Respect includes listening to affected people and rejecting discrimination while keeping the conversation constructive."},
]
EXTRACTION_MODEL = CONFIG["generator_model"]; EXTRACTION_PROMPT_REVISION = "semantic-claims.v1-qwen-constrained-repair1"
RETRIEVAL_THRESHOLDS = {k: CONFIG[k] for k in ["minimum_authority", "minimum_dense_score", "minimum_graph_score", "minimum_rerank_probability", "max_graph_hops"]}
GRAPH_CONFIG = {"max_hops": CONFIG["max_graph_hops"], "hop_decay": 0.75, "minimum_dense_score": CONFIG["minimum_dense_score"], "minimum_graph_score": CONFIG["minimum_graph_score"], "minimum_rerank_probability": CONFIG["minimum_rerank_probability"], "max_evidence": CONFIG["rerank_top_k"], "weights": {"query_entity": 2.0, "predicate": 1.0, "polarity": 0.5, "modality": 0.4, "stance": 0.8, "review_state": 0.8, "authority": 0.8, "extraction_confidence": 0.0, "seed_score": 1.0, "hop_decay": 0.6}, "review_state_scores": {"accepted": 1.0, "reviewed": 0.85, "unknown": 0.0}, "extraction_confidence_status": "diagnostic_only_unweighted"}
GRAPH_CONFIG_HASH = stable_id(json.dumps(GRAPH_CONFIG, sort_keys=True))
SCORING_CALIBRATION_STATUS = "uncalibrated_diagnostic_only"
SELF_CONFIDENCE_STATUS = "uncalibrated_diagnostic_only"
QUALITY_THRESHOLDS = {"minimum_graph_linked_claim_rate": 0.02, "minimum_query_linked_entity_rate": 0.02, "minimum_linked_claims": 1, "minimum_linked_queries": 1}
def validate_checkpoint_identity(saved, expected):
    if saved != expected:
        raise RuntimeError("resume_identity_mismatch")
    return True
def checkpoint_identity(record, variant):
    return {"record_id": str(record["ID"]), "variant": variant, "input_text_sha256": str(record["input_text_sha256"]), "target_sha256": hashlib.sha256(str(record.get("Target", "")).encode()).hexdigest(), "category_sha256": hashlib.sha256(str(record.get("Category", "")).encode()).hexdigest(), "reference_answer_sha256": hashlib.sha256((normalize_optional_text(record.get("Counter Narrative")) or "").encode()).hexdigest(), "corpus_manifest_hash": CORPUS_MANIFEST_HASH, "audit_manifest_hash": AUDIT_MANIFEST_HASH, "chunk_manifest_hash": CHUNK_MANIFEST_HASH, "graph_manifest_hash": GRAPH_MANIFEST_HASH, "extraction_model": EXTRACTION_MODEL, "extraction_prompt_revision": EXTRACTION_PROMPT_REVISION, "retrieval_thresholds": RETRIEVAL_THRESHOLDS, "graph_config": GRAPH_CONFIG, "config_hash": CONFIG_HASH, "run_identity_hash": RUN_IDENTITY_HASH, "lockfile_sha256": LOCKFILE_SHA256, "environment_fingerprint_hash": ENVIRONMENT_FINGERPRINT_HASH, "prompt_template_hash": PROMPT_TEMPLATE_HASH, "variant_prompt_hash": stable_id(PROMPT_TEMPLATE_HASH, FEW_SHOT_PROMPT_REVISION if variant == "qwen_few_shot" else variant), "split_name": CONFIG["split_name"], "split_membership_hash": SPLIT_MEMBERSHIP_HASH, "shard_index": SHARD_INDEX, "shard_count": SHARD_COUNT, "run_name": RUN_NAME}



# %% [notebook cell 6]

# 04 - Dataset audit and manifest-authoritative collision-safe corpus registry.
from bs4 import BeautifulSoup
import fitz
def audit_dataset(path):
    required = ["ID", "Text", "Category", "Target", "Counter Narrative"]; frame = pd.read_excel(path, sheet_name=CONFIG["sheet_name"]); assert set(required).issubset(frame.columns)
    frame = frame[required].copy()
    raw_rows = frame.to_dict("records")
    base = filter_audit(raw_rows, [normalize_optional_text(row.get("ID")) is not None and normalize_optional_text(row.get("Text")) is not None for row in raw_rows], filter_name="required_id_text", reasons=[None if normalize_optional_text(row.get("ID")) is not None and normalize_optional_text(row.get("Text")) is not None else "missing_id_or_text" for row in raw_rows])
    frame = pd.DataFrame(base["kept_rows"], columns=required)
    frame["ID"] = frame["ID"].astype(str).str.replace(r"\.0$", "", regex=True); assert not frame.ID.duplicated().any()
    category_result = filter_rows_by_category(frame.to_dict("records"), CONFIG["evaluation_categories"]); frame = pd.DataFrame(category_result["rows"])
    reference_result = quarantine_missing_references(frame.to_dict("records"), reference_key="Counter Narrative")
    frame["reference_available"] = frame["ID"].astype(str).isin({row["ID"] for row in reference_result["scorable_rows"]})
    frame["CategoryNormalized"] = frame["Category"].map(normalize_category)
    frame["input_text_sha256"] = frame.Text.astype(str).map(lambda x: hashlib.sha256(x.encode()).hexdigest()); frame["script_bucket"] = frame.Text.astype(str).map(lambda x: "tamil" if any("\u0b80" <= c <= "\u0bff" for c in x) else ("devanagari" if any("\u0900" <= c <= "\u097f" for c in x) else "latin_or_mixed")); frame["stratify_key"] = frame.CategoryNormalized.astype(str) + "|" + frame.Target.fillna("No Target").astype(str) + "|" + frame.script_bucket
    write_json(RUN / "artifacts" / "dataset_filter_manifest.json", {"filters": [base, category_result["manifest"], reference_result["manifest"]], "reference_quarantine_count": len(reference_result["quarantined_rows"]), "reference_quarantine_policy": "generation_preserved_scoring_excluded"})
    log_event("dataset", "filter_manifest_written", reference_quarantine_count=len(reference_result["quarantined_rows"]), dataset_rows=len(frame))
    return frame.reset_index(drop=True)
def make_frozen_split(frame):
    path = RUN / "artifacts" / "frozen_split.json"
    membership = [{"ID": str(row["ID"]), "input_text_sha256": hashlib.sha256(str(row["Text"]).encode()).hexdigest(), "target_sha256": hashlib.sha256(str(row.get("Target", "")).encode()).hexdigest(), "category_sha256": hashlib.sha256(str(row.get("Category", "")).encode()).hexdigest(), "reference_answer_sha256": hashlib.sha256((normalize_optional_text(row.get("Counter Narrative")) or "").encode()).hexdigest()} for _, row in frame.sort_values("ID").iterrows()]
    membership_hash = stable_id("frozen-membership.v2", json.dumps(membership, ensure_ascii=False, sort_keys=True))
    if path.exists():
        frozen = json.loads(path.read_text())
        if frozen.get("config_hash") != CONFIG_HASH or frozen.get("split_membership_hash") != membership_hash or frozen.get("membership") != membership:
            raise RuntimeError("frozen_split_identity_mismatch")
        return frozen
    buckets = {"train": [], "dev": [], "test": []}
    for _, group in frame.groupby("stratify_key", dropna=False):
        for i, rid in enumerate(sorted(group.ID.astype(str), key=lambda x: stable_id(x, SEED))): buckets["train" if i % 10 < 7 else ("dev" if i % 10 < 9 else "test")].append(rid)
    value = {"config_hash": CONFIG_HASH, "seed": SEED, "stratify_fields": ["Category", "Target", "script_bucket"], "split_membership_hash": membership_hash, "membership": membership, "splits": buckets}; write_json(path, value); return value
all_dataset = audit_dataset(CONFIG["dataset_xlsx"]); frozen_split = make_frozen_split(all_dataset); SPLIT_MEMBERSHIP_HASH = frozen_split["split_membership_hash"]; dataset = all_dataset[all_dataset.ID.astype(str).isin(frozen_split["splits"][CONFIG["split_name"]])].copy(); dataset = dataset.head(CONFIG["smoke_rows"] if CONFIG["smoke_test"] else CONFIG["max_experiment_rows"]).reset_index(drop=True)
if CONFIG["shard_count"] > 1: dataset = dataset.iloc[CONFIG["shard_index"]::CONFIG["shard_count"]].reset_index(drop=True)
if len(dataset) > int(CONFIG["cache_max_records"]): raise RuntimeError("cache_capacity_underprovisioned_before_work")
log_event("runtime", "cache_capacity_bound", cache_capacity=CONFIG["cache_max_records"], selected_dataset_rows=len(dataset), capacity_hashed=True)
source_registry_list = load_source_registry(CONFIG["corpus_root"]); source_registry = pd.DataFrame(list(source_registry_list));
if source_registry.empty: raise RuntimeError("no_corpus_documents")
source_registry["source_id"] = source_registry["legacy_source_id"]; source_registry["legacy_source_id"] = source_registry["legacy_source_id"].fillna("UNKNOWN");
if "review_status" not in source_registry: source_registry["review_status"] = "unreviewed"
else: source_registry["review_status"] = source_registry["review_status"].fillna("unreviewed")
assert source_registry.document_uid.is_unique
def canonical_manifest_rows(frame):
    fields = ["document_uid", "relative_path", "content_sha256", "source_type", "authority_score", "factual_index_allowed", "status", "status_reason", "quarantine_reasons", "review_status"]
    rows = []
    for row in frame.to_dict("records"):
        rows.append({key: row.get(key) for key in fields})
    return sorted(rows, key=lambda row: (str(row["document_uid"]), str(row["relative_path"]), str(row["content_sha256"])))
canonical_manifest_rows = canonical_manifest_rows(source_registry); document_uid_manifest_hash = stable_id(json.dumps(canonical_manifest_rows, ensure_ascii=False, sort_keys=True, default=str)); CORPUS_MANIFEST_HASH = document_uid_manifest_hash
def parquet_safe_frame(frame):
    safe = frame.copy()
    for column in safe.columns:
        if safe[column].dtype != object: continue
        if any(isinstance(value, (dict, list, tuple, set)) for value in safe[column] if value is not None):
            safe[column] = safe[column].map(lambda value: json.dumps(value, ensure_ascii=False, sort_keys=True, default=str) if isinstance(value, (dict, list, tuple, set)) else value)
    return safe
write_json(RUN / "artifacts" / "source_manifest.json", {"rows": source_registry.to_dict("records"), "audit_events": list(getattr(source_registry_list, "audit_events", [])), "validation_errors": list(getattr(source_registry_list, "validation_errors", [])), "file_records_before_deduplication": int(getattr(source_registry_list, "file_records_before_deduplication", len(source_registry)))}); parquet_safe_frame(source_registry).to_parquet(RUN / "artifacts" / "source_registry.parquet", index=False); parquet_safe_frame(dataset).to_parquet(RUN / "artifacts" / "dataset_scope.parquet", index=False)
print("Corpus files before dedup:", source_registry_list.file_records_before_deduplication, "| documents:", len(source_registry), "| hidden_metadata_file audit retained by registry")
reviewed_orgs = sorted({str(row.get("organisation")) for row in source_registry.to_dict("records") if row.get("organisation") and row.get("status") == "accepted"})
BASE_ENTITY_CATALOG = build_entity_catalog(source_registry.to_dict("records"), dataset["Target"].astype(str).tolist(), reviewed_orgs)
write_json(RUN / "artifacts" / "base_entity_catalog.json", BASE_ENTITY_CATALOG)



# %% [notebook cell 7]

# 05 - Format-aware PDF/HTML extraction, source-level audit, and document_uid chunks.
EXTRACTOR_REVISION = "format_aware_pdf_html_audit_v4_fail_closed"
COOKIE_MARKERS = ("enable cookies", "please enable cookies", "javascript is disabled", "checking your browser", "access denied", "captcha")
RECOVERY_EXPECTED_SOURCES = {"SRC029": "nested-div", "SRC059": "long-heading"}
PARENT_DOCUMENT_UID_FIELD = "parent_document_uid"
def normalize_text(text): return re.sub(r"\s+", " ", str(text or "")).strip()
def canonical_page(value): return None if value is None or (isinstance(value, float) and math.isnan(value)) or pd.isna(value) else int(value)
def split_windows(text, max_chars, overlap):
    # Exact, non-overlapping sentence windows come from the tested evaluation core.
    return sentence_aligned_windows(text, max_chars, overlap)
def content_based_interstitial(text, title=""):
    lowered = normalize_text(text).casefold()
    title_lowered = normalize_text(title).casefold()
    return (any(marker in lowered for marker in COOKIE_MARKERS) and len(lowered) < 700) or ("pubmed" in title_lowered and "cookies-required" in lowered)
def _validated_fallback_path(source):
    metadata = source.get("manifest_metadata") if isinstance(source.get("manifest_metadata"), dict) else {}
    fallback_path = source.get("fallback_path") or metadata.get("fallback_path")
    fallback_sha256 = source.get("fallback_sha256") or metadata.get("fallback_sha256")
    if not fallback_path or not isinstance(fallback_sha256, str): return None
    candidate = Path(str(fallback_path)); candidate = candidate if candidate.is_absolute() else Path(source["path"]).parent / candidate
    if not candidate.is_file(): return None
    actual = hashlib.sha256(candidate.read_bytes()).hexdigest()
    return (candidate, actual) if actual == fallback_sha256.casefold() else None
def _html_root(soup):
    # Prefer semantic content containers, including role=main, before body fallback.
    return soup.find("main") or soup.find("article") or soup.select_one('[role="main"]') or soup.body or soup
def deduplicate_contained_text_rows(rows, return_audit=False):
    kept, seen_hashes, audit = [], set(), []
    for row in rows:
        text = normalize_text(row.get("text"))
        if not text: continue
        identity = {"document_uid": row.get("document_uid"), "relative_path": row.get("relative_path"), "locator": {"paragraph": row.get("paragraph"), "page": row.get("page")}}
        if any(normalize_text(existing.get("text")) == text for existing in kept):
            audit.append({"filter": "html_element", "kept": False, "reason": "duplicate_text", **identity}); continue
        contained_text = len(text) >= 40 and any(text.casefold() in normalize_text(existing.get("text")).casefold() for existing in kept)
        if contained_text:
            audit.append({"filter": "html_element", "kept": False, "reason": "contained_text", **identity}); continue
        removed = [existing for existing in kept if len(normalize_text(existing.get("text"))) >= 40 and normalize_text(existing.get("text")).casefold() in text.casefold()]
        for existing in removed:
            audit.append({"filter": "html_element", "kept": False, "reason": "contained_by_larger_text", "document_uid": existing.get("document_uid"), "relative_path": existing.get("relative_path"), "locator": {"paragraph": existing.get("paragraph"), "page": existing.get("page")}})
        kept = [existing for existing in kept if existing not in removed]
        kept.append(row)
    return (kept, audit) if return_audit else kept
def extract_html_document(source):
    html_path = Path(source["path"]); extraction_method = "html_dom_content_container"; fallback = None; fallback_attempted = False; fallback_path = None; fallback_sha256 = None; fallback_source_identity = None
    while True:
        soup = BeautifulSoup(html_path.read_text(encoding="utf-8", errors="ignore"), "html.parser")
        for element in soup(["script", "style", "nav", "header", "footer", "aside", "form", "noscript"]): element.decompose()
        root = _html_root(soup); title = normalize_text(soup.title.get_text(" ", strip=True) if soup.title else source["relative_path"])
        if not content_based_interstitial(root.get_text(" ", strip=True), title): break
        fallback = _validated_fallback_path(source)
        if fallback_attempted or fallback is None: return [], {"status": "blocked", "method": "html_cookie_interstitial", "reason": "content_based_interstitial", "fallback_path": fallback_path, "fallback_sha256": fallback_sha256, "fallback_source_identity": fallback_source_identity}
        fallback_attempted = True
        html_path, fallback_hash = fallback; extraction_method = "html_hashed_fallback"; fallback_path = str(html_path); fallback_sha256 = fallback_hash; fallback_source_identity = stable_id("html-fallback-source", fallback_path, fallback_hash); source = {**source, "path": str(html_path), "fallback_path": fallback_path, "fallback_sha256": fallback_hash, "fallback_source_identity": fallback_source_identity}
    rows, seen_hashes, section, short_element_events, empty_element_events = [], set(), title or "document", [], []
    for paragraph_index, element in enumerate(root.find_all(["p", "li", "blockquote", "td", "th", "dt", "dd", "div", "section", "h1", "h2", "h3", "h4"])):
        text = normalize_text(element.get_text(" ", strip=True)); is_heading = element.name.startswith("h")
        if not text:
            empty_element_events.append({"filter": "html_element", "kept": False, "reason": "html_empty_element", "document_uid": source.get("document_uid"), "relative_path": source.get("relative_path"), "locator": {"paragraph": paragraph_index, "tag": element.name}})
            continue
        if len(text) < 20:
            short_element_events.append({"filter": "html_element", "kept": False, "reason": "html_short_element", "document_uid": source.get("document_uid"), "relative_path": source.get("relative_path"), "locator": {"paragraph": paragraph_index, "tag": element.name}, "text_length": len(text)})
            if is_heading and not re.search(r"[.!?]", text): section = text
            continue
        if is_heading: section = text
        rows.append({**source, "title": title, "page": None, "section": section, "paragraph": paragraph_index, "text": text, "extraction_method": extraction_method, "quality_label": "native", "parent_document_uid": None})
    rows, dedup_events = deduplicate_contained_text_rows(rows, return_audit=True)
    html_filter_events = empty_element_events + short_element_events + dedup_events
    linked_pdf_rows = []
    for link in root.find_all("a", href=True):
        href = str(link.get("href", "")); candidate = (Path(source["path"]).parent / href.split("?", 1)[0]).resolve()
        if not href.casefold().endswith(".pdf") or not candidate.is_file(): continue
        child_bytes = candidate.read_bytes(); child_hash = hashlib.sha256(child_bytes).hexdigest(); child_source = {**source, "document_uid": stable_id("linked-pdf", source["document_uid"], child_hash), "content_sha256": child_hash, "path": str(candidate), "relative_path": str(candidate), "parent_document_uid": source["document_uid"]}
        child_rows, _ = extract_pdf_document(child_source); linked_pdf_rows.extend([{**row, "parent_document_uid": source["document_uid"], "extraction_method": "linked_pdf_native"} for row in child_rows])
    return rows + linked_pdf_rows, {"status": "extracted" if rows or linked_pdf_rows else "empty", "method": extraction_method, "reason": None, "linked_pdf_count": len(linked_pdf_rows), "fallback_path": fallback_path, "fallback_sha256": fallback_sha256, "fallback_source_identity": fallback_source_identity, "html_candidate_count": len(rows) + len(html_filter_events), "html_empty_element_count": len(empty_element_events), "html_short_element_events": short_element_events, "html_short_element_count": len(short_element_events), "html_filter_events": html_filter_events, "html_filter_kept_count": len(rows)}
def extract_pdf_document(source):
    rows = []; empty_page_events = []
    with fitz.open(source["path"]) as pdf:
        pdf_candidate_count = len(pdf)
        for page_no, page in enumerate(pdf, 1):
            text = normalize_text(page.get_text("text"))
            if text: rows.append({**source, "page": page_no, "section": "page", "paragraph": None, "text": text, "extraction_method": "pdf_native", "quality_label": "native"})
            else: empty_page_events.append({"filter": "pdf_page", "kept": False, "reason": "pdf_empty_page", "document_uid": source.get("document_uid"), "relative_path": source.get("relative_path"), "locator": {"page": page_no}})
    return rows, {"status": "extracted" if rows else "empty", "method": "pdf_native", "reason": None, "pdf_candidate_count": pdf_candidate_count, "pdf_empty_page_count": len(empty_page_events), "pdf_filter_events": empty_page_events, "pdf_filter_kept_count": len(rows)}
pages_rows = []
extraction_audit = []
for source in tqdm(source_registry.to_dict("records"), desc="Extract PDF/HTML corpus"):
    try:
        source_rows, outcome = extract_html_document(source) if source["document_type"] in {"html", "htm"} else extract_pdf_document(source)
        pages_rows.extend(source_rows); extraction_audit.append({"document_uid": source["document_uid"], "content_sha256": source["content_sha256"], "relative_path": source["relative_path"], "method": outcome["method"], "status": outcome["status"], "text_length": sum(len(row["text"]) for row in source_rows), "row_count": len(source_rows), "chunk_count": 0, "reason": outcome.get("reason"), "fallback_path": outcome.get("fallback_path"), "fallback_sha256": outcome.get("fallback_sha256"), "html_candidate_count": int(outcome.get("html_candidate_count", 0)), "html_empty_element_count": int(outcome.get("html_empty_element_count", 0)), "html_short_element_count": int(outcome.get("html_short_element_count", 0)), "html_filter_kept_count": int(outcome.get("html_filter_kept_count", 0)), "filter_events": outcome.get("html_filter_events", outcome.get("pdf_filter_events", [])), "pdf_candidate_count": int(outcome.get("pdf_candidate_count", 0)), "pdf_empty_page_count": int(outcome.get("pdf_empty_page_count", 0)), "pdf_filter_kept_count": int(outcome.get("pdf_filter_kept_count", 0)), "extractor_revision": EXTRACTOR_REVISION})
        extraction_audit[-1].update({"fallback_source_identity": outcome.get("fallback_source_identity"), "source_identity": stable_id("source-identity", source["document_uid"], source["content_sha256"], outcome.get("fallback_source_identity"))})
    except Exception as exc:
        log_event("extract", "source_failed", document_uid=source["document_uid"], error=str(exc)); extraction_audit.append({"document_uid": source["document_uid"], "content_sha256": source["content_sha256"], "relative_path": source["relative_path"], "method": "failed", "status": "failed", "text_length": 0, "row_count": 0, "chunk_count": 0, "reason": str(exc), "filter_events": [], "extractor_revision": EXTRACTOR_REVISION})
pages = pd.DataFrame(pages_rows)
chunk_rows = []; corpus_filter_rows = []
SOURCE_TEXT_BY_KEY = {}
for page in pages.to_dict("records"):
    if len(page["text"]) < 80:
        corpus_filter_rows.append({"filter": "page", "kept": False, "reason": "page_short_text", "document_uid": page.get("document_uid"), "relative_path": page.get("relative_path"), "locator": {"page": page.get("page"), "paragraph": page.get("paragraph")}, "text_length": len(page["text"])})
        continue
    corpus_filter_rows.append({"filter": "page", "kept": True, "reason": None, "document_uid": page.get("document_uid"), "relative_path": page.get("relative_path"), "locator": {"page": page.get("page"), "paragraph": page.get("paragraph")}, "text_length": len(page["text"])})
    page_value = canonical_page(page.get("page")); paragraph = page.get("paragraph")
    source_text_key = stable_id("source-text", page["document_uid"], page_value, paragraph); SOURCE_TEXT_BY_KEY[source_text_key] = str(page["text"]); source_text_sha256 = hashlib.sha256(str(page["text"]).encode("utf-8")).hexdigest()
    for index, window in enumerate(split_windows(page["text"], CONFIG["max_chunk_chars"], CONFIG["chunk_overlap_chars"])):
        text = window["text"]; text_sha256 = hashlib.sha256(text.encode()).hexdigest(); chunk_id = stable_id("chunk", page["document_uid"], page_value, paragraph, index, text_sha256)
        chunk_rows.append({**page, "page": page_value, "chunk_index": index, "chunk_id": chunk_id, "chunk_document_uid": page["document_uid"], "document_sha256": page["content_sha256"], "text_sha256": text_sha256, "text": text, "source_text_key": source_text_key, "source_text_sha256": source_text_sha256, "span_start": window["start_char"], "span_end": window["end_char"], "sentence_start": window["sentence_start"], "sentence_end": window["sentence_end"], "sentence_aligned": window["sentence_aligned"], "split_reason": window["split_reason"], "content_warning": page["source_type"] == "harmful_examples"})
chunks = pd.DataFrame(chunk_rows)
chunk_counts = chunks.groupby("document_uid").size().to_dict() if not chunks.empty else {}
for audit in extraction_audit: audit["chunk_count"] = int(chunk_counts.get(audit["document_uid"], 0)); audit["indexable"] = audit["chunk_count"] > 0
html_events = [event for audit in extraction_audit for event in audit.get("filter_events", []) if event.get("filter") == "html_element"]
pdf_events = [event for audit in extraction_audit for event in audit.get("filter_events", []) if event.get("filter") == "pdf_page"]
def filter_manifest_row(name, input_count, kept_count, events, kept_rows=None):
    all_rows = list(events) + list(kept_rows or [])
    reasons = {}
    for event in events:
        reasons[event["reason"]] = reasons.get(event["reason"], 0) + 1
    return {"filter": name, "input": int(input_count), "kept": int(kept_count), "dropped": int(input_count - kept_count), "reason_counts": reasons, "rows": all_rows}
corpus_filter_manifest = {"filters": [filter_manifest_row("html_element", sum(int(audit.get("html_candidate_count", 0)) for audit in extraction_audit), sum(int(audit.get("html_filter_kept_count", 0)) for audit in extraction_audit), html_events), filter_manifest_row("pdf_page", sum(int(audit.get("pdf_candidate_count", 0)) for audit in extraction_audit), sum(int(audit.get("pdf_filter_kept_count", 0)) for audit in extraction_audit), pdf_events), filter_manifest_row("page", len(corpus_filter_rows), sum(bool(row["kept"]) for row in corpus_filter_rows), [row for row in corpus_filter_rows if not row["kept"]], [row for row in corpus_filter_rows if row["kept"]])], "source_registry_audit_events": list(getattr(source_registry_list, "audit_events", [])), "source_registry_validation_errors": list(getattr(source_registry_list, "validation_errors", [])), "retrieval_corpus_identity_excludes_ignored_files": True}
AUDIT_MANIFEST_HASH = stable_id("audit-manifest.v1", json.dumps(corpus_filter_manifest, ensure_ascii=False, sort_keys=True, default=str))
write_json(RUN / "artifacts" / "corpus_filter_manifest.json", {**corpus_filter_manifest, "audit_manifest_hash": AUDIT_MANIFEST_HASH})
factual_documents = {str(row["document_uid"]): row for row in source_registry[source_registry.factual_index_allowed == True].to_dict("records")}
missing_factual_documents = [dict(audit) for audit in extraction_audit if audit["document_uid"] in factual_documents and audit["chunk_count"] == 0]
extraction_audit_hash = stable_id(EXTRACTOR_REVISION, AUDIT_MANIFEST_HASH, json.dumps(extraction_audit, ensure_ascii=False, sort_keys=True, default=str))
healthy_factual_documents = sorted(set(factual_documents) - {row["document_uid"] for row in missing_factual_documents})
factual_document_coverage = len(healthy_factual_documents) / max(1, len(factual_documents))
factual_availability_gate = {"healthy_count": len(healthy_factual_documents), "expected_count": len(factual_documents), "coverage": factual_document_coverage, "minimum_healthy_count": int(CONFIG["minimum_healthy_factual_documents"]), "minimum_coverage": float(CONFIG["minimum_factual_document_coverage"]), "pass": len(healthy_factual_documents) >= int(CONFIG["minimum_healthy_factual_documents"]) and factual_document_coverage >= float(CONFIG["minimum_factual_document_coverage"])}
write_json(RUN / "artifacts" / "extraction_audit.json", {"extractor_revision": EXTRACTOR_REVISION, "extraction_audit_hash": extraction_audit_hash, "audit_manifest_hash": AUDIT_MANIFEST_HASH, "rows": extraction_audit, "expected_factual_documents": sorted(factual_documents), "extracted_factual_documents": healthy_factual_documents, "missing_factual_documents": missing_factual_documents, "blocked_documents": [row for row in extraction_audit if row["status"] == "blocked"], "factual_availability_gate": factual_availability_gate})
if not factual_availability_gate["pass"]:
    raise RuntimeError("insufficient_factual_corpus:" + json.dumps({"factual_availability_gate": factual_availability_gate, "missing_factual_documents": missing_factual_documents, "message": "Too few locally verified factual documents remain after fail-closed extraction."}, ensure_ascii=False, sort_keys=True))
if chunks.empty: raise RuntimeError("no_indexable_chunks")
assert chunks.chunk_id.is_unique and chunks.chunk_document_uid.notna().all(); CHUNK_MANIFEST_HASH = stable_id(json.dumps(chunks[["document_uid", "chunk_id", "text_sha256"]].to_dict("records"), sort_keys=True))
EXPECTED_EXTRACTION_UNIVERSE = sorted(
    [
        {"document_uid": str(row["document_uid"]), "chunk_id": str(row["chunk_id"]), "text_sha256": str(row["text_sha256"])}
        for row in chunks[chunks.factual_index_allowed].to_dict("records")
    ],
    key=lambda row: (row["document_uid"], row["chunk_id"], row["text_sha256"]),
)
EXPECTED_EXTRACTION_UNIVERSE_HASH = stable_id(json.dumps(EXPECTED_EXTRACTION_UNIVERSE, ensure_ascii=False, sort_keys=True))
EXPECTED_MENTION_UNIVERSE = sorted(
    [
        {"document_uid": str(row["document_uid"]), "chunk_id": str(row["chunk_id"]), "text_sha256": str(row["text_sha256"]), "text": str(row["text"])}
        for row in chunks[chunks.factual_index_allowed].to_dict("records")
    ],
    key=lambda row: (row["document_uid"], row["chunk_id"], row["text_sha256"]),
)
EXPECTED_MENTION_UNIVERSE_HASH = stable_id(json.dumps([{key: row[key] for key in ("document_uid", "chunk_id", "text_sha256")} for row in EXPECTED_MENTION_UNIVERSE], ensure_ascii=False, sort_keys=True))
parquet_safe_frame(pages).to_parquet(RUN / "artifacts" / "pages.parquet", index=False); parquet_safe_frame(chunks).to_parquet(RUN / "artifacts" / "chunks.parquet", index=False)
write_json(RUN / "artifacts" / "extraction_manifest.json", {"identity": {"corpus_manifest_hash": CORPUS_MANIFEST_HASH, "chunk_manifest_hash": CHUNK_MANIFEST_HASH, "extraction_audit_hash": extraction_audit_hash, "extractor_revision": EXTRACTOR_REVISION, "core_source_sha256": CORE_SOURCE_SHA256, "schema_revision": SEMANTIC_EXTRACTION_SCHEMA["version"]}})



# %% [notebook cell 8]

# 06 - One Qwen lifecycle for validated semantic extraction and full query signatures.
from unsloth import FastModel
import transformers
import transformers.tokenization_utils as transformers_tokenization_utils
if not hasattr(transformers_tokenization_utils, "PreTrainedTokenizerBase"):
    transformers_tokenization_utils.PreTrainedTokenizerBase = transformers.PreTrainedTokenizerBase
from lmformatenforcer import JsonSchemaParser
from lmformatenforcer.integrations.transformers import build_transformers_prefix_allowed_tokens_fn
def verify_effective_4bit(model):
    diagnostics = {"requested": bool(CONFIG["load_in_4bit"]), "is_loaded_in_4bit": bool(getattr(model, "is_loaded_in_4bit", False)), "quantization_config": str(getattr(getattr(model, "config", None), "quantization_config", None)), "quantized_module_count": 0, "quantized_parameter_count": 0, "float32_parameter_count": 0, "total_parameter_count": 0, "float32_parameter_fraction": 0.0}
    for module in model.modules():
        name = type(module).__name__.casefold()
        if "4bit" in name or "bitsandbytes" in name or "bnb" in name or "params4bit" in name: diagnostics["quantized_module_count"] += 1
    for parameter in model.parameters():
        count = int(parameter.numel()) if hasattr(parameter, "numel") else 1; diagnostics["total_parameter_count"] += count
        parameter_type = type(parameter).__name__.casefold()
        if "params4bit" in parameter_type or "4bit" in parameter_type or "int8params" in parameter_type: diagnostics["quantized_parameter_count"] += count
        elif str(getattr(parameter, "dtype", "")).casefold().endswith("float32"): diagnostics["float32_parameter_count"] += count
    if diagnostics["total_parameter_count"]: diagnostics["float32_parameter_fraction"] = diagnostics["float32_parameter_count"] / diagnostics["total_parameter_count"]
    # Metadata flags/config strings are diagnostics only. Require inspected
    # quantized modules and reject a meaningful all-fp32 fallback; small fp32
    # norms/heads are allowed.
    diagnostics["effective"] = bool(diagnostics["quantized_module_count"] > 0 and diagnostics["float32_parameter_fraction"] <= 0.25)
    if diagnostics["requested"] and not diagnostics["effective"]: raise RuntimeError("effective_4bit_verification_failed:" + json.dumps(diagnostics, sort_keys=True))
    return diagnostics
def load_qwen_for_extraction():
    global model, tokenizer
    model, tokenizer = FastModel.from_pretrained(model_name=CONFIG["generator_model"], revision=CONFIG["generator_model_revision"], max_seq_length=CONFIG["max_seq_length"], load_in_4bit=CONFIG["load_in_4bit"], dtype=None, full_finetuning=False); verify_effective_4bit(model); tokenizer.pad_token = tokenizer.pad_token or tokenizer.eos_token; tokenizer.padding_side = "left"; model.eval(); MEMORY_SNAPSHOTS.append(gpu_snapshot("qwen_extraction_loaded")); return model, tokenizer
def unload_generator():
    global model, tokenizer
    if "model" in globals(): del model
    if "tokenizer" in globals(): del tokenizer
    gc.collect(); torch.cuda.empty_cache(); MEMORY_SNAPSHOTS.append(gpu_snapshot("qwen_extraction_unloaded"))
def split_visible_thinking(raw):
    splitter = globals().get("split_qwen_thinking")
    if callable(splitter): return splitter(raw)["final_content"]
    value = str(raw or ""); match = re.search(r"<think>.*?</think>", value, flags=re.S | re.I)
    return (value[0 : match.start()] + value[match.end() :]).strip() if match else value.strip()
def qwen_generation_trace(raw):
    trace = dict(split_qwen_thinking(raw)); base_tokenizer = getattr(globals().get("tokenizer"), "tokenizer", globals().get("tokenizer"))
    def token_count(value):
        if base_tokenizer is None or not hasattr(base_tokenizer, "encode"): return None
        return len(base_tokenizer.encode(str(value or ""), add_special_tokens=False))
    trace["reasoning_token_count"] = token_count(trace["reasoning_content"]); trace["answer_token_count"] = token_count(trace["final_content"]); trace["raw_generation"] = str(raw or ""); return trace
def _apply_chat_template(messages, enable_thinking):
    try:
        return tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True, enable_thinking=enable_thinking)
    except TypeError as exc:
        log_event("generation", "chat_template_compatibility_fallback", error=str(exc))
        return tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
class GenerationRowQuarantine(RuntimeError):
    def __init__(self, audit):
        self.audit = dict(audit or {})
        self.reason = str(self.audit.get('reason', 'row_quarantine'))
        super().__init__(self.reason)
class PromptBudgetQuarantine(GenerationRowQuarantine):
    def __init__(self, audit):
        super().__init__(audit)
        self.args = (f"prompt_token_budget_exceeded:prompt_budget_irreducible:{self.reason}",)
class ParseRateQuarantine(GenerationRowQuarantine):
    pass
class SchemaValidationQuarantine(GenerationRowQuarantine):
    pass
def preflight_prompt_token_budget(prompts, max_new_tokens, schema_tail=None):
    if callable(globals().get("fit_prompt_to_budget")):
        for prompt in prompts:
            fit_prompt_to_budget(prompt, tokenizer, int(CONFIG["max_seq_length"]), schema_tail=schema_tail, reserve_output_tokens=int(max_new_tokens))
    encoded = tokenizer(text=prompts, return_tensors="pt", padding=True, truncation=False)
    mask = encoded["attention_mask"]; lengths = mask.sum(dim=1).tolist() if hasattr(mask, "sum") else [sum(int(value) for value in row) for row in mask]
    violations = [{"index": i, "prompt_tokens": int(length), "max_new_tokens": int(max_new_tokens), "total_tokens": int(length) + int(max_new_tokens), "max_seq_length": int(CONFIG["max_seq_length"])} for i, length in enumerate(lengths) if int(length) + int(max_new_tokens) > int(CONFIG["max_seq_length"])]
    if violations:
        audit = {"reason": "prompt_budget_irreducible", "violations": violations, "dropped_evidence_ids": [], "attempts": []}
        raise PromptBudgetQuarantine(audit) if "PromptBudgetQuarantine" in globals() else RuntimeError("prompt_token_budget_exceeded:" + json.dumps(violations, sort_keys=True))
    return encoded, [int(length) for length in lengths]
def _generate_prompt_batch(prompts, max_new_tokens, temperature=0.0, output_schema=None):
    schema_tail = json.dumps(output_schema, ensure_ascii=False, sort_keys=True) if output_schema is not None else None
    if schema_tail and all(schema_tail in prompt for prompt in prompts): schema_tail = None
    encoded, prompt_lengths = preflight_prompt_token_budget(prompts, max_new_tokens, schema_tail=schema_tail); encoded = encoded.to(model.device); base_tokenizer = getattr(tokenizer, "tokenizer", tokenizer)
    generation_kwargs = {"max_new_tokens": max_new_tokens, "do_sample": temperature > 0, "temperature": max(temperature, 1e-5), "pad_token_id": getattr(base_tokenizer, "pad_token_id", getattr(tokenizer, "pad_token_id", None))}
    if output_schema is not None:
        parser = JsonSchemaParser(output_schema); generation_kwargs["prefix_allowed_tokens_fn"] = build_transformers_prefix_allowed_tokens_fn(base_tokenizer, parser)
    with torch.inference_mode(): output = model.generate(**encoded, **generation_kwargs)
    n = encoded["input_ids"].shape[1]; return [base_tokenizer.decode(row[n:], skip_special_tokens=True) for row in output]
def generate_batch(messages_batch, max_new_tokens, temperature=0.0, enable_thinking=False, output_schema=None):
    if not enable_thinking:
        prompts = [_apply_chat_template(x, enable_thinking=False) for x in messages_batch]
        return _generate_prompt_batch(prompts, max_new_tokens, temperature=temperature, output_schema=output_schema)
    reasoning_prompts = [_apply_chat_template(x, enable_thinking=True) for x in messages_batch]
    reasoning_outputs = _generate_prompt_batch(reasoning_prompts, int(CONFIG["reasoning_max_new_tokens"]), temperature=temperature)
    final_messages = []
    for messages, reasoning in zip(messages_batch, reasoning_outputs):
        final_messages.append(list(messages) + [{"role": "user", "content": f"Prior model deliberation (use as private working context):\n{reasoning.strip()}\nNow return only the requested final JSON object."}])
    final_prompts = [_apply_chat_template(x, enable_thinking=False) for x in final_messages]
    final_outputs = _generate_prompt_batch(final_prompts, max_new_tokens, temperature=temperature, output_schema=output_schema)
    return [f"<think>\n{reasoning.strip()}\n</think>\n\n{final.strip()}" for reasoning, final in zip(reasoning_outputs, final_outputs)]
def _repair_payload_candidates(raw, prompt):
    raw_text = split_visible_thinking(raw)
    try:
        semantic_raw = json.dumps(json.loads(raw_text), ensure_ascii=False, sort_keys=True)
    except Exception:
        semantic_raw = raw_text
    full = json.dumps({"raw_output": semantic_raw, "schema_task": str(prompt)}, ensure_ascii=False, sort_keys=True)
    # The compact candidate must retain the original model output. A status-only
    # payload would permit the repair model to invent an unrelated object.
    compact = json.dumps({"raw_output": semantic_raw, "schema_task": "Repair this prior output to the requested JSON schema without adding commentary."}, ensure_ascii=False, sort_keys=True)
    return (full, compact)
def _repair_prompt_token_count(prompt):
    encoded = tokenizer(text=[str(prompt)], return_tensors="pt", padding=False, truncation=False)
    mask = encoded["attention_mask"]
    return int(mask.sum().item()) if hasattr(mask.sum(), "item") else int(mask.sum())
def _adaptive_repair_messages(raw, prompt, output_schema):
    payload_candidates = _repair_payload_candidates(raw, prompt)
    schema_tail = "JSON schema (must be followed exactly): " + json.dumps(output_schema, ensure_ascii=False, sort_keys=True)
    def builder(selected, payload):
        return [{"role": "system", "content": "Return only valid JSON matching the requested schema."}, {"role": "user", "content": f"Repair context: {payload}\n{schema_tail}"}]
    def rendered_builder(selected, payload):
        return _apply_chat_template(builder(selected, payload), enable_thinking=False)
    fit = fit_adaptive_prompt_with_evidence([], prompt_builder=rendered_builder, payload_candidates=payload_candidates, token_counter=_repair_prompt_token_count, max_input_tokens=int(CONFIG["max_seq_length"]), reserve_output_tokens=512)
    if fit.get("status") != "fit":
        audit = dict(fit); audit["reason"] = "repair_prompt_budget"; raise PromptBudgetQuarantine(audit)
    return builder([], fit.get("payload", "")), fit
def parse_json_object(raw):
    value = split_visible_thinking(raw); start, end = value.find("{"), value.rfind("}")
    if start < 0 or end <= start: raise ValueError("invalid_json")
    return json.loads(value[start:end + 1])
def repair_json_output(raw, prompt, output_schema):
    adaptive_builder = globals().get("_adaptive_repair_messages")
    if callable(adaptive_builder):
        messages, _ = adaptive_builder(raw, prompt, output_schema)
    else:
        # Keep lightweight extraction/probe namespaces compatible while still
        # retaining the invalid output and immutable schema tail.
        schema_tail = "JSON schema (must be followed exactly): " + json.dumps(output_schema, ensure_ascii=False, sort_keys=True)
        repair_context = json.dumps({"raw_output": split_visible_thinking(raw), "schema_task": str(prompt)}, ensure_ascii=False, sort_keys=True)
        messages = [{"role": "system", "content": "Return only valid JSON matching the requested schema."}, {"role": "user", "content": f"Repair context: {repair_context}\n{schema_tail}"}]
    return generate_batch([messages], 512, enable_thinking=False, output_schema=output_schema)[0]
def parse_with_one_repair(raw, prompt, validator=None, output_schema=None):
    try:
        parsed = parse_json_object(raw)
        if validator is not None and not validator(parsed)["valid"]: raise ValueError("schema_invalid")
        return parsed, raw, "initial"
    except Exception:
        repaired = repair_json_output(raw, prompt, output_schema)
        try:
            parsed = parse_json_object(repaired)
            if validator is not None and not validator(parsed)["valid"]: raise ValueError("schema_invalid")
            original_trace = split_qwen_thinking(raw); repaired_final = split_qwen_thinking(repaired)["final_content"]
            preserved = (f"<think>\n{original_trace['reasoning_content']}\n</think>\n\n" + repaired_final) if original_trace["reasoning_content"] else repaired_final
            return parsed, preserved, "repair"
        except Exception: return None, repaired, "schema_invalid"
QUERY_SCHEMA_REVISION = "query-signature.v2"
# The derived cache remains validated against QUERY_SIGNATURE_SCHEMA; model decoding uses QUERY_MODEL_SCHEMA (output_schema=QUERY_SIGNATURE_SCHEMA is reserved for the derived cache contract).
ACCEPTED_PARSE_STATUSES = frozenset({"initial", "repair"})
def canonical_extraction_chunks(rows, expected_universe):
    expected_order = sorted(
        [(str(item["document_uid"]), str(item["chunk_id"]), str(item["text_sha256"])) for item in expected_universe],
        key=lambda key: key,
    )
    ordered = sorted(
        rows,
        key=lambda row: (str(row["document_uid"]), str(row["chunk_id"]), str(row["text_sha256"])),
    )
    actual_order = [(str(row.get("document_uid")), str(row.get("chunk_id")), str(row.get("text_sha256"))) for row in ordered]
    if actual_order != expected_order or len(actual_order) != len(set(actual_order)): raise RuntimeError("extraction_universe_mismatch")
    return ordered
def parse_extraction_with_one_repair(raw, prompt, source_text, context):
    """Parse and semantically validate extraction with one total repair budget."""
    try:
        payload = parse_json_object(raw)
    except Exception:
        repaired = repair_json_output(raw, prompt, SEMANTIC_EXTRACTION_SCHEMA)
        try:
            repaired_payload = parse_json_object(repaired)
        except Exception:
            return None, repaired, "schema_invalid", validate_extraction(None, source_text, context)
        repaired_validation = validate_extraction(repaired_payload, source_text, context)
        if repaired_validation.get("status") == "quarantined" or repaired_validation.get("quarantined"):
            return repaired_payload, repaired, "semantic_invalid", repaired_validation
        return repaired_payload, repaired, "repair", repaired_validation
    validation = validate_extraction(payload, source_text, context)
    if validation.get("status") == "quarantined" or validation.get("quarantined"):
        reasons = sorted({reason for item in validation.get("quarantined", []) for reason in item.get("reasons", [])})
        semantic_repair_prompt = f"{prompt}\nSemantic validation reasons: {json.dumps(reasons, ensure_ascii=False)}\nExact source text: {source_text}\nSource context: {json.dumps(context, ensure_ascii=False, sort_keys=True)}"
        repaired = repair_json_output(raw, semantic_repair_prompt, SEMANTIC_EXTRACTION_SCHEMA)
        try:
            repaired_payload = parse_json_object(repaired)
        except Exception:
            return None, repaired, "semantic_invalid", validate_extraction(None, source_text, context)
        repaired_validation = validate_extraction(repaired_payload, source_text, context)
        if repaired_validation.get("status") == "quarantined" or repaired_validation.get("quarantined"):
            return repaired_payload, repaired, "semantic_invalid", repaired_validation
        return repaired_payload, repaired, "repair", repaired_validation
    return payload, raw, "initial", validation
def parse_mentions_with_one_repair(raw, prompt, source_text, context):
    parsed, final_raw, status = parse_with_one_repair(raw, prompt, None, MENTION_DISCOVERY_SCHEMA)
    validation = validate_mentions(parsed, source_text, context) if parsed is not None else validate_mentions(None, source_text, context)
    if validation.get("status") != "accepted" or validation.get("quarantined"): return [], final_raw, "semantic_invalid"
    return validation.get("accepted", []), final_raw, status
def extraction_parse_status(parse_status, validation):
    if parse_status in ACCEPTED_PARSE_STATUSES and isinstance(validation, dict) and (validation.get("status") == "quarantined" or validation.get("quarantined")):
        return "semantic_invalid"
    return parse_status
def extraction_parse_accepted(parse_status, validation):
    return parse_status in ACCEPTED_PARSE_STATUSES and isinstance(validation, dict) and validation.get("status") != "quarantined" and not validation.get("quarantined")
PERSPECTIVE_SCHEMA = {"type": "object", "additionalProperties": False, "required": ["perspective", "rationale", "claims_to_address", "supported_evidence_ids", "response_guidance", "risk_flags", "confidence"], "properties": {"perspective": {"type": "string"}, "rationale": {"type": "string"}, "claims_to_address": {"type": "array", "items": {"type": "string"}}, "supported_evidence_ids": {"type": "array", "items": {"type": "string", "pattern": "^E[1-9][0-9]*$"}}, "response_guidance": {"type": "array", "items": {"type": "string"}}, "risk_flags": {"type": "array", "items": {"type": "string"}}, "confidence": {"type": "number", "minimum": 0, "maximum": 1}}}
QUERY_SIGNATURE_SCHEMA = {"type": "object", "additionalProperties": False, "required": ["entity_ids", "predicates", "polarities", "modalities", "desired_stances"], "properties": {key: {"type": "array", "items": {"type": "string"}} for key in ["entity_ids", "predicates", "polarities", "modalities", "desired_stances"]}}
QUERY_MODEL_SCHEMA = {"type": "object", "additionalProperties": False, "required": ["entity_candidate_indices", "predicates", "polarities", "modalities", "desired_stances"], "properties": {"entity_candidate_indices": {"type": "array", "items": {"type": "object", "additionalProperties": False, "required": ["mention_id", "candidate_index"], "properties": {"mention_id": {"type": "string"}, "candidate_index": {"type": ["integer", "null"], "minimum": 0}}}}, **{key: {"type": "array", "items": {"type": "string"}} for key in ["predicates", "polarities", "modalities", "desired_stances"]}}}
PLAN_SCHEMA = {"type": "object", "additionalProperties": False, "required": ["claim_focus", "selected_evidence_ids", "response_steps", "tone", "factual_constraints", "safety_constraints"], "properties": {"claim_focus": {"type": "string"}, "selected_evidence_ids": {"type": "array", "items": {"type": "string", "pattern": "^E[1-9][0-9]*$"}}, "response_steps": {"type": "array", "items": {"type": "string"}}, "tone": {"type": "string"}, "factual_constraints": {"type": "array", "items": {"type": "string"}}, "safety_constraints": {"type": "array", "items": {"type": "string"}}}}
FINAL_RESPONSE_SCHEMA = {"type": "object", "additionalProperties": False, "required": ["counter_narrative", "cited_evidence_ids", "factual_claims", "safety_notes"], "properties": {"counter_narrative": {"type": "string"}, "cited_evidence_ids": {"type": "array", "items": {"type": "string", "pattern": "^E[1-9][0-9]*$"}}, "factual_claims": {"type": "array", "items": {"type": "string"}}, "safety_notes": {"type": "array", "items": {"type": "string"}}}}
INTERNAL_RESPONSE_METADATA_KEYS = frozenset({"parse_status", "quarantine", "schema_errors", "few_shot", "few_shot_prompt_revision"})
def _strict_payload(payload, schema, label, ledger_ids=None):
    reasons = []
    if not isinstance(payload, dict): return {"valid": False, "reasons": [f"{label}_not_object"], "quarantine": True}
    required, allowed = set(schema["required"]), set(schema["properties"]); reasons.extend(f"missing_{key}" for key in sorted(required - set(payload))); reasons.extend(["additionalProperties"] if set(payload) - allowed else [])
    for key, spec in schema["properties"].items():
        if key not in payload: continue
        value = payload[key]; kind = spec["type"]
        if kind == "string" and not isinstance(value, str): reasons.append(f"{key}_not_string")
        if kind == "array" and (not isinstance(value, list) or not all(isinstance(item, str) for item in value)): reasons.append(f"{key}_not_string_array")
        if kind == "number" and (isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(value) or value < spec.get("minimum", -math.inf) or value > spec.get("maximum", math.inf)): reasons.append(f"{key}_out_of_range")
    for key in ["supported_evidence_ids", "selected_evidence_ids", "cited_evidence_ids"]:
        values = payload.get(key)
        if not isinstance(values, list): continue
        if any(not isinstance(value, str) for value in values):
            reasons.append(f"invalid_{key}")
            continue
        if len(values) != len(set(values)): reasons.append(f"duplicate_{key}")
        if any(not re.fullmatch(r"E[1-9][0-9]*", value) for value in values): reasons.append(f"invalid_{key}")
        if ledger_ids is not None and set(values) - set(ledger_ids): reasons.append("unknown_evidence_ids")
    return {"valid": not reasons, "reasons": sorted(set(reasons)), "quarantine": bool(reasons)}
def validate_query_signature(payload): return _strict_payload(payload, QUERY_SIGNATURE_SCHEMA, "query_signature")
def validate_query_model_signature(payload):
    if not isinstance(payload, dict) or set(payload) != set(QUERY_MODEL_SCHEMA["properties"]): return {"valid": False, "reasons": ["query_model_schema_invalid"]}
    if not isinstance(payload.get("entity_candidate_indices"), list) or any(not isinstance(item, dict) or set(item) != {"mention_id", "candidate_index"} or not isinstance(item.get("mention_id"), str) or (item.get("candidate_index") is not None and (isinstance(item.get("candidate_index"), bool) or not isinstance(item.get("candidate_index"), int) or item.get("candidate_index") < 0)) for item in payload["entity_candidate_indices"]): return {"valid": False, "reasons": ["query_candidate_schema_invalid"]}
    return _strict_payload({key: payload[key] for key in ["predicates", "polarities", "modalities", "desired_stances"]}, {"type": "object", "additionalProperties": False, "required": ["predicates", "polarities", "modalities", "desired_stances"], "properties": {key: {"type": "array", "items": {"type": "string"}} for key in ["predicates", "polarities", "modalities", "desired_stances"]}}, "query_model_signature")
def validate_perspective(payload, ledger_ids=None, expected_perspective=None):
    result = _strict_payload(payload, PERSPECTIVE_SCHEMA, "perspective", ledger_ids)
    if result["valid"] and expected_perspective is not None and payload.get("perspective") != expected_perspective:
        result["valid"] = False; result["quarantine"] = True; result["reasons"].append("perspective_name_mismatch")
    return result
def validate_plan(payload, ledger_ids=None): return _strict_payload(payload, PLAN_SCHEMA, "plan", ledger_ids)
def validate_final_response(payload, ledger_ids=None): return _strict_payload(payload, FINAL_RESPONSE_SCHEMA, "final_response", ledger_ids)
def validate_structured_output(payload, schema=PERSPECTIVE_SCHEMA): return _strict_payload(payload, schema, "structured_output")["valid"]
def build_query_signature(record):
    target_candidates = build_entity_candidates(str(record.get("Target", "")), ENTITY_CATALOG, namespace_preference="corpus", allow_target_fallback=True)
    candidate_context = {"target": target_candidates, "candidate_set_hash": target_candidates["candidate_set_hash"], "catalog_manifest_hash": CATALOG_MANIFEST_HASH}
    return [{"role": "system", "content": "Return JSON only with entity_candidate_indices, predicates, polarities, modalities, desired_stances. Never emit entity IDs; select candidate indices or null."}, {"role": "user", "content": f"Target: {record['Target']}\nPost: {record['Text']}\nCandidate context: {json.dumps(candidate_context, ensure_ascii=False, sort_keys=True)}"}]
MENTION_DISCOVERY_CACHE = RUN / "artifacts" / "mention_discovery.jsonl"; MENTION_DISCOVERY_IDENTITY_PATH = RUN / "artifacts" / "mention_discovery_identity.json"; EXTRACTION_CACHE = RUN / "artifacts" / "validated_extractions.jsonl"; QUERY_SIGNATURE_CACHE = RUN / "artifacts" / "query_signatures.jsonl"
MENTION_DISCOVERY_PROMPT_REVISION = "mention-discovery.v1-exact-spans"
def mention_discovery_identity(expected_universe):
    ordered = [{key: row[key] for key in ("document_uid", "chunk_id", "text_sha256")} for row in expected_universe]
    return {"stage": "mention-discovery", "prompt_revision": MENTION_DISCOVERY_PROMPT_REVISION, "schema_revision": MENTION_DISCOVERY_SCHEMA["version"], "model": EXTRACTION_MODEL, "corpus_manifest_hash": CORPUS_MANIFEST_HASH, "expected_mention_universe": ordered, "expected_mention_universe_hash": stable_id("mention-universe.v1", json.dumps(ordered, ensure_ascii=False, sort_keys=True)), "expected_mention_universe_size": len(ordered), "core_source_sha256": CORE_SOURCE_SHA256}
def verify_mention_cache(rows, expected_universe, expected_identity):
    expected_identity = dict(expected_identity or {}); expected = mention_discovery_identity(expected_universe)
    if expected_identity != expected: return False
    expected_order = [(str(row["document_uid"]), str(row["chunk_id"]), str(row["text_sha256"])) for row in expected_universe]
    required = {"document_uid", "chunk_id", "text_sha256", "text", "mentions", "parse_status", "raw_output"}
    if not isinstance(rows, list) or len(rows) != len(expected_order): return False
    for index, row in enumerate(rows):
        if not isinstance(row, dict) or set(row) != required: return False
        source = expected_universe[index]; identity = (str(row.get("document_uid")), str(row.get("chunk_id")), str(row.get("text_sha256")))
        if identity != expected_order[index] or row.get("text") != source.get("text") or hashlib.sha256(row["text"].encode()).hexdigest() != row["text_sha256"]: return False
        if row.get("parse_status") not in {"initial", "repair"} or not isinstance(row.get("raw_output"), str) or not isinstance(row.get("mentions"), list): return False
        validation = validate_mentions({"schema_version": MENTION_DISCOVERY_SCHEMA["version"], "mentions": row["mentions"]}, row["text"], {"document_uid": row["document_uid"], "chunk_id": row["chunk_id"], "text_sha256": row["text_sha256"]})
        if validation.get("status") != "accepted" or validation.get("quarantined"): return False
    return True
def mention_records_from_cache(rows):
    return [{**mention, "document_uid": row["document_uid"], "chunk_id": row["chunk_id"]} for row in rows for mention in row["mentions"]]
def verify_extraction_cache(rows, expected_universe):
    expected_order = [
        (str(item.get("document_uid")), str(item.get("chunk_id")), str(item.get("text_sha256")))
        for item in expected_universe
        if isinstance(item, dict)
    ]
    expected = set(expected_order)
    if not isinstance(rows, list) or len(rows) != len(expected) or len(expected) != len(expected_order): return False
    seen = set()
    allowed_statuses = {"initial", "repair", "semantic_invalid", "schema_invalid"}
    allowed_validation_statuses = {"accepted", "partial", "reviewed", "quarantined"}
    catalog_bound = "CATALOG_MANIFEST_HASH" in globals()
    required_row_keys = {"document_uid", "chunk_id", "text_sha256", "validation", "parse_status", "raw_output"} | ({"catalog_manifest_hash"} if catalog_bound else set())
    for index, row in enumerate(rows):
        if not isinstance(row, dict) or set(row) != required_row_keys: return False
        if not all(isinstance(row.get(key), str) and row.get(key) for key in ("document_uid", "chunk_id", "text_sha256", "raw_output")): return False
        if catalog_bound and row.get("catalog_manifest_hash") != CATALOG_MANIFEST_HASH: return False
        identity = (row["document_uid"], row["chunk_id"], row["text_sha256"])
        if identity in seen or identity not in expected or identity != expected_order[index] or row.get("parse_status") not in allowed_statuses: return False
        validation = row.get("validation")
        if not isinstance(validation, dict) or validation.get("status") not in allowed_validation_statuses or not isinstance(validation.get("quarantined", []), list): return False
        source_context = validation.get("source_context")
        if not isinstance(source_context, dict) or any(source_context.get(key) != row[key] for key in ("document_uid", "chunk_id", "text_sha256")): return False
        if not isinstance(validation.get("text"), str) or hashlib.sha256(validation["text"].encode()).hexdigest() != row["text_sha256"]: return False
        if validation.get("validation_marker") != "mpkg-rag.validated-extraction.v1" or not isinstance(validation.get("validation_fingerprint"), str) or not validation.get("validation_fingerprint") or validation.get("validation_fingerprint") != _validation_fingerprint(validation): return False
        seen.add(identity)
    return seen == expected
mention_identity = mention_discovery_identity(EXPECTED_MENTION_UNIVERSE); saved_mention_identity = json.loads(MENTION_DISCOVERY_IDENTITY_PATH.read_text()) if MENTION_DISCOVERY_IDENTITY_PATH.exists() else {}; mention_rows = load_jsonl(MENTION_DISCOVERY_CACHE); mention_cache_complete = verify_mention_cache(mention_rows, EXPECTED_MENTION_UNIVERSE, saved_mention_identity.get("identity")); need_mentions = not (saved_mention_identity.get("identity") == mention_identity and mention_cache_complete); model_loaded = False
if need_mentions:
    load_qwen_for_extraction(); model_loaded = True
    extraction_chunks = canonical_extraction_chunks(chunks[chunks.factual_index_allowed].to_dict("records"), EXPECTED_EXTRACTION_UNIVERSE)
    with MENTION_DISCOVERY_CACHE.open("w", encoding="utf-8") as stream:
        for start in tqdm(range(0, len(extraction_chunks), CONFIG["extraction_batch_size"]), desc="Qwen mention discovery"):
            batch = extraction_chunks[start:start + CONFIG["extraction_batch_size"]]; base_contexts = [{k: chunk.get(k) for k in ["document_uid", "chunk_id", "source_id", "source_type", "authority_score", "factual_index_allowed", "status", "content_sha256", "document_sha256", "text_sha256"]} for chunk in batch]; mention_prompts = [build_mention_prompt(chunk["text"], context) for chunk, context in zip(batch, base_contexts)]; mention_raw = generate_batch(mention_prompts, 384, enable_thinking=False, output_schema=MENTION_DISCOVERY_SCHEMA)
            for chunk, context, mention_prompt, raw_mention in zip(batch, base_contexts, mention_prompts, mention_raw):
                mentions, mention_final_raw, mention_status = parse_mentions_with_one_repair(raw_mention, mention_prompt[1]["content"], chunk["text"], context); stream.write(json.dumps({"document_uid": chunk["document_uid"], "chunk_id": chunk["chunk_id"], "text_sha256": chunk["text_sha256"], "text": chunk["text"], "mentions": mentions, "parse_status": mention_status, "raw_output": mention_final_raw}, ensure_ascii=False) + "\n")
    mention_rows = load_jsonl(MENTION_DISCOVERY_CACHE); mention_cache_complete = verify_mention_cache(mention_rows, EXPECTED_MENTION_UNIVERSE, mention_identity); write_json_atomic(MENTION_DISCOVERY_IDENTITY_PATH, {"identity": mention_identity, "cache_state": "complete_ready" if mention_cache_complete else "complete_validation_failed"})
if not mention_cache_complete: raise RuntimeError("mention_cache_identity_mismatch")
mention_records = mention_records_from_cache(mention_rows)
ENTITY_CATALOG = build_entity_catalog(source_registry.to_dict("records"), dataset["Target"].astype(str).tolist(), reviewed_orgs, mention_records); CATALOG_MANIFEST_HASH = ENTITY_CATALOG["catalog_hash"]
write_json(RUN / "artifacts" / "entity_catalog.json", ENTITY_CATALOG); log_event("catalog", "final_catalog_frozen", catalog_manifest_hash=CATALOG_MANIFEST_HASH, mention_identity=mention_identity)
identity = {"corpus_manifest_hash": CORPUS_MANIFEST_HASH, "chunk_manifest_hash": CHUNK_MANIFEST_HASH, "extraction_audit_hash": extraction_audit_hash, "catalog_manifest_hash": CATALOG_MANIFEST_HASH, "mention_identity": mention_identity, "expected_extraction_universe_hash": EXPECTED_EXTRACTION_UNIVERSE_HASH, "expected_extraction_universe_size": len(EXPECTED_EXTRACTION_UNIVERSE), "extraction_model": EXTRACTION_MODEL, "extraction_prompt_revision": EXTRACTION_PROMPT_REVISION, "core_source_sha256": CORE_SOURCE_SHA256, "schema_revision": SEMANTIC_EXTRACTION_SCHEMA["version"]}; identity_path = RUN / "artifacts" / "extraction_identity.json"; saved = json.loads(identity_path.read_text()) if identity_path.exists() else {}
def record_identity(record):
    return {"record_id": str(record["ID"]), "input_text_sha256": str(record["input_text_sha256"]), "target_sha256": hashlib.sha256(str(record.get("Target", "")).encode()).hexdigest(), "category_sha256": hashlib.sha256(str(record.get("Category", "")).encode()).hexdigest()}
record_identities = sorted([record_identity(record) for record in dataset.to_dict("records")], key=lambda row: row["record_id"])
query_identity = {"record_identities": record_identities, "model": EXTRACTION_MODEL, "prompt_revision": EXTRACTION_PROMPT_REVISION, "core_source_sha256": CORE_SOURCE_SHA256, "schema_revision": QUERY_SCHEMA_REVISION, "catalog_manifest_hash": CATALOG_MANIFEST_HASH, "mention_identity": mention_identity}; extraction_rows = load_jsonl(EXTRACTION_CACHE); extraction_cache_complete = verify_extraction_cache(extraction_rows, EXPECTED_EXTRACTION_UNIVERSE)
def cache_reuse_state(saved, expected_identity, expected_query_identity, expected_universe, extraction_cache_complete, signature_cache_complete, mention_cache_complete=True, mention_identity_matches=True):
    identity_matches = saved.get("identity") == expected_identity and saved.get("query_identity") == expected_query_identity and saved.get("expected_extraction_universe") == expected_universe and saved.get("expected_extraction_universe_hash") == EXPECTED_EXTRACTION_UNIVERSE_HASH and mention_cache_complete and mention_identity_matches
    return (not (identity_matches and extraction_cache_complete), not (identity_matches and signature_cache_complete))
def extraction_identity_payload(cache_state, parse_rate, parse_status_counts):
    return {"identity": identity, "query_identity": query_identity, "mention_identity": mention_identity, "expected_extraction_universe": EXPECTED_EXTRACTION_UNIVERSE, "expected_extraction_universe_hash": EXPECTED_EXTRACTION_UNIVERSE_HASH, "cache_state": cache_state, "parse_rate": parse_rate, "parse_status_counts": parse_status_counts}
def persist_extraction_identity_state(cache_state, parse_rate, parse_status_counts):
    write_json_atomic(identity_path, extraction_identity_payload(cache_state, parse_rate, parse_status_counts))
reuse_state = cache_reuse_state(saved, identity, query_identity, EXPECTED_EXTRACTION_UNIVERSE, extraction_cache_complete, QUERY_SIGNATURE_CACHE.exists() and saved.get("query_identity") == query_identity, mention_cache_complete=mention_cache_complete, mention_identity_matches=saved.get("mention_identity", saved_mention_identity.get("identity")) == mention_identity); need_extraction = reuse_state[0]; need_signatures = reuse_state[1]
def query_signature_entity_ids_usable(signature, catalog):
    if not isinstance(signature, dict) or not isinstance(catalog, dict): return {"valid": False, "linked_entity_ids": [], "target_anchor_ids": [], "invalid_entity_ids": ["<invalid>"]}
    by_id = {str(row.get("entity_id")): row for row in catalog.get("entities", []) if isinstance(row, dict) and row.get("entity_id")}
    entity_ids = signature.get("entity_ids")
    if not isinstance(entity_ids, list) or any(not isinstance(value, str) for value in entity_ids) or len(entity_ids) != len(set(entity_ids)): return {"valid": False, "linked_entity_ids": [], "target_anchor_ids": [], "invalid_entity_ids": ["<invalid_entity_ids>"]}
    invalid, linked, target = [], [], []
    for entity_id in entity_ids:
        row = by_id.get(entity_id)
        if row is None or row.get("retrieval_allowed") is not True or row.get("link_status") != "linked": invalid.append(entity_id); continue
        if row.get("namespace") == "target" or row.get("factual_identity_allowed") is not True: target.append(entity_id)
        else: linked.append(entity_id)
    return {"valid": not invalid, "linked_entity_ids": linked, "target_anchor_ids": target, "invalid_entity_ids": sorted(invalid)}
def verify_query_signature_cache(rows, catalog=None):
    expected_by_id = {identity["record_id"]: identity for identity in record_identities}; seen = set()
    if len(rows) != len(expected_by_id): return False
    for row in rows:
        row_id = str(row.get("ID", "")); embedded = row.get("record_identity"); expected = expected_by_id.get(row_id); catalog_bound = "CATALOG_MANIFEST_HASH" in globals()
        if row_id in seen or expected is None or embedded != expected or row.get("input_text_sha256") != expected["input_text_sha256"] or (catalog_bound and row.get("catalog_manifest_hash") != CATALOG_MANIFEST_HASH): return False
        if row.get("parse_status") not in ACCEPTED_PARSE_STATUSES or not validate_query_signature(row.get("query_signature"))["valid"]: return False
        if catalog is not None and not query_signature_entity_ids_usable(row.get("query_signature"), catalog)["valid"]: return False
        seen.add(row_id)
    return seen == set(expected_by_id)
def semantic_linkage_quality_gate(graph_tables, catalog, query_signatures, thresholds=None, smoke_test=False):
    limits = dict(thresholds or QUALITY_THRESHOLDS)
    if smoke_test:
        limits["minimum_graph_linked_claim_rate"] = min(float(limits.get("minimum_graph_linked_claim_rate", 0.02)), 0.10)
        limits["minimum_query_linked_entity_rate"] = min(float(limits.get("minimum_query_linked_entity_rate", 0.02)), 0.10)
    factual_ids = {str(row.get("entity_id")) for row in catalog.get("entities", []) if isinstance(row, dict) and row.get("namespace") == "corpus" and row.get("factual_identity_allowed") is True and row.get("retrieval_allowed") is True and row.get("link_status") == "linked"}
    accepted_claims = [claim for claim in graph_tables.get("Claim", []) if isinstance(claim, dict) and claim.get("review_status") == "accepted"]
    linked_claims = [claim for claim in accepted_claims if any(str(claim.get(key)) in factual_ids for key in ("subject_entity_id", "object_entity_id"))]
    linked_queries = 0; target_only = 0
    for signature in query_signatures or []:
        usable = query_signature_entity_ids_usable(signature, catalog)
        if usable["linked_entity_ids"]: linked_queries += 1
        elif usable["target_anchor_ids"] and usable["valid"]: target_only += 1
    graph_rate = len(linked_claims) / max(1, len(accepted_claims)); query_rate = linked_queries / max(1, len(query_signatures or []))
    result = {"pass": len(linked_claims) >= int(limits.get("minimum_linked_claims", 1)) and graph_rate >= float(limits.get("minimum_graph_linked_claim_rate", 0.02)) and linked_queries >= int(limits.get("minimum_linked_queries", 1)) and query_rate >= float(limits.get("minimum_query_linked_entity_rate", 0.02)), "accepted_claims": len(accepted_claims), "linked_claims": len(linked_claims), "graph_linked_claim_rate": graph_rate, "total_queries": len(query_signatures or []), "linked_queries": linked_queries, "linked_query_rate": query_rate, "target_anchor_only_signatures": target_only, "thresholds": limits, "scoring_calibration_status": SCORING_CALIBRATION_STATUS, "self_confidence_status": SELF_CONFIDENCE_STATUS}
    return result
if need_extraction or need_signatures:
    if not model_loaded: load_qwen_for_extraction(); model_loaded = True
    if need_extraction:
        with EXTRACTION_CACHE.open("w", encoding="utf-8") as stream:
            extraction_chunks = canonical_extraction_chunks(chunks[chunks.factual_index_allowed].to_dict("records"), EXPECTED_EXTRACTION_UNIVERSE)
            for start in tqdm(range(0, len(extraction_chunks), CONFIG["extraction_batch_size"]), desc="Qwen semantic extraction"):
                batch = extraction_chunks[start:start + CONFIG["extraction_batch_size"]]
                base_contexts = [{k: chunk.get(k) for k in ["document_uid", "chunk_id", "source_id", "source_type", "authority_score", "factual_index_allowed", "status", "content_sha256", "document_sha256", "text_sha256"]} for chunk in batch]
                candidate_contexts = []
                for chunk, base_context in zip(batch, base_contexts):
                    candidates = {mention["mention_id"]: build_entity_candidates(mention["text"], ENTITY_CATALOG, namespace_filter="corpus", factual_only=True) for mention in mention_records if mention.get("document_uid") == chunk["document_uid"] and mention.get("chunk_id") == chunk["chunk_id"]}
                    candidate_contexts.append({**base_context, "entity_catalog": ENTITY_CATALOG, "candidate_sets": candidates})
                prompts = [build_extraction_prompt(chunk["text"], context, "corpus") for chunk, context in zip(batch, candidate_contexts)]
                raw_outputs = generate_batch(prompts, CONFIG["perspective_max_new_tokens"], enable_thinking=False, output_schema=SEMANTIC_EXTRACTION_SCHEMA)
                for chunk, context, prompt, raw in zip(batch, candidate_contexts, prompts, raw_outputs):
                    payload, final_raw, parse_status, validation = parse_extraction_with_one_repair(raw, prompt[1]["content"], chunk["text"], context)
                    stream.write(json.dumps({"document_uid": chunk["document_uid"], "chunk_id": chunk["chunk_id"], "text_sha256": chunk["text_sha256"], "catalog_manifest_hash": CATALOG_MANIFEST_HASH, "validation": validation, "parse_status": parse_status, "raw_output": final_raw}, ensure_ascii=False) + "\n")
    if need_signatures:
        with QUERY_SIGNATURE_CACHE.open("w", encoding="utf-8") as stream:
            for record in tqdm(dataset.to_dict("records"), desc="Qwen query signatures"):
                prompt = build_query_signature(record); raw = generate_batch([prompt], 256, enable_thinking=False, output_schema=QUERY_MODEL_SCHEMA)[0]; parsed, final_raw, parse_status = parse_with_one_repair(raw, prompt[1]["content"], validate_query_model_signature, QUERY_MODEL_SCHEMA); target_candidates = build_entity_candidates(str(record.get("Target", "")), ENTITY_CATALOG, namespace_preference="corpus", allow_target_fallback=True); resolved = resolve_query_signature_entities(parsed, {"target": target_candidates}, ENTITY_CATALOG) if parsed is not None else {"valid": False}; signature = ({"entity_ids": resolved["entity_ids"], "predicates": parsed["predicates"], "polarities": parsed["polarities"], "modalities": parsed["modalities"], "desired_stances": parsed["desired_stances"]} if parse_status in ACCEPTED_PARSE_STATUSES and resolved.get("valid") else None); stream.write(json.dumps({"ID": str(record["ID"]), "record_identity": record_identity(record), "input_text_sha256": record["input_text_sha256"], "catalog_manifest_hash": CATALOG_MANIFEST_HASH, "query_signature": signature, "parse_status": parse_status if signature is not None else "semantic_invalid"}, ensure_ascii=False) + "\n")
    unload_generator()
else: print("reusing validated extraction and query-signature caches")
extraction_rows = load_jsonl(EXTRACTION_CACHE); signature_rows = load_jsonl(QUERY_SIGNATURE_CACHE);
if not verify_extraction_cache(extraction_rows, EXPECTED_EXTRACTION_UNIVERSE): raise RuntimeError("cache_identity_mismatch: extraction rows do not match expected factual extraction universe")
if not verify_query_signature_cache(signature_rows, catalog=ENTITY_CATALOG): raise RuntimeError("cache_identity_mismatch: query signature rows do not match frozen record identities or frozen catalog")
if any(row.get("query_signature") is None for row in signature_rows): raise RuntimeError("schema_invalid: query signature cache contains rejected payloads")
QUERY_SIGNATURES = {row["ID"]: row["query_signature"] for row in signature_rows}; assert set(QUERY_SIGNATURES) == set(dataset.ID.astype(str)); query_quality = {"total_queries": len(signature_rows), "valid_signatures": sum(row.get("query_signature") is not None for row in signature_rows), "linked_entity_signatures": sum(bool(query_signature_entity_ids_usable(row.get("query_signature"), ENTITY_CATALOG)["linked_entity_ids"]) for row in signature_rows if isinstance(row.get("query_signature"), dict)), "target_anchor_only_signatures": sum(bool(query_signature_entity_ids_usable(row.get("query_signature"), ENTITY_CATALOG)["target_anchor_ids"]) and not bool(query_signature_entity_ids_usable(row.get("query_signature"), ENTITY_CATALOG)["linked_entity_ids"]) for row in signature_rows if isinstance(row.get("query_signature"), dict)), "coverage_gate": "enforced_by_semantic_linkage_quality_gate"}; query_quality["linked_entity_rate"] = query_quality["linked_entity_signatures"] / max(1, query_quality["total_queries"]); write_json(RUN / "artifacts" / "query_quality_gate.json", query_quality); parse_rate = sum(extraction_parse_accepted(row.get("parse_status"), row.get("validation", {})) for row in extraction_rows) / max(1, len(EXPECTED_EXTRACTION_UNIVERSE)); parse_status_counts = {status: sum(1 for row in extraction_rows if row.get("parse_status") == status) for status in sorted({row.get("parse_status") for row in extraction_rows})}; cache_state = "complete_ready" if parse_rate >= CONFIG["minimum_parse_rate"] else "complete_validation_failed"; persist_extraction_identity_state(cache_state, parse_rate, parse_status_counts); log_event("extraction", "cache_state_persisted", cache_state=cache_state, parse_rate=parse_rate, parse_status_counts=parse_status_counts); accepted_extractions = [row["validation"] for row in extraction_rows]
if cache_state != "complete_ready": log_event("extraction", "parse_rate_gate_failed", cache_state=cache_state, parse_rate=parse_rate, minimum_parse_rate=CONFIG["minimum_parse_rate"], parse_status_counts=parse_status_counts); raise RuntimeError(f"parse_rate gate failed: {parse_rate:.3f}")



# %% [notebook cell 9]

# 07 - Build the validated semantic graph and persist graph manifests.
graph_tables = build_semantic_graph(chunks.to_dict("records"), accepted_extractions)
graph_yield = semantic_linkage_quality_gate(graph_tables, ENTITY_CATALOG, list(QUERY_SIGNATURES.values()), QUALITY_THRESHOLDS, smoke_test=CONFIG["smoke_test"])
graph_yield.update({"accepted_linked_claims": graph_yield["linked_claims"], "graph_claims": len(graph_tables["Claim"]), "graph_evidence": len(graph_tables["EvidenceChunk"]), "entity_count": len(graph_tables["Entity"]), "quarantined": len(graph_tables["quarantined"]), "reviewed": len(graph_tables["reviewed"]), "scoring_calibration_status": SCORING_CALIBRATION_STATUS, "self_confidence_status": SELF_CONFIDENCE_STATUS})
write_json(RUN / "artifacts" / "graph_quality_gate.json", graph_yield)
if not graph_yield["pass"] or not graph_tables["Claim"] or not graph_tables["EvidenceChunk"]: raise RuntimeError("semantic_graph_quality_gate_failed:" + json.dumps(graph_yield, sort_keys=True))
GRAPH_MANIFEST_HASH = stable_id(json.dumps({"graph": graph_tables, "catalog_manifest_hash": CATALOG_MANIFEST_HASH}, ensure_ascii=False, sort_keys=True, default=str))
semantic_kg_nodes = pd.DataFrame([{"node_id": x.get("document_uid"), "node_type": "Document", **x} for x in graph_tables["Document"]] + [{"node_id": x.get("evidence_chunk_id"), "node_type": "EvidenceChunk", **x} for x in graph_tables["EvidenceChunk"]] + [{"node_id": x.get("claim_id"), "node_type": "Claim", **x} for x in graph_tables["Claim"]] + [{"node_id": x.get("entity_id"), "node_type": "Entity", **x} for x in graph_tables["Entity"]]); semantic_kg_edges = pd.DataFrame(graph_tables["edges"])
semantic_kg_nodes.to_parquet(RUN / "artifacts" / "semantic_kg_nodes.parquet", index=False); semantic_kg_edges.to_parquet(RUN / "artifacts" / "semantic_kg_edges.parquet", index=False); write_json(RUN / "artifacts" / "semantic_kg_manifest.json", {"graph_manifest_hash": GRAPH_MANIFEST_HASH, "catalog_manifest_hash": CATALOG_MANIFEST_HASH, "corpus_manifest_hash": CORPUS_MANIFEST_HASH, "chunk_manifest_hash": CHUNK_MANIFEST_HASH, "node_count": len(semantic_kg_nodes), "edge_count": len(semantic_kg_edges)})



# %% [notebook cell 10]

# 08 - Retrieval models, semantic seeds, full-signature graph retrieval, RRF, and reranking.
from sentence_transformers import SentenceTransformer, CrossEncoder
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct
from rank_bm25 import BM25Okapi
def load_retrieval_models():
    global EMBEDDER, RERANKER, BM25, QDRANT, factual_chunks
    factual_chunks = chunks[(chunks.factual_index_allowed == True) & (chunks.authority_score >= CONFIG["minimum_authority"])].reset_index(drop=True);
    if factual_chunks.empty: raise RuntimeError("no_indexable_chunks: no factual chunks met the authority threshold")
    EMBEDDER = SentenceTransformer(CONFIG["embedding_model"], revision=CONFIG["embedding_model_revision"], device="cuda"); RERANKER = CrossEncoder(CONFIG["reranker_model"], revision=CONFIG["reranker_model_revision"], device="cuda", max_length=512); BM25 = BM25Okapi([re.findall(r"[\w']+", str(x).casefold()) for x in factual_chunks.text]); QDRANT = QdrantClient(path=str(RUN / "artifacts" / "qdrant")); return "factual_chunks"
INDEX_MANIFEST_PATH = RUN / "artifacts" / "index_manifest.json"; INDEX_IDENTITY = {"document_uid_manifest_hash": CORPUS_MANIFEST_HASH, "chunk_manifest_hash": CHUNK_MANIFEST_HASH, "graph_manifest_hash": GRAPH_MANIFEST_HASH, "catalog_manifest_hash": CATALOG_MANIFEST_HASH, "extraction_audit_hash": extraction_audit_hash, "embedding_model": CONFIG["embedding_model"], "minimum_authority": CONFIG["minimum_authority"]}
COLLECTION = load_retrieval_models(); old_index_identity = json.loads(INDEX_MANIFEST_PATH.read_text(encoding="utf-8")).get("identity", {}) if INDEX_MANIFEST_PATH.exists() else {}
if old_index_identity != INDEX_IDENTITY and QDRANT.collection_exists(COLLECTION): print("index_identity_mismatch: rebuilding Qdrant collection"); QDRANT.delete_collection(COLLECTION)
if not QDRANT.collection_exists(COLLECTION):
    vectors = EMBEDDER.encode(factual_chunks.text.tolist(), batch_size=8, normalize_embeddings=True, convert_to_numpy=True).astype("float32")
    QDRANT.create_collection(COLLECTION, vectors_config=VectorParams(size=vectors.shape[1], distance=Distance.COSINE))
    points = [PointStruct(id=i, vector=vector.tolist(), payload={**row, "evidence_chunk_id": stable_id("evidence", row["document_uid"], row["chunk_id"]), "page": canonical_page(row.get("page"))}) for i, (row, vector) in enumerate(zip(factual_chunks.to_dict("records"), vectors))]
    for start in range(0, len(points), 128): QDRANT.upsert(COLLECTION, points[start:start + 128], wait=True)
write_json(INDEX_MANIFEST_PATH, {"identity": INDEX_IDENTITY, "document_uid_manifest_hash": CORPUS_MANIFEST_HASH, "chunk_manifest_hash": CHUNK_MANIFEST_HASH})
EVIDENCE_BY_ID = {x["evidence_chunk_id"]: x for x in graph_tables["EvidenceChunk"]}
def dense_search(query, k):
    vector = EMBEDDER.encode([query], normalize_embeddings=True, convert_to_numpy=True)[0].astype("float32").tolist(); points = QDRANT.query_points(COLLECTION, query=vector, limit=k, with_payload=True).points; return [{**dict(p.payload), "branch": "dense", "rank": i, "raw_score": float(p.score), "evidence_chunk_id": dict(p.payload).get("evidence_chunk_id")} for i, p in enumerate(points, 1) if float(p.score) >= CONFIG["minimum_dense_score"]]
def bm25_search(query, k):
    scores = BM25.get_scores(re.findall(r"[\w']+", str(query).casefold())); rows = []
    for rank, index in enumerate(np.argsort(scores)[::-1][:k], 1):
        if float(scores[index]) <= 0: continue
        row = factual_chunks.iloc[int(index)].to_dict(); row.update({"branch": "bm25", "rank": rank, "raw_score": float(scores[index]), "evidence_chunk_id": stable_id("evidence", row["document_uid"], row["chunk_id"]) }); rows.append(row)
    return rows
def graph_search(target, k, post="", category="", query_signature=None, graph_enabled=True):
    if not graph_enabled: return []
    query_signature = query_signature or {"entity_ids": [], "predicates": [], "polarities": [], "modalities": [], "desired_stances": []}; seeds = dense_search(f"evidence addressing {target} {post}", CONFIG["dense_top_k"]) + bm25_search(f"{target} {post}", CONFIG["bm25_top_k"]); return expand_graph_from_seeds(seeds, query_signature=query_signature, graph_tables=graph_tables, config=GRAPH_CONFIG)[:k]
def retrieve_evidence(post, target, category, query_signature, graph_enabled):
    if not query_signature.get("entity_ids"):
        return {"evidence": [], "candidate_evidence_ids": [], "graph_enabled": graph_enabled, "abstention_reason": "no_linked_entities"}
    dense_hits, bm25_hits = dense_search(f"evidence addressing harmful or misleading claims about {target}", CONFIG["dense_top_k"]), bm25_search(f"{target} {post}", CONFIG["bm25_top_k"]); graph_hits = graph_search(target, CONFIG["graph_top_k"], post, category, query_signature, graph_enabled); candidates = reciprocal_rank_fusion({"dense": dense_hits, "bm25": bm25_hits, "graph": graph_hits}, {"dense": 1.1, "bm25": 1.0, "graph": 0.8}, CONFIG["rrf_constant"])
    for item in candidates: item.update(EVIDENCE_BY_ID.get(item["evidence_chunk_id"], {}))
    if not candidates: return {"evidence": [], "candidate_evidence_ids": [], "graph_enabled": graph_enabled, "abstention_reason": "no_positive_candidates"}
    logits = RERANKER.predict([(post, x["text"]) for x in candidates], batch_size=16).tolist(); selection = select_evidence(candidates, {x["evidence_chunk_id"]: float(y) for x, y in zip(candidates, logits)}, {"minimum_rerank_probability": CONFIG["minimum_rerank_probability"], "max_evidence": CONFIG["rerank_top_k"]}); return {"evidence": selection["selected"], "candidate_evidence_ids": [x["evidence_chunk_id"] for x in candidates], "graph_enabled": graph_enabled, "graph_hit_count": len(graph_hits), "retrieval_trace": {"dense_hits": len(dense_hits), "bm25_hits": len(bm25_hits), "graph_hits": len(graph_hits)}, "abstention_reason": selection["reason"]}



# %% [notebook cell 11]

# 09 - Paired graph-on/off retrieval over identical frozen IDs and retrieval-model release.
EVIDENCE_CACHE = BoundedLRU(maxsize=int(CONFIG["cache_max_records"])); PAIRED_RETRIEVAL_CACHE = BoundedLRU(maxsize=int(CONFIG["cache_max_records"]))
def paired_permutation(a, b, rounds=10000, seed=SEED):
    delta = np.asarray(a, dtype=float) - np.asarray(b, dtype=float); rng = np.random.default_rng(seed); observed = abs(delta.mean()); null = [abs((delta * rng.choice([-1, 1], len(delta))).mean()) for _ in range(rounds)]; return {"mean_difference": float(delta.mean()), "p_value": float((np.asarray(null) >= observed).mean())}
def paired_retrieval_metrics(graph_on, graph_off, shared_frozen_universe):
    on = graph_on["evidence"]; off = graph_off["evidence"]; on_by_id = {x["evidence_chunk_id"]: x for x in on}; off_by_id = {x["evidence_chunk_id"]: x for x in off}; universe = list(dict.fromkeys(shared_frozen_universe)); domain_size = len(universe)
    on_ids, off_ids = set(on_by_id), set(off_by_id)
    if on_ids - set(universe) or off_ids - set(universe): raise ValueError("selected evidence outside shared frozen universe")
    on_selection = [1 if evidence_id in on_by_id else 0 for evidence_id in universe]; off_selection = [1 if evidence_id in off_by_id else 0 for evidence_id in universe]
    on_authority = [float(on_by_id[evidence_id].get("authority_score") or 0.0) if evidence_id in on_by_id else 0.0 for evidence_id in universe]; off_authority = [float(off_by_id[evidence_id].get("authority_score") or 0.0) if evidence_id in off_by_id else 0.0 for evidence_id in universe]
    on_accepted = [1 if evidence_id in on_by_id and str(on_by_id[evidence_id].get("status", "")) == "accepted" else 0 for evidence_id in universe]; off_accepted = [1 if evidence_id in off_by_id and str(off_by_id[evidence_id].get("status", "")) == "accepted" else 0 for evidence_id in universe]
    on_score = [float(on_by_id[evidence_id].get("rerank_probability", on_by_id[evidence_id].get("rrf_score", 0.0)) or 0.0) if evidence_id in on_by_id else 0.0 for evidence_id in universe]; off_score = [float(off_by_id[evidence_id].get("rerank_probability", off_by_id[evidence_id].get("rrf_score", 0.0)) or 0.0) if evidence_id in off_by_id else 0.0 for evidence_id in universe]
    denominator = max(1, domain_size); mean = lambda values: float(sum(values) / denominator); traces = [{"evidence_id": evidence_id, "graph_on_selected": on_selection[i], "graph_off_selected": off_selection[i], "graph_on_authority": on_authority[i], "graph_off_authority": off_authority[i], "graph_on_accepted": on_accepted[i], "graph_off_accepted": off_accepted[i], "graph_on_score": on_score[i], "graph_off_score": off_score[i]} for i, evidence_id in enumerate(universe)]
    return {"shared_frozen_universe": universe, "universe_size": domain_size, "universe_traces": traces, "graph_on_selection_vector": on_selection, "graph_off_selection_vector": off_selection, "graph_on_authority_vector": on_authority, "graph_off_authority_vector": off_authority, "graph_on_accepted_vector": on_accepted, "graph_off_accepted_vector": off_accepted, "graph_on_score_vector": on_score, "graph_off_score_vector": off_score, "overlap": mean([a * b for a, b in zip(on_selection, off_selection)]), "graph_only_gain": mean([a - b for a, b in zip(on_selection, off_selection)]), "authority_rate": mean(on_authority), "graph_off_authority_rate": mean(off_authority), "accepted_evidence_rate": mean(on_accepted), "graph_off_accepted_evidence_rate": mean(off_accepted), "selected_score_mean": mean(on_score), "graph_off_selected_score_mean": mean(off_score), "graph_on_selected_count": sum(on_selection), "graph_off_selected_count": sum(off_selection)}
def paired_statistical_comparison(metric_rows):
    on = [float(sum(row.get("graph_on_score_vector", [])) / max(1, row.get("universe_size", len(row.get("graph_on_score_vector", []))))) for row in metric_rows]; off = [float(sum(row.get("graph_off_score_vector", [])) / max(1, row.get("universe_size", len(row.get("graph_off_score_vector", []))))) for row in metric_rows]
    return {"valid": len(on) >= 2, **(paired_permutation(on, off) if len(on) >= 2 else {"mean_difference": None, "p_value": None})}
for record in tqdm(dataset.to_dict("records"), desc="Paired retrieval graph-on/off"):
    rid, signature = str(record["ID"]), QUERY_SIGNATURES[str(record["ID"])]
    graph_on, graph_off = retrieve_evidence(record["Text"], record["Target"], record["Category"], signature, True), retrieve_evidence(record["Text"], record["Target"], record["Category"], signature, False); shared_frozen_universe = sorted(set(graph_on.get("candidate_evidence_ids", [])) | set(graph_off.get("candidate_evidence_ids", []))); graph_on["shared_frozen_universe"] = shared_frozen_universe; graph_off["shared_frozen_universe"] = shared_frozen_universe; metrics_row = paired_retrieval_metrics(graph_on, graph_off, shared_frozen_universe); PAIRED_RETRIEVAL_CACHE[rid] = {"record_identity": record_identity(record), "query_signature": signature, "graph_on": graph_on, "graph_off": graph_off, "frozen_evidence_ids": shared_frozen_universe, "shared_frozen_universe": shared_frozen_universe, "metrics": metrics_row}; EVIDENCE_CACHE[rid] = graph_on if not CONFIG["graph_ablation"] else graph_off
paired_metrics = [row["metrics"] for row in PAIRED_RETRIEVAL_CACHE.values()]; paired_comparison = paired_statistical_comparison(paired_metrics) if paired_metrics else {"valid": False, "mean_difference": None, "p_value": None}; cache_snapshot = {"evidence": EVIDENCE_CACHE.snapshot(), "paired_retrieval": PAIRED_RETRIEVAL_CACHE.snapshot(), "cache_capacity": CONFIG["cache_max_records"], "evictions": {"evidence": EVIDENCE_CACHE.evictions, "paired_retrieval": PAIRED_RETRIEVAL_CACHE.evictions}}; write_json(RUN / "artifacts" / "paired_retrieval_evaluation.json", {"records": dict(PAIRED_RETRIEVAL_CACHE), "paired_statistical_comparison": paired_comparison, "cache_snapshot": cache_snapshot}); write_json(RUN / "artifacts" / "cache_diagnostics.json", cache_snapshot); log_event("retrieval", "cache_snapshot", cache_snapshot=cache_snapshot)
def unload_retrieval_models():
    global EMBEDDER, RERANKER, BM25, QDRANT
    for name in ["EMBEDDER", "RERANKER", "BM25"]:
        if name in globals(): del globals()[name]
    if "QDRANT" in globals():
        try: QDRANT.close()
        except Exception: pass
        del QDRANT
    gc.collect(); torch.cuda.empty_cache(); MEMORY_SNAPSHOTS.append(gpu_snapshot("retrieval_models_unloaded")); print("retrieval_models_unloaded")
unload_retrieval_models()



# %% [notebook cell 12]

# 10 - Reload the same Qwen generator only after retrieval models are unloaded.
print("Qwen generation load deferred until after the shared NLI lifecycle gate.")



# %% [notebook cell 13]

# 11 - Canonical E1...En ledgers for perspectives, synthesis, generation, and metrics.
PERSPECTIVES = [("fact_checking", "Correct only directly supported claims."), ("cultural_context", "Give respectful context."), ("harm_reduction", "Avoid amplifying harmful language."), ("legal_rights", "Use only source-grounded rights claims."), ("persuasion", "Suggest an empathetic response.")]
def build_evidence_ledger(evidence):
    source_lookup = globals().get("SOURCE_TEXT_BY_KEY", {}); rows = []
    for i, x in enumerate(evidence):
        displayed = str(x.get("text", "")); source_key = x.get("source_text_key"); source_text = source_lookup.get(source_key) if source_key else displayed
        if source_key and source_text is None: raise RuntimeError("source_text_lookup_missing")
        start = int(x.get("span_start", 0) or 0); end = int(x.get("span_end", len(displayed)) or 0)
        if not isinstance(source_text, str) or start < 0 or start >= end or end > len(source_text) or source_text[start:end] != displayed: raise RuntimeError("evidence_span_invariant_failed")
        rows.append({"evidence_id": f"E{i + 1}", "ledger_id": f"E{i + 1}", "rank": x.get("rank", i), "evidence_chunk_id": x["evidence_chunk_id"], "chunk_id": x.get("chunk_id"), "document_uid": x.get("document_uid"), "source_id": x.get("source_id"), "locator": f"p. {canonical_page(x.get('page'))}" if canonical_page(x.get("page")) is not None else f"section: {x.get('section', 'document')}", "displayed_text": displayed, "evidence_text": displayed, "source_text_key": source_key or f"inline-{i}", "source_text_sha256": __import__("hashlib").sha256(source_text.encode("utf-8")).hexdigest(), "span_start": start, "span_end": end, "displayed_text_sha256": __import__("hashlib").sha256(displayed.encode("utf-8")).hexdigest()})
    return validate_evidence_ledger(rows, source_lookup) if "validate_evidence_ledger" in globals() else rows
def validate_evidence_ids(ids, ledger):
    if not isinstance(ids, list): return {"valid": False, "unknown": [str(ids)] if ids is not None else [], "evidence_ids": [], "quarantine": True, "reason": "evidence_ids_not_list"}
    values = list(ids); allowed = {x["evidence_id"] for x in ledger}; unknown = [str(value) for value in values if not isinstance(value, str) or value not in allowed]; valid = all(isinstance(value, str) for value in values) and len(values) == len(set(values)) and not unknown and all(re.fullmatch(r"E[1-9][0-9]*", value) for value in values); return {"valid": valid, "unknown": unknown, "evidence_ids": values, "quarantine": not valid}
def evidence_for_prompt(evidence, char_budget=None):
    ledger = build_evidence_ledger(evidence)
    if char_budget is None: return evidence, ledger
    selected = select_evidence_within_budget(ledger, int(char_budget)); selected_ids = {x["evidence_id"] for x in selected}
    return [item for item, entry in zip(evidence, ledger) if entry["evidence_id"] in selected_ids], selected
def evidence_block(evidence, limit=None):
    # Whole spans only: an explicit budget selects a ledger-consistent prefix;
    # it never slices evidence text.
    _, ledger = evidence_for_prompt(evidence, limit)
    return "\n\n".join(f"[{x['evidence_id']}] source={x['source_id']} {x['locator']} document_uid={x['document_uid']} span={x['span_start']}:{x['span_end']}\n{x['displayed_text']}" for x in ledger)
def _ledger_block(ledger):
    return "\n\n".join(f"[{x['evidence_id']}] source={x['source_id']} {x['locator']} document_uid={x['document_uid']} span={x['span_start']}:{x['span_end']}\n{x['displayed_text']}" for x in ledger)
def _prompt_token_count(prompt):
    encoded = tokenizer(text=[str(prompt)], return_tensors="pt", padding=False, truncation=False)
    mask = encoded["attention_mask"]
    return int(mask.sum().item()) if hasattr(mask.sum(), "item") else int(mask.sum())
def _adaptive_messages(evidence, message_builder, payload_candidates, max_new_tokens):
    # Deterministic token fitting drops whole ledger spans, never text slices.
    ledger = build_evidence_ledger(evidence)
    def rendered_builder(selected, payload):
        return _apply_chat_template(message_builder(selected, payload), enable_thinking=False)
    fit = fit_adaptive_prompt_with_evidence(ledger, prompt_builder=rendered_builder, payload_candidates=payload_candidates, token_counter=_prompt_token_count, max_input_tokens=int(CONFIG["max_seq_length"]), reserve_output_tokens=int(max_new_tokens))
    if fit.get("status") != "fit": raise PromptBudgetQuarantine(fit)
    selected_ids = set(fit.get("selected_evidence_ids", [])); selected = [item for item in ledger if item["evidence_id"] in selected_ids]
    return message_builder(selected, fit.get("payload", "")), fit
def adaptive_perspective_messages(post, target, category, evidence):
    prompts = []; schema_tail = "JSON schema (must be followed exactly): " + json.dumps(PERSPECTIVE_SCHEMA, ensure_ascii=False, sort_keys=True)
    for name, goal in PERSPECTIVES:
        def builder(selected, payload, name=name, goal=goal):
            return [{"role": "system", "content": "Think carefully, then return exactly one JSON object after the thinking block."}, {"role": "user", "content": f"Perspective: {name}\nGoal: {goal}\nPost: {post}\nTarget: {target}\nCategory: {category}\nEvidence (complete ledger spans only):\n{_ledger_block(selected)}\nReturn perspective, rationale, claims_to_address, supported_evidence_ids, response_guidance, risk_flags, confidence. Use only E1...En IDs.\n{schema_tail}"}]
        prompt, fit = _adaptive_messages(evidence, builder, ("",), CONFIG["perspective_max_new_tokens"]); prompts.append((prompt, fit))
    return prompts
def enforce_perspective_parse_rate(rows):
    total_count = len(rows)
    parse_status_counts = {}
    for row in rows:
        status = str(row.get("structured_output", {}).get("parse_status", "schema_invalid"))
        parse_status_counts[status] = parse_status_counts.get(status, 0) + 1
    parsed_count = sum(parse_status_counts.get(status, 0) for status in ACCEPTED_PARSE_STATUSES)
    parse_rate = parsed_count / max(1, total_count)
    audit = {"reason": "minimum_parse_rate_failed", "parse_rate": parse_rate, "minimum_parse_rate": float(CONFIG["minimum_parse_rate"]), "parsed_count": parsed_count, "total_count": total_count, "parse_status_counts": parse_status_counts}
    if parse_rate < audit["minimum_parse_rate"]: raise ParseRateQuarantine(audit)
    return audit
def perspective_messages(post, target, category, evidence):
    prompt_evidence, _ = evidence_for_prompt(evidence, CONFIG.get("evidence_char_budget")); return [[{"role": "system", "content": "Return exactly one JSON object and no reasoning trace."}, {"role": "user", "content": f"Perspective: {name}\nGoal: {goal}\nPost: {post}\nTarget: {target}\nCategory: {category}\nEvidence (full exact displayed spans):\n{evidence_block(prompt_evidence)}\nReturn perspective, rationale, claims_to_address, supported_evidence_ids, response_guidance, risk_flags, confidence. Use only E1...En IDs."}] for name, goal in PERSPECTIVES]
def run_perspectives(post, target, category, evidence):
    rows = []; prompt_entries = adaptive_perspective_messages(post, target, category, evidence); prompts = [entry[0] for entry in prompt_entries]; raw_outputs = generate_batch(prompts, CONFIG["perspective_max_new_tokens"], temperature=0.6, enable_thinking=CONFIG["thinking_enabled"], output_schema=PERSPECTIVE_SCHEMA)
    for (name, goal), (prompt, fit), raw in zip(PERSPECTIVES, prompt_entries, raw_outputs):
        allowed = set(fit.get("selected_evidence_ids", []))
        parsed, final_raw, status = parse_with_one_repair(raw, prompt[1]["content"], lambda payload: validate_perspective(payload, allowed, expected_perspective=name), PERSPECTIVE_SCHEMA);
        if parsed is None: parsed = {"perspective": name, "rationale": "", "claims_to_address": [], "supported_evidence_ids": [], "response_guidance": [], "risk_flags": [status], "confidence": 0.0, "parse_status": status, "quarantine": True}
        else: parsed["parse_status"] = status
        rows.append({"perspective": name, "goal": goal, "perspective_rationale": str(parsed.get("rationale", "")), "structured_output": parsed, "supported_evidence_ids": parsed.get("supported_evidence_ids", []), "reasoning_trace": qwen_generation_trace(final_raw), "raw_output": final_raw})
    enforce_perspective_parse_rate(rows)
    return rows



# %% [notebook cell 14]

# 12 - Synthesis, final generation, canonical citation validation, and checkpoint identity.
VARIANT_COLUMNS = {"qwen_zero_shot": "qwen-zero-shot-counter-narrative", "qwen_few_shot": "qwen-few-shot-counter-narrative", "kg_rag": "kg-rag-generated-source-grounded-counter-narrative", "mp_kg_rag": "mp-kg-rag-generated-source-grounded-counter-narrative"}
def resolve_citation_tokens(text, evidence_ledger):
    by_id = {x["evidence_id"]: x for x in evidence_ledger}; tokens = re.findall(r"\[(E\d+)\]", str(text)); return tokens, [by_id[x] for x in tokens if x in by_id], sorted(set(tokens) - set(by_id))
def claim_level_citations(text, evidence_ledger, factual_claims=None):
    if "build_claim_citation_records" in globals():
        return build_claim_citation_records(text, evidence_ledger, factual_claims=factual_claims)
    return [{"claim": sentence, "claim_text": sentence, "evidence_ids": resolve_citation_tokens(sentence, evidence_ledger)[0], "is_factual": bool(factual_claims), "citation_format_valid": True, "unknown_evidence_ids": []} for sentence in re.split(r"(?<=[.!?])\s+", str(text).strip()) if sentence]
def synthesize_plan(post, target, category, evidence, perspective_outputs):
    compact = [{k: p["structured_output"].get(k, []) for k in ["perspective", "rationale", "claims_to_address", "supported_evidence_ids", "response_guidance", "risk_flags"]} for p in perspective_outputs if p["structured_output"].get("parse_status") in ACCEPTED_PARSE_STATUSES]
    compact_minimal = [{"perspective": p.get("perspective"), "supported_evidence_ids": p.get("supported_evidence_ids", [])} for p in compact]
    payload_candidates = (json.dumps(compact, ensure_ascii=False, sort_keys=True), json.dumps(compact_minimal, ensure_ascii=False, sort_keys=True))
    schema_tail = "JSON schema (must be followed exactly): " + json.dumps(PLAN_SCHEMA, ensure_ascii=False, sort_keys=True)
    def builder(selected, payload):
        return [{"role": "system", "content": "Return one JSON evidence plan only."}, {"role": "user", "content": f"Post: {post}\nTarget: {target}\nEvidence (complete ledger spans only):\n{_ledger_block(selected)}\nPerspectives: {payload}\nReturn claim_focus, selected_evidence_ids, response_steps, tone, factual_constraints, safety_constraints. Use only E IDs.\n{schema_tail}"}]
    prompt, fit = _adaptive_messages(evidence, builder, payload_candidates, CONFIG["plan_max_new_tokens"]); ledger_ids = set(fit.get("selected_evidence_ids", [])); raw = generate_batch([prompt], CONFIG["plan_max_new_tokens"], temperature=0.6, enable_thinking=CONFIG["thinking_enabled"], output_schema=PLAN_SCHEMA)[0]; parsed, final_raw, status = parse_with_one_repair(raw, prompt[1]["content"], lambda payload: validate_plan(payload, ledger_ids), PLAN_SCHEMA);
    if parsed is None: raise SchemaValidationQuarantine({'reason': 'plan_schema_invalid', 'stage': 'plan', 'parse_status': status})
    parsed["parse_status"] = status; return parsed, final_raw
def generate_final_counter_narrative(post, target, evidence, plan):
    compact_plan = {key: plan.get(key) for key in ["claim_focus", "selected_evidence_ids", "response_steps", "tone"] if key in plan}
    payload_candidates = (json.dumps(plan, ensure_ascii=False, sort_keys=True), json.dumps(compact_plan, ensure_ascii=False, sort_keys=True))
    schema_tail = "JSON schema (must be followed exactly): " + json.dumps(FINAL_RESPONSE_SCHEMA, ensure_ascii=False, sort_keys=True)
    def builder(selected, payload):
        return [{"role": "system", "content": "Write one safe evidence-grounded counter-narrative as JSON only."}, {"role": "user", "content": f"Post: {post}\nTarget: {target}\nPlan: {payload}\nEvidence (complete ledger spans only):\n{_ledger_block(selected)}\nFactual claims require inline [E1] citations. If there are no factual claims, return exactly one approved safe-abstention template and set factual_claims to []. Return counter_narrative, cited_evidence_ids, factual_claims, safety_notes. Use only E IDs; chunk IDs are metadata only.\n{schema_tail}"}]
    prompt, fit = _adaptive_messages(evidence, builder, payload_candidates, CONFIG["answer_max_new_tokens"]); ledger_ids = set(fit.get("selected_evidence_ids", [])); raw = generate_batch([prompt], CONFIG["answer_max_new_tokens"], temperature=0.6, enable_thinking=CONFIG["thinking_enabled"], output_schema=FINAL_RESPONSE_SCHEMA)[0]; parsed, final_raw, status = parse_with_one_repair(raw, prompt[1]["content"], lambda payload: validate_final_response(payload, ledger_ids), FINAL_RESPONSE_SCHEMA);
    if parsed is None: raise SchemaValidationQuarantine({'reason': 'final_schema_invalid', 'stage': 'final', 'parse_status': status})
    parsed["parse_status"] = status; return parsed, final_raw
def grounding_repair_response(post, target, evidence, plan, failure_reasons):
    payload_candidates = (json.dumps({"plan": plan, "failures": failure_reasons}, ensure_ascii=False, sort_keys=True), json.dumps({"failures": [str(item) for item in failure_reasons[:3]]}, ensure_ascii=False, sort_keys=True))
    schema_tail = "JSON schema (must be followed exactly): " + json.dumps(FINAL_RESPONSE_SCHEMA, ensure_ascii=False, sort_keys=True)
    def builder(selected, payload):
        return [{"role": "system", "content": "Repair one failed evidence-grounded response as strict JSON."}, {"role": "user", "content": f"Post: {post}\nTarget: {target}\nRepair context: {payload}\nEvidence (complete ledger spans only):\n{_ledger_block(selected)}\nReturn counter_narrative, cited_evidence_ids, factual_claims, safety_notes. Cite only supported E IDs.\n{schema_tail}"}]
    prompt, fit = _adaptive_messages(evidence, builder, payload_candidates, CONFIG["answer_max_new_tokens"]); ledger_ids = set(fit.get("selected_evidence_ids", [])); raw = generate_batch([prompt], CONFIG["answer_max_new_tokens"], temperature=0.6, enable_thinking=CONFIG["thinking_enabled"], output_schema=FINAL_RESPONSE_SCHEMA)[0]; parsed, final_raw, status = parse_with_one_repair(raw, prompt[1]["content"], lambda payload: validate_final_response(payload, ledger_ids), FINAL_RESPONSE_SCHEMA)
    if parsed is None: raise SchemaValidationQuarantine({'reason': 'grounding_repair_schema_invalid', 'stage': 'grounding_repair', 'parse_status': status})
    parsed['parse_status'] = status; parsed['grounding_repair_reasons'] = list(failure_reasons); return parsed, final_raw
SAFE_ABSTENTION_TEMPLATES = {"I cannot verify this from the available evidence.", "I can’t verify this from the available evidence.", "இந்தக் கூற்றை கிடைக்கும் ஆதாரங்களிலிருந்து சரிபார்க்க முடியவில்லை.", "उपलब्ध साक्ष्यों से इस दावे की पुष्टि नहीं कर सकता।"}
def safe_abstention_validator(text):
    raw = unicodedata.normalize("NFKC", str(text or "")).strip()
    if re.search(r"\[(?:E[1-9][0-9]*)\]", raw): return False
    normalized = re.sub(r"\s+", " ", raw).strip().rstrip(" .!?！？。॥")
    return any(normalized == re.sub(r"\s+", " ", unicodedata.normalize("NFKC", item).strip()).rstrip(" .!?！？。॥") for item in SAFE_ABSTENTION_TEMPLATES)
NLI_CALIBRATION_ARTIFACT = {}
LANGUAGE_DETECTOR = load_language_detector()
if LANGUAGE_DETECTOR is None: raise RuntimeError('language_evaluation_detector_unavailable')
if CONFIG.get('nli_model_label_mapping') != MODEL_LABEL_MAPPING or CONFIG.get('nli_dataset_label_mapping') != DATASET_LABEL_MAPPING:
    raise RuntimeError('nli_label_mapping_config_mismatch')
NLI_CALIBRATION_PATH = RUN / 'artifacts' / 'nli_calibration.json'
if NLI_CALIBRATION_PATH.exists():
    NLI_CALIBRATION_ARTIFACT = json.loads(NLI_CALIBRATION_PATH.read_text(encoding='utf-8'))
def create_nli_calibration_artifact(language, heldout_rows, predictor, *, dataset_id, dataset_revision, split='validation', label_mapping=None):
    rows = list(heldout_rows)
    if not rows: raise RuntimeError('nli_calibration_dataset_empty')
    example_ids = [str(row['id']) for row in rows]
    example_digest = stable_identity_hash([{key: row.get(key) for key in ['id', 'premise', 'hypothesis', 'label']} for row in rows])
    raw_predictions = [predictor(row['premise'], row['hypothesis']) for row in rows]
    predictions = [int(value.get('label')) if isinstance(value, dict) else int(value) for value in raw_predictions]
    labels = [int(row['label']) for row in rows]; accuracy = sum(prediction == label for prediction, label in zip(predictions, labels)) / len(labels)
    per_label = {}
    for label in sorted(set(labels)):
        support = sum(value == label for value in labels); true_positive = sum(value == label and prediction == label for value, prediction in zip(labels, predictions)); predicted = sum(prediction == label for prediction in predictions)
        per_label[str(label)] = {'n': support, 'correct': true_positive, 'precision': true_positive / predicted if predicted else 0.0, 'recall': true_positive / support if support else 0.0}
    selected_label_mapping = label_mapping or CONFIG.get('nli_dataset_label_mapping', DATASET_LABEL_MAPPING)
    entailment_label = int(selected_label_mapping.get('entailment')) if 'entailment' in selected_label_mapping else int(next(key for key, value in selected_label_mapping.items() if str(value).casefold() == 'entailment'))
    per_label['entailment'] = per_label.get(str(entailment_label), {'n': 0, 'correct': 0, 'precision': 0.0, 'recall': 0.0})
    dataset_hash = normalized_dataset_content_hash(rows)
    scores = [float(value.get('entailment_probability', 0.0)) if isinstance(value, dict) else float(value == 0) for value in raw_predictions]
    indexed_predictions = {str(row['id']): {'label': prediction, 'entailment_probability': score} for row, prediction, score in zip(rows, predictions, scores)}
    calibration = calibrate_nli_threshold(rows, lambda row: indexed_predictions[str(row['id'])], seed=SEED, min_support=int(CONFIG['nli_calibration_min_support']), n_bootstrap=int(CONFIG['nli_calibration_bootstrap']), entailment_label=entailment_label)
    calibration_rows = [row for row in rows if str(row['id']) in calibration['calibration_ids']]
    audit_rows = [row for row in rows if str(row['id']) in calibration['audit_ids']]
    audit_stats = {}
    audit_predictions = [indexed_predictions[str(item['id'])]['label'] for item in audit_rows]
    for raw_label in sorted({int(row['label']) for row in audit_rows}):
        support = sum(int(row['label']) == raw_label for row in audit_rows)
        correct = sum(int(row['label']) == raw_label and prediction == raw_label for row, prediction in zip(audit_rows, audit_predictions))
        predicted = sum(prediction == raw_label for prediction in audit_predictions)
        audit_stats[str(raw_label)] = {'n_total': support, 'positive_support': support, 'correct': correct, 'precision': correct / predicted if predicted else 0.0, 'recall': correct / support if support else 0.0}
    audit_stats['entailment'] = {**calibration['metrics']['audit']['entailment'], 'n_total': len(audit_rows), 'positive_support': calibration['metrics']['audit']['entailment']['support']}
    artifact = build_nli_calibration_artifact(model_id=CONFIG['nli_model_id'], model_revision=CONFIG['nli_model_revision'], dataset_id=dataset_id, dataset_revision=dataset_revision, dataset_content_hash=dataset_hash, language=language, split=split, label_mapping=selected_label_mapping, threshold=calibration['threshold'], n_examples=len(rows), example_ids=example_ids, example_content_digest=example_digest, accuracy=calibration['metrics']['audit']['accuracy'], per_label_stats=audit_stats, calibration_metadata={'method': calibration['method'], 'seed': calibration['seed'], 'criterion': calibration['criterion'], 'entailment_label': calibration['entailment_label'], 'calibration_ids': sorted(calibration['calibration_ids']), 'audit_ids': sorted(calibration['audit_ids']), 'calibration_content_digest': normalized_dataset_content_hash(calibration_rows), 'audit_content_digest': normalized_dataset_content_hash(audit_rows), 'support': calibration['support'], 'metrics': calibration['metrics'], 'dataset_content_hash': dataset_hash}, code_hash=CORE_SOURCE_SHA256, eval_core_hash=EVAL_CORE_SOURCE_SHA256)
    existing = NLI_CALIBRATION_ARTIFACT.get(language) if isinstance(NLI_CALIBRATION_ARTIFACT, dict) else None
    payload = dict(NLI_CALIBRATION_ARTIFACT) if isinstance(NLI_CALIBRATION_ARTIFACT, dict) else {}; payload[language] = artifact; write_json_atomic(NLI_CALIBRATION_PATH, payload)
    verified = verify_nli_calibration_artifact(artifact, model_id=CONFIG['nli_model_id'], model_revision=CONFIG['nli_model_revision'], dataset_id=dataset_id, dataset_revision=dataset_revision, dataset_content_hash=dataset_hash, language=language, split=split, code_hash=CORE_SOURCE_SHA256, eval_core_hash=EVAL_CORE_SOURCE_SHA256, dataset_examples=rows, label_mapping=label_mapping or CONFIG.get('nli_dataset_label_mapping'))
    if not verified.get('enabled'): raise RuntimeError('nli_calibration_self_verification_failed')
    quality = nli_calibration_quality(artifact, min_accuracy=CONFIG['nli_min_accuracy'], min_entailment_precision=CONFIG['nli_min_entailment_precision'], min_entailment_recall=CONFIG['nli_min_entailment_recall'], min_per_label_support=CONFIG['nli_min_per_label_support'], min_accuracy_lower=CONFIG['nli_min_accuracy_lower'], min_entailment_precision_lower=CONFIG['nli_min_entailment_precision_lower'], min_entailment_recall_lower=CONFIG['nli_min_entailment_recall_lower'])
    if not quality.get('enabled'): raise RuntimeError(f"nli_calibration_quality_failed:{language}")
    NLI_CALIBRATION_ARTIFACT[language] = artifact
    return artifact
def _verified_nli_model_indices(classifier):
    raw = getattr(getattr(getattr(classifier, 'model', None), 'config', None), 'id2label', None)
    if not isinstance(raw, dict): raise RuntimeError('nli_model_id2label_unavailable')
    normalized = {}
    for key, value in raw.items():
        key_text = str(key); key_text = f'LABEL_{key_text}' if key_text.isdigit() else key_text
        normalized[key_text] = str(value).casefold()
    if validate_nli_label_mapping(normalized, kind='model')['valid'] is False: raise RuntimeError('nli_model_id2label_mapping_mismatch')
    indices = {}
    for key, value in normalized.items():
        index = int(key.rsplit('_', 1)[1]); indices[key] = index; indices[key.casefold()] = index; indices[value] = index
    return indices
NLI_SHARED_PIPELINE = None
NLI_SHARED_MODEL_INDICES = None
NLI_SHARED_DEVICE = -1  # CPU offload permits the one evaluator to coexist with Qwen without GPU residency.
def _nli_memory_gate():
    if NLI_SHARED_DEVICE != -1 and globals().get("model") is not None and torch.cuda.is_available():
        raise RuntimeError("nli_qwen_gpu_coexistence")
    return {"device": NLI_SHARED_DEVICE, "qwen_gpu_resident": bool(globals().get("model") is not None and torch.cuda.is_available()), "offloaded": NLI_SHARED_DEVICE == -1}
def load_shared_nli_pipeline():
    global NLI_SHARED_PIPELINE, NLI_SHARED_MODEL_INDICES
    _nli_memory_gate()
    if NLI_SHARED_PIPELINE is None:
        from transformers import pipeline
        NLI_SHARED_PIPELINE = pipeline("text-classification", model=CONFIG["nli_model_id"], revision=CONFIG["nli_model_revision"], tokenizer=CONFIG["nli_model_id"], device=NLI_SHARED_DEVICE)
        NLI_SHARED_MODEL_INDICES = _verified_nli_model_indices(NLI_SHARED_PIPELINE)
    return NLI_SHARED_PIPELINE, NLI_SHARED_MODEL_INDICES
def release_shared_nli_pipeline():
    global NLI_SHARED_PIPELINE, NLI_SHARED_MODEL_INDICES
    NLI_SHARED_PIPELINE = None; NLI_SHARED_MODEL_INDICES = None; gc.collect()
    if torch.cuda.is_available(): torch.cuda.empty_cache()
    MEMORY_SNAPSHOTS.append(gpu_snapshot("shared_nli_pipeline_released"))
def load_or_create_nli_calibration(language, classifier=None, model_indices=None):
    provenance = CONFIG['nli_calibration_provenance'][language]
    from datasets import load_dataset
    all_rows, provenance = load_nli_dataset_rows(load_dataset, provenance)
    rows = list(select_stratified_nli_rows(all_rows, max_rows=int(CONFIG.get('nli_calibration_examples', 256)), seed=SEED))
    if len(rows) < 6 or len({int(row['label']) for row in rows}) < 3:
        raise RuntimeError('nli_calibration_stratified_sample_incomplete')
    selected_hash = normalized_dataset_content_hash(rows)
    provenance = {**provenance, 'dataset_content_hash': selected_hash}
    entry = NLI_CALIBRATION_ARTIFACT.get(language) if isinstance(NLI_CALIBRATION_ARTIFACT, dict) else None
    if classifier is None or model_indices is None:
        classifier, model_indices = load_shared_nli_pipeline()
    label_map = DATASET_LABEL_MAPPING
    def predictor(premise, hypothesis):
        scores = classifier({'text': premise, 'text_pair': hypothesis}, top_k=None); values = scores[0] if scores and isinstance(scores[0], list) else scores; best = max(values, key=lambda item: float(item['score']))
        if CONFIG.get('nli_model_label_mapping') != MODEL_LABEL_MAPPING:
            raise RuntimeError('nli_model_label_mapping_mismatch')
        indexed = {model_indices.get(str(item['label']).casefold(), model_indices.get(str(item['label']), -1)): float(item['score']) for item in values}
        return {'label': {0: 2, 1: 1, 2: 0}.get(model_indices.get(str(best['label']).casefold(), model_indices.get(str(best['label']), -1)), -1), 'entailment_probability': indexed.get(2, 0.0)}
    expected = verify_nli_calibration_artifact(entry, model_id=provenance['model_id'], model_revision=provenance['model_revision'], dataset_id=provenance['dataset_id'], dataset_revision=provenance['dataset_revision'], dataset_content_hash=provenance['dataset_content_hash'], language=language, split=provenance['split'], code_hash=provenance['code_hash'], eval_core_hash=provenance['eval_core_hash'], dataset_examples=rows, label_mapping=provenance.get('dataset_label_mapping'), fresh_predictor=lambda row: predictor(row['premise'], row['hypothesis']), calibration_seed=SEED, calibration_min_support=int(CONFIG['nli_calibration_min_support']), calibration_bootstrap=int(CONFIG['nli_calibration_bootstrap']), calibration_criterion='f1', calibration_entailment_label=0) if entry else {'enabled': False}
    if expected.get('enabled'):
        quality = nli_calibration_quality(entry, min_accuracy=CONFIG['nli_min_accuracy'], min_entailment_precision=CONFIG['nli_min_entailment_precision'], min_entailment_recall=CONFIG['nli_min_entailment_recall'], min_per_label_support=CONFIG['nli_min_per_label_support'], min_accuracy_lower=CONFIG['nli_min_accuracy_lower'], min_entailment_precision_lower=CONFIG['nli_min_entailment_precision_lower'], min_entailment_recall_lower=CONFIG['nli_min_entailment_recall_lower'])
        if not quality.get('enabled'): raise RuntimeError(f"nli_calibration_quality_failed:{language}")
        return entry
    return create_nli_calibration_artifact(language, rows, predictor, dataset_id=provenance['dataset_id'], dataset_revision=provenance['dataset_revision'], split=provenance['split'], label_mapping=label_map)
def _load_calibrated_nli_model(**kwargs):
    if MODEL_LABEL_MAPPING != {'LABEL_0': 'contradiction', 'LABEL_1': 'neutral', 'LABEL_2': 'entailment'}: raise RuntimeError('nli_model_label_mapping_mismatch')
    load_shared_nli_pipeline()  # validate the shared lifecycle before returning a non-retaining evaluator
    def score(premise, hypothesis):
        classifier, model_indices = load_shared_nli_pipeline()
        output = classifier({'text': premise, 'text_pair': hypothesis}, top_k=None)
        labels = {model_indices.get(str(item['label']).casefold(), model_indices.get(str(item['label']), -1)): float(item['score']) for item in (output[0] if output and isinstance(output[0], list) else output)}
        return labels.get(2, 0.0)
    return score
NLI_RUNTIME_BY_LANGUAGE = {}
CITATION_NLI_EVALUATORS = {}
NLI_THRESHOLDS = {}
def ensure_nli_calibration_artifacts():
    classifier, model_indices = load_shared_nli_pipeline()
    write_json(RUN / 'artifacts' / 'nli_memory_gate.json', _nli_memory_gate())
    for _language in CONFIG['evaluation_languages']:
        try:
            load_or_create_nli_calibration(_language, classifier, model_indices)
        except Exception as _calibration_error:
            write_json(RUN / 'artifacts' / 'nli_calibration_gate.json', {'status': 'blocked', 'language': _language, 'reason': str(_calibration_error)})
            raise RuntimeError(f'nli_calibration_gate_failed:{_language}')
ensure_nli_calibration_artifacts()
for _language in CONFIG['evaluation_languages']:
    _runtime_provenance = dict(CONFIG['nli_calibration_provenance'][_language])
    _runtime_provenance['dataset_content_hash'] = NLI_CALIBRATION_ARTIFACT[_language].get('dataset_content_hash')
    _runtime = load_nli_runtime_evaluator(_language, calibration_artifact=NLI_CALIBRATION_ARTIFACT, calibration_provenance={_language: _runtime_provenance}, model_loader=_load_calibrated_nli_model, model_id=CONFIG['nli_model_id'], revision=CONFIG['nli_model_revision'])
    NLI_RUNTIME_BY_LANGUAGE[_language] = _runtime
    CITATION_NLI_EVALUATORS[_language] = _runtime.get('evaluator') if _runtime.get('enabled') else None
    NLI_THRESHOLDS[_language] = float(_runtime.get('threshold', 0.5))
if CONFIG.get('require_citation_nli') and any(not NLI_RUNTIME_BY_LANGUAGE[_language].get('enabled') for _language in CONFIG['evaluation_languages']):
    write_json(RUN / 'artifacts' / 'nli_calibration_gate.json', {'status': 'blocked', 'reason': 'missing_or_unverified_language_calibration', 'languages': NLI_RUNTIME_BY_LANGUAGE})
    raise RuntimeError('nli_calibration_gate_failed_before_generation')
NLI_RUNTIME_BY_LANGUAGE = {language: {key: value for key, value in runtime.items() if key != 'evaluator'} for language, runtime in NLI_RUNTIME_BY_LANGUAGE.items()}
def validate_response(parsed, evidence, language=None):
    ledger = build_evidence_ledger(evidence); ledger_ids = {x["evidence_id"] for x in ledger}; clean_payload = {key: value for key, value in parsed.items() if key not in INTERNAL_RESPONSE_METADATA_KEYS} if isinstance(parsed, dict) else parsed; schema = validate_final_response(clean_payload, ledger_ids) if isinstance(parsed, dict) and parsed.get("parse_status") in ACCEPTED_PARSE_STATUSES else {"valid": False, "reasons": ["schema_invalid"], "quarantine": True}; narrative = str(parsed.get("counter_narrative", "")).strip() if isinstance(parsed, dict) else ""; cited_ids = parsed.get("cited_evidence_ids", []) if isinstance(parsed, dict) else []; ids = validate_evidence_ids(cited_ids, ledger); tokens, resolved, unknown = resolve_citation_tokens(narrative, ledger); reasons = list(schema.get("reasons", [])); factual_claims = parsed.get("factual_claims", []) if isinstance(parsed, dict) and isinstance(parsed.get("factual_claims", []), list) else []
    if not narrative: reasons.append("empty_narrative")
    declared_tokens = {value for value in cited_ids if isinstance(value, str)} if isinstance(cited_ids, list) else set()
    if not isinstance(cited_ids, list) or any(not isinstance(value, str) for value in cited_ids) or set(tokens) != declared_tokens: reasons.append("inline_and_declared_evidence_mismatch")
    if unknown: reasons.append("unknown_inline_evidence_ids")
    if "build_claim_citation_records" not in globals() or "aggregate_citation_support" not in globals():
        claim_records = []; support = {"entailment_status": "helper_unavailable", "pass": False, "abstention": False, "format_compliance": {"valid": False}, "citation_precision": None, "syntactic_citation_precision": None, "claim_citation_recall": None, "overcitation_count": 0, "citation_entailment": None}
    else:
        claim_records = build_claim_citation_records(narrative, ledger, factual_claims=factual_claims, safe_non_factual_validator=safe_abstention_validator)
        evaluator = globals().get('CITATION_NLI_EVALUATORS', {}).get(language)
        support = aggregate_citation_support(claim_records, evaluator=evaluator, threshold=float(globals().get('NLI_THRESHOLDS', {}).get(language, 0.5)), require_nli=bool(globals().get("CONFIG", {}).get("require_citation_nli", True)))
    if support["entailment_status"] == "helper_unavailable": reasons.append("citation_helpers_unavailable")
    if support["entailment_status"] == "unavailable": reasons.append("citation_entailment_unavailable")
    if support["entailment_status"] == "evaluator_error": reasons.append("citation_entailment_evaluator_error")
    if support["entailment_status"] == "scored_incomplete": reasons.append("citation_entailment_incomplete")
    if not support["pass"] and not support["abstention"] and support["entailment_status"] == "scored": reasons.append("citation_support_failed")
    status = support.get("entailment_status"); mode = "format_and_entailment" if status == "scored" else "format_only_abstention" if support.get("abstention") else "evaluator_error" if status == "evaluator_error" else "format_only_nli_unavailable"
    return {"citation_count": len(tokens), "valid_citation_count": len(resolved), "unknown_evidence_ids": sorted(set(ids["unknown"]) | set(unknown)), "claim_level_citations": claim_records, "citation_entailment": support.get("citation_entailment"), "entailment_mean": support.get("entailment_mean"), "evaluated_claim_entailment_mean": support.get("evaluated_claim_entailment_mean"), "evaluated_claim_count": support.get("evaluated_claim_count"), "incomplete_claim_count": support.get("incomplete_claim_count"), "citation_entailment_status": status, "citation_validation_mode": mode, "citation_format_compliance": support.get("format_compliance"), "citation_precision": support.get("citation_precision"), "syntactic_citation_precision": support.get("syntactic_citation_precision"), "claim_citation_recall": support.get("claim_citation_recall"), "citation_recall": support.get("citation_recall"), "citation_necessity": support.get("citation_necessity"), "necessary_citation_count": support.get("necessary_citation_count"), "overcitation_count": support.get("overcitation_count"), "schema_valid": schema["valid"], "validation_reasons": sorted(set(reasons)), "pass": schema["valid"] and ids["valid"] and not reasons and support["pass"]}
def authenticated_checkpoint_evidence(record, variant):
    if variant not in {"kg_rag", "mp_kg_rag"}: return []
    bundle = EVIDENCE_CACHE.get(str(record["ID"]))
    if not isinstance(bundle, dict) or not isinstance(bundle.get("evidence"), list): raise RuntimeError("resume_identity_mismatch")
    return bundle["evidence"]
def checkpoint_raw_envelope(final_raw, variant, mp_fields):
    if not isinstance(final_raw, str) or variant not in VARIANT_COLUMNS: raise RuntimeError("resume_identity_mismatch")
    expected_mp_keys = {"perspective_rationale", "perspective_parse_rate", "mp_perspective_outputs", "mp_response_plan", "mp_plan_raw_output"}
    if variant == "mp_kg_rag" and set(mp_fields) != expected_mp_keys: raise RuntimeError("resume_identity_mismatch")
    if variant != "mp_kg_rag" and mp_fields: raise RuntimeError("resume_identity_mismatch")
    return json.dumps({"schema_version": "mpkg-rag.checkpoint-envelope.v2", "variant": variant, "final_raw_output": final_raw, "mp_fields": mp_fields}, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
def checkpoint_materialization(payload, evidence, variant, language=None):
    if variant not in VARIANT_COLUMNS: raise RuntimeError("resume_identity_mismatch")
    narrative = str(payload.get("counter_narrative", "")).strip()
    verification = validate_response(payload, evidence, language=language) if language is not None else validate_response(payload, evidence)
    output_language = LANGUAGE_DETECTOR(narrative) if globals().get('LANGUAGE_DETECTOR') is not None else None
    verification['output_language'] = output_language
    if language is not None and output_language is not None and output_language != language:
        verification.setdefault('validation_reasons', []).append('output_language_mismatch'); verification['pass'] = False
    if not verification.get('schema_valid'): raise SchemaValidationQuarantine({'reason': 'final_schema_invalid', 'stage': 'checkpoint_materialization', 'variant': variant, 'validation_reasons': verification.get('validation_reasons', [])})
    if variant in {"kg_rag", "mp_kg_rag"} and not verification.get('pass'): raise SchemaValidationQuarantine({'reason': 'citation_support_failed', 'stage': 'checkpoint_materialization', 'variant': variant, 'validation_reasons': verification.get('validation_reasons', [])})
    return {"response": narrative, "parsed_counter_narrative": narrative, VARIANT_COLUMNS[variant]: narrative, "evidence": evidence, "evidence_ledger": build_evidence_ledger(evidence), "verification": verification, "input_language": language, "output_language": output_language}
def checkpoint_reasoning_trace(final_raw):
    trace_builder = globals().get("qwen_generation_trace")
    if callable(trace_builder): return trace_builder(final_raw)
    raw = str(final_raw or ""); match = re.search(r"<think>\s*(.*?)\s*</think>", raw, flags=re.S | re.I)
    return {"reasoning_content": match.group(1).strip() if match else "", "final_content": (raw[match.end():].strip() if match else raw.strip()), "thinking_status": "complete" if match else "not_emitted", "reasoning_truncated": False, "reasoning_token_count": None, "answer_token_count": None, "raw_generation": raw}
def canonical_checkpoint_row(record, variant, final_raw, parsed, evidence, mp_fields):
    detector = globals().get('LANGUAGE_DETECTOR')
    language = record.get('language') or record.get('Language') or (detector(record.get('Text', '')) if detector is not None else None)
    materialized = checkpoint_materialization(parsed, evidence, variant, language=language)
    trace_builder = globals().get("qwen_generation_trace"); reasoning_trace = trace_builder(final_raw) if callable(trace_builder) else {"reasoning_content": "", "final_content": str(final_raw or ""), "thinking_status": "not_emitted", "reasoning_truncated": False, "reasoning_token_count": None, "answer_token_count": None, "raw_generation": str(final_raw or "")}
    return {"ID": str(record["ID"]), "Text": record["Text"], "Category": record["Category"], "Target": record["Target"], "Counter Narrative": record["Counter Narrative"], "variant": variant, "split_name": CONFIG["split_name"], "input_text_sha256": record["input_text_sha256"], "config_hash": CONFIG_HASH, "prompt_template_hash": PROMPT_TEMPLATE_HASH, "checkpoint_identity": checkpoint_identity(record, variant), "parse_status": parsed.get("parse_status", "initial"), "reasoning_trace": reasoning_trace, "raw_output": checkpoint_raw_envelope(final_raw, variant, mp_fields), **materialized, **mp_fields}
def baseline_response(post, target, few_shot=False):
    examples = "" if not few_shot else "\nFrozen examples (do not copy identities or answers):\n" + "\n".join(f"Example post: {item['post']}\nTarget: {item['target']}\nSafe response: {item['response']}" for item in FEW_SHOT_EXAMPLES)
    prompt_revision = FEW_SHOT_PROMPT_REVISION if few_shot else "zero-shot.v1"; schema_tail = "JSON schema (must be followed exactly): " + json.dumps(FINAL_RESPONSE_SCHEMA, ensure_ascii=False, sort_keys=True)
    def builder(selected, payload):
        return [{"role": "system", "content": f"Return exactly one JSON object matching the final response schema. Prompt variant: {prompt_revision}."}, {"role": "user", "content": f"Post: {post}\nTarget: {target}{examples}\nReturn counter_narrative, cited_evidence_ids, factual_claims, safety_notes. Do not add wrappers or extra keys.\n{schema_tail}"}]
    if callable(globals().get("_adaptive_messages")):
        prompt, _ = _adaptive_messages([], builder, ("",), CONFIG["answer_max_new_tokens"])
    else:
        prompt = builder([], "")
    raw = generate_batch([prompt], CONFIG["answer_max_new_tokens"], temperature=0.6, enable_thinking=CONFIG.get("thinking_enabled", True), output_schema=FINAL_RESPONSE_SCHEMA)[0]; parsed, final_raw, status = parse_with_one_repair(raw, prompt[1]["content"], validate_final_response, FINAL_RESPONSE_SCHEMA); result = parsed if parsed is not None else {"counter_narrative": "", "cited_evidence_ids": [], "factual_claims": [], "safety_notes": [], "quarantine": True}; result["parse_status"] = status; result["few_shot"] = bool(few_shot); result["few_shot_prompt_revision"] = prompt_revision; return result, final_raw
def generate_variant(record, variant):
    rid, evidence = str(record["ID"]), (EVIDENCE_CACHE[str(record["ID"])]["evidence"] if variant in {"kg_rag", "mp_kg_rag"} else []); post, target, category = record["Text"], record["Target"], record["Category"]
    if variant == "qwen_zero_shot": parsed, raw = baseline_response(post, target)
    elif variant == "qwen_few_shot": parsed, raw = baseline_response(post, target, True)
    elif variant == "kg_rag": parsed, raw = generate_final_counter_narrative(post, target, evidence, {"claim_focus": "Address the claim respectfully.", "selected_evidence_ids": [x["evidence_id"] for x in build_evidence_ledger(evidence)], "response_steps": [], "tone": "empathetic", "factual_constraints": [], "safety_constraints": ["avoid amplification"]})
    elif variant == "mp_kg_rag":
        agents = run_perspectives(post, target, category, evidence); plan, plan_raw = synthesize_plan(post, target, category, evidence, agents)
        parsed, raw = generate_final_counter_narrative(post, target, evidence, plan)
    else: raise ValueError(variant)
    if not isinstance(parsed, dict) or parsed.get("parse_status") not in ACCEPTED_PARSE_STATUSES or parsed.get("quarantine"):
        raise SchemaValidationQuarantine({"reason": "generation_schema_invalid", "stage": variant, "parse_status": parsed.get("parse_status") if isinstance(parsed, dict) else None})
    narrative = str(parsed.get("counter_narrative", "")); ledger = build_evidence_ledger(evidence); extra = {"perspective_rationale": {p["perspective"]: p["perspective_rationale"] for p in agents}, "perspective_parse_rate": sum(p["structured_output"].get("parse_status") in ACCEPTED_PARSE_STATUSES for p in agents) / max(1, len(agents)), "mp_perspective_outputs": agents, "mp_response_plan": {**plan, "reasoning_trace": qwen_generation_trace(plan_raw)}, "mp_plan_raw_output": plan_raw} if variant == "mp_kg_rag" else {}
    try:
        return canonical_checkpoint_row(record, variant, raw, parsed, evidence, extra)
    except SchemaValidationQuarantine as failure:
        if variant not in {"kg_rag", "mp_kg_rag"}: raise
        repaired, repaired_raw = grounding_repair_response(post, target, evidence, plan if variant == "mp_kg_rag" else {"claim_focus": "Use only evidence."}, [str(failure)])
        return canonical_checkpoint_row(record, variant, repaired_raw, repaired, evidence, extra)
def revalidate_checkpoint_row(row, record, variant):
    if not isinstance(row, dict) or not isinstance(row.get("ID"), (str, int)): raise RuntimeError("resume_identity_mismatch")
    validate_checkpoint_identity(row.get("checkpoint_identity"), checkpoint_identity(record, variant))
    if not isinstance(row.get("raw_output"), str) or row.get("parse_status") not in ACCEPTED_PARSE_STATUSES: raise RuntimeError("resume_identity_mismatch")
    try:
        envelope = parse_json_object(row["raw_output"])
        if set(envelope) != {"schema_version", "variant", "final_raw_output", "mp_fields"} or envelope.get("schema_version") != "mpkg-rag.checkpoint-envelope.v2" or envelope.get("variant") != variant or not isinstance(envelope.get("final_raw_output"), str) or not isinstance(envelope.get("mp_fields"), dict): raise RuntimeError("resume_identity_mismatch")
        payload = parse_json_object(envelope["final_raw_output"]); payload["parse_status"] = row["parse_status"]
        authenticated_evidence = authenticated_checkpoint_evidence(record, variant)
        canonical = canonical_checkpoint_row(record, variant, envelope["final_raw_output"], payload, authenticated_evidence, envelope["mp_fields"])
    except Exception as exc:
        if isinstance(exc, RuntimeError) and str(exc) == "resume_identity_mismatch": raise
        raise RuntimeError("resume_identity_mismatch")
    if row != canonical: raise RuntimeError("resume_identity_mismatch")
    return row
def checkpoint_lock(path):
    return Path(path).with_name(Path(path).name + ".lock")
def read_checkpoint_rows_locked(path):
    path = Path(path); lock_path = checkpoint_lock(path); lock_path.parent.mkdir(parents=True, exist_ok=True)
    with lock_path.open("a+", encoding="utf-8") as lock:
        fcntl.flock(lock.fileno(), fcntl.LOCK_EX)
        try: return load_jsonl(path)
        finally: fcntl.flock(lock.fileno(), fcntl.LOCK_UN)
def append_checkpoint_row(path, row):
    path = Path(path); lock_path = checkpoint_lock(path); lock_path.parent.mkdir(parents=True, exist_ok=True)
    with lock_path.open("a+", encoding="utf-8") as lock:
        fcntl.flock(lock.fileno(), fcntl.LOCK_EX)
        try:
            existing_rows = load_jsonl(path); record_id = str(row.get("ID", ""))
            for existing in existing_rows:
                if str(existing.get("ID", "")) != record_id: continue
                if existing == row: return False
                raise RuntimeError("duplicate_checkpoint_conflict")
            with path.open("a", encoding="utf-8") as stream: stream.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n"); stream.flush(); os.fsync(stream.fileno())
            return True
        finally: fcntl.flock(lock.fileno(), fcntl.LOCK_UN)
def load_checkpoint_rows(path, data, variant):
    previous = read_checkpoint_rows_locked(path); expected = {str(row["ID"]): row for row in data.to_dict("records")}; seen = set()
    for row in previous:
        rid = str(row.get("ID", ""))
        if rid in seen: raise RuntimeError("duplicate_checkpoint_ids")
        if rid not in expected: raise RuntimeError("unknown_checkpoint_id")
        revalidate_checkpoint_row(row, expected[rid], variant); seen.add(rid)
    return previous, {str(row["ID"]): row for row in previous}
def row_quarantine_row(record, variant, audit):
    audit = dict(audit or {}); reason = str(audit.get("reason", "row_quarantine")); prompt_quarantine = reason.startswith("prompt_budget") or reason == "repair_prompt_budget"
    return {"ID": str(record["ID"]), "Text": record.get("Text"), "Category": record.get("Category"), "Target": record.get("Target"), "Counter Narrative": record.get("Counter Narrative"), "variant": variant, "split_name": CONFIG["split_name"], "parsed_counter_narrative": None, "response": None, "raw_output": None, "parse_status": "quarantined", "prompt_quarantine": prompt_quarantine, "prompt_quarantine_reason": reason if prompt_quarantine else None, "prompt_budget_audit": audit, "generation_quarantine_reason": reason, "verification": {"pass": False, "validation_reasons": [reason]}, "evidence": [], "evidence_ledger": []}
def prompt_budget_quarantine_row(record, variant, audit):
    return row_quarantine_row(record, variant, audit)
def _excel_cell(value):
    if isinstance(value, (dict, list, tuple)): value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    if value is None: return None
    return str(value)[:32767]
def _style_production_workbook(path):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    workbook = load_workbook(path); header_fill = PatternFill("solid", fgColor="1F4E78"); header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions; sheet.sheet_view.showGridLines = False
        for cell in sheet[1]: cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 32
        for column_cells in sheet.columns:
            letter = column_cells[0].column_letter; header = str(column_cells[0].value or "").casefold(); width = 14
            if any(key in header for key in ["text", "narrative", "reasoning", "evidence", "perspective", "plan", "verification", "raw"]): width = 48
            elif header in {"id", "status", "variant", "category", "target"}: width = 18
            sheet.column_dimensions[letter].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)
def export_production_workbook(frame):
    output_path = RUN / "exports" / "dataset_with_all_rag_counter_narratives.xlsx"
    base_columns = ["ID", "Text", "Category", "Target", "Counter Narrative"]; outputs = dataset[base_columns].copy(); outputs["ID"] = outputs["ID"].astype(str)
    trace_rows = {}; variant_output_columns = {"qwen_zero_shot": "zero-shot-counter-narrative", "qwen_few_shot": "few-shot-counter-narrative", "kg_rag": "kg-rag-counter-narrative", "mp_kg_rag": "mp-kg-rag-counter-narrative"}
    for variant, output_column in variant_output_columns.items():
        selected = frame[frame.variant.eq(variant)].copy(); selected["ID"] = selected["ID"].astype(str); indexed = selected.drop_duplicates("ID").set_index("ID")
        outputs[output_column] = outputs["ID"].map(indexed["parsed_counter_narrative"]); outputs[variant + "-status"] = outputs["ID"].map(indexed["parse_status"])
        for row in selected.to_dict("records"):
            trace = row.get("reasoning_trace") if isinstance(row.get("reasoning_trace"), dict) else {}
            trace_rows.setdefault(variant, []).append({"ID": str(row["ID"]), "reasoning_content": trace.get("reasoning_content"), "thinking_status": trace.get("thinking_status"), "reasoning_truncated": trace.get("reasoning_truncated"), "reasoning_token_count": trace.get("reasoning_token_count"), "answer_token_count": trace.get("answer_token_count"), "parse_status": row.get("parse_status"), "final_counter_narrative": row.get("parsed_counter_narrative"), "evidence_ledger": _excel_cell(row.get("evidence_ledger")), "verification": _excel_cell(row.get("verification")), "mp_perspective_outputs": _excel_cell(row.get("mp_perspective_outputs")), "mp_response_plan": _excel_cell(row.get("mp_response_plan")), "raw_generation": _excel_cell(trace.get("raw_generation"))})
    evidence_rows = []
    for row in frame[frame.variant.isin(["kg_rag", "mp_kg_rag"])].to_dict("records"):
        for evidence in row.get("evidence_ledger", []) if isinstance(row.get("evidence_ledger"), list) else []: evidence_rows.append({"ID": str(row["ID"]), "variant": row["variant"], **{key: _excel_cell(value) for key, value in evidence.items()}})
    quality = frame.groupby("variant", dropna=False).agg(rows=("ID", "size"), unique_ids=("ID", "nunique"), accepted=("parse_status", lambda values: int(sum(value in ACCEPTED_PARSE_STATUSES for value in values)))).reset_index(); quality["parse_rate"] = quality["accepted"] / quality["rows"].clip(lower=1)
    manifest = pd.DataFrame([{"run_name": RUN_NAME, "model": CONFIG["generator_model"], "model_revision": CONFIG["generator_model_revision"], "load_in_4bit": CONFIG["load_in_4bit"], "thinking_enabled": CONFIG["thinking_enabled"], "input_rows": len(dataset), "input_sha256": hashlib.sha256(Path(CONFIG["dataset_xlsx"]).read_bytes()).hexdigest(), "config_hash": CONFIG_HASH, "prompt_template_hash": PROMPT_TEMPLATE_HASH}])
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        outputs.to_excel(writer, sheet_name="Outputs", index=False)
        sheet_names = {"qwen_zero_shot": "Zero-Shot Trace", "qwen_few_shot": "Few-Shot Trace", "kg_rag": "KG-RAG Trace", "mp_kg_rag": "MP-KG-RAG Trace"}
        for variant, sheet_name in sheet_names.items(): pd.DataFrame(trace_rows.get(variant, [])).to_excel(writer, sheet_name=sheet_name, index=False)
        pd.DataFrame(evidence_rows).to_excel(writer, sheet_name="Evidence Ledger", index=False); manifest.to_excel(writer, sheet_name="Run Manifest", index=False); quality.to_excel(writer, sheet_name="Quality Summary", index=False)
    _style_production_workbook(output_path); return output_path
def generate_all_variants(data):
    global GENERATION_QUARANTINE_SUMMARY
    outputs = []; generation_quarantines = []
    for variant in CONFIG["generation_variants"]:
        checkpoint = RUN / "checkpoints" / ("mp_kg_rag_rows.jsonl" if variant == "mp_kg_rag" else f"{variant}_rows.jsonl"); previous, by_id = load_checkpoint_rows(checkpoint, data, variant)
        for record in tqdm(data.to_dict("records"), desc=f"Generate {variant}"):
            if str(record["ID"]) in by_id: outputs.append(by_id[str(record["ID"])]); continue
            try:
                row = generate_variant(record, variant)
            except GenerationRowQuarantine as failure:
                audit = {"record_id": str(record["ID"]), "variant": variant, "quarantine_type": type(failure).__name__, **failure.audit}; generation_quarantines.append(audit); row = row_quarantine_row(record, variant, audit); outputs.append(row); continue
            append_checkpoint_row(checkpoint, row); outputs.append(row)
    type_counts = {}; reason_counts = {}; variant_counts = {}
    for audit in generation_quarantines:
        quarantine_type = str(audit.get("quarantine_type", "GenerationRowQuarantine")); reason = str(audit.get("reason", "row_quarantine")); variant_name = str(audit.get("variant", "unknown"))
        type_counts[quarantine_type] = type_counts.get(quarantine_type, 0) + 1
        reason_counts[reason] = reason_counts.get(reason, 0) + 1
        variant_counts[variant_name] = variant_counts.get(variant_name, 0) + 1
    GENERATION_QUARANTINE_SUMMARY = {"count": len(generation_quarantines), "counts_by_quarantine_type": type_counts, "counts_by_reason": reason_counts, "counts_by_variant": variant_counts}
    write_json(RUN / "artifacts" / "generation_quarantine.json", {**GENERATION_QUARANTINE_SUMMARY, "rows": generation_quarantines})
    prompt_rows = [row for row in generation_quarantines if str(row.get("reason", "")).startswith("prompt_budget") or row.get("reason") == "repair_prompt_budget"]
    write_json(RUN / "artifacts" / "prompt_budget_quarantine.json", {"count": len(prompt_rows), "rows": prompt_rows, "reason": "prompt_budget_irreducible"})
    frame = pd.DataFrame(outputs); frame.to_json(RUN / "exports" / "all_variant_outputs.jsonl", orient="records", lines=True, force_ascii=False); frame.to_excel(RUN / "exports" / "all_variant_outputs.xlsx", index=False); frame[frame.variant.eq("mp_kg_rag")].to_excel(RUN / "exports" / "dataset_with_mp_kg_rag_counter_narratives.xlsx", index=False); export_production_workbook(frame); return frame
def load_qwen_for_generation(): return load_qwen_for_extraction()
release_shared_nli_pipeline()
load_qwen_for_generation(); print("Generator model:", CONFIG["generator_model"], "is_loaded_in_4bit:", getattr(model, "is_loaded_in_4bit", None), "quantization_config:", getattr(getattr(model, "config", None), "quantization_config", None), "first_parameter_dtype:", next(model.parameters()).dtype)
generation_frame = generate_all_variants(dataset)
GENERATION_FRAME_MATERIALIZED = generation_frame.copy(deep=True)
unload_generator()
if "model" in globals() or "tokenizer" in globals(): raise RuntimeError("qwen_generation_unload_failed")
QWEN_METRICS_MEMORY_GATE_GB = float(os.environ.get("QWEN_METRICS_MEMORY_GATE_GB", "1.0"))
QWEN_METRICS_GPU_SNAPSHOT = gpu_snapshot("qwen_generation_unloaded_before_metrics")
if float(QWEN_METRICS_GPU_SNAPSHOT.get("allocated_gb", 0.0)) > QWEN_METRICS_MEMORY_GATE_GB: raise RuntimeError("qwen_metrics_memory_gate")
QWEN_METRICS_GPU_GATE_PASSED = True



# %% [notebook cell 15]

# 13 - Language-aware automatic metrics and corrected paired comparisons.
from rouge_score import rouge_scorer
from nltk.translate.meteor_score import meteor_score
ROUGE = rouge_scorer.RougeScorer(['rougeL'], use_stemmer=True)
LANGUAGE_DETECTOR = load_language_detector()
if LANGUAGE_DETECTOR is None: raise RuntimeError('language_evaluation_detector_unavailable')
BERTSCORE_DEVICE = "cuda" if QWEN_METRICS_GPU_GATE_PASSED else "cpu"
if "model" in globals() or "tokenizer" in globals(): raise RuntimeError("detoxify_model_unloaded_before_metrics")
def _rouge(reference, hypothesis): return ROUGE.score(reference, hypothesis)['rougeL'].fmeasure
def _meteor(reference, hypothesis): return meteor_score([reference.split()], hypothesis.split())
def reference_metric_row(row):
    reference = normalize_optional_text(row["Counter Narrative"] if "Counter Narrative" in row else None)
    response = normalize_optional_text(row.get("parsed_counter_narrative"))
    if reference is None: return {'rouge_l': None, 'reference_metric_status': 'excluded_missing_reference', 'reference_metrics_excluded': True}
    if response is None: return {'rouge_l': None, 'reference_metric_status': 'missing_hypothesis', 'reference_metrics_excluded': False}
    return {'rouge_l': ROUGE.score(reference, response)['rougeL'].fmeasure, 'reference_metric_status': 'scored', 'reference_metrics_excluded': False}
def citation_metrics(row):
    verification = row.get('verification') if isinstance(row.get('verification'), dict) else {}
    if str(row.get('variant', '')).casefold() not in {'kg_rag', 'mp_kg_rag'}:
        return {name: None for name in ('citation_precision', 'citation_recall', 'citation_necessity', 'citation_entailment')} | {name: 'not_applicable_non_rag' for name in ('citation_precision_status', 'citation_recall_status', 'citation_necessity_status', 'citation_entailment_status')}
    return {'citation_precision': verification.get('citation_precision'), 'citation_recall': verification.get('citation_recall'), 'citation_necessity': verification.get('citation_necessity'), 'citation_entailment': verification.get('citation_entailment'), 'citation_precision_status': verification.get('citation_entailment_status'), 'citation_recall_status': verification.get('citation_entailment_status'), 'citation_necessity_status': verification.get('citation_entailment_status'), 'citation_entailment_status': verification.get('citation_entailment_status')}
def _detoxify(hypothesis): return float(DETOXIFY_MODEL.predict(hypothesis).get('toxicity'))
DETOXIFY_MODEL = None
DETOXIFY_DEVICE = 'cpu'
DETOXIFY_STATUS = {'status': 'unavailable', 'device': DETOXIFY_DEVICE, 'required': bool(CONFIG.get('require_detoxify', True))}
try:
    from detoxify import Detoxify
    if "model" in globals() or "tokenizer" in globals(): raise RuntimeError("detoxify_model_unloaded_before_metrics")
    DETOXIFY_MODEL = Detoxify('original', device='cpu')
    DETOXIFY_STATUS = {'status': 'loaded', 'device': DETOXIFY_DEVICE, 'required': bool(CONFIG.get('require_detoxify', True)), 'model': 'original'}
except Exception as error:
    DETOXIFY_MODEL = None
    DETOXIFY_STATUS = {'status': 'unavailable', 'device': DETOXIFY_DEVICE, 'required': bool(CONFIG.get('require_detoxify', True)), 'reason': f'{type(error).__name__}:{error}'}
    # Persist the unavailable contract before the required-model fail-closed
    # raise, so an aborted run still retains an explicit audit record.
    write_json(RUN / 'artifacts' / 'detoxify_metric_contract.json', DETOXIFY_STATUS)
    if CONFIG.get('require_detoxify', True): raise RuntimeError(f'detoxify_required_unavailable:{DETOXIFY_STATUS["reason"]}') from error
write_json(RUN / 'artifacts' / 'detoxify_metric_contract.json', DETOXIFY_STATUS)
_BERTSCORE_LOCAL_MODELS = {}
_BERTSCORE_CONFIG_HASH = stable_identity_hash(multilingual_bertscore_config({'model_id': CONFIG['bertscore_model_id'], 'revision': CONFIG['bertscore_model_revision']}))
def _bertscore(**kwargs):
    from huggingface_hub import snapshot_download
    from bert_score import score
    cache_key = (kwargs['model_id'], kwargs['revision'])
    if cache_key not in _BERTSCORE_LOCAL_MODELS:
        _BERTSCORE_LOCAL_MODELS[cache_key] = snapshot_download(repo_id=kwargs['model_id'], revision=kwargs['revision'])
    local_model = _BERTSCORE_LOCAL_MODELS[cache_key]
    _, _, f1 = score([kwargs['hypothesis']], [kwargs['reference']], lang=kwargs['language'], model_type=local_model, num_layers=BERTSCORE_NUM_LAYERS, device=BERTSCORE_DEVICE, verbose=False)
    return {'f1': float(f1[0].item())}
metric_rows = []
for _, row in generation_frame.iterrows():
    evaluated = evaluate_multilingual_record(row.to_dict(), language_detector=LANGUAGE_DETECTOR, rouge_scorer=_rouge, meteor_scorer=_meteor, bertscore_scorer=_bertscore, detoxify_scorer=_detoxify if DETOXIFY_MODEL is not None else None, bertscore_config={'model_id': CONFIG['bertscore_model_id'], 'revision': CONFIG['bertscore_model_revision']})
    output = {'ID': str(row['ID']), 'variant': str(row['variant']), 'response': normalize_optional_text(row.get('parsed_counter_narrative')), **{key: evaluated.get(key) for key in ['language', 'language_status', 'expected_language', 'input_language', 'reference_language', 'output_language', 'language_match', 'input_language_match', 'reference_language_match', 'script_match', 'reference_available']}}
    for metric_name, metric_value in evaluated['metrics'].items():
        output[metric_name] = metric_value.get('value'); output[f'{metric_name}_status'] = metric_value.get('status'); output[f'{metric_name}_reason'] = metric_value.get('reason')
    output.update(citation_metrics(row.to_dict()))
    metric_rows.append(output)
metrics = pd.DataFrame(metric_rows)
validate_metric_rows_unique(metric_rows)
language_summary = summarize_metric_records([{'language': row.get('language'), 'metrics': {name: {'value': row.get(name), 'status': row.get(f'{name}_status'), 'reason': row.get(f'{name}_reason')} for name in ['chrf', 'bertscore', 'rouge_l', 'meteor', 'detoxify']}} for row in metric_rows])
language_diagnostics = summarize_language_records(metric_rows)
coverage = {language: sum(row.get('bertscore_status') == 'scored' for row in metric_rows if row.get('language') == language) for language in CONFIG['evaluation_languages']}
observed_languages = {row.get('language') for row in metric_rows if row.get('language') in CONFIG['evaluation_languages']}
if any(coverage[language] < 1 for language in observed_languages): raise RuntimeError('bertscore_minimum_language_coverage_failed')
DETERMINISTIC_PLOT_KWARGS = dict(errorbar=None, bootstrap_seed=BOOTSTRAP_SEED)
write_json(RUN / 'artifacts' / 'language_metric_summary.json', {'rows': language_summary, 'language_diagnostics': language_diagnostics, 'bertscore_coverage': coverage, 'languages': CONFIG['evaluation_languages'], 'bertscore_config': multilingual_bertscore_config({'model_id': CONFIG['bertscore_model_id'], 'revision': CONFIG['bertscore_model_revision']}), 'deterministic_plot_kwargs': DETERMINISTIC_PLOT_KWARGS})
metrics.to_parquet(RUN / 'exports' / 'per_record_automatic_metrics.parquet', index=False); metrics.to_excel(RUN / 'exports' / 'per_record_automatic_metrics.xlsx', index=False)
pairwise = pairwise_metric_family(metric_rows, metrics=('chrf', 'bertscore', 'rouge_l', 'meteor', 'detoxify', 'citation_precision', 'citation_recall', 'citation_necessity', 'citation_entailment'), directions={'detoxify': 'lower'}, seed=SEED, permutations=10000)
write_json(RUN / 'artifacts' / 'pairwise_variant_comparisons.json', pairwise)
write_json(RUN / 'artifacts' / 'reference_metric_audit.json', build_reference_metric_audit(metric_rows, metrics=('chrf', 'bertscore', 'rouge_l', 'meteor', 'detoxify')))



# %% [notebook cell 16]

# 14 - Bounded stratified human annotation workbook and reliability gate.
HUMAN_METRICS = ['politeness_respect', 'helpfulness', 'factuality', 'safety', 'empathy', 'fluency', 'inclusivity', 'non_confrontational', 'contextual_coherence', 'persuasiveness']
def build_human_annotation_workbook(frame, output_path):
    dataset_keys = dataset.set_index(dataset['ID'].astype(str))[['script_bucket', 'stratify_key']].to_dict('index') if 'script_bucket' in dataset.columns else {}
    source = []
    for row in frame.to_dict('records'):
        key = dataset_keys.get(str(row['ID']), {})
        source.append({**row, 'script_bucket': key.get('script_bucket', 'unknown'), 'stratify_key': key.get('stratify_key', 'unknown')})
    sample = sample_annotation_records(source, max_ids=int(CONFIG['annotation_max_ids']), seed=SEED, stratify_key='stratify_key')
    ratings_source = pd.DataFrame(sample['rows']).copy()
    key_rows = ratings_source[['ID', 'variant']].copy(); key_rows.insert(0, 'blind_id', [f'R{n:04d}' for n in range(1, len(key_rows) + 1)])
    key_rows['evidence_ids'] = ratings_source.get('evidence_ledger', pd.Series([None] * len(key_rows))).map(lambda value: [item.get('evidence_id') for item in value] if isinstance(value, list) else None)
    key_rows.to_json(RUN / 'artifacts' / 'annotation_blind_key.jsonl', orient='records', lines=True, force_ascii=False)
    ratings = ratings_source.rename(columns={'Text': 'post', 'Target': 'target', 'parsed_counter_narrative': 'response'}).copy()
    ratings.insert(0, 'blind_id', key_rows['blind_id'])
    ratings = ratings[['blind_id', 'post', 'target', 'response'] + [column for column in HUMAN_METRICS if column in ratings.columns]]
    if ratings.empty: raise RuntimeError('annotation_sample_empty')
    for metric in HUMAN_METRICS: ratings[metric] = pd.NA
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        ratings.to_excel(writer, sheet_name='Ratings', index=False)
        pd.DataFrame({'metric': HUMAN_METRICS}).to_excel(writer, sheet_name='Rubric', index=False)
        pd.DataFrame([sample]).to_json(RUN / 'artifacts' / 'annotation_sampling_manifest.json', orient='records')
    return sample
annotation_sample = build_human_annotation_workbook(generation_frame, RUN / 'exports' / 'human_annotation_blinded.xlsx'); print('human ratings are required; at least two distinct raters must complete the workbook')
def export_human_agreement(completed_rows):
    report = human_agreement_report(completed_rows)
    write_json(RUN / 'artifacts' / 'human_human_agreement.json', report)
    return report
COMPLETED_ANNOTATIONS = RUN / 'input' / 'human_annotations_completed.jsonl'
if COMPLETED_ANNOTATIONS.exists():
    completed_rows = load_jsonl(COMPLETED_ANNOTATIONS)
    if len({str(row.get('rater_id')) for row in completed_rows if row.get('rater_id')}) < 2: raise RuntimeError('human_agreement_requires_two_distinct_raters')
    export_human_agreement(completed_rows)
else:
    write_json(RUN / 'artifacts' / 'human_human_agreement.json', {'status': 'pending', 'reason': 'completed_annotations_not_supplied', 'required_distinct_raters': 2, 'weighted_kappa': None, 'krippendorff_alpha_ordinal': None})



# %% [notebook cell 17]

# 15 - Final run manifest.
NLI_CALIBRATION_MANIFEST = {language: {key: value for key, value in runtime.items() if key != 'evaluator'} for language, runtime in NLI_RUNTIME_BY_LANGUAGE.items()}
write_json(RUN / "exports" / "run_manifest.json", {"created_at": now(), "run_name": RUN_NAME, "run_identity_hash": RUN_IDENTITY_HASH, "config": {k: str(v) if isinstance(v, Path) else v for k, v in CONFIG.items()}, "config_hash": CONFIG_HASH, "lockfile_sha256": LOCKFILE_SHA256, "environment_fingerprint_hash": ENVIRONMENT_FINGERPRINT_HASH, "managed_accelerator_contract": MANAGED_ACCELERATOR_CONTRACT, "managed_accelerator_contract_hash": MANAGED_ACCELERATOR_CONTRACT_HASH, "qwen35_transformers_compatibility": QWEN35_TRANSFORMERS_COMPATIBILITY, "prompt_template_hash": PROMPT_TEMPLATE_HASH, "core_source_sha256": CORE_SOURCE_SHA256, "eval_core_source_sha256": EVAL_CORE_SOURCE_SHA256, "corpus_manifest_hash": CORPUS_MANIFEST_HASH, "audit_manifest_hash": AUDIT_MANIFEST_HASH, "chunk_manifest_hash": CHUNK_MANIFEST_HASH, "graph_manifest_hash": GRAPH_MANIFEST_HASH, "checkpoint_identity": ["input_text_sha256", "corpus_manifest_hash", "audit_manifest_hash", "chunk_manifest_hash", "graph_manifest_hash", "run_identity_hash", "lockfile_sha256", "environment_fingerprint_hash", "input_language", "output_language"], "paired_retrieval_evaluation": True, "memory_snapshots": MEMORY_SNAPSHOTS, "environment_fingerprint": ENVIRONMENT_FINGERPRINT, "evaluator_models": {"bertscore": {"model_id": BERTSCORE_MODEL_ID, "revision": BERTSCORE_MODEL_REVISION, "local_cache_keys": [list(key) for key in _BERTSCORE_LOCAL_MODELS]}, "nli": {"model_id": CONFIG["nli_model_id"], "revision": CONFIG["nli_model_revision"]}}, "nli_calibration_status": NLI_RUNTIME_BY_LANGUAGE, "dataset_limitations": {"en": "XNLI English validation", "hi": "IndicXNLI Hindi validation", "ta": "IndicXNLI Tamil validation; no evaluator without passing held-out gate"}})
_run_manifest_path = RUN / "exports" / "run_manifest.json"
_run_manifest = json.loads(_run_manifest_path.read_text(encoding="utf-8"))
_run_manifest["lock_package_counts"] = LOCK_PACKAGE_COUNTS
_run_manifest["detoxify_metric"] = DETOXIFY_STATUS
_run_manifest["detoxify_device"] = DETOXIFY_DEVICE
_run_manifest["detoxify_status"] = DETOXIFY_STATUS["status"]
_run_manifest["metric_devices"] = {"bertscore": BERTSCORE_DEVICE, "detoxify": DETOXIFY_DEVICE}
_run_manifest["generation_quarantine"] = GENERATION_QUARANTINE_SUMMARY
write_json(_run_manifest_path, _run_manifest)

